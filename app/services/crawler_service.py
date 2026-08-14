from __future__ import annotations

import json
import re
import uuid
import base64
import binascii
import calendar
import hashlib
import hmac
import logging
import math
import mimetypes
import random
import shutil
import unicodedata
import xml.etree.ElementTree as ET
import time
import threading
from contextlib import contextmanager
from copy import deepcopy
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from html import unescape
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Callable, Iterable, Iterator
from urllib.parse import parse_qs, quote, unquote, urlencode, urljoin, urlsplit, urlunsplit

import requests
from bs4 import BeautifulSoup, Comment
from sqlalchemy import and_, case, exists, func, inspect, or_, select, update
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import aliased

from app.core.config import settings
from app.core.secure_storage import decrypt_text, encrypt_text, mask_secret
from app.core.text_limits import (
    RAKUTEN_TAGLINE_MAX_BYTES,
    RAKUTEN_TITLE_MAX_BYTES,
    truncate_utf8_bytes,
)
from app.core.task_queue import all_task_queue_names, enqueue_task, enqueue_task_in, redis_connection, task_queue_name_for_kind
from app.db.database import session_scope
from app.db.models import (
    AutoDeletionTaskModel,
    AutoListingScheduleModel,
    CrawlLogModel,
    CrawlSourceModel,
    CrawlTaskModel,
    DeletedProductImageCleanupModel,
    ListingTaskModel,
    ProductSalesDailyModel,
    ProductModel,
    ProductTitleVersionModel,
    RoleModel,
    SalesOrderModel,
    SalesSyncStateModel,
    ScheduledCrawlModel,
    StoreModel,
    SyncTaskModel,
    SystemSettingModel,
    UserAccountModel,
    UserCollectionGenreRuleModel,
    UserCollectionGenreSettingModel,
    canonical_sales_order_item_product_key,
    make_source_url_hash,
)
from app.services import sales_order_sync_history_service
from app.services.product_image_storage import (
    DRAFT_IMAGE_OBJECT_PREFIX,
    DRAFT_IMAGE_URL_PREFIX,
    PRODUCT_IMAGE_OBJECT_PREFIX,
    PRODUCT_IMAGE_URL_PREFIX,
    StoredObject,
    parse_product_image_url,
    product_image_storage,
)
from app.services.sensitive_word_service import active_sensitive_words, sanitize_product_payload
from app.services.sales_time import (
    iso_sales_datetime,
    sales_now_naive,
)
from app.services.task_status_ordering import task_status_order_by
from app.services.user_service import account_crawl_price_rule, normalize_crawl_price_rule

logger = logging.getLogger(__name__)


EMPTY_SENSITIVE_TITLE_SAVE_ERROR = "商品标题命中敏感词后为空，商品未保存。"
RAKUTEN_IMAGE_CDN_HOSTS = {
    "image.rakuten.co.jp",
    "shop.r10s.jp",
    "tshop.r10s.jp",
}
PRODUCT_IMAGE_VISUAL_SIZE = (32, 32)
PRODUCT_IMAGE_VISUAL_MAX_MEAN_DIFFERENCE = 2.0
PRODUCT_IMAGE_VISUAL_MAX_ASPECT_RATIO_DIFFERENCE = 0.03
CRAWL_SAVE_LOCK_RETRY_DELAYS_SECONDS = (0.5, 1.0, 2.0)
WHOLE_SHOP_PARTITION_TARGET_ITEMS = 5000
WHOLE_SHOP_PARTITION_MAX_STEPS = 32


def system_task_dispatch_paused(
    owner_username: str | None = None,
) -> bool:
    return (
        str(owner_username or "").strip()
        in system_task_paused_usernames()
    )


def system_task_paused_usernames() -> set[str]:
    from app.db.models import SystemTaskControlModel
    from app.services.task_control_service import TASK_CONTROL_ROW_ID

    with session_scope() as session:
        row = session.get(SystemTaskControlModel, TASK_CONTROL_ROW_ID)
        if row is None or not row.paused:
            return set()
        try:
            snapshot = json.loads(row.snapshot_json or "{}")
        except (TypeError, ValueError):
            snapshot = {}
        if not isinstance(snapshot, dict):
            snapshot = {}
    return {
        str(username or "").strip()
        for username in snapshot.get("selectedUsernames", [])
        if str(username or "").strip()
    }


def ensure_system_task_dispatch_allowed(owner_username: str) -> None:
    if system_task_dispatch_paused(owner_username):
        raise RuntimeError("系统维护期间任务调度已暂停，请等待超级管理员恢复。")


@dataclass(frozen=True)
class PreparedProductUpsertItem:
    item: dict[str, Any]
    source_url: str
    title: str
    source_url_hash_key: str
    rakuten_manage_number: str | None
    error: str = ""


@dataclass(frozen=True)
class ListingProductAttemptResult:
    product_id: int
    success: bool
    error: str = ""


@dataclass(frozen=True)
class CollectionGenrePolicySnapshot:
    default_policy: str = "allow"
    unknown_genre_policy: str = "allow"
    rules_by_path: dict[str, str] | None = None


@dataclass(frozen=True)
class CollectedItemPlan:
    total_count: int
    items: Iterable[dict[str, Any]]


@dataclass(frozen=True)
class ProductImageVisualSignature:
    width: int
    height: int
    pixels: bytes


PROXY_USAGE_CACHE_LOCK = threading.Lock()
PROXY_USAGE_REFRESH_LOCK = threading.Lock()
PROXY_USAGE_CACHE: dict[str, Any] = {}

RAKUTEN_SEARCH_BASE = "https://search.rakuten.co.jp/search/mall/"
RAKUTEN_RANKING_BASE = "https://ranking.rakuten.co.jp/search"
RAKUTEN_REALTIME_RANKING_URL = "https://ranking.rakuten.co.jp/realtime/"
RAKUTEN_SHOP_MASTER_URL = "https://api.rms.rakuten.co.jp/es/1.0/shop/shopMaster"
RAKUTEN_CABINET_USAGE_URL = "https://api.rms.rakuten.co.jp/es/1.0/cabinet/usage/get"
RAKUTEN_ITEM_SEARCH_URL = "https://api.rms.rakuten.co.jp/es/2.0/items/search"
RAKUTEN_ITEM_PATCH_URL = "https://api.rms.rakuten.co.jp/es/2.0/items/manage-numbers/{manageNumber}"
RAKUTEN_CABINET_FILE_SEARCH_URL = "https://api.rms.rakuten.co.jp/es/1.0/cabinet/files/search"
RAKUTEN_CABINET_FILE_DELETE_URL = "https://api.rms.rakuten.co.jp/es/1.0/cabinet/file/delete"
RAKUTEN_CABINET_FILE_INSERT_URL = "https://api.rms.rakuten.co.jp/es/1.0/cabinet/file/insert"
RAKUTEN_CABINET_FOLDERS_GET_URL = "https://api.rms.rakuten.co.jp/es/1.0/cabinet/folders/get"
RAKUTEN_CABINET_FOLDER_INSERT_URL = "https://api.rms.rakuten.co.jp/es/1.0/cabinet/folder/insert"
RAKUTEN_INVENTORY_BULK_UPSERT_URL = "https://api.rms.rakuten.co.jp/es/2.1/inventories/bulk-upsert"
RAKUTEN_ITEM_SEARCH_HITS = 100
RAKUTEN_ITEM_SEARCH_MAX_RETRIES = 4
RAKUTEN_WRITE_MAX_RETRIES = 5
RAKUTEN_WRITE_RETRY_STATUS_CODES = {429, 500, 502, 503, 504}
RAKUTEN_WRITE_QPS_BACKOFF_SECONDS = (1.5, 3.0, 5.0, 8.0)
RAKUTEN_WRITE_REQUEST_MIN_INTERVAL_SECONDS = 0.2
RAKUTEN_ITEM_DELETE_MIN_INTERVAL_SECONDS = 0.35
RAKUTEN_INVENTORY_BULK_UPSERT_LIMIT = 400
BATCH_TASK_PRODUCT_LIMIT = 50
LISTED_STORE_NONE_FILTER = "__none__"
_TASK_DETAIL_UNSET = object()
_STORE_SNAPSHOT_UNSET = object()
LOCAL_PRODUCT_IMAGE_URL_PREFIX = PRODUCT_IMAGE_URL_PREFIX
LOCAL_PRODUCT_IMAGE_DIR = settings.backend_dir / "data" / "product-images"
LOCAL_PRODUCT_IMAGE_DRAFT_URL_PREFIX = DRAFT_IMAGE_URL_PREFIX
LOCAL_PRODUCT_IMAGE_DRAFT_DIR = settings.backend_dir / "data" / "product-image-drafts"
RAKUTEN_ATTRIBUTE_RULES_PATH = settings.backend_dir / "app" / "resources" / "rakuten_attribute_rules.json"
RAKUTEN_GENRE_ZH_MAP_PATH = settings.backend_dir / "app" / "resources" / "rakuten_genre_zh_map.json"
ALLOWED_PRODUCT_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif"}
ALLOWED_PRODUCT_IMAGE_MIME_TYPES = {"image/jpeg", "image/png", "image/gif"}
MAX_PRODUCT_IMAGE_BYTES = 2 * 1024 * 1024
MAX_PRODUCT_IMAGE_DOWNLOAD_BYTES = 20 * 1024 * 1024
RAKUTEN_CABINET_MAX_IMAGE_BYTES = MAX_PRODUCT_IMAGE_BYTES
RAKUTEN_CABINET_MAX_IMAGE_DIMENSION = 3840
RAKUTEN_LISTING_IMAGE_LIMIT = 20
LISTING_PREPARATION_CACHE_KEY = "ltRakutenListingPreparation"
LISTING_PREPARATION_CACHE_VERSION = 1
RAKUTEN_SP_DESCRIPTION_IMAGE_LIMIT = 20
RAKUTEN_SP_DESCRIPTION_ALLOWED_TAGS = {
    "a",
    "b",
    "br",
    "center",
    "font",
    "img",
    "p",
    "table",
    "td",
    "th",
    "tr",
}
RAKUTEN_SP_DESCRIPTION_ALLOWED_ATTRIBUTES = {
    "*": {"align"},
    "a": {"href"},
    "font": {"color", "size"},
    "img": {"alt", "border", "height", "src", "width"},
    "table": {"bgcolor", "border", "cellpadding", "cellspacing", "height", "width"},
    "td": {"align", "bgcolor", "colspan", "height", "rowspan", "valign", "width"},
    "th": {"align", "bgcolor", "colspan", "height", "rowspan", "valign", "width"},
    "tr": {"align", "bgcolor", "valign"},
}
RAKUTEN_SP_DESCRIPTION_DROP_TAGS = {
    "audio",
    "button",
    "canvas",
    "embed",
    "form",
    "iframe",
    "input",
    "link",
    "map",
    "meta",
    "object",
    "script",
    "select",
    "source",
    "style",
    "svg",
    "textarea",
    "video",
}
RAKUTEN_CABINET_FOLDER_PAGE_SIZE = 100
RAKUTEN_CABINET_BATCH_FOLDER_IMAGE_LIMIT = 500
RAKUTEN_CABINET_FOLDER_CREATE_ATTEMPTS = 10
RAKUTEN_CABINET_FOLDER_LOCK_TIMEOUT_SECONDS = 120
RAKUTEN_CABINET_FOLDER_VISIBILITY_DELAYS_SECONDS = (1.0, 2.0, 3.0)
RAKUTEN_CABINET_QPS_COOLDOWN_SECONDS = 5 * 60
RAKUTEN_CABINET_REQUEST_MAX_RETRIES = 6
RAKUTEN_CABINET_QPS_BACKOFF_SECONDS = (1.5, 3.0, 5.0, 8.0, 13.0)
RAKUTEN_CABINET_RETRY_STATUS_CODES = {408, 425, 429, 500, 502, 503, 504}
RAKUTEN_CABINET_UPSTREAM_ERROR_MARKERS = (
    "upstream connect error",
    "disconnect/reset before headers",
    "connection termination",
)
DEFAULT_PAGE_SIZE = 30
MAX_PAGE_SIZE = 500
IGNORED_CABINET_IMAGE_FILENAMES = {"bg_logo.gif", "bg_logo2.gif", "bg_logo3.gif", "spacer.gif", "blank.gif"}
RAKUTEN_ATTRIBUTE_PLACEHOLDER_VALUES = {"-", "ー", "－", "―", "なし", "無し", "無", "不明", "n/a", "N/A", "na", "NA"}
RAKUTEN_ATTRIBUTE_ALLOW_PLACEHOLDER_NAMES = {"ブランド名", "シリーズ名", "メーカー型番", "原産国／製造国"}
RAKUTEN_ATTRIBUTE_TEXT_ONLY_NAMES = {"ブランド名", "シリーズ名", "メーカー型番"}
RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES = {
    "ブランド名": "-",
    "シリーズ名": "-",
    "メーカー型番": "-",
    "原産国／製造国": "-",
    "総本数": "1",
    "単品容量": "1",
}
RAKUTEN_ATTRIBUTE_DEFAULT_UNITS = {
    "本体横幅": "cm",
    "本体縦幅": "cm",
    "本体高さ": "cm",
    "本体奥行": "cm",
    "着丈": "cm",
    "身丈": "cm",
    "総丈": "cm",
    "袖丈": "cm",
    "裄丈": "cm",
    "肩幅": "cm",
    "身幅": "cm",
    "胸囲": "cm",
    "バスト": "cm",
    "ウエスト": "cm",
    "ヒップ": "cm",
    "股上": "cm",
    "股下": "cm",
    "わたり幅": "cm",
    "渡り幅": "cm",
    "裾幅": "cm",
    "もも周り": "cm",
    "ボール径": "mm",
    "芯径": "mm",
    "線幅": "mm",
    "印鑑の直径": "mm",
    "印面サイズ": "mm",
    "印面直径": "mm",
    "直径": "mm",
    "腕時計のベルトの幅": "mm",
    "腕時計のフェイスの縦のサイズ": "mm",
    "腕時計のフェイスの横のサイズ": "mm",
    "腕時計のケースの厚み": "mm",
    "腕時計のケースの直径": "mm",
    "スマートウォッチの画面サイズ": "インチ",
    "バッテリー容量": "mAh",
    "単品容量": "ml",
}
RAKUTEN_ATTRIBUTE_UNIT_ALIASES = {
    "mm": "mm",
    "ミリ": "mm",
    "ミリメートル": "mm",
    "cm": "cm",
    "㎝": "cm",
    "センチ": "cm",
    "センチメートル": "cm",
    "m": "m",
    "メートル": "m",
    "g": "g",
    "グラム": "g",
    "kg": "kg",
    "キログラム": "kg",
    "ml": "ml",
    "ミリリットル": "ml",
    "l": "L",
    "リットル": "L",
    "mah": "mAh",
    "ミリアンペア時": "mAh",
    "インチ": "インチ",
    "inch": "インチ",
    "in": "インチ",
}
RAKUTEN_BRAND_INFERENCE_PATTERNS = (
    (r"\bCASIO\b|カシオ|G-SHOCK|ジーショック|BABY-G", "CASIO"),
    (r"\bSEIKO\b|セイコー", "SEIKO"),
    (r"\bCITIZEN\b|シチズン", "CITIZEN"),
    (r"\bHAMILTON\b|ハミルトン", "HAMILTON"),
    (r"\bSUUNTO\b|スント", "SUUNTO"),
    (r"\bAMAZFIT\b|アマズフィット", "Amazfit"),
    (r"\bTIMEX\b|タイメックス", "TIMEX"),
    (r"\bORIENT\b|オリエント", "ORIENT"),
    (r"\bPUMA\b|プーマ", "PUMA"),
)
RAKUTEN_SERIES_INFERENCE_PATTERNS = (
    (r"G-SHOCK|ジーショック", "G-SHOCK"),
    (r"BABY-G", "BABY-G"),
    (r"PRO TREK|プロトレック", "PRO TREK"),
    (r"EDIFICE|エディフィス", "EDIFICE"),
    (r"OCEANUS|オシアナス", "OCEANUS"),
    (r"ASTRON|アストロン", "ASTRON"),
    (r"PROMASTER|プロマスター", "PROMASTER"),
)
RAKUTEN_ORIGIN_INFERENCE_PATTERNS = (
    (r"日本製|MADE\s+IN\s+JAPAN|JAPAN\s+MADE|原産国[:：]?\s*日本|製造国[:：]?\s*日本", "日本"),
    (r"中国製|MADE\s+IN\s+CHINA|CHINA\s+MADE|原産国[:：]?\s*中国|製造国[:：]?\s*中国", "中国"),
    (r"タイ製|MADE\s+IN\s+THAILAND|原産国[:：]?\s*タイ|製造国[:：]?\s*タイ", "タイ"),
    (r"ベトナム製|MADE\s+IN\s+VIETNAM|原産国[:：]?\s*ベトナム|製造国[:：]?\s*ベトナム", "ベトナム"),
    (r"フィリピン製|MADE\s+IN\s+PHILIPPINES|原産国[:：]?\s*フィリピン|製造国[:：]?\s*フィリピン", "フィリピン"),
    (r"インドネシア製|MADE\s+IN\s+INDONESIA|原産国[:：]?\s*インドネシア|製造国[:：]?\s*インドネシア", "インドネシア"),
    (r"韓国製|MADE\s+IN\s+KOREA|原産国[:：]?\s*韓国|製造国[:：]?\s*韓国", "韓国"),
    (r"台湾製|MADE\s+IN\s+TAIWAN|原産国[:：]?\s*台湾|製造国[:：]?\s*台湾", "台湾"),
)
RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE = "代表カラー"
RAKUTEN_REPRESENTATIVE_COLOR_FALLBACK = "マルチカラー"
RAKUTEN_REPRESENTATIVE_COLOR_ALLOWED_VALUES = {
    "ブラック",
    "ホワイト",
    "グレー",
    "シルバー",
    "ゴールド",
    "レッド",
    "ピンク",
    "オレンジ",
    "イエロー",
    "グリーン",
    "ブルー",
    "ネイビー",
    "パープル",
    "ブラウン",
    "ベージュ",
    "カーキグリーン",
    "ワインレッド",
    "透明",
    "マルチカラー",
}
RAKUTEN_CHEST_HEIGHT_LOW_VALUE = "ロー（～ 99cm）"
RAKUTEN_CHEST_HEIGHT_HIGH_VALUE = "ハイ（100cm ～）"
SINGLE_PRODUCT_VARIANT_ID = "default"
LISTING_MANAGE_NUMBER_PREFIX = "fashiongoods"
RAKUTEN_COLOR_SELECTOR_KEYWORDS = ("カラー", "色", "color", "colour")
RAKUTEN_COLOR_VALUE_MAP = {
    "ブラック": "ブラック",
    "黒": "ブラック",
    "くろ": "ブラック",
    "クロ": "ブラック",
    "black": "ブラック",
    "ホワイト": "ホワイト",
    "白": "ホワイト",
    "しろ": "ホワイト",
    "white": "ホワイト",
    "グレー": "グレー",
    "灰": "グレー",
    "gray": "グレー",
    "grey": "グレー",
    "グレイ": "グレー",
    "シルバー": "シルバー",
    "銀": "シルバー",
    "silver": "シルバー",
    "ゴールド": "ゴールド",
    "金": "ゴールド",
    "gold": "ゴールド",
    "レッド": "レッド",
    "赤": "レッド",
    "red": "レッド",
    "ピンク": "ピンク",
    "桃": "ピンク",
    "pink": "ピンク",
    "オレンジ": "オレンジ",
    "orange": "オレンジ",
    "イエロー": "イエロー",
    "黄": "イエロー",
    "yellow": "イエロー",
    "カーキグリーン": "カーキグリーン",
    "カーキ": "カーキグリーン",
    "khaki": "カーキグリーン",
    "グリーン": "グリーン",
    "緑": "グリーン",
    "green": "グリーン",
    "ブルー": "ブルー",
    "青": "ブルー",
    "blue": "ブルー",
    "ネイビー": "ネイビー",
    "紺": "ネイビー",
    "navy": "ネイビー",
    "パープル": "パープル",
    "紫": "パープル",
    "purple": "パープル",
    "ワインレッド": "ワインレッド",
    "ボルドー": "ワインレッド",
    "バーガンディ": "ワインレッド",
    "wine": "ワインレッド",
    "bordeaux": "ワインレッド",
    "burgundy": "ワインレッド",
    "ブラウン": "ブラウン",
    "茶": "ブラウン",
    "brown": "ブラウン",
    "ベージュ": "ベージュ",
    "beige": "ベージュ",
    "クリア": "透明",
    "透明": "透明",
    "無色": "透明",
    "clear": "透明",
    "アイボリー": "ホワイト",
    "ivory": "ホワイト",
    "マルチ": "マルチカラー",
    "ミックス": "マルチカラー",
    "複数色": "マルチカラー",
    "多色": "マルチカラー",
    "multi": "マルチカラー",
    "mix": "マルチカラー",
}
RAKUTEN_MACHINE_DEPENDENT_TRANSLATION = str.maketrans(
    {
        "\u301c": "\uff5e",
        "\u2212": "-",
        "\u2010": "-",
        "\u2011": "-",
        "\u2012": "-",
        "\u2013": "-",
        "\u2014": "-",
        "\u2015": "-",
        "\u00a0": " ",
        "\u200b": "",
        "\u200c": "",
        "\u200d": "",
        "\ufeff": "",
        "\u2460": "(1)",
        "\u2461": "(2)",
        "\u2462": "(3)",
        "\u2463": "(4)",
        "\u2464": "(5)",
        "\u2465": "(6)",
        "\u2466": "(7)",
        "\u2467": "(8)",
        "\u2468": "(9)",
        "\u2469": "(10)",
        "\u2474": "(1)",
        "\u2475": "(2)",
        "\u2476": "(3)",
        "\u2477": "(4)",
        "\u2478": "(5)",
        "\u2479": "(6)",
        "\u247a": "(7)",
        "\u247b": "(8)",
        "\u247c": "(9)",
        "\u247d": "(10)",
        "\u3231": "(株)",
        "\u3232": "(有)",
        "\u337f": "株式会社",
    }
)
RAKUTEN_PRODUCT_TARGET_ERROR = "单个商品采集支持普通乐天商品链接、Rakuten Fashion 商品链接、带参数链接、店铺编码/商品编号。"
RAKUTEN_SHOP_TARGET_ERROR = "店铺采集请输入店铺url代码、店铺url或sid，不能只填写店铺展示名称。"
WHOLE_SHOP_TARGET_ERROR = "整店采集请输入店铺url代码、店铺url、商品url或sid，不能只填写店铺展示名称。"
RAKUTEN_FASHION_IMAGE_BASE = "https://tshop.r10s.jp/stylife/cabinet/item"
CRAWLER_HTTP_RETRY_STATUS_CODES = {403, 408, 429, 500, 502, 503, 504}
SCHEDULE_RUN_LOCK = threading.Lock()
AUTO_LISTING_SCHEDULE_LOCK = threading.Lock()
SALES_ORDER_SYNC_RUN_LOCK = threading.Lock()
STORE_PRODUCT_SYNC_SCHEDULE_LOCK = threading.Lock()
SCHEDULED_CRAWL_TASK_CLEANUP_LOCK = threading.Lock()
STORE_UNLISTED_PRODUCT_CLEANUP_LOCK = threading.Lock()
DELETED_PRODUCT_IMAGE_CLEANUP_LOCK = threading.Lock()
CRAWLER_REQUEST_LOCK = threading.Lock()
RAKUTEN_CABINET_REQUEST_LOCK = threading.Lock()
RAKUTEN_CABINET_FOLDER_CREATE_LOCK = threading.Lock()
RAKUTEN_WRITE_REQUEST_LOCK = threading.Lock()
RAKUTEN_ITEM_DELETE_REQUEST_LOCK = threading.Lock()
CRAWLER_SESSION_LOCAL = threading.local()
CRAWLER_LAST_REQUEST_AT = 0.0
RAKUTEN_CABINET_LAST_REQUEST_AT = 0.0
RAKUTEN_WRITE_LAST_REQUEST_AT_BY_ACCOUNT: dict[str, float] = {}
RAKUTEN_ITEM_DELETE_LAST_REQUEST_AT = 0.0
SCHEDULE_RUNNER_STARTED = False
SCHEDULE_RUNNER_HEALTH_LOCK = threading.Lock()
SCHEDULE_RUNNER_HEALTH: dict[str, Any] = {
    "lastTickAt": None,
    "lastSuccessfulTickAt": None,
    "lastError": "",
    "consecutiveFailures": 0,
}
SCHEDULE_RUNNER_ACTIVE = False
SCHEDULE_RUNNER_REDIS_LOCK_NAME = "lt:schedule-runner:tick"
DRAFT_IMAGE_CLEANUP_LAST_RUN_AT = 0.0
ORPHAN_IMAGE_CLEANUP_LAST_RUN_AT = 0.0
TASK_CANCEL_REQUESTED_MARKER = "__LT_CANCEL_REQUESTED__"
TASK_CANCEL_REQUESTED_MESSAGE = "已请求终止，等待任务停止"
TASK_CANCELLED_MESSAGE = "任务已终止"
TASK_START_RETRY_DELAY_SECONDS = 5.0
TASK_STALE_CANCEL_REQUEST_SECONDS = 10 * 60
TASK_REDIS_MISSING_JOB_GRACE_SECONDS = 60
TASK_QUEUED_REDIS_MISSING_JOB_GRACE_SECONDS = 2 * 60
TITLE_OPTIMIZATION_TASK_TYPES = frozenset({"title_optimization"})
IMAGE_CLEANUP_TASK_TYPES = frozenset({"deleted_product_image_cleanup"})
LISTING_IMAGE_UPLOAD_TASK_TYPES = frozenset({"listing_image_upload"})
SPECIALIZED_SYNC_TASK_TYPES = (
    TITLE_OPTIMIZATION_TASK_TYPES
    | IMAGE_CLEANUP_TASK_TYPES
    | LISTING_IMAGE_UPLOAD_TASK_TYPES
)
PRODUCT_SCOPED_SYNC_TASK_TYPES = frozenset(
    {
        "product_delete",
        "product_listing_status",
        "product_replace",
        "title_optimization",
        "listing_image_upload",
    }
)
CRAWL_DISPATCH_LOCK_NAME = "lt:crawl-dispatch"
CRAWL_DISPATCH_LOCK_TIMEOUT_SECONDS = 30
CRAWL_DISPATCH_LOCK_BLOCKING_TIMEOUT_SECONDS = 3
SCHEDULE_IMPORT_TEMPLATE_FILENAME = "scheduled-crawl-template.xlsx"
SCHEDULE_EXPORT_FILENAME = "scheduled-crawl-shops.xlsx"
MANUAL_CRAWL_IMPORT_TEMPLATE_FILENAME = "manual-crawl-import-template.xlsx"
SCHEDULE_FALLBACK_SHOP_URL_KEY = "fallbackShopUrl"
SCHEDULE_IMPORTED_NOTE_KEY = "importedSchedule"
SCHEDULE_FALLBACK_TARGET_PREFIX = "__LT_FALLBACK_SHOP_URL__:"
SCHEDULED_CRAWL_TASK_CLEANUP_SETTING_KEY = "scheduledCrawlTaskCleanup"
USER_TIME_SETTINGS_PREFIX = "userTimeSettings:"
DELETED_PRODUCT_IMAGE_CLEANUP_USER_SETTING_PREFIX = "deletedProductImageCleanup:"
USER_TIME_SETTINGS_OWNER_SCOPE_MIGRATION_KEY = "migration:userTimeSettingsOwnerScopeV2"
SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_WEEKDAY = 6
SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_TIME = "09:00"
SCHEDULED_CRAWL_TASK_CLEANUP_RETENTION_DAYS = 7
STORE_PRODUCT_SYNC_DEFAULT_WEEKDAY = 6
STORE_PRODUCT_SYNC_DEFAULT_TIME = "21:00"
STORE_UNLISTED_PRODUCT_CLEANUP_MONTH_DAY = 1
STORE_UNLISTED_PRODUCT_CLEANUP_TIME = "01:00"
DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY = 5
DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME = "09:00"
SALES_ORDER_SYNC_INTERVAL = timedelta(minutes=30)
SALES_ORDER_SYNC_FAILURE_COOLDOWN = timedelta(minutes=30)
SALES_ORDER_SYNC_BATCH_SIZE = 20
SALES_ORDER_SYNC_LOCAL_MAX_PENDING = 20
SALES_ORDER_SYNC_TASK_PREFIX = "order-"
SALES_ORDER_SYNC_EXECUTOR = ThreadPoolExecutor(
    max_workers=max(
        1,
        min(4, int(settings.max_running_sync_tasks_per_user)),
    ),
    thread_name_prefix="lt-order-sync",
)
LISTING_IMAGE_PREPARE_EXECUTOR = ThreadPoolExecutor(
    max_workers=max(1, int(settings.listing_image_prepare_workers)),
    thread_name_prefix="listing-image-prepare",
)
SALES_ORDER_SYNC_LOCAL_SLOTS = threading.BoundedSemaphore(
    SALES_ORDER_SYNC_LOCAL_MAX_PENDING
)
sales_sync_service: Any | None = None


class ProductImageUnavailableError(RuntimeError):
    def __init__(self, status_code: int) -> None:
        self.status_code = int(status_code)
        super().__init__(f"图片不存在或已失效（HTTP {self.status_code}）。")


class TaskCancelled(RuntimeError):
    pass


def start_background_task(target: Callable[..., Any], *args: Any, delay_seconds: float = 0.0) -> None:
    if delay_seconds > 0:
        timer = threading.Timer(delay_seconds, target, args=args)
        timer.daemon = True
        timer.start()
        return
    worker = threading.Thread(target=target, args=args, daemon=True)
    worker.start()


def normalize_crawl_task_mode(value: Any) -> str:
    return "scheduled" if normalize_text(value).lower() == "scheduled" else "manual"


def crawl_task_mode_filter(mode: str) -> Any:
    if normalize_crawl_task_mode(mode) == "scheduled":
        return CrawlTaskModel.mode == "scheduled"
    return or_(CrawlTaskModel.mode.is_(None), CrawlTaskModel.mode != "scheduled")


def crawl_task_queue_kind(mode: str | None) -> str:
    return "scheduled-crawl" if normalize_crawl_task_mode(mode) == "scheduled" else "manual-crawl"


def crawl_task_concurrency_limit(mode: str | None) -> int:
    if normalize_crawl_task_mode(mode) == "scheduled":
        return max(1, int(settings.max_running_scheduled_crawl_tasks_per_user))
    return max(1, int(settings.max_running_manual_crawl_tasks_per_user))


def crawl_task_queue_kind_for_id(task_id: str) -> str:
    with session_scope() as session:
        mode = session.scalar(
            select(CrawlTaskModel.mode).where(CrawlTaskModel.id == task_id)
        )
    return crawl_task_queue_kind(mode)


def crawl_task_owner_for_id(task_id: str) -> str:
    with session_scope() as session:
        return str(
            session.scalar(
                select(CrawlTaskModel.owner_username).where(
                    CrawlTaskModel.id == task_id
                )
            )
            or ""
        )


def dispatch_crawl_task(
    task_id: str,
    *,
    delay_seconds: float = 0.0,
    job_id: str | None = None,
    queue_kind: str | None = None,
    mark_failed_on_error: bool = True,
) -> None:
    owner_username = crawl_task_owner_for_id(task_id)
    if should_use_redis_task_queue():
        try:
            resolved_queue_kind = queue_kind or crawl_task_queue_kind_for_id(task_id)
            enqueue = enqueue_task_in if delay_seconds > 0 else enqueue_task
            enqueue_args = (
                (delay_seconds, run_task, task_id, job_id)
                if delay_seconds > 0
                else (run_task, task_id, job_id)
            )
            enqueue(
                *enqueue_args,
                job_id=job_id,
                description=f"采集任务 {task_id}",
                queue_name=task_queue_name_for_kind(resolved_queue_kind),
                owner_username=owner_username,
            )
        except Exception as exc:
            if mark_failed_on_error:
                mark_background_task_dispatch_failed(CrawlTaskModel, task_id, exc)
            raise
        return
    start_background_task(run_task, task_id, delay_seconds=delay_seconds)


def crawl_dispatch_job_id(task_id: str) -> str:
    return f"crawl-{task_id}-{uuid.uuid4().hex[:8]}"


def crawl_dispatch_available_slots(
    running_count: int,
    reserved_count: int,
    limit: int | None = None,
) -> int:
    capacity = max(1, int(limit or settings.max_running_crawl_tasks_per_user))
    return max(0, capacity - max(0, int(running_count or 0)) - max(0, int(reserved_count or 0)))


def reserve_queued_crawl_tasks(
    session: Any,
    owner_username: str | None = None,
) -> list[tuple[str, str, str]]:
    paused_owners = system_task_paused_usernames()
    reservations: list[tuple[str, str, str]] = []
    for task_mode in ("manual", "scheduled"):
        mode_filter = crawl_task_mode_filter(task_mode)
        owner_query = (
            select(
                CrawlTaskModel.owner_username,
                func.min(CrawlTaskModel.created_at).label("oldest_created_at"),
            )
            .where(
                CrawlTaskModel.status == "queued",
                mode_filter,
            )
            .group_by(CrawlTaskModel.owner_username)
            .order_by(func.min(CrawlTaskModel.created_at).asc(), CrawlTaskModel.owner_username.asc())
        )
        if owner_username:
            owner_query = owner_query.where(CrawlTaskModel.owner_username == owner_username)
        owners = [str(row.owner_username) for row in session.execute(owner_query)]
        for task_owner in owners:
            if task_owner in paused_owners:
                continue
            running_count = int(
                session.scalar(
                    select(func.count()).where(
                        CrawlTaskModel.owner_username == task_owner,
                        CrawlTaskModel.status == "running",
                        mode_filter,
                    )
                )
                or 0
            )
            reserved_count = int(
                session.scalar(
                    select(func.count()).where(
                        CrawlTaskModel.owner_username == task_owner,
                        CrawlTaskModel.status == "queued",
                        CrawlTaskModel.queue_job_id.is_not(None),
                        mode_filter,
                    )
                )
                or 0
            )
            available_slots = crawl_dispatch_available_slots(
                running_count,
                reserved_count,
                crawl_task_concurrency_limit(task_mode),
            )
            if available_slots <= 0:
                continue
            rows = session.scalars(
                select(CrawlTaskModel)
                .where(
                    CrawlTaskModel.owner_username == task_owner,
                    CrawlTaskModel.status == "queued",
                    CrawlTaskModel.queue_job_id.is_(None),
                    mode_filter,
                )
                .order_by(CrawlTaskModel.created_at.asc(), CrawlTaskModel.id.asc())
                .limit(available_slots)
                .with_for_update()
            ).all()
            queue_kind = crawl_task_queue_kind(task_mode)
            for task in rows:
                job_id = crawl_dispatch_job_id(str(task.id))
                task.queue_job_id = job_id
                task.message = (
                    "已进入定时采集队列，等待 Worker"
                    if task_mode == "scheduled"
                    else "已进入手动采集队列，等待 Worker"
                )
                reservations.append((str(task.id), job_id, queue_kind))
    return reservations


def dispatch_queued_crawl_tasks(owner_username: str | None = None) -> int:
    if owner_username and system_task_dispatch_paused(owner_username):
        return 0
    if not should_use_redis_task_queue():
        return 0
    connection = redis_connection()
    lock = connection.lock(
        CRAWL_DISPATCH_LOCK_NAME,
        timeout=CRAWL_DISPATCH_LOCK_TIMEOUT_SECONDS,
    )
    acquired = lock.acquire(
        blocking=True,
        blocking_timeout=CRAWL_DISPATCH_LOCK_BLOCKING_TIMEOUT_SECONDS,
    )
    if not acquired:
        return 0
    try:
        with session_scope() as session:
            reservations = reserve_queued_crawl_tasks(session, owner_username)
        dispatched_count = 0
        for task_id, job_id, queue_kind in reservations:
            try:
                dispatch_crawl_task(
                    task_id,
                    job_id=job_id,
                    queue_kind=queue_kind,
                    mark_failed_on_error=False,
                )
                dispatched_count += 1
            except Exception as exc:
                logger.warning("采集任务 %s 投递失败，等待系统重试：%s", task_id, exc)
                with session_scope() as session:
                    task = session.get(CrawlTaskModel, task_id)
                    if task is None or task.status != "queued" or task.queue_job_id != job_id:
                        continue
                    task.queue_job_id = None
                    task.message = "采集队列投递失败，等待系统重试"
        return dispatched_count
    finally:
        try:
            lock.release()
        except Exception:
            logger.warning("释放采集派发锁失败", exc_info=True)


def dispatch_queued_crawl_tasks_safely(owner_username: str | None = None) -> int:
    try:
        return dispatch_queued_crawl_tasks(owner_username)
    except Exception:
        logger.warning("采集队列容量派发失败", exc_info=True)
        return 0


def sync_task_queue_kind(task_type: str | None) -> str:
    normalized_type = normalize_text(task_type)
    if normalized_type in TITLE_OPTIMIZATION_TASK_TYPES:
        return "title-optimization"
    if normalized_type in IMAGE_CLEANUP_TASK_TYPES:
        return "image-cleanup"
    if normalized_type in LISTING_IMAGE_UPLOAD_TASK_TYPES:
        return "listing"
    return "sync"


def sync_task_type_label(task_type: str | None) -> str:
    normalized_type = normalize_text(task_type)
    if normalized_type in TITLE_OPTIMIZATION_TASK_TYPES:
        return "标题优化任务"
    if normalized_type in IMAGE_CLEANUP_TASK_TYPES:
        return "图片清理任务"
    if normalized_type in LISTING_IMAGE_UPLOAD_TASK_TYPES:
        return "图片上传任务"
    return "同步任务"


def sync_task_type_for_id(task_id: str) -> str:
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        return normalize_text(task.task_type) if task is not None else ""


def dispatch_sync_task(
    owner_username: str,
    task_id: str,
    *,
    task_type: str | None = None,
    delay_seconds: float = 0.0,
) -> None:
    resolved_task_type = normalize_text(task_type) or sync_task_type_for_id(task_id)
    queue_kind = sync_task_queue_kind(resolved_task_type)
    if should_use_redis_task_queue():
        try:
            enqueue = enqueue_task_in if delay_seconds > 0 else enqueue_task
            enqueue_args = (delay_seconds, run_sync_task, owner_username, task_id) if delay_seconds > 0 else (run_sync_task, owner_username, task_id)
            enqueue(
                *enqueue_args,
                description=f"{sync_task_type_label(resolved_task_type)} {task_id}",
                queue_name=task_queue_name_for_kind(queue_kind),
                owner_username=owner_username,
            )
        except Exception as exc:
            mark_background_task_dispatch_failed(SyncTaskModel, task_id, exc)
            raise
        return
    start_background_task(run_sync_task, owner_username, task_id, delay_seconds=delay_seconds)


def sync_task_has_active_background_job(task_id: str, task_type: str | None = None) -> bool:
    if not should_use_redis_task_queue():
        return False
    try:
        queue_kind = sync_task_queue_kind(
            normalize_text(task_type) or sync_task_type_for_id(task_id)
        )
        state = redis_task_states(
            {str(task_id)},
            queue_kind,
            job_id_prefixes=task_model_job_id_prefixes(SyncTaskModel),
        ).get(str(task_id))
    except Exception:
        return False
    return state is not None and state.get("status") in {"queued", "started", "deferred", "scheduled"}


def dispatch_next_sync_task() -> None:
    next_tasks: list[tuple[str, str, str]] = []
    with session_scope() as session:
        finalize_stale_cancel_requested_tasks(session, SyncTaskModel, action_label="同步")
        reconcile_interrupted_running_tasks(session, SyncTaskModel)
        running_rows = session.scalars(
            select(SyncTaskModel).where(
                SyncTaskModel.status == "running",
                SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
            )
        ).all()
        queued_rows = session.scalars(
            select(SyncTaskModel)
            .where(
                SyncTaskModel.status == "queued",
                SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
            )
            .order_by(SyncTaskModel.created_at.asc(), SyncTaskModel.id.asc())
        ).all()
        active_job_ids = {
            str(task.id)
            for task in queued_rows
            if sync_task_has_active_background_job(task.id, task.task_type)
        }
        active_rows_by_id = {str(task.id): task for task in running_rows}
        active_rows_by_id.update(
            {
                str(task.id): task
                for task in queued_rows
                if str(task.id) in active_job_ids
            }
        )
        active_listing_rows = session.scalars(
            select(ListingTaskModel).where(
                ListingTaskModel.status.in_(("queued", "running"))
            )
        ).all()
        active_scopes = [
            sync_task_resource_scope(session, task)
            for task in active_rows_by_id.values()
        ]
        active_scopes.extend(
            listing_task_resource_scope(session, task)
            for task in active_listing_rows
            if task.status == "running" or listing_task_has_active_background_job(task.id)
        )
        owner_counts: dict[str, int] = {}
        for task in active_rows_by_id.values():
            owner_counts[task.owner_username] = owner_counts.get(task.owner_username, 0) + 1
        available_slots = max(
            0,
            int(settings.max_running_sync_tasks_global) - len(active_rows_by_id),
        )
        for task in queued_rows:
            if available_slots <= 0:
                break
            if system_task_dispatch_paused(task.owner_username):
                continue
            if task_cancel_requested(task):
                task.status = "cancelled"
                task.message = TASK_CANCELLED_MESSAGE
                task.error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
                task.finished_at = datetime.now()
                continue
            if str(task.id) in active_job_ids:
                continue
            if owner_counts.get(task.owner_username, 0) >= int(settings.max_running_sync_tasks_per_user):
                task.message = "排队中，等待该用户当前同步任务释放并发额度"
                continue
            scope = sync_task_resource_scope(session, task)
            wait_reason = task_scope_conflict_reason(scope, active_scopes)
            if wait_reason:
                task.message = wait_reason
                continue
            next_tasks.append(
                (task.owner_username, task.id, task.task_type or "store_sync")
            )
            owner_counts[task.owner_username] = owner_counts.get(task.owner_username, 0) + 1
            active_scopes.append(scope)
            available_slots -= 1
    for owner_username, task_id, task_type in next_tasks:
        dispatch_sync_task(owner_username, task_id, task_type=task_type)


def dispatch_next_sync_task_safely() -> None:
    try:
        dispatch_next_sync_task()
    except Exception:
        return


def dispatch_listing_task(owner_username: str, task_id: str, *, delay_seconds: float = 0.0) -> None:
    if should_use_redis_task_queue():
        try:
            enqueue = enqueue_task_in if delay_seconds > 0 else enqueue_task
            enqueue_args = (delay_seconds, run_listing_task, owner_username, task_id) if delay_seconds > 0 else (run_listing_task, owner_username, task_id)
            enqueue(
                *enqueue_args,
                description=f"上架任务 {task_id}",
                queue_name=task_queue_name_for_kind("listing"),
                owner_username=owner_username,
            )
        except Exception as exc:
            mark_background_task_dispatch_failed(ListingTaskModel, task_id, exc)
            raise
        return
    start_background_task(run_listing_task, owner_username, task_id, delay_seconds=delay_seconds)


def listing_task_has_active_background_job(task_id: str) -> bool:
    if not should_use_redis_task_queue():
        return False
    try:
        state = redis_task_states({str(task_id)}, "listing", job_id_prefixes=task_model_job_id_prefixes(ListingTaskModel)).get(str(task_id))
    except Exception:
        return False
    return state is not None and state.get("status") in {"queued", "started", "deferred", "scheduled"}


def listing_task_store_ids(task: ListingTaskModel) -> list[int]:
    payload = listing_task_product_ids_payload(task.product_ids_json)
    store_ids = payload["storeIds"]
    if store_ids:
        return store_ids
    return [int(task.store_id)] if task.store_id else []


def dispatch_next_listing_task() -> None:
    next_tasks: list[tuple[str, str]] = []
    with session_scope() as session:
        finalize_stale_cancel_requested_tasks(session, ListingTaskModel, action_label="上架")
        reconcile_interrupted_running_tasks(session, ListingTaskModel)
        running_rows = session.scalars(
            select(ListingTaskModel).where(ListingTaskModel.status == "running")
        ).all()
        queued_rows = session.scalars(
            select(ListingTaskModel)
            .where(ListingTaskModel.status == "queued")
            .order_by(ListingTaskModel.created_at.asc(), ListingTaskModel.id.asc())
        ).all()
        active_job_ids: set[str] = set()
        if queued_rows and should_use_redis_task_queue():
            try:
                states = redis_task_states(
                    {str(task.id) for task in queued_rows},
                    "listing",
                    job_id_prefixes=task_model_job_id_prefixes(ListingTaskModel),
                )
                active_job_ids = {
                    task_id
                    for task_id, state in states.items()
                    if state.get("status") in {"queued", "started", "deferred", "scheduled"}
                }
            except Exception:
                active_job_ids = {
                    str(task.id)
                    for task in queued_rows
                    if listing_task_has_active_background_job(task.id)
                }

        active_rows_by_id = {str(task.id): task for task in running_rows}
        active_rows_by_id.update({
            str(task.id): task
            for task in queued_rows
            if str(task.id) in active_job_ids
        })
        owner_counts: dict[str, int] = {}
        store_counts: dict[int, int] = {}
        active_scopes: list[TaskResourceScope] = []
        for task in active_rows_by_id.values():
            owner_counts[task.owner_username] = owner_counts.get(task.owner_username, 0) + 1
            for store_id in listing_task_store_ids(task):
                store_counts[store_id] = store_counts.get(store_id, 0) + 1
            active_scopes.append(listing_task_resource_scope(session, task))
        active_sync_rows = session.scalars(
            select(SyncTaskModel).where(
                SyncTaskModel.status.in_(("queued", "running")),
                SyncTaskModel.task_type.notin_(IMAGE_CLEANUP_TASK_TYPES),
            )
        ).all()
        active_scopes.extend(
            sync_task_resource_scope(session, task)
            for task in active_sync_rows
            if task.status == "running"
            or sync_task_has_active_background_job(task.id, task.task_type)
        )

        available_slots = max(
            0,
            int(settings.max_running_listing_tasks_global) - len(active_rows_by_id),
        )
        for task in queued_rows:
            if available_slots <= 0:
                break
            if system_task_dispatch_paused(task.owner_username):
                continue
            if task_cancel_requested(task):
                release_listing_task_locks(session, task.owner_username, task)
                task.status = "cancelled"
                task.message = TASK_CANCELLED_MESSAGE
                task.error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
                task.finished_at = datetime.now()
                continue
            if str(task.id) in active_job_ids:
                continue
            if owner_counts.get(task.owner_username, 0) >= int(settings.max_running_listing_tasks_per_user):
                task.message = "排队中，等待该用户当前上架任务释放并发额度"
                continue
            task_store_ids = listing_task_store_ids(task)
            if any(
                store_counts.get(store_id, 0) >= int(settings.max_running_listing_tasks_per_store)
                for store_id in task_store_ids
            ):
                task.message = "排队中，等待该店铺当前上架任务释放并发额度"
                continue
            scope = listing_task_resource_scope(session, task)
            wait_reason = task_scope_conflict_reason(scope, active_scopes)
            if wait_reason:
                task.message = wait_reason
                continue
            next_tasks.append((task.owner_username, task.id))
            owner_counts[task.owner_username] = owner_counts.get(task.owner_username, 0) + 1
            for store_id in task_store_ids:
                store_counts[store_id] = store_counts.get(store_id, 0) + 1
            active_scopes.append(scope)
            available_slots -= 1
    for owner_username, task_id in next_tasks:
        dispatch_listing_task(owner_username, task_id)


def dispatch_next_listing_task_safely() -> None:
    try:
        dispatch_next_listing_task()
    except Exception:
        return


def run_auto_listing_schedule_job(
    schedule_id: int,
    owner_username: str | None = None,
    advance_next_run: bool = False,
) -> None:
    execute_auto_listing_schedule(
        schedule_id,
        owner_username=owner_username,
        advance_next_run=advance_next_run,
    )


def mark_auto_listing_schedule_dispatch_failed(schedule_id: int, exc: Exception) -> None:
    with session_scope() as session:
        row = session.get(AutoListingScheduleModel, schedule_id)
        if row is None:
            return
        row.status = "failed"
        row.last_message = "上架任务投递失败。"
        row.last_error = f"Redis 队列投递失败：{exc}"


def dispatch_auto_listing_schedule(
    schedule_id: int,
    *,
    owner_username: str | None = None,
    advance_next_run: bool = False,
) -> None:
    if should_use_redis_task_queue():
        try:
            enqueue_task(
                run_auto_listing_schedule_job,
                schedule_id,
                owner_username,
                advance_next_run,
                job_id=f"auto-listing-schedule-{schedule_id}",
                description=f"创建上架任务 {schedule_id}",
                queue_name=task_queue_name_for_kind("schedule"),
                owner_username=owner_username,
            )
        except Exception as exc:
            mark_auto_listing_schedule_dispatch_failed(schedule_id, exc)
            raise
        return
    start_background_task(
        run_auto_listing_schedule_job,
        schedule_id,
        owner_username,
        advance_next_run,
    )


def run_auto_deletion_task_job(
    task_id: int,
    owner_username: str | None = None,
    advance_next_run: bool = False,
) -> None:
    execute_auto_deletion_task(
        task_id,
        owner_username=owner_username,
        advance_next_run=advance_next_run,
    )


def mark_auto_deletion_task_dispatch_failed(task_id: int, exc: Exception) -> None:
    with session_scope() as session:
        row = session.get(AutoDeletionTaskModel, task_id)
        if row is None:
            return
        row.status = "failed"
        row.last_message = "删除任务投递失败。"
        row.last_error = f"Redis 队列投递失败：{exc}"


def dispatch_auto_deletion_task(
    task_id: int,
    *,
    owner_username: str | None = None,
    advance_next_run: bool = False,
) -> None:
    if should_use_redis_task_queue():
        try:
            enqueue_task(
                run_auto_deletion_task_job,
                task_id,
                owner_username,
                advance_next_run,
                job_id=f"auto-deletion-task-{task_id}",
                description=f"创建删除任务 {task_id}",
                queue_name=task_queue_name_for_kind("schedule"),
                owner_username=owner_username,
            )
        except Exception as exc:
            mark_auto_deletion_task_dispatch_failed(task_id, exc)
            raise
        return
    start_background_task(
        run_auto_deletion_task_job,
        task_id,
        owner_username,
        advance_next_run,
    )


def dispatch_scheduled_crawl(owner_username: str, schedule_id: int) -> None:
    if should_use_redis_task_queue():
        try:
            enqueue_task(
                run_scheduled_crawl_job,
                owner_username,
                schedule_id,
                job_id=f"schedule-{schedule_id}-{uuid.uuid4().hex[:8]}",
                description=f"定时采集 {schedule_id}",
                queue_name=task_queue_name_for_kind("schedule"),
                owner_username=owner_username,
            )
        except Exception as exc:
            mark_scheduled_crawl_dispatch_failed(schedule_id, exc)
            raise
        return
    worker = threading.Thread(target=run_scheduled_crawl_job, args=(owner_username, schedule_id), daemon=True)
    worker.start()


def should_use_redis_task_queue() -> bool:
    return settings.task_queue_mode == "redis"


def mark_background_task_dispatch_failed(model: Any, task_id: str, exc: Exception) -> None:
    with session_scope() as session:
        task = session.get(model, task_id)
        if task is None:
            return
        task.status = "failed"
        task.failed_count = max(1, int(getattr(task, "failed_count", 0) or 0))
        task.message = "任务投递失败"
        task.error_detail = f"Redis 队列投递失败：{exc}"
        task.finished_at = datetime.now()
        if model is ListingTaskModel:
            release_listing_task_locks(session, task.owner_username, task)


def mark_scheduled_crawl_dispatch_failed(schedule_id: int, exc: Exception) -> None:
    with session_scope() as session:
        row = session.get(ScheduledCrawlModel, schedule_id)
        if row is None:
            return
        row.status = "failed"
        row.notes = f"Redis 队列投递失败：{exc}"


def log_event(owner_username: str, task_id: str | None, level: str, message: str) -> None:
    with session_scope() as session:
        session.add(CrawlLogModel(owner_username=owner_username, task_id=task_id, level=level, message=message))


def normalize_page_params(page: int | None, page_size: int | None) -> tuple[int, int | None]:
    normalized_page = max(1, int(page or 1))
    normalized_page_size = min(MAX_PAGE_SIZE, max(1, int(page_size or 0))) if page_size else None
    return normalized_page, normalized_page_size


def paginate_query(
    session: Any,
    query: Any,
    *,
    order_by: Any | tuple[Any, ...],
    page: int | None,
    page_size: int | None,
    response_key: str,
    serializer: Any,
) -> list[dict[str, Any]] | dict[str, Any]:
    normalized_page, normalized_page_size = normalize_page_params(page, page_size)
    order_values = order_by if isinstance(order_by, tuple) else (order_by,)
    if not normalized_page_size:
        rows = session.scalars(query.order_by(*order_values)).all()
        return [serializer(row) for row in rows]

    total = int(session.scalar(select(func.count()).select_from(query.order_by(None).subquery())) or 0)
    if total:
        max_page = max(1, (total + normalized_page_size - 1) // normalized_page_size)
        normalized_page = min(normalized_page, max_page)
    rows = session.scalars(
        query.order_by(*order_values)
        .offset((normalized_page - 1) * normalized_page_size)
        .limit(normalized_page_size)
    ).all()
    return {
        response_key: [serializer(row) for row in rows],
        "total": total,
        "page": normalized_page,
        "pageSize": normalized_page_size,
    }


def source_to_public(row: CrawlSourceModel) -> dict[str, Any]:
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "name": row.name,
        "sourceType": row.source_type,
        "target": row.target,
        "enabled": bool(row.enabled),
        "scheduleEnabled": bool(row.schedule_enabled),
        "intervalMinutes": row.interval_minutes,
        "lastRunAt": row.last_run_at.isoformat(sep=" ") if row.last_run_at else None,
        "notes": row.notes,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


def crawl_task_storage_counts(row: CrawlTaskModel) -> tuple[int, int]:
    saved_count = max(0, int(getattr(row, "saved_count", 0) or 0))
    skipped_count = max(0, int(getattr(row, "skipped_count", 0) or 0))
    if saved_count == 0 and skipped_count == 0:
        match = re.search(r"入库\s*(\d+)\s*条，跳过\s*(\d+)\s*条", normalize_text(row.message))
        if match:
            saved_count = int(match.group(1))
            skipped_count = int(match.group(2))
    return saved_count, skipped_count


def task_stored_crawl_price_rule(row: CrawlTaskModel) -> dict[str, Any] | None:
    raw_rule = getattr(row, "crawl_price_rule_json", None)
    if not normalize_text(raw_rule):
        return None
    try:
        return normalize_crawl_price_rule(raw_rule)
    except RuntimeError:
        logger.warning("Invalid crawl price rule stored on task %s", row.id)
        return {"operator": "all"}


def task_to_public(row: CrawlTaskModel) -> dict[str, Any]:
    saved_count, skipped_count = crawl_task_storage_counts(row)
    warning_count = max(0, int(getattr(row, "warning_count", 0) or 0))
    warning_detail = task_public_warning_detail(
        row,
        skipped_count=skipped_count,
        warning_count=warning_count,
    )
    if skipped_count > 0 and warning_count == 0:
        warning_count = skipped_count
    if skipped_count > 0 and not warning_detail:
        warning_detail = (
            f"本次有 {skipped_count} 条商品未入库，可能已存在于商品管理"
            "或低于采集价格门槛。重新采集后将显示具体跳过原因。"
        )
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "sourceId": row.source_id,
        "sourceType": row.source_type,
        "target": row.target,
        "crawlPriceRule": task_stored_crawl_price_rule(row),
        "mode": row.mode,
        "status": resolve_crawl_task_status(row.status, row.total_count, row.success_count, row.failed_count),
        "totalCount": row.total_count,
        "successCount": row.success_count,
        "failedCount": row.failed_count,
        "warningCount": warning_count,
        "savedCount": saved_count,
        "skippedCount": skipped_count,
        "cancelRequested": task_cancel_requested(row),
        "message": row.message,
        "errorDetail": task_public_error_detail(row),
        "warningDetail": warning_detail,
        "startedAt": row.started_at.isoformat(sep=" ") if row.started_at else None,
        "finishedAt": row.finished_at.isoformat(sep=" ") if row.finished_at else None,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
    }


def resolve_crawl_task_status(status: str, total_count: int, success_count: int, failed_count: int) -> str:
    if status in {"queued", "running", "cancelled"}:
        return status
    total = max(0, int(total_count or 0))
    success = max(0, int(success_count or 0))
    failed = max(0, int(failed_count or 0))
    if total > 0 and failed >= total and success == 0:
        return "failed"
    if success > 0 and failed > 0:
        return "partial"
    if failed > 0 and success == 0:
        return "failed"
    if success > 0 and failed == 0:
        return "success"
    return status


STORE_PRODUCT_SALES_PERIOD_DAYS = {7, 14, 30, 60, 90, 180, 365}
STORE_PRODUCT_SALES_MAX_RANGE_DAYS = 365


def product_to_public(
    row: ProductModel,
    *,
    period_sales_count: int | None = None,
    title_optimization_count: int = 0,
    title_optimization_task_id: str | None = None,
    product_delete_task_id: str | None = None,
) -> dict[str, Any]:
    listed_at = product_listed_at_text(row)
    raw_payload = product_raw_payload(row)
    shop_code = product_shop_code(row, raw_payload)
    image_urls = product_editable_image_urls(raw_payload, shop_code=shop_code)
    if row.image_url and row.image_url not in image_urls:
        image_urls.insert(0, row.image_url)
    price_range = price_range_from_rakuten_item(raw_payload)
    fallback_price = float(row.price) if row.price is not None else None
    price_min = price_range[0] if price_range else fallback_price
    price_max = price_range[1] if price_range else fallback_price
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "taskId": row.task_id,
        "scheduledCrawlId": row.scheduled_crawl_id,
        "collectionSource": getattr(row, "collection_source", "manual") or "manual",
        "parentProductId": row.parent_product_id,
        "listingTaskId": row.listing_task_id,
        "storeId": row.store_id,
        "rakutenManageNumber": row.rakuten_manage_number,
        "storeProductStatus": row.store_product_status,
        "rakutenListingStatus": row.rakuten_listing_status,
        "listedStores": product_listed_stores(raw_payload),
        "storeLastSeenAt": row.store_last_seen_at.isoformat(sep=" ") if row.store_last_seen_at else None,
        "title": row.title,
        "tagline": product_tagline(raw_payload),
        "sourceUrl": row.source_url,
        "rakutenItemUrl": product_rakuten_item_url(row, raw_payload),
        "itemNumber": row.item_number,
        "shopName": row.shop_name,
        "imageUrl": row.image_url,
        "images": image_urls,
        "price": price_min,
        "priceMin": price_min,
        "priceMax": price_max,
        "currency": row.currency,
        "salesCount": product_sales_count(raw_payload),
        "reviewCount": int(getattr(row, "review_count", None) or 0),
        "periodSalesCount": period_sales_count,
        "titleOptimizationCount": title_optimization_count,
        "titleOptimizationTaskId": title_optimization_task_id,
        "productDeleteTaskId": product_delete_task_id,
        "genreId": row.genre_id,
        "genrePath": rakuten_genre_path(row.genre_id),
        "genrePathZh": rakuten_genre_zh_path(rakuten_genre_path(row.genre_id)),
        "reviewStatus": row.review_status,
        "replacementTaskId": normalize_text(product_replacement_metadata(raw_payload).get("taskId")) or None,
        "replacementTargetProductId": (
            int(product_replacement_metadata(raw_payload).get("targetProductId") or 0) or None
        ),
        "replacementTargetManageNumber": (
            normalize_text(product_replacement_metadata(raw_payload).get("targetManageNumber")) or None
        ),
        "replacementTargetStoreId": (
            int(product_replacement_metadata(raw_payload).get("targetStoreId") or 0) or None
        ),
        "replacementTargetStoreName": (
            normalize_text(product_replacement_metadata(raw_payload).get("targetStoreName")) or None
        ),
        "lastError": row.last_error,
        "listedAt": listed_at,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


def product_listed_stores(raw_payload: dict[str, Any]) -> list[dict[str, Any]]:
    listed_stores = raw_payload.get("listedStores") if isinstance(raw_payload.get("listedStores"), list) else []
    result: list[dict[str, Any]] = []
    seen: set[int] = set()
    for item in listed_stores:
        if not isinstance(item, dict):
            continue
        try:
            store_id = int(item.get("storeId") or 0)
        except (TypeError, ValueError):
            store_id = 0
        if not store_id or store_id in seen:
            continue
        seen.add(store_id)
        result.append(
            {
                "storeId": store_id,
                "storeCode": normalize_text(item.get("storeCode")),
                "storeName": normalize_text(item.get("storeName")),
                "aliasName": normalize_text(item.get("aliasName")),
                "manageNumber": normalize_text(item.get("manageNumber")),
                "itemNumber": normalize_text(item.get("itemNumber")),
                "productId": int(item.get("productId") or 0) if str(item.get("productId") or "").isdigit() else None,
                "listedAt": normalize_text(item.get("listedAt")),
            }
        )
    return result


def product_rakuten_item_url(row: ProductModel, raw_payload: dict[str, Any]) -> str:
    if row.review_status != "listed":
        return row.source_url
    listing_store = raw_payload.get("listingStore") if isinstance(raw_payload.get("listingStore"), dict) else {}
    shop_code = (
        first_text_from_keys(listing_store, ("storeCode", "shopCode"))
        or normalize_shop_code(row.image_url)
        or normalize_shop_code(first_text_from_keys(raw_payload, ("itemUrl", "itemPageUrl", "url")))
    )
    item_number = (
        normalize_text(row.item_number)
        or first_text_from_keys(raw_payload, ("itemNumber", "manageNumber"))
        or normalize_text(row.rakuten_manage_number)
    )
    return build_public_item_page_url(shop_code, item_number) or row.source_url


def product_raw_payload(row: ProductModel) -> dict[str, Any]:
    try:
        payload = json.loads(row.raw_payload_json or "{}")
    except ValueError:
        return {}
    return payload if isinstance(payload, dict) else {}


RAKUTEN_TAGLINE_KEYS = ("tagline", "catchcopy", "catchCopy", "catchCopyTrans", "subTitle", "subtitle", "saleComment", "sale_comment")


def product_tagline(raw_payload: dict[str, Any]) -> str:
    tagline = first_text_from_keys(raw_payload, RAKUTEN_TAGLINE_KEYS)
    if tagline:
        return tagline
    embedded_item = raw_payload.get("embeddedItem")
    if isinstance(embedded_item, dict):
        tagline = first_text_from_keys(embedded_item, RAKUTEN_TAGLINE_KEYS)
        if tagline:
            return tagline
    source_product = raw_payload.get("sourceProduct")
    if isinstance(source_product, dict):
        return first_text_from_keys(source_product, RAKUTEN_TAGLINE_KEYS)
    return ""


def parse_rakuten_datetime_value(value: Any) -> datetime | None:
    text = first_text_value(value)
    if not text:
        return None
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed


def product_rakuten_created_at(row: ProductModel) -> datetime | None:
    return row.listed_at or parse_rakuten_datetime_value(product_raw_payload(row).get("created"))


def product_listed_at_text(row: ProductModel) -> str | None:
    value = row.listed_at or product_rakuten_created_at(row)
    return value.isoformat(sep=" ", timespec="seconds") if value else None


def product_sales_count(raw_payload: dict[str, Any]) -> int | None:
    for key in ("salesCount", "salesQuantity", "soldCount", "orderCount", "sales_count", "sold_count"):
        value = raw_payload.get(key)
        if value is None:
            continue
        try:
            return int(float(str(value).replace(",", "")))
        except (TypeError, ValueError):
            continue
    return None


def product_review_count(raw_payload: dict[str, Any]) -> int | None:
    if "reviewCount" not in raw_payload:
        return None
    value = raw_payload.get("reviewCount")
    try:
        return max(0, int(float(str(value).replace(",", ""))))
    except (TypeError, ValueError):
        return None


def product_detail_to_public(row: ProductModel) -> dict[str, Any]:
    raw_payload = product_raw_payload(row)
    shop_code = product_shop_code(row, raw_payload)
    public = product_to_public(row)
    public["detail"] = {
        "manageNumber": first_text_from_keys(raw_payload, ("manageNumber", "itemNumber")) or row.rakuten_manage_number,
        "itemNumber": first_text_from_keys(raw_payload, ("itemNumber", "manageNumber")) or row.item_number,
        "title": first_text_from_keys(raw_payload, ("itemName", "title", "name")) or row.title,
        "tagline": product_tagline(raw_payload),
        "genreId": first_text_from_keys(raw_payload, ("genreId", "genre_id", "genre")) or row.genre_id,
        "shopName": row.shop_name,
        "sourceUrl": row.source_url,
        "rakutenItemUrl": product_rakuten_item_url(row, raw_payload),
        "listingStatus": row.rakuten_listing_status,
        "salesCount": product_sales_count(raw_payload),
        "created": first_text_from_keys(raw_payload, ("created",)),
        "updated": first_text_from_keys(raw_payload, ("updated",)),
        "descriptions": product_descriptions(raw_payload),
        "images": product_editable_image_urls(raw_payload, shop_code=shop_code),
        "variantSelectors": product_variant_selectors(raw_payload),
        "variants": product_variants(raw_payload),
        "raw": raw_payload,
    }
    return public


def product_descriptions(raw_payload: dict[str, Any]) -> list[dict[str, str]]:
    descriptions: list[dict[str, str]] = []
    seen_values: set[tuple[str, str]] = set()

    def append(label: str, value: Any, *, keep_empty: bool = False) -> None:
        normalized_label = normalize_text(label) or "商品说明"
        if not has_description_source(value):
            return
        text = normalize_listing_detail_html(value)
        text_key = description_content_key(text)
        seen_key = (normalized_label, text_key)
        if not text_key and not keep_empty:
            return
        if seen_key in seen_values:
            return
        seen_values.add(seen_key)
        descriptions.append({"label": normalized_label, "value": text})

    source_fields = source_rakuten_description_fields(raw_payload)
    append("PC用 商品説明文", source_fields.get("PC用 商品説明文"), keep_empty=True)
    append("スマートフォン用 商品説明文", source_fields.get("スマートフォン用 商品説明文"), keep_empty=True)
    append("PC用 販売説明文", source_fields.get("PC用 販売説明文"), keep_empty=True)

    product_description = raw_payload.get("productDescription")
    if isinstance(product_description, dict):
        append("PC用 商品説明文", product_description.get("pc"), keep_empty=True)
        append("スマートフォン用 商品説明文", product_description.get("sp"), keep_empty=True)
        append("智能手机商品说明", product_description.get("smartphone"))
        append("商品说明", product_description.get("value"))
    else:
        append("商品说明", product_description)

    raw_descriptions = raw_payload.get("descriptions")
    if isinstance(raw_descriptions, list):
        for index, item in enumerate(raw_descriptions, start=1):
            if isinstance(item, dict):
                append(first_text_from_keys(item, ("label", "name")) or f"商品说明 {index}", item.get("value"))
            else:
                append(f"商品说明 {index}", item)

    fields = (
        ("商品说明", "description"),
        ("PC用商品说明", "pcDescription"),
        ("スマートフォン用 商品説明文", "spDescription"),
        ("智能手机商品说明", "smartphoneDescription"),
    )
    for label, key in fields:
        append(label, raw_payload.get(key))
    append("PC用 販売説明文", raw_payload.get("salesDescription"), keep_empty=True)
    return normalize_rakuten_description_fields(
        clean_market_product_descriptions(descriptions, keep_empty_labels=RAKUTEN_DESCRIPTION_FIELD_LABELS)
    )


def source_rakuten_description_fields(raw_payload: dict[str, Any]) -> dict[str, str]:
    fields = {label: "" for label in RAKUTEN_STANDARD_DESCRIPTION_LABELS}
    embedded_item = raw_payload.get("embeddedItem") if isinstance(raw_payload.get("embeddedItem"), dict) else {}
    pc_fields = embedded_item.get("pcFields") if isinstance(embedded_item.get("pcFields"), dict) else {}
    source_values = {
        "PC用 商品説明文": pc_fields.get("productDescription"),
        "スマートフォン用 商品説明文": embedded_item.get("newProductDescription"),
        "PC用 販売説明文": embedded_item.get("salesDescription"),
    }
    for label, value in source_values.items():
        if has_description_source(value):
            fields[label] = str(value or "")

    product_description = raw_payload.get("productDescription")
    if isinstance(product_description, dict):
        fallback_values = {
            "PC用 商品説明文": product_description.get("pc"),
            "スマートフォン用 商品説明文": product_description.get("sp") or product_description.get("smartphone"),
            "PC用 販売説明文": raw_payload.get("salesDescription"),
        }
        for label, value in fallback_values.items():
            if not fields[label] and has_description_source(value):
                fields[label] = str(value or "")

    raw_descriptions = raw_payload.get("descriptions")
    if isinstance(raw_descriptions, list):
        for item in raw_descriptions:
            if not isinstance(item, dict):
                continue
            target_label = standard_rakuten_description_label(first_text_from_keys(item, ("label", "name")))
            if target_label and not fields[target_label] and has_description_source(item.get("value")):
                fields[target_label] = str(item.get("value") or "")

    if not fields["PC用 販売説明文"] and has_description_source(raw_payload.get("salesDescription")):
        fields["PC用 販売説明文"] = str(raw_payload.get("salesDescription") or "")
    return fields


def product_descriptions_for_display(raw_payload: dict[str, Any], images: list[str] | None = None, *, shop_code: str = "") -> list[dict[str, str]]:
    return product_descriptions(raw_payload)


def product_shop_code(row: ProductModel, raw_payload: dict[str, Any]) -> str:
    for value in (
        first_text_from_keys(raw_payload, ("shopCode", "shop_code", "shopUrl", "shop_url")),
        row.source_url,
        row.image_url,
    ):
        shop_code = normalize_shop_code(value)
        if shop_code:
            return shop_code
    return ""


def product_image_urls(raw_payload: dict[str, Any], *, shop_code: str = "") -> list[str]:
    urls: list[str] = []
    skipped_description_keys = {
        "description",
        "descriptions",
        "productdescription",
        "pcdescription",
        "spdescription",
        "smartphonedescription",
        "salesdescription",
        "descriptionimages",
    }

    def remember(value: Any) -> None:
        url = normalize_product_image_url(value, shop_code=shop_code)
        if url and url not in urls:
            urls.append(url)

    def collect(value: Any) -> None:
        if isinstance(value, str):
            text = unescape(str(value or "")).replace("\\/", "/")
            remember(text)
            for match in re.findall(r"https?://[^\s\"'<>)]+'?", text):
                remember(match)
            return
        if isinstance(value, dict):
            for key in ("url", "imageUrl", "location", "value"):
                collect(value.get(key))
            for key, child in value.items():
                if str(key).lower() in skipped_description_keys:
                    continue
                collect(child)
        elif isinstance(value, list):
            for child in value:
                collect(child)

    explicit_images = raw_payload.get("images")
    if isinstance(explicit_images, list) and not isinstance(raw_payload.get("ltEditedImages"), list):
        collect(explicit_images)
        if urls:
            return urls

    for image_url in trusted_product_main_image_urls(raw_payload, shop_code=shop_code):
        remember(image_url)
    return urls


def trusted_product_main_image_urls(
    raw_payload: dict[str, Any],
    *,
    shop_code: str = "",
) -> list[str]:
    urls: list[str] = []

    def remember(value: Any) -> None:
        url = normalize_product_image_url(value, shop_code=shop_code)
        if url and url not in urls:
            urls.append(url)

    def collect(value: Any) -> None:
        if isinstance(value, str):
            remember(value)
            return
        if isinstance(value, dict):
            for key in ("location", "url", "imageUrl", "src", "value"):
                collect(value.get(key))
            return
        if isinstance(value, list):
            for child in value:
                collect(child)

    explicit_images = raw_payload.get("images")
    if isinstance(explicit_images, list) and not isinstance(raw_payload.get("ltEditedImages"), list):
        explicit_remote_urls: list[str] = []
        for image in explicit_images:
            url = normalize_product_image_url(image, shop_code=shop_code)
            if url and not is_local_product_image_url(url) and not is_product_image_draft_url(url):
                explicit_remote_urls.append(url)
        if explicit_remote_urls:
            return unique_texts(explicit_remote_urls)

    embedded_item = raw_payload.get("embeddedItem")
    if isinstance(embedded_item, dict):
        pc_fields = embedded_item.get("pcFields")
        if isinstance(pc_fields, dict):
            collect(pc_fields.get("images"))
        media = embedded_item.get("media")
        if urls:
            if isinstance(media, dict):
                collect(media.get("skuImages"))
            for sku in embedded_item.get("sku") if isinstance(embedded_item.get("sku"), list) else []:
                if isinstance(sku, dict):
                    collect(sku.get("images"))
            return urls
        if isinstance(media, dict):
            collect(media.get("images"))
            collect(media.get("skuImages"))
        collect(embedded_item.get("picImageUrl"))
        for sku in embedded_item.get("sku") if isinstance(embedded_item.get("sku"), list) else []:
            if isinstance(sku, dict):
                collect(sku.get("images"))

    media = raw_payload.get("media")
    if isinstance(media, dict):
        collect(media.get("images"))
        collect(media.get("skuImages"))
    collect(raw_payload.get("imageUrl"))
    collect(raw_payload.get("imageUrls"))

    variants = raw_payload.get("variants")
    if isinstance(variants, dict):
        for variant in variants.values():
            if isinstance(variant, dict):
                collect(variant.get("images"))

    json_ld = raw_payload.get("jsonLd")
    if isinstance(json_ld, dict):
        collect(json_ld.get("image"))
    elif isinstance(json_ld, list):
        for item in json_ld:
            if isinstance(item, dict):
                collect(item.get("image"))
    return urls


def product_editable_image_urls(raw_payload: dict[str, Any], *, shop_code: str = "") -> list[str]:
    edited_images = raw_payload.get("ltEditedImages")
    if isinstance(edited_images, list):
        urls = []
        for image in edited_images:
            url = normalize_product_image_url(image, shop_code=shop_code)
            if url and url not in urls:
                urls.append(url)
        return urls
    return product_image_urls(raw_payload, shop_code=shop_code)


def set_product_image_urls(raw_payload: dict[str, Any], images: list[str]) -> dict[str, Any]:
    updated_payload = dict(raw_payload)
    normalized_images = unique_texts([image for image in images if normalize_product_image_url(image)])
    updated_payload["ltEditedImages"] = normalized_images
    updated_payload["images"] = normalized_images
    media = updated_payload.get("media")
    if isinstance(media, dict):
        updated_media = dict(media)
        updated_media["images"] = [{"type": "CABINET" if is_cabinet_image_url(image) else "ABSOLUTE", "location": image} for image in normalized_images]
        updated_payload["media"] = updated_media
    updated_payload["updated"] = datetime.now().isoformat(timespec="seconds")
    return updated_payload


def set_product_image_urls_with_description_updates(
    raw_payload: dict[str, Any],
    images: list[str],
    *,
    replace_map: dict[str, str] | None = None,
    remove_urls: list[str] | None = None,
) -> dict[str, Any]:
    updated_payload = set_product_image_urls(raw_payload, images)
    if replace_map:
        updated_payload = replace_product_description_image_urls(updated_payload, replace_map)
    if remove_urls:
        updated_payload = remove_product_description_image_urls(updated_payload, remove_urls)
    return updated_payload


def rakuten_image_identity_key(image_url: str) -> str:
    normalized_url = normalize_product_image_url(image_url)
    if not normalized_url:
        return ""
    parsed = urlsplit(normalized_url)
    if parsed.hostname not in RAKUTEN_IMAGE_CDN_HOSTS:
        return ""
    path = unquote(parsed.path or "").replace("\\", "/")
    return path.lower().rstrip("/")


def rakuten_image_quality_score(image_url: str) -> tuple[int, int]:
    parsed = urlsplit(image_url)
    query = parse_qs(parsed.query)
    resized = any(key.lower() in {"_ex", "fitin", "resize", "width", "height"} for key in query)
    preferred_host = 1 if parsed.hostname == "image.rakuten.co.jp" else 0
    return (0 if resized else 1, preferred_host)


def preferred_rakuten_image_urls(image_urls: list[str]) -> tuple[list[str], dict[str, str]]:
    source_urls = unique_texts(image_urls)
    best_by_identity: dict[str, str] = {}
    for image_url in source_urls:
        identity = rakuten_image_identity_key(image_url)
        if not identity:
            continue
        current = best_by_identity.get(identity)
        if current is None or rakuten_image_quality_score(image_url) > rakuten_image_quality_score(current):
            best_by_identity[identity] = image_url

    selected_urls: list[str] = []
    aliases: dict[str, str] = {}
    emitted_identities: set[str] = set()
    for image_url in source_urls:
        identity = rakuten_image_identity_key(image_url)
        if not identity:
            selected_urls.append(image_url)
            continue
        preferred_url = best_by_identity[identity]
        aliases[image_url] = preferred_url
        if identity not in emitted_identities:
            selected_urls.append(preferred_url)
            emitted_identities.add(identity)
    return unique_texts(selected_urls), aliases


def product_image_visual_signature(content: bytes) -> ProductImageVisualSignature | None:
    try:
        from PIL import Image, ImageOps, UnidentifiedImageError

        with Image.open(BytesIO(content)) as source:
            image = ImageOps.exif_transpose(source).convert("RGB")
            width, height = image.size
            if width < 1 or height < 1:
                return None
            pixels = image.resize(PRODUCT_IMAGE_VISUAL_SIZE, Image.Resampling.LANCZOS).tobytes()
            return ProductImageVisualSignature(width=width, height=height, pixels=pixels)
    except (OSError, UnidentifiedImageError):
        return None


def product_images_visually_equal(
    first: ProductImageVisualSignature,
    second: ProductImageVisualSignature,
) -> bool:
    first_ratio = first.width / first.height
    second_ratio = second.width / second.height
    if abs(first_ratio - second_ratio) / max(first_ratio, second_ratio) > PRODUCT_IMAGE_VISUAL_MAX_ASPECT_RATIO_DIFFERENCE:
        return False
    if len(first.pixels) != len(second.pixels) or not first.pixels:
        return False
    difference = sum(abs(a - b) for a, b in zip(first.pixels, second.pixels)) / len(first.pixels)
    return difference <= PRODUCT_IMAGE_VISUAL_MAX_MEAN_DIFFERENCE


def canonical_product_image_url_by_visual_content(
    content: bytes,
    visual_content_urls: list[tuple[ProductImageVisualSignature, str]],
) -> tuple[str, ProductImageVisualSignature | None]:
    signature = product_image_visual_signature(content)
    if signature is None:
        return "", None
    for known_signature, known_url in visual_content_urls:
        if product_images_visually_equal(signature, known_signature):
            return known_url, signature
    return "", signature


def localize_collected_product_images(owner_username: str, product_id: int) -> str:
    referenced_local_urls: list[str] = []
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            return ""
        if product.review_status not in {"pending", "approved", "error", "listed_master"}:
            return ""
        raw_payload = product_raw_payload(product)
        shop_code = product_shop_code(product, raw_payload)
        product_title = product.title
        source_images = product_editable_image_urls(raw_payload, shop_code=shop_code)
        if product.image_url and product.image_url not in source_images:
            source_images.insert(0, product.image_url)

    content_hash_urls: dict[str, str] = {}
    visual_content_urls: list[tuple[ProductImageVisualSignature, str]] = []
    image_result = localize_product_image_urls(
        product_id,
        source_images,
        prefix="p",
        content_hash_urls=content_hash_urls,
        visual_content_urls=visual_content_urls,
    )
    description_result = localize_product_description_images(
        product_id,
        raw_payload,
        existing_replacements=image_result["replacementMap"],
        content_hash_urls=content_hash_urls,
        visual_content_urls=visual_content_urls,
    )
    replacement_map = {**image_result["replacementMap"], **description_result["replacementMap"]}

    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            return ""
        if product.review_status not in {"pending", "approved", "error", "listed_master"}:
            return ""
        raw_payload = product_raw_payload(product)
        updated_payload = set_product_image_urls(raw_payload, image_result["urls"]) if image_result["urls"] else dict(raw_payload)
        if replacement_map:
            updated_payload = replace_product_description_image_urls(updated_payload, replacement_map)
            updated_payload = replace_payload_image_url_texts(updated_payload, replacement_map)
        removed_description_urls = description_result.get("removedUrls") or []
        if removed_description_urls:
            updated_payload = remove_product_description_image_urls(updated_payload, removed_description_urls)
        updated_payload["ltLocalImagesReady"] = True
        updated_payload["ltLocalImageUpdatedAt"] = datetime.now().isoformat(timespec="seconds")
        description_warnings = description_result.get("warnings") or []
        if description_warnings:
            updated_payload["ltLocalImageWarnings"] = description_warnings[:20]
        else:
            updated_payload.pop("ltLocalImageWarnings", None)
        if image_result["errors"] or description_result["errors"]:
            image_errors = [*image_result["errors"], *description_result["errors"]]
            updated_payload["ltLocalImageErrors"] = image_errors[:20]
            product.last_error = summarize_local_image_errors(product_title or product.title, product.source_url, product.id, image_errors)
        else:
            updated_payload.pop("ltLocalImageErrors", None)
            if product.last_error and product.last_error.startswith("图片本地化"):
                product.last_error = None
        original_images = [
            image
            for image in source_images
            if normalize_product_image_url(image, shop_code=shop_code)
            and not is_local_product_image_url(image)
            and not is_product_image_draft_url(image)
            and not is_gif_image_url(image)
        ]
        if original_images:
            updated_payload["ltOriginalImages"] = unique_texts(original_images)
        product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        if image_result["urls"]:
            product.image_url = image_result["urls"][0]
        referenced_local_urls = collect_local_product_image_urls(updated_payload)
        session.flush()
        result = product.last_error or ""
    remove_unused_local_product_images(product_id, referenced_local_urls)
    return result


def localize_product_image_urls(
    product_id: int,
    image_urls: list[str],
    *,
    prefix: str,
    content_hash_urls: dict[str, str] | None = None,
    visual_content_urls: list[tuple[ProductImageVisualSignature, str]] | None = None,
) -> dict[str, Any]:
    local_urls: list[str] = []
    replacement_map: dict[str, str] = {}
    errors: list[str] = []
    known_content_hash_urls = content_hash_urls if content_hash_urls is not None else {}
    known_visual_content_urls = visual_content_urls if visual_content_urls is not None else []
    source_urls, source_aliases = preferred_rakuten_image_urls(image_urls)
    if not source_urls:
        errors.append("未采集到商品主图。")
    for index, image_url in enumerate(source_urls, start=1):
        if is_product_image_draft_url(image_url):
            continue
        if is_gif_image_url(image_url):
            continue
        if is_local_product_image_url(image_url):
            canonical_url = localized_product_image_url_by_content(
                image_url,
                known_content_hash_urls,
                known_visual_content_urls,
            )
            replacement_map[image_url] = canonical_url
            if canonical_url not in local_urls:
                local_urls.append(canonical_url)
            continue
        try:
            local_url = save_remote_product_image(
                product_id,
                image_url,
                f"{prefix}{index:02d}",
                content_hash_urls=known_content_hash_urls,
                visual_content_urls=known_visual_content_urls,
            )
        except Exception as exc:
            errors.append(f"{image_url}: {exc}")
            continue
        replacement_map[image_url] = local_url
        local_urls.append(local_url)
    for alias_url, preferred_url in source_aliases.items():
        canonical_url = replacement_map.get(preferred_url)
        if canonical_url:
            replacement_map[alias_url] = canonical_url
    return {"urls": unique_texts(local_urls), "replacementMap": replacement_map, "errors": errors}


def localize_product_description_images(
    product_id: int,
    raw_payload: dict[str, Any],
    *,
    existing_replacements: dict[str, str] | None = None,
    content_hash_urls: dict[str, str] | None = None,
    visual_content_urls: list[tuple[ProductImageVisualSignature, str]] | None = None,
) -> dict[str, Any]:
    description_urls = unique_texts(
        [
            url
            for description in product_descriptions(raw_payload)
            for url in description_image_urls(description.get("value"))
            if not is_product_image_draft_url(url)
        ]
    )
    replacement_map: dict[str, str] = {}
    errors: list[str] = []
    warnings: list[str] = []
    removed_urls: list[str] = []
    known_replacements = existing_replacements or {}
    known_content_hash_urls = content_hash_urls if content_hash_urls is not None else {}
    known_visual_content_urls = visual_content_urls if visual_content_urls is not None else []
    for index, image_url in enumerate(description_urls, start=1):
        if is_gif_image_url(image_url):
            removed_urls.append(image_url)
            warnings.append(f"{image_url}: GIF 图片已从详情说明移除。")
            continue
        if is_local_product_image_url(image_url):
            replacement_map[image_url] = localized_product_image_url_by_content(
                image_url,
                known_content_hash_urls,
                known_visual_content_urls,
            )
            continue
        if image_url in known_replacements:
            replacement_map[image_url] = known_replacements[image_url]
            continue
        try:
            replacement_map[image_url] = save_remote_product_image(
                product_id,
                image_url,
                f"d{index:02d}",
                content_hash_urls=known_content_hash_urls,
                visual_content_urls=known_visual_content_urls,
            )
        except ProductImageUnavailableError as exc:
            removed_urls.append(image_url)
            warnings.append(f"{image_url}: {exc}，已从详情说明移除。")
        except Exception as exc:
            warnings.append(f"{image_url}: {exc}")
    return {"replacementMap": replacement_map, "errors": errors, "warnings": warnings, "removedUrls": unique_texts(removed_urls)}


def localized_product_image_url_by_content(
    image_url: str,
    content_hash_urls: dict[str, str],
    visual_content_urls: list[tuple[ProductImageVisualSignature, str]],
) -> str:
    try:
        image_data = load_product_image_bytes(
            image_url,
            max_bytes=MAX_PRODUCT_IMAGE_DOWNLOAD_BYTES,
            size_error_message="图片下载大小不能超过 20MB。",
        )
    except Exception:
        return image_url
    digest = hashlib.sha256(image_data["content"]).hexdigest()
    canonical_url = content_hash_urls.get(digest)
    if canonical_url:
        return canonical_url
    canonical_url, visual_signature = canonical_product_image_url_by_visual_content(
        image_data["content"],
        visual_content_urls,
    )
    if canonical_url:
        content_hash_urls[digest] = canonical_url
        return canonical_url
    content_hash_urls[digest] = image_url
    if visual_signature is not None:
        visual_content_urls.append((visual_signature, image_url))
    return image_url


def save_remote_product_image(
    product_id: int,
    image_url: str,
    name_prefix: str,
    *,
    content_hash_urls: dict[str, str] | None = None,
    visual_content_urls: list[tuple[ProductImageVisualSignature, str]] | None = None,
) -> str:
    image_data = load_product_image_bytes(
        image_url,
        max_bytes=MAX_PRODUCT_IMAGE_DOWNLOAD_BYTES,
        size_error_message="图片下载大小不能超过 20MB。",
    )
    digest = hashlib.sha256(image_data["content"]).hexdigest()
    known_content_hash_urls = content_hash_urls if content_hash_urls is not None else {}
    canonical_url = known_content_hash_urls.get(digest)
    if canonical_url:
        return canonical_url
    known_visual_content_urls = visual_content_urls if visual_content_urls is not None else []
    canonical_url, visual_signature = canonical_product_image_url_by_visual_content(
        image_data["content"],
        known_visual_content_urls,
    )
    if canonical_url:
        known_content_hash_urls[digest] = canonical_url
        return canonical_url
    safe_name = f"{name_prefix}-{uuid.uuid4().hex[:12]}{image_data['suffix']}"
    target_url = local_product_image_url(product_id, safe_name)
    store_product_image_content(
        target_url,
        image_data["content"],
        image_data["contentType"],
        LOCAL_PRODUCT_IMAGE_DIR / str(int(product_id)) / safe_name,
    )
    known_content_hash_urls[digest] = target_url
    if visual_signature is not None:
        known_visual_content_urls.append((visual_signature, target_url))
    return target_url


def is_local_product_image_url(image_url: str) -> bool:
    stored_image = parse_product_image_url(image_url)
    return bool(stored_image and stored_image.kind == PRODUCT_IMAGE_OBJECT_PREFIX)


def summarize_local_image_errors(product_title: str, source_url: str, product_id: int, errors: list[str]) -> str:
    if not errors:
        return ""
    display_name = normalize_text(product_title or source_url or product_id)
    first_error = errors[0]
    suffix = f"，另有 {len(errors) - 1} 张失败" if len(errors) > 1 else ""
    return f"图片本地化失败：{display_name}: {first_error[:300]}{suffix}"


def mark_product_local_image_error(owner_username: str, product_id: int, message: str) -> None:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            return
        product.last_error = message[:1000]


def replace_payload_image_url_texts(value: Any, replacement_map: dict[str, str]) -> Any:
    if not replacement_map:
        return value
    if isinstance(value, str):
        next_value = replace_payload_image_urls_in_text(value, replacement_map)
        for old_url, new_url in replacement_map.items():
            if old_url and new_url:
                next_value = next_value.replace(old_url, new_url)
        return next_value
    if isinstance(value, list):
        return [replace_payload_image_url_texts(item, replacement_map) for item in value]
    if isinstance(value, dict):
        return {key: replace_payload_image_url_texts(child, replacement_map) for key, child in value.items()}
    return value


def replace_payload_image_urls_in_text(value: str, replacement_map: dict[str, str]) -> str:
    def replace_match(match: re.Match[str]) -> str:
        matched = match.group(0)
        image_url = matched.rstrip(".,;")
        trailing = matched[len(image_url):]
        replacement = (
            replacement_map.get(image_url)
            or replacement_map.get(normalize_product_image_url(image_url))
            or replacement_map.get(normalize_description_image_url(image_url))
        )
        return f"{replacement}{trailing}" if replacement else matched

    return re.sub(r"https?://[^\s\"'<>)]+'?", replace_match, value)


def collect_local_product_image_urls(value: Any) -> list[str]:
    urls: list[str] = []

    def remember(candidate: Any) -> None:
        url = normalize_product_image_url(candidate)
        if is_local_product_image_url(url) and url not in urls:
            urls.append(url)

    def walk(item: Any) -> None:
        if isinstance(item, str):
            text = unescape(item).replace("\\/", "/")
            remember(text)
            pattern = rf"{re.escape(LOCAL_PRODUCT_IMAGE_URL_PREFIX)}/\d+/[^\s\"'<>),]+"
            for match in re.findall(pattern, text):
                remember(match.rstrip(".,;"))
            return
        if isinstance(item, list):
            for child in item:
                walk(child)
            return
        if isinstance(item, dict):
            for child in item.values():
                walk(child)

    walk(value)
    return urls


def remove_unused_local_product_images(product_id: int, referenced_urls: list[str]) -> None:
    image_dir = (LOCAL_PRODUCT_IMAGE_DIR / str(int(product_id))).resolve()
    root = LOCAL_PRODUCT_IMAGE_DIR.resolve()
    try:
        image_dir.relative_to(root)
    except ValueError:
        return
    if not image_dir.exists() or not image_dir.is_dir():
        return
    referenced_paths = {
        path.resolve()
        for path in (local_product_image_path_from_url(url) for url in referenced_urls)
        if path is not None
    }
    for path in image_dir.iterdir():
        if not path.is_file():
            continue
        try:
            if path.resolve() not in referenced_paths:
                path.unlink(missing_ok=True)
        except OSError:
            continue


def normalize_product_image_url(value: Any, *, shop_code: str = "") -> str:
    text = unescape(str(value or "")).replace("\\/", "/").strip().strip("'\"")
    if text.startswith((LOCAL_PRODUCT_IMAGE_URL_PREFIX, LOCAL_PRODUCT_IMAGE_DRAFT_URL_PREFIX)):
        return text.split("?", 1)[0].split("#", 1)[0]
    if not text.startswith(("http://", "https://")):
        if not re.search(r"\.(apng|avif|bmp|gif|jpe?g|png|webp)(?:$|[?#])", text, flags=re.I):
            return ""
        normalized_location = text.lstrip("/")
        thumbnail_match = re.match(r"@0_mall/([^/]+)/cabinet/(.+)", normalized_location, flags=re.I)
        if thumbnail_match:
            matched_shop_code = normalize_shop_code(thumbnail_match.group(1))
            if shop_code and matched_shop_code and matched_shop_code != normalize_shop_code(shop_code):
                return ""
            shop_code = shop_code or matched_shop_code
            normalized_location = thumbnail_match.group(2)
        if not shop_code:
            return ""
        if normalized_location.startswith("cabinet/"):
            normalized_location = normalized_location.removeprefix("cabinet/")
        text = build_rakuten_cabinet_image_url(shop_code, normalized_location)
    text = text.rstrip(".,;")
    try:
        parsed = urlsplit(text)
        path = parsed.path.lower()
    except Exception:
        return ""
    if not re.search(r"\.(apng|avif|bmp|gif|jpe?g|png|webp)$", path):
        return ""
    filename = path.rsplit("/", 1)[-1]
    if is_ignored_cabinet_image_filename(filename):
        return ""
    return f"{parsed.scheme}://{parsed.netloc}{parsed.path}"


def is_gif_image_url(value: Any) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    try:
        path = urlsplit(text).path
    except Exception:
        path = text.split("?", 1)[0].split("#", 1)[0]
    return path.lower().endswith(".gif")


def is_cabinet_image_url(value: str) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    if text.startswith("/cabinet/") or text.startswith("cabinet/"):
        return True
    try:
        parsed = urlsplit(text)
    except Exception:
        return False
    return "/cabinet/" in parsed.path.lower() and parsed.netloc.lower() in {
        "image.rakuten.co.jp",
        "thumbnail.image.rakuten.co.jp",
        "cabinet.rms.rakuten.co.jp",
    }


def is_ignored_cabinet_image_filename(value: str) -> bool:
    return value.strip().lower().rsplit("/", 1)[-1] in IGNORED_CABINET_IMAGE_FILENAMES


def product_cabinet_file_targets(raw_payload: dict[str, Any], shop_code: str) -> list[dict[str, str]]:
    normalized_shop_code = normalize_shop_code(shop_code)
    targets: list[dict[str, str]] = []
    seen: set[str] = set()

    def remember(value: Any) -> None:
        for path in cabinet_paths_from_text(value, normalized_shop_code):
            if not path or path in seen:
                continue
            if is_ignored_cabinet_image_filename(path):
                continue
            target = cabinet_target_from_path(path)
            target_key = "|".join([target.get("folderPath", ""), target.get("filePath", ""), target.get("fileName", "")])
            if target_key in seen:
                continue
            seen.add(target_key)
            targets.append(target)

    def walk(value: Any) -> None:
        if isinstance(value, str):
            remember(value)
            return
        if isinstance(value, dict):
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(raw_payload)
    for url in product_image_urls(raw_payload, shop_code=shop_code):
        remember(url)
    return targets


def cabinet_target_from_path(path: str) -> dict[str, str]:
    normalized_path = normalize_cabinet_path(path)
    without_cabinet = normalized_path
    if without_cabinet.lower().startswith("/cabinet/"):
        without_cabinet = without_cabinet[len("/cabinet/") :]
    else:
        without_cabinet = without_cabinet.lstrip("/")
    folder_path, _, file_path = without_cabinet.rpartition("/")
    file_name = file_path or without_cabinet
    return {
        "folderPath": f"/{folder_path}" if folder_path else "",
        "filePath": file_path or file_name,
        "fileName": file_name,
        "cabinetPath": normalized_path,
    }


def cabinet_paths_from_text(value: Any, shop_code: str) -> list[str]:
    text = str(value or "").replace("\\/", "/")
    if not text:
        return []

    patterns = [
        r"https?://image\.rakuten\.co\.jp/([^/]+)/cabinet/[^\"'\s<>]+?\.(?:jpg|jpeg|png|webp|gif)",
        r"https?://(?:shop|tshop)\.r10s\.jp/([^/]+)/cabinet/[^\"'\s<>]+?\.(?:jpg|jpeg|png|webp|gif)",
        r"https?://thumbnail\.image\.rakuten\.co\.jp/@0_mall/([^/]+)/cabinet/[^\"'\s<>]+?\.(?:jpg|jpeg|png|webp|gif)",
        r"@0_mall/([^/]+)/cabinet/[^\"'\s<>]+?\.(?:jpg|jpeg|png|webp|gif)",
    ]
    paths: list[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            matched_shop = normalize_shop_code(match.group(1))
            if shop_code and matched_shop and matched_shop != shop_code:
                continue
            full_match = match.group(0)
            cabinet_index = full_match.lower().find("/cabinet/")
            if cabinet_index < 0:
                cabinet_index = full_match.lower().find("cabinet/")
            if cabinet_index < 0:
                continue
            path = "/" + full_match[cabinet_index:].lstrip("/")
            path = path.split("?", 1)[0].split("#", 1)[0]
            if path not in paths:
                paths.append(path)

    for match in re.finditer(r"(?<![\w/])cabinet/[^\"'\s<>]+?\.(?:jpg|jpeg|png|webp|gif)", text, flags=re.IGNORECASE):
        path = "/" + match.group(0).split("?", 1)[0].split("#", 1)[0].lstrip("/")
        if path not in paths:
            paths.append(path)
    return paths


def product_variant_selectors(raw_payload: dict[str, Any]) -> list[dict[str, Any]]:
    selectors = raw_payload.get("variantSelectors")
    if not isinstance(selectors, list):
        return []
    result: list[dict[str, Any]] = []
    for selector in selectors:
        if not isinstance(selector, dict):
            continue
        result.append(
            {
                "key": first_text_from_keys(selector, ("key", "id", "selectorId")),
                "name": first_text_from_keys(selector, ("name", "displayName", "label")),
                "values": selector_values_to_public(selector.get("values")),
            }
        )
    return result


def product_variants(raw_payload: dict[str, Any]) -> list[dict[str, Any]]:
    variants = raw_payload.get("variants")
    if isinstance(variants, dict):
        variant_items = variants.items()
    elif isinstance(variants, list):
        variant_items = ((first_text_from_keys(item, ("variantId", "skuId", "merchantDefinedSkuId")) if isinstance(item, dict) else "", item) for item in variants)
    else:
        variant_items = []

    result: list[dict[str, Any]] = []
    for variant_id, variant in variant_items:
        if not isinstance(variant, dict):
            continue
        selector_values = variant.get("selectorValues")
        result.append(
            {
                "variantId": normalize_text(variant_id) or first_text_from_keys(variant, ("variantId", "skuId")),
                "merchantDefinedSkuId": first_text_from_keys(variant, ("merchantDefinedSkuId",)),
                "articleNumber": first_text_value(variant.get("articleNumber")),
                "standardPrice": first_text_from_keys(variant, ("standardPrice", "price", "displayPrice")),
                "hidden": variant.get("hidden"),
                "selectorValues": selector_values if isinstance(selector_values, dict) else {},
                "specs": variant.get("specs") if isinstance(variant.get("specs"), list) else [],
                "attributes": variant.get("attributes") if isinstance(variant.get("attributes"), list) else [],
                "material": first_text_from_keys(variant, ("material",)),
            }
        )
    if result:
        return result
    price = price_from_rakuten_item_without_variants(raw_payload)
    if price is None:
        return []
    return [
        {
            "variantId": SINGLE_PRODUCT_VARIANT_ID,
            "merchantDefinedSkuId": "",
            "articleNumber": "",
            "standardPrice": str(int(price)) if price == int(price) else str(price),
            "hidden": False,
            "selectorValues": {},
            "specs": [],
            "attributes": [],
            "material": "",
            "singleProduct": True,
        }
    ]


def selector_values_to_public(values: Any) -> list[Any]:
    if not isinstance(values, list):
        return []
    result: list[Any] = []
    for item in values:
        if isinstance(item, dict):
            result.append(first_text_from_keys(item, ("label", "value", "name")) or item)
        else:
            result.append(item)
    return result


def parse_datetime_filter(value: str | None) -> datetime | None:
    text = normalize_text(value)
    if not text:
        return None
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed


def parse_sales_date_filter(value: str | None, *, field_name: str) -> date | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"{field_name}格式不正确，应为YYYY-MM-DD。") from exc


def normalize_store_product_sales_range(
    sales_period_days: int | None,
    sales_period_from: str | None,
    sales_period_to: str | None,
    *,
    reference_date: date | None = None,
) -> tuple[date, date] | None:
    period_from = parse_sales_date_filter(
        sales_period_from,
        field_name="销量开始日期",
    )
    period_to = parse_sales_date_filter(
        sales_period_to,
        field_name="销量结束日期",
    )
    if (period_from is None) != (period_to is None):
        raise ValueError("销量自定义时间范围必须同时填写开始日期和结束日期。")

    today = reference_date or sales_now_naive().date()
    earliest_allowed = today - timedelta(days=STORE_PRODUCT_SALES_MAX_RANGE_DAYS - 1)
    if period_from is not None and period_to is not None:
        if period_from > period_to:
            raise ValueError("销量开始日期不能晚于结束日期。")
        if period_to > today:
            raise ValueError("销量结束日期不能晚于今天。")
        if period_from < earliest_allowed:
            raise ValueError("销量时间范围只能选择最近365天。")
        if (period_to - period_from).days + 1 > STORE_PRODUCT_SALES_MAX_RANGE_DAYS:
            raise ValueError("销量时间范围不能超过365天。")
        return period_from, period_to

    if sales_period_days is None:
        return None
    try:
        normalized_days = int(sales_period_days)
    except (TypeError, ValueError) as exc:
        raise ValueError("销量周期不合法。") from exc
    if normalized_days not in STORE_PRODUCT_SALES_PERIOD_DAYS:
        raise ValueError(
            "销量周期只能选择近1周、近2周、近1个月、近2个月、近3个月、近半年或近一年。"
        )
    return today - timedelta(days=normalized_days - 1), today


def normalize_schedule_time(value: Any) -> str:
    text = normalize_text(value) or "09:00"
    match = re.fullmatch(r"([0-9]{1,2}):([0-9]{1,2})(?::[0-9]{1,2})?", text)
    if not match:
        raise RuntimeError("定时执行时间格式不正确。")
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour > 23 or minute > 59:
        raise RuntimeError("定时执行时间不合法。")
    return f"{hour:02d}:{minute:02d}"


def next_daily_run_at(schedule_time: str, *, now: datetime | None = None) -> datetime:
    reference = now or datetime.now()
    normalized = normalize_schedule_time(schedule_time)
    hour, minute = [int(part) for part in normalized.split(":", 1)]
    candidate = reference.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if candidate <= reference:
        candidate += timedelta(days=1)
    return candidate


def datetime_to_public(value: datetime | None) -> str | None:
    return value.isoformat(sep=" ", timespec="seconds") if value else None


def parse_public_datetime(value: Any) -> datetime | None:
    text = normalize_text(value)
    if not text:
        return None
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed


def manual_task_execution_settings(
    payload: Any,
    *,
    task_name: str,
) -> tuple[str, datetime | None]:
    execution_mode = normalize_text(
        getattr(payload, "executionMode", None)
    ).lower() or "immediate"
    if execution_mode not in {"immediate", "scheduled"}:
        raise RuntimeError(f"{task_name}执行方式不正确。")
    if execution_mode == "immediate":
        return execution_mode, None
    execute_at = parse_public_datetime(getattr(payload, "executeAt", None))
    if execute_at is None:
        raise RuntimeError(f"请选择{task_name}的到期执行时间。")
    if execute_at <= datetime.now():
        raise RuntimeError(f"{task_name}的到期执行时间必须晚于当前时间。")
    return execution_mode, execute_at


def normalize_cleanup_weekday(value: Any) -> int:
    try:
        weekday = int(value)
    except (TypeError, ValueError):
        raise RuntimeError("定时清理星期不合法。") from None
    if weekday < 0 or weekday > 6:
        raise RuntimeError("定时清理星期不合法。")
    return weekday


def next_weekly_run_at(weekday: int, schedule_time: str, *, now: datetime | None = None) -> datetime:
    reference = now or datetime.now()
    normalized_weekday = normalize_cleanup_weekday(weekday)
    normalized_time = normalize_schedule_time(schedule_time)
    hour, minute = [int(part) for part in normalized_time.split(":", 1)]
    days_until = (normalized_weekday - reference.weekday()) % 7
    candidate = (reference + timedelta(days=days_until)).replace(hour=hour, minute=minute, second=0, microsecond=0)
    if candidate <= reference:
        candidate += timedelta(days=7)
    return candidate


def next_monthly_run_at(month_day: int, schedule_time: str, *, now: datetime | None = None) -> datetime:
    reference = now or datetime.now()
    normalized_time = normalize_schedule_time(schedule_time)
    hour, minute = [int(part) for part in normalized_time.split(":", 1)]
    requested_day = max(1, min(31, int(month_day or 1)))
    current_day = min(requested_day, calendar.monthrange(reference.year, reference.month)[1])
    candidate = reference.replace(day=current_day, hour=hour, minute=minute, second=0, microsecond=0)
    if candidate <= reference:
        year = reference.year + (1 if reference.month == 12 else 0)
        month = 1 if reference.month == 12 else reference.month + 1
        next_day = min(requested_day, calendar.monthrange(year, month)[1])
        candidate = candidate.replace(year=year, month=month, day=next_day)
    return candidate


def next_auto_listing_run_at(
    schedule_type: str,
    schedule_time: str,
    *,
    weekday: int | None = None,
    month_day: int | None = None,
    now: datetime | None = None,
) -> datetime:
    if schedule_type == "daily":
        return next_daily_run_at(schedule_time, now=now)
    if schedule_type == "weekly":
        if weekday is None or weekday < 1 or weekday > 7:
            raise RuntimeError("请选择每周自动上架的星期。")
        return next_weekly_run_at(weekday - 1, schedule_time, now=now)
    if schedule_type == "monthly":
        if month_day is None or month_day < 1 or month_day > 31:
            raise RuntimeError("请选择每月自动上架的日期。")
        return next_monthly_run_at(month_day, schedule_time, now=now)
    raise RuntimeError("自动上架周期不正确。")


def ranking_period_label(value: Any) -> str:
    period = normalize_text(value) or "daily"
    return {
        "realtime": "实时",
        "daily": "日榜",
        "weekly": "周榜",
        "monthly": "月榜",
    }.get(period, "日榜")


def crawl_limit_label(value: Any, *, default: str = "全部") -> str:
    normalized = normalize_text(value)
    if not normalized:
        return default
    if normalized.lower() in {"all", "none"} or normalized in {"全部", "全量"}:
        return "全部"
    match = re.search(r"([0-9]{1,5})", normalized)
    if not match:
        return default
    return f"前 {max(1, int(match.group(1)))}"


def default_imported_schedule_target(shop_name: str) -> str:
    return f"店铺:{normalize_text(shop_name)} 日榜 全部"


def schedule_import_notes(shop_name: str, fallback_shop_url: str) -> str:
    return json.dumps(
        {
            SCHEDULE_IMPORTED_NOTE_KEY: True,
            "shopName": normalize_text(shop_name),
            SCHEDULE_FALLBACK_SHOP_URL_KEY: normalize_text(fallback_shop_url),
        },
        ensure_ascii=False,
    )


def schedule_fallback_shop_url(notes: Any) -> str:
    text = normalize_text(notes)
    if not text:
        return ""
    try:
        payload = json.loads(text)
    except ValueError:
        payload = {}
    if isinstance(payload, dict):
        return normalize_text(payload.get(SCHEDULE_FALLBACK_SHOP_URL_KEY))
    return ""


def append_shop_fallback_target(target: str, fallback_shop_url: str) -> str:
    base_target, _ = split_shop_fallback_target(target)
    fallback = normalize_text(fallback_shop_url)
    if not fallback:
        return base_target
    return f"{base_target}\n{SCHEDULE_FALLBACK_TARGET_PREFIX}{fallback}"


def split_shop_fallback_target(target: str) -> tuple[str, str]:
    raw = str(target or "").replace("\r\n", "\n").replace("\r", "\n")
    marker_index = raw.find(SCHEDULE_FALLBACK_TARGET_PREFIX)
    if marker_index >= 0:
        base = raw[:marker_index].strip()
        fallback_tail = raw[marker_index + len(SCHEDULE_FALLBACK_TARGET_PREFIX):].strip()
        match = re.match(r"https?://\S+", fallback_tail)
        fallback = normalize_text(match.group(0) if match else fallback_tail)
        return base, fallback

    lines = [line.strip() for line in raw.split("\n")]
    visible_lines: list[str] = []
    for line in lines:
        if line:
            visible_lines.append(line)
    return "\n".join(visible_lines).strip(), ""


def fallback_shop_target(primary_target: str, fallback_shop_url: str) -> str:
    base_target, _ = split_shop_fallback_target(primary_target)
    parsed_target, limit, period = parse_ranking_target(strip_shop_ranking_prefix(base_target))
    fallback_target = safe_shop_ranking_target(fallback_shop_url)
    if not fallback_target:
        fallback_target = parsed_target or base_target
    return f"店铺:{fallback_target} {ranking_period_label(period)} {crawl_limit_label(limit, default='全部')}"


def safe_shop_ranking_target(value: Any) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    try:
        target = normalize_rakuten_shop_target(normalized)
    except RuntimeError:
        return normalized if not normalized.startswith(("http://", "https://")) else ""
    return target


def is_rakuten_search_url(value: Any) -> bool:
    normalized = normalize_text(value)
    if not normalized.startswith(("http://", "https://")):
        return False
    try:
        parsed = urlsplit(normalized)
    except Exception:
        return False
    return parsed.netloc.lower() == "search.rakuten.co.jp" and parsed.path.rstrip("/").endswith("/search/mall")


def scheduled_crawl_task_target(row: ScheduledCrawlModel) -> str:
    fallback = schedule_fallback_shop_url(row.notes)
    return append_shop_fallback_target(row.target, fallback)


def store_to_public(
    row: StoreModel,
    *,
    reveal: bool = False,
    recent_year_order_count: int | None = None,
) -> dict[str, Any]:
    service_secret = decrypt_text(row.rakuten_service_secret_encrypted)
    license_key = decrypt_text(row.rakuten_license_key_encrypted)
    checked_at = row.last_checked_at or row.last_synced_at
    product_synced_at = row.last_product_synced_at
    availability_status = "unchecked"
    if row.last_error:
        availability_status = "error"
    elif checked_at or product_synced_at:
        availability_status = "available"
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "storeCode": row.store_code,
        "storeName": row.store_name,
        "aliasName": row.alias_name,
        "platform": row.platform,
        "storeUrl": row.store_url,
        "enabled": bool(row.enabled),
        "description": row.description,
        "rakutenServiceSecret": service_secret if reveal else "",
        "rakutenLicenseKey": license_key if reveal else "",
        "masked": {
            "rakutenServiceSecret": mask_secret(service_secret),
            "rakutenLicenseKey": mask_secret(license_key),
        },
        "cabinetUsedFolderCount": row.cabinet_used_folder_count,
        "cabinetRemainingFolderCount": row.cabinet_remaining_folder_count,
        "recentYearOrderCount": recent_year_order_count,
        "cabinetUsageCheckedAt": row.cabinet_usage_checked_at.isoformat(sep=" ") if row.cabinet_usage_checked_at else None,
        "rakutenProductTotalCount": row.rakuten_product_total_count,
        "rakutenProductListedCount": row.rakuten_product_listed_count,
        "rakutenProductUnlistedCount": row.rakuten_product_unlisted_count,
        "lastCheckedAt": checked_at.isoformat(sep=" ") if checked_at else None,
        "lastProductSyncedAt": product_synced_at.isoformat(sep=" ") if product_synced_at else None,
        "lastSyncedAt": product_synced_at.isoformat(sep=" ") if product_synced_at else None,
        "lastError": row.last_error,
        "availabilityStatus": availability_status,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


def scheduled_crawl_to_public(row: ScheduledCrawlModel) -> dict[str, Any]:
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "sourceId": row.source_id,
        "name": row.name,
        "crawlContent": row.crawl_content,
        "crawlCondition": row.crawl_condition,
        "sourceType": row.source_type,
        "target": row.target,
        "enabled": bool(row.enabled),
        "intervalMinutes": row.interval_minutes,
        "scheduleTime": row.schedule_time,
        "lastRunAt": row.last_run_at.isoformat(sep=" ") if row.last_run_at else None,
        "nextRunAt": row.next_run_at.isoformat(sep=" ") if row.next_run_at else None,
        "status": row.status,
        "notes": row.notes,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


def auto_listing_schedule_to_public(
    row: AutoListingScheduleModel,
    store: StoreModel | None = None,
) -> dict[str, Any]:
    try:
        last_task_ids = json.loads(row.last_task_ids_json or "[]")
    except (TypeError, ValueError, json.JSONDecodeError):
        last_task_ids = []
    if not isinstance(last_task_ids, list):
        last_task_ids = []
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "storeId": row.store_id,
        "storeName": store.store_name if store is not None else "",
        "storeAliasName": store.alias_name if store is not None else "",
        "taskType": row.task_type or "automatic",
        "executionMode": (
            "scheduled"
            if row.task_type == "manual" and row.schedule_type == "once"
            else ("immediate" if row.task_type == "manual" else None)
        ),
        "scheduleType": row.schedule_type,
        "scheduleTime": row.schedule_time,
        "weekday": row.weekday,
        "monthDay": row.month_day,
        "quantity": row.quantity,
        "enabled": bool(row.enabled),
        "status": row.status,
        "lastRunAt": datetime_to_public(row.last_run_at),
        "nextRunAt": datetime_to_public(row.next_run_at),
        "lastTaskIds": [str(value) for value in last_task_ids if str(value).strip()],
        "lastMessage": row.last_message,
        "lastError": row.last_error,
        "createdAt": datetime_to_public(row.created_at),
        "updatedAt": datetime_to_public(row.updated_at),
    }


def list_auto_listing_schedules(
    owner_username: str,
    *,
    store_id: int | None = None,
    task_type: str | None = None,
) -> list[dict[str, Any]]:
    normalized_task_type = normalize_text(task_type).lower()
    if normalized_task_type not in {"", "automatic", "manual"}:
        raise RuntimeError("上架任务类型筛选条件无效。")
    with session_scope() as session:
        query = (
            select(AutoListingScheduleModel, StoreModel)
            .join(StoreModel, StoreModel.id == AutoListingScheduleModel.store_id)
            .where(AutoListingScheduleModel.owner_username == owner_username)
        )
        if store_id is not None:
            query = query.where(AutoListingScheduleModel.store_id == store_id)
        if normalized_task_type:
            query = query.where(AutoListingScheduleModel.task_type == normalized_task_type)
        rows = session.execute(
            query.order_by(
                *task_status_order_by(
                    AutoListingScheduleModel,
                    disabled_condition=and_(
                        AutoListingScheduleModel.task_type == "automatic",
                        AutoListingScheduleModel.enabled.is_(False),
                    ),
                ),
            )
        ).all()
        return [
            auto_listing_schedule_to_public(schedule, store)
            for schedule, store in rows
        ]


def create_auto_listing_schedule(owner_username: str, payload: Any) -> dict[str, Any]:
    schedule_type = normalize_text(getattr(payload, "scheduleType", None))
    schedule_time = normalize_schedule_time(getattr(payload, "scheduleTime", None))
    weekday = getattr(payload, "weekday", None)
    month_day = getattr(payload, "monthDay", None)
    quantity = int(getattr(payload, "quantity", 0) or 0)
    store_id = int(getattr(payload, "storeId", 0) or 0)
    if quantity < 1 or quantity > 10000:
        raise RuntimeError("自动上架数量只能设置为 1 到 10000。")
    next_run_at = next_auto_listing_run_at(
        schedule_type,
        schedule_time,
        weekday=weekday,
        month_day=month_day,
    )
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None or store.owner_username != owner_username:
            raise RuntimeError("上架店铺不存在或不属于当前用户。")
        if not store.enabled:
            raise RuntimeError(f"上架店铺「{store.alias_name or store.store_name}」已停用。")
        if (
            not decrypt_text(store.rakuten_service_secret_encrypted)
            or not decrypt_text(store.rakuten_license_key_encrypted)
        ):
            raise RuntimeError(f"上架店铺「{store.alias_name or store.store_name}」缺少乐天 Secret 或乐天 Key。")
        row = AutoListingScheduleModel(
            owner_username=owner_username,
            store_id=store_id,
            automatic_store_id=store_id,
            task_type="automatic",
            schedule_type=schedule_type,
            schedule_time=schedule_time,
            weekday=int(weekday) if schedule_type == "weekly" else None,
            month_day=int(month_day) if schedule_type == "monthly" else None,
            quantity=quantity,
            enabled=True,
            status="idle",
            next_run_at=next_run_at,
        )
        session.add(row)
        try:
            session.flush()
        except IntegrityError as exc:
            raise RuntimeError("该店铺已经创建过自动上架任务，不能重复创建。") from exc
        return auto_listing_schedule_to_public(row, store)


def create_manual_listing_task(owner_username: str, payload: Any) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    quantity = int(getattr(payload, "quantity", 0) or 0)
    store_id = int(getattr(payload, "storeId", 0) or 0)
    execution_mode, execute_at = manual_task_execution_settings(
        payload,
        task_name="手动上架任务",
    )
    if quantity < 1 or quantity > 10000:
        raise RuntimeError("上架数量只能设置为 1 到 10000。")
    with session_scope() as session:
        store = session.scalar(
            select(StoreModel)
            .where(StoreModel.id == store_id)
            .with_for_update()
        )
        if store is None or store.owner_username != owner_username:
            raise RuntimeError("上架店铺不存在或不属于当前用户。")
        if not store.enabled:
            raise RuntimeError(f"上架店铺「{store.alias_name or store.store_name}」已停用。")
        if (
            not decrypt_text(store.rakuten_service_secret_encrypted)
            or not decrypt_text(store.rakuten_license_key_encrypted)
        ):
            raise RuntimeError(f"上架店铺「{store.alias_name or store.store_name}」缺少乐天 Secret 或乐天 Key。")
        if execution_mode == "immediate":
            active_manual_task_id = session.scalar(
                select(AutoListingScheduleModel.id)
                .where(
                    AutoListingScheduleModel.owner_username == owner_username,
                    AutoListingScheduleModel.store_id == store_id,
                    AutoListingScheduleModel.task_type == "manual",
                    AutoListingScheduleModel.schedule_type == "",
                    AutoListingScheduleModel.status.in_(("idle", "running")),
                )
                .order_by(AutoListingScheduleModel.id.asc())
                .limit(1)
            )
            if active_manual_task_id is not None:
                raise RuntimeError("该店铺已有正在创建或执行的手动上架任务，请勿重复提交。")
        row = AutoListingScheduleModel(
            owner_username=owner_username,
            store_id=store_id,
            automatic_store_id=None,
            task_type="manual",
            schedule_type="once" if execution_mode == "scheduled" else "",
            schedule_time=execute_at.strftime("%H:%M") if execute_at else "",
            quantity=quantity,
            enabled=execution_mode == "scheduled",
            status="idle",
            next_run_at=execute_at,
            last_message=(
                f"任务将在 {datetime_to_public(execute_at)} 到期执行。"
                if execute_at
                else "任务已受理，后台正在创建上架任务。"
            ),
        )
        session.add(row)
        session.flush()
        schedule_id = int(row.id)
        if execution_mode == "scheduled":
            return auto_listing_schedule_to_public(row, store)
        created = auto_listing_schedule_to_public(row, store)
    try:
        dispatch_auto_listing_schedule(
            schedule_id,
            owner_username=owner_username,
            advance_next_run=False,
        )
    except Exception as exc:
        raise RuntimeError(f"上架任务投递失败：{exc}") from exc
    return created


def update_auto_listing_schedule_status(
    owner_username: str,
    schedule_id: int,
    enabled: bool,
) -> dict[str, Any]:
    with session_scope() as session:
        row = session.scalar(
            select(AutoListingScheduleModel).where(
                AutoListingScheduleModel.id == schedule_id,
                AutoListingScheduleModel.owner_username == owner_username,
            )
        )
        if row is None:
            raise RuntimeError("自动上架任务不存在。")
        if row.task_type != "automatic":
            raise RuntimeError("手动上架任务不支持启用或停用。")
        if row.status == "running":
            raise RuntimeError("自动上架任务正在执行，暂时不能启用或关闭。")
        store = session.get(StoreModel, row.store_id)
        row.enabled = bool(enabled)
        row.status = "idle" if enabled else "disabled"
        row.next_run_at = (
            next_auto_listing_run_at(
                row.schedule_type,
                row.schedule_time,
                weekday=row.weekday,
                month_day=row.month_day,
            )
            if enabled
            else None
        )
        row.last_error = None
        session.flush()
        return auto_listing_schedule_to_public(row, store)


def update_auto_listing_schedule(
    owner_username: str,
    schedule_id: int,
    payload: Any,
) -> dict[str, Any]:
    schedule_type = normalize_text(getattr(payload, "scheduleType", None))
    schedule_time = normalize_schedule_time(getattr(payload, "scheduleTime", None))
    weekday = getattr(payload, "weekday", None)
    month_day = getattr(payload, "monthDay", None)
    quantity = int(getattr(payload, "quantity", 0) or 0)
    if quantity < 1 or quantity > 10000:
        raise RuntimeError("自动上架数量只能设置为 1 到 10000。")
    next_run_at = next_auto_listing_run_at(
        schedule_type,
        schedule_time,
        weekday=weekday,
        month_day=month_day,
    )
    with session_scope() as session:
        row = session.scalar(
            select(AutoListingScheduleModel).where(
                AutoListingScheduleModel.id == schedule_id,
                AutoListingScheduleModel.owner_username == owner_username,
            )
        )
        if row is None:
            raise RuntimeError("自动上架任务不存在。")
        if row.task_type != "automatic":
            raise RuntimeError("手动上架任务不支持编辑。")
        if row.status == "running":
            raise RuntimeError("自动上架任务正在执行，暂时不能编辑。")
        store = session.get(StoreModel, row.store_id)
        row.schedule_type = schedule_type
        row.schedule_time = schedule_time
        row.weekday = int(weekday) if schedule_type == "weekly" else None
        row.month_day = int(month_day) if schedule_type == "monthly" else None
        row.quantity = quantity
        row.status = "idle" if row.enabled else "disabled"
        row.next_run_at = next_run_at if row.enabled else None
        row.last_error = None
        session.flush()
        return auto_listing_schedule_to_public(row, store)


def delete_auto_listing_schedule(owner_username: str, schedule_id: int) -> None:
    with session_scope() as session:
        row = session.scalar(
            select(AutoListingScheduleModel).where(
                AutoListingScheduleModel.id == schedule_id,
                AutoListingScheduleModel.owner_username == owner_username,
            )
        )
        if row is None:
            raise RuntimeError("自动上架任务不存在。")
        if row.status == "running" or (
            row.task_type == "manual"
            and row.schedule_type == ""
            and row.status == "idle"
        ):
            raise RuntimeError("该上架任务正在等待或执行，暂时不能删除。")
        session.delete(row)


def auto_deletion_task_to_public(row: AutoDeletionTaskModel, store: StoreModel | None = None) -> dict[str, Any]:
    try:
        task_ids = json.loads(row.last_task_ids_json or "[]")
    except (TypeError, ValueError, json.JSONDecodeError):
        task_ids = []
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "storeId": row.store_id,
        "storeName": store.store_name if store else "",
        "storeAliasName": store.alias_name if store else "",
        "taskType": row.task_type,
        "executionMode": (
            "scheduled"
            if row.task_type == "manual" and row.schedule_type == "once"
            else ("immediate" if row.task_type == "manual" else None)
        ),
        "scheduleType": row.schedule_type,
        "scheduleTime": row.schedule_time,
        "weekday": row.weekday,
        "monthDay": row.month_day,
        "quantity": row.quantity,
        "enabled": bool(row.enabled),
        "status": row.status,
        "lastRunAt": datetime_to_public(row.last_run_at),
        "nextRunAt": datetime_to_public(row.next_run_at),
        "lastTaskIds": task_ids if isinstance(task_ids, list) else [],
        "lastMessage": row.last_message,
        "lastError": row.last_error,
        "createdAt": datetime_to_public(row.created_at),
        "updatedAt": datetime_to_public(row.updated_at),
    }


def list_auto_deletion_tasks(owner_username: str, *, store_id: int | None = None, task_type: str | None = None) -> list[dict[str, Any]]:
    normalized_type = normalize_text(task_type).lower()
    if normalized_type not in {"", "automatic", "manual"}:
        raise RuntimeError("删除任务类型筛选条件无效。")
    with session_scope() as session:
        query = select(AutoDeletionTaskModel, StoreModel).join(StoreModel, StoreModel.id == AutoDeletionTaskModel.store_id).where(AutoDeletionTaskModel.owner_username == owner_username)
        if store_id is not None:
            query = query.where(AutoDeletionTaskModel.store_id == store_id)
        if normalized_type:
            query = query.where(AutoDeletionTaskModel.task_type == normalized_type)
        rows = session.execute(
            query.order_by(
                *task_status_order_by(
                    AutoDeletionTaskModel,
                    disabled_condition=and_(
                        AutoDeletionTaskModel.task_type == "automatic",
                        AutoDeletionTaskModel.enabled.is_(False),
                    ),
                )
            )
        ).all()
        return [auto_deletion_task_to_public(row, store) for row, store in rows]


def _validate_deletion_store(
    session: Any,
    owner_username: str,
    store_id: int,
    *,
    for_update: bool = False,
) -> StoreModel:
    query = select(StoreModel).where(StoreModel.id == store_id)
    if for_update:
        query = query.with_for_update()
    store = session.scalar(query)
    if store is None or store.owner_username != owner_username:
        raise RuntimeError("删除店铺不存在或不属于当前用户。")
    if not store.enabled:
        raise RuntimeError(f"店铺「{store.alias_name or store.store_name}」已停用。")
    if not decrypt_text(store.rakuten_service_secret_encrypted) or not decrypt_text(store.rakuten_license_key_encrypted):
        raise RuntimeError(f"店铺「{store.alias_name or store.store_name}」缺少乐天 Secret 或乐天 Key。")
    return store


def create_auto_deletion_task(owner_username: str, payload: Any) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    store_id = int(getattr(payload, "storeId", 0) or 0)
    quantity = int(getattr(payload, "quantity", 0) or 0)
    schedule_type = normalize_text(getattr(payload, "scheduleType", None))
    schedule_time = normalize_schedule_time(getattr(payload, "scheduleTime", None))
    weekday = getattr(payload, "weekday", None)
    month_day = getattr(payload, "monthDay", None)
    if quantity < 1 or quantity > 10000:
        raise RuntimeError("删除数量只能设置为 1 到 10000。")
    next_run = next_auto_listing_run_at(schedule_type, schedule_time, weekday=weekday, month_day=month_day)
    with session_scope() as session:
        store = _validate_deletion_store(session, owner_username, store_id)
        row = AutoDeletionTaskModel(owner_username=owner_username, store_id=store_id, automatic_store_id=store_id, task_type="automatic", schedule_type=schedule_type, schedule_time=schedule_time, weekday=int(weekday) if schedule_type == "weekly" else None, month_day=int(month_day) if schedule_type == "monthly" else None, quantity=quantity, enabled=True, status="idle", next_run_at=next_run)
        session.add(row)
        try:
            session.flush()
        except IntegrityError as exc:
            raise RuntimeError("该店铺已经创建过自动删除任务，不能重复创建。") from exc
        return auto_deletion_task_to_public(row, store)


def update_auto_deletion_task(
    owner_username: str,
    task_id: int,
    payload: Any,
) -> dict[str, Any]:
    schedule_type = normalize_text(getattr(payload, "scheduleType", None))
    schedule_time = normalize_schedule_time(getattr(payload, "scheduleTime", None))
    weekday = getattr(payload, "weekday", None)
    month_day = getattr(payload, "monthDay", None)
    quantity = int(getattr(payload, "quantity", 0) or 0)
    if quantity < 1 or quantity > 10000:
        raise RuntimeError("删除数量只能设置为 1 到 10000。")
    next_run_at = next_auto_listing_run_at(
        schedule_type,
        schedule_time,
        weekday=weekday,
        month_day=month_day,
    )
    with session_scope() as session:
        row = session.scalar(
            select(AutoDeletionTaskModel).where(
                AutoDeletionTaskModel.id == task_id,
                AutoDeletionTaskModel.owner_username == owner_username,
            )
        )
        if row is None:
            raise RuntimeError("自动删除任务不存在。")
        if row.task_type != "automatic":
            raise RuntimeError("手动删除任务不支持编辑。")
        if row.status == "running":
            raise RuntimeError("自动删除任务正在执行，暂时不能编辑。")
        store = session.get(StoreModel, row.store_id)
        row.schedule_type = schedule_type
        row.schedule_time = schedule_time
        row.weekday = int(weekday) if schedule_type == "weekly" else None
        row.month_day = int(month_day) if schedule_type == "monthly" else None
        row.quantity = quantity
        row.status = "idle" if row.enabled else "disabled"
        row.next_run_at = next_run_at if row.enabled else None
        row.last_error = None
        session.flush()
        return auto_deletion_task_to_public(row, store)


def update_auto_deletion_task_status(
    owner_username: str,
    task_id: int,
    enabled: bool,
) -> dict[str, Any]:
    with session_scope() as session:
        row = session.scalar(
            select(AutoDeletionTaskModel).where(
                AutoDeletionTaskModel.id == task_id,
                AutoDeletionTaskModel.owner_username == owner_username,
            )
        )
        if row is None:
            raise RuntimeError("自动删除任务不存在。")
        if row.task_type != "automatic":
            raise RuntimeError("手动删除任务不支持启用或关闭。")
        if row.status == "running":
            raise RuntimeError("自动删除任务正在执行，暂时不能启用或关闭。")
        store = session.get(StoreModel, row.store_id)
        row.enabled = bool(enabled)
        row.status = "idle" if enabled else "disabled"
        row.next_run_at = (
            next_auto_listing_run_at(
                row.schedule_type,
                row.schedule_time,
                weekday=row.weekday,
                month_day=row.month_day,
            )
            if enabled
            else None
        )
        row.last_error = None
        session.flush()
        return auto_deletion_task_to_public(row, store)


def delete_auto_deletion_task(owner_username: str, task_id: int) -> None:
    with session_scope() as session:
        row = session.scalar(select(AutoDeletionTaskModel).where(AutoDeletionTaskModel.id == task_id, AutoDeletionTaskModel.owner_username == owner_username))
        if row is None:
            raise RuntimeError("自动删除任务不存在。")
        if row.status == "running" or (
            row.task_type == "manual"
            and row.schedule_type == ""
            and row.status == "idle"
        ):
            raise RuntimeError("该删除任务正在等待或执行，暂时不能删除。")
        session.delete(row)


def normalize_listing_task_product_ids(value: Any) -> list[int]:
    if not isinstance(value, list):
        return []
    result: list[int] = []
    seen: set[int] = set()
    for item in value:
        try:
            product_id = int(item)
        except (TypeError, ValueError):
            continue
        if product_id in seen:
            continue
        seen.add(product_id)
        result.append(product_id)
    return result


def normalize_listing_task_store_ids(value: Any) -> list[int]:
    if not isinstance(value, list):
        return []
    result: list[int] = []
    seen: set[int] = set()
    for item in value:
        try:
            store_id = int(item)
        except (TypeError, ValueError):
            continue
        if store_id <= 0 or store_id in seen:
            continue
        seen.add(store_id)
        result.append(store_id)
    return result


def listing_task_payload_store_ids(payload: Any) -> list[int]:
    store_ids = normalize_listing_task_store_ids(getattr(payload, "storeIds", None))
    if store_ids:
        return store_ids
    store_id = getattr(payload, "storeId", None)
    try:
        normalized_store_id = int(store_id or 0)
    except (TypeError, ValueError):
        normalized_store_id = 0
    return [normalized_store_id] if normalized_store_id > 0 else []


def listing_task_product_ids_payload(product_ids_json: str | None) -> dict[str, list[int]]:
    try:
        product_ids_payload = json.loads(product_ids_json or "[]")
    except (TypeError, ValueError):
        product_ids_payload = []
    if isinstance(product_ids_payload, dict):
        return {
            "productIds": normalize_listing_task_product_ids(product_ids_payload.get("productIds")),
            "successIds": normalize_listing_task_product_ids(product_ids_payload.get("successIds")),
            "failedIds": normalize_listing_task_product_ids(product_ids_payload.get("failedIds")),
            "retryIds": normalize_listing_task_product_ids(product_ids_payload.get("retryIds")),
            "storeIds": normalize_listing_task_store_ids(product_ids_payload.get("storeIds")),
        }
    return {
        "productIds": normalize_listing_task_product_ids(product_ids_payload),
        "successIds": [],
        "failedIds": [],
        "retryIds": [],
        "storeIds": [],
    }


def merge_listing_task_product_ids(*groups: list[int]) -> list[int]:
    result: list[int] = []
    seen: set[int] = set()
    for group in groups:
        for product_id in group:
            if product_id in seen:
                continue
            seen.add(product_id)
            result.append(product_id)
    return result


def listing_task_result_payload(
    product_ids: list[int],
    success_ids: list[int],
    failed_ids: list[int],
    *,
    retry_ids: list[int] | None = None,
    store_ids: list[int] | None = None,
) -> dict[str, list[int]]:
    result = {
        "productIds": normalize_listing_task_product_ids(product_ids),
        "successIds": normalize_listing_task_product_ids(success_ids),
        "failedIds": normalize_listing_task_product_ids(failed_ids),
    }
    if retry_ids is not None:
        result["retryIds"] = normalize_listing_task_product_ids(retry_ids)
    if store_ids is not None:
        result["storeIds"] = normalize_listing_task_store_ids(store_ids)
    return result


def listing_task_retry_product_ids(task: ListingTaskModel) -> list[int]:
    product_ids_payload = listing_task_product_ids_payload(task.product_ids_json)
    failed_ids = product_ids_payload["failedIds"]
    if task.status in {"partial", "failed"} and failed_ids:
        return failed_ids
    return product_ids_payload["productIds"]


def listing_task_product_worker_count(product_count: int, *, retry: bool) -> int:
    configured_workers = (
        settings.listing_retry_product_workers
        if retry
        else settings.listing_product_workers
    )
    return min(max(0, int(product_count)), max(1, int(configured_workers)))


def listing_task_store_snapshot(row: StoreModel | None) -> dict[str, str]:
    if row is None:
        return {"storeCode": "", "storeName": "", "aliasName": ""}
    return {
        "storeCode": row.store_code,
        "storeName": row.store_name,
        "aliasName": row.alias_name,
    }


def listing_task_store_snapshot_by_id(store_id: int | None) -> dict[str, str]:
    if not store_id:
        return listing_task_store_snapshot(None)
    with session_scope() as session:
        return listing_task_store_snapshot(session.get(StoreModel, store_id))


def listing_task_store_snapshots(session: Any, rows: list[ListingTaskModel]) -> dict[int, dict[str, str]]:
    store_ids = sorted({int(row.store_id) for row in rows if row.store_id})
    if not store_ids:
        return {}
    stores = session.scalars(select(StoreModel).where(StoreModel.id.in_(store_ids))).all()
    return {int(store.id): listing_task_store_snapshot(store) for store in stores}


def listing_task_store_snapshots_by_ids(store_ids: list[int]) -> list[dict[str, Any]]:
    normalized_ids = normalize_listing_task_store_ids(store_ids)
    if not normalized_ids:
        return []
    with session_scope() as session:
        stores = session.scalars(select(StoreModel).where(StoreModel.id.in_(normalized_ids))).all()
        stores_by_id = {int(store.id): store for store in stores}
        result: list[dict[str, Any]] = []
        for store_id in normalized_ids:
            store = stores_by_id.get(store_id)
            if store is None:
                continue
            result.append({
                "storeId": int(store.id),
                "storeCode": store.store_code,
                "storeName": store.store_name,
                "aliasName": store.alias_name,
            })
        return result


def listing_task_to_public(
    row: ListingTaskModel,
    store_snapshot: dict[str, str] | object = _STORE_SNAPSHOT_UNSET,
) -> dict[str, Any]:
    if store_snapshot is _STORE_SNAPSHOT_UNSET:
        store_snapshot = listing_task_store_snapshot_by_id(row.store_id)
    if not isinstance(store_snapshot, dict):
        store_snapshot = listing_task_store_snapshot(None)
    product_ids_payload = listing_task_product_ids_payload(row.product_ids_json)
    product_ids = product_ids_payload["productIds"]
    success_ids = product_ids_payload["successIds"]
    failed_ids = product_ids_payload["failedIds"]
    store_ids = product_ids_payload["storeIds"] or ([int(row.store_id)] if row.store_id else [])
    listed_stores = listing_task_store_snapshots_by_ids(store_ids) if len(store_ids) > 1 else []
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "storeId": row.store_id,
        "storeIds": store_ids,
        "stores": listed_stores,
        "storeCode": store_snapshot.get("storeCode", ""),
        "storeName": store_snapshot.get("storeName", ""),
        "aliasName": store_snapshot.get("aliasName", ""),
        "taskName": row.task_name,
        "taskGroupId": row.task_group_id,
        "taskGroupIndex": row.task_group_index,
        "taskGroupSize": row.task_group_size,
        "status": row.status,
        "totalCount": row.total_count,
        "successCount": row.success_count,
        "failedCount": row.failed_count,
        "productIds": product_ids,
        "successIds": success_ids,
        "failedIds": failed_ids,
        "cancelRequested": task_cancel_requested(row),
        "message": row.message,
        "errorDetail": task_public_error_detail(row),
        "startedAt": row.started_at.isoformat(sep=" ") if row.started_at else None,
        "finishedAt": row.finished_at.isoformat(sep=" ") if row.finished_at else None,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


def sync_task_payload_product_count(payload: Any) -> int:
    if not isinstance(payload, dict):
        return 0
    values = payload.get("cleanupRecordIds") if payload.get("cleanupRecordIds") is not None else payload.get("productIds")
    return len(normalize_listing_task_product_ids(values))


def sync_task_known_total_count(row: SyncTaskModel, payload: Any | None = None) -> int:
    total_count = max(0, int(row.total_count or 0))
    if payload is None:
        payload = sync_task_payload(row)
    payload_count = sync_task_payload_product_count(payload)
    return max(total_count, payload_count)


def sync_task_to_public(row: SyncTaskModel) -> dict[str, Any]:
    payload = sync_task_payload(row)
    task_result = payload.get("result") if isinstance(payload, dict) and isinstance(payload.get("result"), dict) else {}
    total_count = sync_task_known_total_count(row, payload)
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "storeId": row.store_id,
        "storeName": row.store_name,
        "taskName": row.task_name,
        "taskGroupId": row.task_group_id,
        "taskGroupIndex": row.task_group_index,
        "taskGroupSize": row.task_group_size,
        "taskType": row.task_type,
        "status": row.status,
        "totalCount": total_count,
        "successCount": row.success_count,
        "failedCount": row.failed_count,
        "payload": payload if isinstance(payload, dict) else {},
        "successIds": task_result.get("successIds") if isinstance(task_result.get("successIds"), list) else [],
        "failedIds": task_result.get("failedIds") if isinstance(task_result.get("failedIds"), list) else [],
        "cancelRequested": task_cancel_requested(row),
        "message": row.message,
        "errorDetail": task_public_error_detail(row),
        "startedAt": row.started_at.isoformat(sep=" ") if row.started_at else None,
        "finishedAt": row.finished_at.isoformat(sep=" ") if row.finished_at else None,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


TASK_GROUP_NAME_PATTERN = re.compile(r"^(?P<prefix>.*?)(?:\s+)(?P<index>\d+)/(?P<size>\d+)(?P<suffix>(?:\s+.*)?)$")
TASK_GROUP_TERMINAL_STATUSES = {"success", "partial", "failed", "cancelled", "completed"}


def task_group_name_parts(task_name: str) -> tuple[str, int, int] | None:
    match = TASK_GROUP_NAME_PATTERN.fullmatch(normalize_text(task_name))
    if match is None:
        return None
    index = int(match.group("index"))
    size = int(match.group("size"))
    if size <= 1 or index < 1 or index > size:
        return None
    base_name = f"{match.group('prefix')}{match.group('suffix')}".strip()
    return base_name, index, size


def task_group_status(rows: list[Any]) -> str:
    statuses = [normalize_text(row.status) for row in rows]
    if any(status in {"running", "processing"} for status in statuses):
        return "running"
    if any(status in {"queued", "idle", "pending", "waiting", "preview_ready"} for status in statuses):
        return "queued"
    if statuses and all(status in {"success", "completed"} for status in statuses):
        return "success"
    if statuses and all(status == "cancelled" for status in statuses):
        return "cancelled"
    if any(status in {"success", "completed", "partial"} for status in statuses):
        return "partial"
    return "failed"


def task_group_status_priority(status: str) -> int:
    if status in {"running", "processing"}:
        return 0
    if status in {"queued", "idle", "pending", "waiting", "preview_ready"}:
        return 1
    if status in {"failed", "partial", "error", "cancelled"}:
        return 2
    if status in {"success", "completed"}:
        return 3
    return 5


def task_group_rows(rows: list[Any]) -> list[list[Any]]:
    grouped_rows: dict[tuple[Any, ...], list[Any]] = {}
    standalone_rows: list[list[Any]] = []
    legacy_candidates: dict[tuple[Any, ...], list[tuple[Any, int, int]]] = {}

    for row in rows:
        if normalize_text(getattr(row, "task_group_id", None)):
            key = ("persisted", row.owner_username, row.task_group_id)
            grouped_rows.setdefault(key, []).append(row)
            continue
        name_parts = task_group_name_parts(row.task_name)
        if name_parts is None:
            standalone_rows.append([row])
            continue
        base_name, group_index, group_size = name_parts
        key = (
            "legacy",
            row.owner_username,
            getattr(row, "store_id", None),
            getattr(row, "task_type", None),
            row.created_at,
            base_name,
            group_size,
        )
        legacy_candidates.setdefault(key, []).append((row, group_index, group_size))

    for key, candidates in legacy_candidates.items():
        expected_size = candidates[0][2]
        indexes = [candidate[1] for candidate in candidates]
        if len(candidates) == expected_size and sorted(indexes) == list(range(1, expected_size + 1)):
            grouped_rows[key] = [candidate[0] for candidate in candidates]
        else:
            standalone_rows.extend([[candidate[0]] for candidate in candidates])

    result = [*standalone_rows, *grouped_rows.values()]
    for group in result:
        group.sort(
            key=lambda row: (
                int(getattr(row, "task_group_index", 0) or (task_group_name_parts(row.task_name) or ("", 0, 0))[1]),
                row.created_at or datetime.min,
                row.id,
            )
        )
    result.sort(key=lambda group: min(row.id for row in group))
    result.sort(
        key=lambda group: max((row.created_at or datetime.min) for row in group),
        reverse=True,
    )
    result.sort(key=lambda group: task_group_status_priority(task_group_status(group)))
    return result


def task_group_to_public(
    rows: list[Any],
    *,
    serializer: Callable[[Any], dict[str, Any]],
    group_kind: str,
) -> dict[str, Any]:
    children = [serializer(row) for row in rows]
    if len(children) == 1:
        return children[0]

    for row, child in zip(rows, children):
        name_parts = task_group_name_parts(row.task_name)
        child["taskGroupIndex"] = (
            int(getattr(row, "task_group_index", 0) or 0)
            or (name_parts[1] if name_parts else None)
        )
        child["taskGroupSize"] = (
            int(getattr(row, "task_group_size", 0) or 0)
            or (name_parts[2] if name_parts else len(rows))
        )

    first_row = rows[0]
    first_child = children[0]
    name_parts = task_group_name_parts(first_row.task_name)
    task_name = name_parts[0] if name_parts else first_row.task_name
    group_id = normalize_text(getattr(first_row, "task_group_id", None)) or hashlib.sha256(
        "|".join(
            [
                str(first_row.owner_username),
                str(getattr(first_row, "store_id", "") or ""),
                str(getattr(first_row, "task_type", "") or ""),
                str(first_row.created_at or ""),
                task_name,
            ]
        ).encode("utf-8")
    ).hexdigest()[:32]
    status = task_group_status(rows)
    completed_count = sum(
        1 for row in rows if normalize_text(row.status) in TASK_GROUP_TERMINAL_STATUSES
    )
    failed_task_count = sum(
        1 for row in rows if normalize_text(row.status) in {"failed", "partial", "error", "cancelled"}
    )
    started_values = [row.started_at for row in rows if row.started_at is not None]
    finished_values = [row.finished_at for row in rows if row.finished_at is not None]
    all_terminal = all(normalize_text(row.status) in TASK_GROUP_TERMINAL_STATUSES for row in rows)
    aggregate = {
        **first_child,
        "id": f"group:{group_kind}:{group_id}",
        "taskGroupId": group_id,
        "taskGroupIndex": None,
        "taskGroupSize": len(rows),
        "taskName": task_name,
        "status": status,
        "totalCount": sum(max(0, int(child.get("totalCount") or 0)) for child in children),
        "successCount": sum(max(0, int(child.get("successCount") or 0)) for child in children),
        "failedCount": sum(max(0, int(child.get("failedCount") or 0)) for child in children),
        "cancelRequested": any(bool(child.get("cancelRequested")) for child in children),
        "message": f"共 {len(rows)} 个分任务，已完成 {completed_count} / {len(rows)} 个",
        "errorDetail": f"{failed_task_count} 个分任务存在失败或终止，请展开查看。" if failed_task_count else None,
        "startedAt": min(started_values).isoformat(sep=" ") if started_values else None,
        "finishedAt": max(finished_values).isoformat(sep=" ") if all_terminal and finished_values else None,
        "createdAt": min(row.created_at for row in rows if row.created_at is not None).isoformat(sep=" "),
        "updatedAt": max(row.updated_at for row in rows if row.updated_at is not None).isoformat(sep=" "),
        "isGroup": True,
        "childTaskIds": [row.id for row in rows],
        "children": children,
    }
    if group_kind == "listing":
        aggregate["productIds"] = [
            product_id
            for child in children
            for product_id in child.get("productIds") or []
        ]
        aggregate["successIds"] = [
            product_id
            for child in children
            for product_id in child.get("successIds") or []
        ]
        aggregate["failedIds"] = [
            product_id
            for child in children
            for product_id in child.get("failedIds") or []
        ]
    return aggregate


def paginate_task_groups(
    rows: list[Any],
    *,
    page: int | None,
    page_size: int | None,
    response_key: str,
    serializer: Callable[[Any], dict[str, Any]],
    group_kind: str,
) -> list[dict[str, Any]] | dict[str, Any]:
    groups = task_group_rows(rows)
    normalized_page, normalized_page_size = normalize_page_params(page, page_size)
    if not normalized_page_size:
        return [
            task_group_to_public(group, serializer=serializer, group_kind=group_kind)
            for group in groups
        ]
    total = len(groups)
    if total:
        max_page = max(1, (total + normalized_page_size - 1) // normalized_page_size)
        normalized_page = min(normalized_page, max_page)
    start = (normalized_page - 1) * normalized_page_size
    page_groups = groups[start : start + normalized_page_size]
    return {
        response_key: [
            task_group_to_public(group, serializer=serializer, group_kind=group_kind)
            for group in page_groups
        ],
        "total": total,
        "page": normalized_page,
        "pageSize": normalized_page_size,
    }


def update_task_progress(
    model: Any,
    task_id: str,
    *,
    total_count: int | None = None,
    success_count: int | None = None,
    failed_count: int | None = None,
    warning_count: int | None = None,
    saved_count: int | None = None,
    skipped_count: int | None = None,
    message: str | None = None,
    status: str | None = None,
    error_detail: Any = _TASK_DETAIL_UNSET,
    warning_detail: Any = _TASK_DETAIL_UNSET,
) -> None:
    last_error: OperationalError | None = None
    for attempt in range(3):
        try:
            with session_scope() as session:
                task = session.get(model, task_id)
                if task is None:
                    return
                cancel_requested = task_cancel_requested(task)
                if total_count is not None:
                    task.total_count = max(0, int(total_count))
                if success_count is not None:
                    task.success_count = max(0, int(success_count))
                if failed_count is not None:
                    task.failed_count = max(0, int(failed_count))
                if warning_count is not None and hasattr(task, "warning_count"):
                    task.warning_count = max(0, int(warning_count))
                if saved_count is not None and hasattr(task, "saved_count"):
                    task.saved_count = max(0, int(saved_count))
                if skipped_count is not None and hasattr(task, "skipped_count"):
                    task.skipped_count = max(0, int(skipped_count))
                if message is not None and not cancel_requested:
                    task.message = message
                if status is not None and not cancel_requested:
                    task.status = status
                if error_detail is not _TASK_DETAIL_UNSET and hasattr(task, "error_detail"):
                    task.error_detail = with_task_cancel_marker(error_detail) if cancel_requested else error_detail
                if warning_detail is not _TASK_DETAIL_UNSET and hasattr(task, "warning_detail"):
                    task.warning_detail = warning_detail
            return
        except OperationalError as exc:
            last_error = exc
            if not is_mysql_retryable_lock_error(exc) or attempt >= 2:
                raise
            time.sleep(0.25 * (attempt + 1))
    if last_error is not None:
        raise last_error


def task_cancel_requested(row: Any | None) -> bool:
    if row is None:
        return False
    detail = normalize_task_detail_text(getattr(row, "error_detail", ""))
    if not detail:
        return False
    if detail == TASK_CANCEL_REQUESTED_MARKER or detail.startswith(f"{TASK_CANCEL_REQUESTED_MARKER} "):
        return True
    first_line = detail.splitlines()[0].strip()
    return first_line == TASK_CANCEL_REQUESTED_MARKER


def task_public_error_detail(row: Any | None) -> str | None:
    if row is None:
        return None
    detail = strip_task_cancel_marker(getattr(row, "error_detail", ""))
    return humanize_task_error_detail(detail)


def task_public_warning_detail(
    row: Any | None,
    *,
    skipped_count: int = 0,
    warning_count: int = 0,
) -> str | None:
    if row is None:
        return None
    detail = normalize_task_detail_text(getattr(row, "warning_detail", ""))
    return humanize_task_warning_detail(
        detail,
        skipped_count=skipped_count,
        warning_count=warning_count,
    )


def humanize_task_error_detail(error_detail: Any) -> str | None:
    detail = normalize_task_detail_text(error_detail)
    if not detail:
        return None
    normalized = detail.lower()
    if "lock wait timeout exceeded" in normalized or "deadlock found" in normalized:
        return (
            "原因：数据库当时正在处理其他写入，商品保存等待超时。\n"
            "影响：任务提前结束，但此前已经入库的商品不会丢失。\n"
            "处理建议：请重新执行任务；系统会自动跳过已有商品，并对短暂锁等待自动重试。"
        )
    if "duplicate entry" in normalized or ("1062" in normalized and "integrityerror" in normalized):
        return (
            "原因：待保存的数据与系统中的现有记录重复。\n"
            "影响：重复数据没有再次写入。\n"
            "处理建议：通常无需处理；刷新列表确认原记录即可。"
        )
    if (
        "can't connect to mysql" in normalized
        or "mysql server has gone away" in normalized
        or "lost connection to mysql" in normalized
    ):
        return (
            "原因：任务执行时暂时无法连接数据库。\n"
            "影响：任务未能继续处理。\n"
            "处理建议：请稍后重新执行；如果连续出现，请联系管理员检查数据库服务。"
        )
    if "proxyerror" in normalized or "proxy connection" in normalized:
        return (
            "原因：采集代理暂时无法连接目标网站。\n"
            "影响：部分或全部商品未能获取。\n"
            "处理建议：请稍后重新执行；如果连续出现，请联系管理员检查代理配置。"
        )
    if "too many requests" in normalized or re.search(
        r"(?:http|status|response)[^0-9]{0,20}429\b",
        normalized,
    ):
        return (
            "原因：目标网站限制了过于频繁的请求。\n"
            "影响：部分数据暂时未能获取。\n"
            "处理建议：请稍后重新执行，系统会继续按照限速规则采集。"
        )
    if "unauthorized" in normalized or re.search(
        r"(?:http|status|response)[^0-9]{0,20}401\b",
        normalized,
    ):
        return (
            "原因：店铺授权信息无效或已经过期。\n"
            "影响：系统无法访问对应店铺数据。\n"
            "处理建议：请检查店铺授权信息并重新保存，然后再次执行任务。"
        )
    if "forbidden" in normalized or re.search(
        r"(?:http|status|response)[^0-9]{0,20}403\b",
        normalized,
    ):
        return (
            "原因：目标服务拒绝了当前访问请求。\n"
            "影响：相关数据未能获取或提交。\n"
            "处理建议：请检查店铺权限、授权范围或采集网络后重新执行。"
        )
    if "read timed out" in normalized or "readtimeout" in normalized or "connecttimeout" in normalized:
        return (
            "原因：访问目标网站时等待响应超时。\n"
            "影响：当前请求未完成。\n"
            "处理建议：请稍后重新执行；已完成的数据不会重复处理。"
        )
    if "connection refused" in normalized or "failed to establish a new connection" in normalized:
        return (
            "原因：目标服务当前无法连接。\n"
            "影响：任务未能继续处理。\n"
            "处理建议：请稍后重新执行；如果连续出现，请联系管理员检查相关服务。"
        )
    if "jsondecodeerror" in normalized or "expecting value:" in normalized:
        return (
            "原因：目标服务返回的数据格式异常。\n"
            "影响：当前请求结果无法识别。\n"
            "处理建议：请稍后重新执行；如果连续出现，请联系管理员。"
        )

    visible_lines: list[str] = []
    for line in detail.splitlines():
        normalized_line = line.strip()
        if not normalized_line:
            continue
        lowered_line = normalized_line.lower()
        if (
            normalized_line.startswith("[SQL:")
            or normalized_line.startswith("[parameters:")
            or lowered_line.startswith("(background on this error")
            or lowered_line.startswith("traceback (most recent call last)")
            or re.match(r'^file ".+", line \d+', lowered_line)
        ):
            break
        visible_lines.append(normalized_line)
    visible_detail = "\n".join(visible_lines).strip()
    if visible_detail and re.search(r"[\u4e00-\u9fff]", visible_detail):
        return visible_detail
    return (
        "原因：任务执行过程中发生系统异常。\n"
        "影响：任务未能继续完成。\n"
        "处理建议：请稍后重新执行；如果重复出现，请联系管理员并提供任务编号。"
    )


def humanize_task_warning_detail(
    warning_detail: Any,
    *,
    skipped_count: int = 0,
    warning_count: int = 0,
) -> str | None:
    detail = normalize_task_detail_text(warning_detail)
    if not detail:
        return None
    reason_labels: list[str] = []
    reason_markers = (
        ("已存在于", "商品此前已经采集，为避免重复入库已自动跳过"),
        ("不符合", "商品价格不符合本次设置的价格条件"),
        ("禁止采集", "商品所属品类被当前用户设置为禁止采集"),
        ("商品详情采集失败", "部分商品详情获取失败，系统使用了列表页可用信息"),
        ("图片本地化失败", "部分商品图片保存失败，但商品基础信息已经入库"),
    )
    for marker, label in reason_markers:
        if marker in detail and label not in reason_labels:
            reason_labels.append(label)
    affected_count = skipped_count or warning_count
    if affected_count > 0 and reason_labels:
        reason_text = "\n".join(f"- {label}。" for label in reason_labels)
        summary = (
            f"本次有 {skipped_count} 件商品未入库，这些属于正常过滤或可恢复提示，不影响其他商品。"
            if skipped_count > 0
            else f"本次有 {warning_count} 件商品需要注意，但任务中的其他商品已正常处理。"
        )
        return (
            f"{summary}\n"
            f"主要原因：\n{reason_text}\n"
            "处理建议：重复商品无需处理；如需采集被价格或品类条件过滤的商品，请调整任务设置后重新执行。"
        )
    return humanize_task_error_detail(detail)


def strip_task_cancel_marker(error_detail: Any) -> str:
    detail = normalize_task_detail_text(error_detail)
    if not detail:
        return ""
    if detail == TASK_CANCEL_REQUESTED_MARKER:
        return ""
    if detail.startswith(f"{TASK_CANCEL_REQUESTED_MARKER} "):
        return detail[len(TASK_CANCEL_REQUESTED_MARKER):].strip()
    lines = detail.splitlines()
    if lines and lines[0].strip() == TASK_CANCEL_REQUESTED_MARKER:
        return "\n".join(line for line in lines[1:] if line.strip()).strip()
    return detail


def with_task_cancel_marker(error_detail: Any) -> str:
    detail = normalize_task_detail_text(error_detail)
    if detail == TASK_CANCEL_REQUESTED_MARKER or detail.startswith(f"{TASK_CANCEL_REQUESTED_MARKER} "):
        return detail
    if detail.splitlines() and detail.splitlines()[0].strip() == TASK_CANCEL_REQUESTED_MARKER:
        return detail
    return f"{TASK_CANCEL_REQUESTED_MARKER}\n{detail}".strip() if detail else TASK_CANCEL_REQUESTED_MARKER


def normalize_task_detail_text(value: Any) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    return "\n".join(line.strip() for line in text.splitlines() if line.strip())


def cancelled_task_error_detail(errors: list[str] | None = None, existing_error_detail: Any = None) -> str | None:
    values: list[str] = []
    existing = strip_task_cancel_marker(existing_error_detail)
    if existing:
        values.extend(existing.splitlines())
    values.extend(str(error or "").strip() for error in (errors or []) if str(error or "").strip())
    return summarize_task_errors(values, limit=50)


def cancelled_task_warning_detail(warnings: list[str] | None = None, existing_warning_detail: Any = None) -> str | None:
    values: list[str] = []
    existing = normalize_task_detail_text(existing_warning_detail)
    if existing:
        values.extend(existing.splitlines())
    values.extend(str(warning or "").strip() for warning in (warnings or []) if str(warning or "").strip())
    return summarize_task_errors(values, limit=50, item_label="警告")


def sync_task_payload(row: SyncTaskModel) -> dict[str, Any]:
    try:
        payload = json.loads(row.payload_json or "{}")
    except ValueError:
        payload = {}
    return payload if isinstance(payload, dict) else {}


@dataclass(frozen=True)
class TaskResourceScope:
    owner_username: str
    store_ids: frozenset[int]
    product_ids: frozenset[int]
    store_wide: bool = False


def logical_product_ids(
    session: Any,
    owner_username: str,
    product_ids: list[int],
) -> set[int]:
    normalized_ids = normalize_product_ids(product_ids)
    if not normalized_ids:
        return set()
    rows = session.execute(
        select(ProductModel.id, ProductModel.parent_product_id).where(
            ProductModel.owner_username == owner_username,
            ProductModel.id.in_(normalized_ids),
        )
    ).all()
    roots_by_id = {
        int(row.id): int(row.parent_product_id or row.id)
        for row in rows
    }
    return {
        roots_by_id.get(product_id, product_id)
        for product_id in normalized_ids
    }


def sync_task_resource_product_ids(task: SyncTaskModel) -> list[int]:
    payload = sync_task_payload(task)
    task_type = normalize_text(task.task_type) or "store_sync"
    if task_type == "product_replace":
        return normalize_product_ids(
            [
                payload.get("targetProductId"),
                payload.get("pendingProductId"),
            ]
        )
    return normalize_product_ids(list(payload.get("productIds") or []))


def sync_task_resource_scope(session: Any, task: SyncTaskModel) -> TaskResourceScope:
    task_type = normalize_text(task.task_type) or "store_sync"
    store_ids = frozenset({int(task.store_id)}) if task.store_id else frozenset()
    product_ids = (
        logical_product_ids(
            session,
            task.owner_username,
            sync_task_resource_product_ids(task),
        )
        if task_type in PRODUCT_SCOPED_SYNC_TASK_TYPES
        else set()
    )
    return TaskResourceScope(
        owner_username=task.owner_username,
        store_ids=store_ids,
        product_ids=frozenset(product_ids),
        store_wide=task_type not in PRODUCT_SCOPED_SYNC_TASK_TYPES,
    )


def listing_task_resource_scope(
    session: Any,
    task: ListingTaskModel,
) -> TaskResourceScope:
    payload = listing_task_product_ids_payload(task.product_ids_json)
    store_ids = payload["storeIds"] or ([int(task.store_id)] if task.store_id else [])
    product_ids = payload["retryIds"] or payload["productIds"]
    return TaskResourceScope(
        owner_username=task.owner_username,
        store_ids=frozenset(store_ids),
        product_ids=frozenset(
            logical_product_ids(session, task.owner_username, product_ids)
        ),
    )


def task_resource_scopes_conflict(
    left: TaskResourceScope,
    right: TaskResourceScope,
) -> bool:
    if left.owner_username != right.owner_username:
        return False
    if left.store_wide or right.store_wide:
        return bool(left.store_ids & right.store_ids)
    return bool(left.product_ids & right.product_ids)


def task_scope_conflict_reason(
    scope: TaskResourceScope,
    active_scopes: list[TaskResourceScope],
) -> str:
    for active_scope in active_scopes:
        if not task_resource_scopes_conflict(scope, active_scope):
            continue
        if scope.store_wide or active_scope.store_wide:
            return "排队中，等待同店铺整店任务完成"
        return "排队中，等待同一商品当前任务完成"
    return ""


def partition_product_ids_by_active_task_conflicts(
    session: Any,
    owner_username: str,
    store_ids: list[int],
    product_ids: list[int],
) -> tuple[list[int], list[int]]:
    normalized_ids = normalize_product_ids(product_ids)
    if not normalized_ids:
        return [], []
    active_scopes = [
        sync_task_resource_scope(session, task)
        for task in session.scalars(
            select(SyncTaskModel).where(
                SyncTaskModel.owner_username == owner_username,
                SyncTaskModel.status.in_(("queued", "running")),
                SyncTaskModel.task_type.notin_(IMAGE_CLEANUP_TASK_TYPES),
            )
        ).all()
    ]
    active_scopes.extend(
        listing_task_resource_scope(session, task)
        for task in session.scalars(
            select(ListingTaskModel).where(
                ListingTaskModel.owner_username == owner_username,
                ListingTaskModel.status.in_(("queued", "running")),
            )
        ).all()
    )
    roots_by_id: dict[int, int] = {}
    rows = session.execute(
        select(ProductModel.id, ProductModel.parent_product_id).where(
            ProductModel.owner_username == owner_username,
            ProductModel.id.in_(normalized_ids),
        )
    ).all()
    for row in rows:
        roots_by_id[int(row.id)] = int(row.parent_product_id or row.id)
    ready_ids: list[int] = []
    blocked_ids: list[int] = []
    normalized_store_ids = frozenset(normalize_listing_task_store_ids(store_ids))
    for product_id in normalized_ids:
        scope = TaskResourceScope(
            owner_username=owner_username,
            store_ids=normalized_store_ids,
            product_ids=frozenset({roots_by_id.get(product_id, product_id)}),
        )
        if task_scope_conflict_reason(scope, active_scopes):
            blocked_ids.append(product_id)
        else:
            ready_ids.append(product_id)
    return ready_ids, blocked_ids


def conflict_aware_product_chunks(
    ready_ids: list[int],
    blocked_ids: list[int],
) -> list[tuple[list[int], bool]]:
    result: list[tuple[list[int], bool]] = []
    for product_ids, blocked in ((ready_ids, False), (blocked_ids, True)):
        for chunk_ids in chunk_product_ids(product_ids):
            if chunk_ids:
                result.append((chunk_ids, blocked))
    return result


def sync_task_action_label(row: SyncTaskModel) -> str:
    task_type = row.task_type or "store_sync"
    if task_type == "title_optimization":
        return "标题优化"
    if task_type == "product_delete":
        return "删除"
    if task_type == "deleted_product_image_cleanup":
        return "图片清理"
    if task_type == "listing_image_upload":
        return "图片上传"
    if task_type in {"listing_status", "product_listing_status"}:
        payload = sync_task_payload(row)
        listing_status = normalize_text(payload.get("listingStatus"))
        return "上架" if listing_status == "listed" else "下架"
    return "同步"


def failed_sync_task_progress_message(
    action_label: str,
    total_count: int,
    success_count: int,
    failed_count: int,
    *,
    unfinished_count: int = 0,
) -> str:
    total = max(0, int(total_count or 0))
    success = max(0, int(success_count or 0))
    failed = max(0, int(failed_count or 0))
    unfinished = max(0, int(unfinished_count or 0))
    if total > 0:
        processed = min(total, success + failed)
        return (
            f"{action_label}异常中断，已处理 {processed} / {total} 条，"
            f"成功 {success} 条，异常 {failed} 条，未完成 {unfinished} 条"
        )
    return f"{action_label}异常中断"


def interrupted_task_error_detail(reason: str, existing_error_detail: Any = None) -> str | None:
    values: list[str] = []
    existing = strip_task_cancel_marker(existing_error_detail)
    if existing:
        values.extend(existing.splitlines())
    if reason:
        values.append(reason)
    return summarize_task_errors(values, limit=50)


def finalize_stale_cancel_requested_tasks(
    session: Any,
    model: Any,
    *,
    action_label: str,
    owner_username: str | None = None,
    store_id: int | None = None,
) -> int:
    cutoff = datetime.now() - timedelta(seconds=TASK_STALE_CANCEL_REQUEST_SECONDS)
    query = select(model).where(
        model.status == "running",
        model.updated_at <= cutoff,
    )
    if owner_username is not None:
        query = query.where(model.owner_username == owner_username)
    if store_id is not None and hasattr(model, "store_id"):
        query = query.where(model.store_id == store_id)
    rows = session.scalars(query).all()
    finalized = 0
    for task in rows:
        if not task_cancel_requested(task):
            continue
        if model is ListingTaskModel:
            release_listing_task_locks(session, task.owner_username, task)
        total_count = int(getattr(task, "total_count", 0) or 0)
        success_count = int(getattr(task, "success_count", 0) or 0)
        failed_count = int(getattr(task, "failed_count", 0) or 0)
        task.status = "cancelled"
        task.message = cancelled_task_progress_message(action_label, total_count, success_count, failed_count)
        task.error_detail = cancelled_task_error_detail(existing_error_detail=getattr(task, "error_detail", None))
        if hasattr(task, "warning_detail"):
            task.warning_detail = cancelled_task_warning_detail(existing_warning_detail=getattr(task, "warning_detail", None))
        task.finished_at = datetime.now()
        finalized += 1
    return finalized


def reconcile_interrupted_running_tasks(
    session: Any,
    model: Any,
    *,
    owner_username: str | None = None,
    store_id: int | None = None,
) -> int:
    if not should_use_redis_task_queue():
        return 0
    queue_kind = task_model_queue_kind(model)
    if not queue_kind:
        return 0
    query = select(model).where(model.status == "running")
    if owner_username is not None:
        query = query.where(model.owner_username == owner_username)
    if store_id is not None and hasattr(model, "store_id"):
        query = query.where(model.store_id == store_id)
    rows = session.scalars(query).all()
    if not rows:
        return 0
    try:
        if model is CrawlTaskModel:
            task_ids = {str(row.id) for row in rows}
            task_states: dict[str, dict[str, Any]] = {}
            for row_queue_kind in ("manual-crawl", "scheduled-crawl", "crawl"):
                task_states.update(
                    redis_task_states(
                        task_ids,
                        row_queue_kind,
                        job_id_prefixes=task_model_job_id_prefixes(model),
                    )
                )
        elif model is SyncTaskModel:
            task_states: dict[str, dict[str, Any]] = {}
            rows_by_queue: dict[str, set[str]] = {}
            for row in rows:
                rows_by_queue.setdefault(
                    sync_task_queue_kind(row.task_type),
                    set(),
                ).add(str(row.id))
            for row_queue_kind, task_ids in rows_by_queue.items():
                task_states.update(
                    redis_task_states(
                        task_ids,
                        row_queue_kind,
                        job_id_prefixes=task_model_job_id_prefixes(model),
                    )
                )
        else:
            task_ids = {str(row.id) for row in rows}
            task_states = redis_task_states(
                task_ids,
                queue_kind,
                job_id_prefixes=task_model_job_id_prefixes(model),
            )
    except Exception:
        return 0
    finalized = 0
    missing_cutoff = interrupted_task_missing_cutoff()
    for task in rows:
        task_id = str(task.id)
        state = task_states.get(task_id)
        if state and state.get("status") == "failed":
            finalize_interrupted_task(
                session,
                model,
                task,
                redis_failed_task_reason(state, task_model_action_label(model, task)),
            )
            finalized += 1
            continue
        if state is not None:
            continue
        task_touched_at = task.updated_at or task.started_at or task.created_at
        if task_touched_at and task_touched_at > missing_cutoff:
            continue
        finalize_interrupted_task(
            session,
            model,
            task,
            redis_missing_task_reason(task_model_action_label(model, task)),
        )
        finalized += 1
    return finalized


def reconcile_redis_failed_sync_tasks(
    session: Any,
    *,
    owner_username: str | None = None,
    store_id: int | None = None,
) -> int:
    return reconcile_interrupted_running_tasks(
        session,
        SyncTaskModel,
        owner_username=owner_username,
        store_id=store_id,
    )


def reconcile_interrupted_scheduled_crawls(session: Any, *, owner_username: str | None = None) -> int:
    if not should_use_redis_task_queue():
        return 0
    query = select(ScheduledCrawlModel).where(ScheduledCrawlModel.status == "running")
    if owner_username is not None:
        query = query.where(ScheduledCrawlModel.owner_username == owner_username)
    rows = session.scalars(query).all()
    if not rows:
        return 0
    schedule_ids = {str(row.id) for row in rows}
    try:
        schedule_states = redis_task_states(schedule_ids, "schedule", job_id_prefixes=("schedule-",))
    except Exception:
        return 0
    finalized = 0
    missing_cutoff = interrupted_task_missing_cutoff()
    for row in rows:
        schedule_id = str(row.id)
        state = schedule_states.get(schedule_id)
        if state and state.get("status") == "failed":
            finalize_interrupted_scheduled_crawl(row, redis_failed_task_reason(state, "定时采集"))
            finalized += 1
            continue
        if state is not None:
            continue
        task_touched_at = row.updated_at or row.last_run_at or row.created_at
        if task_touched_at and task_touched_at > missing_cutoff:
            continue
        finalize_interrupted_scheduled_crawl(row, redis_missing_task_reason("定时采集"))
        finalized += 1
    return finalized


def reconcile_interrupted_background_tasks_once() -> int:
    with session_scope() as session:
        return (
            reconcile_interrupted_running_tasks(session, CrawlTaskModel)
            + reconcile_interrupted_running_tasks(session, SyncTaskModel)
            + reconcile_interrupted_running_tasks(session, ListingTaskModel)
            + reconcile_missing_queued_tasks(session, CrawlTaskModel)
            + reconcile_missing_queued_tasks(session, SyncTaskModel)
            + reconcile_missing_queued_tasks(session, ListingTaskModel)
            + reconcile_interrupted_scheduled_crawls(session)
        )


def task_model_queue_kind(model: Any) -> str:
    if model is CrawlTaskModel:
        return "crawl"
    if model is SyncTaskModel:
        return "sync"
    if model is ListingTaskModel:
        return "listing"
    return ""


def task_model_action_label(model: Any, task: Any | None = None) -> str:
    if model is CrawlTaskModel:
        return "采集"
    if model is SyncTaskModel and task is not None:
        return sync_task_action_label(task)
    if model is SyncTaskModel:
        return "同步"
    if model is ListingTaskModel:
        return "上架"
    return "任务"


def task_model_job_id_prefixes(model: Any) -> tuple[str, ...]:
    if model is CrawlTaskModel:
        return ("crawl-", "crawl:")
    if model is SyncTaskModel:
        return ("sync-", "sync:")
    if model is ListingTaskModel:
        return ("listing-", "listing:")
    return ()


def interrupted_task_missing_cutoff() -> datetime:
    return datetime.now() - timedelta(
        seconds=max(
            settings.task_queue_job_timeout_seconds + TASK_REDIS_MISSING_JOB_GRACE_SECONDS,
            TASK_STALE_CANCEL_REQUEST_SECONDS,
        )
    )


def queued_task_missing_cutoff() -> datetime:
    return datetime.now() - timedelta(seconds=TASK_QUEUED_REDIS_MISSING_JOB_GRACE_SECONDS)


def reconcile_missing_queued_tasks(
    session: Any,
    model: Any,
    *,
    owner_username: str | None = None,
    store_id: int | None = None,
    limit: int | None = None,
) -> int:
    if not should_use_redis_task_queue():
        return 0
    queue_kind = task_model_queue_kind(model)
    if not queue_kind:
        return 0
    normalized_limit = max(1, int(limit or settings.task_queue_queued_requeue_limit))
    cutoff = queued_task_missing_cutoff()
    query = (
        select(model)
        .where(model.status == "queued", model.created_at <= cutoff)
        .order_by(model.created_at.asc(), model.id.asc())
        .limit(normalized_limit)
    )
    if model is CrawlTaskModel:
        query = query.where(CrawlTaskModel.queue_job_id.is_not(None))
    if owner_username is not None:
        query = query.where(model.owner_username == owner_username)
    if store_id is not None and hasattr(model, "store_id"):
        query = query.where(model.store_id == store_id)
    rows = session.scalars(query).all()
    if not rows:
        return 0
    if model is CrawlTaskModel:
        connection = redis_connection()
        task_states = {
            str(row.id): state
            for row in rows
            if row.queue_job_id
            and (state := reserved_crawl_job_state(connection, row.queue_job_id)) is not None
        }
    else:
        try:
            if model is SyncTaskModel:
                task_states = {}
                rows_by_queue: dict[str, set[str]] = {}
                for row in rows:
                    rows_by_queue.setdefault(
                        sync_task_queue_kind(row.task_type),
                        set(),
                    ).add(str(row.id))
                for row_queue_kind, task_ids in rows_by_queue.items():
                    task_states.update(
                        redis_task_states(
                            task_ids,
                            row_queue_kind,
                            job_id_prefixes=task_model_job_id_prefixes(model),
                        )
                    )
            else:
                task_ids = {str(row.id) for row in rows}
                task_states = redis_task_states(
                    task_ids,
                    queue_kind,
                    job_id_prefixes=task_model_job_id_prefixes(model),
                )
        except Exception:
            return 0
    requeued = 0
    for task in rows:
        task_id = str(task.id)
        state = task_states.get(task_id)
        if model is CrawlTaskModel:
            if state is not None and state.get("status") in {"queued", "started", "deferred", "scheduled", "unknown"}:
                continue
        elif state is not None:
            continue
        if task_cancel_requested(task):
            if model is ListingTaskModel:
                release_listing_task_locks(session, task.owner_username, task)
            if model is CrawlTaskModel:
                task.queue_job_id = None
            task.status = "cancelled"
            task.message = TASK_CANCELLED_MESSAGE
            task.error_detail = cancelled_task_error_detail(existing_error_detail=getattr(task, "error_detail", None))
            if hasattr(task, "warning_detail"):
                task.warning_detail = cancelled_task_warning_detail(existing_warning_detail=getattr(task, "warning_detail", None))
            task.finished_at = datetime.now()
            continue
        if model is CrawlTaskModel:
            task.queue_job_id = None
            task.message = "采集队列已恢复，等待重新投递"
        elif model is SyncTaskModel:
            task.message = f"{task_model_action_label(model, task)}队列已恢复，系统已重新投递任务"
            dispatch_sync_task(task.owner_username, task_id)
        elif model is ListingTaskModel:
            task.message = f"{task_model_action_label(model, task)}队列已恢复，系统已重新投递任务"
            dispatch_listing_task(task.owner_username, task_id)
        requeued += 1
    return requeued


def redis_task_states(
    task_ids: set[str],
    queue_kind: str,
    *,
    job_id_prefixes: tuple[str, ...] = (),
) -> dict[str, dict[str, Any]]:
    if not task_ids:
        return {}
    from rq import Queue
    from rq.registry import DeferredJobRegistry, FailedJobRegistry, ScheduledJobRegistry, StartedJobRegistry

    connection = redis_connection()
    queue_name = task_queue_name_for_kind(queue_kind)
    queue = Queue(queue_name, connection=connection)
    states: dict[str, dict[str, Any]] = {}
    active_sources = [
        ("queued", list(queue.job_ids)),
        ("started", StartedJobRegistry(queue_name, connection=connection).get_job_ids()),
        ("deferred", DeferredJobRegistry(queue_name, connection=connection).get_job_ids()),
        ("scheduled", ScheduledJobRegistry(queue_name, connection=connection).get_job_ids()),
    ]
    for status, job_ids in active_sources:
        for job_id in job_ids:
            job = fetch_rq_job(connection, job_id)
            if job is None:
                continue
            task_id = task_id_from_rq_job(job, task_ids, job_id_prefixes=job_id_prefixes)
            if task_id:
                states[task_id] = rq_job_state(job, status)
    failed_registry = FailedJobRegistry(queue_name, connection=connection)
    for job_id in failed_registry.get_job_ids():
        job = fetch_rq_job(connection, job_id)
        if job is None:
            continue
        task_id = task_id_from_rq_job(job, task_ids, job_id_prefixes=job_id_prefixes)
        if task_id and task_id not in states:
            states[task_id] = rq_job_state(job, "failed")
    return states


def fetch_rq_job(connection: Any, job_id: str) -> Any | None:
    from rq.job import Job

    try:
        return Job.fetch(job_id, connection=connection)
    except Exception:
        return None


def normalized_rq_job_status(value: Any) -> str:
    raw = getattr(value, "value", value)
    return str(raw or "").strip().lower()


def reserved_crawl_job_state(connection: Any, job_id: str) -> dict[str, Any] | None:
    job = fetch_rq_job(connection, job_id)
    if job is None:
        return None
    try:
        status = normalized_rq_job_status(job.get_status(refresh=True))
    except Exception:
        status = "unknown"
    return rq_job_state(job, status or "unknown")


def task_queue_health_kind_by_name() -> dict[str, str]:
    return {
        task_queue_name_for_kind("manual-crawl"): "手动采集",
        task_queue_name_for_kind("scheduled-crawl"): "定时采集执行",
        task_queue_name_for_kind("crawl"): "采集兼容队列",
        task_queue_name_for_kind("sync"): "同步",
        task_queue_name_for_kind("title-optimization"): "标题优化",
        task_queue_name_for_kind("image-cleanup"): "图片清理",
        task_queue_name_for_kind("listing"): "上架",
        task_queue_name_for_kind("schedule"): "定时",
        settings.task_queue_name: "默认",
    }


def task_queue_health() -> dict[str, Any]:
    checked_at = datetime.now()
    base = {
        "mode": settings.task_queue_mode,
        "status": "ok",
        "ok": True,
        "summary": "后台队列正常",
        "checkedAt": datetime_to_public(checked_at),
        "workerCount": 0,
        "redis": None,
        "queues": [],
        "error": "",
    }
    if not should_use_redis_task_queue():
        return {
            **base,
            "status": "disabled",
            "ok": True,
            "summary": "当前使用本地线程模式，未启用 Redis 队列",
        }
    try:
        from rq import Queue, Worker
        from rq.registry import DeferredJobRegistry, FailedJobRegistry, ScheduledJobRegistry, StartedJobRegistry

        connection = redis_connection()
        connection.ping()
        redis_info = connection.info(section="memory")
        workers = Worker.all(connection=connection)
        worker_count_by_queue: dict[str, int] = {}
        for worker in workers:
            for queue in getattr(worker, "queues", []) or []:
                queue_name = str(getattr(queue, "name", "") or "")
                if queue_name:
                    worker_count_by_queue[queue_name] = worker_count_by_queue.get(queue_name, 0) + 1

        queue_kind_by_name = task_queue_health_kind_by_name()
        expected_queue_names = set(queue_kind_by_name) - {
            settings.task_queue_name,
            settings.task_queue_crawl_name,
        }
        queues: list[dict[str, Any]] = []
        issues: list[str] = []
        for queue_name in all_task_queue_names():
            queue = Queue(queue_name, connection=connection)
            started = len(StartedJobRegistry(queue_name, connection=connection))
            failed = len(FailedJobRegistry(queue_name, connection=connection))
            deferred = len(DeferredJobRegistry(queue_name, connection=connection))
            scheduled = len(ScheduledJobRegistry(queue_name, connection=connection))
            queued = len(queue)
            worker_count = int(worker_count_by_queue.get(queue_name, 0))
            expected = queue_name in expected_queue_names
            pending = queued + started + deferred + scheduled
            queue_ok = (not expected or worker_count > 0) and not (pending > 0 and worker_count <= 0)
            if not queue_ok:
                if pending > 0 and worker_count <= 0:
                    issues.append(f"{queue_name} 有待执行任务但没有可用 worker")
                elif expected and worker_count <= 0:
                    issues.append(f"{queue_name} 没有可用 worker")
            queues.append(
                {
                    "name": queue_name,
                    "kind": queue_kind_by_name.get(queue_name, queue_name),
                    "workerCount": worker_count,
                    "queued": queued,
                    "started": started,
                    "failed": failed,
                    "deferred": deferred,
                    "scheduled": scheduled,
                    "pending": pending,
                    "ok": queue_ok,
                }
            )

        status = "ok" if not issues else "degraded"
        return {
            **base,
            "status": status,
            "ok": not issues,
            "summary": "后台队列正常" if not issues else "；".join(issues[:3]),
            "workerCount": len(workers),
            "redis": {
                "usedMemory": int(redis_info.get("used_memory", 0) or 0),
                "usedMemoryHuman": str(redis_info.get("used_memory_human", "") or ""),
                "maxMemory": int(redis_info.get("maxmemory", 0) or 0),
                "maxMemoryHuman": str(redis_info.get("maxmemory_human", "") or ""),
            },
            "queues": queues,
        }
    except Exception as exc:
        return {
            **base,
            "status": "error",
            "ok": False,
            "summary": "Redis 队列连接异常",
            "error": str(exc),
        }


def task_id_from_rq_job(
    job: Any,
    task_ids: set[str],
    *,
    job_id_prefixes: tuple[str, ...] = (),
) -> str | None:
    job_id = str(getattr(job, "id", "") or "")
    if job_id in task_ids:
        return job_id
    for prefix in job_id_prefixes:
        if not job_id.startswith(prefix):
            continue
        normalized = job_id[len(prefix):]
        for task_id in task_ids:
            if normalized == task_id or normalized.startswith(f"{task_id}-"):
                return task_id
    for value in getattr(job, "args", ()) or ():
        text = str(value)
        if text in task_ids:
            return text
    description = str(getattr(job, "description", "") or "")
    for task_id in task_ids:
        if re.search(rf"(?<![\w-]){re.escape(task_id)}(?![\w-])", description):
            return task_id
    return None


def rq_job_state(job: Any, status: str) -> dict[str, Any]:
    exc_info = normalize_task_detail_text(str(getattr(job, "exc_info", "") or ""))
    return {
        "jobId": str(getattr(job, "id", "") or ""),
        "status": status,
        "description": str(getattr(job, "description", "") or ""),
        "startedAt": getattr(job, "started_at", None),
        "endedAt": getattr(job, "ended_at", None),
        "excInfo": exc_info,
    }


def redis_failed_task_reason(state: dict[str, Any], action_label: str) -> str:
    description = str(state.get("description") or f"{action_label}任务")
    job_id = str(state.get("jobId") or "")
    ended_at = state.get("endedAt")
    exc_info = normalize_task_detail_text(state.get("excInfo"))
    reason = f"后台{action_label}任务进程异常退出，系统已自动标记为失败。"
    if "Work-horse terminated unexpectedly" in exc_info:
        reason = f"后台{action_label}任务进程异常退出或超时，系统已自动标记为失败。"
    details = [reason]
    if description:
        details.append(f"任务：{description}")
    if job_id:
        details.append(f"Redis Job：{job_id}")
    if ended_at:
        details.append(f"结束时间：{ended_at}")
    if exc_info:
        details.append(f"队列错误：{exc_info[-500:]}")
    return "\n".join(details)


def redis_missing_task_reason(action_label: str) -> str:
    return f"后台{action_label}任务已不在 Redis 队列中，且超过任务超时时间，系统已自动标记为失败。"


def finalize_interrupted_task(session: Any, model: Any, task: Any, reason: str) -> None:
    if model is SyncTaskModel:
        finalize_interrupted_sync_task(task, reason)
        return
    if model is ListingTaskModel:
        finalize_interrupted_listing_task(session, task, reason)
        return
    if model is CrawlTaskModel:
        finalize_interrupted_crawl_task(task, reason)
        return


def finalize_interrupted_sync_task(task: SyncTaskModel, reason: str) -> None:
    total_count = int(task.total_count or 0)
    success_count = int(task.success_count or 0)
    failed_count = int(task.failed_count or 0)
    remaining_count = max(0, total_count - success_count - failed_count)
    task.failed_count = max(failed_count + remaining_count, 1)
    task.status = "failed"
    task.message = failed_sync_task_progress_message(
        sync_task_action_label(task),
        total_count,
        success_count,
        failed_count,
        unfinished_count=remaining_count,
    )
    task.error_detail = interrupted_task_error_detail(reason, task.error_detail)
    task.finished_at = datetime.now()


def finalize_interrupted_crawl_task(task: CrawlTaskModel, reason: str) -> None:
    total_count = int(task.total_count or 0)
    success_count = int(task.success_count or 0)
    failed_count = int(task.failed_count or 0)
    unfinished_count = max(0, total_count - success_count - failed_count)
    task.failed_count = max(failed_count + unfinished_count, 1)
    task.status = "failed"
    task.message = interrupted_progress_message("采集", total_count, success_count, failed_count, unfinished_count=unfinished_count)
    task.error_detail = interrupted_task_error_detail(reason, task.error_detail)
    task.finished_at = datetime.now()


def finalize_interrupted_listing_task(session: Any, task: ListingTaskModel, reason: str) -> None:
    product_ids_payload = listing_task_product_ids_payload(task.product_ids_json)
    task_product_ids = product_ids_payload["productIds"]
    success_ids = product_ids_payload["successIds"]
    failed_ids = product_ids_payload["failedIds"]
    retry_ids = product_ids_payload["retryIds"]
    store_ids = product_ids_payload["storeIds"] or ([int(task.store_id)] if task.store_id else [])
    active_ids = retry_ids or task_product_ids
    handled_ids = set(success_ids) | set(failed_ids)
    unfinished_ids = [product_id for product_id in active_ids if product_id not in handled_ids]
    final_failed_ids = merge_listing_task_product_ids(failed_ids, unfinished_ids)
    release_listing_task_locks(session, task.owner_username, task)

    total_count = max(int(task.total_count or 0), len(task_product_ids) * max(1, len(store_ids)))
    success_count = len(success_ids)
    failed_count = max(len(final_failed_ids), 0 if success_count else 1)
    task.status = "partial" if success_count else "failed"
    task.total_count = total_count
    task.success_count = success_count
    task.failed_count = failed_count
    task.message = interrupted_progress_message(
        "上架",
        total_count,
        success_count,
        failed_count,
        unfinished_count=len(unfinished_ids),
    )
    task.error_detail = interrupted_task_error_detail(reason, task.error_detail)
    task.product_ids_json = json.dumps(
        listing_task_result_payload(task_product_ids, success_ids, final_failed_ids, store_ids=store_ids),
        ensure_ascii=False,
    )
    task.finished_at = datetime.now()


def finalize_interrupted_scheduled_crawl(row: ScheduledCrawlModel, reason: str) -> None:
    row.status = "failed"
    row.notes = interrupted_task_error_detail(reason, row.notes) or reason
    if row.enabled and row.next_run_at is None:
        row.next_run_at = next_daily_run_at(row.schedule_time)


def interrupted_progress_message(
    action_label: str,
    total_count: int,
    success_count: int,
    failed_count: int,
    *,
    unfinished_count: int = 0,
) -> str:
    total = max(0, int(total_count or 0))
    success = max(0, int(success_count or 0))
    failed = max(0, int(failed_count or 0))
    unfinished = max(0, int(unfinished_count or 0))
    if total > 0:
        processed = min(total, success + failed)
        return (
            f"{action_label}异常中断，已处理 {processed} / {total} 条，"
            f"成功 {success} 条，异常 {failed} 条，未完成 {unfinished} 条"
        )
    return f"{action_label}异常中断"


def finalize_stale_store_cancel_requested_tasks(session: Any, store_id: int | None) -> int:
    if not store_id:
        return 0
    sync_count = finalize_stale_cancel_requested_tasks(
        session,
        SyncTaskModel,
        action_label="同步",
        store_id=store_id,
    )
    sync_count += reconcile_interrupted_running_tasks(session, SyncTaskModel, store_id=store_id)
    listing_count = finalize_stale_cancel_requested_tasks(
        session,
        ListingTaskModel,
        action_label="上架",
        store_id=store_id,
    )
    listing_count += reconcile_interrupted_running_tasks(session, ListingTaskModel, store_id=store_id)
    return sync_count + listing_count


def is_task_cancel_requested(model: Any, task_id: str) -> bool:
    with session_scope() as session:
        task = session.get(model, task_id)
        if task is None:
            return False
        return getattr(task, "status", "") == "cancelled" or task_cancel_requested(task)


def raise_if_task_cancelled(model: Any, task_id: str | None) -> None:
    if task_id and is_task_cancel_requested(model, task_id):
        raise TaskCancelled(TASK_CANCELLED_MESSAGE)


def is_mysql_retryable_lock_error(exc: OperationalError) -> bool:
    original = getattr(exc, "orig", None)
    code = getattr(original, "args", [None])[0] if original is not None else None
    normalized = str(exc).lower()
    return (
        code in {1205, 1213}
        or "lock wait timeout exceeded" in normalized
        or "deadlock found" in normalized
    )


def running_user_task_count(
    session: Any,
    model: Any,
    owner_username: str,
    *,
    exclude_task_id: str | None = None,
) -> int:
    query = select(func.count()).where(
        model.owner_username == owner_username,
        model.status == "running",
    )
    if exclude_task_id:
        query = query.where(model.id != exclude_task_id)
    return int(session.scalar(query) or 0)


def running_sync_task_count(session: Any, *, exclude_task_id: str | None = None) -> int:
    query = select(func.count()).where(
        SyncTaskModel.status == "running",
        SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
    )
    if exclude_task_id:
        query = query.where(SyncTaskModel.id != exclude_task_id)
    return int(session.scalar(query) or 0)


def specialized_sync_task_wait_reason(
    session: Any,
    task: SyncTaskModel,
) -> str:
    task_type = normalize_text(task.task_type)
    if task_type in TITLE_OPTIMIZATION_TASK_TYPES:
        session.execute(
            select(UserAccountModel.username)
            .where(UserAccountModel.username == task.owner_username)
            .with_for_update()
        ).first()
        running_task_id = session.scalar(
            select(SyncTaskModel.id)
            .where(
                SyncTaskModel.owner_username == task.owner_username,
                SyncTaskModel.task_type.in_(TITLE_OPTIMIZATION_TASK_TYPES),
                SyncTaskModel.status == "running",
                SyncTaskModel.id != task.id,
            )
            .limit(1)
            .with_for_update()
        )
        if running_task_id is not None:
            return "排队中，等待该用户当前标题优化任务完成"
    if task_type in IMAGE_CLEANUP_TASK_TYPES and task.store_id is not None:
        session.execute(
            select(StoreModel.id)
            .where(StoreModel.id == task.store_id)
            .with_for_update()
        ).first()
        running_task_id = session.scalar(
            select(SyncTaskModel.id)
            .where(
                SyncTaskModel.store_id == task.store_id,
                SyncTaskModel.task_type.in_(IMAGE_CLEANUP_TASK_TYPES),
                SyncTaskModel.status == "running",
                SyncTaskModel.id != task.id,
            )
            .limit(1)
            .with_for_update()
        )
        if running_task_id is not None:
            return "排队中，等待该店铺当前图片清理任务完成"
    if task_type in LISTING_IMAGE_UPLOAD_TASK_TYPES and task.store_id is not None:
        running_task_id = session.scalar(
            select(SyncTaskModel.id)
            .where(
                SyncTaskModel.store_id == task.store_id,
                SyncTaskModel.task_type.in_(LISTING_IMAGE_UPLOAD_TASK_TYPES),
                SyncTaskModel.status == "running",
                SyncTaskModel.id != task.id,
            )
            .limit(1)
            .with_for_update()
        )
        if running_task_id is not None:
            return "排队中，等待该店铺当前图片完善任务完成"
    if task_type in PRODUCT_SCOPED_SYNC_TASK_TYPES:
        scope = sync_task_resource_scope(session, task)
        active_scopes = [
            sync_task_resource_scope(session, row)
            for row in session.scalars(
                select(SyncTaskModel).where(
                    SyncTaskModel.status == "running",
                    SyncTaskModel.id != task.id,
                    SyncTaskModel.task_type.notin_(IMAGE_CLEANUP_TASK_TYPES),
                )
            ).all()
        ]
        active_scopes.extend(
            listing_task_resource_scope(session, row)
            for row in session.scalars(
                select(ListingTaskModel).where(
                    ListingTaskModel.status.in_(("queued", "running")),
                )
            ).all()
        )
        wait_reason = task_scope_conflict_reason(scope, active_scopes)
        if wait_reason:
            return wait_reason
    return ""


def running_listing_task_count(session: Any, *, exclude_task_id: str | None = None) -> int:
    query = select(func.count()).where(ListingTaskModel.status == "running")
    if exclude_task_id:
        query = query.where(ListingTaskModel.id != exclude_task_id)
    return int(session.scalar(query) or 0)


def running_user_listing_task_count(
    session: Any,
    owner_username: str,
    *,
    exclude_task_id: str | None = None,
) -> int:
    query = select(func.count()).where(
        ListingTaskModel.owner_username == owner_username,
        ListingTaskModel.status == "running",
    )
    if exclude_task_id:
        query = query.where(ListingTaskModel.id != exclude_task_id)
    return int(session.scalar(query) or 0)


def running_store_listing_task_count(
    session: Any,
    store_id: int,
    *,
    exclude_task_id: str | None = None,
) -> int:
    query = select(func.count()).where(
        ListingTaskModel.store_id == store_id,
        ListingTaskModel.status == "running",
    )
    if exclude_task_id:
        query = query.where(ListingTaskModel.id != exclude_task_id)
    return int(session.scalar(query) or 0)


def running_store_sync_task_count(session: Any, store_id: int) -> int:
    return int(session.scalar(
        select(func.count()).where(
            SyncTaskModel.store_id == store_id,
            SyncTaskModel.status == "running",
            SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
        )
    ) or 0)


def sync_task_start_wait_reason(session: Any, task_id: str, store_id: int | None) -> str:
    finalize_stale_cancel_requested_tasks(session, SyncTaskModel, action_label="同步")
    finalize_stale_store_cancel_requested_tasks(session, store_id)
    reconcile_interrupted_running_tasks(session, SyncTaskModel)
    task = session.get(SyncTaskModel, task_id)
    if task is None:
        return "同步任务不存在"
    session.execute(
        select(UserAccountModel.username)
        .where(UserAccountModel.username == task.owner_username)
        .with_for_update()
    ).first()
    if running_sync_task_count(
        session,
        exclude_task_id=task_id,
    ) >= int(settings.max_running_sync_tasks_global):
        return "排队中，等待全局同步并发额度"
    if running_user_task_count(
        session,
        SyncTaskModel,
        task.owner_username,
        exclude_task_id=task_id,
    ) >= int(settings.max_running_sync_tasks_per_user):
        return "排队中，等待该用户当前同步任务释放并发额度"
    scope = sync_task_resource_scope(session, task)
    active_scopes = [
        sync_task_resource_scope(session, row)
        for row in session.scalars(
            select(SyncTaskModel).where(
                SyncTaskModel.status == "running",
                SyncTaskModel.id != task_id,
                SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
            )
        ).all()
    ]
    active_scopes.extend(
        listing_task_resource_scope(session, row)
        for row in session.scalars(
            select(ListingTaskModel).where(
                ListingTaskModel.status == "running",
            )
        ).all()
    )
    return task_scope_conflict_reason(scope, active_scopes)


def listing_task_start_wait_reason(session: Any, task: ListingTaskModel) -> str:
    task_id = str(task.id)
    finalize_stale_cancel_requested_tasks(session, ListingTaskModel, action_label="上架")
    reconcile_interrupted_running_tasks(session, ListingTaskModel)
    session.execute(
        select(UserAccountModel.username)
        .where(UserAccountModel.username == task.owner_username)
        .with_for_update()
    ).first()
    store_ids = listing_task_store_ids(task)
    if store_ids:
        session.execute(
            select(StoreModel.id)
            .where(StoreModel.id.in_(store_ids))
            .order_by(StoreModel.id.asc())
            .with_for_update()
        ).all()
    for store_id in store_ids:
        finalize_stale_store_cancel_requested_tasks(session, store_id)
    if running_listing_task_count(session, exclude_task_id=task_id) >= int(settings.max_running_listing_tasks_global):
        return "排队中，等待全局上架并发额度"
    if running_user_listing_task_count(
        session,
        task.owner_username,
        exclude_task_id=task_id,
    ) >= int(settings.max_running_listing_tasks_per_user):
        return "排队中，等待该用户当前上架任务释放并发额度"
    for store_id in store_ids:
        if running_store_listing_task_count(
            session,
            store_id,
            exclude_task_id=task_id,
        ) >= int(settings.max_running_listing_tasks_per_store):
            return "排队中，等待该店铺当前上架任务释放并发额度"
    scope = listing_task_resource_scope(session, task)
    active_sync_scopes = [
        sync_task_resource_scope(session, row)
        for row in session.scalars(
            select(SyncTaskModel).where(
                SyncTaskModel.status == "running",
                SyncTaskModel.task_type.notin_(IMAGE_CLEANUP_TASK_TYPES),
            )
        ).all()
    ]
    return task_scope_conflict_reason(scope, active_sync_scopes)


def running_store_task_count(
    session: Any,
    store_id: int | None,
    *,
    exclude_sync_task_id: str | None = None,
    exclude_listing_task_id: str | None = None,
) -> int:
    if not store_id:
        return 0
    sync_query = select(func.count()).where(
        SyncTaskModel.store_id == store_id,
        SyncTaskModel.status == "running",
        SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
    )
    if exclude_sync_task_id:
        sync_query = sync_query.where(SyncTaskModel.id != exclude_sync_task_id)
    listing_query = select(func.count()).where(
        ListingTaskModel.store_id == store_id,
        ListingTaskModel.status == "running",
    )
    if exclude_listing_task_id:
        listing_query = listing_query.where(ListingTaskModel.id != exclude_listing_task_id)
    return int(session.scalar(sync_query) or 0) + int(session.scalar(listing_query) or 0)


def task_start_wait_reason(
    session: Any,
    model: Any,
    owner_username: str,
    task_id: str,
    *,
    limit: int,
    label: str,
    store_id: int | None = None,
    exclude_sync_task_id: str | None = None,
    exclude_listing_task_id: str | None = None,
) -> str:
    finalize_stale_cancel_requested_tasks(session, model, action_label=label, owner_username=owner_username)
    finalize_stale_store_cancel_requested_tasks(session, store_id)
    running_count = running_user_task_count(session, model, owner_username, exclude_task_id=task_id)
    if running_count >= limit:
        return f"排队中，等待当前{label}任务完成"
    store_running_count = running_store_task_count(
        session,
        store_id,
        exclude_sync_task_id=exclude_sync_task_id,
        exclude_listing_task_id=exclude_listing_task_id,
    )
    if store_running_count > 0:
        return "排队中，等待该店铺当前同步、上架、上下架或删除任务完成"
    return ""


def crawl_task_start_wait_reason(
    session: Any,
    task: CrawlTaskModel,
) -> str:
    task_id = str(task.id)
    task_mode = normalize_crawl_task_mode(task.mode)
    action_label = "定时采集" if task_mode == "scheduled" else "手动采集"
    finalize_stale_cancel_requested_tasks(
        session,
        CrawlTaskModel,
        action_label=action_label,
        owner_username=task.owner_username,
    )
    running_count = int(
        session.scalar(
            select(func.count()).where(
                CrawlTaskModel.owner_username == task.owner_username,
                CrawlTaskModel.status == "running",
                CrawlTaskModel.id != task_id,
                crawl_task_mode_filter(task_mode),
            )
        )
        or 0
    )
    if running_count >= crawl_task_concurrency_limit(task_mode):
        return f"排队中，等待当前{action_label}任务完成"
    return ""


def role_to_public(row: RoleModel) -> dict[str, Any]:
    try:
        permissions = json.loads(row.permissions_json or "[]")
    except ValueError:
        permissions = []
    return {
        "id": row.id,
        "name": row.name,
        "code": row.code,
        "scope": row.scope,
        "enabled": bool(row.enabled),
        "permissions": permissions if isinstance(permissions, list) else [],
        "notes": row.notes,
        "createdAt": row.created_at.isoformat(sep=" ") if row.created_at else None,
        "updatedAt": row.updated_at.isoformat(sep=" ") if row.updated_at else None,
    }


def _product_status_filter(status: str | None) -> str | None:
    if not status:
        return None
    status_map = {
        "pending": "pending",
        "approved": "approved",
        "error": "error",
        "listed_master": "listed_master",
        "listed": "listed",
        "rejected": "rejected",
    }
    return status_map.get(status, status)


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_shop_code(value: Any) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    if normalized.startswith(("http://", "https://")):
        try:
            path_parts = [part for part in urlsplit(normalized).path.split("/") if part]
        except Exception:
            path_parts = []
        return normalize_text(path_parts[0])
    return normalized.strip("/")


def build_rakuten_store_url(shop_code: str) -> str:
    normalized_shop_code = normalize_shop_code(shop_code)
    if not normalized_shop_code:
        return ""
    return f"https://www.rakuten.co.jp/{normalized_shop_code}/"


def build_public_item_page_url(shop_code: str, item_number: str) -> str:
    normalized_shop_code = normalize_shop_code(shop_code)
    normalized_item_number = normalize_text(item_number)
    if not normalized_shop_code or not normalized_item_number:
        return ""
    return f"https://item.rakuten.co.jp/{quote(normalized_shop_code, safe='')}/{quote(normalized_item_number, safe='')}/"


def is_rakuten_product_url(value: str) -> bool:
    normalized = normalize_text(value)
    if not normalized.startswith(("http://", "https://")):
        return False
    try:
        parsed = urlsplit(normalized)
    except Exception:
        return False
    hostname = parsed.netloc.lower()
    if hostname == "item.rakuten.co.jp":
        return parse_rakuten_product_target(normalized) is not None
    if hostname == "brandavenue.rakuten.co.jp":
        return parse_rakuten_fashion_product_code(normalized) != ""
    return False


def parse_rakuten_fashion_product_code(target: str) -> str:
    normalized = normalize_text(target)
    if not normalized.startswith(("http://", "https://")):
        return ""
    try:
        parsed = urlsplit(normalized)
    except Exception:
        return ""
    if parsed.netloc.lower() != "brandavenue.rakuten.co.jp":
        return ""
    parts = [unquote(part.strip()) for part in parsed.path.split("/") if part.strip()]
    if len(parts) >= 2 and parts[0] == "item":
        return normalize_text(parts[1])
    return ""


def parse_rakuten_product_target(target: str) -> tuple[str, str] | None:
    normalized = normalize_text(target)
    if not normalized:
        return None
    if normalized.startswith(("http://", "https://")):
        try:
            parsed = urlsplit(normalized)
        except Exception:
            return None
        if parsed.netloc.lower() != "item.rakuten.co.jp":
            return None
        parts = [unquote(part.strip()) for part in parsed.path.split("/") if part.strip()]
        if len(parts) < 2:
            return None
        shop_code, item_number = parts[0], parts[1]
    else:
        parts = [unquote(part.strip()) for part in normalized.strip("/").split("/") if part.strip()]
        if len(parts) != 2:
            return None
        shop_code, item_number = parts
    if not shop_code or not item_number:
        return None
    if item_number.lower() == "c":
        return None
    if any(part.startswith(("http:", "https:")) for part in (shop_code, item_number)):
        return None
    return shop_code, item_number


def normalize_rakuten_product_targets(target: Any) -> list[str]:
    normalized: list[str] = []
    for value in re.split(r"[\r\n]+", str(target or "")):
        item = normalize_text(value)
        if not item:
            continue
        url = normalize_rakuten_product_target(item)
        if url not in normalized:
            normalized.append(url)
    if not normalized:
        raise RuntimeError(RAKUTEN_PRODUCT_TARGET_ERROR)
    return normalized


def normalize_rakuten_product_target(target: str) -> str:
    normalized = normalize_text(target)
    fashion_code = parse_rakuten_fashion_product_code(normalized)
    if fashion_code:
        return f"https://brandavenue.rakuten.co.jp/item/{quote(fashion_code, safe='')}/"
    parsed = parse_rakuten_product_target(target)
    if parsed is None:
        raise RuntimeError(RAKUTEN_PRODUCT_TARGET_ERROR)
    shop_code, item_number = parsed
    return build_public_item_page_url(shop_code, item_number)


def normalize_rakuten_shop_target(target: str) -> str:
    normalized = normalize_text(target)
    if re.fullmatch(r"[0-9]+", normalized):
        return normalized
    if not normalized.startswith(("http://", "https://")):
        return normalized
    try:
        parsed = urlsplit(normalized)
    except Exception as exc:
        raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR) from exc
    if parsed.netloc.lower() == "search.rakuten.co.jp" and parsed.path.rstrip("/").endswith("/search/mall"):
        params = parse_qs(parsed.query)
        shop_code = normalize_text((params.get("su") or [""])[0])
        sid = normalize_text((params.get("sid") or [""])[0])
        shop_name = normalize_text((params.get("sn") or [""])[0])
        if looks_like_rakuten_shop_code(shop_code):
            return shop_code
        if shop_code.startswith(("http://", "https://")):
            try:
                nested_target = normalize_rakuten_shop_target(shop_code)
            except RuntimeError:
                nested_target = ""
            if looks_like_rakuten_shop_code(nested_target):
                return nested_target
        if re.fullmatch(r"[0-9]+", sid):
            return sid
        return shop_name or shop_code or sid
    if parsed.netloc.lower() in {"www.rakuten.co.jp", "item.rakuten.co.jp"}:
        parts = [unquote(part.strip()) for part in parsed.path.split("/") if part.strip()]
        if parts:
            return normalize_text(parts[0])
    raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR)


def resolve_rakuten_shop_search_keyword(target: str) -> str:
    normalized = normalize_rakuten_shop_target(target)
    if not normalized:
        raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR)
    if re.fullmatch(r"[0-9]+", normalized):
        display_name = fetch_rakuten_shop_display_name_by_sid(normalized)
        return display_name or normalized
    if looks_like_rakuten_shop_code(normalized):
        display_name = fetch_rakuten_shop_display_name_by_code(normalized)
        return display_name or normalized
    return normalized


def looks_like_rakuten_shop_code(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_-]{1,80}", normalize_text(value)))


def fetch_rakuten_shop_display_name_by_code(shop_code: str) -> str:
    normalized_shop_code = normalize_shop_code(shop_code)
    if not normalized_shop_code:
        return ""
    try:
        html = fetch_html(build_rakuten_store_url(normalized_shop_code))
    except requests.RequestException:
        return ""
    return parse_rakuten_shop_display_name(html)


def fetch_rakuten_shop_display_name_by_sid(sid: str) -> str:
    _, display_name = fetch_rakuten_shop_identity_by_sid(sid)
    return display_name


def fetch_rakuten_shop_identity_by_sid(sid: str) -> tuple[str, str]:
    normalized_sid = normalize_text(sid)
    if not re.fullmatch(r"[0-9]+", normalized_sid):
        return "", ""
    try:
        html = fetch_html(f"{RAKUTEN_SEARCH_BASE}?sid={quote(normalized_sid)}")
    except requests.RequestException:
        return "", ""
    return parse_rakuten_search_shop_identity(html)


def parse_rakuten_search_shop_identity(html: str) -> tuple[str, str]:
    soup = BeautifulSoup(html or "", "lxml")
    store_link_codes: set[str] = set()
    for link in soup.select("a[href*='www.rakuten.co.jp']"):
        try:
            parsed_link = urlsplit(urljoin(RAKUTEN_SEARCH_BASE, str(link.get("href") or "")))
        except Exception:
            continue
        if parsed_link.netloc.lower() != "www.rakuten.co.jp":
            continue
        path_parts = [unquote(part.strip()) for part in parsed_link.path.split("/") if part.strip()]
        if len(path_parts) == 1 and looks_like_rakuten_shop_code(path_parts[0]):
            store_link_codes.add(normalize_shop_code(path_parts[0]))

    product_link_codes: set[str] = set()
    for link in soup.select("a[href*='item.rakuten.co.jp']"):
        parsed_product = parse_rakuten_product_target(urljoin(RAKUTEN_SEARCH_BASE, str(link.get("href") or "")))
        if parsed_product:
            product_link_codes.add(normalize_shop_code(parsed_product[0]))
    shared_codes = store_link_codes & product_link_codes
    if len(shared_codes) == 1:
        shop_code = next(iter(shared_codes))
    elif len(store_link_codes) == 1 and not product_link_codes:
        shop_code = next(iter(store_link_codes))
    elif not store_link_codes and len(product_link_codes) == 1:
        shop_code = next(iter(product_link_codes))
    else:
        shop_code = ""
    display_name = parse_rakuten_search_shop_name(html) or parse_rakuten_shop_display_name(html)
    return shop_code, display_name


def parse_rakuten_shop_display_name(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    h1 = soup.select_one("h1")
    if h1:
        display_name = normalize_text(h1.get_text(" ", strip=True))
        if display_name:
            return display_name
    og_title = soup.select_one("meta[property='og:title'], meta[name='og:title']")
    if og_title:
        display_name = shop_name_from_title(str(og_title.get("content") or ""))
        if display_name:
            return display_name
    if soup.title:
        display_name = shop_name_from_title(soup.title.get_text(" ", strip=True))
        if display_name:
            return display_name
    return ""


def parse_rakuten_search_shop_name(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    for selector in ("h1, h2", "a"):
        for node in soup.select(selector):
            text = normalize_text(node.get_text(" ", strip=True))
            if text and not is_generic_rakuten_shop_label(text) and len(text) <= 80:
                return text
    return ""


def is_generic_rakuten_shop_label(value: str) -> bool:
    normalized = normalize_text(value)
    return normalized in {
        "ログイン",
        "会員登録",
        "買い物かご",
        "閲覧履歴",
        "お気に入り",
        "ショップへ問い合わせ",
        "すべてのショップ",
        "ショップ内から探す",
    }


def shop_name_from_title(title: str) -> str:
    normalized = normalize_text(title)
    if not normalized:
        return ""
    match = re.search(r"楽天市場\s*[|｜]\s*(.+?)(?:\s*[-－|｜]\s*.+)?$", normalized)
    if match:
        return normalize_text(match.group(1))
    return ""


def build_rakuten_authorization_header(service_secret: str, license_key: str) -> str:
    authorization = base64.b64encode(f"{service_secret}:{license_key}".encode("utf-8")).decode("ascii")
    return f"ESA {authorization}"


def rakuten_request_account_key(headers: Any) -> str:
    authorization = normalize_text(headers.get("Authorization")) if isinstance(headers, dict) else ""
    return hashlib.sha256(authorization.encode("utf-8")).hexdigest()[:24] if authorization else "default"


def wait_for_distributed_request_slot(bucket: str, min_interval: float) -> bool:
    if not should_use_redis_task_queue() or min_interval <= 0:
        return False
    connection = None
    lock = None
    try:
        connection = redis_connection()
        lock = connection.lock(
            f"lt:request-rate:{bucket}:lock",
            timeout=max(5.0, min_interval * 10),
            blocking_timeout=max(5.0, min_interval * 10),
        )
        if not lock.acquire(blocking=True):
            return False
        timestamp_key = f"lt:request-rate:{bucket}:last"
        raw_last = connection.get(timestamp_key)
        last_at = float(raw_last or 0.0)
        wait_seconds = max(0.0, min_interval - (time.time() - last_at))
        if wait_seconds > 0:
            time.sleep(wait_seconds)
        connection.set(timestamp_key, str(time.time()), ex=max(60, int(min_interval * 100)))
        return True
    except Exception:
        return False
    finally:
        if lock is not None:
            try:
                lock.release()
            except Exception:
                pass


def rakuten_cabinet_request_interval(account_key: str) -> float:
    normal_interval = max(
        0.0,
        float(settings.rakuten_cabinet_request_min_interval_seconds),
    )
    if not should_use_redis_task_queue():
        return normal_interval
    try:
        if redis_connection().exists(f"lt:request-rate:cabinet:{account_key}:cooldown"):
            return max(
                normal_interval,
                float(settings.rakuten_cabinet_qps_cooldown_interval_seconds),
            )
    except Exception:
        pass
    return normal_interval


def mark_rakuten_cabinet_qps_limited(account_key: str) -> None:
    if not should_use_redis_task_queue():
        return
    try:
        redis_connection().set(
            f"lt:request-rate:cabinet:{account_key}:cooldown",
            "1",
            ex=RAKUTEN_CABINET_QPS_COOLDOWN_SECONDS,
        )
    except Exception:
        pass


def throttle_rakuten_cabinet_request(account_key: str = "default") -> None:
    global RAKUTEN_CABINET_LAST_REQUEST_AT
    min_interval = rakuten_cabinet_request_interval(account_key)
    if min_interval <= 0:
        return
    if wait_for_distributed_request_slot(f"cabinet:{account_key}", min_interval):
        return
    with RAKUTEN_CABINET_REQUEST_LOCK:
        now = time.monotonic()
        elapsed = now - RAKUTEN_CABINET_LAST_REQUEST_AT
        if elapsed < min_interval:
            time.sleep(min_interval - elapsed)
        RAKUTEN_CABINET_LAST_REQUEST_AT = time.monotonic()


def rakuten_cabinet_request(method: str, url: str, **kwargs: Any) -> requests.Response:
    attempts = max(1, RAKUTEN_CABINET_REQUEST_MAX_RETRIES)
    last_exc: requests.RequestException | None = None
    account_key = rakuten_request_account_key(kwargs.get("headers"))
    for attempt in range(1, attempts + 1):
        throttle_rakuten_cabinet_request(account_key)
        try:
            response = requests.request(method, url, **kwargs)
        except requests.RequestException as exc:
            last_exc = exc
            if attempt >= attempts:
                raise
            time.sleep(rakuten_cabinet_backoff_seconds(attempt))
            continue
        transient_error = rakuten_cabinet_transient_error(response)
        if transient_error:
            if transient_error in {"QPSLimit", "SystemError(107)"}:
                mark_rakuten_cabinet_qps_limited(account_key)
            if attempt >= attempts:
                detail = normalize_text(response.text)[:500]
                response.close()
                if transient_error == "QPSLimit":
                    raise RuntimeError(f"R-Cabinet 请求触发 QPSLimit，已重试 {attempts} 次：{detail}")
                raise RuntimeError(f"R-Cabinet 请求持续返回 {transient_error}，已重试 {attempts} 次：{detail}")
            response.close()
            time.sleep(rakuten_cabinet_backoff_seconds(attempt))
            continue
        return response
    if last_exc is not None:
        raise last_exc
    raise RuntimeError("R-Cabinet 请求失败。")


def rakuten_cabinet_backoff_seconds(attempt: int) -> float:
    index = max(0, min(len(RAKUTEN_CABINET_QPS_BACKOFF_SECONDS) - 1, attempt - 1))
    base = float(RAKUTEN_CABINET_QPS_BACKOFF_SECONDS[index])
    return base + random.uniform(0, 0.5)


def is_rakuten_cabinet_qps_limited_response(response: requests.Response) -> bool:
    return rakuten_cabinet_transient_error(response) == "QPSLimit"


def rakuten_cabinet_transient_error(response: requests.Response) -> str:
    text = normalize_text(getattr(response, "text", ""))
    if response.status_code == 429 or "QPSLimit" in text:
        return "QPSLimit"
    if "SystemError(107)" in text:
        return "SystemError(107)"
    normalized_text = text.lower()
    if any(marker in normalized_text for marker in RAKUTEN_CABINET_UPSTREAM_ERROR_MARKERS):
        return "UpstreamConnectionError"
    if int(getattr(response, "status_code", 0) or 0) in RAKUTEN_CABINET_RETRY_STATUS_CODES:
        return f"HTTP {response.status_code}"
    return ""


def cabinet_xml_error_message(xml_text: str) -> str:
    text = normalize_text(xml_text)
    if not text:
        return ""
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return ""
    system_status = ""
    messages: list[str] = []
    result_codes: list[str] = []
    for element in root.iter():
        local_name = element.tag.split("}", 1)[-1].lower()
        value = normalize_text(element.text)
        if not value:
            continue
        if local_name == "systemstatus":
            system_status = value.upper()
        elif local_name == "message":
            messages.append(value)
        elif local_name in {"resultcode", "code"}:
            result_codes.append(value)
    if system_status and system_status != "OK":
        parts = [*messages, *[f"resultCode={code}" for code in result_codes]]
        return "，".join(unique_texts(parts)) or text[:500]
    return ""


def fetch_rakuten_shop_meta(service_secret: str, license_key: str) -> dict[str, str]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 和乐天 Key 不能为空。")
    try:
        response = requests.get(
            RAKUTEN_SHOP_MASTER_URL,
            timeout=settings.crawler_timeout_seconds,
            headers={
                "Authorization": build_rakuten_authorization_header(service_secret, license_key),
                "Accept": "application/xml, text/xml",
            },
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError("乐天密钥检测失败，请检查 Secret / Key 是否正确。") from exc
    meta = parse_rakuten_shop_master_xml(response.text)
    if not meta.get("shopCode") or not meta.get("shopName"):
        raise RuntimeError("未能从乐天接口读取到店铺编号和店铺名称。")
    return meta


def fetch_rakuten_cabinet_usage(service_secret: str, license_key: str) -> dict[str, int]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 和乐天 Key 不能为空。")
    try:
        response = rakuten_cabinet_request(
            "GET",
            RAKUTEN_CABINET_USAGE_URL,
            timeout=settings.crawler_timeout_seconds,
            headers={
                "Authorization": build_rakuten_authorization_header(service_secret, license_key),
                "Accept": "application/xml, text/xml",
            },
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError("乐天 Cabinet 使用量读取失败，请检查 Secret / Key 权限。") from exc
    return parse_rakuten_cabinet_usage_xml(response.text)


def parse_rakuten_cabinet_usage_xml(xml_text: str) -> dict[str, int]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError("乐天 Cabinet 使用量返回格式无法解析。") from exc

    values = {
        "usedFolderCount": 0,
        "remainingFolderCount": 0,
    }
    tag_map = {
        "UseFolderCount": "usedFolderCount",
        "useFolderCount": "usedFolderCount",
        "AvailFolderCount": "remainingFolderCount",
        "availFolderCount": "remainingFolderCount",
    }
    for element in root.iter():
        local_name = element.tag.split("}", 1)[-1]
        target_key = tag_map.get(local_name)
        if not target_key:
            continue
        try:
            values[target_key] = int(float(normalize_text(element.text) or 0))
        except ValueError:
            values[target_key] = 0
    return values


def fetch_rakuten_cabinet_folders(service_secret: str, license_key: str) -> list[dict[str, Any]]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    folders: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    offset = 1
    while True:
        response = rakuten_cabinet_request(
            "GET",
            RAKUTEN_CABINET_FOLDERS_GET_URL,
            timeout=settings.crawler_timeout_seconds,
            headers={
                "Authorization": build_rakuten_authorization_header(service_secret, license_key),
                "Accept": "application/xml, text/xml",
            },
            params={"offset": offset, "limit": RAKUTEN_CABINET_FOLDER_PAGE_SIZE},
        )
        try:
            response.raise_for_status()
        except requests.RequestException as exc:
            detail = normalize_text(response.text)
            message = "R-Cabinet 文件夹列表读取失败"
            if detail:
                message = f"{message}：{detail[:500]}"
            raise RuntimeError(message) from exc

        page_folders = parse_rakuten_cabinet_folders_xml(response.text)
        new_count = 0
        for folder in page_folders:
            folder_id = folder.get("folderId")
            if folder_id is None or folder_id in seen_ids:
                continue
            folders.append(folder)
            seen_ids.add(folder_id)
            new_count += 1
        if len(page_folders) < RAKUTEN_CABINET_FOLDER_PAGE_SIZE or new_count == 0:
            break
        offset += 1
    return folders


def parse_rakuten_cabinet_folders_xml(xml_text: str) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError("R-Cabinet 文件夹列表返回格式无法解析。") from exc
    folders: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    for element in root.iter():
        children = list(element)
        if not children:
            continue
        values: dict[str, str] = {}
        for child in children:
            local_name = child.tag.split("}", 1)[-1].lower()
            if local_name in {
                "folderid",
                "foldername",
                "folderpath",
                "foldernode",
                "directoryname",
                "filecount",
                "imagecount",
                "folderfilecount",
            }:
                values[local_name] = normalize_text(child.text)
        folder_id = parse_optional_int(values.get("folderid"))
        if folder_id is None or folder_id in seen_ids:
            continue
        seen_ids.add(folder_id)
        folder_name = (
            values.get("foldername")
            or values.get("foldernode")
            or values.get("directoryname")
            or Path(values.get("folderpath") or "").name
        )
        folders.append(
            {
                "folderId": folder_id,
                "folderName": folder_name,
                "directoryName": values.get("directoryname", ""),
                "folderPath": values.get("folderpath", ""),
                "fileCount": parse_optional_int(
                    values.get("filecount") or values.get("imagecount") or values.get("folderfilecount")
                )
                or 0,
            }
        )
    return folders


def parse_optional_int(value: Any) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def create_rakuten_cabinet_folder(
    service_secret: str,
    license_key: str,
    *,
    folder_name: str,
    directory_name: str,
) -> dict[str, Any]:
    normalized_folder_name = normalize_cabinet_folder_name(folder_name)
    normalized_directory_name = normalize_cabinet_directory_name(directory_name)
    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<request>"
        "<folderInsertRequest>"
        "<folder>"
        f"<folderName>{xml_escape_text(normalized_folder_name)}</folderName>"
        f"<directoryName>{xml_escape_text(normalized_directory_name)}</directoryName>"
        "</folder>"
        "</folderInsertRequest>"
        "</request>"
    )
    response = rakuten_cabinet_request(
        "POST",
        RAKUTEN_CABINET_FOLDER_INSERT_URL,
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/xml, text/xml",
            "Content-Type": "text/xml; charset=utf-8",
        },
        data=xml_body.encode("utf-8"),
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        result: dict[str, Any] = {}
        if detail:
            try:
                result = parse_rakuten_cabinet_folder_insert_xml(response.text, allow_error=True)
            except RuntimeError:
                result = {}
        if is_cabinet_same_folder_path_result(result):
            raise CabinetFolderAlreadyExistsError(normalized_directory_name, detail) from exc
        message = f"R-Cabinet 文件夹 {normalized_directory_name} 创建失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc
    result = parse_rakuten_cabinet_folder_insert_xml(response.text, allow_error=True)
    if is_cabinet_same_folder_path_result(result):
        raise CabinetFolderAlreadyExistsError(normalized_directory_name, response.text)
    if result.get("folderId") is None:
        result_code = normalize_text(result.get("resultCode"))
        result_message = normalize_text(result.get("message"))
        detail_parts = [part for part in [f"resultCode={result_code}" if result_code else "", result_message] if part]
        detail = "，".join(detail_parts) or normalize_text(response.text)[:500]
        raise RuntimeError(f"R-Cabinet 文件夹 {normalized_directory_name} 创建失败：{detail}")
    result["folderName"] = result.get("folderName") or normalized_folder_name
    result["directoryName"] = result.get("directoryName") or normalized_directory_name
    result["folderPath"] = result.get("folderPath") or normalized_directory_name
    result["fileCount"] = int(result.get("fileCount") or 0)
    return result


def parse_rakuten_cabinet_folder_insert_xml(xml_text: str, *, allow_error: bool = False) -> dict[str, Any]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError("R-Cabinet 文件夹创建返回格式无法解析。") from exc
    result: dict[str, Any] = {}
    for element in root.iter():
        local_name = element.tag.split("}", 1)[-1].lower()
        text = normalize_text(element.text)
        if not text:
            continue
        if local_name in {"systemstatus", "status"}:
            result["systemStatus"] = text
        elif local_name == "message":
            result["message"] = text
        elif local_name in {"resultcode", "code"}:
            result["resultCode"] = text
        elif local_name == "folderid":
            folder_id = parse_optional_int(text)
            if folder_id is not None:
                result["folderId"] = folder_id
        elif local_name in {"foldername", "foldernode"}:
            result["folderName"] = text
        elif local_name == "directoryname":
            result["directoryName"] = text
        elif local_name == "folderpath":
            result["folderPath"] = text
        elif local_name in {"filecount", "imagecount", "folderfilecount"}:
            result["fileCount"] = parse_optional_int(text) or 0
    if result.get("folderId") is None:
        if allow_error:
            return result
        raise RuntimeError("R-Cabinet 文件夹创建成功但未返回 folderId。")
    return result


class CabinetFolderAlreadyExistsError(RuntimeError):
    def __init__(self, directory_name: str, detail: str = "") -> None:
        self.directory_name = normalize_cabinet_directory_name(directory_name)
        self.detail = normalize_text(detail)
        super().__init__(f"R-Cabinet 文件夹 {self.directory_name} 已存在。")


def is_cabinet_same_folder_path_result(result: dict[str, Any]) -> bool:
    result_code = normalize_text(result.get("resultCode"))
    message = normalize_text(result.get("message") or result.get("detail")).lower()
    return result_code == "3015" or "same folder path" in message


@contextmanager
def rakuten_cabinet_folder_creation_lock(
    service_secret: str,
    license_key: str,
) -> Iterator[None]:
    local_acquired = RAKUTEN_CABINET_FOLDER_CREATE_LOCK.acquire(
        timeout=RAKUTEN_CABINET_FOLDER_LOCK_TIMEOUT_SECONDS,
    )
    if not local_acquired:
        raise RuntimeError("等待 R-Cabinet 文件夹创建锁超时，请稍后重试。")
    distributed_lock = None
    distributed_acquired = False
    try:
        if should_use_redis_task_queue():
            account_key = rakuten_request_account_key(
                {
                    "Authorization": build_rakuten_authorization_header(
                        service_secret,
                        license_key,
                    )
                }
            )
            try:
                distributed_lock = redis_connection().lock(
                    f"lt:lock:cabinet-folder:{account_key}",
                    timeout=RAKUTEN_CABINET_FOLDER_LOCK_TIMEOUT_SECONDS,
                    blocking_timeout=RAKUTEN_CABINET_FOLDER_LOCK_TIMEOUT_SECONDS,
                )
                distributed_acquired = bool(distributed_lock.acquire(blocking=True))
                if not distributed_acquired:
                    raise RuntimeError("等待 R-Cabinet 店铺文件夹创建锁超时，请稍后重试。")
            except RuntimeError:
                raise
            except Exception:
                logger.warning("获取 R-Cabinet 文件夹分布式锁失败，使用本进程锁继续", exc_info=True)
                distributed_lock = None
        yield
    finally:
        if distributed_lock is not None and distributed_acquired:
            try:
                distributed_lock.release()
            except Exception:
                logger.warning("释放 R-Cabinet 文件夹分布式锁失败", exc_info=True)
        RAKUTEN_CABINET_FOLDER_CREATE_LOCK.release()


def wait_for_visible_listing_cabinet_folder(
    service_secret: str,
    license_key: str,
    directory_name: str,
    required_slots: int,
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    folders: list[dict[str, Any]] = []
    for delay_seconds in RAKUTEN_CABINET_FOLDER_VISIBILITY_DELAYS_SECONDS:
        time.sleep(delay_seconds)
        folders = fetch_rakuten_cabinet_folders(service_secret, license_key)
        existing = find_listing_cabinet_folder_by_directory(folders, directory_name)
        if existing and cabinet_folder_remaining_slots(existing) >= required_slots:
            return prepare_listing_cabinet_folder(existing), folders
    return None, folders


def ensure_listing_cabinet_folder(
    service_secret: str,
    license_key: str,
    store: StoreModel,
    required_slots: int,
    *,
    usage: dict[str, int] | None = None,
) -> dict[str, Any]:
    with rakuten_cabinet_folder_creation_lock(service_secret, license_key):
        return ensure_listing_cabinet_folder_locked(
            service_secret,
            license_key,
            store,
            required_slots,
            usage=usage,
        )


def ensure_listing_cabinet_folder_locked(
    service_secret: str,
    license_key: str,
    store: StoreModel,
    required_slots: int,
    *,
    usage: dict[str, int] | None = None,
) -> dict[str, Any]:
    folders = fetch_rakuten_cabinet_folders(service_secret, license_key)
    candidates = [
        folder
        for folder in folders
        if listing_cabinet_folder_identity(folder) is not None
    ]
    for folder in sorted(candidates, key=cabinet_listing_folder_sort_key):
        if cabinet_folder_remaining_slots(folder) > 0:
            return prepare_listing_cabinet_folder(folder)

    usage = usage or fetch_rakuten_cabinet_usage(service_secret, license_key)
    if int(usage.get("remainingFolderCount") or 0) <= 0:
        raise RuntimeError("R-Cabinet 没有可用文件夹数量，不能自动创建新的图片文件夹。")

    created_date = datetime.now().strftime("%Y%m%d")
    next_batch = next_listing_cabinet_batch_number(candidates, created_date)
    last_error: Exception | None = None
    for batch in range(next_batch, next_batch + RAKUTEN_CABINET_FOLDER_CREATE_ATTEMPTS):
        directory_name = listing_cabinet_directory_name(created_date, batch)
        existing = find_listing_cabinet_folder_by_directory(folders, directory_name)
        if existing:
            if cabinet_folder_remaining_slots(existing) > 0:
                return prepare_listing_cabinet_folder(existing)
            continue
        try:
            folder = create_rakuten_cabinet_folder(
                service_secret,
                license_key,
                folder_name=listing_cabinet_folder_display_name(store, batch),
                directory_name=directory_name,
            )
            folder["directoryName"] = directory_name
            return folder
        except CabinetFolderAlreadyExistsError as exc:
            last_error = exc
            existing, folders = wait_for_visible_listing_cabinet_folder(
                service_secret,
                license_key,
                directory_name,
                required_slots,
            )
            if existing:
                return existing
            continue
        except Exception as exc:
            last_error = exc
            raise
    if last_error is not None:
        raise RuntimeError(f"R-Cabinet 自动创建图片文件夹失败：{last_error}") from last_error
    raise RuntimeError("R-Cabinet 自动创建图片文件夹失败。")


def listing_cabinet_folder_directory(folder: dict[str, Any]) -> str:
    for value in listing_cabinet_folder_directory_candidates(folder):
        if value:
            return value
    return ""


def listing_cabinet_folder_directory_candidates(folder: dict[str, Any]) -> list[str]:
    values = [
        normalize_text(folder.get("folderPath")),
        normalize_text(folder.get("directoryName")),
        normalize_text(folder.get("folderName")),
    ]
    return [value for value in values if value]


def listing_cabinet_folder_identity(folder: dict[str, Any]) -> tuple[str, int] | None:
    for value in listing_cabinet_folder_directory_candidates(folder):
        normalized = normalize_cabinet_directory_name(value)
        match = re.fullmatch(r"yx(\d{8})-(\d+)", normalized)
        if match:
            return match.group(1), int(match.group(2))
    return None


def prepare_listing_cabinet_folder(folder: dict[str, Any]) -> dict[str, Any]:
    folder["directoryName"] = listing_cabinet_folder_directory(folder)
    return folder


def find_listing_cabinet_folder_by_directory(folders: list[dict[str, Any]], directory_name: str) -> dict[str, Any] | None:
    normalized_directory = normalize_cabinet_directory_name(directory_name)
    for folder in folders:
        if any(
            normalize_cabinet_directory_name(value) == normalized_directory
            for value in listing_cabinet_folder_directory_candidates(folder)
        ):
            return folder
    return None


def listing_cabinet_directory_name(created_date: str, batch: int) -> str:
    return normalize_cabinet_directory_name(f"YX{created_date}-{max(1, int(batch))}")


def listing_cabinet_folder_display_name(store: StoreModel, batch: int) -> str:
    return normalize_cabinet_folder_name(f"YX{datetime.now():%Y%m%d}-{max(1, int(batch))}")


def cabinet_listing_folder_sort_key(folder: dict[str, Any]) -> tuple[str, int, int]:
    identity = listing_cabinet_folder_identity(folder)
    if identity is None:
        return ("99999999", 0, int(folder.get("folderId") or 0))
    created_date, batch = identity
    return (created_date, batch, int(folder.get("folderId") or 0))


def next_listing_cabinet_batch_number(folders: list[dict[str, Any]], created_date: str) -> int:
    max_batch = 0
    for folder in folders:
        identity = listing_cabinet_folder_identity(folder)
        if identity is not None and identity[0] == created_date:
            max_batch = max(max_batch, identity[1])
    return max_batch + 1


def cabinet_folder_remaining_slots(folder: dict[str, Any]) -> int:
    return max(0, RAKUTEN_CABINET_BATCH_FOLDER_IMAGE_LIMIT - int(folder.get("fileCount") or 0))


def normalize_cabinet_folder_name(value: str) -> str:
    text = normalize_text(value) or "LT Images"
    text = re.sub(r"<[^>]*>", "", text).replace("　", " ")
    text = re.sub(r"[\x00-\x1f\x7f]", "", text).strip()
    if not text:
        text = "LT Images"
    encoded = text.encode("utf-8")
    while len(encoded) > 50 and text:
        text = text[:-1]
        encoded = text.encode("utf-8")
    return text or "LT Images"


def normalize_cabinet_directory_name(value: str) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9_-]+", "-", text).strip("-_")
    if not text:
        text = f"lt-{uuid.uuid4().hex[:8]}"
    return text[:20]


def normalize_cabinet_directory_segment(value: str, *, max_length: int) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9_-]+", "-", text).strip("-_")
    if not text:
        text = hashlib.sha1(normalize_text(value).encode("utf-8")).hexdigest()[:max_length]
    return text[:max_length] or uuid.uuid4().hex[:max_length]


def parse_rakuten_shop_master_xml(xml_text: str) -> dict[str, str]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError("乐天店铺信息返回格式无法解析。") from exc

    shop_meta = {
        "shopId": "",
        "shopCode": "",
        "shopName": "",
    }
    shop_name_tags = {"shopName", "shopname", "shop_name", "storeName", "storename", "name", "title"}
    shop_code_tags = {"url", "shopUrl", "shopURL", "shop_url", "shopCode", "shopcode", "shop_code"}
    shop_id_tags = {"shopId", "shopid", "shop_id"}

    for element in root.iter():
        local_name = element.tag.split("}", 1)[-1]
        text_value = normalize_text(element.text)
        if not text_value:
            continue
        if not shop_meta["shopName"] and local_name in shop_name_tags:
            shop_meta["shopName"] = text_value
        if not shop_meta["shopCode"] and local_name in shop_code_tags:
            shop_meta["shopCode"] = normalize_shop_code(text_value)
        if not shop_meta["shopId"] and local_name in shop_id_tags:
            shop_meta["shopId"] = text_value
    return shop_meta


def fetch_rakuten_store_items(service_secret: str, license_key: str) -> list[dict[str, Any]]:
    items, _ = fetch_rakuten_store_items_with_total(service_secret, license_key)
    return items


def fetch_rakuten_store_items_with_total(
    service_secret: str,
    license_key: str,
) -> tuple[list[dict[str, Any]], int | None]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    headers = {
        "Authorization": build_rakuten_authorization_header(service_secret, license_key),
        "Accept": "application/json",
    }
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    offset = 0
    reported_total_count: int | None = None
    while True:
        payload = request_rakuten_items_page(headers, offset)
        reported_total_count = (
            reported_total_count
            if reported_total_count is not None
            else parse_rakuten_total_count(payload)
        )

        page_items = extract_rakuten_item_candidates(payload)
        for item in page_items:
            item_key = normalize_text(
                first_text_from_keys(item, ("manageNumber", "itemNumber", "itemUrl", "itemPageUrl"))
            )
            if not item_key or item_key in seen:
                continue
            seen.add(item_key)
            items.append(item)

        raw_results = payload.get("results")
        raw_page_count = len(raw_results) if isinstance(raw_results, list) else len(page_items)
        offset += RAKUTEN_ITEM_SEARCH_HITS
        if raw_page_count == 0:
            break
        if reported_total_count is not None and offset >= reported_total_count:
            break
        if reported_total_count is None and raw_page_count < RAKUTEN_ITEM_SEARCH_HITS:
            break

    # Search results can temporarily contain manage-number-only entries for
    # recently deleted products. They are not readable products and must not
    # inflate store totals or prevent a complete local reconciliation.
    return items, len(items)


def request_rakuten_items_page(headers: dict[str, str], offset: int) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(RAKUTEN_ITEM_SEARCH_MAX_RETRIES):
        try:
            response = requests.get(
                RAKUTEN_ITEM_SEARCH_URL,
                timeout=settings.crawler_timeout_seconds,
                headers=headers,
                params={"hits": RAKUTEN_ITEM_SEARCH_HITS, "offset": offset},
            )
            if response.status_code == 429 and attempt < RAKUTEN_ITEM_SEARCH_MAX_RETRIES - 1:
                retry_after = response.headers.get("Retry-After")
                wait_seconds = float(retry_after) if retry_after and retry_after.isdecimal() else 1.5 * (attempt + 1)
                time.sleep(wait_seconds)
                continue
            response.raise_for_status()
            payload = response.json()
            if not isinstance(payload, dict):
                raise RuntimeError("乐天商品接口返回格式无法解析。")
            return payload
        except ValueError as exc:
            raise RuntimeError("乐天商品接口返回格式无法解析。") from exc
        except requests.RequestException as exc:
            last_error = exc
            if attempt < RAKUTEN_ITEM_SEARCH_MAX_RETRIES - 1:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise RuntimeError(f"乐天商品更新失败，读取 offset={offset} 分页时失败，请检查店铺密钥权限或稍后重试。") from exc
    raise RuntimeError(f"乐天商品更新失败，读取 offset={offset} 分页时失败：{last_error}")


def patch_rakuten_item_listing_status(
    service_secret: str,
    license_key: str,
    manage_number: str,
    *,
    listing_status: str,
) -> None:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能更新上架状态。")
    if listing_status not in {"listed", "unlisted"}:
        raise RuntimeError("上架状态不合法。")

    visible = listing_status == "listed"
    payload = {
        "hideItem": not visible,
        "features": {
            "searchVisibility": "ALWAYS_VISIBLE" if visible else "ALWAYS_HIDDEN",
        },
    }
    response = requests.patch(
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        },
        json=payload,
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 状态更新失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc


def patch_rakuten_item_price(
    service_secret: str,
    license_key: str,
    manage_number: str,
    raw_payload: dict[str, Any],
    price: Decimal,
) -> dict[str, Any]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能修改价格。")

    variants = raw_payload.get("variants")
    if not isinstance(variants, dict) or not variants:
        raise RuntimeError("当前商品没有可修改的 SKU 款式价格，不能同步到乐天。")

    price_text = str(int(price)) if price == price.to_integral_value() else format(price, "f")
    patch_variants: dict[str, dict[str, str]] = {}
    for variant_id, variant in variants.items():
        if not isinstance(variant, dict):
            continue
        normalized_variant_id = normalize_text(variant_id)
        if normalized_variant_id:
            patch_variants[normalized_variant_id] = {"standardPrice": price_text}
    if not patch_variants:
        raise RuntimeError("当前商品没有可修改的 SKU 款式价格，不能同步到乐天。")

    response = requests.patch(
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        },
        json={"variants": patch_variants},
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 价格更新失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc

    updated_payload = dict(raw_payload)
    updated_variants = dict(variants)
    for variant_id, variant in updated_variants.items():
        if isinstance(variant, dict):
            next_variant = dict(variant)
            next_variant["standardPrice"] = price_text
            updated_variants[variant_id] = next_variant
    updated_payload["variants"] = updated_variants
    updated_payload["updated"] = datetime.now().isoformat(timespec="seconds")
    return updated_payload


def patch_rakuten_item_detail(
    service_secret: str,
    license_key: str,
    manage_number: str,
    raw_payload: dict[str, Any],
    *,
    title: str,
    tagline: str,
    genre_id: str,
    variants: list[Any],
) -> dict[str, Any]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能同步修改乐天商品。")

    normalized_title = normalize_text(title)
    if not normalized_title:
        raise RuntimeError("商品标题不能为空。")
    normalized_genre_id = normalize_text(genre_id)
    if not re.fullmatch(r"\d{6}", normalized_genre_id) or not rakuten_genre_path(normalized_genre_id):
        raise RuntimeError("请选择有效品类。")

    raw_variants = raw_payload.get("variants")
    if not isinstance(raw_variants, dict) or not raw_variants:
        raise RuntimeError("当前商品没有可修改的 SKU 款式，不能同步到乐天。")

    patch_variants: dict[str, dict[str, Any]] = {}
    for variant in variants:
        variant_id = normalize_text(getattr(variant, "variantId", ""))
        if not variant_id or variant_id not in raw_variants:
            raise RuntimeError(f"SKU {variant_id or '-'} 不存在，不能同步修改。")
        standard_price = getattr(variant, "standardPrice", None)
        if standard_price is None or standard_price <= 0:
            raise RuntimeError(f"SKU {variant_id} 价格必须大于 0。")
        if standard_price != standard_price.to_integral_value():
            raise RuntimeError(f"SKU {variant_id} 价格必须为日元整数。")
        patch_variants[variant_id] = {
            "standardPrice": str(int(standard_price)),
            "hidden": bool(getattr(variant, "hidden", False)),
        }

    if not patch_variants:
        raise RuntimeError("请至少保留一个可修改的 SKU 款式。")

    payload = {
        "title": normalized_title,
        "tagline": str(tagline or "").strip(),
        "genreId": normalized_genre_id,
        "variants": patch_variants,
    }
    response = requests.patch(
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        },
        json=payload,
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 详情更新失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc

    updated_payload = dict(raw_payload)
    updated_payload["title"] = normalized_title
    updated_payload["tagline"] = str(tagline or "").strip()
    updated_payload["genreId"] = normalized_genre_id
    updated_variants = dict(raw_variants)
    for variant_id, variant_patch in patch_variants.items():
        current_variant = raw_variants.get(variant_id)
        if isinstance(current_variant, dict):
            next_variant = dict(current_variant)
            next_variant.update(variant_patch)
            updated_variants[variant_id] = next_variant
    updated_payload["variants"] = updated_variants
    updated_payload["updated"] = datetime.now().isoformat(timespec="seconds")
    return updated_payload


def patch_rakuten_item_images(
    service_secret: str,
    license_key: str,
    manage_number: str,
    uploaded_images: list[dict[str, str]],
    title: str,
) -> None:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能同步修改乐天商品图片。")
    images = build_rakuten_listing_images(uploaded_images, title)
    if not images:
        raise RuntimeError("商品缺少可同步的 R-Cabinet 图片。")
    response = requests.patch(
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        },
        json={"images": images},
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 图片更新失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc


def patch_rakuten_item_visibility(
    service_secret: str,
    license_key: str,
    manage_number: str,
    *,
    hide_item: bool,
) -> None:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能更新乐天商品显示状态。")
    response = request_rakuten_write(
        "PATCH",
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        operation=f"乐天商品 {normalized_manage_number} 显示状态更新",
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        },
        json={"hideItem": bool(hide_item)},
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 显示状态更新失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc


def patch_local_item_detail(raw_payload: dict[str, Any], *, title: str, tagline: str, variants: list[Any]) -> dict[str, Any]:
    normalized_title = normalize_text(title)
    if not normalized_title:
        raise RuntimeError("商品标题不能为空。")

    updated_payload = dict(raw_payload)
    updated_payload["title"] = normalized_title
    updated_payload["itemName"] = normalized_title
    updated_payload["tagline"] = str(tagline or "").strip()

    raw_variants = updated_payload.get("variants")
    variant_updates: dict[str, Any] = {}
    single_product_price: Decimal | None = None
    for variant in variants:
        variant_id = normalize_text(getattr(variant, "variantId", ""))
        standard_price = getattr(variant, "standardPrice", None)
        if not variant_id:
            continue
        price_label = "价格" if variant_id == SINGLE_PRODUCT_VARIANT_ID and not raw_variants else f"SKU {variant_id} 价格"
        if standard_price is None or standard_price <= 0:
            raise RuntimeError(f"{price_label}必须大于 0。")
        if standard_price != standard_price.to_integral_value():
            raise RuntimeError(f"{price_label}必须为日元整数。")
        if variant_id == SINGLE_PRODUCT_VARIANT_ID and not raw_variants:
            single_product_price = standard_price.to_integral_value()
            continue
        variant_updates[variant_id] = {
            "standardPrice": str(int(standard_price)),
            "price": str(int(standard_price)),
            "hidden": bool(getattr(variant, "hidden", False)),
        }

    if isinstance(raw_variants, dict):
        updated_variants = dict(raw_variants)
        for variant_id, variant_patch in variant_updates.items():
            current_variant = updated_variants.get(variant_id)
            if isinstance(current_variant, dict):
                next_variant = dict(current_variant)
                next_variant.update(variant_patch)
                updated_variants[variant_id] = next_variant
        updated_payload["variants"] = updated_variants
    elif isinstance(raw_variants, list):
        updated_variants = []
        for index, current_variant in enumerate(raw_variants):
            if not isinstance(current_variant, dict):
                updated_variants.append(current_variant)
                continue
            variant_id = first_text_from_keys(current_variant, ("variantId", "skuId", "merchantDefinedSkuId")) or f"sku-{index + 1}"
            next_variant = dict(current_variant)
            if variant_id in variant_updates:
                next_variant.update(variant_updates[variant_id])
            updated_variants.append(next_variant)
        updated_payload["variants"] = updated_variants
    elif variant_updates:
        updated_payload["variants"] = {
            variant_id: {"variantId": variant_id, **variant_patch}
            for variant_id, variant_patch in variant_updates.items()
        }
    if single_product_price is not None:
        price_text = str(int(single_product_price))
        updated_payload["price"] = price_text
        updated_payload["standardPrice"] = price_text
        updated_payload["itemPrice"] = price_text

    updated_payload["updated"] = datetime.now().isoformat(timespec="seconds")
    return updated_payload


def rakuten_item_delete_target_missing(response: requests.Response) -> bool:
    try:
        payload = json.loads(response.text or "{}")
    except (TypeError, ValueError):
        return False
    if not isinstance(payload, dict):
        return False
    errors = payload.get("errors")
    if not isinstance(errors, list):
        return False
    return any(
        isinstance(error, dict)
        and normalize_text(error.get("code")).upper() == "GE0014"
        for error in errors
    )


def delete_rakuten_item(service_secret: str, license_key: str, manage_number: str) -> None:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能删除乐天商品。")

    wait_for_rakuten_item_delete_slot()
    response = request_rakuten_write(
        "DELETE",
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        operation=f"乐天商品 {normalized_manage_number} 删除",
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
        },
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        if rakuten_item_delete_target_missing(response):
            return
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 删除失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc


def bulk_upsert_rakuten_inventories(
    service_secret: str,
    license_key: str,
    inventories: list[dict[str, Any]],
) -> None:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    if not inventories:
        return
    for offset in range(0, len(inventories), RAKUTEN_INVENTORY_BULK_UPSERT_LIMIT):
        chunk = inventories[offset : offset + RAKUTEN_INVENTORY_BULK_UPSERT_LIMIT]
        response = request_rakuten_write(
            "POST",
            RAKUTEN_INVENTORY_BULK_UPSERT_URL,
            operation="乐天库存/发货信息登记",
            headers={
                "Authorization": build_rakuten_authorization_header(service_secret, license_key),
                "Accept": "application/json",
                "Content-Type": "application/json; charset=utf-8",
            },
            json={"inventories": chunk},
        )
        try:
            response.raise_for_status()
        except requests.RequestException as exc:
            detail = normalize_text(response.text)
            message = "乐天库存/发货信息登记失败"
            if detail:
                message = f"{message}：{detail[:800]}"
            raise RuntimeError(message) from exc


def delete_rakuten_cabinet_file(service_secret: str, license_key: str, file_id: int) -> None:
    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<request>"
        "<fileDeleteRequest>"
        "<file>"
        f"<fileId>{int(file_id)}</fileId>"
        "</file>"
        "</fileDeleteRequest>"
        "</request>"
    )
    response = rakuten_cabinet_request(
        "POST",
        RAKUTEN_CABINET_FILE_DELETE_URL,
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/xml, text/xml",
            "Content-Type": "application/xml; charset=utf-8",
        },
        data=xml_body.encode("utf-8"),
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"R-Cabinet 图片 {file_id} 删除失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc
    error_message = cabinet_xml_error_message(response.text)
    if error_message:
        raise RuntimeError(f"R-Cabinet 图片 {file_id} 删除失败：{error_message[:500]}")


def insert_rakuten_cabinet_file(
    service_secret: str,
    license_key: str,
    *,
    file_name: str,
    file_path: str,
    content: bytes,
    content_type: str,
    folder_id: int = 0,
    overwrite: bool = True,
) -> dict[str, Any]:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    if not content:
        raise RuntimeError("图片内容为空，不能上传到 R-Cabinet。")
    normalized_file_name = normalize_cabinet_file_name(file_name)
    normalized_file_path = normalize_cabinet_file_path(file_path)
    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<request>"
        "<fileInsertRequest>"
        "<file>"
        f"<fileName>{xml_escape_text(normalized_file_name)}</fileName>"
        f"<folderId>{int(folder_id)}</folderId>"
        f"<filePath>{xml_escape_text(normalized_file_path)}</filePath>"
        f"<overWrite>{str(bool(overwrite)).lower()}</overWrite>"
        "</file>"
        "</fileInsertRequest>"
        "</request>"
    )
    response = rakuten_cabinet_request(
        "POST",
        RAKUTEN_CABINET_FILE_INSERT_URL,
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/xml, text/xml",
        },
        data={"xml": xml_body},
        files={"file": (normalized_file_path, content, content_type or "application/octet-stream")},
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"R-Cabinet 图片 {normalized_file_path} 上传失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc
    error_message = cabinet_xml_error_message(response.text)
    if error_message:
        raise RuntimeError(f"R-Cabinet 图片 {normalized_file_path} 上传失败：{error_message[:500]}")
    result = parse_cabinet_insert_xml(response.text)
    result["fileName"] = normalized_file_name
    result["filePath"] = normalized_file_path
    return result


def parse_cabinet_insert_xml(xml_text: str) -> dict[str, Any]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError("R-Cabinet 图片上传返回格式无法解析。") from exc
    result: dict[str, Any] = {}
    for element in root.iter():
        local_name = element.tag.split("}", 1)[-1].lower()
        text = normalize_text(element.text)
        if local_name == "fileid" and text:
            try:
                result["fileId"] = int(float(text))
            except ValueError:
                pass
        elif local_name == "message" and text:
            result["message"] = text
        elif local_name == "systemstatus" and text:
            result["systemStatus"] = text
        elif local_name in {"fileurl", "fileurlssl", "url"} and text:
            result["fileUrl"] = text
        elif local_name in {"resultcode", "code"} and text:
            result["resultCode"] = text
    return result


def xml_escape_text(value: Any) -> str:
    return (
        str(value or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def normalize_cabinet_file_name(value: str) -> str:
    text = normalize_text(value) or "product"
    text = unicodedata.normalize("NFKC", re.sub(r"<[^>]*>", "", text)).replace("　", " ")
    text = re.sub(r"[\x00-\x1f\x7f]", "", text).strip()
    text = re.sub(r"[^A-Za-z0-9_-]+", "-", text).strip("-_")
    if not text:
        text = "product"
    encoded = text.encode("utf-8")
    while len(encoded) > 50 and text:
        text = text[:-1]
        encoded = text.encode("utf-8")
    return text or "product"


def normalize_cabinet_file_path(value: str) -> str:
    text = normalize_text(value).lower()
    suffix = Path(text).suffix.lower()
    stem = Path(text).stem.lower()
    if suffix == ".jpeg":
        suffix = ".jpg"
    if suffix not in {".jpg", ".png", ".gif"}:
        suffix = ".jpg"
    stem = re.sub(r"[^a-z0-9_-]+", "-", stem).strip("-_")
    if not stem:
        stem = uuid.uuid4().hex[:12]
    stem = stem[: max(1, 20 - len(suffix))]
    if re.fullmatch(r"img\d{8}|imgrc\d{10}", stem):
        stem = f"lt-{stem}"[: max(1, 20 - len(suffix))]
    return f"{stem}{suffix}"


def search_rakuten_cabinet_file_ids(
    service_secret: str,
    license_key: str,
    *,
    file_path: str = "",
    file_name: str = "",
) -> list[int]:
    params: dict[str, Any] = {"offset": 1, "limit": 100}
    if file_path:
        params["filePath"] = file_path
    if file_name:
        params["fileName"] = file_name
    response = rakuten_cabinet_request(
        "GET",
        RAKUTEN_CABINET_FILE_SEARCH_URL,
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/xml, text/xml",
        },
        params=params,
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = "R-Cabinet 图片搜索失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc
    return [record["fileId"] for record in parse_cabinet_files_xml(response.text) if record.get("fileId")]


def parse_cabinet_file_ids_xml(xml_text: str) -> list[int]:
    return [record["fileId"] for record in parse_cabinet_files_xml(xml_text) if record.get("fileId")]


def search_rakuten_cabinet_files(
    service_secret: str,
    license_key: str,
    *,
    file_path: str = "",
    file_name: str = "",
) -> list[dict[str, Any]]:
    params: dict[str, Any] = {"offset": 1, "limit": 100}
    if file_path:
        params["filePath"] = file_path
    if file_name:
        params["fileName"] = file_name
    response = rakuten_cabinet_request(
        "GET",
        RAKUTEN_CABINET_FILE_SEARCH_URL,
        timeout=settings.crawler_timeout_seconds,
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/xml, text/xml",
        },
        params=params,
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = "R-Cabinet 图片搜索失败"
        if detail:
            message = f"{message}：{detail[:500]}"
        raise RuntimeError(message) from exc
    return parse_cabinet_files_xml(response.text)


def parse_cabinet_files_xml(xml_text: str) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError("R-Cabinet 图片搜索返回格式无法解析。") from exc
    records: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    for element in root.iter():
        children = list(element)
        if not children:
            continue
        values: dict[str, str] = {}
        for child in children:
            local_name = child.tag.split("}", 1)[-1].lower()
            if local_name in {"fileid", "filename", "filepath", "fileurl", "folderpath"}:
                values[local_name] = normalize_text(child.text)
        raw_file_id = values.get("fileid", "")
        if not raw_file_id:
            continue
        try:
            file_id = int(float(raw_file_id))
        except ValueError:
            continue
        if file_id in seen_ids:
            continue
        seen_ids.add(file_id)
        records.append(
            {
                "fileId": file_id,
                "fileName": values.get("filename", ""),
                "filePath": values.get("filepath", ""),
                "fileUrl": values.get("fileurl", ""),
                "folderPath": values.get("folderpath", ""),
            }
        )
    return records


def parse_rakuten_total_count(payload: dict[str, Any]) -> int | None:
    for key in ("numFound", "totalCount", "total", "count"):
        value = payload.get(key)
        if value is None:
            continue
        try:
            return int(value)
        except (TypeError, ValueError):
            continue
    return None


def extract_rakuten_item_candidates(payload: Any) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            if is_rakuten_item_candidate(value):
                candidates.append(value)
                return
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(payload)
    return candidates


def is_rakuten_item_candidate(value: dict[str, Any]) -> bool:
    identity = first_text_from_keys(value, ("manageNumber", "itemNumber", "itemUrl", "itemPageUrl"))
    title = first_text_from_keys(value, ("itemName", "title", "name"))
    return bool(identity and title)


def first_text_from_keys(source: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        text = first_text_value(source.get(key))
        if text:
            return text
    return ""


def first_text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_text(value)
    if isinstance(value, (int, float, Decimal)):
        return normalize_text(value)
    if isinstance(value, dict):
        for key in ("value", "text", "name", "title", "url"):
            text = first_text_value(value.get(key))
            if text:
                return text
    if isinstance(value, list):
        for item in value:
            text = first_text_value(item)
            if text:
                return text
    return ""


def first_url_from_keys(source: dict[str, Any], keys: tuple[str, ...], *, shop_code: str = "") -> str:
    for key in keys:
        url = first_url_value(source.get(key), shop_code=shop_code)
        if url:
            return url
    return ""


def first_url_value(value: Any, *, shop_code: str = "") -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = normalize_text(value)
        if text.startswith(("http://", "https://")):
            return text
        if shop_code and text.startswith("/"):
            return build_rakuten_cabinet_image_url(shop_code, text)
        return ""
    if isinstance(value, dict):
        for key in ("url", "imageUrl", "itemUrl", "itemPageUrl", "location", "value"):
            url = first_url_value(value.get(key), shop_code=shop_code)
            if url:
                return url
        for child in value.values():
            url = first_url_value(child, shop_code=shop_code)
            if url:
                return url
    if isinstance(value, list):
        for item in value:
            url = first_url_value(item, shop_code=shop_code)
            if url:
                return url
    return ""


def first_rakuten_image_url(item: dict[str, Any], shop_code: str) -> str:
    for key in ("images", "imageUrl", "imageUrls", "mediumImageUrls", "smallImageUrls"):
        url = first_url_value(item.get(key), shop_code=shop_code)
        if url:
            return url
    return ""


def build_rakuten_cabinet_image_url(shop_code: str, location: str) -> str:
    normalized_shop_code = normalize_shop_code(shop_code)
    normalized_location = normalize_text(location).lstrip("/")
    if not normalized_shop_code or not normalized_location:
        return ""
    return f"https://image.rakuten.co.jp/{quote(normalized_shop_code, safe='')}/cabinet/{quote(normalized_location, safe='/')}"


RAKUTEN_PRICE_KEYS = (
    "itemPrice",
    "price",
    "standardPrice",
    "displayPrice",
    "taxIncludedPrice",
    "tax_included_price",
    "taxIncludedPriceMin",
    "taxIncludedPriceMax",
    "sellingPrice",
    "selling_price",
    "selling_price_no_format",
    "sellingPriceWithTax",
    "selling_price_with_tax",
    "selling_price_tax_included",
    "tax_included_selling_price",
)


def price_from_rakuten_item(item: dict[str, Any]) -> float | None:
    value = first_text_from_keys(item, RAKUTEN_PRICE_KEYS) or first_variant_price(item)
    if not value:
        return None
    normalized = re.sub(r"[^0-9.]", "", value)
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def price_from_rakuten_item_without_variants(item: dict[str, Any]) -> float | None:
    value = first_text_from_keys(item, RAKUTEN_PRICE_KEYS)
    if not value:
        return None
    normalized = re.sub(r"[^0-9.]", "", value)
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def price_range_from_rakuten_item(item: dict[str, Any]) -> tuple[float, float] | None:
    prices = variant_prices(item)
    if prices:
        return min(prices), max(prices)
    price = price_from_rakuten_item(item)
    if price is None:
        return None
    return price, price


def variant_prices(item: dict[str, Any]) -> list[float]:
    variants = item.get("variants")
    if isinstance(variants, dict):
        variant_items = variants.values()
    elif isinstance(variants, list):
        variant_items = variants
    else:
        variant_items = []

    prices: list[float] = []
    for variant in variant_items:
        if not isinstance(variant, dict):
            continue
        value = first_text_from_keys(variant, RAKUTEN_PRICE_KEYS)
        normalized = re.sub(r"[^0-9.]", "", value)
        if not normalized:
            continue
        try:
            prices.append(float(normalized))
        except ValueError:
            continue
    return prices


def first_variant_price(item: dict[str, Any]) -> str:
    prices = variant_prices(item)
    if not prices:
        return ""
    return str(min(prices))


def rakuten_listing_status_from_item(item: dict[str, Any]) -> str:
    features = item.get("features")
    if isinstance(features, dict):
        search_visibility = normalize_text(features.get("searchVisibility")).upper()
        if "HIDDEN" in search_visibility:
            return "unlisted"
    hide_item = item.get("hideItem")
    if isinstance(hide_item, str):
        return "unlisted" if hide_item.strip().lower() in {"1", "true", "yes", "on"} else "listed"
    return "unlisted" if bool(hide_item) else "listed"


def list_sources(owner_username: str, *, page: int | None = None, page_size: int | None = None) -> list[dict[str, Any]] | dict[str, Any]:
    with session_scope() as session:
        query = select(CrawlSourceModel).where(CrawlSourceModel.owner_username == owner_username)
        return paginate_query(
            session,
            query,
            order_by=CrawlSourceModel.created_at.desc(),
            page=page,
            page_size=page_size,
            response_key="sources",
            serializer=source_to_public,
        )


def save_source(owner_username: str, payload: Any, source_id: int | None = None) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(CrawlSourceModel, source_id) if source_id else None
        if row is None:
            row = CrawlSourceModel(owner_username=owner_username)
            session.add(row)
        if row.owner_username != owner_username:
            raise RuntimeError("不能修改其他用户的采集源。")
        row.name = str(getattr(payload, "name", "") or "").strip()
        row.source_type = str(getattr(payload, "sourceType", "") or "keyword").strip()
        row.target = str(getattr(payload, "target", "") or "").strip()
        if row.source_type == "product_url":
            row.target = normalize_rakuten_product_target(row.target)
        row.enabled = bool(getattr(payload, "enabled", True))
        row.schedule_enabled = bool(getattr(payload, "scheduleEnabled", False))
        row.interval_minutes = int(getattr(payload, "intervalMinutes", 60) or 60)
        row.notes = str(getattr(payload, "notes", "") or "").strip()
        if not row.name or not row.target:
            raise RuntimeError("采集源名称和目标不能为空。")
        session.flush()
        return source_to_public(row)


def delete_source(owner_username: str, source_id: int) -> None:
    with session_scope() as session:
        row = session.get(CrawlSourceModel, source_id)
        if row is None:
            return
        if row.owner_username != owner_username:
            raise RuntimeError("不能删除其他用户的采集源。")
        session.delete(row)


def today_range() -> tuple[datetime, datetime]:
    start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    return start, start + timedelta(days=1)


def _count_grouped_status(
    session: Any,
    model: Any,
    owner_username: str,
    statuses: list[str],
    start_at: datetime,
    end_at: datetime,
    *timestamp_columns: Any,
) -> dict[str, int]:
    if not timestamp_columns:
        timestamp_columns = (model.created_at,)
    time_conditions = [
        and_(column >= start_at, column < end_at)
        for column in timestamp_columns
    ]
    rows = session.execute(
        select(model.status, func.count())
        .where(
            model.owner_username == owner_username,
            or_(*time_conditions),
        )
        .group_by(model.status)
    ).all()
    counts = {status: 0 for status in statuses}
    for status, count in rows:
        if status in counts:
            counts[str(status)] = int(count or 0)
    return counts


def _count_grouped_product_status(
    session: Any,
    owner_username: str,
    statuses: list[str],
    start_at: datetime,
    end_at: datetime,
) -> dict[str, int]:
    rows = session.execute(
        select(ProductModel.review_status, func.count())
        .where(
            ProductModel.owner_username == owner_username,
            or_(
                and_(ProductModel.created_at >= start_at, ProductModel.created_at < end_at),
                and_(ProductModel.updated_at >= start_at, ProductModel.updated_at < end_at),
            ),
        )
        .group_by(ProductModel.review_status)
    ).all()
    counts = {status: 0 for status in statuses}
    for status, count in rows:
        if status in counts:
            counts[str(status)] = int(count or 0)
    return counts


def dashboard_summary(
    owner_username: str,
    *,
    include_stores: bool = True,
    include_crawler: bool = True,
    include_products: bool = True,
    include_sync_tasks: bool = True,
) -> dict[str, Any]:
    start_at, end_at = today_range()
    task_statuses = ["queued", "running", "success", "failed"]
    empty_task_counts = {status: 0 for status in task_statuses}
    empty_product_counts = {"pending": 0, "approved": 0, "error": 0}
    with session_scope() as session:
        total_stores = enabled_stores = error_stores = 0
        if include_stores:
            total_stores = int(session.scalar(
                select(func.count()).select_from(StoreModel).where(StoreModel.owner_username == owner_username)
            ) or 0)
            enabled_stores = int(session.scalar(
                select(func.count()).select_from(StoreModel).where(
                    StoreModel.owner_username == owner_username,
                    StoreModel.enabled.is_(True),
                )
            ) or 0)
            error_stores = int(session.scalar(
                select(func.count()).select_from(StoreModel).where(
                    StoreModel.owner_username == owner_username,
                    StoreModel.last_error.is_not(None),
                    StoreModel.last_error != "",
                )
            ) or 0)
        crawl_tasks = empty_task_counts.copy()
        if include_crawler:
            reconcile_interrupted_running_tasks(session, CrawlTaskModel, owner_username=owner_username)
            crawl_tasks = _count_grouped_status(
                session,
                CrawlTaskModel,
                owner_username,
                task_statuses,
                start_at,
                end_at,
                CrawlTaskModel.created_at,
                CrawlTaskModel.started_at,
                CrawlTaskModel.finished_at,
            )
        products = empty_product_counts.copy()
        listing_tasks = empty_task_counts.copy()
        if include_products:
            products = _count_grouped_product_status(
                session,
                owner_username,
                ["pending", "approved", "error"],
                start_at,
                end_at,
            )
            reconcile_interrupted_running_tasks(session, ListingTaskModel, owner_username=owner_username)
            listing_tasks = _count_grouped_status(
                session,
                ListingTaskModel,
                owner_username,
                task_statuses,
                start_at,
                end_at,
                ListingTaskModel.created_at,
                ListingTaskModel.started_at,
                ListingTaskModel.finished_at,
                ListingTaskModel.updated_at,
            )
        sync_tasks = empty_task_counts.copy()
        if include_sync_tasks:
            reconcile_interrupted_running_tasks(session, SyncTaskModel, owner_username=owner_username)
            sync_tasks = _count_grouped_status(
                session,
                SyncTaskModel,
                owner_username,
                task_statuses,
                start_at,
                end_at,
                SyncTaskModel.created_at,
                SyncTaskModel.started_at,
                SyncTaskModel.finished_at,
                SyncTaskModel.updated_at,
            )
    return {
        "dateRange": {
            "from": start_at.isoformat(sep=" "),
            "to": end_at.isoformat(sep=" "),
        },
        "stores": {
            "total": total_stores,
            "enabled": enabled_stores,
            "error": error_stores,
        },
        "products": {
            "pending": products["pending"],
            "approved": products["approved"],
            "error": products["error"],
        },
        "crawlTasks": {
            "queued": crawl_tasks["queued"],
            "running": crawl_tasks["running"],
            "success": crawl_tasks["success"],
            "failed": crawl_tasks["failed"],
        },
        "listingTasks": {
            "queued": listing_tasks["queued"],
            "running": listing_tasks["running"],
            "success": listing_tasks["success"],
            "failed": listing_tasks["failed"],
        },
        "syncTasks": {
            "queued": sync_tasks["queued"],
            "running": sync_tasks["running"],
            "success": sync_tasks["success"],
            "failed": sync_tasks["failed"],
        },
    }


def list_tasks(
    owner_username: str,
    *,
    page: int | None = None,
    page_size: int | None = None,
    target: str | None = None,
    status: str | None = None,
    source_type: str | None = None,
    mode: str | None = None,
    created_at_from: str | None = None,
    created_at_to: str | None = None,
) -> list[dict[str, Any]] | dict[str, Any]:
    with session_scope() as session:
        finalize_stale_cancel_requested_tasks(session, CrawlTaskModel, action_label="采集", owner_username=owner_username)
        reconcile_interrupted_running_tasks(session, CrawlTaskModel, owner_username=owner_username)
        query = select(CrawlTaskModel).where(CrawlTaskModel.owner_username == owner_username)
        created_at_from_value = parse_datetime_filter(created_at_from)
        created_at_to_value = parse_datetime_filter(created_at_to)
        if target:
            query = query.where(CrawlTaskModel.target.like(f"%{target}%"))
        if status:
            query = query.where(crawl_task_status_filter(status))
        if source_type:
            query = query.where(CrawlTaskModel.source_type == source_type)
        if mode:
            query = query.where(CrawlTaskModel.mode == mode)
        if created_at_from_value is not None:
            query = query.where(CrawlTaskModel.created_at >= created_at_from_value)
        if created_at_to_value is not None:
            query = query.where(CrawlTaskModel.created_at <= created_at_to_value)
        return paginate_query(
            session,
            query,
            order_by=task_status_order_by(
                CrawlTaskModel,
                failure_condition=CrawlTaskModel.failed_count > 0,
            ),
            page=page,
            page_size=page_size,
            response_key="tasks",
            serializer=task_to_public,
        )


def crawl_task_status_filter(status: str) -> Any:
    normalized = normalize_text(status)
    cancel_requested = func.coalesce(
        CrawlTaskModel.error_detail,
        "",
    ).like(f"{TASK_CANCEL_REQUESTED_MARKER}%")
    if normalized in {"queued", "running"}:
        return and_(
            CrawlTaskModel.status == normalized,
            ~cancel_requested,
        )
    if normalized == "cancelling":
        return and_(
            CrawlTaskModel.status.in_(("queued", "running")),
            cancel_requested,
        )
    if normalized == "skipped":
        return and_(
            CrawlTaskModel.status.notin_(("queued", "running", "cancelled")),
            CrawlTaskModel.saved_count == 0,
            CrawlTaskModel.skipped_count > 0,
            CrawlTaskModel.failed_count == 0,
        )
    if normalized == "partial_saved":
        return and_(
            CrawlTaskModel.status.notin_(("queued", "running", "cancelled")),
            CrawlTaskModel.saved_count > 0,
            CrawlTaskModel.skipped_count > 0,
            CrawlTaskModel.failed_count == 0,
        )
    if normalized == "failed":
        return and_(
            CrawlTaskModel.status.notin_(("queued", "running", "cancelled")),
            CrawlTaskModel.failed_count > 0,
            CrawlTaskModel.success_count == 0,
        )
    if normalized == "partial":
        return and_(
            CrawlTaskModel.status.notin_(("queued", "running", "cancelled")),
            CrawlTaskModel.success_count > 0,
            CrawlTaskModel.failed_count > 0,
        )
    if normalized == "success":
        return and_(
            CrawlTaskModel.status.notin_(("queued", "running", "cancelled")),
            or_(
                CrawlTaskModel.status == "success",
                CrawlTaskModel.success_count > 0,
            ),
            CrawlTaskModel.failed_count == 0,
            CrawlTaskModel.skipped_count == 0,
        )
    return CrawlTaskModel.status == normalized


def normalize_task_ids(task_ids: list[str]) -> list[str]:
    normalized: list[str] = []
    for value in task_ids or []:
        task_id = str(value or "").strip()
        if task_id and task_id not in normalized:
            normalized.append(task_id)
    if not normalized:
        raise RuntimeError("请选择要删除的任务。")
    return normalized


def delete_tasks(owner_username: str, task_ids: list[str]) -> dict[str, Any]:
    normalized_ids = normalize_task_ids(task_ids)
    with session_scope() as session:
        rows = session.scalars(
            select(CrawlTaskModel).where(
                CrawlTaskModel.owner_username == owner_username,
                CrawlTaskModel.id.in_(normalized_ids),
            )
        ).all()
        found_ids = {row.id for row in rows}
        products = session.scalars(select(ProductModel).where(ProductModel.task_id.in_(list(found_ids)))).all() if found_ids else []
        for product in products:
            product.task_id = None
        logs = session.scalars(
            select(CrawlLogModel).where(
                CrawlLogModel.owner_username == owner_username,
                CrawlLogModel.task_id.in_(list(found_ids)),
            )
        ).all() if found_ids else []
        for log in logs:
            log.task_id = None
        for row in rows:
            session.delete(row)
        deleted_ids = [row.id for row in rows]
        return {
            "deletedIds": deleted_ids,
            "failedIds": [task_id for task_id in normalized_ids if task_id not in found_ids],
            "deletedCount": len(deleted_ids),
        }


def delete_pending_products_for_crawl_task(
    owner_username: str,
    task_id: str,
) -> dict[str, Any]:
    normalized_task_id = normalize_text(task_id)
    if not normalized_task_id:
        raise RuntimeError("采集任务不存在。")
    deleted_ids: list[int] = []
    with session_scope() as session:
        task = session.scalar(
            select(CrawlTaskModel).where(
                CrawlTaskModel.id == normalized_task_id,
                CrawlTaskModel.owner_username == owner_username,
            )
        )
        if task is None:
            raise RuntimeError("采集任务不存在或无权操作。")
        if task.status in {"queued", "running"}:
            raise RuntimeError("待执行或执行中的采集任务不能删除商品。")
        if task.source_type == "product_replace":
            raise RuntimeError("商品替换任务不能通过采集记录删除商品。")

        rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.task_id == normalized_task_id,
                ProductModel.review_status == "pending",
            )
        ).all()
        deleted_ids = [int(row.id) for row in rows]
        for row in rows:
            session.delete(row)
        session.flush()

    cleanup_product_image_ids(deleted_ids)
    return {
        "taskId": normalized_task_id,
        "deletedCount": len(deleted_ids),
        "deletedProductIds": deleted_ids,
    }


def parse_subscription_userinfo(value: Any) -> dict[str, int]:
    parsed: dict[str, int] = {}
    key_map = {
        "upload": "uploadBytes",
        "download": "downloadBytes",
        "total": "totalBytes",
    }
    for part in str(value or "").split(";"):
        if "=" not in part:
            continue
        key, raw_value = part.strip().split("=", 1)
        public_key = key_map.get(key.strip().lower())
        if not public_key:
            continue
        try:
            parsed[public_key] = max(0, int(raw_value.strip()))
        except ValueError:
            continue
    return parsed


def next_proxy_traffic_reset_at(reset_day: int, *, now: datetime | None = None) -> datetime:
    reference = now or datetime.now()
    normalized_day = min(28, max(1, int(reset_day or 1)))
    candidate = reference.replace(day=normalized_day, hour=0, minute=0, second=0, microsecond=0)
    if candidate > reference:
        return candidate
    if reference.month == 12:
        return candidate.replace(year=reference.year + 1, month=1)
    return candidate.replace(month=reference.month + 1)


def proxy_usage_public_payload(
    *,
    upload_bytes: int,
    download_bytes: int,
    total_bytes: int,
    source: str,
    stale: bool,
    checked_at: datetime,
    now: datetime | None = None,
) -> dict[str, Any]:
    used_bytes = max(0, int(upload_bytes or 0)) + max(0, int(download_bytes or 0))
    normalized_total = max(0, int(total_bytes or 0))
    remaining_bytes = max(0, normalized_total - used_bytes)
    usage_percent = round((used_bytes / normalized_total) * 100, 2) if normalized_total else 0.0
    reference = now or datetime.now()
    reset_at = next_proxy_traffic_reset_at(settings.proxy_traffic_reset_day, now=reference)
    return {
        "uploadBytes": max(0, int(upload_bytes or 0)),
        "downloadBytes": max(0, int(download_bytes or 0)),
        "usedBytes": used_bytes,
        "totalBytes": normalized_total,
        "remainingBytes": remaining_bytes,
        "usagePercent": min(100.0, usage_percent),
        "resetDay": settings.proxy_traffic_reset_day,
        "resetAt": datetime_to_public(reset_at),
        "resetRemainingSeconds": max(0, int((reset_at - reference).total_seconds())),
        "checkedAt": datetime_to_public(checked_at),
        "source": source,
        "stale": stale,
    }


def fetch_proxy_subscription_usage() -> dict[str, Any]:
    subscription_url = normalize_text(settings.proxy_subscription_url)
    if not subscription_url:
        raise RuntimeError("代理订阅地址未配置。")
    response = requests.get(
        subscription_url,
        headers={"User-Agent": "clash.meta"},
        timeout=settings.crawler_timeout_seconds,
        stream=True,
    )
    try:
        response.raise_for_status()
        parsed = parse_subscription_userinfo(response.headers.get("subscription-userinfo"))
        if not all(key in parsed for key in ("uploadBytes", "downloadBytes", "totalBytes")):
            raise RuntimeError("订阅响应未提供完整流量信息。")
        return proxy_usage_public_payload(
            upload_bytes=parsed["uploadBytes"],
            download_bytes=parsed["downloadBytes"],
            total_bytes=parsed["totalBytes"],
            source="subscription",
            stale=False,
            checked_at=datetime.now(),
        )
    finally:
        response.close()


def proxy_usage_from_mihomo_config() -> dict[str, Any]:
    config_path = Path(settings.mihomo_config_path)
    if not config_path.exists():
        raise RuntimeError("Mihomo 配置文件不存在。")
    text = config_path.read_text(encoding="utf-8", errors="ignore")
    remaining_match = re.search(r"剩余流量[：:]\s*([0-9.]+)\s*(GB|MB|TB)", text, re.IGNORECASE)
    if not remaining_match:
        raise RuntimeError("Mihomo 配置中没有剩余流量信息。")
    amount = float(remaining_match.group(1))
    unit = remaining_match.group(2).upper()
    unit_bytes = {"MB": 1024 ** 2, "GB": 1024 ** 3, "TB": 1024 ** 4}[unit]
    remaining_bytes = int(amount * unit_bytes)
    total_bytes = int(settings.proxy_traffic_total_gib * 1024 ** 3)
    used_bytes = max(0, total_bytes - remaining_bytes)
    checked_at = datetime.fromtimestamp(config_path.stat().st_mtime)
    return proxy_usage_public_payload(
        upload_bytes=0,
        download_bytes=used_bytes,
        total_bytes=total_bytes,
        source="mihomo_config",
        stale=True,
        checked_at=checked_at,
    )


def stale_proxy_usage_payload(payload: dict[str, Any], *, now: datetime) -> dict[str, Any]:
    stale_payload = dict(payload)
    reset_at = next_proxy_traffic_reset_at(settings.proxy_traffic_reset_day, now=now)
    stale_payload.update({
        "stale": True,
        "resetDay": settings.proxy_traffic_reset_day,
        "resetAt": datetime_to_public(reset_at),
        "resetRemainingSeconds": max(0, int((reset_at - now).total_seconds())),
    })
    return stale_payload


def get_proxy_resource_usage(*, force: bool = False) -> dict[str, Any]:
    requested_monotonic = time.monotonic()
    with PROXY_USAGE_CACHE_LOCK:
        cached = PROXY_USAGE_CACHE.get("payload")
        cached_monotonic = PROXY_USAGE_CACHE.get("cachedMonotonic")
        requested_generation = int(PROXY_USAGE_CACHE.get("generation") or 0)
        if (
            not force
            and isinstance(cached, dict)
            and isinstance(cached_monotonic, (int, float))
            and requested_monotonic - cached_monotonic < settings.proxy_usage_cache_seconds
        ):
            return dict(cached)

    with PROXY_USAGE_REFRESH_LOCK:
        now = datetime.now()
        current_monotonic = time.monotonic()
        with PROXY_USAGE_CACHE_LOCK:
            cached = PROXY_USAGE_CACHE.get("payload")
            cached_monotonic = PROXY_USAGE_CACHE.get("cachedMonotonic")
            current_generation = int(PROXY_USAGE_CACHE.get("generation") or 0)
            if (
                isinstance(cached, dict)
                and (
                    current_generation > requested_generation
                    or (
                        not force
                        and isinstance(cached_monotonic, (int, float))
                        and current_monotonic - cached_monotonic < settings.proxy_usage_cache_seconds
                    )
                )
            ):
                return dict(cached)

        try:
            payload = fetch_proxy_subscription_usage()
        except Exception:
            try:
                payload = proxy_usage_from_mihomo_config()
            except Exception:
                with PROXY_USAGE_CACHE_LOCK:
                    cached = PROXY_USAGE_CACHE.get("payload")
                if not isinstance(cached, dict):
                    raise
                payload = stale_proxy_usage_payload(cached, now=now)

        with PROXY_USAGE_CACHE_LOCK:
            PROXY_USAGE_CACHE["payload"] = dict(payload)
            PROXY_USAGE_CACHE["cachedAt"] = datetime.now()
            PROXY_USAGE_CACHE["cachedMonotonic"] = time.monotonic()
            PROXY_USAGE_CACHE["generation"] = int(PROXY_USAGE_CACHE.get("generation") or 0) + 1
        return payload


def default_time_settings_value(*, now: datetime | None = None) -> dict[str, Any]:
    reference = now or datetime.now()
    next_cleanup_at = next_weekly_run_at(
        SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_WEEKDAY,
        SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_TIME,
        now=reference,
    )
    next_product_sync_at = next_weekly_run_at(
        STORE_PRODUCT_SYNC_DEFAULT_WEEKDAY,
        STORE_PRODUCT_SYNC_DEFAULT_TIME,
        now=reference,
    )
    next_unlisted_cleanup_at = next_monthly_run_at(
        STORE_UNLISTED_PRODUCT_CLEANUP_MONTH_DAY,
        STORE_UNLISTED_PRODUCT_CLEANUP_TIME,
        now=reference,
    )
    next_deleted_image_cleanup_at = next_weekly_run_at(
        DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY,
        DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME,
        now=reference,
    )
    return {
        "cleanupWeekday": SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_WEEKDAY,
        "cleanupTime": SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_TIME,
        "cleanupEnabled": True,
        "retentionDays": SCHEDULED_CRAWL_TASK_CLEANUP_RETENTION_DAYS,
        "nextCleanupAt": datetime_to_public(next_cleanup_at),
        "lastCleanupAt": None,
        "lastCleanupDeletedCount": 0,
        "productSyncEnabled": True,
        "productSyncWeekday": STORE_PRODUCT_SYNC_DEFAULT_WEEKDAY,
        "productSyncTime": STORE_PRODUCT_SYNC_DEFAULT_TIME,
        "productSyncNextAt": datetime_to_public(next_product_sync_at),
        "productSyncLastAt": None,
        "productSyncLastTaskCount": 0,
        "unlistedCleanupEnabled": True,
        "unlistedCleanupMonthDay": STORE_UNLISTED_PRODUCT_CLEANUP_MONTH_DAY,
        "unlistedCleanupTime": STORE_UNLISTED_PRODUCT_CLEANUP_TIME,
        "unlistedNextCleanupAt": datetime_to_public(next_unlisted_cleanup_at),
        "unlistedLastCleanupAt": None,
        "unlistedLastDeletedCount": 0,
        "unlistedLastTaskCount": 0,
        "deletedImageCleanupEnabled": True,
        "deletedImageCleanupWeekday": DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY,
        "deletedImageCleanupTime": DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME,
        "deletedImageCleanupNextAt": datetime_to_public(next_deleted_image_cleanup_at),
        "deletedImageCleanupLastAt": None,
        "deletedImageCleanupLastProductCount": 0,
        "deletedImageCleanupLastTaskCount": 0,
    }


def load_time_settings_payload(row: SystemSettingModel | None, *, now: datetime | None = None) -> dict[str, Any]:
    reference = now or datetime.now()
    try:
        raw_payload = json.loads(row.value_json or "{}") if row is not None else {}
    except ValueError:
        raw_payload = {}
    payload = raw_payload if isinstance(raw_payload, dict) else {}
    default_payload = default_time_settings_value(now=reference)

    try:
        cleanup_weekday = normalize_cleanup_weekday(payload.get("cleanupWeekday", default_payload["cleanupWeekday"]))
    except RuntimeError:
        cleanup_weekday = default_payload["cleanupWeekday"]
    try:
        cleanup_time = normalize_schedule_time(payload.get("cleanupTime", default_payload["cleanupTime"]))
    except RuntimeError:
        cleanup_time = default_payload["cleanupTime"]
    cleanup_enabled = payload.get("cleanupEnabled", default_payload["cleanupEnabled"])
    if not isinstance(cleanup_enabled, bool):
        cleanup_enabled = default_payload["cleanupEnabled"]
    product_sync_enabled = payload.get("productSyncEnabled", default_payload["productSyncEnabled"])
    if not isinstance(product_sync_enabled, bool):
        product_sync_enabled = default_payload["productSyncEnabled"]
    try:
        product_sync_weekday = normalize_cleanup_weekday(
            payload.get("productSyncWeekday", default_payload["productSyncWeekday"])
        )
    except RuntimeError:
        product_sync_weekday = default_payload["productSyncWeekday"]
    try:
        product_sync_time = normalize_schedule_time(
            payload.get("productSyncTime", default_payload["productSyncTime"])
        )
    except RuntimeError:
        product_sync_time = default_payload["productSyncTime"]

    next_cleanup_at = parse_public_datetime(payload.get("nextCleanupAt"))
    if next_cleanup_at is None:
        next_cleanup_at = next_weekly_run_at(cleanup_weekday, cleanup_time, now=reference)
    last_cleanup_deleted_count = payload.get("lastCleanupDeletedCount", 0)
    try:
        last_cleanup_deleted_count = max(0, int(last_cleanup_deleted_count or 0))
    except (TypeError, ValueError):
        last_cleanup_deleted_count = 0
    product_sync_next_at = parse_public_datetime(payload.get("productSyncNextAt"))
    if product_sync_next_at is None:
        product_sync_next_at = next_weekly_run_at(product_sync_weekday, product_sync_time, now=reference)
    product_sync_last_task_count = payload.get("productSyncLastTaskCount", 0)
    try:
        product_sync_last_task_count = max(0, int(product_sync_last_task_count or 0))
    except (TypeError, ValueError):
        product_sync_last_task_count = 0
    unlisted_cleanup_enabled = payload.get(
        "unlistedCleanupEnabled",
        default_payload["unlistedCleanupEnabled"],
    )
    if not isinstance(unlisted_cleanup_enabled, bool):
        unlisted_cleanup_enabled = default_payload["unlistedCleanupEnabled"]
    unlisted_month_day = STORE_UNLISTED_PRODUCT_CLEANUP_MONTH_DAY
    try:
        unlisted_cleanup_time = normalize_schedule_time(payload.get("unlistedCleanupTime", STORE_UNLISTED_PRODUCT_CLEANUP_TIME))
    except RuntimeError:
        unlisted_cleanup_time = STORE_UNLISTED_PRODUCT_CLEANUP_TIME
    unlisted_next_cleanup_at = parse_public_datetime(payload.get("unlistedNextCleanupAt"))
    if unlisted_next_cleanup_at is None:
        unlisted_next_cleanup_at = next_monthly_run_at(unlisted_month_day, unlisted_cleanup_time, now=reference)
    unlisted_last_deleted_count = payload.get("unlistedLastDeletedCount", 0)
    try:
        unlisted_last_deleted_count = max(0, int(unlisted_last_deleted_count or 0))
    except (TypeError, ValueError):
        unlisted_last_deleted_count = 0
    unlisted_last_task_count = payload.get("unlistedLastTaskCount", 0)
    try:
        unlisted_last_task_count = max(0, int(unlisted_last_task_count or 0))
    except (TypeError, ValueError):
        unlisted_last_task_count = 0
    deleted_image_cleanup_enabled = payload.get("deletedImageCleanupEnabled", True)
    if not isinstance(deleted_image_cleanup_enabled, bool):
        deleted_image_cleanup_enabled = True
    try:
        deleted_image_cleanup_weekday = normalize_cleanup_weekday(
            payload.get("deletedImageCleanupWeekday", DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY)
        )
    except RuntimeError:
        deleted_image_cleanup_weekday = DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY
    try:
        deleted_image_cleanup_time = normalize_schedule_time(
            payload.get("deletedImageCleanupTime", DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME)
        )
    except RuntimeError:
        deleted_image_cleanup_time = DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME
    deleted_image_cleanup_next_at = parse_public_datetime(payload.get("deletedImageCleanupNextAt"))
    if deleted_image_cleanup_next_at is None:
        deleted_image_cleanup_next_at = next_weekly_run_at(
            deleted_image_cleanup_weekday,
            deleted_image_cleanup_time,
            now=reference,
        )

    return {
        "cleanupWeekday": cleanup_weekday,
        "cleanupTime": cleanup_time,
        "cleanupEnabled": cleanup_enabled,
        "retentionDays": SCHEDULED_CRAWL_TASK_CLEANUP_RETENTION_DAYS,
        "nextCleanupAt": datetime_to_public(next_cleanup_at),
        "lastCleanupAt": datetime_to_public(parse_public_datetime(payload.get("lastCleanupAt"))),
        "lastCleanupDeletedCount": last_cleanup_deleted_count,
        "productSyncEnabled": product_sync_enabled,
        "productSyncWeekday": product_sync_weekday,
        "productSyncTime": product_sync_time,
        "productSyncNextAt": datetime_to_public(product_sync_next_at),
        "productSyncLastAt": datetime_to_public(parse_public_datetime(payload.get("productSyncLastAt"))),
        "productSyncLastTaskCount": product_sync_last_task_count,
        "unlistedCleanupEnabled": unlisted_cleanup_enabled,
        "unlistedCleanupMonthDay": unlisted_month_day,
        "unlistedCleanupTime": unlisted_cleanup_time,
        "unlistedNextCleanupAt": datetime_to_public(unlisted_next_cleanup_at),
        "unlistedLastCleanupAt": datetime_to_public(parse_public_datetime(payload.get("unlistedLastCleanupAt"))),
        "unlistedLastDeletedCount": unlisted_last_deleted_count,
        "unlistedLastTaskCount": unlisted_last_task_count,
        "deletedImageCleanupEnabled": deleted_image_cleanup_enabled,
        "deletedImageCleanupWeekday": deleted_image_cleanup_weekday,
        "deletedImageCleanupTime": deleted_image_cleanup_time,
        "deletedImageCleanupNextAt": datetime_to_public(deleted_image_cleanup_next_at),
        "deletedImageCleanupLastAt": datetime_to_public(parse_public_datetime(payload.get("deletedImageCleanupLastAt"))),
        "deletedImageCleanupLastProductCount": max(0, int(payload.get("deletedImageCleanupLastProductCount") or 0)),
        "deletedImageCleanupLastTaskCount": max(0, int(payload.get("deletedImageCleanupLastTaskCount") or 0)),
    }


def time_settings_to_public(
    row: SystemSettingModel | None,
    payload: dict[str, Any],
    *,
    include_queue_health: bool = True,
    deleted_image_cleanup_pending_count: int = 0,
) -> dict[str, Any]:
    result = {
        **payload,
        "deletedImageCleanupPendingCount": max(0, int(deleted_image_cleanup_pending_count)),
        "serverNow": datetime_to_public(datetime.now()),
        "updatedAt": datetime_to_public(row.updated_at if row is not None else None),
    }
    if include_queue_health:
        result["queueHealth"] = task_queue_health()
    return result


def user_time_settings_key(owner_username: str) -> str:
    return f"{USER_TIME_SETTINGS_PREFIX}{normalize_text(owner_username)}"


def time_settings_row_key(owner_username: str | None) -> str:
    return (
        user_time_settings_key(owner_username)
        if normalize_text(owner_username)
        else SCHEDULED_CRAWL_TASK_CLEANUP_SETTING_KEY
    )


def initial_user_time_settings_payload(
    session: Any,
    owner_username: str,
    *,
    now: datetime,
) -> dict[str, Any]:
    payload = default_time_settings_value(now=now)
    deleted_row = session.get(
        SystemSettingModel,
        deleted_product_image_cleanup_user_setting_key(owner_username),
    )
    if deleted_row is not None:
        payload.update(
            load_deleted_product_image_cleanup_user_payload(
                deleted_row,
                payload,
                now=now,
            )
        )
    return payload


def migrate_user_time_settings_to_owner_scope(
    session: Any,
    *,
    now: datetime | None = None,
) -> int:
    if not hasattr(session, "get") or not hasattr(session, "scalars"):
        return 0
    bind = session.get_bind()
    user_columns = {
        column["name"]
        for column in inspect(bind).get_columns(UserAccountModel.__tablename__)
    }
    if "username" not in user_columns:
        return 0
    if session.get(
        SystemSettingModel,
        USER_TIME_SETTINGS_OWNER_SCOPE_MIGRATION_KEY,
    ) is not None:
        return 0
    reference = now or datetime.now()
    migrated_count = 0
    owner_usernames = session.scalars(
        select(UserAccountModel.username).order_by(UserAccountModel.username.asc())
    ).all()
    for owner_username_value in owner_usernames:
        owner_username = str(owner_username_value)
        time_row = session.get(
            SystemSettingModel,
            user_time_settings_key(owner_username),
        )
        existing = (
            load_time_settings_payload(time_row, now=reference)
            if time_row is not None
            else default_time_settings_value(now=reference)
        )
        migrated = default_time_settings_value(now=reference)
        migrated.update(
            {
                "cleanupWeekday": existing["cleanupWeekday"],
                "cleanupTime": existing["cleanupTime"],
                "cleanupEnabled": existing["cleanupEnabled"],
                "nextCleanupAt": datetime_to_public(
                    next_weekly_run_at(
                        existing["cleanupWeekday"],
                        existing["cleanupTime"],
                        now=reference,
                    )
                ),
                "productSyncEnabled": existing["productSyncEnabled"],
                "productSyncWeekday": existing["productSyncWeekday"],
                "productSyncTime": existing["productSyncTime"],
                "productSyncNextAt": datetime_to_public(
                    next_weekly_run_at(
                        existing["productSyncWeekday"],
                        existing["productSyncTime"],
                        now=reference,
                    )
                ),
                "unlistedCleanupEnabled": existing["unlistedCleanupEnabled"],
                "unlistedNextCleanupAt": datetime_to_public(
                    next_monthly_run_at(
                        existing["unlistedCleanupMonthDay"],
                        existing["unlistedCleanupTime"],
                        now=reference,
                    )
                ),
                "deletedImageCleanupEnabled": existing["deletedImageCleanupEnabled"],
                "deletedImageCleanupWeekday": existing["deletedImageCleanupWeekday"],
                "deletedImageCleanupTime": existing["deletedImageCleanupTime"],
                "deletedImageCleanupNextAt": datetime_to_public(
                    next_weekly_run_at(
                        existing["deletedImageCleanupWeekday"],
                        existing["deletedImageCleanupTime"],
                        now=reference,
                    )
                ),
            }
        )
        upsert_time_settings_row(session, migrated, owner_username)

        deleted_key = deleted_product_image_cleanup_user_setting_key(owner_username)
        deleted_row = session.get(SystemSettingModel, deleted_key)
        if deleted_row is None:
            deleted_row = SystemSettingModel(key=deleted_key)
            session.add(deleted_row)
        deleted_row.value_json = json.dumps(
            {
                "deletedImageCleanupEnabled": migrated["deletedImageCleanupEnabled"],
                "deletedImageCleanupWeekday": migrated["deletedImageCleanupWeekday"],
                "deletedImageCleanupTime": migrated["deletedImageCleanupTime"],
                "deletedImageCleanupNextAt": migrated["deletedImageCleanupNextAt"],
                "deletedImageCleanupLastAt": None,
                "deletedImageCleanupLastProductCount": 0,
                "deletedImageCleanupLastTaskCount": 0,
            },
            ensure_ascii=False,
        )
        migrated_count += 1

    session.add(
        SystemSettingModel(
            key=USER_TIME_SETTINGS_OWNER_SCOPE_MIGRATION_KEY,
            value_json=json.dumps(
                {
                    "migratedAt": datetime_to_public(reference),
                    "migratedUsers": migrated_count,
                },
                ensure_ascii=False,
            ),
        )
    )
    return migrated_count


def upsert_time_settings_row(
    session: Any,
    payload: dict[str, Any],
    owner_username: str | None = None,
) -> SystemSettingModel:
    setting_key = time_settings_row_key(owner_username)
    row = session.get(SystemSettingModel, setting_key)
    if row is None:
        row = SystemSettingModel(key=setting_key)
        session.add(row)
    row.value_json = json.dumps(payload, ensure_ascii=False)
    return row


def get_time_settings(
    owner_username: str | None = None,
    *,
    include_queue_health: bool = True,
) -> dict[str, Any]:
    now = datetime.now()
    with session_scope() as session:
        row = session.get(
            SystemSettingModel,
            time_settings_row_key(owner_username),
        )
        payload = (
            initial_user_time_settings_payload(
                session,
                owner_username,
                now=now,
            )
            if owner_username and row is None
            else load_time_settings_payload(row, now=now)
        )
        if row is None or (row.value_json or "") != json.dumps(payload, ensure_ascii=False):
            row = upsert_time_settings_row(session, payload, owner_username)
            session.flush()
        pending_query = select(func.count()).select_from(
            DeletedProductImageCleanupModel
        )
        if owner_username:
            pending_query = pending_query.where(
                DeletedProductImageCleanupModel.owner_username == owner_username
            )
        pending_count = session.scalar(pending_query) or 0
        return time_settings_to_public(
            row,
            payload,
            include_queue_health=include_queue_health,
            deleted_image_cleanup_pending_count=pending_count,
        )


def save_time_settings(
    owner_username: str | Any,
    payload: Any | None = None,
    *,
    include_queue_health: bool = True,
) -> dict[str, Any]:
    if payload is None:
        payload = owner_username
        owner_username = None
    cleanup_weekday = normalize_cleanup_weekday(getattr(payload, "cleanupWeekday", SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_WEEKDAY))
    cleanup_time = normalize_schedule_time(getattr(payload, "cleanupTime", SCHEDULED_CRAWL_TASK_CLEANUP_DEFAULT_TIME))
    cleanup_enabled = bool(getattr(payload, "cleanupEnabled", True))
    product_sync_enabled = bool(getattr(payload, "productSyncEnabled", True))
    product_sync_weekday = normalize_cleanup_weekday(
        getattr(payload, "productSyncWeekday", STORE_PRODUCT_SYNC_DEFAULT_WEEKDAY)
    )
    product_sync_time = normalize_schedule_time(
        getattr(payload, "productSyncTime", STORE_PRODUCT_SYNC_DEFAULT_TIME)
    )
    unlisted_cleanup_enabled = bool(getattr(payload, "unlistedCleanupEnabled", True))
    deleted_image_cleanup_enabled = bool(getattr(payload, "deletedImageCleanupEnabled", True))
    deleted_image_cleanup_weekday = normalize_cleanup_weekday(
        getattr(payload, "deletedImageCleanupWeekday", DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY)
    )
    deleted_image_cleanup_time = normalize_schedule_time(
        getattr(payload, "deletedImageCleanupTime", DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME)
    )
    now = datetime.now()
    with session_scope() as session:
        row = session.get(
            SystemSettingModel,
            time_settings_row_key(owner_username),
        )
        existing = (
            initial_user_time_settings_payload(
                session,
                str(owner_username),
                now=now,
            )
            if owner_username and row is None
            else load_time_settings_payload(row, now=now)
        )
        updated_payload = {
            **existing,
            "cleanupWeekday": cleanup_weekday,
            "cleanupTime": cleanup_time,
            "cleanupEnabled": cleanup_enabled,
            "retentionDays": SCHEDULED_CRAWL_TASK_CLEANUP_RETENTION_DAYS,
            "nextCleanupAt": datetime_to_public(next_weekly_run_at(cleanup_weekday, cleanup_time, now=now)),
            "productSyncEnabled": product_sync_enabled,
            "productSyncWeekday": product_sync_weekday,
            "productSyncTime": product_sync_time,
            "productSyncNextAt": datetime_to_public(
                next_weekly_run_at(product_sync_weekday, product_sync_time, now=now)
            ),
            "unlistedCleanupEnabled": unlisted_cleanup_enabled,
            "deletedImageCleanupEnabled": deleted_image_cleanup_enabled,
            "deletedImageCleanupWeekday": deleted_image_cleanup_weekday,
            "deletedImageCleanupTime": deleted_image_cleanup_time,
            "deletedImageCleanupNextAt": datetime_to_public(
                next_weekly_run_at(deleted_image_cleanup_weekday, deleted_image_cleanup_time, now=now)
            ),
        }
        row = upsert_time_settings_row(
            session,
            updated_payload,
            str(owner_username) if owner_username else None,
        )
        if owner_username:
            deleted_setting_key = (
                deleted_product_image_cleanup_user_setting_key(
                    str(owner_username)
                )
            )
            deleted_row = session.get(
                SystemSettingModel,
                deleted_setting_key,
            )
            if deleted_row is None:
                deleted_row = SystemSettingModel(key=deleted_setting_key)
                session.add(deleted_row)
            deleted_row.value_json = json.dumps(
                {
                    "deletedImageCleanupEnabled": (
                        updated_payload["deletedImageCleanupEnabled"]
                    ),
                    "deletedImageCleanupWeekday": (
                        updated_payload["deletedImageCleanupWeekday"]
                    ),
                    "deletedImageCleanupTime": (
                        updated_payload["deletedImageCleanupTime"]
                    ),
                    "deletedImageCleanupNextAt": (
                        updated_payload["deletedImageCleanupNextAt"]
                    ),
                    "deletedImageCleanupLastAt": (
                        updated_payload.get(
                            "deletedImageCleanupLastAt"
                        )
                    ),
                    "deletedImageCleanupLastProductCount": (
                        updated_payload.get(
                            "deletedImageCleanupLastProductCount",
                            0,
                        )
                    ),
                    "deletedImageCleanupLastTaskCount": (
                        updated_payload.get(
                            "deletedImageCleanupLastTaskCount",
                            0,
                        )
                    ),
                },
                ensure_ascii=False,
            )
        session.flush()
        pending_query = select(func.count()).select_from(
            DeletedProductImageCleanupModel
        )
        if owner_username:
            pending_query = pending_query.where(
                DeletedProductImageCleanupModel.owner_username
                == str(owner_username)
            )
        pending_count = session.scalar(pending_query) or 0
        return time_settings_to_public(
            row,
            updated_payload,
            include_queue_health=include_queue_health,
            deleted_image_cleanup_pending_count=pending_count,
        )


def cleanup_completed_scheduled_crawl_tasks(
    *,
    force: bool = False,
    include_recent_terminal: bool = False,
    owner_username: str | None = None,
) -> int:
    if not SCHEDULED_CRAWL_TASK_CLEANUP_LOCK.acquire(blocking=False):
        return 0
    try:
        now = datetime.now()
        with session_scope() as session:
            row = session.get(
                SystemSettingModel,
                time_settings_row_key(owner_username),
            )
            payload = (
                initial_user_time_settings_payload(
                    session,
                    owner_username,
                    now=now,
                )
                if owner_username and row is None
                else load_time_settings_payload(row, now=now)
            )
            row = upsert_time_settings_row(
                session,
                payload,
                owner_username,
            )
            next_cleanup_at = parse_public_datetime(payload.get("nextCleanupAt"))
            if not force and (
                not payload["cleanupEnabled"]
                or (next_cleanup_at is not None and next_cleanup_at > now)
            ):
                return 0

            retention_days = max(
                1,
                int(
                    payload.get("retentionDays")
                    or SCHEDULED_CRAWL_TASK_CLEANUP_RETENTION_DAYS
                ),
            )
            cleanup_cutoff = now - timedelta(days=retention_days)
            task_query = select(CrawlTaskModel).where(
                CrawlTaskModel.mode == "scheduled",
                CrawlTaskModel.status.in_(
                    ("success", "partial", "failed", "cancelled")
                ),
            )
            if owner_username:
                task_query = task_query.where(
                    CrawlTaskModel.owner_username == owner_username
                )
            if not include_recent_terminal:
                task_query = task_query.where(
                    func.coalesce(
                        CrawlTaskModel.finished_at,
                        CrawlTaskModel.updated_at,
                        CrawlTaskModel.created_at,
                    )
                    <= cleanup_cutoff,
                )
            rows = session.scalars(task_query).all()
            task_ids = [task.id for task in rows]
            if task_ids:
                remove_crawl_queue_jobs_for_task_ids(set(task_ids))
                session.execute(update(ProductModel).where(ProductModel.task_id.in_(task_ids)).values(task_id=None))
                session.execute(update(CrawlLogModel).where(CrawlLogModel.task_id.in_(task_ids)).values(task_id=None))
                for task in rows:
                    session.delete(task)

            payload["lastCleanupAt"] = datetime_to_public(now)
            payload["lastCleanupDeletedCount"] = len(task_ids)
            payload["nextCleanupAt"] = datetime_to_public(
                next_weekly_run_at(payload["cleanupWeekday"], payload["cleanupTime"], now=now)
            )
            row.value_json = json.dumps(payload, ensure_ascii=False)
            return len(task_ids)
    finally:
        SCHEDULED_CRAWL_TASK_CLEANUP_LOCK.release()


def remove_crawl_queue_jobs_for_task_ids(task_ids: set[str]) -> int:
    if not task_ids or not should_use_redis_task_queue():
        return 0
    try:
        from rq import Queue
        from rq.registry import DeferredJobRegistry, FailedJobRegistry, ScheduledJobRegistry
    except Exception:
        return 0
    try:
        connection = redis_connection()
        removed = 0

        for queue_kind in ("manual-crawl", "scheduled-crawl", "crawl"):
            queue_name = task_queue_name_for_kind(queue_kind)
            queue = Queue(queue_name, connection=connection)

            for job_id in list(queue.job_ids):
                job = fetch_rq_job(connection, job_id)
                if job is None or not task_id_from_rq_job(job, task_ids):
                    continue
                try:
                    job.cancel()
                    job.delete(remove_from_queue=True)
                    removed += 1
                except Exception:
                    continue

            for registry in (
                DeferredJobRegistry(queue_name, connection=connection),
                ScheduledJobRegistry(queue_name, connection=connection),
                FailedJobRegistry(queue_name, connection=connection),
            ):
                for job_id in list(registry.get_job_ids()):
                    job = fetch_rq_job(connection, job_id)
                    if job is None or not task_id_from_rq_job(job, task_ids):
                        continue
                    try:
                        registry.remove(job_id, delete_job=True)
                        removed += 1
                    except Exception:
                        continue
        return removed
    except Exception:
        return 0


def cleanup_completed_scheduled_crawl_tasks_if_due() -> int:
    with session_scope() as session:
        owners = session.scalars(
            select(UserAccountModel.username).where(
                UserAccountModel.enabled.is_(True)
            )
        ).all()
    return sum(
        cleanup_completed_scheduled_crawl_tasks(
            force=False,
            owner_username=str(owner_username),
        )
        for owner_username in owners
    )


def run_completed_scheduled_crawl_tasks_cleanup_now(
    owner_username: str | None = None,
    *,
    include_queue_health: bool = True,
) -> dict[str, Any]:
    cleanup_completed_scheduled_crawl_tasks(
        force=True,
        include_recent_terminal=True,
        owner_username=owner_username,
    )
    return get_time_settings(
        owner_username,
        include_queue_health=include_queue_health,
    )


def create_store_unlisted_product_delete_tasks(
    session: Any,
    now: datetime,
    *,
    owner_username: str | None = None,
) -> tuple[list[tuple[str, str]], int]:
    rows_query = (
        select(
            ProductModel.owner_username,
            ProductModel.store_id,
            ProductModel.id,
            StoreModel.alias_name,
            StoreModel.store_name,
        )
        .join(StoreModel, ProductModel.store_id == StoreModel.id)
        .where(
            StoreModel.enabled.is_(True),
            ProductModel.owner_username == StoreModel.owner_username,
            ProductModel.store_id.is_not(None),
            ProductModel.review_status == "listed",
            ProductModel.rakuten_listing_status == "unlisted",
        )
        .order_by(ProductModel.store_id.asc(), ProductModel.id.asc())
    )
    if owner_username:
        rows_query = rows_query.where(
            ProductModel.owner_username == owner_username
        )
    rows = session.execute(rows_query).all()
    groups: dict[tuple[str, int], dict[str, Any]] = {}
    for owner_username, store_id, product_id, alias_name, store_name in rows:
        if store_id is None:
            continue
        key = (str(owner_username), int(store_id))
        group = groups.setdefault(
            key,
            {
                "storeName": normalize_text(alias_name or store_name) or f"店铺 {store_id}",
                "productIds": [],
            },
        )
        group["productIds"].append(int(product_id))

    task_refs: list[tuple[str, str]] = []
    product_count = 0
    for (owner_username, store_id), group in groups.items():
        product_ids = normalize_product_ids(group["productIds"])
        if not product_ids:
            continue
        product_count += len(product_ids)
        chunks = chunk_product_ids(product_ids)
        task_group_id = uuid.uuid4().hex if len(chunks) > 1 else None
        for index, chunk_ids in enumerate(chunks, start=1):
            task_id = uuid.uuid4().hex
            part_label = "" if len(chunks) == 1 else f" {index}/{len(chunks)}"
            task = SyncTaskModel(
                id=task_id,
                owner_username=owner_username,
                store_id=store_id,
                task_group_id=task_group_id,
                task_group_index=index if task_group_id else None,
                task_group_size=len(chunks) if task_group_id else None,
                store_name=group["storeName"],
                task_name=f"月度删除未上架{part_label} {group['storeName']} {now:%Y-%m-%d %H:%M}",
                task_type="product_delete",
                payload_json=json.dumps({"productIds": chunk_ids, "autoMonthlyUnlistedCleanup": True}, ensure_ascii=False),
                status="queued",
                total_count=len(chunk_ids),
                message="等待执行月度未上架商品删除",
            )
            session.add(task)
            task_refs.append((owner_username, task_id))
    return task_refs, product_count


def cleanup_store_unlisted_products(
    *,
    force: bool = False,
    owner_username: str | None = None,
) -> dict[str, int]:
    if not STORE_UNLISTED_PRODUCT_CLEANUP_LOCK.acquire(blocking=False):
        return {"taskCount": 0, "productCount": 0}
    task_refs: list[tuple[str, str]] = []
    product_count = 0
    try:
        now = datetime.now()
        with session_scope() as session:
            row = session.get(
                SystemSettingModel,
                time_settings_row_key(owner_username),
            )
            payload = (
                initial_user_time_settings_payload(
                    session,
                    owner_username,
                    now=now,
                )
                if owner_username and row is None
                else load_time_settings_payload(row, now=now)
            )
            row = upsert_time_settings_row(
                session,
                payload,
                owner_username,
            )
            next_cleanup_at = parse_public_datetime(payload.get("unlistedNextCleanupAt"))
            if not force and (
                not payload["unlistedCleanupEnabled"]
                or (next_cleanup_at is not None and next_cleanup_at > now)
            ):
                return {"taskCount": 0, "productCount": 0}

            task_refs, product_count = create_store_unlisted_product_delete_tasks(
                session,
                now,
                owner_username=owner_username,
            )
            payload["unlistedLastCleanupAt"] = datetime_to_public(now)
            payload["unlistedLastDeletedCount"] = product_count
            payload["unlistedLastTaskCount"] = len(task_refs)
            payload["unlistedNextCleanupAt"] = datetime_to_public(
                next_monthly_run_at(
                    payload["unlistedCleanupMonthDay"],
                    payload["unlistedCleanupTime"],
                    now=now,
                )
            )
            row.value_json = json.dumps(payload, ensure_ascii=False)
        if task_refs:
            dispatch_next_sync_task()
        return {"taskCount": len(task_refs), "productCount": product_count}
    finally:
        STORE_UNLISTED_PRODUCT_CLEANUP_LOCK.release()


def cleanup_store_unlisted_products_if_due() -> dict[str, int]:
    with session_scope() as session:
        owners = session.scalars(
            select(UserAccountModel.username).where(
                UserAccountModel.enabled.is_(True)
            )
        ).all()
    summaries = [
        cleanup_store_unlisted_products(
            force=False,
            owner_username=str(owner_username),
        )
        for owner_username in owners
    ]
    return {
        "taskCount": sum(item["taskCount"] for item in summaries),
        "productCount": sum(item["productCount"] for item in summaries),
    }


def run_store_unlisted_product_cleanup_now(
    owner_username: str | None = None,
    *,
    include_queue_health: bool = True,
) -> dict[str, Any]:
    summary = cleanup_store_unlisted_products(
        force=True,
        owner_username=owner_username,
    )
    return {
        "settings": get_time_settings(
            owner_username,
            include_queue_health=include_queue_health,
        ),
        "summary": {
            "taskCount": summary["taskCount"],
            "productCount": summary["productCount"],
        },
    }


def cleanup_record_json_list(row: DeletedProductImageCleanupModel, field: str) -> list[Any]:
    try:
        value = json.loads(getattr(row, field) or "[]")
    except ValueError:
        return []
    return value if isinstance(value, list) else []


def deleted_product_image_cleanup_to_public(row: DeletedProductImageCleanupModel) -> dict[str, Any]:
    return {
        "id": row.id,
        "ownerUsername": row.owner_username,
        "storeId": row.store_id,
        "storeName": row.store_name,
        "originalProductId": row.original_product_id,
        "productCode": row.product_code,
        "cabinetImageCount": len(cleanup_record_json_list(row, "cabinet_targets_json")),
        "localImageCount": len(cleanup_record_json_list(row, "local_image_urls_json")),
        "status": row.status,
        "syncTaskId": row.sync_task_id,
        "lastError": row.last_error,
        "deletedAt": datetime_to_public(row.deleted_at),
        "createdAt": datetime_to_public(row.created_at),
        "updatedAt": datetime_to_public(row.updated_at),
    }


def deleted_product_image_cleanup_user_setting_key(owner_username: str) -> str:
    return f"{DELETED_PRODUCT_IMAGE_CLEANUP_USER_SETTING_PREFIX}{normalize_text(owner_username)}"


def load_deleted_product_image_cleanup_user_payload(
    row: SystemSettingModel | None,
    fallback_payload: dict[str, Any],
    *,
    now: datetime,
) -> dict[str, Any]:
    try:
        stored = json.loads(row.value_json or "{}") if row is not None else {}
    except ValueError:
        stored = {}
    if not isinstance(stored, dict):
        stored = {}
    enabled = bool(
        stored.get(
            "deletedImageCleanupEnabled",
            fallback_payload.get("deletedImageCleanupEnabled", True),
        )
    )
    weekday = normalize_cleanup_weekday(
        stored.get(
            "deletedImageCleanupWeekday",
            fallback_payload.get(
                "deletedImageCleanupWeekday",
                DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY,
            ),
        )
    )
    cleanup_time = normalize_schedule_time(
        stored.get(
            "deletedImageCleanupTime",
            fallback_payload.get(
                "deletedImageCleanupTime",
                DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME,
            ),
        )
    )
    next_at = parse_public_datetime(stored.get("deletedImageCleanupNextAt"))
    if next_at is None:
        next_at = next_weekly_run_at(weekday, cleanup_time, now=now)
    return {
        **stored,
        "deletedImageCleanupEnabled": enabled,
        "deletedImageCleanupWeekday": weekday,
        "deletedImageCleanupTime": cleanup_time,
        "deletedImageCleanupNextAt": datetime_to_public(next_at),
        "deletedImageCleanupLastAt": datetime_to_public(
            parse_public_datetime(stored.get("deletedImageCleanupLastAt"))
        ),
        "deletedImageCleanupLastProductCount": max(
            0,
            int(stored.get("deletedImageCleanupLastProductCount") or 0),
        ),
        "deletedImageCleanupLastTaskCount": max(
            0,
            int(stored.get("deletedImageCleanupLastTaskCount") or 0),
        ),
    }


def get_deleted_product_image_cleanup_settings(
    owner_username: str,
) -> dict[str, Any]:
    now = datetime.now()
    base_settings = get_time_settings(
        owner_username,
        include_queue_health=False,
    )
    with session_scope() as session:
        fallback_payload = default_time_settings_value(now=now)
        setting_key = deleted_product_image_cleanup_user_setting_key(owner_username)
        row = session.get(SystemSettingModel, setting_key)
        personal_payload = load_deleted_product_image_cleanup_user_payload(
            row,
            fallback_payload,
            now=now,
        )
        if row is None:
            row = SystemSettingModel(
                key=setting_key,
                value_json=json.dumps(personal_payload, ensure_ascii=False),
            )
            session.add(row)
            session.flush()
        pending_count = session.scalar(
            select(func.count())
            .select_from(DeletedProductImageCleanupModel)
            .where(DeletedProductImageCleanupModel.owner_username == owner_username)
        ) or 0
        return {
            **base_settings,
            **personal_payload,
            "deletedImageCleanupPendingCount": max(0, int(pending_count)),
            "queueHealth": None,
            "updatedAt": datetime_to_public(row.updated_at),
            "serverNow": datetime_to_public(now),
        }


def save_deleted_product_image_cleanup_settings(
    owner_username: str,
    payload: Any,
) -> dict[str, Any]:
    now = datetime.now()
    weekday = normalize_cleanup_weekday(
        getattr(
            payload,
            "deletedImageCleanupWeekday",
            DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_WEEKDAY,
        )
    )
    cleanup_time = normalize_schedule_time(
        getattr(
            payload,
            "deletedImageCleanupTime",
            DELETED_PRODUCT_IMAGE_CLEANUP_DEFAULT_TIME,
        )
    )
    with session_scope() as session:
        fallback_payload = default_time_settings_value(now=now)
        setting_key = deleted_product_image_cleanup_user_setting_key(owner_username)
        row = session.get(SystemSettingModel, setting_key)
        personal_payload = load_deleted_product_image_cleanup_user_payload(
            row,
            fallback_payload,
            now=now,
        )
        personal_payload.update(
            {
                "deletedImageCleanupEnabled": bool(
                    getattr(payload, "deletedImageCleanupEnabled", True)
                ),
                "deletedImageCleanupWeekday": weekday,
                "deletedImageCleanupTime": cleanup_time,
                "deletedImageCleanupNextAt": datetime_to_public(
                    next_weekly_run_at(weekday, cleanup_time, now=now)
                ),
            }
        )
        if row is None:
            row = SystemSettingModel(key=setting_key)
            session.add(row)
        row.value_json = json.dumps(personal_payload, ensure_ascii=False)
        time_row = session.get(
            SystemSettingModel,
            time_settings_row_key(owner_username),
        )
        time_payload = (
            initial_user_time_settings_payload(
                session,
                owner_username,
                now=now,
            )
            if time_row is None
            else load_time_settings_payload(time_row, now=now)
        )
        time_payload.update(personal_payload)
        upsert_time_settings_row(
            session,
            time_payload,
            owner_username,
        )
    return get_deleted_product_image_cleanup_settings(owner_username)


def list_deleted_product_image_cleanups(
    *,
    owner_username: str | None = None,
    page: int | None = None,
    page_size: int | None = None,
) -> dict[str, Any]:
    with session_scope() as session:
        query = select(DeletedProductImageCleanupModel)
        if owner_username is not None:
            query = query.where(
                DeletedProductImageCleanupModel.owner_username == owner_username
            )
        return paginate_query(
            session,
            query,
            order_by=task_status_order_by(DeletedProductImageCleanupModel),
            page=page,
            page_size=page_size,
            response_key="records",
            serializer=deleted_product_image_cleanup_to_public,
        )


def create_deleted_product_image_cleanup_tasks(
    session: Any,
    now: datetime,
    *,
    owner_username: str | None = None,
    excluded_owner_usernames: set[str] | None = None,
) -> tuple[list[tuple[str, str]], int]:
    queued_query = select(DeletedProductImageCleanupModel).where(
        DeletedProductImageCleanupModel.status == "queued"
    )
    if owner_username is not None:
        queued_query = queued_query.where(
            DeletedProductImageCleanupModel.owner_username == owner_username
        )
    elif excluded_owner_usernames:
        queued_query = queued_query.where(
            DeletedProductImageCleanupModel.owner_username.notin_(
                excluded_owner_usernames
            )
        )
    queued_rows = session.scalars(queued_query).all()
    for row in queued_rows:
        task = session.get(SyncTaskModel, row.sync_task_id) if row.sync_task_id else None
        if task is None or task.status not in {"queued", "running"}:
            row.status = "failed"
            row.sync_task_id = None
            row.last_error = "上次图片清理任务未完成，已重新进入待清理队列。"
    pending_query = select(DeletedProductImageCleanupModel).where(
        DeletedProductImageCleanupModel.status.in_(("pending", "failed", "cancelled"))
    )
    if owner_username is not None:
        pending_query = pending_query.where(
            DeletedProductImageCleanupModel.owner_username == owner_username
        )
    elif excluded_owner_usernames:
        pending_query = pending_query.where(
            DeletedProductImageCleanupModel.owner_username.notin_(
                excluded_owner_usernames
            )
        )
    rows = session.scalars(
        pending_query
        .order_by(
            DeletedProductImageCleanupModel.owner_username.asc(),
            DeletedProductImageCleanupModel.store_id.asc(),
            DeletedProductImageCleanupModel.id.asc(),
        )
    ).all()
    groups: dict[tuple[str, int], list[DeletedProductImageCleanupModel]] = {}
    for row in rows:
        if row.store_id is None:
            row.status = "failed"
            row.last_error = "店铺已不存在，无法清理 R-Cabinet 图片。"
            continue
        groups.setdefault((row.owner_username, int(row.store_id)), []).append(row)

    task_refs: list[tuple[str, str]] = []
    record_count = 0
    for (owner_username, store_id), group_rows in groups.items():
        store = session.get(StoreModel, store_id)
        if store is None or store.owner_username != owner_username or not store.enabled:
            for row in group_rows:
                row.status = "failed"
                row.last_error = "店铺不存在或已停用，无法创建图片清理任务。"
            continue
        chunks = [group_rows[index:index + BATCH_TASK_PRODUCT_LIMIT] for index in range(0, len(group_rows), BATCH_TASK_PRODUCT_LIMIT)]
        task_group_id = uuid.uuid4().hex if len(chunks) > 1 else None
        for index, chunk_rows in enumerate(chunks, start=1):
            task_id = uuid.uuid4().hex
            cleanup_ids = [int(row.id) for row in chunk_rows]
            part_label = "" if len(chunks) == 1 else f" {index}/{len(chunks)}"
            task = SyncTaskModel(
                id=task_id,
                owner_username=owner_username,
                store_id=store_id,
                task_group_id=task_group_id,
                task_group_index=index if task_group_id else None,
                task_group_size=len(chunks) if task_group_id else None,
                store_name=store.alias_name or store.store_name,
                task_name=f"已删除商品图片清理{part_label} {store.alias_name or store.store_name} {now:%Y-%m-%d %H:%M}",
                task_type="deleted_product_image_cleanup",
                payload_json=json.dumps({"cleanupRecordIds": cleanup_ids}, ensure_ascii=False),
                status="queued",
                total_count=len(cleanup_ids),
                message="等待清理已删除商品图片",
            )
            session.add(task)
            for row in chunk_rows:
                row.status = "queued"
                row.sync_task_id = task_id
                row.last_error = None
            task_refs.append((owner_username, task_id))
            record_count += len(cleanup_ids)
    return task_refs, record_count


def cleanup_deleted_product_images_scope(
    *,
    force: bool,
    owner_username: str | None,
    excluded_owner_usernames: set[str] | None = None,
) -> tuple[dict[str, int], list[tuple[str, str]]]:
    now = datetime.now()
    with session_scope() as session:
        global_payload = default_time_settings_value(now=now)
        if owner_username is None:
            settings_row = upsert_time_settings_row(session, global_payload)
            payload = global_payload
        else:
            setting_key = deleted_product_image_cleanup_user_setting_key(
                owner_username
            )
            settings_row = session.get(SystemSettingModel, setting_key)
            payload = load_deleted_product_image_cleanup_user_payload(
                settings_row,
                global_payload,
                now=now,
            )
            if settings_row is None:
                settings_row = SystemSettingModel(key=setting_key)
                session.add(settings_row)
        next_at = parse_public_datetime(payload.get("deletedImageCleanupNextAt"))
        if not force and (
            not payload["deletedImageCleanupEnabled"]
            or (next_at is not None and next_at > now)
        ):
            return {"taskCount": 0, "productCount": 0}, []
        task_refs, product_count = create_deleted_product_image_cleanup_tasks(
            session,
            now,
            owner_username=owner_username,
            excluded_owner_usernames=excluded_owner_usernames,
        )
        payload["deletedImageCleanupLastAt"] = datetime_to_public(now)
        payload["deletedImageCleanupLastProductCount"] = product_count
        payload["deletedImageCleanupLastTaskCount"] = len(task_refs)
        payload["deletedImageCleanupNextAt"] = datetime_to_public(
            next_weekly_run_at(
                payload["deletedImageCleanupWeekday"],
                payload["deletedImageCleanupTime"],
                now=now,
            )
        )
        settings_row.value_json = json.dumps(payload, ensure_ascii=False)
    return {
        "taskCount": len(task_refs),
        "productCount": product_count,
    }, task_refs


def cleanup_deleted_product_images(
    *,
    force: bool = False,
    owner_username: str | None = None,
) -> dict[str, int]:
    if owner_username and system_task_dispatch_paused(owner_username):
        return {"taskCount": 0, "productCount": 0}
    if not DELETED_PRODUCT_IMAGE_CLEANUP_LOCK.acquire(blocking=False):
        return {"taskCount": 0, "productCount": 0}
    try:
        summaries: list[dict[str, int]] = []
        task_refs: list[tuple[str, str]] = []
        if owner_username is not None:
            summary, scoped_refs = cleanup_deleted_product_images_scope(
                force=force,
                owner_username=owner_username,
            )
            summaries.append(summary)
            task_refs.extend(scoped_refs)
        else:
            paused_owners = system_task_paused_usernames()
            with session_scope() as session:
                owners = session.scalars(
                    select(UserAccountModel.username).where(
                        UserAccountModel.enabled.is_(True)
                    )
                ).all()
            for personal_owner in sorted(str(owner) for owner in owners):
                if personal_owner in paused_owners:
                    continue
                summary, scoped_refs = cleanup_deleted_product_images_scope(
                    force=force,
                    owner_username=personal_owner,
                )
                summaries.append(summary)
                task_refs.extend(scoped_refs)
        for task_owner, task_id in task_refs:
            dispatch_sync_task(
                task_owner,
                task_id,
                task_type="deleted_product_image_cleanup",
            )
        return {
            "taskCount": sum(item["taskCount"] for item in summaries),
            "productCount": sum(item["productCount"] for item in summaries),
        }
    finally:
        DELETED_PRODUCT_IMAGE_CLEANUP_LOCK.release()


def run_deleted_product_image_cleanup_now(
    *,
    owner_username: str | None = None,
    include_queue_health: bool = True,
) -> dict[str, Any]:
    summary = cleanup_deleted_product_images(
        force=True,
        owner_username=owner_username,
    )
    return {
        "settings": (
            get_time_settings(include_queue_health=include_queue_health)
            if owner_username is None
            else get_deleted_product_image_cleanup_settings(owner_username)
        ),
        "summary": summary,
    }


def perform_deleted_product_image_cleanup(
    owner_username: str,
    store_id: int,
    cleanup_ids: list[int],
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    if not cleanup_ids:
        raise RuntimeError("图片清理任务缺少待清理记录。")
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None or store.owner_username != owner_username:
            raise RuntimeError("图片清理任务关联店铺不存在。")
        service_secret = decrypt_text(store.rakuten_service_secret_encrypted)
        license_key = decrypt_text(store.rakuten_license_key_encrypted)
        rows = session.scalars(
            select(DeletedProductImageCleanupModel).where(
                DeletedProductImageCleanupModel.owner_username == owner_username,
                DeletedProductImageCleanupModel.store_id == store_id,
                DeletedProductImageCleanupModel.id.in_(cleanup_ids),
            )
        ).all()
        rows_by_id = {int(row.id): row for row in rows}
        success_ids: list[int] = []
        failed_ids: list[int] = []
        errors: list[str] = []
        cancelled = False
        for index, cleanup_id in enumerate(cleanup_ids, start=1):
            if task_id and is_task_cancel_requested(SyncTaskModel, task_id):
                cancelled = True
                break
            row = rows_by_id.get(cleanup_id)
            if row is None:
                failed_ids.append(cleanup_id)
                errors.append(f"记录 {cleanup_id}：待清理记录不存在")
                continue
            try:
                _, warnings = delete_cabinet_targets(
                    service_secret,
                    license_key,
                    [target for target in cleanup_record_json_list(row, "cabinet_targets_json") if isinstance(target, dict)],
                )
                if warnings:
                    raise RuntimeError("；".join(warnings[:10]))
                cleanup_queued_local_image_urls(
                    int(row.original_product_id),
                    [str(value) for value in cleanup_record_json_list(row, "local_image_urls_json")],
                )
                success_ids.append(cleanup_id)
                session.delete(row)
            except Exception as exc:
                row.status = "failed"
                row.sync_task_id = task_id
                row.last_error = str(exc)
                failed_ids.append(cleanup_id)
                errors.append(f"{row.product_code or cleanup_id}：{exc}")
            if task_id:
                update_task_progress(
                    SyncTaskModel,
                    task_id,
                    total_count=len(cleanup_ids),
                    success_count=len(success_ids),
                    failed_count=len(failed_ids),
                    message=f"图片清理中，已处理 {index} / {len(cleanup_ids)} 个商品",
                )
        if cancelled:
            processed = set(success_ids) | set(failed_ids)
            for cleanup_id in cleanup_ids:
                row = rows_by_id.get(cleanup_id)
                if row is not None and cleanup_id not in processed:
                    row.status = "pending"
                    row.sync_task_id = None
        return {
            "totalCount": len(cleanup_ids),
            "successCount": len(success_ids),
            "failedCount": len(failed_ids),
            "successIds": success_ids,
            "failedIds": failed_ids,
            "errors": errors,
            "cancelled": cancelled,
        }


def cleanup_queued_local_image_urls(product_id: int, image_urls: list[str]) -> None:
    errors: list[str] = []
    for image_url in unique_texts(image_urls):
        try:
            remove_local_product_image_if_unused(
                image_url,
                [],
                expected_product_id=product_id,
            )
        except Exception as exc:
            errors.append(f"{image_url}: {exc}")
    if errors:
        raise RuntimeError("；".join(errors[:10]))


def cancel_crawl_task(owner_username: str, task_id: str) -> dict[str, Any]:
    return request_task_cancel(CrawlTaskModel, owner_username, task_id, serializer=task_to_public)


def listing_preflight_issue(
    severity: str,
    code: str,
    message: str,
    *,
    field: str = "",
    attribute_name: str = "",
    variant_id: str = "",
) -> dict[str, Any]:
    issue = {
        "severity": severity,
        "code": code,
        "message": message,
    }
    if field:
        issue["field"] = field
    if attribute_name:
        issue["attributeName"] = attribute_name
    if variant_id:
        issue["variantId"] = variant_id
    return issue


def listing_preflight_product_stub(product_id: int, issue: dict[str, Any]) -> dict[str, Any]:
    return {
        "productId": product_id,
        "productCode": str(product_id),
        "productTitle": "",
        "status": "blocked",
        "issueCount": 1,
        "blockerCount": 1 if issue.get("severity") == "blocker" else 0,
        "warningCount": 1 if issue.get("severity") == "warning" else 0,
        "issues": [issue],
        "preview": {},
    }


def listing_preparation_source_fingerprint(
    product: ProductModel,
    raw_payload: dict[str, Any] | None = None,
) -> str:
    payload = deepcopy(raw_payload if isinstance(raw_payload, dict) else product_raw_payload(product))
    payload.pop(LISTING_PREPARATION_CACHE_KEY, None)
    try:
        rules_stat = RAKUTEN_ATTRIBUTE_RULES_PATH.stat()
        attribute_rules_version = f"{rules_stat.st_size}:{rules_stat.st_mtime_ns}"
    except OSError:
        attribute_rules_version = "missing"
    source = {
        "title": normalize_text(product.title),
        "genreId": normalize_text(product.genre_id),
        "price": str(product.price) if product.price is not None else "",
        "imageUrl": normalize_text(product.image_url),
        "attributeRulesVersion": attribute_rules_version,
        "payload": payload,
    }
    serialized = json.dumps(source, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def listing_preparation_cache(
    product: ProductModel,
    raw_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload = raw_payload if isinstance(raw_payload, dict) else product_raw_payload(product)
    cache = payload.get(LISTING_PREPARATION_CACHE_KEY)
    if not isinstance(cache, dict):
        return {}
    if int(cache.get("version") or 0) != LISTING_PREPARATION_CACHE_VERSION:
        return {}
    if normalize_text(cache.get("sourceFingerprint")) != listing_preparation_source_fingerprint(product, payload):
        return {}
    return cache


def listing_prepared_image_map(
    product: ProductModel,
    raw_payload: dict[str, Any] | None = None,
) -> dict[str, str]:
    cache = listing_preparation_cache(product, raw_payload)
    rows = cache.get("images")
    if not isinstance(rows, list):
        return {}
    result: dict[str, str] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        source_url = normalize_text(row.get("sourceUrl"))
        prepared_url = normalize_text(row.get("preparedUrl"))
        if not source_url or not prepared_url or is_missing_local_product_image_url(prepared_url):
            return {}
        result[source_url] = prepared_url
    return result


def listing_preparation_expected_image_urls(
    product: ProductModel,
    raw_payload: dict[str, Any] | None = None,
) -> list[str]:
    payload = raw_payload if isinstance(raw_payload, dict) else product_raw_payload(product)
    main_images = [
        image
        for image in product_images_for_edit(product)
        if not is_gif_image_url(image)
    ][:RAKUTEN_LISTING_IMAGE_LIMIT]
    description_images = [
        image_url
        for description in product_descriptions(payload)
        for image_url in description_image_urls(description.get("value"))
        if not is_gif_image_url(image_url)
    ]
    return unique_texts([*main_images, *description_images])


def listing_preparation_ready(
    product: ProductModel,
    raw_payload: dict[str, Any] | None = None,
) -> bool:
    payload = raw_payload if isinstance(raw_payload, dict) else product_raw_payload(product)
    cache = listing_preparation_cache(product, payload)
    if not cache or int(cache.get("missingImageCount") or 0) > 0:
        return False
    expected_urls = listing_preparation_expected_image_urls(product, payload)
    prepared_map = listing_prepared_image_map(product, payload)
    return len(prepared_map) == len(expected_urls) and all(
        image_url in prepared_map
        for image_url in expected_urls
    )


def _listing_preflight_product_check_uncached(product: ProductModel, store: StoreModel | None) -> dict[str, Any]:
    raw_payload = product_raw_payload(product)
    issues: list[dict[str, Any]] = []
    title = first_text_from_keys(raw_payload, ("itemName", "title", "name")) or product.title
    genre_id = first_text_from_keys(raw_payload, ("genreId", "genre_id", "genre")) or product.genre_id
    product._listing_store_id = store.id if store is not None else 0

    if not normalize_text(title):
        issues.append(listing_preflight_issue("blocker", "missing_title", "商品标题为空，不能上架。", field="title"))
    if not re.fullmatch(r"\d{6}", normalize_text(genre_id)):
        issues.append(listing_preflight_issue("blocker", "invalid_genre_id", "商品缺少 6 位乐天ジャンルID。", field="genreId"))
    elif not rakuten_attribute_group_rule_for_payload({"genreId": normalize_text(genre_id)}):
        issues.append(
            listing_preflight_issue(
                "blocker",
                "unknown_genre_id",
                f"ジャンルID {normalize_text(genre_id)} 未匹配到本地商品属性定义书。",
                field="genreId",
            )
        )

    variants = build_rakuten_listing_variants(raw_payload, product)
    if not variants:
        issues.append(listing_preflight_issue("blocker", "missing_sku_price", "商品缺少可上架的 SKU 价格信息。", field="variants"))
    else:
        invalid_variant_ids = [
            variant_id
            for variant_id, variant in variants.items()
            if not normalize_rakuten_price(variant.get("standardPrice") if isinstance(variant, dict) else None)
        ]
        for variant_id in invalid_variant_ids[:5]:
            issues.append(
                listing_preflight_issue(
                    "blocker",
                    "invalid_sku_price",
                    f"SKU {variant_id} 价格必须为大于 0 的日元整数。",
                    field="variants",
                    variant_id=variant_id,
                )
            )

    images = product_images_for_edit(product)
    non_gif_images = [image for image in images if not is_gif_image_url(image)]
    usable_images = [image for image in non_gif_images if not is_missing_local_product_image_url(image)]
    if not images:
        issues.append(listing_preflight_issue("blocker", "missing_image", "商品缺少图片，不能上架。", field="images"))
    elif not non_gif_images:
        issues.append(listing_preflight_issue("blocker", "only_gif_images", "商品图片全部为 GIF，上架需替换为 jpg/png。", field="images"))
    elif not usable_images:
        issues.append(listing_preflight_issue("blocker", "missing_local_image", "商品本地图片文件不存在或已失效。", field="images"))
    else:
        skipped_gif_count = len(images) - len(non_gif_images)
        if skipped_gif_count > 0:
            issues.append(listing_preflight_issue("warning", "gif_images_skipped", f"{skipped_gif_count} 张 GIF 图片不会参与上架。", field="images"))
        if len(non_gif_images) > RAKUTEN_LISTING_IMAGE_LIMIT:
            issues.append(
                listing_preflight_issue(
                    "warning",
                    "image_limit_trimmed",
                    f"商品图超过 {RAKUTEN_LISTING_IMAGE_LIMIT} 张，正式上架只会使用前 {RAKUTEN_LISTING_IMAGE_LIMIT} 张。",
                    field="images",
                )
            )

    description_gif_count = sum(
        1
        for description in product_descriptions(raw_payload)
        for image_url in description_image_urls(description.get("value"))
        if is_gif_image_url(image_url)
    )
    if description_gif_count > 0:
        issues.append(
            listing_preflight_issue(
                "warning",
                "description_gif_removed",
                f"{description_gif_count} 张说明图 GIF 会在上架时从商品说明中移除。",
                field="descriptions",
            )
        )

    if not any(issue.get("severity") == "blocker" and issue.get("field") in {"title", "genreId", "variants"} for issue in issues):
        try:
            payload = build_rakuten_item_upsert_payload(
                product,
                raw_payload,
                [],
                manage_number=generate_listing_manage_number(product, raw_payload),
                hide_item=True,
            )
        except RuntimeError as exc:
            issues.append(listing_preflight_issue("blocker", "payload_build_failed", str(exc)))
        else:
            issues.extend(listing_preflight_attribute_issues(payload))

    blocker_count = sum(1 for issue in issues if issue.get("severity") == "blocker")
    warning_count = sum(1 for issue in issues if issue.get("severity") == "warning")
    if blocker_count:
        status = "blocked"
    elif warning_count:
        status = "warning"
    else:
        status = "passed"
    return {
        "productId": product.id,
        "productCode": productCodeForError(product),
        "productTitle": product.title,
        "status": status,
        "issueCount": len(issues),
        "blockerCount": blocker_count,
        "warningCount": warning_count,
        "issues": issues,
        "preview": {
            "title": normalize_text(title),
            "genreId": normalize_text(genre_id),
            "variantCount": len(variants),
            "imageCount": len(usable_images),
            "attributeGroup": normalize_text(rakuten_attribute_group_rule_for_payload({"genreId": normalize_text(genre_id)}).get("group")),
        },
    }


def listing_preflight_product_check(product: ProductModel, store: StoreModel | None) -> dict[str, Any]:
    product._listing_store_id = store.id if store is not None else 0
    raw_payload = product_raw_payload(product)
    cached = listing_preparation_cache(product)
    cached_preflight = cached.get("preflight")
    if isinstance(cached_preflight, dict):
        result = deepcopy(cached_preflight)
        result["productId"] = product.id
        result["productCode"] = productCodeForError(product)
        result["productTitle"] = product.title
    else:
        result = _listing_preflight_product_check_uncached(product, store)
    if not listing_preparation_ready(product, raw_payload):
        preparation_issue = listing_preflight_issue(
            "blocker",
            "listing_preparation_required",
            "商品尚未完成上架预处理，请等待预处理任务完成后再上架。",
        )
        result["issues"] = [*list(result.get("issues") or []), preparation_issue]
        result["issueCount"] = int(result.get("issueCount") or 0) + 1
        result["blockerCount"] = int(result.get("blockerCount") or 0) + 1
        result["status"] = "blocked"
    return result


def listing_preflight_attribute_issues(payload: dict[str, Any]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    rule_map = rakuten_attribute_rule_map_for_payload(payload)
    if not rule_map:
        return issues
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return issues
    required_names = [name for name, rule in rule_map.items() if bool(rule.get("required"))]
    either_required_names = rakuten_either_required_attribute_names(rule_map)
    for variant_id, variant in variants.items():
        if not isinstance(variant, dict):
            continue
        attributes = variant.get("attributes") if isinstance(variant.get("attributes"), list) else []
        attributes_by_name = {
            normalize_text(attribute.get("name")): attribute
            for attribute in attributes
            if isinstance(attribute, dict) and normalize_text(attribute.get("name"))
        }
        for name in required_names:
            rule = rule_map.get(name) or {}
            attribute = attributes_by_name.get(name)
            if not attribute or not rakuten_attribute_has_effective_values(attribute, rule):
                issues.append(
                    listing_preflight_issue(
                        "blocker",
                        "missing_required_attribute",
                        f"SKU {variant_id} 缺少必填属性「{name}」。",
                        field="attributes",
                        attribute_name=name,
                        variant_id=normalize_text(variant_id),
                    )
                )
        if either_required_names and not any(
            name in attributes_by_name and rakuten_attribute_has_effective_values(attributes_by_name[name], rule_map.get(name) or {})
            for name in either_required_names
        ):
            issues.append(
                listing_preflight_issue(
                    "blocker",
                    "missing_either_required_attribute",
                    f"SKU {variant_id} 缺少几选一必填属性：{' / '.join(either_required_names)}。",
                    field="attributes",
                    attribute_name=" / ".join(either_required_names),
                    variant_id=normalize_text(variant_id),
                )
            )
        for name, attribute in attributes_by_name.items():
            rule = rule_map.get(name)
            if not rule:
                continue
            values = normalize_rakuten_attribute_values_for_rule(attribute.get("values"), rule)
            if not values:
                if bool(rule.get("required")):
                    issues.append(
                        listing_preflight_issue(
                            "blocker",
                            "empty_required_attribute",
                            f"SKU {variant_id} 属性「{name}」没有有效值。",
                            field="attributes",
                            attribute_name=name,
                            variant_id=normalize_text(variant_id),
                        )
                    )
                continue
            if any(value in RAKUTEN_ATTRIBUTE_PLACEHOLDER_VALUES for value in values) and (
                bool(rule.get("required")) or name in either_required_names
            ):
                issues.append(
                    listing_preflight_issue(
                        "warning",
                        "placeholder_attribute_value",
                        f"SKU {variant_id} 属性「{name}」使用了占位值，建议补真实值。",
                        field="attributes",
                        attribute_name=name,
                        variant_id=normalize_text(variant_id),
                    )
                )
            if bool(rule.get("unitRequired")) and not normalize_text(attribute.get("unit")):
                issues.append(
                    listing_preflight_issue(
                        "blocker",
                        "missing_attribute_unit",
                        f"SKU {variant_id} 属性「{name}」需要单位「{normalize_text(rule.get('unit')) or '指定单位'}」。",
                        field="attributes",
                        attribute_name=name,
                        variant_id=normalize_text(variant_id),
                    )
                )
            recommended_values = rule.get("recommendedValues")
            if (
                normalize_text(rule.get("inputMethod")) == "選択式"
                and isinstance(recommended_values, list)
                and recommended_values
            ):
                recommended_set = {normalize_text(value) for value in recommended_values if normalize_text(value)}
                invalid_values = [value for value in values if value not in recommended_set]
                if invalid_values:
                    issues.append(
                        listing_preflight_issue(
                            "blocker",
                            "invalid_recommended_attribute_value",
                            f"SKU {variant_id} 属性「{name}」的值不在楽天推荐值中：{', '.join(invalid_values[:3])}。",
                            field="attributes",
                            attribute_name=name,
                            variant_id=normalize_text(variant_id),
                        )
                    )
    return issues


def listing_preflight_result(
    product_checks: list[dict[str, Any]],
    global_issues: list[dict[str, Any]],
) -> dict[str, Any]:
    product_issue_count = sum(int(check.get("issueCount") or 0) for check in product_checks)
    product_blocker_count = sum(int(check.get("blockerCount") or 0) for check in product_checks)
    product_warning_count = sum(int(check.get("warningCount") or 0) for check in product_checks)
    global_blocker_count = sum(1 for issue in global_issues if issue.get("severity") == "blocker")
    global_warning_count = sum(1 for issue in global_issues if issue.get("severity") == "warning")
    blocker_count = product_blocker_count + global_blocker_count
    warning_count = product_warning_count + global_warning_count
    passed_count = sum(1 for check in product_checks if check.get("status") == "passed")
    blocked_count = sum(1 for check in product_checks if check.get("status") == "blocked")
    warning_product_count = sum(1 for check in product_checks if check.get("status") == "warning")
    if blocker_count:
        message = f"体检未通过：{blocker_count} 个阻断项，{warning_count} 个警告。"
    elif warning_count:
        message = f"体检通过但有 {warning_count} 个警告，建议确认后再上架。"
    else:
        message = "体检通过，可以创建上架任务。"
    return {
        "canProceed": blocker_count == 0,
        "message": message,
        "summary": {
            "productCount": len(product_checks),
            "passedCount": passed_count,
            "blockedCount": blocked_count,
            "warningProductCount": warning_product_count,
            "issueCount": product_issue_count + len(global_issues),
            "blockerCount": blocker_count,
            "warningCount": warning_count,
        },
        "globalIssues": global_issues,
        "products": product_checks,
    }


def listing_preflight_blocking_messages(product_checks: list[dict[str, Any]], global_issues: list[dict[str, Any]] | None = None) -> list[str]:
    messages = [
        normalize_text(issue.get("message"))
        for issue in (global_issues or [])
        if issue.get("severity") == "blocker" and normalize_text(issue.get("message"))
    ]
    for check in product_checks:
        product_code = normalize_text(check.get("productCode")) or str(check.get("productId") or "")
        for issue in check.get("issues") or []:
            if not isinstance(issue, dict) or issue.get("severity") != "blocker":
                continue
            message = normalize_text(issue.get("message"))
            if message:
                messages.append(f"{product_code}: {message}" if product_code else message)
    return messages


def preflight_listing_task(owner_username: str, payload: Any) -> dict[str, Any]:
    product_ids = [int(value) for value in (getattr(payload, "productIds", None) or [])]
    store_ids = listing_task_payload_store_ids(payload)
    global_issues: list[dict[str, Any]] = []
    product_checks: list[dict[str, Any]] = []
    if not product_ids:
        global_issues.append(listing_preflight_issue("blocker", "missing_products", "请选择要上架的商品。"))
    if not store_ids:
        global_issues.append(listing_preflight_issue("blocker", "missing_store", "请选择上架店铺。"))
    with session_scope() as session:
        stores = session.scalars(select(StoreModel).where(StoreModel.id.in_(store_ids or [-1]))).all()
        stores_by_id = {int(store.id): store for store in stores}
        missing_store_ids = [store_id for store_id in store_ids if store_id not in stores_by_id]
        if missing_store_ids:
            raise RuntimeError("上架店铺不存在。")
        for store in stores:
            if store.owner_username != owner_username:
                raise RuntimeError("不能使用其他用户的店铺上架。")
            if not store.enabled:
                global_issues.append(listing_preflight_issue("blocker", "store_disabled", f"上架店铺「{store.alias_name or store.store_name}」已停用。"))
            if not decrypt_text(store.rakuten_service_secret_encrypted) or not decrypt_text(store.rakuten_license_key_encrypted):
                global_issues.append(listing_preflight_issue("blocker", "missing_store_credentials", f"上架店铺「{store.alias_name or store.store_name}」缺少乐天 Secret 或乐天 Key。"))
        primary_store = stores_by_id.get(store_ids[0]) if store_ids else None

        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(product_ids or [-1]),
            )
        ).all()
        products_by_id = {int(product.id): product for product in products}
        for product_id in product_ids:
            product = products_by_id.get(int(product_id))
            if product is None:
                product_checks.append(
                    listing_preflight_product_stub(
                        int(product_id),
                        listing_preflight_issue("blocker", "product_not_found", "商品不存在或不属于当前用户。"),
                    )
                )
                continue
            base_issues: list[dict[str, Any]] = []
            if product.review_status not in {"approved", "listed_master"}:
                base_issues.append(listing_preflight_issue("blocker", "invalid_review_status", "只有已审核或已上架管理商品可以创建上架任务。"))
            if product.listing_task_id:
                base_issues.append(listing_preflight_issue("blocker", "product_listing_locked", "商品正在上架中，请等待当前任务完成。"))
            listed_store_ids = {int(item.get("storeId") or 0) for item in product_listed_stores(product_raw_payload(product))}
            duplicated_store_names = [
                stores_by_id[store_id].alias_name or stores_by_id[store_id].store_name
                for store_id in store_ids
                if store_id in stores_by_id and store_id in listed_store_ids
            ]
            if duplicated_store_names:
                base_issues.append(listing_preflight_issue("blocker", "duplicate_store_listing", f"商品已上架过以下店铺：{'、'.join(duplicated_store_names[:5])}。"))
            check = listing_preflight_product_check(product, primary_store)
            if base_issues:
                check["issues"] = [*base_issues, *(check.get("issues") or [])]
                check["issueCount"] = len(check["issues"])
                check["blockerCount"] = sum(1 for issue in check["issues"] if issue.get("severity") == "blocker")
                check["warningCount"] = sum(1 for issue in check["issues"] if issue.get("severity") == "warning")
                check["status"] = "blocked" if check["blockerCount"] else ("warning" if check["warningCount"] else "passed")
            product_checks.append(check)
    return listing_preflight_result(product_checks, global_issues)


def cancel_listing_task(owner_username: str, task_id: str) -> dict[str, Any]:
    result = request_task_cancel(ListingTaskModel, owner_username, task_id, serializer=listing_task_to_public)
    dispatch_next_listing_task_safely()
    dispatch_next_sync_task_safely()
    return result


def cancel_sync_task(
    owner_username: str,
    task_id: str,
    *,
    allow_all_owners: bool = False,
) -> dict[str, Any]:
    task_owner = owner_username
    if allow_all_owners:
        with session_scope() as session:
            task = session.get(SyncTaskModel, task_id)
            if task is None:
                raise RuntimeError("同步任务不存在。")
            task_owner = task.owner_username
    result = request_task_cancel(
        SyncTaskModel,
        task_owner,
        task_id,
        serializer=sync_task_to_public,
    )
    dispatch_next_sync_task_safely()
    dispatch_next_listing_task_safely()
    return result


def listing_task_cancel_requested(task_id: str) -> bool:
    return is_task_cancel_requested(ListingTaskModel, task_id)


def request_task_cancel(model: Any, owner_username: str, task_id: str, *, serializer: Callable[[Any], dict[str, Any]]) -> dict[str, Any]:
    with session_scope() as session:
        task = session.get(model, task_id)
        if task is None:
            raise RuntimeError("任务不存在。")
        if task.owner_username != owner_username:
            raise RuntimeError("不能终止其他用户的任务。")
        if task.status == "cancelled":
            return serializer(task)
        if task.status not in {"queued", "running"}:
            raise RuntimeError("只有待执行或执行中的任务可以终止。")
        if task.status == "queued":
            task.status = "cancelled"
            task.message = TASK_CANCELLED_MESSAGE
            task.error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
            if hasattr(task, "warning_detail"):
                task.warning_detail = cancelled_task_warning_detail(existing_warning_detail=task.warning_detail)
            task.finished_at = datetime.now()
            if model is ListingTaskModel:
                release_listing_task_locks(session, owner_username, task)
        else:
            task.message = TASK_CANCEL_REQUESTED_MESSAGE
            task.error_detail = with_task_cancel_marker(task.error_detail)
        session.flush()
        return serializer(task)


def release_listing_task_locks(session: Any, owner_username: str, task: ListingTaskModel) -> None:
    product_ids_payload = listing_task_product_ids_payload(task.product_ids_json)
    success_ids = set(product_ids_payload["successIds"])
    product_ids = [product_id for product_id in product_ids_payload["productIds"] if product_id not in success_ids]
    products = session.scalars(
        select(ProductModel).where(
            ProductModel.owner_username == owner_username,
            ProductModel.id.in_(product_ids or [-1]),
        )
    ).all()
    for product in products:
        clear_listing_product_lock(product, task.id)


def list_products(
    owner_username: str,
    *,
    status: str | None = None,
    collection_source: str | None = None,
    keyword: str | None = None,
    task_id: str | None = None,
    store_id: int | None = None,
    listed_store_id: str | None = None,
    listing_status: str | None = None,
    listed_at_from: str | None = None,
    listed_at_to: str | None = None,
    price_min: Decimal | None = None,
    price_max: Decimal | None = None,
    collected_at_from: str | None = None,
    collected_at_to: str | None = None,
    genre_status: str | None = None,
    genre_path: str | None = None,
    review_filter: str | None = None,
    sort: str | None = None,
    sales_period_days: int | None = None,
    sales_period_from: str | None = None,
    sales_period_to: str | None = None,
    sales_sort: str | None = None,
    sales_min: int | None = None,
    sales_max: int | None = None,
    zero_filter: str | None = None,
    page: int | None = None,
    page_size: int | None = None,
) -> list[dict[str, Any]] | dict[str, Any]:
    with session_scope() as session:
        query = select(ProductModel).where(ProductModel.owner_username == owner_username)
        listed_at_from_value = parse_datetime_filter(listed_at_from)
        listed_at_to_value = parse_datetime_filter(listed_at_to)
        collected_at_from_value = parse_datetime_filter(collected_at_from)
        collected_at_to_value = parse_datetime_filter(collected_at_to)
        normalized_page = max(1, int(page or 1))
        normalized_page_size = min(500, max(1, int(page_size or 0))) if page_size else None
        product_status = _product_status_filter(status)
        normalized_genre_status = normalize_text(genre_status).lower()
        normalized_genre_path = normalize_text(genre_path)
        if normalized_genre_status not in {"", "missing", "present"}:
            raise ValueError("品类状态筛选条件无效")
        if (normalized_genre_status or normalized_genre_path) and product_status != "pending":
            raise ValueError("品类筛选仅支持待审核商品")
        normalized_review_filter = normalize_text(review_filter).lower()
        if normalized_review_filter not in {"", "has", "none"}:
            raise ValueError("评论状态筛选条件无效")
        normalized_sort = normalize_text(sort).lower()
        if normalized_sort not in {
            "",
            "default",
            "price_asc",
            "price_desc",
            "review_count_desc",
        }:
            raise ValueError("商品排序条件无效")
        normalized_collection_source = normalize_text(collection_source).lower()
        if normalized_collection_source not in {"", "manual", "scheduled"}:
            raise ValueError("采集来源筛选条件无效")
        if normalized_collection_source and product_status not in {"pending", "approved"}:
            raise ValueError("采集来源筛选仅支持待审核或已审核商品")
        if (
            normalized_review_filter
            or normalized_sort in {"price_asc", "price_desc", "review_count_desc"}
        ) and not (
            product_status == "pending"
            and normalized_collection_source == "manual"
        ):
            raise ValueError("评论筛选和商品排序仅支持手动采集待审核商品")
        sales_period_range = (
            normalize_store_product_sales_range(
                sales_period_days,
                sales_period_from,
                sales_period_to,
            )
            if product_status == "listed"
            else None
        )
        normalized_sales_sort = (
            sales_sort if sales_sort in {"asc", "desc"} else None
        )
        normalized_zero_filter = normalize_text(zero_filter)
        if normalized_zero_filter not in {
            "",
            "sales",
            "optimization",
            "sales_and_optimization",
        }:
            raise ValueError("零值筛选条件无效")
        if normalized_zero_filter and product_status != "listed":
            raise ValueError("零值筛选仅支持店铺商品")
        if normalized_zero_filter in {"sales", "sales_and_optimization"}:
            sales_min = 0
            sales_max = 0
        listed_store_filter = normalize_listed_store_filter(listed_store_id)
        if product_status:
            query = query.where(ProductModel.review_status == product_status)
        if product_status == "approved":
            query = query.where(ProductModel.listing_task_id.is_(None))
        if normalized_collection_source == "manual":
            query = query.where(ProductModel.collection_source == "manual")
        elif normalized_collection_source == "scheduled":
            query = query.where(ProductModel.collection_source == "scheduled")
        if product_status == "listed":
            query = query.where(ProductModel.store_product_status != "removed")
            if normalized_zero_filter in {
                "optimization",
                "sales_and_optimization",
            }:
                optimized_version_exists = exists(
                    select(1).where(
                        ProductTitleVersionModel.owner_username
                        == ProductModel.owner_username,
                        ProductTitleVersionModel.product_id
                        == ProductModel.id,
                        ProductTitleVersionModel.source != "original",
                    )
                )
                query = query.where(~optimized_version_exists)
        if normalize_text(task_id):
            query = query.where(
                ProductModel.task_id == normalize_text(task_id),
                ProductModel.parent_product_id.is_(None),
            )
        if store_id is not None:
            query = query.where(ProductModel.store_id == store_id)
        if listing_status in {"listed", "unlisted"}:
            query = query.where(ProductModel.rakuten_listing_status == listing_status)
        if keyword:
            if product_status == "listed":
                query = query.where(
                    ProductModel.title.like(f"%{keyword}%")
                    | ProductModel.item_number.like(f"%{keyword}%")
                    | ProductModel.rakuten_manage_number.like(f"%{keyword}%")
                )
            else:
                query = query.where(ProductModel.title.like(f"%{keyword}%"))
        if price_min is not None:
            query = query.where(ProductModel.price >= price_min)
        if price_max is not None:
            query = query.where(ProductModel.price <= price_max)
        if normalized_review_filter == "has":
            query = query.where(ProductModel.review_count > 0)
        elif normalized_review_filter == "none":
            query = query.where(
                or_(
                    ProductModel.review_count == 0,
                    ProductModel.review_count.is_(None),
                )
            )
        if collected_at_from_value is not None:
            query = query.where(ProductModel.created_at >= collected_at_from_value)
        if collected_at_to_value is not None:
            query = query.where(ProductModel.created_at <= collected_at_to_value)
        if product_status == "pending" and (normalized_genre_status or normalized_genre_path):
            genres = load_rakuten_attribute_rules().get("genres") or {}
            valid_genre_ids = list(genres.keys())
            if normalized_genre_path:
                genre_path_prefix = f"{normalized_genre_path}>"
                selected_genre_ids = [
                    normalize_text(genre_id)
                    for genre_id, genre in genres.items()
                    if isinstance(genre, dict)
                    and (
                        normalize_text(genre.get("genrePath")) == normalized_genre_path
                        or normalize_text(genre.get("genrePath")).startswith(genre_path_prefix)
                    )
                ]
                query = (
                    query.where(ProductModel.genre_id.in_(selected_genre_ids))
                    if selected_genre_ids
                    else query.where(ProductModel.id == -1)
                )
            if normalized_genre_status == "missing":
                if valid_genre_ids:
                    query = query.where(
                        or_(
                            ProductModel.genre_id == "",
                            ProductModel.genre_id.is_(None),
                            ProductModel.genre_id.notin_(valid_genre_ids),
                        )
                    )
                else:
                    query = query.where(
                        or_(
                            ProductModel.genre_id == "",
                            ProductModel.genre_id.is_(None),
                        )
                    )
            elif normalized_genre_status == "present":
                query = (
                    query.where(ProductModel.genre_id.in_(valid_genre_ids))
                    if valid_genre_ids
                    else query.where(ProductModel.id == -1)
                )
        if listed_at_from_value is not None:
            query = query.where(ProductModel.listed_at >= listed_at_from_value)
        if listed_at_to_value is not None:
            query = query.where(ProductModel.listed_at <= listed_at_to_value)
        if product_status == "listed_master" and listed_store_filter is not None:
            listed_store_product = aliased(ProductModel)
            listed_store_exists = exists(
                select(1).where(
                    listed_store_product.owner_username
                    == ProductModel.owner_username,
                    listed_store_product.parent_product_id
                    == ProductModel.id,
                    listed_store_product.review_status == "listed",
                    listed_store_product.store_product_status != "removed",
                )
            )
            if listed_store_filter == LISTED_STORE_NONE_FILTER:
                query = query.where(~listed_store_exists)
            else:
                query = query.where(
                    listed_store_exists.where(
                        listed_store_product.store_id
                        == listed_store_filter
                    )
                )
        order_by = product_list_order_by(product_status, normalized_sort)
        if (
            product_status == "listed"
            and sales_period_range is not None
        ):
            synced_store_query = select(SalesSyncStateModel.store_id).where(
                SalesSyncStateModel.owner_username == owner_username,
                SalesSyncStateModel.initial_sync_completed.is_(True),
            )
            if store_id is not None:
                synced_store_query = synced_store_query.where(
                    SalesSyncStateModel.store_id == store_id
                )
            synced_store_ids = set(session.scalars(synced_store_query).all())
            cutoff_date, period_end_date = sales_period_range
            sales_counts = {
                (int(result.store_id), str(result.manage_number)): int(
                    result.effective_units or 0
                )
                for result in session.execute(
                    select(
                        ProductSalesDailyModel.store_id,
                        ProductSalesDailyModel.manage_number,
                        func.sum(
                            ProductSalesDailyModel.effective_units
                        ).label("effective_units"),
                    )
                    .where(
                        ProductSalesDailyModel.owner_username
                        == owner_username,
                        ProductSalesDailyModel.store_id.in_(
                            synced_store_ids or {-1}
                        ),
                        ProductSalesDailyModel.sales_date >= cutoff_date,
                        ProductSalesDailyModel.sales_date
                        <= period_end_date,
                    )
                    .group_by(
                        ProductSalesDailyModel.store_id,
                        ProductSalesDailyModel.manage_number,
                    )
                )
            }
            identity_query = query.with_only_columns(
                ProductModel.id,
                ProductModel.store_id,
                ProductModel.rakuten_manage_number,
                ProductModel.item_number,
                maintain_column_froms=True,
            )

            def identity_sales_count(identity: Any) -> int | None:
                identity_store_id = (
                    int(identity.store_id)
                    if identity.store_id is not None
                    else None
                )
                if identity_store_id not in synced_store_ids:
                    return None
                product_key = canonical_sales_order_item_product_key(
                    manage_number=identity.rakuten_manage_number,
                    item_number=identity.item_number,
                )
                return sales_counts.get(
                    (identity_store_id, product_key),
                    0,
                )

            needs_all_identities = (
                sales_min is not None
                or sales_max is not None
                or normalized_sales_sort is not None
                or normalized_page_size is None
            )
            if needs_all_identities:
                identities = session.execute(
                    identity_query.order_by(*order_by)
                ).all()
                identities_with_sales = [
                    (identity, identity_sales_count(identity))
                    for identity in identities
                ]
                if sales_min is not None or sales_max is not None:
                    identities_with_sales = [
                        (identity, period_sales_count)
                        for identity, period_sales_count
                        in identities_with_sales
                        if period_sales_count is not None
                        and (
                            sales_min is None
                            or period_sales_count >= sales_min
                        )
                        and (
                            sales_max is None
                            or period_sales_count <= sales_max
                        )
                    ]
                if normalized_sales_sort is not None:
                    identities_with_sales.sort(
                        key=lambda item: (
                            item[1] is None,
                            (
                                int(item[1] or 0)
                                if normalized_sales_sort == "asc"
                                else -int(item[1] or 0)
                            ),
                            int(item[0].id),
                        )
                    )
                total = len(identities_with_sales)
                if normalized_page_size:
                    if total:
                        max_page = max(
                            1,
                            (
                                total + normalized_page_size - 1
                            )
                            // normalized_page_size,
                        )
                        normalized_page = min(
                            normalized_page,
                            max_page,
                        )
                    start = (
                        normalized_page - 1
                    ) * normalized_page_size
                    identities_with_sales = identities_with_sales[
                        start:start + normalized_page_size
                    ]
            else:
                total = int(
                    session.scalar(
                        query.with_only_columns(
                            func.count(),
                            maintain_column_froms=True,
                        ).order_by(None)
                    )
                    or 0
                )
                if total:
                    max_page = max(
                        1,
                        (
                            total + normalized_page_size - 1
                        )
                        // normalized_page_size,
                    )
                    normalized_page = min(
                        normalized_page,
                        max_page,
                    )
                identities = session.execute(
                    identity_query.order_by(*order_by)
                    .offset(
                        (
                            normalized_page - 1
                        )
                        * normalized_page_size
                    )
                    .limit(normalized_page_size)
                ).all()
                identities_with_sales = [
                    (identity, identity_sales_count(identity))
                    for identity in identities
                ]

            page_ids = [
                int(identity.id)
                for identity, _ in identities_with_sales
            ]
            rows_by_id = {
                int(row.id): row
                for row in session.scalars(
                    select(ProductModel).where(
                        ProductModel.id.in_(page_ids or [-1]),
                        ProductModel.owner_username == owner_username,
                    )
                ).all()
            }
            rows = [
                rows_by_id[product_id]
                for product_id in page_ids
                if product_id in rows_by_id
            ]
            period_sales_counts = {
                int(identity.id): period_sales_count
                for identity, period_sales_count
                in identities_with_sales
            }
            public_rows = _products_to_public_with_period_sales(
                session,
                rows,
                owner_username,
                sales_period_range,
                period_sales_counts=period_sales_counts,
            )
            if normalized_page_size:
                return {
                    "products": public_rows,
                    "total": total,
                    "page": normalized_page,
                    "pageSize": normalized_page_size,
                }
            return public_rows
        if normalized_page_size:
            total = session.scalar(
                query.with_only_columns(
                    func.count(),
                    maintain_column_froms=True,
                ).order_by(None)
            ) or 0
            if total:
                max_page = max(1, (int(total) + normalized_page_size - 1) // normalized_page_size)
                normalized_page = min(normalized_page, max_page)
            page_ids = session.scalars(
                query.with_only_columns(
                    ProductModel.id,
                    maintain_column_froms=True,
                )
                .order_by(*order_by)
                .offset((normalized_page - 1) * normalized_page_size)
                .limit(normalized_page_size)
            ).all()
            rows_by_id = {
                int(row.id): row
                for row in session.scalars(
                    select(ProductModel).where(
                        ProductModel.id.in_(page_ids or [-1]),
                        ProductModel.owner_username == owner_username,
                    )
                ).all()
            }
            rows = [
                rows_by_id[int(product_id)]
                for product_id in page_ids
                if int(product_id) in rows_by_id
            ]
            return {
                "products": _products_to_public_with_period_sales(
                    session,
                    rows,
                    owner_username,
                    sales_period_range,
                ),
                "total": int(total),
                "page": normalized_page,
                "pageSize": normalized_page_size,
            }

        rows = session.scalars(query.order_by(*order_by)).all()
        return _products_to_public_with_period_sales(
            session,
            rows,
            owner_username,
            sales_period_range,
        )


def _products_to_public_with_period_sales(
    session: Any,
    rows: list[ProductModel],
    owner_username: str,
    sales_period_range: tuple[date, date] | None,
    *,
    period_sales_counts: dict[int, int | None] | None = None,
) -> list[dict[str, Any]]:
    if not rows:
        return []

    title_optimization_task_ids = active_title_optimization_task_ids(
        session,
        owner_username,
        [int(row.id) for row in rows],
    )
    product_delete_task_ids = active_product_delete_task_ids(
        session,
        owner_username,
        [int(row.id) for row in rows],
    )
    if sales_period_range is None and period_sales_counts is None:
        return [
            product_to_public(
                row,
                title_optimization_task_id=title_optimization_task_ids.get(int(row.id)),
                product_delete_task_id=product_delete_task_ids.get(int(row.id)),
            )
            for row in rows
        ]

    title_optimization_counts = {
        int(result.product_id): int(result.optimization_count or 0)
        for result in session.execute(
            select(
                ProductTitleVersionModel.product_id,
                func.count(ProductTitleVersionModel.id).label("optimization_count"),
            )
            .where(
                ProductTitleVersionModel.owner_username == owner_username,
                ProductTitleVersionModel.product_id.in_([row.id for row in rows]),
                ProductTitleVersionModel.source != "original",
            )
            .group_by(ProductTitleVersionModel.product_id)
        )
    }

    sales_counts: dict[tuple[int, str], int] = {}
    synced_store_ids: set[int] = set()
    if period_sales_counts is None:
        store_ids = {
            int(row.store_id)
            for row in rows
            if row.store_id is not None
        }
        synced_store_ids = set(
            session.scalars(
                select(SalesSyncStateModel.store_id).where(
                    SalesSyncStateModel.owner_username == owner_username,
                    SalesSyncStateModel.store_id.in_(store_ids or {-1}),
                    SalesSyncStateModel.initial_sync_completed.is_(True),
                )
            ).all()
        )
        product_keys = {
            canonical_sales_order_item_product_key(
                manage_number=row.rakuten_manage_number,
                item_number=row.item_number,
            )
            for row in rows
            if row.store_id in synced_store_ids
        }
        cutoff_date, period_end_date = sales_period_range
        sales_counts = {
            (int(result.store_id), str(result.manage_number)): int(
                result.effective_units or 0
            )
            for result in session.execute(
                select(
                    ProductSalesDailyModel.store_id,
                    ProductSalesDailyModel.manage_number,
                    func.sum(
                        ProductSalesDailyModel.effective_units
                    ).label("effective_units"),
                )
                .where(
                    ProductSalesDailyModel.owner_username == owner_username,
                    ProductSalesDailyModel.store_id.in_(
                        synced_store_ids or {-1}
                    ),
                    ProductSalesDailyModel.manage_number.in_(
                        product_keys or {""}
                    ),
                    ProductSalesDailyModel.sales_date >= cutoff_date,
                    ProductSalesDailyModel.sales_date <= period_end_date,
                )
                .group_by(
                    ProductSalesDailyModel.store_id,
                    ProductSalesDailyModel.manage_number,
                )
            )
        }
    return [
        product_to_public(
            row,
            period_sales_count=(
                period_sales_counts.get(int(row.id))
                if period_sales_counts is not None
                else (
                    sales_counts.get(
                        (
                            int(row.store_id),
                            canonical_sales_order_item_product_key(
                                manage_number=row.rakuten_manage_number,
                                item_number=row.item_number,
                            ),
                        ),
                        0,
                    )
                    if row.store_id in synced_store_ids
                    else None
                )
            ),
            title_optimization_count=title_optimization_counts.get(int(row.id), 0),
            title_optimization_task_id=title_optimization_task_ids.get(int(row.id)),
            product_delete_task_id=product_delete_task_ids.get(int(row.id)),
        )
        for row in rows
    ]


def store_product_sales_summary(
    session: Any,
    owner_username: str,
    store_id: int | None,
    sales_period_range: tuple[date, date],
    *,
    synced_store_ids: set[int],
    sales_counts: dict[tuple[int, str], int],
) -> dict[str, Any] | None:
    if store_id is None:
        return None
    normalized_store_id = int(store_id)
    period_from, period_to = sales_period_range
    if normalized_store_id not in synced_store_ids:
        return {
            "storeId": normalized_store_id,
            "periodFrom": period_from.isoformat(),
            "periodTo": period_to.isoformat(),
            "syncCompleted": False,
            "totalEffectiveUnits": None,
            "currentProductEffectiveUnits": None,
            "outsideCurrentProductEffectiveUnits": None,
            "currentProductCount": 0,
            "outsideCurrentProductCount": 0,
        }

    current_product_keys = {
        canonical_sales_order_item_product_key(
            manage_number=row.rakuten_manage_number,
            item_number=row.item_number,
        )
        for row in session.execute(
            select(
                ProductModel.rakuten_manage_number,
                ProductModel.item_number,
            ).where(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == normalized_store_id,
                ProductModel.review_status == "listed",
                ProductModel.store_product_status != "removed",
            )
        )
    }
    store_sales_counts = {
        product_key: max(0, int(effective_units or 0))
        for (sales_store_id, product_key), effective_units in sales_counts.items()
        if int(sales_store_id) == normalized_store_id
    }
    total_effective_units = sum(store_sales_counts.values())
    current_product_effective_units = sum(
        store_sales_counts.get(product_key, 0)
        for product_key in current_product_keys
    )
    outside_current_product_keys = {
        product_key
        for product_key, effective_units in store_sales_counts.items()
        if effective_units > 0 and product_key not in current_product_keys
    }
    return {
        "storeId": normalized_store_id,
        "periodFrom": period_from.isoformat(),
        "periodTo": period_to.isoformat(),
        "syncCompleted": True,
        "totalEffectiveUnits": total_effective_units,
        "currentProductEffectiveUnits": current_product_effective_units,
        "outsideCurrentProductEffectiveUnits": max(
            0,
            total_effective_units - current_product_effective_units,
        ),
        "currentProductCount": len(current_product_keys),
        "outsideCurrentProductCount": len(outside_current_product_keys),
    }


def get_store_product_sales_summary(owner_username: str, store_id: int) -> dict[str, Any]:
    sales_period_range = normalize_store_product_sales_range(365, None, None)
    if sales_period_range is None:
        raise RuntimeError("销量周期不正确。")
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None:
            raise RuntimeError("店铺不存在。")
        if store.owner_username != owner_username:
            raise RuntimeError("不能查看其他用户的店铺销量。")
        sync_completed = bool(
            session.scalar(
                select(SalesSyncStateModel.store_id).where(
                    SalesSyncStateModel.owner_username == owner_username,
                    SalesSyncStateModel.store_id == store_id,
                    SalesSyncStateModel.initial_sync_completed.is_(True),
                )
            )
        )
        synced_store_ids = {int(store_id)} if sync_completed else set()
        period_from, period_to = sales_period_range
        sales_counts = {
            (int(result.store_id), str(result.manage_number)): int(
                result.effective_units or 0
            )
            for result in session.execute(
                select(
                    ProductSalesDailyModel.store_id,
                    ProductSalesDailyModel.manage_number,
                    func.sum(
                        ProductSalesDailyModel.effective_units
                    ).label("effective_units"),
                )
                .where(
                    ProductSalesDailyModel.owner_username == owner_username,
                    ProductSalesDailyModel.store_id == store_id,
                    ProductSalesDailyModel.sales_date >= period_from,
                    ProductSalesDailyModel.sales_date <= period_to,
                )
                .group_by(
                    ProductSalesDailyModel.store_id,
                    ProductSalesDailyModel.manage_number,
                )
            )
        }
        return store_product_sales_summary(
            session,
            owner_username,
            store_id,
            sales_period_range,
            synced_store_ids=synced_store_ids,
            sales_counts=sales_counts,
        ) or {}


def active_title_optimization_task_ids(
    session: Any,
    owner_username: str,
    product_ids: list[int],
) -> dict[int, str]:
    normalized_ids = set(normalize_product_ids(product_ids))
    if not normalized_ids:
        return {}
    result: dict[int, str] = {}
    rows = session.scalars(
        select(SyncTaskModel).where(
            SyncTaskModel.owner_username == owner_username,
            SyncTaskModel.task_type == "title_optimization",
            SyncTaskModel.status.in_(("queued", "running")),
        )
    ).all()
    for task in rows:
        payload = sync_task_payload(task)
        for product_id in normalize_product_ids(list(payload.get("productIds") or [])):
            if product_id in normalized_ids:
                result.setdefault(product_id, task.id)
    return result


def active_product_delete_task_ids(
    session: Any,
    owner_username: str,
    product_ids: list[int],
) -> dict[int, str]:
    normalized_ids = set(normalize_product_ids(product_ids))
    if not normalized_ids:
        return {}
    result: dict[int, str] = {}
    rows = session.scalars(
        select(SyncTaskModel).where(
            SyncTaskModel.owner_username == owner_username,
            SyncTaskModel.task_type == "product_delete",
            SyncTaskModel.status.in_(("queued", "running")),
        )
    ).all()
    for task in rows:
        payload = sync_task_payload(task)
        for product_id in normalize_product_ids(list(payload.get("productIds") or [])):
            if product_id in normalized_ids:
                result.setdefault(product_id, task.id)
    return result


def normalize_listed_store_filter(value: str | None) -> int | str | None:
    normalized = normalize_text(value)
    if not normalized:
        return None
    if normalized == LISTED_STORE_NONE_FILTER:
        return LISTED_STORE_NONE_FILTER
    if normalized.isdigit():
        store_id = int(normalized)
        return store_id if store_id > 0 else None
    return None


def product_matches_listed_store_filter(row: ProductModel, listed_store_filter: int | str) -> bool:
    listed_stores = product_listed_stores(product_raw_payload(row))
    if listed_store_filter == LISTED_STORE_NONE_FILTER:
        return len(listed_stores) < 1
    return any(int(item.get("storeId") or 0) == listed_store_filter for item in listed_stores)


def product_list_order_by(status: str | None, sort: str = "") -> tuple[Any, ...]:
    if sort == "price_asc":
        return (
            ProductModel.price.is_(None).asc(),
            ProductModel.price.asc(),
            ProductModel.created_at.desc(),
            ProductModel.id.desc(),
        )
    if sort == "price_desc":
        return (
            ProductModel.price.is_(None).asc(),
            ProductModel.price.desc(),
            ProductModel.created_at.desc(),
            ProductModel.id.desc(),
        )
    if sort == "review_count_desc":
        return (
            func.coalesce(ProductModel.review_count, 0).desc(),
            ProductModel.created_at.desc(),
            ProductModel.id.desc(),
        )
    if status == "listed":
        return (
            ProductModel.listed_at.is_(None).asc(),
            ProductModel.listed_at.asc(),
            ProductModel.id.asc(),
        )
    return (ProductModel.created_at.desc(), ProductModel.id.desc())


def ensure_store_owner(owner_username: str, store_id: int) -> StoreModel:
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None:
            raise RuntimeError("店铺不存在。")
        if row.owner_username != owner_username:
            raise RuntimeError("不能操作其他用户的店铺。")
        session.expunge(row)
        return row


def list_stores(
    owner_username: str,
    *,
    page: int | None = None,
    page_size: int | None = None,
    reveal: bool = False,
) -> list[dict[str, Any]] | dict[str, Any]:
    with session_scope() as session:
        query = select(StoreModel).where(StoreModel.owner_username == owner_username)
        synced_store_ids = set(
            session.scalars(
                select(SalesSyncStateModel.store_id).where(
                    SalesSyncStateModel.owner_username == owner_username,
                    SalesSyncStateModel.initial_sync_completed.is_(True),
                )
            ).all()
        )
        recent_year_order_counts: dict[int, int] = {}
        if synced_store_ids:
            cutoff = sales_now_naive() - timedelta(days=365)
            recent_year_order_counts = {
                int(row.store_id): int(row.order_count or 0)
                for row in session.execute(
                    select(
                        SalesOrderModel.store_id,
                        func.count(SalesOrderModel.id).label("order_count"),
                    )
                    .where(
                        SalesOrderModel.owner_username == owner_username,
                        SalesOrderModel.store_id.in_(synced_store_ids),
                        SalesOrderModel.ordered_at >= cutoff,
                    )
                    .group_by(SalesOrderModel.store_id)
                )
            }

        def serialize_store(row: StoreModel) -> dict[str, Any]:
            return store_to_public(
                row,
                reveal=reveal,
                recent_year_order_count=(
                    recent_year_order_counts.get(int(row.id), 0)
                    if row.id in synced_store_ids
                    else None
                ),
            )

        return paginate_query(
            session,
            query,
            order_by=StoreModel.id.asc(),
            page=page,
            page_size=page_size,
            response_key="stores",
            serializer=serialize_store,
        )


def save_store(owner_username: str, payload: Any, store_id: int | None = None) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(StoreModel, store_id) if store_id else None
        if row is not None and row.owner_username != owner_username:
            raise RuntimeError("不能操作其他用户的店铺。")
        if row is None:
            row = StoreModel(owner_username=owner_username)
            session.add(row)

        row.alias_name = str(getattr(payload, "aliasName", "") or "").strip()
        row.platform = str(getattr(payload, "platform", "") or "rakuten").strip()
        row.enabled = bool(getattr(payload, "enabled", True))
        row.description = str(getattr(payload, "description", "") or "").strip()

        incoming_service_secret = str(getattr(payload, "rakutenServiceSecret", "") or "").strip()
        incoming_license_key = str(getattr(payload, "rakutenLicenseKey", "") or "").strip()
        service_secret = incoming_service_secret or decrypt_text(row.rakuten_service_secret_encrypted)
        license_key = incoming_license_key or decrypt_text(row.rakuten_license_key_encrypted)

        if row.id is None and (not incoming_service_secret or not incoming_license_key):
            raise RuntimeError("新增店铺时必须填写乐天 Secret 和乐天 Key。")
        shop_meta = fetch_rakuten_shop_meta(service_secret, license_key)
        row.store_code = shop_meta["shopCode"]
        row.store_name = shop_meta["shopName"]
        if not row.alias_name:
            row.alias_name = row.store_name
        row.store_url = build_rakuten_store_url(row.store_code)
        if incoming_service_secret:
            row.rakuten_service_secret_encrypted = encrypt_text(incoming_service_secret)
        if incoming_license_key:
            row.rakuten_license_key_encrypted = encrypt_text(incoming_license_key)
        with session.no_autoflush:
            duplicated_query = select(StoreModel).where(
                StoreModel.owner_username == owner_username,
                StoreModel.store_code == row.store_code,
            )
            if row.id is not None:
                duplicated_query = duplicated_query.where(StoreModel.id != row.id)
            duplicated_store = session.scalar(duplicated_query)
        if duplicated_store is not None:
            raise RuntimeError("店铺编号已存在。")
        session.flush()
        return store_to_public(row)


def update_owned_store_settings(
    owner_username: str,
    store_id: int,
    *,
    alias_name: str,
    enabled: bool,
) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None or row.owner_username != owner_username:
            raise RuntimeError("店铺不存在或无权操作。")
        normalized_alias = normalize_text(alias_name)
        row.alias_name = normalized_alias or row.store_name
        row.enabled = bool(enabled)
        session.flush()
        return store_to_public(row, reveal=True)


def delete_store(owner_username: str, store_id: int) -> None:
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None:
            return
        if row.owner_username != owner_username:
            raise RuntimeError("不能删除其他用户的店铺。")
        product_ids = session.scalars(
            select(ProductModel.id).where(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == store_id,
            )
        ).all()
        parent_products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.review_status == "listed_master",
            )
        ).all()
        for product in parent_products:
            remove_product_listed_store_mark(product, store_id)
        if product_ids:
            session.query(ProductModel).filter(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == store_id,
            ).delete(synchronize_session=False)
        session.delete(row)
    cleanup_product_image_ids([int(product_id) for product_id in product_ids])


def store_product_sync_identity(
    *,
    manage_number: str | None,
    item_number: str | None,
) -> str:
    return normalize_text(manage_number) or normalize_text(item_number)


def store_product_local_title_override(session: Any, product: ProductModel) -> bool:
    selected = session.scalar(
        select(ProductTitleVersionModel).where(
            ProductTitleVersionModel.product_id == product.id,
            ProductTitleVersionModel.owner_username == product.owner_username,
            ProductTitleVersionModel.is_selected.is_(True),
        )
    )
    return selected is not None and selected.source != "original"


def apply_store_product_remote_update(
    session: Any,
    owner_username: str,
    store: StoreModel,
    row: ProductModel,
    item: dict[str, Any],
    *,
    active_words: list[str] | None = None,
) -> bool:
    item_number = first_text_from_keys(item, ("itemNumber", "manageNumber"))
    manage_number = first_text_from_keys(item, ("manageNumber", "itemNumber"))
    source_url = (
        first_url_from_keys(item, ("itemUrl", "itemPageUrl", "url"))
        or build_public_item_page_url(store.store_code, item_number or manage_number)
    )
    title = first_text_from_keys(item, ("itemName", "title", "name"))
    raw_payload = item if isinstance(item, dict) else {}
    remote_title_override = store_product_local_title_override(session, row)
    current_payload = product_raw_payload(row)
    current_tagline = product_tagline(current_payload)
    normalized = {
        "sourceUrl": source_url,
        "sourceUrlHash": make_source_url_hash(
            f"{source_url}#store={store.id}&manage={quote(manage_number or item_number, safe='')}"
        ),
        "imageUrl": first_rakuten_image_url(item, store.store_code),
        "price": price_from_rakuten_item(item),
        "itemNumber": item_number or manage_number,
        "rakutenManageNumber": manage_number,
        "listingStatus": rakuten_listing_status_from_item(item),
        "genreId": first_text_from_keys(item, ("genreId", "genre_id", "genre")),
        "shopName": store.store_name,
    }
    changed = any(
        (
            row.source_url != normalized["sourceUrl"],
            row.source_url_hash != normalized["sourceUrlHash"],
            row.image_url != normalized["imageUrl"],
            row.price != normalized["price"],
            row.item_number != normalized["itemNumber"],
            row.rakuten_manage_number != normalized["rakutenManageNumber"],
            row.rakuten_listing_status != normalized["listingStatus"],
            row.genre_id != normalized["genreId"],
            row.shop_name != normalized["shopName"],
            not remote_title_override and row.title != title[:500],
            row.store_product_status == "removed",
        )
    )
    row.source_url = normalized["sourceUrl"]
    row.source_url_hash = normalized["sourceUrlHash"]
    row.store_id = store.id
    row.rakuten_manage_number = normalized["rakutenManageNumber"]
    row.item_number = normalized["itemNumber"]
    row.shop_name = normalized["shopName"]
    row.image_url = normalized["imageUrl"]
    row.price = Decimal(str(normalized["price"])) if normalized["price"] is not None else None
    row.currency = "JPY"
    row.genre_id = normalized["genreId"]
    row.rakuten_listing_status = normalized["listingStatus"]
    row.review_status = "listed"
    row.store_product_status = "active"
    row.store_last_seen_at = datetime.now()
    row.listed_at = (
        parse_rakuten_datetime_value(raw_payload.get("created"))
        or row.listed_at
    )
    if not remote_title_override:
        row.title = title[:500]
        row.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
    else:
        preserved_payload = patch_local_item_detail(
            raw_payload,
            title=row.title,
            tagline=current_tagline,
            variants=[],
        )
        row.raw_payload_json = json.dumps(preserved_payload, ensure_ascii=False)
    row.last_error = None
    if active_words:
        sanitized_payload, _ = sanitize_product_payload(
            json.loads(row.raw_payload_json or "{}"),
            active_words,
        )
        row.raw_payload_json = json.dumps(sanitized_payload, ensure_ascii=False)
    return changed


def mark_missing_store_products_removed(
    session: Any,
    rows: list[ProductModel],
    remote_identities: set[str],
) -> int:
    removed_count = 0
    for row in rows:
        identity = store_product_sync_identity(
            manage_number=row.rakuten_manage_number,
            item_number=row.item_number,
        )
        if not identity or identity in remote_identities:
            continue
        if row.store_product_status != "removed":
            removed_count += 1
        remove_listed_store_mark_for_store_product(session, row)
        row.store_product_status = "removed"
        row.rakuten_listing_status = "unlisted"
        row.store_last_seen_at = None
    return removed_count

def verify_store_credentials(
    row: StoreModel,
    *,
    include_product_counts: bool = True,
    include_cabinet_usage: bool = True,
) -> None:
    service_secret = decrypt_text(row.rakuten_service_secret_encrypted)
    license_key = decrypt_text(row.rakuten_license_key_encrypted)
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    checked_at = datetime.now()
    verify_store_key_and_update_meta(row, service_secret, license_key)
    if include_cabinet_usage:
        update_store_cabinet_usage(row, service_secret, license_key, checked_at=checked_at)
    if include_product_counts:
        items, _ = fetch_rakuten_store_items_with_total(service_secret, license_key)
        apply_store_product_counts(row, items, checked_at=checked_at)
    row.last_checked_at = checked_at
    row.last_error = None


def verify_store_key_and_update_meta(row: StoreModel, service_secret: str, license_key: str) -> None:
    """Validate RMS credentials before any quantity or usage API calls."""
    shop_meta = fetch_rakuten_shop_meta(service_secret, license_key)
    row.store_code = shop_meta["shopCode"]
    row.store_name = shop_meta["shopName"]
    if not row.alias_name:
        row.alias_name = row.store_name
    row.store_url = build_rakuten_store_url(row.store_code)


def update_store_cabinet_usage(
    row: StoreModel,
    service_secret: str,
    license_key: str,
    *,
    checked_at: datetime | None = None,
) -> None:
    usage = fetch_rakuten_cabinet_usage(service_secret, license_key)
    apply_store_cabinet_usage(row, usage, checked_at=checked_at)


def apply_store_cabinet_usage(row: StoreModel, usage: dict[str, int], *, checked_at: datetime | None = None) -> None:
    row.cabinet_used_folder_count = usage["usedFolderCount"]
    row.cabinet_remaining_folder_count = usage["remainingFolderCount"]
    row.cabinet_usage_checked_at = checked_at or datetime.now()


def apply_store_product_counts(
    row: StoreModel,
    items: list[dict[str, Any]],
    *,
    checked_at: datetime | None = None,
) -> None:
    fetched_count = len(items)
    total_count = fetched_count
    listed_count = sum(1 for item in items if rakuten_listing_status_from_item(item) == "listed")
    row.rakuten_product_total_count = total_count
    row.rakuten_product_listed_count = listed_count
    row.rakuten_product_unlisted_count = max(0, fetched_count - listed_count)
    row.last_checked_at = checked_at or datetime.now()


def sync_store_cabinet_usage_fields(row: StoreModel, service_secret: str, license_key: str) -> None:
    try:
        update_store_cabinet_usage(row, service_secret, license_key)
    except Exception:
        # R-Cabinet usage is advisory for UI statistics; listing still surfaces hard failures later.
        pass


def sync_store(owner_username: str, store_id: int) -> dict[str, Any]:
    task = create_sync_task(owner_username, store_id)
    return {
        "store": task.get("store"),
        "syncTask": task.get("syncTask"),
        "syncedCount": task.get("syncedCount", 0),
    }


def list_store_empty_cabinet_folders(owner_username: str, store_id: int) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None:
            raise RuntimeError("店铺不存在。")
        if row.owner_username != owner_username:
            raise RuntimeError("不能查看其他用户店铺的 R-Cabinet 文件夹。")
        service_secret = decrypt_text(row.rakuten_service_secret_encrypted)
        license_key = decrypt_text(row.rakuten_license_key_encrypted)
        if not service_secret or not license_key:
            raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
        folders = fetch_rakuten_cabinet_folders(service_secret, license_key)
        folder_prefix = "YXYYYYMMDD-N"
        empty_folders = [
            {
                "folderId": int(folder.get("folderId") or 0),
                "folderName": normalize_text(folder.get("folderName")),
                "folderPath": normalize_text(folder.get("folderPath")),
                "fileCount": 0,
            }
            for folder in folders
            if int(folder.get("folderId") or 0) > 0
            and int(folder.get("fileCount") or 0) == 0
            and listing_cabinet_folder_identity(folder) is not None
        ]
        empty_folders.sort(key=cabinet_listing_folder_sort_key)
        return {
            "store": {
                "id": int(row.id),
                "storeCode": row.store_code,
                "storeName": row.store_name,
                "aliasName": row.alias_name,
            },
            "folders": empty_folders,
            "total": len(empty_folders),
            "folderPrefix": folder_prefix,
            "manualCleanupRequired": True,
        }


def perform_store_sync(owner_username: str, store_id: int, *, task_id: str | None = None) -> dict[str, Any]:
    raise_if_task_cancelled(SyncTaskModel, task_id)
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None:
            raise RuntimeError("店铺不存在。")
        if row.owner_username != owner_username:
            raise RuntimeError("不能同步其他用户的店铺。")
        if not row.enabled:
            raise RuntimeError("店铺已停用，不能更新商品。")
        synced_count = 0
        failed_count = 0
        service_secret = decrypt_text(row.rakuten_service_secret_encrypted)
        license_key = decrypt_text(row.rakuten_license_key_encrypted)
        verify_store_credentials(row, include_product_counts=False)
        items, rakuten_total_count = fetch_rakuten_store_items_with_total(service_secret, license_key)
        apply_store_product_counts(row, items)
        active_words = active_sensitive_words(session)
        local_rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == row.id,
            )
        ).all()
        local_by_manage_number: dict[str, ProductModel] = {}
        local_by_item_number: dict[str, ProductModel] = {}
        for local_row in local_rows:
            manage_identity = normalize_text(local_row.rakuten_manage_number)
            item_identity = normalize_text(local_row.item_number)
            if manage_identity:
                local_by_manage_number.setdefault(manage_identity, local_row)
            if item_identity:
                local_by_item_number.setdefault(item_identity, local_row)
        remote_identities = {
            store_product_sync_identity(
                manage_number=first_text_from_keys(item, ("manageNumber", "itemNumber")),
                item_number=first_text_from_keys(item, ("itemNumber", "manageNumber")),
            )
            for item in items
        }
        remote_identities.discard("")
        remote_listing_complete = (
            rakuten_total_count is None
            or len(items) >= int(rakuten_total_count)
        )
        removed_count = (
            mark_missing_store_products_removed(
                session,
                local_rows,
                remote_identities,
            )
            if remote_listing_complete
            else 0
        )
        added_count = 0
        updated_count = 0
        unchanged_count = 0
        failed_count = 0
        errors: list[str] = []
        total_count = len(items) + removed_count
        processed_count = removed_count
        if task_id:
            update_task_progress(
                SyncTaskModel,
                task_id,
                total_count=total_count,
                success_count=processed_count - failed_count,
                failed_count=0,
                message=(
                    f"同步中，新增 0，更新 0，移除 {removed_count}，"
                    f"已处理 {processed_count} / {total_count} 条"
                ),
            )
        for index, item in enumerate(items, start=1):
            raise_if_task_cancelled(SyncTaskModel, task_id)
            manage_number = first_text_from_keys(item, ("manageNumber", "itemNumber"))
            item_number = first_text_from_keys(item, ("itemNumber", "manageNumber"))
            identity = store_product_sync_identity(
                manage_number=manage_number,
                item_number=item_number,
            )
            existing = (
                local_by_manage_number.get(normalize_text(manage_number))
                or local_by_item_number.get(normalize_text(item_number))
            )
            try:
                item_outcome = "unchanged"
                with session.begin_nested():
                    if existing is None:
                        if upsert_store_product(
                            session,
                            owner_username,
                            row,
                            item,
                            active_words=active_words,
                        ):
                            item_outcome = "added"
                        else:
                            item_outcome = "failed"
                    elif apply_store_product_remote_update(
                        session,
                        owner_username,
                        row,
                        existing,
                        item,
                        active_words=active_words,
                    ):
                        item_outcome = "updated"
                    session.flush()
                if item_outcome == "added":
                    added_count += 1
                elif item_outcome == "updated":
                    updated_count += 1
                elif item_outcome == "failed":
                    failed_count += 1
                else:
                    unchanged_count += 1
            except Exception as exc:
                failed_count += 1
                errors.append(f"{identity or item_number or manage_number or index}: {exc}")
                logger.exception("店铺商品差异同步失败 store_id=%s identity=%s", store_id, identity)
                if existing is not None:
                    existing.last_error = str(exc)[:1000]
            processed_count += 1
            if task_id:
                update_task_progress(
                    SyncTaskModel,
                    task_id,
                    total_count=total_count,
                    success_count=processed_count - failed_count,
                    failed_count=failed_count,
                    message=(
                        f"同步中，新增 {added_count}，更新 {updated_count}，"
                        f"移除 {removed_count}，已处理 {processed_count} / {total_count} 条"
                    ),
                )
        row.last_product_synced_at = datetime.now()
        session.flush()
        synced_count = added_count + updated_count + unchanged_count + removed_count
        result = {
            "store": store_to_public(row),
            "totalCount": total_count,
            "syncedCount": synced_count,
            "failedCount": failed_count,
            "addedCount": added_count,
            "updatedCount": updated_count,
            "removedCount": removed_count,
            "unchangedCount": unchanged_count,
            "errors": errors[:50],
            "remoteListingComplete": remote_listing_complete,
            "rakutenTotalCount": int(rakuten_total_count or len(items)),
            "fetchedCount": len(items),
            "cancelled": False,
        }
    return result


def list_sync_tasks(
    owner_username: str,
    *,
    page: int | None = None,
    page_size: int | None = None,
    task_ids: list[str] | None = None,
    task_group: str = "sync",
    all_owners: bool = False,
) -> list[dict[str, Any]] | dict[str, Any]:
    dispatch_next_sync_task_safely()
    normalized_group = normalize_text(task_group) or "sync"
    if normalized_group not in {"sync", "title_optimization", "image_cleanup", "listing_image_upload"}:
        raise RuntimeError("任务分组无效。")
    with session_scope() as session:
        scoped_owner = None if all_owners else owner_username
        finalize_stale_cancel_requested_tasks(
            session,
            SyncTaskModel,
            action_label="任务",
            owner_username=scoped_owner,
        )
        reconcile_interrupted_running_tasks(
            session,
            SyncTaskModel,
            owner_username=scoped_owner,
        )
        query = select(SyncTaskModel)
        if not all_owners:
            query = query.where(SyncTaskModel.owner_username == owner_username)
        normalized_task_ids = normalize_task_ids(task_ids) if task_ids else []
        if normalized_task_ids:
            query = query.where(SyncTaskModel.id.in_(normalized_task_ids))
        elif normalized_group == "title_optimization":
            query = query.where(
                SyncTaskModel.task_type.in_(TITLE_OPTIMIZATION_TASK_TYPES)
            )
        elif normalized_group == "image_cleanup":
            query = query.where(
                SyncTaskModel.task_type.in_(IMAGE_CLEANUP_TASK_TYPES)
            )
        elif normalized_group == "listing_image_upload":
            query = query.where(
                SyncTaskModel.task_type.in_(LISTING_IMAGE_UPLOAD_TASK_TYPES)
            )
        else:
            query = query.where(
                SyncTaskModel.task_type.notin_(SPECIALIZED_SYNC_TASK_TYPES),
                or_(
                    SyncTaskModel.task_type != "product_replace",
                    SyncTaskModel.status != "preview_ready",
                ),
            )
        if normalized_task_ids:
            rows = session.scalars(query.order_by(*task_status_order_by(SyncTaskModel))).all()
            return [sync_task_to_public(row) for row in rows]
        rows = session.scalars(query).all()
        return paginate_task_groups(
            rows,
            page=page,
            page_size=page_size,
            response_key="syncTasks",
            serializer=sync_task_to_public,
            group_kind="sync",
        )


def sync_task_visible_in_list(task_type: str, status: str) -> bool:
    return not (normalize_text(task_type) == "product_replace" and normalize_text(status) == "preview_ready")


def delete_sync_tasks(
    owner_username: str,
    task_ids: list[str],
    *,
    allow_all_owners: bool = False,
) -> dict[str, Any]:
    normalized_ids = normalize_task_ids(task_ids)
    with session_scope() as session:
        query = select(SyncTaskModel).where(
            SyncTaskModel.id.in_(normalized_ids),
        )
        if not allow_all_owners:
            query = query.where(SyncTaskModel.owner_username == owner_username)
        rows = session.scalars(query).all()
        found_ids = {row.id for row in rows}
        for row in rows:
            session.delete(row)
        deleted_ids = [row.id for row in rows]
        return {
            "deletedIds": deleted_ids,
            "failedIds": [task_id for task_id in normalized_ids if task_id not in found_ids],
            "deletedCount": len(deleted_ids),
        }


def create_sync_task(owner_username: str, store_id: int) -> dict[str, Any]:
    ensure_store_owner(owner_username, store_id)
    task_id = create_sync_task_record(
        owner_username,
        store_id,
        task_type="store_sync",
        task_name_prefix="商品同步",
        message="等待同步店铺商品",
    )
    dispatch_next_sync_task()
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        store = session.get(StoreModel, task.store_id) if task and task.store_id else None
        return {
            "syncTask": sync_task_to_public(task) if task else {"id": task_id},
            "store": store_to_public(store) if store else None,
            "syncedCount": 0,
        }


def create_listing_status_sync_task(owner_username: str, store_id: int, listing_status: str) -> dict[str, Any]:
    if listing_status not in {"listed", "unlisted"}:
        raise RuntimeError("上架状态不合法。")
    ensure_store_owner(owner_username, store_id)
    action_label = "全部上架" if listing_status == "listed" else "全部下架"
    task_id = create_sync_task_record(
        owner_username,
        store_id,
        task_type="listing_status",
        task_name_prefix=action_label,
        message=f"等待执行{action_label}",
        payload={"listingStatus": listing_status},
    )
    dispatch_next_sync_task()
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        store = session.get(StoreModel, task.store_id) if task and task.store_id else None
        return {
            "syncTask": sync_task_to_public(task) if task else {"id": task_id},
            "store": store_to_public(store) if store else None,
            "summary": {
                "total": 0,
                "successCount": 0,
                "failedCount": 0,
                "message": f"{action_label}任务已创建",
                "errors": [],
            },
        }


def create_product_listing_status_sync_task(owner_username: str, product_ids: list[int], listing_status: str) -> dict[str, Any]:
    if listing_status not in {"listed", "unlisted"}:
        raise RuntimeError("上架状态不合法。")
    normalized_ids = normalize_product_ids(product_ids)
    if not normalized_ids:
        raise RuntimeError("请先选择商品。")
    action_label = "批量上架" if listing_status == "listed" else "批量下架"
    store_id = validate_sync_task_products(owner_username, normalized_ids)
    with session_scope() as session:
        ready_ids, blocked_ids = partition_product_ids_by_active_task_conflicts(
            session,
            owner_username,
            [store_id],
            normalized_ids,
        )
    chunks = conflict_aware_product_chunks(ready_ids, blocked_ids)
    task_group_id = uuid.uuid4().hex if len(chunks) > 1 else None
    task_ids: list[str] = []
    for index, (chunk_ids, blocked) in enumerate(chunks, start=1):
        task_name_prefix = action_label if len(chunks) == 1 else f"{action_label} {index}/{len(chunks)}"
        task_id = create_sync_task_record(
            owner_username,
            store_id,
            task_type="product_listing_status",
            task_name_prefix=task_name_prefix,
            message=(
                "排队中，等待同一商品当前任务完成"
                if blocked
                else f"等待执行{action_label}"
            ),
            payload={"listingStatus": listing_status, "productIds": chunk_ids},
            total_count=len(chunk_ids),
            task_group_id=task_group_id,
            task_group_index=index if task_group_id else None,
            task_group_size=len(chunks) if task_group_id else None,
        )
        task_ids.append(task_id)
    dispatch_next_sync_task()
    return created_sync_tasks_response(task_ids, message=f"{action_label}任务已创建", total=len(normalized_ids))


def create_product_delete_sync_task(owner_username: str, product_ids: list[int]) -> dict[str, Any]:
    normalized_ids = normalize_product_ids(product_ids)
    if not normalized_ids:
        raise RuntimeError("请先选择商品。")
    store_id = validate_sync_task_products(owner_username, normalized_ids)
    with session_scope() as session:
        active_delete_tasks = active_product_delete_task_ids(
            session,
            owner_username,
            normalized_ids,
        )
        existing_task_ids = list(dict.fromkeys(active_delete_tasks.values()))
        pending_ids = [
            product_id
            for product_id in normalized_ids
            if product_id not in active_delete_tasks
        ]
        ready_ids, blocked_ids = partition_product_ids_by_active_task_conflicts(
            session,
            owner_username,
            [store_id],
            pending_ids,
        )
    chunks = conflict_aware_product_chunks(ready_ids, blocked_ids)
    if not chunks and existing_task_ids:
        return created_sync_tasks_response(
            existing_task_ids,
            message="所选商品已有删除任务，已继续跟踪原任务",
            total=len(normalized_ids),
        )
    task_group_id = uuid.uuid4().hex if len(chunks) > 1 else None
    task_ids: list[str] = list(existing_task_ids)
    for index, (chunk_ids, blocked) in enumerate(chunks, start=1):
        task_name_prefix = "批量删除" if len(chunks) == 1 else f"批量删除 {index}/{len(chunks)}"
        task_id = create_sync_task_record(
            owner_username,
            store_id,
            task_type="product_delete",
            task_name_prefix=task_name_prefix,
            message=(
                "排队中，等待同一商品当前任务完成"
                if blocked
                else "等待执行批量删除"
            ),
            payload={"productIds": chunk_ids},
            total_count=len(chunk_ids),
            task_group_id=task_group_id,
            task_group_index=index if task_group_id else None,
            task_group_size=len(chunks) if task_group_id else None,
        )
        task_ids.append(task_id)
    dispatch_next_sync_task()
    message = (
        "批量删除任务已创建"
        if not existing_task_ids
        else "批量删除任务已创建，重复商品继续跟踪原删除任务"
    )
    return created_sync_tasks_response(task_ids, message=message, total=len(normalized_ids))


def create_product_title_optimization_task(
    owner_username: str,
    product_ids: list[int],
) -> dict[str, Any]:
    normalized_ids = normalize_product_ids(product_ids)
    if not normalized_ids:
        raise RuntimeError("请先选择商品。")
    chunks = chunk_product_ids(normalized_ids)
    task_group_id = uuid.uuid4().hex if len(chunks) > 1 else None
    task_ids: list[str] = []
    with session_scope() as session:
        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(normalized_ids),
                ProductModel.review_status == "listed",
            )
        ).all()
        found_ids = {int(product.id) for product in products}
        missing_ids = [product_id for product_id in normalized_ids if product_id not in found_ids]
        if missing_ids:
            raise RuntimeError("存在不可优化的商品，请刷新列表后重新选择。")
        store_ids = {int(product.store_id) for product in products if product.store_id is not None}
        if len(store_ids) != 1:
            raise RuntimeError("请选择同一个店铺下的店铺商品。")
        store_id = next(iter(store_ids))
        store = session.get(StoreModel, store_id)
        if store is None or not store.enabled:
            raise RuntimeError("商品关联店铺不存在或已停用。")

        active_tasks = session.scalars(
            select(SyncTaskModel).where(
                SyncTaskModel.owner_username == owner_username,
                SyncTaskModel.task_type == "title_optimization",
                SyncTaskModel.status.in_(("queued", "running")),
            )
        ).all()
        active_product_ids: set[int] = set()
        for task in active_tasks:
            active_product_ids.update(
                normalize_product_ids(list(sync_task_payload(task).get("productIds") or []))
            )
        duplicated_ids = [product_id for product_id in normalized_ids if product_id in active_product_ids]
        if duplicated_ids:
            raise RuntimeError("所选商品中有标题优化任务正在执行，请等待任务完成后再试。")

        created_at = datetime.now()
        for index, chunk_ids in enumerate(chunks, start=1):
            task_id = uuid.uuid4().hex
            part_label = "" if len(chunks) == 1 else f" {index}/{len(chunks)}"
            task = SyncTaskModel(
                id=task_id,
                owner_username=owner_username,
                store_id=store.id,
                task_group_id=task_group_id,
                task_group_index=index if task_group_id else None,
                task_group_size=len(chunks) if task_group_id else None,
                store_name=store.alias_name or store.store_name,
                task_name=(
                    f"批量优化标题{part_label} "
                    f"{store.alias_name or store.store_name} {created_at:%Y-%m-%d %H:%M}"
                ),
                task_type="title_optimization",
                payload_json=json.dumps({"productIds": chunk_ids}, ensure_ascii=False),
                status="queued",
                total_count=len(chunk_ids),
                message="等待执行标题优化",
            )
            session.add(task)
            task_ids.append(task_id)
        session.flush()

    for task_id in task_ids:
        dispatch_sync_task(
            owner_username,
            task_id,
            task_type="title_optimization",
        )
    return created_sync_tasks_response(
        task_ids,
        message="批量标题优化任务已创建",
        total=len(normalized_ids),
    )


def normalize_product_ids(product_ids: list[int]) -> list[int]:
    result: list[int] = []
    seen: set[int] = set()
    for value in product_ids or []:
        product_id = int(value)
        if product_id in seen:
            continue
        seen.add(product_id)
        result.append(product_id)
    return result


def chunk_product_ids(product_ids: list[int], *, chunk_size: int = BATCH_TASK_PRODUCT_LIMIT) -> list[list[int]]:
    normalized_ids = normalize_product_ids(product_ids)
    size = max(1, int(chunk_size or BATCH_TASK_PRODUCT_LIMIT))
    return [normalized_ids[index : index + size] for index in range(0, len(normalized_ids), size)]


def product_review_statuses(owner_username: str, product_ids: list[int]) -> set[str]:
    normalized_ids = normalize_product_ids(product_ids)
    if not normalized_ids:
        raise RuntimeError("请先选择商品。")
    with session_scope() as session:
        statuses = session.scalars(
            select(ProductModel.review_status).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(normalized_ids),
            )
        ).all()
    if len(statuses) != len(normalized_ids):
        raise RuntimeError("部分商品不存在，不能执行该操作。")
    return {str(value or "") for value in statuses}


def validate_sync_task_products(owner_username: str, product_ids: list[int]) -> int:
    with session_scope() as session:
        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(product_ids),
                ProductModel.review_status == "listed",
            )
        ).all()
        found_ids = {product.id for product in products}
        missing_ids = [product_id for product_id in product_ids if product_id not in found_ids]
        if missing_ids:
            raise RuntimeError(f"存在不可操作的店铺商品：{', '.join(str(value) for value in missing_ids[:10])}")
        store_ids = {product.store_id for product in products if product.store_id}
        if not products or len(store_ids) != 1:
            raise RuntimeError("请选择同一个店铺下的店铺商品。")
        store_id = int(next(iter(store_ids)))
        store = session.get(StoreModel, store_id)
        if store is None:
            raise RuntimeError("商品关联店铺不存在。")
        if not store.enabled:
            raise RuntimeError("商品关联店铺已停用，不能创建同步任务。")
        return store_id


def created_sync_task_response(task_id: str, *, message: str) -> dict[str, Any]:
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        store = session.get(StoreModel, task.store_id) if task and task.store_id else None
        return {
            "syncTask": sync_task_to_public(task) if task else {"id": task_id},
            "store": store_to_public(store) if store else None,
            "summary": {
                "total": 0,
                "successCount": 0,
                "failedCount": 0,
                "message": message,
                "errors": [],
            },
        }


def created_sync_tasks_response(task_ids: list[str], *, message: str, total: int = 0) -> dict[str, Any]:
    normalized_ids = [task_id for task_id in task_ids if task_id]
    if not normalized_ids:
        raise RuntimeError("同步任务创建失败。")
    with session_scope() as session:
        rows = session.scalars(
            select(SyncTaskModel)
            .where(SyncTaskModel.id.in_(normalized_ids))
            .order_by(SyncTaskModel.created_at.asc())
        ).all()
        task_by_id = {row.id: row for row in rows}
        tasks = [sync_task_to_public(task_by_id[task_id]) for task_id in normalized_ids if task_id in task_by_id]
        first_task = task_by_id.get(normalized_ids[0])
        store = session.get(StoreModel, first_task.store_id) if first_task and first_task.store_id else None
        split_message = message if len(tasks) == 1 else f"{message}，已拆分为 {len(tasks)} 个任务，每个最多 {BATCH_TASK_PRODUCT_LIMIT} 条"
        return {
            "syncTask": tasks[0] if tasks else {"id": normalized_ids[0]},
            "syncTasks": tasks,
            "store": store_to_public(store) if store else None,
            "summary": {
                "total": int(total or 0),
                "successCount": 0,
                "failedCount": 0,
                "message": split_message,
                "errors": [],
            },
        }


def create_sync_task_record(
    owner_username: str,
    store_id: int,
    *,
    task_type: str,
    task_name_prefix: str,
    message: str,
    payload: dict[str, Any] | None = None,
    total_count: int = 0,
    task_group_id: str | None = None,
    task_group_index: int | None = None,
    task_group_size: int | None = None,
) -> str:
    ensure_system_task_dispatch_allowed(owner_username)
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None:
            raise RuntimeError("店铺不存在。")
        if not store.enabled:
            raise RuntimeError("店铺已停用，不能创建同步任务。")
        task_payload = payload or {}
        initial_total_count = max(0, int(total_count or 0))
        if initial_total_count <= 0:
            initial_total_count = sync_task_payload_product_count(task_payload)
        task = SyncTaskModel(
            id=uuid.uuid4().hex,
            owner_username=owner_username,
            store_id=store.id,
            task_group_id=task_group_id,
            task_group_index=task_group_index,
            task_group_size=task_group_size,
            store_name=store.alias_name or store.store_name,
            task_name=f"{task_name_prefix} {store.alias_name or store.store_name} {datetime.now():%Y-%m-%d %H:%M}",
            task_type=task_type,
            payload_json=json.dumps(task_payload, ensure_ascii=False),
            status="queued",
            total_count=initial_total_count,
            message=message,
        )
        session.add(task)
        session.flush()
        return task.id


def replacement_draft_from_collected_item(item: dict[str, Any]) -> dict[str, Any]:
    raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
    raw = json.loads(json.dumps(raw, ensure_ascii=False))
    title = normalize_text(item.get("title")) or first_text_from_keys(raw, ("itemName", "title", "name"))
    genre_id = normalize_text(item.get("genre_id")) or first_text_from_keys(raw, ("genreId", "genre_id", "genre"))
    images = product_editable_image_urls(raw) or unique_texts([item.get("image_url")])
    images, _ = preferred_rakuten_image_urls(images)
    price = price_from_rakuten_item(raw)
    if price is None:
        price = item.get("price")
    variants = build_rakuten_listing_variants(raw, SimpleNamespace(price=price))
    raw.update({"title": title, "itemName": title, "genreId": genre_id, "images": images})
    return {
        "title": title,
        "tagline": product_tagline(raw),
        "genreId": genre_id,
        "genrePath": rakuten_genre_path(genre_id),
        "genrePathZh": rakuten_genre_zh_path(rakuten_genre_path(genre_id)),
        "price": float(price) if price is not None else None,
        "images": images,
        "descriptions": product_descriptions(raw),
        "variants": variants,
        "raw": raw,
    }


def product_replacement_metadata(raw_payload: dict[str, Any]) -> dict[str, Any]:
    metadata = raw_payload.get("_replacement")
    return metadata if isinstance(metadata, dict) else {}


def replacement_draft_from_pending_product(product: ProductModel) -> dict[str, Any]:
    raw = product_raw_payload(product)
    images = product_editable_image_urls(raw, shop_code=product_shop_code(product, raw))
    variants = build_rakuten_listing_variants(raw, product)
    return {
        "title": normalize_text(product.title),
        "tagline": product_tagline(raw),
        "genreId": normalize_text(product.genre_id),
        "genrePath": rakuten_genre_path(product.genre_id),
        "genrePathZh": rakuten_genre_zh_path(rakuten_genre_path(product.genre_id)),
        "price": float(product.price) if product.price is not None else None,
        "images": images,
        "descriptions": product_descriptions(raw),
        "variants": variants,
        "raw": raw,
    }


def replacement_detail_from_product_public(product: dict[str, Any]) -> dict[str, Any]:
    detail = product.get("detail") if isinstance(product.get("detail"), dict) else {}
    raw = detail.get("raw") if isinstance(detail.get("raw"), dict) else {}
    return {
        "title": normalize_text(detail.get("title") or product.get("title")),
        "tagline": normalize_text(detail.get("tagline") or product.get("tagline")),
        "genreId": normalize_text(detail.get("genreId") or product.get("genreId")),
        "genrePath": normalize_text(product.get("genrePath")),
        "genrePathZh": normalize_text(product.get("genrePathZh")),
        "price": product.get("price"),
        "images": list(detail.get("images") or product.get("images") or []),
        "descriptions": list(detail.get("descriptions") or []),
        "variants": raw.get("variants") if isinstance(raw.get("variants"), dict) else {},
        "raw": raw,
    }


def product_replacement_difference(before: dict[str, Any], after: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for key in ("title", "tagline", "price"):
        result[key] = {"changed": before.get(key) != after.get(key), "before": before.get(key), "after": after.get(key)}
    result["genre"] = {
        "changed": normalize_text(before.get("genreId")) != normalize_text(after.get("genreId")),
        "before": before.get("genrePathZh") or before.get("genrePath") or before.get("genreId"),
        "after": after.get("genrePathZh") or after.get("genrePath") or after.get("genreId"),
    }
    for key in ("images", "variants", "descriptions"):
        before_value = before.get(key) or ([] if key != "variants" else {})
        after_value = after.get(key) or ([] if key != "variants" else {})
        result[key] = {
            "changed": before_value != after_value,
            "beforeCount": len(before_value),
            "afterCount": len(after_value),
        }
    return result


def product_replacement_payload(row: SyncTaskModel) -> dict[str, Any]:
    try:
        payload = json.loads(row.payload_json or "{}")
    except ValueError:
        payload = {}
    return payload if isinstance(payload, dict) else {}


def product_replacement_to_public(row: SyncTaskModel) -> dict[str, Any]:
    payload = product_replacement_payload(row)
    return {
        "task": sync_task_to_public(row),
        "targetProductId": payload.get("targetProductId"),
        "sourceUrl": payload.get("sourceUrl"),
        "before": payload.get("targetSnapshot") or {},
        "after": payload.get("draftPayload") or {},
        "difference": payload.get("difference") or {},
        "pendingProduct": payload.get("pendingProduct") or None,
        "result": payload.get("result") or {},
    }


def create_product_replacement_preview(owner_username: str, product_id: int, source_url: str) -> dict[str, Any]:
    normalized_url = normalize_rakuten_product_target(source_url)
    crawl_task_id = uuid.uuid4().hex
    with session_scope() as session:
        session.add(CrawlTaskModel(
            id=crawl_task_id,
            owner_username=owner_username,
            source_type="product_replace",
            target=normalized_url,
            mode="manual",
            status="running",
            total_count=1,
            success_count=0,
            failed_count=0,
            message="正在采集替换商品",
            started_at=datetime.now(),
        ))
        session.flush()
    try:
        collected_item = collect_product_detail(normalized_url)
        draft = replacement_draft_from_collected_item(collected_item)
        if not draft["title"]:
            raise RuntimeError("来源商品没有有效标题。")
        with session_scope() as session:
            crawl_task = session.get(CrawlTaskModel, crawl_task_id)
            target = session.get(ProductModel, product_id)
            if target is None or target.owner_username != owner_username:
                raise RuntimeError("目标店铺商品不存在。")
            if target.review_status != "listed" or not target.store_id:
                raise RuntimeError("只有店铺商品可以执行商品替换。")
            store = session.get(StoreModel, target.store_id)
            if store is None or not store.enabled:
                raise RuntimeError("目标商品所属店铺不存在或已停用。")
            active = session.scalar(
                select(SyncTaskModel.id).where(
                    SyncTaskModel.owner_username == owner_username,
                    SyncTaskModel.task_type == "product_replace",
                    SyncTaskModel.status.in_(["preview_ready", "queued", "running"]),
                    SyncTaskModel.payload_json.like(f'%"targetProductId": {product_id}%'),
                )
            )
            if active:
                raise RuntimeError("当前商品已有进行中的替换任务。")
            cleaned_raw, _ = sanitize_product_payload(draft.get("raw") or {}, active_sensitive_words(session))
            draft["raw"] = cleaned_raw
            draft["title"] = first_text_from_keys(cleaned_raw, ("itemName", "title", "name")) or draft["title"]
            draft["tagline"] = product_tagline(cleaned_raw)
            if not normalize_text(draft["title"]):
                raise RuntimeError(EMPTY_SENSITIVE_TITLE_SAVE_ERROR)
            target_public = product_detail_to_public(target)
            before = replacement_detail_from_product_public(target_public)
            payload = {
                "targetProductId": target.id,
                "sourceUrl": normalized_url,
                "crawlTaskId": crawl_task_id,
                "targetSnapshot": target_public,
                "sourcePayload": draft,
                "draftPayload": draft,
                "difference": product_replacement_difference(before, draft),
            }
            task = SyncTaskModel(
                id=uuid.uuid4().hex,
                owner_username=owner_username,
                store_id=store.id,
                store_name=store.alias_name or store.store_name,
                task_name=f"替换商品 {target.rakuten_manage_number or target.item_number}",
                task_type="product_replace",
                payload_json=json.dumps(payload, ensure_ascii=False),
                status="preview_ready",
                total_count=1,
                message="采集完成，等待确认替换",
            )
            session.add(task)
            session.flush()
            pending_raw = json.loads(json.dumps(draft.get("raw") or {}, ensure_ascii=False))
            pending_raw.update({
                "title": draft["title"],
                "itemName": draft["title"],
                "tagline": draft["tagline"],
                "genreId": draft["genreId"],
                "images": list(draft["images"]),
                "variants": draft["variants"],
                "_replacement": {
                    "taskId": task.id,
                    "targetProductId": target.id,
                    "targetManageNumber": target.rakuten_manage_number or target.item_number,
                    "targetStoreId": store.id,
                    "targetStoreName": store.alias_name or store.store_name,
                },
            })
            pending = ProductModel(
                owner_username=owner_username,
                task_id=crawl_task_id,
                store_id=None,
                source_url=normalized_url,
                source_url_hash=make_source_url_hash(f"{normalized_url}#replacement:{target.id}:{task.id}"),
                title=draft["title"][:500],
                image_url=normalize_text((draft["images"] or [""])[0]),
                item_number=first_text_from_keys(pending_raw, ("itemNumber", "manageNumber")),
                shop_name=first_text_from_keys(pending_raw, ("shopName", "shop")),
                genre_id=draft["genreId"],
                price=Decimal(str(draft["price"])) if draft["price"] is not None else None,
                currency="JPY",
                review_status="pending",
                raw_payload_json=json.dumps(pending_raw, ensure_ascii=False),
            )
            session.add(pending)
            session.flush()
            payload["pendingProductId"] = pending.id
            payload["pendingProduct"] = product_to_public(pending)
            task.payload_json = json.dumps(payload, ensure_ascii=False)
            if crawl_task is not None:
                crawl_task.status = "success"
                crawl_task.success_count = 1
                crawl_task.failed_count = 0
                crawl_task.message = "替换商品采集完成，已进入待审核商品"
                crawl_task.finished_at = datetime.now()
            session.flush()
            return product_replacement_to_public(task)
    except Exception as exc:
        with session_scope() as session:
            crawl_task = session.get(CrawlTaskModel, crawl_task_id)
            if crawl_task is not None:
                crawl_task.status = "failed"
                crawl_task.success_count = 0
                crawl_task.failed_count = 1
                crawl_task.message = "替换商品采集失败"
                crawl_task.error_detail = str(exc)
                crawl_task.finished_at = datetime.now()
                session.flush()
        raise


def get_product_replacement(owner_username: str, task_id: str) -> dict[str, Any]:
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is None or task.owner_username != owner_username or task.task_type != "product_replace":
            raise RuntimeError("商品替换任务不存在。")
        return product_replacement_to_public(task)


def update_product_replacement_draft(owner_username: str, task_id: str, updates: dict[str, Any]) -> dict[str, Any]:
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is None or task.owner_username != owner_username or task.task_type != "product_replace":
            raise RuntimeError("商品替换任务不存在。")
        if task.status != "preview_ready":
            raise RuntimeError("当前替换任务不能再编辑。")
        payload = product_replacement_payload(task)
        draft = dict(payload.get("draftPayload") or {})
        for key in ("title", "tagline", "genreId", "price", "images", "descriptions", "variants"):
            if key in updates and updates[key] is not None:
                draft[key] = float(updates[key]) if key == "price" else updates[key]
        genre_id = normalize_text(draft.get("genreId"))
        draft["genrePath"] = rakuten_genre_path(genre_id)
        draft["genrePathZh"] = rakuten_genre_zh_path(draft["genrePath"])
        raw = dict(draft.get("raw") or {})
        raw.update({
            "title": normalize_text(draft.get("title")),
            "itemName": normalize_text(draft.get("title")),
            "tagline": normalize_text(draft.get("tagline")),
            "genreId": genre_id,
            "images": list(draft.get("images") or []),
            "descriptions": list(draft.get("descriptions") or []),
            "variants": draft.get("variants") if isinstance(draft.get("variants"), dict) else {},
        })
        draft["raw"] = raw
        before = replacement_detail_from_product_public(payload.get("targetSnapshot") or {})
        payload["draftPayload"] = draft
        payload["difference"] = product_replacement_difference(before, draft)
        task.payload_json = json.dumps(payload, ensure_ascii=False)
        session.flush()
        return product_replacement_to_public(task)


def confirm_product_replacement(owner_username: str, task_id: str, manage_number: str) -> dict[str, Any]:
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is None or task.owner_username != owner_username or task.task_type != "product_replace":
            raise RuntimeError("商品替换任务不存在。")
        if task.status != "preview_ready":
            raise RuntimeError("当前替换任务不能确认。")
        payload = product_replacement_payload(task)
        target = session.get(ProductModel, int(payload.get("targetProductId") or 0))
        if target is None or target.owner_username != owner_username or target.review_status != "listed":
            raise RuntimeError("目标店铺商品不存在或状态已变化。")
        pending = session.get(ProductModel, int(payload.get("pendingProductId") or 0))
        if pending is None or pending.owner_username != owner_username or pending.review_status != "pending":
            raise RuntimeError("待审核替换商品不存在或状态已变化。")
        metadata = product_replacement_metadata(product_raw_payload(pending))
        if normalize_text(metadata.get("taskId")) != task.id:
            raise RuntimeError("待审核替换商品与替换任务不匹配。")
        expected = normalize_text(target.rakuten_manage_number or target.item_number)
        if normalize_text(manage_number) != expected:
            raise RuntimeError("商品管理编号输入不正确。")
        draft = replacement_draft_from_pending_product(pending)
        if not normalize_text(draft.get("title")):
            raise RuntimeError("替换后商品标题不能为空。")
        if not rakuten_genre_path(draft.get("genreId")):
            raise RuntimeError("替换后商品缺少有效品类。")
        if not draft.get("images"):
            raise RuntimeError("替换后商品缺少图片。")
        if not isinstance(draft.get("variants"), dict) or not draft.get("variants"):
            raise RuntimeError("替换后商品缺少 SKU。")
        payload["draftPayload"] = draft
        payload["pendingProduct"] = product_to_public(pending)
        task.payload_json = json.dumps(payload, ensure_ascii=False)
        task.status = "queued"
        task.message = "等待执行商品替换"
        task.error_detail = None
        task.finished_at = None
        session.flush()
        result = product_replacement_to_public(task)
    dispatch_next_sync_task()
    return result


def confirm_pending_product_replacement(owner_username: str, product_id: int, manage_number: str) -> dict[str, Any]:
    with session_scope() as session:
        pending = session.get(ProductModel, product_id)
        if pending is None or pending.owner_username != owner_username or pending.review_status != "pending":
            raise RuntimeError("待审核替换商品不存在或状态已变化。")
        raw_payload = product_raw_payload(pending)
        metadata = product_replacement_metadata(raw_payload)
        target_product_id = int(metadata.get("targetProductId") or 0)
        if not target_product_id:
            raise RuntimeError("当前待审核商品不是替换商品。")
        target = session.get(ProductModel, target_product_id)
        if target is None or target.owner_username != owner_username or target.review_status != "listed":
            raise RuntimeError("目标店铺商品不存在或状态已变化。")
        store = session.get(StoreModel, target.store_id) if target.store_id else None
        if store is None or not store.enabled:
            raise RuntimeError("目标商品所属店铺不存在或已停用。")
        expected = normalize_text(target.rakuten_manage_number or target.item_number)
        if normalize_text(manage_number) != expected:
            raise RuntimeError("商品管理编号输入不正确。")

        draft = replacement_draft_from_pending_product(pending)
        if not normalize_text(draft.get("title")):
            raise RuntimeError("替换后商品标题不能为空。")
        if not rakuten_genre_path(draft.get("genreId")):
            raise RuntimeError("替换后商品缺少有效品类。")
        if not draft.get("images"):
            raise RuntimeError("替换后商品缺少图片。")
        if not isinstance(draft.get("variants"), dict) or not draft.get("variants"):
            raise RuntimeError("替换后商品缺少 SKU。")

        previous_task_id = normalize_text(metadata.get("taskId"))
        task = session.get(SyncTaskModel, previous_task_id) if previous_task_id else None
        if task is not None and task.owner_username == owner_username and task.task_type == "product_replace":
            if task.status in {"queued", "running"}:
                raise RuntimeError("商品替换任务已创建，请到同步任务中查看进度。")
            if task.status not in {"preview_ready"}:
                task = None
        else:
            task = None

        target_public = product_detail_to_public(target)
        payload = {
            "targetProductId": target.id,
            "pendingProductId": pending.id,
            "crawlTaskId": getattr(pending, "task_id", None),
            "sourceUrl": pending.source_url,
            "targetSnapshot": target_public,
            "sourcePayload": draft,
            "draftPayload": draft,
            "difference": product_replacement_difference(
                replacement_detail_from_product_public(target_public),
                draft,
            ),
            "pendingProduct": product_to_public(pending),
        }
        if task is None:
            task = SyncTaskModel(
                id=uuid.uuid4().hex,
                owner_username=owner_username,
                store_id=store.id,
                store_name=store.alias_name or store.store_name,
                task_name=f"替换商品 {expected}",
                task_type="product_replace",
                status="queued",
                total_count=1,
                message="等待执行商品替换",
            )
            session.add(task)
        task.payload_json = json.dumps(payload, ensure_ascii=False)
        task.status = "queued"
        task.total_count = 1
        task.success_count = 0
        task.failed_count = 0
        task.message = "等待执行商品替换"
        task.error_detail = None
        task.started_at = None
        task.finished_at = None
        metadata.update({
            "taskId": task.id,
            "targetProductId": target.id,
            "targetManageNumber": expected,
            "targetStoreId": store.id,
            "targetStoreName": store.alias_name or store.store_name,
        })
        raw_payload["_replacement"] = metadata
        pending.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
        session.flush()
        result = product_replacement_to_public(task)
    dispatch_next_sync_task()
    return result


def cancel_product_replacement(owner_username: str, task_id: str) -> dict[str, Any]:
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is None or task.owner_username != owner_username or task.task_type != "product_replace":
            raise RuntimeError("商品替换任务不存在。")
        if task.status not in {"preview_ready", "failed"}:
            raise RuntimeError("当前替换任务不能取消。")
        task.status = "cancelled"
        task.message = "已取消商品替换"
        task.finished_at = datetime.now()
        payload = product_replacement_payload(task)
        pending_product_id = int(payload.get("pendingProductId") or 0)
        pending = session.get(ProductModel, pending_product_id) if pending_product_id else None
        if pending is not None and pending.owner_username == owner_username and pending.review_status == "pending":
            session.delete(pending)
        session.flush()
        return product_replacement_to_public(task)


def perform_product_replacement(
    owner_username: str,
    store_id: int,
    payload: dict[str, Any],
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    raise_if_task_cancelled(SyncTaskModel, task_id)
    target_product_id = int(payload.get("targetProductId") or 0)
    draft = payload.get("draftPayload") if isinstance(payload.get("draftPayload"), dict) else {}
    with session_scope() as session:
        target = session.get(ProductModel, target_product_id)
        store = session.get(StoreModel, store_id)
        if target is None or target.owner_username != owner_username:
            raise RuntimeError("目标店铺商品不存在。")
        if target.review_status != "listed" or target.store_id != store_id:
            raise RuntimeError("目标店铺商品状态或店铺已变化。")
        if store is None or not store.enabled:
            raise RuntimeError("目标店铺不存在或已停用。")
        service_secret = decrypt_text(store.rakuten_service_secret_encrypted)
        license_key = decrypt_text(store.rakuten_license_key_encrypted)
        if not service_secret or not license_key:
            raise RuntimeError("目标店铺缺少乐天 Secret 或乐天 Key。")
        manage_number = normalize_text(target.rakuten_manage_number or target.item_number)
        if not manage_number:
            raise RuntimeError("目标商品缺少商品管理编号。")
        target_identity = {
            "id": target.id,
            "parentProductId": target.parent_product_id,
            "sourceUrl": target.source_url,
            "manageNumber": target.rakuten_manage_number,
            "itemNumber": target.item_number,
            "storeId": target.store_id,
            "reviewStatus": target.review_status,
            "listingStatus": target.rakuten_listing_status,
        }
        raw = json.loads(json.dumps(draft.get("raw") or {}, ensure_ascii=False))
        raw.update({
            "title": normalize_text(draft.get("title")),
            "itemName": normalize_text(draft.get("title")),
            "tagline": normalize_text(draft.get("tagline")),
            "genreId": normalize_text(draft.get("genreId")),
            "images": list(draft.get("images") or []),
            "descriptions": list(draft.get("descriptions") or []),
            "variants": draft.get("variants") if isinstance(draft.get("variants"), dict) else {},
        })
        transient = SimpleNamespace(
            id=target.id,
            owner_username=target.owner_username,
            store_id=target.store_id,
            title=normalize_text(draft.get("title")),
            genre_id=normalize_text(draft.get("genreId")),
            price=Decimal(str(draft.get("price"))) if draft.get("price") is not None else target.price,
            image_url=normalize_text((draft.get("images") or [""])[0]),
            raw_payload_json=json.dumps(raw, ensure_ascii=False),
            rakuten_manage_number=manage_number,
            item_number=target.item_number,
            source_url=target.source_url,
            currency=target.currency,
        )

    uploaded_product_images: list[dict[str, str]] = []
    uploaded_description_images: list[dict[str, str]] = []
    remote_write_started = False
    try:
        update_task_progress(SyncTaskModel, task_id, total_count=1, message="正在上传替换商品图片")
        uploaded_product_images = upload_product_images_to_rakuten(
            service_secret,
            license_key,
            store,
            transient,
            manage_number,
            cabinet_context={},
            cancel_check=lambda: is_task_cancel_requested(SyncTaskModel, task_id),
        )
        raise_if_task_cancelled(SyncTaskModel, task_id)
        description_result = upload_product_description_images_to_rakuten(
            service_secret,
            license_key,
            store,
            transient,
            manage_number,
            raw,
            cabinet_context={},
            cancel_check=lambda: is_task_cancel_requested(SyncTaskModel, task_id),
        )
        raw = description_result["rawPayload"]
        uploaded_description_images = description_result["uploadedImages"]
        update_task_progress(SyncTaskModel, task_id, total_count=1, message="正在更新乐天商品内容")
        item_payload = build_rakuten_item_upsert_payload(
            transient,
            raw,
            uploaded_product_images,
            manage_number=manage_number,
            hide_item=target_identity["listingStatus"] == "unlisted",
        )
        item_payload["itemNumber"] = normalize_text(target_identity["itemNumber"] or manage_number)[:32]
        item_payload = put_rakuten_item_with_attribute_retry(
            service_secret,
            license_key,
            manage_number,
            item_payload,
        )
        remote_write_started = True
        inventory_payloads = build_rakuten_inventory_upsert_payloads(
            manage_number,
            item_payload.get("variants") if isinstance(item_payload.get("variants"), dict) else {},
        )
        bulk_upsert_rakuten_inventories(service_secret, license_key, inventory_payloads)
        patch_rakuten_item_visibility(
            service_secret,
            license_key,
            manage_number,
            hide_item=target_identity["listingStatus"] == "unlisted",
        )
    except Exception as exc:
        if not remote_write_started:
            rollback_uploaded_listing_images(
                service_secret,
                license_key,
                [*uploaded_product_images, *uploaded_description_images],
            )
        else:
            payload["recoveryRequired"] = True
            payload["uploadedImages"] = [*uploaded_product_images, *uploaded_description_images]
        raise RuntimeError(f"商品替换失败：{exc}") from exc

    now = datetime.now()
    updated_raw = dict(raw)
    updated_raw.update(item_payload)
    updated_raw.update({
        "manageNumber": target_identity["manageNumber"] or manage_number,
        "itemNumber": target_identity["itemNumber"] or manage_number,
        "images": uploaded_product_images,
        "descriptionImages": uploaded_description_images,
        "listingStore": product_replacement_payload_store_snapshot(store),
        "updated": now.isoformat(timespec="seconds"),
    })
    image_url = (
        build_rakuten_cabinet_image_url(store.store_code, uploaded_product_images[0]["location"])
        if uploaded_product_images else transient.image_url
    )
    with session_scope() as session:
        target = session.get(ProductModel, target_product_id)
        if target is None or target.owner_username != owner_username:
            raise RuntimeError("乐天更新成功，但本地目标商品不存在，请立即同步店铺。")
        target.title = transient.title
        target.genre_id = transient.genre_id
        target.price = price_from_rakuten_item(updated_raw) or transient.price
        target.image_url = image_url
        target.raw_payload_json = json.dumps(updated_raw, ensure_ascii=False)
        target.store_last_seen_at = now
        target.last_error = None
        parent_product_id = int(target_identity.get("parentProductId") or 0)
        parent = session.get(ProductModel, parent_product_id) if parent_product_id else None
        if (
            parent is not None
            and parent.owner_username == owner_username
            and parent.review_status == "listed_master"
        ):
            remove_product_listed_store_mark(parent, store_id)
        pending_product_id = int(payload.get("pendingProductId") or 0)
        pending = session.get(ProductModel, pending_product_id) if pending_product_id else None
        if pending is None or pending.owner_username != owner_username or pending.review_status != "pending":
            raise RuntimeError("乐天更新成功，但待审核替换商品不存在，请立即同步店铺。")
        master_raw = json.loads(json.dumps(updated_raw, ensure_ascii=False))
        master_raw.pop("_replacement", None)
        for key in ("manageNumber", "itemNumber", "listingStore"):
            master_raw.pop(key, None)
        pending.title = transient.title
        pending.genre_id = transient.genre_id
        pending.price = price_from_rakuten_item(master_raw) or transient.price
        pending.image_url = image_url
        pending.raw_payload_json = json.dumps(master_raw, ensure_ascii=False)
        pending.review_status = "listed_master"
        pending.listing_task_id = None
        pending.listed_at = pending.listed_at or now
        pending.last_error = None
        target.parent_product_id = pending.id
        record_product_listed_store(
            pending,
            target,
            store,
            {
                "manageNumber": target_identity["manageNumber"] or manage_number,
                "itemNumber": target_identity["itemNumber"] or manage_number,
            },
        )
        session.flush()
        updated_product = product_detail_to_public(target)
    payload["result"] = {
        "product": updated_product,
        "preservedIdentity": target_identity,
        "uploadedImageCount": len(uploaded_product_images) + len(uploaded_description_images),
    }
    return {"product": updated_product}


def product_replacement_payload_store_snapshot(store: StoreModel) -> dict[str, Any]:
    return {
        "storeId": store.id,
        "storeCode": store.store_code,
        "storeName": store.store_name,
        "aliasName": store.alias_name,
    }


def perform_product_title_optimization(
    owner_username: str,
    store_id: int,
    product_ids: list[int],
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    from app.services import ai_title_service

    normalized_ids = normalize_product_ids(product_ids)
    success_ids: list[int] = []
    failed_ids: list[int] = []
    errors: list[str] = []
    cancelled = False
    total_count = len(normalized_ids)
    for index, product_id in enumerate(normalized_ids, start=1):
        if task_id and is_task_cancel_requested(SyncTaskModel, task_id):
            cancelled = True
            break
        try:
            with session_scope() as session:
                product = session.get(ProductModel, product_id)
                if (
                    product is None
                    or product.owner_username != owner_username
                    or product.review_status != "listed"
                    or int(product.store_id or 0) != int(store_id)
                ):
                    raise RuntimeError("商品已不存在、已移出店铺商品或店铺已发生变化。")
            generated_version = ai_title_service.generate_version(
                owner_username,
                product_id,
                owner_username,
            )
            version_id = int(generated_version.get("id") or 0)
            if version_id <= 0:
                raise RuntimeError("标题优化已生成，但没有返回可应用的版本。")
            ai_title_service.save_title_version(
                owner_username,
                product_id,
                version_id,
            )
            success_ids.append(product_id)
        except Exception as exc:
            failed_ids.append(product_id)
            errors.append(f"商品 {product_id}：{exc}")
        if task_id:
            update_task_progress(
                SyncTaskModel,
                task_id,
                total_count=total_count,
                success_count=len(success_ids),
                failed_count=len(failed_ids),
                message=f"标题优化中，已处理 {index} / {total_count} 条",
            )
    return {
        "totalCount": total_count,
        "successCount": len(success_ids),
        "failedCount": len(failed_ids),
        "successIds": success_ids,
        "failedIds": failed_ids,
        "errors": errors,
        "cancelled": cancelled,
    }


def run_sync_task(owner_username: str, task_id: str) -> None:
    defer_start = False
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is None:
            return
        if task.owner_username != owner_username:
            raise RuntimeError("不能执行其他用户的同步任务。")
        if task.status == "cancelled":
            return
        if task.status != "queued":
            return
        if task_cancel_requested(task):
            task.status = "cancelled"
            task.message = TASK_CANCELLED_MESSAGE
            task.error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
            task.finished_at = datetime.now()
            return
        task_type = task.task_type or "store_sync"
        wait_reason = (
            specialized_sync_task_wait_reason(session, task)
            if task_type in SPECIALIZED_SYNC_TASK_TYPES
            else sync_task_start_wait_reason(session, task_id, task.store_id)
        )
        if wait_reason:
            task.message = wait_reason
            defer_start = True
        else:
            task.status = "running"
            task.message = sync_task_running_message(task)
            task.error_detail = None
            task.started_at = datetime.now()
            task.finished_at = None
        store_id = task.store_id
        try:
            payload = json.loads(task.payload_json or "{}")
        except ValueError:
            payload = {}

    if defer_start:
        if task_type in SPECIALIZED_SYNC_TASK_TYPES:
            dispatch_sync_task(
                owner_username,
                task_id,
                task_type=task_type,
                delay_seconds=TASK_START_RETRY_DELAY_SECONDS,
            )
        else:
            dispatch_next_sync_task_safely()
        return

    try:
        if store_id is None:
            raise RuntimeError("同步任务没有关联店铺。")
        if task_type == "product_replace":
            result = perform_product_replacement(owner_username, store_id, payload, task_id=task_id)
            total_count = 1
            success_count = 1
            failed_count = 0
            status = "success"
            message = "商品替换完成"
            error_detail = None
            payload["result"] = result
        elif task_type == "title_optimization":
            product_ids = normalize_product_ids(list(payload.get("productIds") or []))
            result = perform_product_title_optimization(
                owner_username,
                store_id,
                product_ids,
                task_id=task_id,
            )
            payload["result"] = {
                "successIds": list(result.get("successIds") or []),
                "failedIds": list(result.get("failedIds") or []),
            }
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("successCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            if result.get("cancelled"):
                status = "cancelled"
                message = cancelled_task_progress_message(
                    "标题优化",
                    total_count,
                    success_count,
                    failed_count,
                )
            else:
                status = "success" if failed_count == 0 else "partial"
                message = f"完成，标题优化成功 {success_count} 条，失败 {failed_count} 条"
            error_detail = summarize_task_errors(list(result.get("errors") or []), limit=50)
        elif task_type == "listing_status":
            listing_status = normalize_text(payload.get("listingStatus"))
            result = perform_store_listing_status_sync(owner_username, store_id, listing_status, task_id=task_id)
            payload["result"] = {
                "successIds": list(result.get("successIds") or []),
                "failedIds": list(result.get("failedIds") or []),
            }
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("successCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            action_label = "上架" if listing_status == "listed" else "下架"
            if result.get("cancelled"):
                status = "cancelled"
                message = cancelled_task_progress_message(action_label, total_count, success_count, failed_count)
            else:
                status = "success" if failed_count == 0 else "partial"
                message = f"完成，{action_label} {success_count} 条，异常 {failed_count} 条"
            error_detail = summarize_task_errors(list(result.get("errors") or []), limit=50)
        elif task_type == "product_listing_status":
            listing_status = normalize_text(payload.get("listingStatus"))
            product_ids = normalize_product_ids(list(payload.get("productIds") or []))
            result = perform_product_listing_status_sync(owner_username, store_id, product_ids, listing_status, task_id=task_id)
            payload["result"] = {
                "successIds": list(result.get("successIds") or []),
                "failedIds": list(result.get("failedIds") or []),
            }
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("successCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            action_label = "上架" if listing_status == "listed" else "下架"
            if result.get("cancelled"):
                status = "cancelled"
                message = cancelled_task_progress_message(action_label, total_count, success_count, failed_count)
            else:
                status = "success" if failed_count == 0 else "partial"
                message = f"完成，{action_label} {success_count} 条，异常 {failed_count} 条"
            error_detail = summarize_task_errors(list(result.get("errors") or []), limit=50)
        elif task_type == "product_delete":
            product_ids = normalize_product_ids(list(payload.get("productIds") or []))
            result = perform_product_delete_sync(owner_username, store_id, product_ids, task_id=task_id)
            payload["result"] = {
                "successIds": list(result.get("successIds") or []),
                "failedIds": list(result.get("failedIds") or []),
            }
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("successCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            cabinet_deleted_count = int(result.get("cabinetDeletedCount") or 0)
            if result.get("cancelled"):
                status = "cancelled"
                message = cancelled_task_progress_message("删除", total_count, success_count, failed_count)
            else:
                status = "success" if failed_count == 0 else "partial"
                message = f"完成，删除 {success_count} 条，异常 {failed_count} 条"
            if cabinet_deleted_count:
                message = f"{message}，同步删除图片 {cabinet_deleted_count} 张"
            error_detail = summarize_task_errors(
                [*list(result.get("errors") or []), *list(result.get("warnings") or [])],
                limit=50,
            )
        elif task_type == "deleted_product_image_cleanup":
            cleanup_ids = normalize_product_ids(list(payload.get("cleanupRecordIds") or []))
            result = perform_deleted_product_image_cleanup(owner_username, store_id, cleanup_ids, task_id=task_id)
            payload["result"] = {
                "successIds": list(result.get("successIds") or []),
                "failedIds": list(result.get("failedIds") or []),
            }
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("successCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            status = "cancelled" if result.get("cancelled") else ("success" if failed_count == 0 else "partial")
            message = (
                cancelled_task_progress_message("图片清理", total_count, success_count, failed_count)
                if result.get("cancelled")
                else f"完成，清理成功 {success_count} 个商品，异常 {failed_count} 个"
            )
            error_detail = summarize_task_errors(list(result.get("errors") or []), limit=50)
        elif task_type == "listing_image_upload":
            product_ids = normalize_product_ids(list(payload.get("productIds") or []))
            result = perform_listing_image_upload(
                owner_username,
                store_id,
                product_ids,
                task_id=task_id,
            )
            payload["result"] = {
                "successIds": list(result.get("successIds") or []),
                "failedIds": list(result.get("failedIds") or []),
            }
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("successCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            if result.get("cancelled"):
                status = "cancelled"
                message = cancelled_task_progress_message(
                    "图片完善",
                    total_count,
                    success_count,
                    failed_count,
                )
            else:
                status = "success" if failed_count == 0 else "partial"
                message = f"完成，图片完善成功 {success_count} 个商品，异常 {failed_count} 个"
            error_detail = summarize_task_errors(list(result.get("errors") or []), limit=50)
        else:
            result = perform_store_sync(owner_username, store_id, task_id=task_id)
            total_count = int(result.get("totalCount") or 0)
            success_count = int(result.get("syncedCount") or 0)
            failed_count = int(result.get("failedCount") or 0)
            payload["result"] = result
            if result.get("cancelled"):
                status = "cancelled"
                message = cancelled_task_progress_message("同步", total_count, success_count, failed_count)
            else:
                status = "success" if failed_count == 0 else "partial"
                message = (
                    f"完成，新增 {int(result.get('addedCount') or 0)} 条，"
                    f"更新 {int(result.get('updatedCount') or 0)} 条，"
                    f"移除 {int(result.get('removedCount') or 0)} 条，"
                    f"未变化 {int(result.get('unchangedCount') or 0)} 条，"
                    f"异常 {failed_count} 条"
                )
            error_detail = summarize_task_errors(list(result.get("errors") or []), limit=50)
    except TaskCancelled:
        with session_scope() as session:
            task = session.get(SyncTaskModel, task_id)
            if task is None:
                return
            total_count = int(task.total_count or 0)
            success_count = int(task.success_count or 0)
            failed_count = int(task.failed_count or 0)
            status = "cancelled"
            message = cancelled_task_progress_message("同步", total_count, success_count, failed_count)
            error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
    except Exception as exc:
        total_count = sync_task_payload_product_count(payload)
        success_count = 0
        failed_count = 1
        status = "failed"
        message = "同步失败"
        error_detail = str(exc)
        with session_scope() as session:
            task = session.get(SyncTaskModel, task_id)
            store = session.get(StoreModel, task.store_id) if task and task.store_id else None
            if task is not None:
                if task.status == "cancelled":
                    return
                total_count = sync_task_known_total_count(task, payload)
                success_count = max(0, int(task.success_count or 0))
                failed_count = max(1, int(task.failed_count or 0))
                if success_count > 0:
                    status = "partial"
                    unfinished_count = max(
                        0,
                        total_count - success_count - failed_count,
                    )
                    message = failed_sync_task_progress_message(
                        sync_task_action_label(task),
                        total_count,
                        success_count,
                        failed_count,
                        unfinished_count=unfinished_count,
                    )
            if store is not None:
                store.last_error = error_detail

    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is not None:
            if task.status == "cancelled" and status != "cancelled":
                return
            task.total_count = total_count
            task.success_count = success_count
            task.failed_count = failed_count
            task.status = status
            task.message = message
            task.error_detail = error_detail
            task.payload_json = json.dumps(payload, ensure_ascii=False)
            task.finished_at = datetime.now()
    dispatch_next_sync_task_safely()
    dispatch_next_listing_task_safely()


def sync_task_running_message(task: SyncTaskModel) -> str:
    task_type = task.task_type or "store_sync"
    if task_type == "title_optimization":
        return "正在执行批量标题优化"
    if task_type == "product_replace":
        return "正在替换店铺商品"
    if task_type in {"listing_status", "product_listing_status"}:
        try:
            payload = json.loads(task.payload_json or "{}")
        except ValueError:
            payload = {}
        action_label = "上架" if normalize_text(payload.get("listingStatus")) == "listed" else "下架"
        return f"正在执行{'全部' if task_type == 'listing_status' else '批量'}{action_label}"
    if task_type == "product_delete":
        return "正在执行批量删除"
    if task_type == "deleted_product_image_cleanup":
        return "正在清理已删除商品图片"
    if task_type == "listing_image_upload":
        return "正在上传上架商品图片和详情图"
    return "正在同步店铺商品"


def cancelled_task_progress_message(action_label: str, total_count: int, success_count: int, failed_count: int) -> str:
    total = max(0, int(total_count or 0))
    success = max(0, int(success_count or 0))
    failed = max(0, int(failed_count or 0))
    processed = min(total, success + failed) if total else success + failed
    return f"已终止，{action_label}已处理 {processed} / {total} 条，成功 {success} 条，失败 {failed} 条"


def perform_store_listing_status_sync(
    owner_username: str,
    store_id: int,
    listing_status: str,
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    raise_if_task_cancelled(SyncTaskModel, task_id)
    if listing_status not in {"listed", "unlisted"}:
        raise RuntimeError("上架状态不合法。")
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None:
            raise RuntimeError("店铺不存在。")
        if not store.enabled:
            raise RuntimeError("店铺已停用，不能更新上架状态。")
        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == store_id,
                ProductModel.review_status == "listed",
            )
        ).all()
        if not products:
            raise RuntimeError("当前店铺没有可操作的店铺商品。")
        result = apply_products_listing_status(
            session,
            products,
            listing_status,
            progress_callback=sync_task_progress_callback(task_id, len(products), "上下架同步") if task_id else None,
            cancel_check=(lambda: is_task_cancel_requested(SyncTaskModel, task_id)) if task_id else None,
        )
        session.flush()
        summary = listing_status_result_summary(result, len(products), cancelled=bool(result.get("cancelled")))
        return {
            "store": store_to_public(store),
            "totalCount": summary["total"],
            "successCount": summary["successCount"],
            "failedCount": summary["failedCount"],
            "successIds": summary["successIds"],
            "failedIds": summary["failedIds"],
            "errors": summary["errors"],
            "cancelled": bool(result.get("cancelled")),
        }


def perform_product_listing_status_sync(
    owner_username: str,
    store_id: int,
    product_ids: list[int],
    listing_status: str,
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    raise_if_task_cancelled(SyncTaskModel, task_id)
    if listing_status not in {"listed", "unlisted"}:
        raise RuntimeError("上架状态不合法。")
    if not product_ids:
        raise RuntimeError("同步任务缺少商品。")
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None:
            raise RuntimeError("店铺不存在。")
        if not store.enabled:
            raise RuntimeError("店铺已停用，不能更新上架状态。")
        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == store_id,
                ProductModel.id.in_(product_ids),
                ProductModel.review_status == "listed",
            )
        ).all()
        found_ids = {product.id for product in products}
        missing_ids = [product_id for product_id in product_ids if product_id not in found_ids]
        if task_id:
            update_task_progress(
                SyncTaskModel,
                task_id,
                total_count=len(product_ids),
                success_count=0,
                failed_count=len(missing_ids),
                message=f"上下架同步中，已处理 0 / {len(product_ids)} 条",
            )
        result = (
            apply_products_listing_status(
                session,
                products,
                listing_status,
                progress_callback=sync_task_progress_callback(task_id, len(product_ids), "上下架同步", initial_failed=len(missing_ids)) if task_id else None,
                cancel_check=(lambda: is_task_cancel_requested(SyncTaskModel, task_id)) if task_id else None,
            )
            if products
            else {"successIds": [], "errors": []}
        )
        errors = list(result.get("errors") or [])
        errors.extend(f"{product_id}: 商品不存在或不是店铺商品" for product_id in missing_ids)
        session.flush()
        success_ids = list(result.get("successIds") or [])
        failed_ids = [*list(result.get("failedIds") or []), *missing_ids]
        failed_ids = list(dict.fromkeys(failed_ids))
        success_count = len(success_ids)
        failed_count = len(failed_ids) if result.get("cancelled") else max(0, len(product_ids) - success_count)
        return {
            "store": store_to_public(store),
            "totalCount": len(product_ids),
            "successCount": success_count,
            "failedCount": failed_count,
            "successIds": success_ids,
            "failedIds": failed_ids,
            "errors": errors,
            "cancelled": bool(result.get("cancelled")),
        }


def perform_product_delete_sync(
    owner_username: str,
    store_id: int,
    product_ids: list[int],
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    raise_if_task_cancelled(SyncTaskModel, task_id)
    if not product_ids:
        raise RuntimeError("同步任务缺少商品。")
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None:
            raise RuntimeError("店铺不存在。")
        if not store.enabled:
            raise RuntimeError("店铺已停用，不能删除商品。")
        rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.store_id == store_id,
                ProductModel.id.in_(product_ids),
                ProductModel.review_status == "listed",
            )
        ).all()
        found_ids = {row.id for row in rows}
        missing_ids = [product_id for product_id in product_ids if product_id not in found_ids]
        success_ids: list[int] = []
        failed_ids: list[int] = list(missing_ids)
        success_count = 0
        failed_count = len(missing_ids)
        cabinet_deleted_count = 0
        errors = [f"{product_id}: 商品不存在或不是店铺商品" for product_id in missing_ids]
        warnings: list[str] = []
        cancelled = False
        credential_cache: dict[int, tuple[StoreModel, str, str]] = {}
        if task_id:
            update_task_progress(
                SyncTaskModel,
                task_id,
                total_count=len(product_ids),
                success_count=0,
                failed_count=failed_count,
                message=f"删除中，已处理 0 / {len(product_ids)} 条",
            )
        for index, row in enumerate(rows, start=1):
            if task_id and is_task_cancel_requested(SyncTaskModel, task_id):
                cancelled = True
                break
            try:
                delete_store_product_from_rakuten(session, row, credential_cache)
                cabinet_deleted_count += int(getattr(row, "_deleted_cabinet_count", 0) or 0)
            except Exception as exc:
                failed_count += 1
                failed_ids.append(row.id)
                error_text = str(exc)
                row.last_error = error_text
                errors.append(f"{productCodeForError(row)}: {error_text}")
                if task_id:
                    update_task_progress(
                        SyncTaskModel,
                        task_id,
                        total_count=len(product_ids),
                        success_count=success_count,
                        failed_count=failed_count,
                        message=f"删除中，已处理 {index + len(missing_ids)} / {len(product_ids)} 条",
                    )
                continue
            warnings.extend(getattr(row, "_delete_warnings", []) or [])
            remove_listed_store_mark_for_store_product(session, row)
            success_ids.append(row.id)
            session.delete(row)
            success_count += 1
            if task_id:
                update_task_progress(
                    SyncTaskModel,
                    task_id,
                    total_count=len(product_ids),
                    success_count=success_count,
                    failed_count=failed_count,
                    message=f"删除中，已处理 {index + len(missing_ids)} / {len(product_ids)} 条",
                )
        session.flush()
        result = {
            "store": store_to_public(store),
            "totalCount": len(product_ids),
            "successCount": success_count,
            "failedCount": failed_count,
            "successIds": success_ids,
            "failedIds": failed_ids,
            "cabinetDeletedCount": cabinet_deleted_count,
            "errors": errors,
            "warnings": warnings,
            "cancelled": cancelled,
        }
    return result


def retry_sync_task(
    owner_username: str,
    task_id: str,
    *,
    allow_all_owners: bool = False,
) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        if task is None:
            raise RuntimeError("同步任务不存在。")
        if task.owner_username != owner_username and not allow_all_owners:
            raise RuntimeError("不能重试其他用户的同步任务。")
        task_owner = task.owner_username
        task_type = task.task_type
        if task.status in {"queued", "running"}:
            raise RuntimeError("同步任务正在执行中，不能重试。")
        if task.status == "success":
            raise RuntimeError("成功的同步任务不需要重试。")
        payload = sync_task_payload(task)
        result_payload = payload.get("result") if isinstance(payload.get("result"), dict) else {}
        if task.task_type in {
            "product_delete",
            "product_listing_status",
            "title_optimization",
            "listing_image_upload",
        }:
            original_ids = normalize_product_ids(list(payload.get("productIds") or []))
            failed_ids = normalize_product_ids(list(result_payload.get("failedIds") or []))
            success_ids = set(normalize_product_ids(list(result_payload.get("successIds") or [])))
            retry_ids = normalize_product_ids(
                [
                    *failed_ids,
                    *[
                        product_id
                        for product_id in original_ids
                        if product_id not in success_ids
                    ],
                ]
            )
            if not retry_ids:
                raise RuntimeError("该同步任务没有可重试的失败商品。")
            payload["productIds"] = retry_ids
            task.total_count = len(retry_ids)
        elif task.task_type == "deleted_product_image_cleanup":
            original_ids = normalize_product_ids(list(payload.get("cleanupRecordIds") or []))
            failed_ids = normalize_product_ids(list(result_payload.get("failedIds") or []))
            success_ids = set(normalize_product_ids(list(result_payload.get("successIds") or [])))
            retry_ids = normalize_product_ids(
                [
                    *failed_ids,
                    *[
                        record_id
                        for record_id in original_ids
                        if record_id not in success_ids
                    ],
                ]
            )
            if not retry_ids:
                raise RuntimeError("该同步任务没有可重试的失败记录。")
            payload["cleanupRecordIds"] = retry_ids
            task.total_count = len(retry_ids)
        else:
            task.total_count = 0
        payload.pop("result", None)
        task.status = "queued"
        task.message = "等待重新同步"
        task.error_detail = None
        task.payload_json = json.dumps(payload, ensure_ascii=False)
        task.success_count = 0
        task.failed_count = 0
        task.started_at = None
        task.finished_at = None
    if task_type in SPECIALIZED_SYNC_TASK_TYPES:
        dispatch_sync_task(
            task_owner,
            task_id,
            task_type=task_type,
        )
    else:
        dispatch_next_sync_task()
    with session_scope() as session:
        task = session.get(SyncTaskModel, task_id)
        return sync_task_to_public(task) if task else {"id": task_id}


def verify_all_stores(owner_username: str) -> dict[str, Any]:
    with session_scope() as session:
        rows = session.scalars(
            select(StoreModel)
            .where(StoreModel.owner_username == owner_username)
            .order_by(StoreModel.id.asc())
        ).all()
        for row in rows:
            try:
                verify_store_credentials(row, include_product_counts=False, include_cabinet_usage=False)
            except Exception as exc:
                row.last_checked_at = datetime.now()
                row.last_error = str(exc)
        session.flush()
        stores = [store_to_public(row) for row in rows]
        return {
            "stores": stores,
            "summary": {
                "total": len(stores),
                "available": sum(1 for store in stores if store["availabilityStatus"] == "available"),
                "error": sum(1 for store in stores if store["availabilityStatus"] == "error"),
                "unchecked": sum(1 for store in stores if store["availabilityStatus"] == "unchecked"),
            },
    }


def verify_store(
    owner_username: str,
    store_id: int,
    *,
    reveal: bool = False,
) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None:
            raise RuntimeError("店铺不存在。")
        if row.owner_username != owner_username:
            raise RuntimeError("不能检测其他用户的店铺。")
        try:
            verify_store_credentials(row, include_product_counts=False, include_cabinet_usage=False)
        except Exception as exc:
            row.last_checked_at = datetime.now()
            row.last_error = str(exc)
        session.flush()
        return store_to_public(row, reveal=reveal)


def refresh_all_store_product_counts(owner_username: str) -> dict[str, Any]:
    with session_scope() as session:
        rows = session.scalars(
            select(StoreModel)
            .where(StoreModel.owner_username == owner_username)
            .order_by(StoreModel.id.asc())
        ).all()
        for row in rows:
            try:
                verify_store_credentials(row, include_product_counts=True)
            except Exception as exc:
                row.last_checked_at = datetime.now()
                row.last_error = str(exc)
        session.flush()
        stores = [store_to_public(row) for row in rows]
        return {
            "stores": stores,
            "summary": {
                "total": len(stores),
                "available": sum(1 for store in stores if store["availabilityStatus"] == "available"),
                "error": sum(1 for store in stores if store["availabilityStatus"] == "error"),
                "unchecked": sum(1 for store in stores if store["availabilityStatus"] == "unchecked"),
            },
        }


def refresh_store_product_counts(owner_username: str, store_id: int) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(StoreModel, store_id)
        if row is None:
            raise RuntimeError("店铺不存在。")
        if row.owner_username != owner_username:
            raise RuntimeError("不能获取其他用户店铺的数量。")
        try:
            verify_store_credentials(row, include_product_counts=True)
        except Exception as exc:
            row.last_checked_at = datetime.now()
            row.last_error = str(exc)
        session.flush()
        return store_to_public(row)


def manual_crawl_import_template_bytes() -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法生成导入模板。") from exc

    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = "单个商品采集"
    product_sheet.append(["商品URL", "备注"])
    product_sheet.append(["", ""])
    product_sheet["A1"].comment = Comment(
        "每行填写一个乐天商品 URL。所有有效 URL 会合并创建为一个手动采集任务。",
        "商品采集系统",
    )
    product_sheet.column_dimensions["A"].width = 72
    product_sheet.column_dimensions["B"].width = 32

    shop_sheet = workbook.create_sheet("店铺采集")
    shop_sheet.append(["店铺名称或URL", "榜单时间", "采集数量", "备注"])
    shop_sheet.append(["", "日榜", "全部", ""])
    shop_sheet["A1"].comment = Comment(
        "每行填写一个店铺展示名称、URL 代码、完整 URL 或 SID。每个有效店铺创建一个任务。",
        "商品采集系统",
    )
    shop_sheet.column_dimensions["A"].width = 52
    shop_sheet.column_dimensions["B"].width = 18
    shop_sheet.column_dimensions["C"].width = 18
    shop_sheet.column_dimensions["D"].width = 32

    period_validation = DataValidation(type="list", formula1='"日榜,周榜,月榜"', allow_blank=False)
    shop_sheet.add_data_validation(period_validation)
    period_validation.add("B2:B1000")
    count_validation = DataValidation(type="list", formula1='"全部,30,50,100"', allow_blank=False)
    shop_sheet.add_data_validation(count_validation)
    count_validation.add("C2:C1000")

    header_fill = PatternFill("solid", fgColor="E9EEF5")
    example_fill = PatternFill("solid", fgColor="FFF7D6")
    input_font = Font(name="Arial", color="0000FF")
    header_font = Font(name="Arial", bold=True, color="1F2937")
    warning_fill = PatternFill("solid", fgColor="FDECEC")
    for sheet in (product_sheet, shop_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 24
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in sheet[2]:
            cell.fill = example_fill
            cell.font = input_font
        sheet.conditional_formatting.add(
            f"A2:A1000",
            FormulaRule(formula=["LEN(TRIM(A2))=0"], fill=warning_fill),
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_manual_crawl_tasks(owner_username: str, filename: str, content: bytes) -> dict[str, Any]:
    normalized_filename = normalize_text(filename).lower()
    if not content:
        raise RuntimeError("导入文件为空。")
    if not normalized_filename.endswith((".xlsx", ".xlsm")):
        raise RuntimeError("手动采集导入只支持 .xlsx 文件。")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法读取 xlsx 文件。") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    product_rows = manual_product_import_rows(workbook)
    shop_rows = manual_shop_import_rows(workbook)
    if not product_rows and not shop_rows:
        raise RuntimeError("导入文件没有可创建的手动采集数据。")

    failed_rows: list[dict[str, Any]] = []
    normalized_product_urls: list[str] = []
    seen_product_urls: set[str] = set()
    for row in product_rows:
        row_number = int(row["rowNumber"])
        raw_url = normalize_text(row["target"])
        try:
            url = normalize_rakuten_product_target(raw_url)
        except RuntimeError as exc:
            failed_rows.append({"sheet": "单个商品采集", "rowNumber": row_number, "message": str(exc)})
            continue
        if url in seen_product_urls:
            continue
        seen_product_urls.add(url)
        normalized_product_urls.append(url)

    product_task: dict[str, Any] | None = None
    if normalized_product_urls:
        product_task = create_task(
            owner_username,
            SimpleNamespace(
                sourceId=None,
                sourceType="product_url",
                target="\n".join(normalized_product_urls),
                mode="manual",
                rankingPeriod=None,
                crawlLimit=None,
            ),
        )

    shop_tasks: list[dict[str, Any]] = []
    for row in shop_rows:
        row_number = int(row["rowNumber"])
        raw_target = normalize_text(row["target"])
        try:
            normalized_target = normalize_rakuten_shop_target(raw_target)
            if not normalized_target:
                raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR)
            task = create_task(
                owner_username,
                SimpleNamespace(
                    sourceId=None,
                    sourceType="shop",
                    target=normalized_target,
                    mode="manual",
                    rankingPeriod=manual_import_ranking_period(row.get("rankingPeriod")),
                    crawlLimit=manual_import_crawl_limit(row.get("crawlLimit")),
                ),
            )
            shop_tasks.append(task)
        except RuntimeError as exc:
            failed_rows.append({"sheet": "店铺采集", "rowNumber": row_number, "message": str(exc)})

    return {
        "productUrlCount": len(normalized_product_urls),
        "productTaskCreated": product_task is not None,
        "shopTaskCount": len(shop_tasks),
        "createdTaskCount": (1 if product_task else 0) + len(shop_tasks),
        "failedCount": len(failed_rows),
        "failedRows": failed_rows,
    }


def manual_product_import_rows(workbook: Any) -> list[dict[str, Any]]:
    if "单个商品采集" not in workbook.sheetnames:
        return []
    values = list(workbook["单个商品采集"].iter_rows(values_only=True))
    return manual_import_rows_from_table(
        values,
        sheet_name="单个商品采集",
        required_headers={"target": {"商品url", "商品链接", "url"}},
    )


def manual_shop_import_rows(workbook: Any) -> list[dict[str, Any]]:
    if "店铺采集" not in workbook.sheetnames:
        return []
    values = list(workbook["店铺采集"].iter_rows(values_only=True))
    return manual_import_rows_from_table(
        values,
        sheet_name="店铺采集",
        required_headers={
            "target": {"店铺名称或url", "店铺", "店铺url", "店铺链接"},
            "rankingPeriod": {"榜单时间", "榜单", "排行时间"},
            "crawlLimit": {"采集数量", "数量"},
        },
    )


def manual_import_rows_from_table(
    values: list[Any],
    *,
    sheet_name: str,
    required_headers: dict[str, set[str]],
) -> list[dict[str, Any]]:
    if not values:
        return []
    normalized_header = [re.sub(r"\s+", "", normalize_text(value)).lower() for value in values[0]]
    indexes: dict[str, int] = {}
    for key, aliases in required_headers.items():
        index = next((position for position, value in enumerate(normalized_header) if value in aliases), None)
        if index is None:
            required = "、".join(next(iter(aliases)) for aliases in required_headers.values())
            raise RuntimeError(f"工作表“{sheet_name}”表头必须包含：{required}。")
        indexes[key] = index
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(values[1:], start=2):
        row_values = list(row or [])
        parsed = {
            key: import_cell_text(row_values[index] if index < len(row_values) else "")
            for key, index in indexes.items()
        }
        if not normalize_text(parsed.get("target")):
            continue
        parsed["rowNumber"] = row_number
        rows.append(parsed)
    return rows


def manual_import_ranking_period(value: Any) -> str:
    normalized = normalize_text(value).lower()
    mapping = {
        "日榜": "daily",
        "daily": "daily",
        "周榜": "weekly",
        "weekly": "weekly",
        "月榜": "monthly",
        "monthly": "monthly",
    }
    if normalized not in mapping:
        raise RuntimeError("榜单时间只能填写：日榜、周榜、月榜。")
    return mapping[normalized]


def manual_import_crawl_limit(value: Any) -> str | int:
    normalized = normalize_text(value)
    if normalized in {"全部", "all"}:
        return "all"
    try:
        count = int(float(normalized))
    except (TypeError, ValueError) as exc:
        raise RuntimeError("采集数量必须填写“全部”或大于 0 的整数。") from exc
    if count <= 0:
        raise RuntimeError("采集数量必须填写“全部”或大于 0 的整数。")
    return count


def scheduled_crawl_import_template_bytes() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法生成导入模板。") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "定时采集导入"
    sheet.append(["店铺名称", "店铺URL"])
    sheet.append(["示例店铺名称", "https://www.rakuten.co.jp/example-shop/"])
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 52
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def scheduled_crawl_export_bytes(
    owner_username: str,
    *,
    keyword: str | None = None,
    enabled_status: str | None = None,
    status: str | None = None,
    schedule_time: str | None = None,
    created_at_from: str | None = None,
    created_at_to: str | None = None,
) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法生成导出表格。") from exc

    with session_scope() as session:
        query = scheduled_crawls_query(
            owner_username,
            keyword=keyword,
            enabled_status=enabled_status,
            status=status,
            schedule_time=schedule_time,
            created_at_from=created_at_from,
            created_at_to=created_at_to,
        )
        rows = session.scalars(query.order_by(ScheduledCrawlModel.created_at.desc())).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采集店铺导出"
    sheet.append(["店铺名称", "店铺URL"])
    for row in rows:
        sheet.append([scheduled_crawl_export_shop_name(row), scheduled_crawl_export_shop_url(row)])
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 52
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_scheduled_crawls(owner_username: str, filename: str, content: bytes) -> dict[str, Any]:
    rows = parse_scheduled_crawl_import_rows(filename, content)
    created_count = 0
    updated_count = 0
    failed_rows: list[dict[str, Any]] = []
    imported_rows: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    with session_scope() as session:
        for row in rows:
            row_number = int(row.get("rowNumber") or 0)
            shop_name = normalize_text(row.get("shopName"))
            shop_url = normalize_text(row.get("shopUrl"))
            if not shop_name or not shop_url:
                failed_rows.append({"rowNumber": row_number, "message": "店铺名称和店铺URL不能为空。"})
                continue
            if shop_name in seen_names:
                failed_rows.append({"rowNumber": row_number, "message": f"表格中店铺名称重复：{shop_name}"})
                continue
            seen_names.add(shop_name)
            try:
                normalize_rakuten_shop_target(shop_url)
            except RuntimeError as exc:
                failed_rows.append({"rowNumber": row_number, "message": str(exc)})
                continue
            existing = session.scalar(
                select(ScheduledCrawlModel).where(
                    ScheduledCrawlModel.owner_username == owner_username,
                    ScheduledCrawlModel.source_type == "shop",
                    ScheduledCrawlModel.crawl_content == shop_name,
                )
            )
            is_new = existing is None
            schedule = existing or ScheduledCrawlModel(owner_username=owner_username)
            if is_new:
                session.add(schedule)
            apply_imported_scheduled_crawl(schedule, owner_username, shop_name, shop_url)
            imported_rows.append(
                {
                    "rowNumber": row_number,
                    "shopName": shop_name,
                    "shopUrl": shop_url,
                    "action": "created" if is_new else "updated",
                }
            )
            if is_new:
                created_count += 1
            else:
                updated_count += 1
        session.flush()
    return {
        "totalRows": len(rows),
        "createdCount": created_count,
        "updatedCount": updated_count,
        "failedCount": len(failed_rows),
        "failedRows": failed_rows,
        "importedRows": imported_rows,
    }


def apply_imported_scheduled_crawl(
    row: ScheduledCrawlModel,
    owner_username: str,
    shop_name: str,
    shop_url: str,
) -> None:
    normalized_shop_name = normalize_text(shop_name)
    normalized_shop_url = normalize_text(shop_url)
    row.owner_username = owner_username
    row.source_id = None
    row.source_type = "shop"
    row.name = f"{normalized_shop_name} 每日20:00定时采集"
    row.crawl_content = normalized_shop_name
    row.crawl_condition = "店铺采集；名称失败时使用店铺URL"
    row.target = default_imported_schedule_target(normalized_shop_name)
    row.enabled = True
    row.interval_minutes = 1440
    row.schedule_time = "20:00"
    row.notes = schedule_import_notes(normalized_shop_name, normalized_shop_url)
    row.status = "idle"
    row.next_run_at = next_daily_run_at(row.schedule_time)


def parse_scheduled_crawl_import_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    normalized_filename = normalize_text(filename).lower()
    if not content:
        raise RuntimeError("导入文件为空。")
    if normalized_filename.endswith((".xlsx", ".xlsm")):
        return parse_scheduled_crawl_xlsx_rows(content)
    if normalized_filename.endswith(".xls"):
        return parse_scheduled_crawl_xls_rows(content)
    raise RuntimeError("只支持导入 .xls 或 .xlsx 文件。")


def parse_scheduled_crawl_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法读取 xlsx 文件。") from exc
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    return scheduled_crawl_rows_from_table(values)


def parse_scheduled_crawl_xls_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("服务器缺少 xlrd，无法读取 xls 文件。") from exc
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    values = [[sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)] for row_index in range(sheet.nrows)]
    return scheduled_crawl_rows_from_table(values)


def scheduled_crawl_rows_from_table(values: list[Any]) -> list[dict[str, Any]]:
    if not values:
        raise RuntimeError("导入文件没有内容。")
    header = [schedule_import_header_key(value) for value in values[0]]
    try:
        name_index = header.index("shopName")
        url_index = header.index("shopUrl")
    except ValueError as exc:
        raise RuntimeError("导入文件表头必须包含：店铺名称、店铺URL。") from exc
    rows: list[dict[str, Any]] = []
    for row_offset, row in enumerate(values[1:], start=2):
        row_values = list(row or [])
        shop_name = import_cell_text(row_values[name_index] if name_index < len(row_values) else "")
        shop_url = import_cell_text(row_values[url_index] if url_index < len(row_values) else "")
        if not shop_name and not shop_url:
            continue
        rows.append({"rowNumber": row_offset, "shopName": shop_name, "shopUrl": shop_url})
    if not rows:
        raise RuntimeError("导入文件没有可导入的店铺数据。")
    return rows


def schedule_import_header_key(value: Any) -> str:
    normalized = re.sub(r"\s+", "", normalize_text(value)).lower()
    if normalized in {"店铺名称", "店铺名", "名称", "shopname", "storename"}:
        return "shopName"
    if normalized in {"店铺url", "店铺链接", "店铺网址", "url", "shopurl", "storeurl"}:
        return "shopUrl"
    return normalized


def import_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def schedule_import_shop_name(notes: Any) -> str:
    text = normalize_text(notes)
    if not text:
        return ""
    try:
        payload = json.loads(text)
    except ValueError:
        payload = {}
    if isinstance(payload, dict):
        return normalize_text(payload.get("shopName"))
    return ""


def scheduled_crawl_export_shop_name(row: ScheduledCrawlModel) -> str:
    return (
        schedule_import_shop_name(row.notes)
        or normalize_text(row.crawl_content)
        or scheduled_crawl_target_value(row)
        or normalize_text(row.name)
    )


def scheduled_crawl_export_shop_url(row: ScheduledCrawlModel) -> str:
    fallback_url = schedule_fallback_shop_url(row.notes)
    if fallback_url:
        return fallback_url
    for value in (row.crawl_content, row.target):
        normalized = exportable_shop_target_value(value)
        if normalized:
            return normalized
    return ""


def scheduled_crawl_target_value(row: ScheduledCrawlModel) -> str:
    return exportable_shop_target_value(row.target, build_shop_url=False)


def exportable_shop_target_value(value: Any, *, build_shop_url: bool = True) -> str:
    base_target, fallback_target = split_shop_fallback_target(normalize_text(value))
    if fallback_target:
        return fallback_target
    parsed_target, _, _ = parse_ranking_target(strip_shop_ranking_prefix(base_target))
    normalized = normalize_text(parsed_target or base_target)
    if not normalized:
        return ""
    if normalized.startswith(("http://", "https://")):
        return normalized
    if re.fullmatch(r"[0-9]+", normalized):
        return normalized
    if build_shop_url and looks_like_rakuten_shop_code(normalized):
        return build_rakuten_store_url(normalized)
    return normalized if not build_shop_url else ""


def scheduled_crawls_query(
    owner_username: str,
    *,
    keyword: str | None = None,
    enabled_status: str | None = None,
    status: str | None = None,
    schedule_time: str | None = None,
    created_at_from: str | None = None,
    created_at_to: str | None = None,
) -> Any:
    query = select(ScheduledCrawlModel).where(
        ScheduledCrawlModel.owner_username == owner_username,
        ScheduledCrawlModel.source_type == "shop",
    )
    created_at_from_value = parse_datetime_filter(created_at_from)
    created_at_to_value = parse_datetime_filter(created_at_to)
    normalized_keyword = normalize_text(keyword)
    if normalized_keyword:
        like_value = f"%{normalized_keyword}%"
        query = query.where(
            or_(
                ScheduledCrawlModel.name.like(like_value),
                ScheduledCrawlModel.crawl_content.like(like_value),
                ScheduledCrawlModel.target.like(like_value),
                ScheduledCrawlModel.notes.like(like_value),
            )
        )
    if enabled_status == "enabled":
        query = query.where(ScheduledCrawlModel.enabled.is_(True))
    elif enabled_status == "disabled":
        query = query.where(ScheduledCrawlModel.enabled.is_(False))
    if status in {"idle", "running", "disabled", "failed"}:
        query = query.where(ScheduledCrawlModel.status == status)
    normalized_schedule_time = normalize_text(schedule_time)
    if normalized_schedule_time:
        query = query.where(ScheduledCrawlModel.schedule_time == normalized_schedule_time)
    if created_at_from_value is not None:
        query = query.where(ScheduledCrawlModel.created_at >= created_at_from_value)
    if created_at_to_value is not None:
        query = query.where(ScheduledCrawlModel.created_at <= created_at_to_value)
    return query


def list_scheduled_crawls(
    owner_username: str,
    *,
    page: int | None = None,
    page_size: int | None = None,
    keyword: str | None = None,
    enabled_status: str | None = None,
    status: str | None = None,
    schedule_time: str | None = None,
    created_at_from: str | None = None,
    created_at_to: str | None = None,
) -> list[dict[str, Any]] | dict[str, Any]:
    with session_scope() as session:
        reconcile_interrupted_scheduled_crawls(session, owner_username=owner_username)
        query = scheduled_crawls_query(
            owner_username,
            keyword=keyword,
            enabled_status=enabled_status,
            status=status,
            schedule_time=schedule_time,
            created_at_from=created_at_from,
            created_at_to=created_at_to,
        )
        return paginate_query(
            session,
            query,
            order_by=task_status_order_by(
                ScheduledCrawlModel,
                disabled_condition=ScheduledCrawlModel.enabled.is_(False),
            ),
            page=page,
            page_size=page_size,
            response_key="schedules",
            serializer=scheduled_crawl_to_public,
        )


def save_scheduled_crawl(owner_username: str, payload: Any, schedule_id: int | None = None) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(ScheduledCrawlModel, schedule_id) if schedule_id else None
        if row is None:
            row = ScheduledCrawlModel(owner_username=owner_username)
            session.add(row)
        if row.owner_username != owner_username:
            raise RuntimeError("不能修改其他用户的定时任务。")

        raw_target = str(getattr(payload, "target", "") or "").strip()
        parsed_target, existing_limit, _ = parse_ranking_target(strip_shop_ranking_prefix(raw_target))
        normalized_target = normalize_rakuten_shop_target(parsed_target)
        if not normalized_target:
            raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR)
        schedule_time = normalize_schedule_time(getattr(payload, "scheduleTime", "09:00"))
        period_label = ranking_period_label(getattr(payload, "rankingPeriod", "daily"))
        limit_label = crawl_limit_label(
            getattr(payload, "crawlLimit", None),
            default="全部" if existing_limit is None else f"前 {existing_limit}",
        )

        row.source_id = None
        row.source_type = "shop"
        row.target = f"店铺:{normalized_target} {period_label} {limit_label}"
        row.name = str(getattr(payload, "name", "") or "").strip() or f"{normalized_target} 每日定时采集"
        row.crawl_content = normalized_target
        row.crawl_condition = "店铺采集"
        row.enabled = bool(getattr(payload, "enabled", True))
        row.interval_minutes = 1440
        row.schedule_time = schedule_time
        row.notes = str(getattr(payload, "notes", "") or "").strip()
        row.status = "idle" if row.enabled else "disabled"
        row.next_run_at = next_daily_run_at(row.schedule_time) if row.enabled else None
        if not row.name or not row.target:
            raise RuntimeError("定时任务名称和采集目标不能为空。")
        session.flush()
        return scheduled_crawl_to_public(row)


def delete_scheduled_crawl(
    owner_username: str,
    schedule_id: int,
    *,
    delete_collected_products: bool = False,
) -> dict[str, Any]:
    return delete_scheduled_crawls(
        owner_username,
        [schedule_id],
        delete_collected_products=delete_collected_products,
    )


def update_scheduled_crawl_statuses(
    owner_username: str,
    schedule_ids: list[int],
    enabled: bool,
) -> dict[str, Any]:
    normalized_ids: list[int] = []
    seen: set[int] = set()
    for value in schedule_ids or []:
        try:
            schedule_id = int(value)
        except (TypeError, ValueError):
            continue
        if schedule_id <= 0 or schedule_id in seen:
            continue
        seen.add(schedule_id)
        normalized_ids.append(schedule_id)
    if not normalized_ids:
        raise RuntimeError("请选择要启用或停用的采集店铺。")

    with session_scope() as session:
        rows = session.scalars(
            select(ScheduledCrawlModel).where(
                ScheduledCrawlModel.owner_username == owner_username,
                ScheduledCrawlModel.id.in_(normalized_ids),
            )
        ).all()
        found_ids = {int(row.id) for row in rows}
        for row in rows:
            row.enabled = bool(enabled)
            row.status = "idle" if enabled else "disabled"
            row.next_run_at = next_daily_run_at(row.schedule_time) if enabled else None
        return {
            "updatedIds": sorted(found_ids),
            "failedIds": [
                schedule_id
                for schedule_id in normalized_ids
                if schedule_id not in found_ids
            ],
            "updatedCount": len(found_ids),
            "enabled": bool(enabled),
        }


def update_all_scheduled_crawl_statuses(
    owner_username: str,
    enabled: bool,
) -> dict[str, Any]:
    target_enabled = bool(enabled)
    with session_scope() as session:
        rows = session.scalars(
            select(ScheduledCrawlModel).where(
                ScheduledCrawlModel.owner_username == owner_username,
                ScheduledCrawlModel.source_type == "shop",
            )
        ).all()
        updated_count = 0
        for row in rows:
            if bool(row.enabled) == target_enabled:
                continue
            row.enabled = target_enabled
            row.status = "idle" if target_enabled else "disabled"
            row.next_run_at = next_daily_run_at(row.schedule_time) if target_enabled else None
            updated_count += 1
        return {
            "matchedCount": len(rows),
            "updatedCount": updated_count,
            "enabled": target_enabled,
        }


def delete_scheduled_crawls(
    owner_username: str,
    schedule_ids: list[int],
    *,
    delete_collected_products: bool = False,
) -> dict[str, Any]:
    normalized_ids: list[int] = []
    seen: set[int] = set()
    for value in schedule_ids or []:
        try:
            schedule_id = int(value)
        except (TypeError, ValueError):
            continue
        if schedule_id <= 0 or schedule_id in seen:
            continue
        seen.add(schedule_id)
        normalized_ids.append(schedule_id)
    if not normalized_ids:
        raise RuntimeError("请选择要删除的采集店铺。")
    deleted_product_ids: list[int] = []
    with session_scope() as session:
        rows = session.scalars(
            select(ScheduledCrawlModel).where(
                ScheduledCrawlModel.owner_username == owner_username,
                ScheduledCrawlModel.id.in_(normalized_ids),
            )
        ).all()
        found_ids = {int(row.id) for row in rows}
        if found_ids:
            product_rows = session.scalars(
                select(ProductModel).where(
                    ProductModel.owner_username == owner_username,
                    ProductModel.scheduled_crawl_id.in_(found_ids),
                )
            ).all()
            for product in product_rows:
                if delete_collected_products and product.review_status in {"pending", "approved"}:
                    deleted_product_ids.append(int(product.id))
                    session.delete(product)
                else:
                    product.scheduled_crawl_id = None

            task_rows = session.scalars(
                select(CrawlTaskModel).where(
                    CrawlTaskModel.owner_username == owner_username,
                    CrawlTaskModel.scheduled_crawl_id.in_(found_ids),
                )
            ).all()
            for task in task_rows:
                task.scheduled_crawl_id = None
        for row in rows:
            session.delete(row)
        result = {
            "deletedIds": sorted(found_ids),
            "failedIds": [schedule_id for schedule_id in normalized_ids if schedule_id not in found_ids],
            "deletedCount": len(found_ids),
            "deletedProductCount": len(deleted_product_ids),
        }
    cleanup_product_image_ids(deleted_product_ids)
    return result


def run_scheduled_crawl(owner_username: str, schedule_id: int) -> dict[str, Any]:
    run_scheduled_crawl_job(owner_username, schedule_id)
    with session_scope() as session:
        row = session.get(ScheduledCrawlModel, schedule_id)
        if row is None:
            raise RuntimeError("定时任务不存在。")
        return scheduled_crawl_to_public(row)


def run_scheduled_crawls_now(
    owner_username: str,
    *,
    keyword: str | None = None,
    enabled_status: str | None = None,
    status: str | None = None,
    schedule_time: str | None = None,
    created_at_from: str | None = None,
    created_at_to: str | None = None,
) -> dict[str, Any]:
    if not SCHEDULE_RUN_LOCK.acquire(blocking=False):
        raise RuntimeError("已有定时采集调度正在执行，请稍后再试。")
    try:
        now = datetime.now()
        with session_scope() as session:
            reconcile_interrupted_scheduled_crawls(session, owner_username=owner_username)
            query = scheduled_crawls_query(
                owner_username,
                keyword=keyword,
                enabled_status=enabled_status,
                status=status,
                schedule_time=schedule_time,
                created_at_from=created_at_from,
                created_at_to=created_at_to,
            ).where(
                ScheduledCrawlModel.enabled.is_(True),
                ScheduledCrawlModel.status != "running",
            )
            rows = session.scalars(query.order_by(ScheduledCrawlModel.created_at.asc())).all()
            if not rows:
                raise RuntimeError("当前没有可立即执行的已启用采集店铺。")
            batch_size = max(1, int(settings.scheduled_crawl_dispatch_batch_size))
            batch_rows = rows[:batch_size]
            due_items = [(row.owner_username, int(row.id)) for row in batch_rows]
            for row in rows:
                row.next_run_at = now
            for row in batch_rows:
                row.status = "running"
                row.last_run_at = now
                row.next_run_at = next_daily_run_at(row.schedule_time, now=now)

        dispatched_count = 0
        failed_ids: list[int] = []
        for item_owner_username, schedule_id in due_items:
            try:
                dispatch_scheduled_crawl(item_owner_username, schedule_id)
                dispatched_count += 1
            except Exception as exc:
                failed_ids.append(schedule_id)
                with session_scope() as session:
                    row = session.get(ScheduledCrawlModel, schedule_id)
                    if row is not None:
                        row.status = "failed"
                        row.notes = str(exc)
            time.sleep(settings.scheduled_crawl_dispatch_pause_seconds)
        return {
            "total": len(rows),
            "dispatchedCount": dispatched_count,
            "matchedCount": len(rows),
            "pendingDispatchCount": max(0, len(rows) - dispatched_count),
            "batchSize": batch_size,
            "failedIds": failed_ids,
            "failedCount": len(failed_ids),
        }
    finally:
        SCHEDULE_RUN_LOCK.release()


def run_scheduled_crawl_job(owner_username: str, schedule_id: int) -> None:
    row_enabled = False
    schedule_time = "09:00"
    with session_scope() as session:
        row = session.get(ScheduledCrawlModel, schedule_id)
        if row is None:
            raise RuntimeError("定时任务不存在。")
        if row.owner_username != owner_username:
            raise RuntimeError("不能执行其他用户的定时任务。")
        row.status = "running"
        row.last_run_at = datetime.now()
        session.flush()
        source_type = row.source_type
        target = scheduled_crawl_task_target(row)
        row_enabled = bool(row.enabled)
        schedule_time = row.schedule_time

    task_payload = type(
        "TaskPayload",
        (),
        {
            "sourceId": None,
            "sourceType": source_type,
            "target": target,
            "mode": "scheduled",
            "scheduledCrawlId": schedule_id,
        },
    )()
    try:
        create_task(owner_username, task_payload)
    except Exception as exc:
        with session_scope() as session:
            row = session.get(ScheduledCrawlModel, schedule_id)
            if row is not None:
                row.status = "failed"
                row.notes = str(exc)
                row.next_run_at = next_daily_run_at(row.schedule_time or schedule_time) if row.enabled or row_enabled else None
        raise
    else:
        with session_scope() as session:
            row = session.get(ScheduledCrawlModel, schedule_id)
            if row is None:
                raise RuntimeError("定时任务不存在。")
            row.status = "idle" if row.enabled else "disabled"
            if not schedule_fallback_shop_url(row.notes):
                row.notes = ""
            row.next_run_at = next_daily_run_at(row.schedule_time) if row.enabled else None
            session.flush()


def _sales_order_sync_task_id(store_id: int) -> str:
    return f"{SALES_ORDER_SYNC_TASK_PREFIX}{int(store_id)}"


def _sales_order_sync_store_id(task_id: str) -> int:
    normalized = str(task_id or "").strip()
    if not normalized.startswith(SALES_ORDER_SYNC_TASK_PREFIX):
        raise LookupError("销量同步任务不存在或无权访问。")
    store_id_text = normalized[len(SALES_ORDER_SYNC_TASK_PREFIX) :]
    if not store_id_text.isdigit():
        raise LookupError("销量同步任务不存在或无权访问。")
    store_id = int(store_id_text)
    if store_id <= 0:
        raise LookupError("销量同步任务不存在或无权访问。")
    return store_id


def _owned_sales_order_sync_store(
    session: Any,
    owner_username: str,
    store_id: int,
) -> StoreModel:
    store = session.scalar(
        select(StoreModel).where(
            StoreModel.id == int(store_id),
            StoreModel.owner_username == str(owner_username or "").strip(),
        )
    )
    if store is None:
        raise LookupError("店铺不存在或无权访问。")
    return store


def _ensure_sales_order_sync_state(
    session: Any,
    *,
    owner_username: str,
    store_id: int,
) -> SalesSyncStateModel:
    state = session.scalar(
        select(SalesSyncStateModel).where(
            SalesSyncStateModel.store_id == store_id,
            SalesSyncStateModel.owner_username == owner_username,
        )
    )
    if state is not None:
        return state
    try:
        with session.begin_nested():
            state = SalesSyncStateModel(
                owner_username=owner_username,
                store_id=store_id,
                initial_sync_completed=False,
                sync_status="idle",
            )
            session.add(state)
            session.flush()
    except IntegrityError:
        state = session.scalar(
            select(SalesSyncStateModel).where(
                SalesSyncStateModel.store_id == store_id,
                SalesSyncStateModel.owner_username == owner_username,
            )
        )
    if state is None:
        raise RuntimeError("店铺销量同步状态不存在。")
    return state


def _sales_order_sync_public_status(
    state: SalesSyncStateModel,
) -> str:
    raw_status = str(state.sync_status or "idle")
    if raw_status.startswith("running"):
        return "running"
    if raw_status == "idle" and state.last_successful_sync_at is not None:
        return "completed"
    return raw_status


def _sales_order_sync_state_to_public(
    state: SalesSyncStateModel,
    *,
    already_running: bool = False,
) -> dict[str, Any]:
    return {
        "id": _sales_order_sync_task_id(state.store_id),
        "storeId": state.store_id,
        "status": _sales_order_sync_public_status(state),
        "alreadyRunning": bool(already_running),
        "initialSyncCompleted": bool(state.initial_sync_completed),
        "progressCurrent": int(state.progress_current or 0),
        "progressTotal": int(state.progress_total or 0),
        "lastSuccessfulSyncAt": iso_sales_datetime(
            state.last_successful_sync_at
        ),
        "lastRemoteUpdatedAt": iso_sales_datetime(
            state.last_remote_updated_at
        ),
        "lastError": str(state.last_error or ""),
    }


def _sales_order_sync_active_timeout() -> timedelta:
    sync_service = _loaded_sales_sync_service()
    timeout = getattr(
        sync_service,
        "SALES_SYNC_LEASE_TIMEOUT",
        timedelta(minutes=10),
    )
    return (
        timeout
        if isinstance(timeout, timedelta)
        else timedelta(minutes=10)
    )


def _sales_order_sync_status_is_active(
    sync_status: str | None,
    state_updated_at: datetime | None,
    *,
    now: datetime | None = None,
) -> bool:
    normalized_status = str(sync_status or "")
    if not (
        normalized_status == "queued"
        or normalized_status.startswith("running")
    ):
        return False
    if state_updated_at is None:
        return False
    current = now or sales_now_naive()
    return state_updated_at >= (
        current - _sales_order_sync_active_timeout()
    )


def _sales_order_sync_error_is_cooling_down(
    sync_status: str | None,
    state_updated_at: datetime | None,
    *,
    now: datetime,
) -> bool:
    if str(sync_status or "") != "error":
        return False
    if state_updated_at is None:
        return True
    return state_updated_at > (
        now - SALES_ORDER_SYNC_FAILURE_COOLDOWN
    )

def list_order_sync_stores(
    owner_username: str | None = None,
) -> list[dict[str, Any]]:
    with session_scope() as session:
        query = select(StoreModel)
        if owner_username:
            query = query.where(
                StoreModel.owner_username == owner_username
            )
        rows = session.scalars(
            query.order_by(
                StoreModel.owner_username.asc(),
                StoreModel.id.asc(),
            )
        ).all()
        return [
            {
                "id": int(row.id),
                "name": row.alias_name or row.store_name,
                "code": row.store_code,
                "enabled": bool(row.enabled),
            }
            for row in rows
        ]


def get_sales_order_sync_state(
    owner_username: str,
    store_id: int,
) -> dict[str, Any]:
    normalized_owner = str(owner_username or "").strip()
    normalized_store_id = int(store_id)
    with session_scope() as session:
        _owned_sales_order_sync_store(
            session,
            normalized_owner,
            normalized_store_id,
        )
        state = _ensure_sales_order_sync_state(
            session,
            owner_username=normalized_owner,
            store_id=normalized_store_id,
        )
        return _sales_order_sync_state_to_public(
            state,
            already_running=_sales_order_sync_status_is_active(
                state.sync_status,
                state.updated_at,
                now=sales_now_naive(),
            ),
        )


def get_sales_order_sync_task(
    owner_username: str,
    task_id: str,
) -> dict[str, Any]:
    store_id = _sales_order_sync_store_id(task_id)
    return get_sales_order_sync_state(owner_username, store_id)


def _loaded_sales_sync_service() -> Any:
    global sales_sync_service
    if sales_sync_service is None:
        from app.services import sales_sync_service as loaded_service

        sales_sync_service = loaded_service
    return sales_sync_service


def run_sales_order_sync_task(
    owner_username: str,
    store_id: int,
    run_id: str | None = None,
) -> None:
    normalized_owner = str(owner_username or "").strip()
    normalized_store_id = int(store_id)
    try:
        _loaded_sales_sync_service().sync_owned_store(
            normalized_owner,
            normalized_store_id,
            run_id=run_id,
        )
    except Exception:
        with session_scope() as session:
            state = session.scalar(
                select(SalesSyncStateModel).where(
                    SalesSyncStateModel.store_id == normalized_store_id,
                    SalesSyncStateModel.owner_username == normalized_owner,
                    SalesSyncStateModel.sync_status == "queued",
                )
            )
            if state is not None:
                failed_at = sales_now_naive()
                state.sync_status = "error"
                state.last_error = "销量同步失败，请稍后重试。"
                state.updated_at = failed_at
        if run_id is not None:
            sales_order_sync_history_service.fail_run(
                normalized_owner,
                run_id,
            )
        raise


def dispatch_sales_order_sync_task(
    owner_username: str,
    store_id: int,
    run_id: str,
) -> None:
    task_id = _sales_order_sync_task_id(store_id)
    if should_use_redis_task_queue():
        enqueue_task(
            run_sales_order_sync_task,
            owner_username,
            store_id,
            run_id,
            job_id=f"{task_id}-{uuid.uuid4().hex[:12]}",
            description=f"销量同步 {store_id}",
            queue_name=task_queue_name_for_kind("sync"),
            owner_username=owner_username,
        )
        return
    if not SALES_ORDER_SYNC_LOCAL_SLOTS.acquire(blocking=False):
        raise RuntimeError("本地销量同步队列已满，请稍后重试。")
    try:
        future = SALES_ORDER_SYNC_EXECUTOR.submit(
            run_sales_order_sync_task,
            owner_username,
            store_id,
            run_id,
        )
    except Exception:
        SALES_ORDER_SYNC_LOCAL_SLOTS.release()
        raise

    add_done_callback = getattr(future, "add_done_callback", None)
    if callable(add_done_callback):
        add_done_callback(
            lambda _future: SALES_ORDER_SYNC_LOCAL_SLOTS.release()
        )
    else:
        SALES_ORDER_SYNC_LOCAL_SLOTS.release()


def queue_sales_order_sync(
    owner_username: str,
    store_id: int,
    *,
    trigger_type: str = "manual",
    parent_run_id: str | None = None,
) -> dict[str, Any]:
    normalized_owner = str(owner_username or "").strip()
    ensure_system_task_dispatch_allowed(normalized_owner)
    normalized_store_id = int(store_id)
    with session_scope() as session:
        store = _owned_sales_order_sync_store(
            session,
            normalized_owner,
            normalized_store_id,
        )
        if not store.enabled:
            raise ValueError(
                "店铺已停用，无法立即同步销量；请先启用店铺后再重试。"
            )
        sync_service = _loaded_sales_sync_service()
        if not sync_service.store_has_sync_credentials(store):
            raise ValueError(
                sync_service.MISSING_SALES_SYNC_CREDENTIALS_MESSAGE
            )
        _ensure_sales_order_sync_state(
            session,
            owner_username=normalized_owner,
            store_id=normalized_store_id,
        )
        state = session.scalar(
            select(SalesSyncStateModel)
            .where(
                SalesSyncStateModel.store_id == normalized_store_id,
                SalesSyncStateModel.owner_username == normalized_owner,
            )
            .with_for_update()
        )
        if state is None:
            raise RuntimeError("店铺销量同步状态不存在。")
        queue_decision_at = sales_now_naive()
        active = _sales_order_sync_status_is_active(
            state.sync_status,
            state.updated_at,
            now=queue_decision_at,
        )
        if active:
            if trigger_type == "retry":
                raise ValueError(
                    "该店铺的订单同步任务正在执行，请完成后再重试。"
                )
            return _sales_order_sync_state_to_public(
                state,
                already_running=True,
            )
        state.sync_status = "queued"
        state.progress_current = 0
        state.progress_total = 0
        state.last_error = None
        state.updated_at = queue_decision_at
        run = sales_order_sync_history_service.create_run(
            session,
            owner_username=normalized_owner,
            store=store,
            trigger_type=trigger_type,
            parent_run_id=parent_run_id,
            initial_sync=not bool(state.initial_sync_completed),
        )
        queued_task = _sales_order_sync_state_to_public(state)
        queued_task["runId"] = run.id

    try:
        dispatch_sales_order_sync_task(
            normalized_owner,
            normalized_store_id,
            run.id,
        )
    except Exception:
        with session_scope() as session:
            state = session.scalar(
                select(SalesSyncStateModel).where(
                    SalesSyncStateModel.store_id == normalized_store_id,
                    SalesSyncStateModel.owner_username == normalized_owner,
                    SalesSyncStateModel.sync_status == "queued",
                )
            )
            if state is not None:
                failed_at = sales_now_naive()
                state.sync_status = "error"
                state.last_error = "销量同步任务投递失败，请稍后重试。"
                state.updated_at = failed_at
        sales_order_sync_history_service.fail_run(
            normalized_owner,
            run.id,
            message="销量同步任务投递失败，请稍后重试。",
        )
        raise
    return queued_task


def sales_order_sync_is_due(
    last_successful_sync_at: datetime | None,
    sync_status: str | None,
    *,
    now: datetime | None = None,
    state_updated_at: datetime | None = None,
    interval: timedelta = SALES_ORDER_SYNC_INTERVAL,
) -> bool:
    current = now or sales_now_naive()
    if _sales_order_sync_status_is_active(
        sync_status,
        state_updated_at,
        now=current,
    ):
        return False
    if _sales_order_sync_error_is_cooling_down(
        sync_status,
        state_updated_at,
        now=current,
    ):
        return False
    if last_successful_sync_at is None:
        return True
    return last_successful_sync_at <= (
        current - interval
    )


def sales_order_sync_due_candidates(
    now: datetime,
    *,
    interval: timedelta = SALES_ORDER_SYNC_INTERVAL,
    owner_username: str | None = None,
) -> list[tuple[str, int]]:
    cutoff = now - interval
    active_after = (
        now - _sales_order_sync_active_timeout()
    )
    failed_before = (
        now - SALES_ORDER_SYNC_FAILURE_COOLDOWN
    )
    with session_scope() as session:
        active_count_query = select(func.count()).where(
                    and_(
                        or_(
                            SalesSyncStateModel.sync_status == "queued",
                            SalesSyncStateModel.sync_status.like(
                                "running%"
                            ),
                        ),
                        SalesSyncStateModel.updated_at >= active_after,
                    )
                )
        if owner_username:
            active_count_query = active_count_query.where(
                SalesSyncStateModel.owner_username == owner_username
            )
        active_count = int(session.scalar(active_count_query) or 0)
        available_slots = max(
            0,
            SALES_ORDER_SYNC_BATCH_SIZE - active_count,
        )
        if available_slots <= 0:
            return []
        due_query = (
            select(
                StoreModel.owner_username,
                StoreModel.id,
            )
            .outerjoin(
                SalesSyncStateModel,
                and_(
                    SalesSyncStateModel.store_id == StoreModel.id,
                    SalesSyncStateModel.owner_username
                    == StoreModel.owner_username,
                ),
            )
            .where(
                StoreModel.enabled.is_(True),
                func.length(
                    func.trim(
                        StoreModel.rakuten_service_secret_encrypted
                    )
                )
                > 0,
                func.length(
                    func.trim(
                        StoreModel.rakuten_license_key_encrypted
                    )
                )
                > 0,
                or_(
                    SalesSyncStateModel.store_id.is_(None),
                    SalesSyncStateModel.last_successful_sync_at.is_(None),
                    SalesSyncStateModel.last_successful_sync_at <= cutoff,
                ),
                or_(
                    SalesSyncStateModel.sync_status.is_(None),
                    SalesSyncStateModel.updated_at < active_after,
                    and_(
                        SalesSyncStateModel.sync_status
                        != "queued",
                        SalesSyncStateModel.sync_status.not_like(
                            "running%"
                        ),
                    ),
                ),
                or_(
                    SalesSyncStateModel.store_id.is_(None),
                    SalesSyncStateModel.sync_status != "error",
                    SalesSyncStateModel.updated_at <= failed_before,
                ),
            )
            .order_by(StoreModel.id.asc())
            .limit(available_slots)
        )
        if owner_username:
            due_query = due_query.where(
                StoreModel.owner_username == owner_username
            )
        rows = session.execute(due_query).all()
    return [
        (str(row.owner_username), int(row.id))
        for row in rows
    ]


def run_due_sales_order_syncs_once() -> int:
    if not SALES_ORDER_SYNC_RUN_LOCK.acquire(blocking=False):
        return 0
    try:
        sales_order_sync_history_service.recover_stale_runs(
            stale_after=_sales_order_sync_active_timeout(),
        )
        with session_scope() as session:
            owners = session.scalars(
                select(UserAccountModel.username).where(
                    UserAccountModel.enabled.is_(True)
                )
            ).all()
        queued_count = 0
        for owner_username in owners:
            owner = str(owner_username)
            if system_task_dispatch_paused(owner):
                continue
            try:
                sales_order_sync_history_service.cleanup_successful_runs_if_due(
                    owner_username=owner,
                )
            except Exception:
                logger.warning(
                    "用户 %s 的订单同步成功记录清理失败",
                    owner,
                    exc_info=True,
                )
            user_settings = (
                sales_order_sync_history_service.get_user_settings(owner)
            )
            interval = timedelta(
                minutes=int(user_settings["intervalMinutes"])
            )
            for candidate_owner, store_id in sales_order_sync_due_candidates(
                sales_now_naive(),
                interval=interval,
                owner_username=owner,
            ):
                try:
                    queue_sales_order_sync(
                        candidate_owner,
                        store_id,
                        trigger_type="automatic",
                    )
                except Exception:
                    logger.warning(
                        "店铺 %s 的定时销量同步投递失败",
                        store_id,
                        exc_info=True,
                    )
                    continue
                queued_count += 1
        return queued_count
    finally:
        SALES_ORDER_SYNC_RUN_LOCK.release()


def run_due_store_product_syncs_once() -> int:
    if not STORE_PRODUCT_SYNC_SCHEDULE_LOCK.acquire(blocking=False):
        return 0
    try:
        now = datetime.now()
        with session_scope() as session:
            owners = session.scalars(
                select(UserAccountModel.username).where(
                    UserAccountModel.enabled.is_(True)
                )
            ).all()
        queued_count = 0
        for owner_username in owners:
            owner = str(owner_username)
            if system_task_dispatch_paused(owner):
                continue
            with session_scope() as session:
                row = session.get(
                    SystemSettingModel,
                    time_settings_row_key(owner),
                )
                payload = (
                    initial_user_time_settings_payload(
                        session,
                        owner,
                        now=now,
                    )
                    if row is None
                    else load_time_settings_payload(row, now=now)
                )
                row = upsert_time_settings_row(session, payload, owner)
                next_sync_at = parse_public_datetime(
                    payload.get("productSyncNextAt")
                )
                if not payload["productSyncEnabled"] or (
                    next_sync_at is not None and next_sync_at > now
                ):
                    continue
                due_store_ids = session.scalars(
                    select(StoreModel.id)
                    .where(
                        StoreModel.owner_username == owner,
                        StoreModel.enabled.is_(True),
                        func.length(
                            func.trim(
                                StoreModel.rakuten_service_secret_encrypted
                            )
                        )
                        > 0,
                        func.length(
                            func.trim(
                                StoreModel.rakuten_license_key_encrypted
                            )
                        )
                        > 0,
                    )
                    .order_by(StoreModel.id.asc())
                ).all()
                payload["productSyncNextAt"] = datetime_to_public(
                    next_weekly_run_at(
                        payload["productSyncWeekday"],
                        payload["productSyncTime"],
                        now=now,
                    )
                )
                row.value_json = json.dumps(payload, ensure_ascii=False)
            owner_queued_count = 0
            for store_id in due_store_ids:
                try:
                    create_sync_task(owner, int(store_id))
                except Exception:
                    logger.warning(
                        "店铺 %s 的定时商品同步投递失败",
                        store_id,
                        exc_info=True,
                    )
                    continue
                owner_queued_count += 1
            queued_count += owner_queued_count
            with session_scope() as session:
                row = session.get(
                    SystemSettingModel,
                    time_settings_row_key(owner),
                )
                payload = load_time_settings_payload(row, now=now)
                payload["productSyncLastAt"] = datetime_to_public(now)
                payload["productSyncLastTaskCount"] = owner_queued_count
                upsert_time_settings_row(session, payload, owner)
        return queued_count
    finally:
        STORE_PRODUCT_SYNC_SCHEDULE_LOCK.release()


def run_due_scheduled_crawls_once() -> int:
    if not SCHEDULE_RUN_LOCK.acquire(blocking=False):
        return 0
    try:
        now = datetime.now()
        with session_scope() as session:
            reconcile_interrupted_scheduled_crawls(session)
            rows = session.scalars(
                select(ScheduledCrawlModel).where(
                    ScheduledCrawlModel.enabled.is_(True),
                    ScheduledCrawlModel.source_type == "shop",
                    ScheduledCrawlModel.next_run_at.is_not(None),
                    ScheduledCrawlModel.next_run_at <= now,
                    ScheduledCrawlModel.status != "running",
                )
                .order_by(ScheduledCrawlModel.next_run_at.asc(), ScheduledCrawlModel.created_at.asc(), ScheduledCrawlModel.id.asc())
                .limit(max(1, int(settings.scheduled_crawl_dispatch_batch_size)))
            ).all()
            due_rows = [
                row
                for row in rows
                if not system_task_dispatch_paused(row.owner_username)
            ]
            due_items = [(row.owner_username, row.id) for row in due_rows]
            for row in due_rows:
                row.status = "running"
                row.last_run_at = now
                row.next_run_at = next_daily_run_at(row.schedule_time, now=now)

        for owner_username, schedule_id in due_items:
            try:
                dispatch_scheduled_crawl(owner_username, schedule_id)
            except Exception as exc:
                with session_scope() as session:
                    row = session.get(ScheduledCrawlModel, schedule_id)
                    if row is not None:
                        row.status = "failed"
                        row.notes = str(exc)
            time.sleep(settings.scheduled_crawl_dispatch_pause_seconds)
        return len(due_items)
    finally:
        SCHEDULE_RUN_LOCK.release()


def auto_listing_candidate_product_ids(
    owner_username: str,
    store_id: int,
    quantity: int,
) -> list[int]:
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None or store.owner_username != owner_username:
            raise RuntimeError("自动上架店铺不存在或不属于当前用户。")
        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.review_status == "approved",
                ProductModel.parent_product_id.is_(None),
                ProductModel.listing_task_id.is_(None),
            )
            .order_by(ProductModel.created_at.asc(), ProductModel.id.asc())
        ).all()
        selected_ids: list[int] = []
        for product in products:
            listed_store_ids = {
                int(item.get("storeId") or 0)
                for item in product_listed_stores(product_raw_payload(product))
            }
            if store_id in listed_store_ids:
                continue
            if listing_preflight_blocking_messages(
                [listing_preflight_product_check(product, store)]
            ):
                continue
            selected_ids.append(int(product.id))
            if len(selected_ids) >= quantity:
                break
        return selected_ids


def execute_auto_listing_schedule(
    schedule_id: int,
    *,
    owner_username: str | None = None,
    advance_next_run: bool = False,
) -> dict[str, Any]:
    now = datetime.now()
    with session_scope() as session:
        query = select(AutoListingScheduleModel).where(
            AutoListingScheduleModel.id == schedule_id,
        )
        if owner_username is not None:
            query = query.where(
                AutoListingScheduleModel.owner_username == owner_username,
            )
        row = session.scalar(query.with_for_update())
        if row is None:
            raise RuntimeError("自动上架任务不存在。")
        if row.status == "running":
            raise RuntimeError("该自动上架任务正在执行，请勿重复操作。")
        resolved_owner_username = row.owner_username
        store_id = int(row.store_id)
        quantity = int(row.quantity)
        task_type = row.task_type or "automatic"
        row.status = "running"
        row.last_run_at = now
        row.last_error = None
        if advance_next_run and row.enabled:
            if row.task_type == "manual" and row.schedule_type == "once":
                row.enabled = False
            else:
                row.next_run_at = next_auto_listing_run_at(
                    row.schedule_type,
                    row.schedule_time,
                    weekday=row.weekday,
                    month_day=row.month_day,
                    now=now,
                )

    try:
        product_ids = auto_listing_candidate_product_ids(
            resolved_owner_username,
            store_id,
            quantity,
        )
        if not product_ids:
            with session_scope() as session:
                row = session.get(AutoListingScheduleModel, schedule_id)
                if row is None:
                    raise RuntimeError("自动上架任务不存在。")
                store = session.get(StoreModel, row.store_id)
                row.status = (
                    "completed"
                    if row.task_type == "manual"
                    else ("idle" if row.enabled else "disabled")
                )
                row.last_task_ids_json = "[]"
                row.last_message = "当前没有可自动上架的已审核商品。"
                return auto_listing_schedule_to_public(row, store)

        result = create_listing_task(
            resolved_owner_username,
            SimpleNamespace(
                productIds=product_ids,
                storeIds=[store_id],
                taskName=(
                    f"手动上架 {now:%Y-%m-%d %H:%M}"
                    if task_type == "manual"
                    else f"自动上架 {now:%Y-%m-%d %H:%M}"
                ),
            ),
        )
        task_rows = result.get("listingTasks") or []
        task_ids = [
            str(item.get("id") or "")
            for item in task_rows
            if isinstance(item, dict) and item.get("id")
        ]
        if not task_ids and isinstance(result.get("listingTask"), dict):
            first_task_id = str(result["listingTask"].get("id") or "")
            if first_task_id:
                task_ids.append(first_task_id)
        actual_count = int((result.get("summary") or {}).get("total") or len(product_ids))
        with session_scope() as session:
            row = session.get(AutoListingScheduleModel, schedule_id)
            if row is None:
                raise RuntimeError("自动上架任务不存在。")
            store = session.get(StoreModel, row.store_id)
            row.status = (
                "completed"
                if row.task_type == "manual"
                else ("idle" if row.enabled else "disabled")
            )
            row.last_task_ids_json = json.dumps(task_ids, ensure_ascii=False)
            row.last_message = (
                f"计划上架 {quantity} 件，实际已创建 {actual_count} 件商品的上架任务"
                f"，共 {len(task_ids)} 个任务。"
            )
            row.last_error = None
            return auto_listing_schedule_to_public(row, store)
    except Exception as exc:
        error_message = humanize_task_error_detail(str(exc)) or "自动上架任务创建失败。"
        logger.exception("Auto listing schedule failed: %s", schedule_id)
        with session_scope() as session:
            row = session.get(AutoListingScheduleModel, schedule_id)
            if row is not None:
                row.status = "failed"
                row.last_task_ids_json = "[]"
                row.last_message = "本次自动上架任务创建失败。"
                row.last_error = error_message
        raise RuntimeError(error_message) from exc


def run_auto_listing_schedule_now(
    owner_username: str,
    schedule_id: int,
) -> dict[str, Any]:
    return execute_auto_listing_schedule(
        schedule_id,
        owner_username=owner_username,
        advance_next_run=False,
    )


def run_due_auto_listing_schedules_once() -> int:
    if not AUTO_LISTING_SCHEDULE_LOCK.acquire(blocking=False):
        return 0
    try:
        now = datetime.now()
        with session_scope() as session:
            schedule_rows = session.scalars(
                select(AutoListingScheduleModel).where(
                    AutoListingScheduleModel.enabled.is_(True),
                    AutoListingScheduleModel.next_run_at.is_not(None),
                    AutoListingScheduleModel.next_run_at <= now,
                    or_(
                        and_(
                            AutoListingScheduleModel.task_type == "automatic",
                            AutoListingScheduleModel.status != "running",
                        ),
                        and_(
                            AutoListingScheduleModel.task_type == "manual",
                            AutoListingScheduleModel.schedule_type == "once",
                            AutoListingScheduleModel.status == "idle",
                        ),
                    ),
                )
                .order_by(
                    AutoListingScheduleModel.next_run_at.asc(),
                    AutoListingScheduleModel.created_at.asc(),
                    AutoListingScheduleModel.id.asc(),
                )
                .limit(max(1, int(settings.scheduled_crawl_dispatch_batch_size)))
            ).all()

        runnable_rows = [
            row
            for row in schedule_rows
            if not system_task_dispatch_paused(row.owner_username)
        ]
        for schedule_row in runnable_rows:
            try:
                execute_auto_listing_schedule(
                    int(schedule_row.id),
                    advance_next_run=True,
                )
            except RuntimeError:
                continue
        return len(runnable_rows)
    finally:
        AUTO_LISTING_SCHEDULE_LOCK.release()


def auto_deletion_candidate_product_ids(owner_username: str, store_id: int, quantity: int) -> list[int]:
    period_end = sales_now_naive().date()
    cutoff = period_end - timedelta(days=364)
    sales = (
        select(
            ProductSalesDailyModel.store_id.label("store_id"),
            ProductSalesDailyModel.manage_number.label("manage_number"),
            func.sum(ProductSalesDailyModel.effective_units).label("effective_units"),
        )
        .where(
            ProductSalesDailyModel.owner_username == owner_username,
            ProductSalesDailyModel.store_id == store_id,
            ProductSalesDailyModel.sales_date >= cutoff,
            ProductSalesDailyModel.sales_date <= period_end,
        )
        .group_by(ProductSalesDailyModel.store_id, ProductSalesDailyModel.manage_number)
        .subquery()
    )
    with session_scope() as session:
        return [
            int(value)
            for value in session.scalars(
                select(ProductModel.id)
                .outerjoin(
                    sales,
                    and_(
                        sales.c.store_id == ProductModel.store_id,
                        sales.c.manage_number == ProductModel.rakuten_manage_number,
                    ),
                )
                .where(
                    ProductModel.owner_username == owner_username,
                    ProductModel.store_id == store_id,
                    ProductModel.review_status == "listed",
                    ProductModel.rakuten_listing_status == "listed",
                    ProductModel.listed_at.is_not(None),
                    func.coalesce(sales.c.effective_units, 0) == 0,
                )
                .order_by(ProductModel.listed_at.asc(), ProductModel.id.asc())
                .limit(quantity)
            ).all()
        ]


def execute_auto_deletion_task(task_id: int, *, owner_username: str | None = None, advance_next_run: bool = False) -> dict[str, Any]:
    now = datetime.now()
    with session_scope() as session:
        query = select(AutoDeletionTaskModel).where(AutoDeletionTaskModel.id == task_id)
        if owner_username is not None:
            query = query.where(AutoDeletionTaskModel.owner_username == owner_username)
        row = session.scalar(query.with_for_update())
        if row is None:
            raise RuntimeError("自动删除任务不存在。")
        if row.status == "running":
            raise RuntimeError("该自动删除任务正在执行。")
        resolved_owner = row.owner_username
        store_id = int(row.store_id)
        quantity = int(row.quantity)
        task_type = row.task_type
        row.status = "running"
        row.last_run_at = now
        row.last_error = None
        if advance_next_run and row.enabled:
            if row.task_type == "manual" and row.schedule_type == "once":
                row.enabled = False
            else:
                row.next_run_at = next_auto_listing_run_at(row.schedule_type, row.schedule_time, weekday=row.weekday, month_day=row.month_day, now=now)
    try:
        product_ids = auto_deletion_candidate_product_ids(resolved_owner, store_id, quantity)
        task_ids: list[str] = []
        if product_ids:
            result = create_product_delete_sync_task(resolved_owner, product_ids)
            rows = result.get("syncTasks") or ([result.get("syncTask")] if result.get("syncTask") else [])
            task_ids = [str(item.get("id")) for item in rows if isinstance(item, dict) and item.get("id")]
        with session_scope() as session:
            row = session.get(AutoDeletionTaskModel, task_id)
            if row is None:
                raise RuntimeError("自动删除任务不存在。")
            store = session.get(StoreModel, row.store_id)
            row.status = "completed" if task_type == "manual" else "idle"
            row.last_task_ids_json = json.dumps(task_ids, ensure_ascii=False)
            row.last_message = (
                f"计划删除 {quantity} 件，实际已创建 {len(product_ids)} 件商品的删除任务，共 {len(task_ids)} 个任务。"
                if product_ids
                else "当前没有符合条件的已上架零销量商品。"
            )
            return auto_deletion_task_to_public(row, store)
    except Exception as exc:
        message = humanize_task_error_detail(str(exc)) or "自动删除任务创建失败。"
        with session_scope() as session:
            row = session.get(AutoDeletionTaskModel, task_id)
            if row is not None:
                row.status = "failed"
                row.last_error = message
                row.last_message = "本次自动删除任务创建失败。"
        raise RuntimeError(message) from exc


def create_manual_deletion_task(owner_username: str, payload: Any) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    store_id = int(getattr(payload, "storeId", 0) or 0)
    quantity = int(getattr(payload, "quantity", 0) or 0)
    execution_mode, execute_at = manual_task_execution_settings(
        payload,
        task_name="手动删除任务",
    )
    if quantity < 1 or quantity > 10000:
        raise RuntimeError("删除数量只能设置为 1 到 10000。")
    with session_scope() as session:
        store = _validate_deletion_store(
            session,
            owner_username,
            store_id,
            for_update=True,
        )
        if execution_mode == "immediate":
            active_manual_task_id = session.scalar(
                select(AutoDeletionTaskModel.id)
                .where(
                    AutoDeletionTaskModel.owner_username == owner_username,
                    AutoDeletionTaskModel.store_id == store_id,
                    AutoDeletionTaskModel.task_type == "manual",
                    AutoDeletionTaskModel.schedule_type == "",
                    AutoDeletionTaskModel.status.in_(("idle", "running")),
                )
                .order_by(AutoDeletionTaskModel.id.asc())
                .limit(1)
            )
            if active_manual_task_id is not None:
                raise RuntimeError("该店铺已有正在创建或执行的手动删除任务，请勿重复提交。")
        row = AutoDeletionTaskModel(
            owner_username=owner_username,
            store_id=store_id,
            automatic_store_id=None,
            task_type="manual",
            schedule_type="once" if execution_mode == "scheduled" else "",
            schedule_time=execute_at.strftime("%H:%M") if execute_at else "",
            quantity=quantity,
            enabled=execution_mode == "scheduled",
            status="idle",
            next_run_at=execute_at,
            last_message=(
                f"任务将在 {datetime_to_public(execute_at)} 到期执行。"
                if execute_at
                else "任务已受理，后台正在创建删除任务。"
            ),
        )
        session.add(row)
        session.flush()
        task_id = int(row.id)
        if execution_mode == "scheduled":
            return auto_deletion_task_to_public(row, store)
        created = auto_deletion_task_to_public(row, store)
    try:
        dispatch_auto_deletion_task(
            task_id,
            owner_username=owner_username,
            advance_next_run=False,
        )
    except Exception as exc:
        raise RuntimeError(f"删除任务投递失败：{exc}") from exc
    return created


def run_due_auto_deletion_tasks_once() -> int:
    now = datetime.now()
    with session_scope() as session:
        task_rows = session.scalars(
            select(AutoDeletionTaskModel).where(
                AutoDeletionTaskModel.enabled.is_(True),
                AutoDeletionTaskModel.next_run_at.is_not(None),
                AutoDeletionTaskModel.next_run_at <= now,
                or_(
                    and_(
                        AutoDeletionTaskModel.task_type == "automatic",
                        AutoDeletionTaskModel.status != "running",
                    ),
                    and_(
                        AutoDeletionTaskModel.task_type == "manual",
                        AutoDeletionTaskModel.schedule_type == "once",
                        AutoDeletionTaskModel.status == "idle",
                    ),
                ),
            ).order_by(AutoDeletionTaskModel.next_run_at.asc(), AutoDeletionTaskModel.id.asc()).limit(max(1, int(settings.scheduled_crawl_dispatch_batch_size)))
        ).all()
    runnable_rows = [
        row
        for row in task_rows
        if not system_task_dispatch_paused(row.owner_username)
    ]
    for task_row in runnable_rows:
        try:
            execute_auto_deletion_task(
                int(task_row.id),
                advance_next_run=True,
            )
        except RuntimeError:
            continue
    return len(runnable_rows)


def run_periodic_maintenance_once() -> None:
    reconcile_interrupted_background_tasks_once()
    dispatch_queued_crawl_tasks_safely()
    cleanup_expired_product_image_drafts_if_due()
    cleanup_orphan_product_image_dirs_if_due()
    cleanup_completed_scheduled_crawl_tasks_if_due()
    cleanup_store_unlisted_products_if_due()
    cleanup_deleted_product_images(force=False)


def cleanup_expired_product_image_drafts_if_due() -> int:
    global DRAFT_IMAGE_CLEANUP_LAST_RUN_AT
    now = time.time()
    if DRAFT_IMAGE_CLEANUP_LAST_RUN_AT and now - DRAFT_IMAGE_CLEANUP_LAST_RUN_AT < 24 * 60 * 60:
        return 0
    DRAFT_IMAGE_CLEANUP_LAST_RUN_AT = now
    return cleanup_expired_product_image_drafts()


def cleanup_orphan_product_image_dirs_if_due() -> int:
    global ORPHAN_IMAGE_CLEANUP_LAST_RUN_AT
    now = time.time()
    if ORPHAN_IMAGE_CLEANUP_LAST_RUN_AT and now - ORPHAN_IMAGE_CLEANUP_LAST_RUN_AT < 24 * 60 * 60:
        return 0
    ORPHAN_IMAGE_CLEANUP_LAST_RUN_AT = now
    return cleanup_orphan_product_image_dirs()


def schedule_runner_health() -> dict[str, Any]:
    with SCHEDULE_RUNNER_HEALTH_LOCK:
        return {
            **SCHEDULE_RUNNER_HEALTH,
            "active": SCHEDULE_RUNNER_ACTIVE,
        }


def _record_schedule_runner_tick(*, errors: list[str]) -> None:
    tick_at = datetime.now()
    with SCHEDULE_RUNNER_HEALTH_LOCK:
        SCHEDULE_RUNNER_HEALTH["lastTickAt"] = datetime_to_public(tick_at)
        if errors:
            SCHEDULE_RUNNER_HEALTH["lastError"] = "; ".join(errors)
            SCHEDULE_RUNNER_HEALTH["consecutiveFailures"] = (
                int(SCHEDULE_RUNNER_HEALTH["consecutiveFailures"]) + 1
            )
            return
        SCHEDULE_RUNNER_HEALTH["lastSuccessfulTickAt"] = datetime_to_public(tick_at)
        SCHEDULE_RUNNER_HEALTH["lastError"] = ""
        SCHEDULE_RUNNER_HEALTH["consecutiveFailures"] = 0


def run_schedule_runner_tick() -> bool:
    global SCHEDULE_RUNNER_ACTIVE
    with SCHEDULE_RUNNER_HEALTH_LOCK:
        SCHEDULE_RUNNER_ACTIVE = True
    tasks = (
        ("scheduled crawls", run_due_scheduled_crawls_once),
        ("auto listing schedules", run_due_auto_listing_schedules_once),
        ("auto deletion tasks", run_due_auto_deletion_tasks_once),
        ("sales order syncs", run_due_sales_order_syncs_once),
        ("store product syncs", run_due_store_product_syncs_once),
        ("periodic maintenance", run_periodic_maintenance_once),
    )
    errors: list[str] = []
    try:
        for task_name, task in tasks:
            try:
                task()
            except Exception as exc:
                errors.append(f"{task_name}: {type(exc).__name__}")
                logger.exception("Schedule runner task failed: %s", task_name)
        _record_schedule_runner_tick(errors=errors)
        return not errors
    finally:
        with SCHEDULE_RUNNER_HEALTH_LOCK:
            SCHEDULE_RUNNER_ACTIVE = False


def _run_schedule_runner_tick_with_lock(interval_seconds: int) -> bool:
    if settings.task_queue_mode != "redis":
        return run_schedule_runner_tick()
    try:
        from app.core.task_queue import redis_connection

        lock = redis_connection().lock(
            SCHEDULE_RUNNER_REDIS_LOCK_NAME,
            timeout=max(30, int(interval_seconds) * 2),
            blocking_timeout=0,
        )
        if not lock.acquire(blocking=False):
            return False
    except Exception:
        logger.exception("Failed to acquire the schedule runner Redis lock")
        return False
    try:
        return run_schedule_runner_tick()
    finally:
        try:
            lock.release()
        except Exception:
            logger.exception("Failed to release the schedule runner Redis lock")


def start_schedule_runner(interval_seconds: int = 60) -> None:
    global SCHEDULE_RUNNER_STARTED
    if SCHEDULE_RUNNER_STARTED:
        return
    SCHEDULE_RUNNER_STARTED = True

    def loop() -> None:
        while True:
            _run_schedule_runner_tick_with_lock(interval_seconds)
            time.sleep(max(10, interval_seconds))

    threading.Thread(target=loop, name="lt-schedule-runner", daemon=True).start()


def update_product_status(owner_username: str, product_ids: list[int], status: str, *, message: str = "") -> list[dict[str, Any]]:
    if status not in {"pending", "approved", "error", "listed", "listed_master", "rejected"}:
        raise RuntimeError("商品状态不合法。")
    with session_scope() as session:
        rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(product_ids or [-1]),
            )
        ).all()
        if not rows:
            raise RuntimeError("没有找到可操作的商品。")
        if status == "error":
            invalid_rows = [row for row in rows if row.review_status != "pending"]
            if invalid_rows:
                raise RuntimeError("只有待审核商品可以标记异常。")
        if status == "approved":
            replacement_rows = [
                row for row in rows
                if product_replacement_metadata(product_raw_payload(row))
            ]
            if replacement_rows:
                raise RuntimeError("替换采集商品请使用“确认替换”，不能执行普通审核通过。")
            invalid_rows = [row for row in rows if not rakuten_genre_path(row.genre_id)]
            if invalid_rows:
                names = "、".join(normalize_text(row.title) or f"商品 {row.id}" for row in invalid_rows[:3])
                suffix = "等" if len(invalid_rows) > 3 else ""
                raise RuntimeError(f"{len(invalid_rows)} 个商品缺少有效品类：{names}{suffix}。请先选择有效品类。")
        for row in rows:
            row.review_status = status
            if message:
                row.last_error = message if status in {"error", "rejected"} else None
            if status == "approved":
                from app.services.ai_title_service import cleanup_title_versions_for_approved_product

                cleanup_title_versions_for_approved_product(session, row)
        session.flush()
        return [product_to_public(row) for row in rows]


def approve_pending_products_by_count(
    owner_username: str,
    *,
    collection_source: str,
    count: int | None,
) -> dict[str, Any]:
    normalized_source = normalize_text(collection_source).lower()
    if normalized_source not in {"manual", "scheduled"}:
        raise RuntimeError("采集来源不合法。")
    if count is not None and count < 1:
        raise RuntimeError("指定通过数量必须大于 0。")

    with session_scope() as session:
        rows = session.scalars(
            select(ProductModel)
            .where(
                ProductModel.owner_username == owner_username,
                ProductModel.review_status == "pending",
                ProductModel.collection_source == normalized_source,
                ProductModel.listing_task_id.is_(None),
            )
            .order_by(*product_list_order_by("pending"))
        ).all()

        approved_ids: list[int] = []
        skipped: list[dict[str, Any]] = []
        for row in rows:
            if count is not None and len(approved_ids) >= count:
                break
            if product_replacement_metadata(product_raw_payload(row)):
                skipped.append({
                    "productId": row.id,
                    "reason": "替换采集商品不能执行普通审核通过",
                })
                continue
            if not rakuten_genre_path(row.genre_id):
                skipped.append({
                    "productId": row.id,
                    "reason": "缺少有效品类",
                })
                continue

            row.review_status = "approved"
            from app.services.ai_title_service import cleanup_title_versions_for_approved_product

            cleanup_title_versions_for_approved_product(session, row)
            approved_ids.append(row.id)

        session.flush()
        return {
            "approvedCount": len(approved_ids),
            "skippedCount": len(skipped),
            "approvedProductIds": approved_ids,
            "skipped": skipped,
        }


def update_pending_product_genre(owner_username: str, product_id: int, genre_id: str) -> dict[str, Any]:
    normalized_genre_id = normalize_text(genre_id)
    if not re.fullmatch(r"\d{6}", normalized_genre_id) or not rakuten_genre_path(normalized_genre_id):
        raise RuntimeError("请选择有效品类。")
    with session_scope() as session:
        row = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id == product_id,
            )
        ).first()
        if not row:
            raise RuntimeError("没有找到可操作的商品。")
        if row.review_status != "pending":
            raise RuntimeError("只有待审核商品可以修改品类。")
        raw_payload = product_raw_payload(row)
        raw_payload["genreId"] = normalized_genre_id
        row.genre_id = normalized_genre_id
        row.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
        session.flush()
        return product_to_public(row)


def update_pending_product_genres(owner_username: str, products: list[Any]) -> list[dict[str, Any]]:
    normalized: dict[int, str] = {}
    for item in products or []:
        try:
            product_id = int(getattr(item, "productId", 0) or 0)
        except (TypeError, ValueError):
            product_id = 0
        genre_id = normalize_text(getattr(item, "genreId", ""))
        if product_id <= 0:
            raise RuntimeError("商品信息不完整，请重新打开品类设置。")
        if not re.fullmatch(r"\d{6}", genre_id) or not rakuten_genre_path(genre_id):
            raise RuntimeError(f"商品 {product_id} 请选择有效品类。")
        normalized[product_id] = genre_id
    if not normalized:
        raise RuntimeError("请先为商品选择品类。")

    with session_scope() as session:
        rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(normalized),
            )
        ).all()
        rows_by_id = {int(row.id): row for row in rows}
        missing_ids = [product_id for product_id in normalized if product_id not in rows_by_id]
        if missing_ids:
            raise RuntimeError("部分商品不存在或无权操作，请刷新后重试。")
        if any(row.review_status != "pending" for row in rows):
            raise RuntimeError("只有待审核商品可以修改品类。")

        for product_id, genre_id in normalized.items():
            row = rows_by_id[product_id]
            raw_payload = product_raw_payload(row)
            raw_payload["genreId"] = genre_id
            row.genre_id = genre_id
            row.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
        session.flush()
        return [product_to_public(rows_by_id[product_id]) for product_id in normalized]


def delete_products(owner_username: str, product_ids: list[int]) -> dict[str, Any]:
    normalized_ids = [int(value) for value in (product_ids or [])]
    if not normalized_ids:
        raise RuntimeError("请先选择商品。")
    with session_scope() as session:
        rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(normalized_ids),
            )
        ).all()
        if not rows:
            raise RuntimeError("没有找到可删除的商品。")
        if any(row.review_status == "listed" for row in rows):
            if any(row.review_status != "listed" for row in rows):
                raise RuntimeError("店铺商品删除任务不能和其他状态商品混选。")
            if len({row.store_id for row in rows if row.store_id}) != 1:
                raise RuntimeError("请选择同一个店铺下的店铺商品。")
            return create_product_delete_sync_task(owner_username, normalized_ids)

    with session_scope() as session:
        rows = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(normalized_ids),
            )
        ).all()
        if not rows:
            raise RuntimeError("没有找到可删除的商品。")

        success_count = 0
        failed_count = 0
        cabinet_deleted_count = 0
        deleted_ids: list[int] = []
        failed_ids: list[int] = []
        failed_products: list[dict[str, Any]] = []
        errors: list[str] = []
        warnings: list[str] = []
        credential_cache: dict[int, tuple[StoreModel, str, str]] = {}
        for row in rows:
            if row.review_status == "listed":
                try:
                    delete_store_product_from_rakuten(session, row, credential_cache)
                    cabinet_deleted_count += int(getattr(row, "_deleted_cabinet_count", 0) or 0)
                except Exception as exc:
                    failed_count += 1
                    failed_ids.append(row.id)
                    error_text = str(exc)
                    row.last_error = error_text
                    errors.append(f"{productCodeForError(row)}: {error_text}")
                    failed_products.append(product_to_public(row))
                    continue
                warnings.extend(getattr(row, "_delete_warnings", []) or [])
                remove_listed_store_mark_for_store_product(session, row)
            if row.review_status == "listed_master":
                child_rows = session.scalars(
                    select(ProductModel).where(ProductModel.parent_product_id == row.id)
                ).all()
                for child in child_rows:
                    child.parent_product_id = None
            replacement = product_replacement_metadata(product_raw_payload(row))
            replacement_task_id = normalize_text(replacement.get("taskId"))
            if replacement_task_id:
                replacement_task = session.get(SyncTaskModel, replacement_task_id)
                if (
                    replacement_task is not None
                    and replacement_task.owner_username == owner_username
                    and replacement_task.task_type == "product_replace"
                    and replacement_task.status in {"preview_ready", "failed"}
                ):
                    replacement_task.status = "cancelled"
                    replacement_task.message = "待审核替换商品已删除"
                    replacement_task.finished_at = datetime.now()
            deleted_ids.append(row.id)
            session.delete(row)
            success_count += 1
        session.flush()
        message = f"完成，成功删除 {success_count} 个，失败 {failed_count} 个"
        if cabinet_deleted_count:
            message = f"{message}，同步删除图片 {cabinet_deleted_count} 个"
        result = {
            "deletedIds": deleted_ids,
            "failedIds": failed_ids,
            "products": failed_products,
            "summary": {
                "total": len(rows),
                "successCount": success_count,
                "failedCount": failed_count,
                "cabinetDeletedCount": cabinet_deleted_count,
                "message": message,
                "errors": errors[:20],
                "warnings": warnings[:20],
            }
        }
    cleanup_product_image_ids(deleted_ids)
    return result


def productCodeForError(row: ProductModel) -> str:
    return normalize_text(row.rakuten_manage_number or row.item_number or row.title or row.id)


def delete_store_product_from_rakuten(
    session: Any,
    product: ProductModel,
    credential_cache: dict[int, tuple[StoreModel, str, str]],
) -> None:
    if not product.store_id:
        raise RuntimeError("商品未关联店铺，不能删除乐天商品。")
    manage_number = normalize_text(product.rakuten_manage_number or product.item_number)
    if not manage_number:
        raise RuntimeError("商品缺少商品管理编号，不能删除乐天商品。")

    credentials = credential_cache.get(product.store_id)
    if credentials is None:
        store = session.get(StoreModel, product.store_id)
        if store is None:
            raise RuntimeError("商品关联店铺不存在。")
        if not store.enabled:
            raise RuntimeError("商品关联店铺已停用，不能删除乐天商品。")
        credentials = (
            store,
            decrypt_text(store.rakuten_service_secret_encrypted),
            decrypt_text(store.rakuten_license_key_encrypted),
        )
        credential_cache[product.store_id] = credentials

    store, service_secret, license_key = credentials
    raw_payload = product_raw_payload(product)
    delete_rakuten_item(service_secret, license_key, manage_number)
    queue_deleted_product_image_cleanup(session, product, store, raw_payload)
    setattr(product, "_deleted_cabinet_count", 0)
    setattr(product, "_delete_warnings", [])


def queue_deleted_product_image_cleanup(
    session: Any,
    product: ProductModel,
    store: StoreModel,
    raw_payload: dict[str, Any],
) -> None:
    cabinet_targets = product_cabinet_file_targets(raw_payload, store.store_code)
    local_image_urls = []
    for image_url in product_images_for_edit(product):
        stored_image = parse_product_image_url(image_url)
        if stored_image is not None and stored_image.product_id == int(product.id):
            local_image_urls.append(image_url)
    existing = session.scalar(
        select(DeletedProductImageCleanupModel).where(
            DeletedProductImageCleanupModel.owner_username == product.owner_username,
            DeletedProductImageCleanupModel.original_product_id == product.id,
        )
    )
    row = existing or DeletedProductImageCleanupModel(
        owner_username=product.owner_username,
        original_product_id=product.id,
    )
    if existing is None:
        session.add(row)
    row.store_id = store.id
    row.store_name = store.alias_name or store.store_name
    row.product_code = productCodeForError(product)
    row.cabinet_targets_json = json.dumps(cabinet_targets, ensure_ascii=False)
    row.local_image_urls_json = json.dumps(unique_texts(local_image_urls), ensure_ascii=False)
    row.status = "pending"
    row.sync_task_id = None
    row.last_error = None
    row.deleted_at = datetime.now()
    row.cleaned_at = None


def delete_product_cabinet_images(
    service_secret: str,
    license_key: str,
    raw_payload: dict[str, Any],
    shop_code: str,
) -> tuple[int, list[str]]:
    targets = product_cabinet_file_targets(raw_payload, shop_code)
    return delete_cabinet_targets(service_secret, license_key, targets)


def delete_cabinet_targets(
    service_secret: str,
    license_key: str,
    targets: list[dict[str, Any]],
) -> tuple[int, list[str]]:
    deleted_ids: set[int] = set()
    warnings: list[str] = []
    for target in targets:
        try:
            file_ids = resolve_cabinet_file_ids(service_secret, license_key, target)
        except Exception as exc:
            warnings.append(f"{target.get('filePath') or target.get('fileName')}: {exc}")
            continue
        for file_id in file_ids:
            if file_id in deleted_ids:
                continue
            try:
                delete_rakuten_cabinet_file(service_secret, license_key, file_id)
                deleted_ids.add(file_id)
            except Exception as exc:
                warnings.append(f"R-Cabinet 图片 {file_id}: {exc}")
    return len(deleted_ids), warnings


def resolve_cabinet_file_ids(service_secret: str, license_key: str, target: dict[str, str]) -> list[int]:
    file_path = normalize_text(target.get("filePath"))
    file_name = normalize_text(target.get("fileName"))
    folder_path = normalize_text(target.get("folderPath"))
    cabinet_path = normalize_text(target.get("cabinetPath"))
    if file_name:
        records = search_rakuten_cabinet_files(service_secret, license_key, file_name=file_name)
        if file_path or folder_path or cabinet_path:
            exact_ids = [
                int(record["fileId"])
                for record in records
                if cabinet_record_matches_target(record, file_path, folder_path=folder_path, cabinet_path=cabinet_path)
            ]
            if exact_ids:
                return exact_ids
        if len(records) == 1:
            return [int(records[0]["fileId"])]
    if file_path:
        try:
            records = search_rakuten_cabinet_files(service_secret, license_key, file_path=file_path)
        except RuntimeError:
            records = []
        exact_ids = [
            int(record["fileId"])
            for record in records
            if cabinet_record_matches_target(record, file_path, folder_path=folder_path, cabinet_path=cabinet_path)
        ]
        if exact_ids:
            return exact_ids
    return []


def cabinet_record_matches_target(
    record: dict[str, Any],
    file_path: str,
    *,
    folder_path: str = "",
    cabinet_path: str = "",
) -> bool:
    expected = normalize_cabinet_path(cabinet_path or file_path)
    expected_file_path = normalize_text(file_path).strip("/").lower()
    expected_folder_path = normalize_text(folder_path).strip("/").lower()
    record_file_path = normalize_text(record.get("filePath")).strip("/").lower()
    record_folder_path = normalize_text(record.get("folderPath")).strip("/").lower()
    if expected_file_path and record_file_path != expected_file_path:
        return False
    if expected_folder_path and record_folder_path != expected_folder_path:
        return False
    if expected_file_path or expected_folder_path:
        return True
    for value in (
        record.get("fileUrl"),
        cabinet_record_path(record),
        record.get("filePath"),
    ):
        if normalize_cabinet_path(value) == expected:
            return True
    return False


def cabinet_record_path(record: dict[str, Any]) -> str:
    file_name = normalize_text(record.get("fileName"))
    folder_path = normalize_text(record.get("folderPath"))
    if folder_path and file_name:
        return f"/cabinet/{folder_path.strip('/')}/{file_name}"
    file_path = normalize_text(record.get("filePath"))
    if file_path and folder_path and not file_path.lower().startswith(folder_path.strip("/").lower() + "/"):
        return f"/cabinet/{folder_path.strip('/')}/{file_path.strip('/')}"
    return file_path


def normalize_cabinet_path(value: Any) -> str:
    text = normalize_text(value).replace("\\/", "/")
    if not text:
        return ""
    if text.startswith(("http://", "https://")):
        try:
            text = urlsplit(text).path
        except Exception:
            return ""
    cabinet_index = text.lower().find("/cabinet/")
    if cabinet_index >= 0:
        text = text[cabinet_index:]
    elif text.lower().startswith("cabinet/"):
        text = "/" + text
    elif not text.startswith("/"):
        text = "/" + text
    return "/" + text.lstrip("/").split("?", 1)[0].split("#", 1)[0].lower()


def get_product_detail(owner_username: str, product_id: int) -> dict[str, Any]:
    with session_scope() as session:
        row = session.get(ProductModel, product_id)
        if row is None or row.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        return product_detail_to_public(row)


def update_store_product_price(owner_username: str, product_id: int, price: Decimal) -> dict[str, Any]:
    if price <= 0:
        raise RuntimeError("商品价格必须大于 0。")
    if price != price.to_integral_value():
        raise RuntimeError("乐天商品价格必须为日元整数，不能包含小数。")
    normalized_price = price.to_integral_value()
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        if product.review_status != "listed":
            raise RuntimeError("只有店铺商品可以同步修改乐天价格。")
        if not product.store_id:
            raise RuntimeError("商品未关联店铺，不能同步修改乐天价格。")
        manage_number = normalize_text(product.rakuten_manage_number or product.item_number)
        if not manage_number:
            raise RuntimeError("商品缺少商品管理编号，不能同步修改乐天价格。")

        store = session.get(StoreModel, product.store_id)
        if store is None:
            raise RuntimeError("商品关联店铺不存在。")
        if not store.enabled:
            raise RuntimeError("商品关联店铺已停用，不能同步修改乐天价格。")

        raw_payload = product_raw_payload(product)
        try:
            updated_payload = patch_rakuten_item_price(
                decrypt_text(store.rakuten_service_secret_encrypted),
                decrypt_text(store.rakuten_license_key_encrypted),
                manage_number,
                raw_payload,
                normalized_price,
            )
        except Exception as exc:
            product.last_error = str(exc)
            raise

        product.price = normalized_price
        product.listed_at = parse_rakuten_datetime_value(updated_payload.get("created")) or product.listed_at
        product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        product.store_last_seen_at = datetime.now()
        product.last_error = None
        session.flush()
        return product_to_public(product)


def update_store_product_detail(owner_username: str, product_id: int, payload: Any) -> dict[str, Any]:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        if product.review_status != "listed":
            raise RuntimeError("只有店铺商品可以同步修改乐天商品详情。")
        if not product.store_id:
            raise RuntimeError("商品未关联店铺，不能同步修改乐天商品详情。")
        manage_number = normalize_text(product.rakuten_manage_number or product.item_number)
        if not manage_number:
            raise RuntimeError("商品缺少商品管理编号，不能同步修改乐天商品详情。")

        store = session.get(StoreModel, product.store_id)
        if store is None:
            raise RuntimeError("商品关联店铺不存在。")
        if not store.enabled:
            raise RuntimeError("商品关联店铺已停用，不能同步修改乐天商品详情。")

        raw_payload = product_raw_payload(product)
        genre_id = normalize_text(getattr(payload, "genreId", None)) or product.genre_id
        if not re.fullmatch(r"\d{6}", genre_id) or not rakuten_genre_path(genre_id):
            raise RuntimeError("请选择有效品类。")
        try:
            updated_payload = patch_rakuten_item_detail(
                decrypt_text(store.rakuten_service_secret_encrypted),
                decrypt_text(store.rakuten_license_key_encrypted),
                manage_number,
                raw_payload,
                title=getattr(payload, "title", ""),
                tagline=getattr(payload, "tagline", ""),
                genre_id=genre_id,
                variants=list(getattr(payload, "variants", []) or []),
            )
        except Exception as exc:
            product.last_error = str(exc)
            raise

        product.title = first_text_from_keys(updated_payload, ("itemName", "title", "name")) or product.title
        product.genre_id = genre_id
        product.price = price_from_rakuten_item(updated_payload)
        product.listed_at = parse_rakuten_datetime_value(updated_payload.get("created")) or product.listed_at
        product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        product.store_last_seen_at = datetime.now()
        product.last_error = None
        session.flush()
        return product_detail_to_public(product)


def repair_store_product_images(owner_username: str, product_id: int) -> dict[str, Any]:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        if product.review_status != "listed":
            raise RuntimeError("只有店铺商品可以修复同步图片。")
        if not product.store_id:
            raise RuntimeError("商品未关联店铺，不能修复同步图片。")
        manage_number = normalize_text(product.rakuten_manage_number or product.item_number)
        if not manage_number:
            raise RuntimeError("商品缺少商品管理编号，不能修复同步图片。")
        store = session.get(StoreModel, product.store_id)
        if store is None:
            raise RuntimeError("商品关联店铺不存在。")
        if not store.enabled:
            raise RuntimeError("商品关联店铺已停用，不能修复同步图片。")

        raw_payload = product_raw_payload(product)
        source_images = listing_image_sources_for_repair(raw_payload, product_shop_code(product, raw_payload))
        if not source_images:
            raise RuntimeError("没有找到可重新同步的原始商品图片。")
        service_secret = decrypt_text(store.rakuten_service_secret_encrypted)
        license_key = decrypt_text(store.rakuten_license_key_encrypted)
        cabinet_context: dict[str, Any] = {}
        uploaded_images: list[dict[str, str]] = []
        try:
            uploaded_images = upload_product_images_to_rakuten(
                service_secret,
                license_key,
                store,
                product,
                manage_number,
                cabinet_context=cabinet_context,
                source_images=source_images,
            )
            patch_rakuten_item_images(service_secret, license_key, manage_number, uploaded_images, product.title)
        except Exception as exc:
            rollback_message = rollback_uploaded_listing_images(service_secret, license_key, uploaded_images)
            product.last_error = f"{exc}；已回滚本次上传图片：{rollback_message}" if rollback_message else str(exc)
            raise RuntimeError(product.last_error) from exc

        edited_images = [
            build_rakuten_cabinet_image_url(store.store_code, image["location"])
            for image in uploaded_images
            if image.get("location")
        ]
        updated_payload = dict(raw_payload)
        updated_payload["images"] = uploaded_images
        updated_payload["ltEditedImages"] = edited_images
        updated_payload["updated"] = datetime.now().isoformat(timespec="seconds")
        product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        product.image_url = edited_images[0] if edited_images else product.image_url
        product.store_last_seen_at = datetime.now()
        product.last_error = None
        session.flush()
        return product_detail_to_public(product)


def listing_image_sources_for_repair(raw_payload: dict[str, Any], shop_code: str) -> list[str]:
    sources: list[str] = []
    for image in raw_payload.get("images") if isinstance(raw_payload.get("images"), list) else []:
        source_url = image.get("sourceUrl") if isinstance(image, dict) else None
        url = normalize_product_image_url(source_url, shop_code=shop_code)
        if url and url not in sources:
            sources.append(url)
    if sources:
        return sources[:RAKUTEN_LISTING_IMAGE_LIMIT]
    fallback_payload = dict(raw_payload)
    fallback_payload.pop("ltEditedImages", None)
    fallback_payload.pop("images", None)
    return product_image_urls(fallback_payload, shop_code=shop_code)[:RAKUTEN_LISTING_IMAGE_LIMIT]


def request_rakuten_write(
    method: str,
    url: str,
    *,
    headers: dict[str, str],
    operation: str,
    **kwargs: Any,
) -> requests.Response:
    timeout = max(int(settings.rakuten_write_timeout_seconds), int(settings.crawler_timeout_seconds))
    last_error: Exception | None = None
    account_key = rakuten_request_account_key(headers)
    throttle_rakuten_write_request(account_key)
    for attempt in range(RAKUTEN_WRITE_MAX_RETRIES):
        try:
            response = requests.request(method, url, timeout=timeout, headers=headers, **kwargs)
        except (requests.Timeout, requests.ConnectionError) as exc:
            last_error = exc
            if attempt < RAKUTEN_WRITE_MAX_RETRIES - 1:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise RuntimeError(f"{operation}超时或连接失败，已重试 {RAKUTEN_WRITE_MAX_RETRIES} 次：{exc}") from exc

        qps_limited = is_rakuten_write_qps_limited_response(response)
        retryable = response.status_code in RAKUTEN_WRITE_RETRY_STATUS_CODES or qps_limited
        if retryable and attempt < RAKUTEN_WRITE_MAX_RETRIES - 1:
            response.close()
            time.sleep(rakuten_write_backoff_seconds(attempt, qps_limited=qps_limited))
            continue
        return response
    if last_error is not None:
        raise RuntimeError(f"{operation}失败：{last_error}") from last_error
    raise RuntimeError(f"{operation}失败。")


def throttle_rakuten_write_request(account_key: str = "default") -> None:
    min_interval = max(0.0, RAKUTEN_WRITE_REQUEST_MIN_INTERVAL_SECONDS)
    if min_interval <= 0:
        return
    if wait_for_distributed_request_slot(f"write:{account_key}", min_interval):
        return
    with RAKUTEN_WRITE_REQUEST_LOCK:
        now = time.monotonic()
        last_at = RAKUTEN_WRITE_LAST_REQUEST_AT_BY_ACCOUNT.get(account_key, 0.0)
        elapsed = now - last_at
        if elapsed < min_interval:
            time.sleep(min_interval - elapsed)
        RAKUTEN_WRITE_LAST_REQUEST_AT_BY_ACCOUNT[account_key] = time.monotonic()


def wait_for_rakuten_item_delete_slot(
    min_interval_seconds: float = RAKUTEN_ITEM_DELETE_MIN_INTERVAL_SECONDS,
) -> None:
    global RAKUTEN_ITEM_DELETE_LAST_REQUEST_AT
    min_interval = max(0.0, float(min_interval_seconds or 0.0))
    if min_interval <= 0:
        return
    with RAKUTEN_ITEM_DELETE_REQUEST_LOCK:
        elapsed = time.monotonic() - RAKUTEN_ITEM_DELETE_LAST_REQUEST_AT
        if elapsed < min_interval:
            time.sleep(min_interval - elapsed)
        RAKUTEN_ITEM_DELETE_LAST_REQUEST_AT = time.monotonic()


def is_rakuten_write_qps_limited_response(response: requests.Response) -> bool:
    if int(getattr(response, "status_code", 0) or 0) == 429:
        return True
    detail = normalize_text(getattr(response, "text", "")).upper()
    return any(
        marker in detail
        for marker in (
            "GA0003",
            "QPSLIMIT",
            "QPS LIMIT",
            "EXCEEDED QPS",
        )
    )


def rakuten_write_backoff_seconds(attempt: int, *, qps_limited: bool = False) -> float:
    if not qps_limited:
        return 1.5 * (max(0, int(attempt)) + 1)
    index = max(0, min(len(RAKUTEN_WRITE_QPS_BACKOFF_SECONDS) - 1, int(attempt)))
    return float(RAKUTEN_WRITE_QPS_BACKOFF_SECONDS[index])


def put_rakuten_item(
    service_secret: str,
    license_key: str,
    manage_number: str,
    payload: dict[str, Any],
) -> None:
    if not service_secret or not license_key:
        raise RuntimeError("乐天 Secret 或乐天 Key 未配置。")
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number:
        raise RuntimeError("商品管理编号为空，不能创建乐天商品。")
    response = request_rakuten_write(
        "PUT",
        RAKUTEN_ITEM_PATCH_URL.format(manageNumber=quote(normalized_manage_number, safe="")),
        operation=f"乐天商品 {normalized_manage_number} 创建",
        headers={
            "Authorization": build_rakuten_authorization_header(service_secret, license_key),
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        },
        json=payload,
    )
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = normalize_text(response.text)
        message = f"乐天商品 {normalized_manage_number} 创建失败"
        if detail:
            message = f"{message}：{detail[:800]}"
        raise RuntimeError(message) from exc


def put_rakuten_item_with_attribute_retry(
    service_secret: str,
    license_key: str,
    manage_number: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    current_payload = payload
    last_error: RuntimeError | None = None
    for _ in range(3):
        try:
            put_rakuten_item(service_secret, license_key, manage_number, current_payload)
            return current_payload
        except RuntimeError as exc:
            last_error = exc
            patched_payload = patch_payload_for_attribute_errors(current_payload, str(exc))
            if patched_payload == current_payload:
                raise
            current_payload = patched_payload
    if last_error is not None:
        raise last_error
    return current_payload


def patch_payload_for_attribute_errors(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    patched_payload = patch_payload_for_machine_dependent_character_errors(payload, error_text)
    patched_payload = patch_payload_for_invalid_selective_attribute_values(patched_payload, error_text)
    patched_payload = patch_payload_for_unknown_attribute_name_errors(patched_payload, error_text)
    patched_payload = patch_payload_for_missing_mandatory_attributes(patched_payload, error_text)
    patched_payload = patch_payload_for_attribute_unit_errors(patched_payload, error_text)
    patched_payload = patch_payload_for_attribute_string_value_errors(patched_payload, error_text)
    return patched_payload


def normalize_rakuten_machine_dependent_characters(value: Any) -> str:
    text = str(value or "")
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKC", text.translate(RAKUTEN_MACHINE_DEPENDENT_TRANSLATION))
    cleaned: list[str] = []
    for char in normalized:
        codepoint = ord(char)
        category = unicodedata.category(char)
        if char in {"\n", "\r", "\t"}:
            cleaned.append(char)
            continue
        if category in {"Cc", "Cf", "Co", "Cs", "Cn"}:
            continue
        if (
            0x1F000 <= codepoint <= 0x1FAFF
            or 0x2600 <= codepoint <= 0x27BF
            or 0xFE00 <= codepoint <= 0xFE0F
            or 0xE0100 <= codepoint <= 0xE01EF
            or codepoint in {0x3030, 0x303D, 0x3297, 0x3299}
        ):
            continue
        cleaned.append(char)
    return "".join(cleaned)


def sanitize_rakuten_payload_text(value: Any) -> Any:
    if isinstance(value, str):
        return normalize_rakuten_machine_dependent_characters(value)
    if isinstance(value, list):
        return [sanitize_rakuten_payload_text(item) for item in value]
    if isinstance(value, dict):
        return {key: sanitize_rakuten_payload_text(item) for key, item in value.items()}
    return value


def sanitize_rakuten_image_alt(value: Any, *, max_length: int = 255) -> str:
    text = unescape(str(value or ""))
    if not text:
        return ""
    soup = BeautifulSoup(text, "lxml")
    text = soup.get_text(" ", strip=True)
    text = re.sub(r"<[^>]*>", " ", text)
    text = normalize_rakuten_machine_dependent_characters(normalize_text(text))
    return truncate_text(text, max_length)


def patch_payload_for_machine_dependent_character_errors(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    paths = extract_machine_dependent_character_error_paths(error_text)
    if not paths:
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    changed = False
    for path in paths:
        if sanitize_payload_value_at_path(patched, path):
            changed = True
    return patched if changed else payload


def extract_machine_dependent_character_error_paths(error_text: str) -> list[str]:
    paths: list[str] = []

    def collect(value: Any) -> None:
        if isinstance(value, dict):
            if normalize_text(value.get("code")) == "IE0270":
                metadata = value.get("metadata") if isinstance(value.get("metadata"), dict) else {}
                path = normalize_text(metadata.get("propertyPath"))
                if path and path not in paths:
                    paths.append(path)
            for child in value.values():
                collect(child)
        elif isinstance(value, list):
            for item in value:
                collect(item)

    text = normalize_text(error_text)
    try:
        json_start = text.find("{")
        if json_start >= 0:
            collect(json.loads(text[json_start:]))
    except Exception:
        pass
    for match in re.finditer(r'"code"\s*:\s*"IE0270".*?"propertyPath"\s*:\s*"([^"]+)"', text):
        path = normalize_text(match.group(1))
        if path and path not in paths:
            paths.append(path)
    return paths


def sanitize_payload_value_at_path(payload: dict[str, Any], path: str) -> bool:
    target = payload
    parts = [part for part in normalize_text(path).split(".") if part]
    if not parts:
        return False
    for part in parts[:-1]:
        if not isinstance(target, dict):
            return False
        next_value = target.get(part)
        if not isinstance(next_value, dict):
            return False
        target = next_value
    if not isinstance(target, dict):
        return False
    key = parts[-1]
    if key not in target:
        return False
    next_value = sanitize_rakuten_payload_text(target.get(key))
    if next_value == target.get(key):
        return False
    target[key] = next_value
    return True


def patch_payload_for_invalid_selective_attribute_values(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    errors = extract_invalid_selective_attribute_value_errors(error_text)
    if not errors:
        return payload
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    patched_variants = patched.get("variants")
    if not isinstance(patched_variants, dict):
        return payload
    rule_map = rakuten_attribute_rule_map_for_payload(patched)
    changed = False
    optional_attribute_names = {
        attribute_name
        for error in errors
        if (attribute_name := normalize_text(error.get("attributeName")))
        and attribute_name != RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE
        and not bool(rule_map.get(attribute_name, {}).get("required"))
    }
    if optional_attribute_names:
        for variant in patched_variants.values():
            if not isinstance(variant, dict):
                continue
            attributes = variant.get("attributes")
            if not isinstance(attributes, list):
                continue
            next_attributes = [
                attribute
                for attribute in attributes
                if not (
                    isinstance(attribute, dict)
                    and normalize_text(attribute.get("name")) in optional_attribute_names
                )
            ]
            if len(next_attributes) != len(attributes):
                variant["attributes"] = next_attributes
                changed = True
    for error in errors:
        attribute_name = normalize_text(error.get("attributeName"))
        if not attribute_name or attribute_name in optional_attribute_names:
            continue
        variant = patched_variants.get(normalize_text(error.get("variantId")))
        if not isinstance(variant, dict):
            continue
        attributes = variant.get("attributes")
        if not isinstance(attributes, list):
            continue
        attribute_index = error.get("attributeIndex")
        targets: list[dict[str, Any]] = []
        if isinstance(attribute_index, int) and 0 <= attribute_index < len(attributes):
            attribute = attributes[attribute_index]
            if isinstance(attribute, dict) and (
                not attribute_name or normalize_text(attribute.get("name")) == attribute_name
            ):
                targets.append(attribute)
        targets.extend(
            attribute
            for attribute in attributes
            if isinstance(attribute, dict)
            and normalize_text(attribute.get("name")) == attribute_name
            and attribute not in targets
        )
        for attribute in targets:
            current_value = first_text_value(attribute.get("values"))
            if attribute_name == RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE:
                next_value = normalize_rakuten_representative_color(current_value) or RAKUTEN_REPRESENTATIVE_COLOR_FALLBACK
                if next_value not in RAKUTEN_REPRESENTATIVE_COLOR_ALLOWED_VALUES:
                    next_value = RAKUTEN_REPRESENTATIVE_COLOR_FALLBACK
                if attribute.get("values") != [next_value]:
                    attribute["values"] = [next_value]
                    changed = True
                if "unit" in attribute:
                    attribute.pop("unit", None)
                    changed = True
                continue
            next_value = normalize_rakuten_selective_attribute_value(attribute_name, current_value)
            if next_value:
                if attribute.get("values") != [next_value]:
                    attribute["values"] = [next_value]
                    changed = True
                if "unit" in attribute:
                    attribute.pop("unit", None)
                    changed = True
                continue
    return patched if changed else payload


def extract_invalid_selective_attribute_value_errors(error_text: str) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []

    def add_error(property_path: Any, attribute_name: Any) -> None:
        normalized_path = normalize_text(property_path)
        match = re.fullmatch(r"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]", normalized_path)
        if not match:
            return
        variant_id, attribute_index = match.groups()
        error = {
            "variantId": normalize_text(variant_id),
            "attributeIndex": int(attribute_index),
            "attributeName": normalize_text(attribute_name),
        }
        if error["variantId"] and error["attributeName"] and error not in errors:
            errors.append(error)

    def collect(value: Any) -> None:
        if isinstance(value, dict):
            metadata = value.get("metadata") if isinstance(value.get("metadata"), dict) else {}
            property_path = metadata.get("propertyPath")
            details = metadata.get("details")
            detail_items = details if isinstance(details, list) else [value]
            for detail in detail_items:
                if not isinstance(detail, dict):
                    continue
                if normalize_text(detail.get("code")) != "invalidSelectiveValue":
                    continue
                properties = detail.get("properties") if isinstance(detail.get("properties"), dict) else {}
                add_error(property_path, properties.get("attributeName"))
            for child in value.values():
                collect(child)
        elif isinstance(value, list):
            for item in value:
                collect(item)

    text = normalize_text(error_text)
    try:
        json_start = text.find("{")
        if json_start >= 0:
            collect(json.loads(text[json_start:]))
    except Exception:
        pass
    pattern = (
        r'"code"\s*:\s*"invalidSelectiveValue".*?'
        r'"attributeName"\s*:\s*"([^"]+)".*?'
        r'"propertyPath"\s*:\s*"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]"'
    )
    for match in re.finditer(pattern, text):
        attribute_name, variant_id, attribute_index = match.groups()
        error = {
            "variantId": normalize_text(variant_id),
            "attributeIndex": int(attribute_index),
            "attributeName": normalize_text(attribute_name),
        }
        if error["variantId"] and error["attributeName"] and error not in errors:
            errors.append(error)
    if not errors and "invalidSelectiveValue" in text:
        path_pattern = r"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]"
        name_match = re.search(r'"attributeName"\s*:\s*"([^"]+)"', text)
        if name_match:
            for match in re.finditer(path_pattern, text):
                variant_id, attribute_index = match.groups()
                error = {
                    "variantId": normalize_text(variant_id),
                    "attributeIndex": int(attribute_index),
                    "attributeName": normalize_text(name_match.group(1)),
                }
                if error["variantId"] and error not in errors:
                    errors.append(error)
    return errors


_RAKUTEN_ATTRIBUTE_RULES_CACHE: dict[str, Any] | None = None
_RAKUTEN_GENRE_ZH_MAP_CACHE: dict[str, str] | None = None
_RAKUTEN_GENRE_NODE_INDEX_CACHE: dict[str, dict[str, Any]] | None = None


def load_rakuten_attribute_rules() -> dict[str, Any]:
    global _RAKUTEN_ATTRIBUTE_RULES_CACHE
    if _RAKUTEN_ATTRIBUTE_RULES_CACHE is not None:
        return _RAKUTEN_ATTRIBUTE_RULES_CACHE
    if not RAKUTEN_ATTRIBUTE_RULES_PATH.exists():
        _RAKUTEN_ATTRIBUTE_RULES_CACHE = {}
        return _RAKUTEN_ATTRIBUTE_RULES_CACHE
    try:
        _RAKUTEN_ATTRIBUTE_RULES_CACHE = json.loads(RAKUTEN_ATTRIBUTE_RULES_PATH.read_text(encoding="utf-8"))
    except Exception:
        _RAKUTEN_ATTRIBUTE_RULES_CACHE = {}
    return _RAKUTEN_ATTRIBUTE_RULES_CACHE


def load_rakuten_genre_zh_map() -> dict[str, str]:
    global _RAKUTEN_GENRE_ZH_MAP_CACHE
    if _RAKUTEN_GENRE_ZH_MAP_CACHE is not None:
        return _RAKUTEN_GENRE_ZH_MAP_CACHE
    if not RAKUTEN_GENRE_ZH_MAP_PATH.exists():
        _RAKUTEN_GENRE_ZH_MAP_CACHE = {}
        return _RAKUTEN_GENRE_ZH_MAP_CACHE
    try:
        payload = json.loads(RAKUTEN_GENRE_ZH_MAP_PATH.read_text(encoding="utf-8"))
    except Exception:
        payload = {}
    translations = payload.get("translations") if isinstance(payload, dict) else {}
    _RAKUTEN_GENRE_ZH_MAP_CACHE = {
        normalize_text(source): normalize_text(target)
        for source, target in translations.items()
        if normalize_text(source) and normalize_text(target)
    } if isinstance(translations, dict) else {}
    return _RAKUTEN_GENRE_ZH_MAP_CACHE


def rakuten_genre_zh_path(genre_path: Any) -> str:
    normalized_path = normalize_text(genre_path)
    if not normalized_path:
        return ""
    translations = load_rakuten_genre_zh_map()
    return ">".join(
        translations.get(segment, segment)
        for segment in (normalize_text(part) for part in normalized_path.split(">"))
        if segment
    )


def rakuten_genre_path(genre_id: Any) -> str:
    normalized_genre_id = normalize_text(genre_id)
    if not normalized_genre_id:
        return ""
    rules = load_rakuten_attribute_rules()
    genres = rules.get("genres") if isinstance(rules.get("genres"), dict) else {}
    genre = genres.get(normalized_genre_id)
    return normalize_text(genre.get("genrePath")) if isinstance(genre, dict) else ""


def search_rakuten_genres(keyword: str = "", limit: int = 30) -> list[dict[str, str]]:
    folded_keyword = normalize_text(keyword).casefold()
    bounded_limit = min(max(int(limit or 30), 1), 100)
    rules = load_rakuten_attribute_rules()
    genres = rules.get("genres") if isinstance(rules.get("genres"), dict) else {}
    results: list[dict[str, str]] = []
    for genre_id, genre in genres.items():
        if not isinstance(genre, dict):
            continue
        normalized_genre_id = normalize_text(genre_id)
        genre_path = normalize_text(genre.get("genrePath"))
        genre_path_zh = rakuten_genre_zh_path(genre_path)
        if not normalized_genre_id or not genre_path:
            continue
        if folded_keyword and (
            folded_keyword not in normalized_genre_id.casefold()
            and folded_keyword not in genre_path.casefold()
            and folded_keyword not in genre_path_zh.casefold()
        ):
            continue
        results.append({
            "genreId": normalized_genre_id,
            "genrePath": genre_path,
            "genrePathZh": genre_path_zh,
        })
        if len(results) >= bounded_limit:
            break
    return results


def list_rakuten_genre_children(parent_path: str = "") -> list[dict[str, Any]]:
    normalized_parent = normalize_text(parent_path)
    parent_parts = [normalize_text(part) for part in normalized_parent.split(">") if normalize_text(part)]
    rules = load_rakuten_attribute_rules()
    genres = rules.get("genres") if isinstance(rules.get("genres"), dict) else {}
    nodes: dict[str, dict[str, Any]] = {}
    for genre_id, genre in genres.items():
        if not isinstance(genre, dict):
            continue
        genre_path = normalize_text(genre.get("genrePath"))
        parts = [normalize_text(part) for part in genre_path.split(">") if normalize_text(part)]
        if len(parts) <= len(parent_parts) or parts[:len(parent_parts)] != parent_parts:
            continue
        label = parts[len(parent_parts)]
        child_parts = [*parent_parts, label]
        child_path = ">".join(child_parts)
        node = nodes.setdefault(
            child_path,
            {
                "label": label,
                "labelZh": load_rakuten_genre_zh_map().get(label, label),
                "genrePath": child_path,
                "genrePathZh": rakuten_genre_zh_path(child_path),
                "genreId": "",
                "leaf": True,
            },
        )
        if len(parts) > len(child_parts):
            node["leaf"] = False
        elif genre_path == child_path:
            node["genreId"] = normalize_text(genre_id)
    return sorted(nodes.values(), key=lambda item: normalize_text(item["label"]).casefold())


def normalize_collection_genre_policy(value: Any, *, field_label: str = "品类策略") -> str:
    policy = normalize_text(value).lower()
    if policy not in {"allow", "deny"}:
        raise RuntimeError(f"{field_label}必须是允许采集或禁止采集。")
    return policy


def collection_genre_path_hash(genre_path: Any) -> str:
    return hashlib.sha256(normalize_text(genre_path).encode("utf-8")).hexdigest()


def rakuten_genre_node_index() -> dict[str, dict[str, Any]]:
    global _RAKUTEN_GENRE_NODE_INDEX_CACHE
    if _RAKUTEN_GENRE_NODE_INDEX_CACHE is not None:
        return _RAKUTEN_GENRE_NODE_INDEX_CACHE
    rules = load_rakuten_attribute_rules()
    genres = rules.get("genres") if isinstance(rules.get("genres"), dict) else {}
    nodes: dict[str, dict[str, Any]] = {}
    translations = load_rakuten_genre_zh_map()
    for genre_id, genre in genres.items():
        if not isinstance(genre, dict):
            continue
        genre_path = normalize_text(genre.get("genrePath"))
        parts = [normalize_text(part) for part in genre_path.split(">") if normalize_text(part)]
        for index, label in enumerate(parts):
            path_parts = parts[:index + 1]
            path = ">".join(path_parts)
            node = nodes.setdefault(
                path,
                {
                    "label": label,
                    "labelZh": translations.get(label, label),
                    "genrePath": path,
                    "genrePathZh": rakuten_genre_zh_path(path),
                    "genreId": "",
                    "leaf": True,
                },
            )
            if index < len(parts) - 1:
                node["leaf"] = False
            elif genre_path == path:
                node["genreId"] = normalize_text(genre_id)
    _RAKUTEN_GENRE_NODE_INDEX_CACHE = nodes
    return _RAKUTEN_GENRE_NODE_INDEX_CACHE


def collection_genre_ancestor_paths(genre_path: Any) -> list[str]:
    parts = [normalize_text(part) for part in normalize_text(genre_path).split(">") if normalize_text(part)]
    return [">".join(parts[:index]) for index in range(1, len(parts) + 1)]


def collection_genre_policy_snapshot(session: Any, owner_username: str) -> CollectionGenrePolicySnapshot:
    setting = session.get(UserCollectionGenreSettingModel, owner_username)
    default_policy = normalize_collection_genre_policy(
        setting.default_policy if setting is not None else "allow",
        field_label="默认品类策略",
    )
    unknown_genre_policy = normalize_collection_genre_policy(
        setting.unknown_genre_policy if setting is not None else "allow",
        field_label="未识别品类策略",
    )
    rows = session.scalars(
        select(UserCollectionGenreRuleModel).where(
            UserCollectionGenreRuleModel.owner_username == owner_username,
        )
    ).all()
    return CollectionGenrePolicySnapshot(
        default_policy=default_policy,
        unknown_genre_policy=unknown_genre_policy,
        rules_by_path={
            normalize_text(row.genre_path): normalize_collection_genre_policy(row.policy)
            for row in rows
            if normalize_text(row.genre_path)
        },
    )


def resolve_collection_genre_policy(
    snapshot: CollectionGenrePolicySnapshot,
    *,
    genre_id: Any = "",
    genre_path: Any = "",
) -> dict[str, Any]:
    normalized_genre_id = normalize_text(genre_id)
    normalized_genre_path = normalize_text(genre_path) or rakuten_genre_path(normalized_genre_id)
    if not normalized_genre_path:
        return {
            "allowed": snapshot.unknown_genre_policy == "allow",
            "policy": snapshot.unknown_genre_policy,
            "genreId": normalized_genre_id,
            "genrePath": "",
            "genrePathZh": "",
            "sourcePath": "",
            "sourcePathZh": "",
            "sourceType": "unknown",
        }
    effective_policy = snapshot.default_policy
    source_path = ""
    rules_by_path = snapshot.rules_by_path or {}
    for ancestor_path in collection_genre_ancestor_paths(normalized_genre_path):
        policy = rules_by_path.get(ancestor_path)
        if policy in {"allow", "deny"}:
            effective_policy = policy
            source_path = ancestor_path
    return {
        "allowed": effective_policy == "allow",
        "policy": effective_policy,
        "genreId": normalized_genre_id,
        "genrePath": normalized_genre_path,
        "genrePathZh": rakuten_genre_zh_path(normalized_genre_path),
        "sourcePath": source_path,
        "sourcePathZh": rakuten_genre_zh_path(source_path),
        "sourceType": "rule" if source_path else "default",
    }


def collected_item_genre_policy_decision(
    item: dict[str, Any],
    snapshot: CollectionGenrePolicySnapshot,
) -> dict[str, Any]:
    raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
    genre_id = (
        first_text_from_keys(item, ("genreId", "genre_id", "genre"))
        or first_text_from_keys(raw, ("genreId", "genre_id", "genre", "rCategoryId"))
    )
    return resolve_collection_genre_policy(snapshot, genre_id=genre_id)


def user_collection_genre_config(owner_username: str) -> dict[str, Any]:
    with session_scope() as session:
        snapshot = collection_genre_policy_snapshot(session, owner_username)
        rule_count = session.scalar(
            select(func.count(UserCollectionGenreRuleModel.id)).where(
                UserCollectionGenreRuleModel.owner_username == owner_username,
            )
        )
    return {
        "defaultPolicy": snapshot.default_policy,
        "unknownGenrePolicy": snapshot.unknown_genre_policy,
        "ruleCount": int(rule_count or 0),
    }


def save_user_collection_genre_config(
    owner_username: str,
    *,
    default_policy: Any,
    unknown_genre_policy: Any,
) -> dict[str, Any]:
    normalized_default = normalize_collection_genre_policy(default_policy, field_label="默认品类策略")
    normalized_unknown = normalize_collection_genre_policy(unknown_genre_policy, field_label="未识别品类策略")
    with session_scope() as session:
        row = session.get(UserCollectionGenreSettingModel, owner_username)
        if row is None:
            row = UserCollectionGenreSettingModel(owner_username=owner_username)
            session.add(row)
        row.default_policy = normalized_default
        row.unknown_genre_policy = normalized_unknown
        session.flush()
        rule_count = session.scalar(
            select(func.count(UserCollectionGenreRuleModel.id)).where(
                UserCollectionGenreRuleModel.owner_username == owner_username,
            )
        )
    return {
        "defaultPolicy": normalized_default,
        "unknownGenrePolicy": normalized_unknown,
        "ruleCount": int(rule_count or 0),
    }


def collection_genre_rule_rows_by_path(session: Any, owner_username: str) -> dict[str, UserCollectionGenreRuleModel]:
    rows = session.scalars(
        select(UserCollectionGenreRuleModel).where(
            UserCollectionGenreRuleModel.owner_username == owner_username,
        )
    ).all()
    return {normalize_text(row.genre_path): row for row in rows if normalize_text(row.genre_path)}


def collection_genre_node_to_public(
    node: dict[str, Any],
    snapshot: CollectionGenrePolicySnapshot,
    rule_rows_by_path: dict[str, UserCollectionGenreRuleModel],
) -> dict[str, Any]:
    genre_path = normalize_text(node.get("genrePath"))
    decision = resolve_collection_genre_policy(
        snapshot,
        genre_id=node.get("genreId"),
        genre_path=genre_path,
    )
    rule = rule_rows_by_path.get(genre_path)
    return {
        **node,
        "ruleId": int(rule.id) if rule is not None else None,
        "explicitPolicy": normalize_collection_genre_policy(rule.policy) if rule is not None else "inherit",
        "effectivePolicy": decision["policy"],
        "inheritedFromPath": decision["sourcePath"],
        "inheritedFromPathZh": decision["sourcePathZh"],
        "policySourceType": decision["sourceType"],
    }


def list_user_collection_genre_children(owner_username: str, parent_path: str = "") -> list[dict[str, Any]]:
    nodes = list_rakuten_genre_children(parent_path)
    with session_scope() as session:
        snapshot = collection_genre_policy_snapshot(session, owner_username)
        rule_rows = collection_genre_rule_rows_by_path(session, owner_username)
        return [collection_genre_node_to_public(node, snapshot, rule_rows) for node in nodes]


def search_user_collection_genres(
    owner_username: str,
    keyword: str = "",
    limit: int = 50,
) -> list[dict[str, Any]]:
    folded_keyword = normalize_text(keyword).casefold()
    bounded_limit = min(max(int(limit or 50), 1), 100)
    nodes = list(rakuten_genre_node_index().values())
    if folded_keyword:
        nodes = [
            node
            for node in nodes
            if (
                folded_keyword in normalize_text(node.get("genreId")).casefold()
                or folded_keyword in normalize_text(node.get("genrePath")).casefold()
                or folded_keyword in normalize_text(node.get("genrePathZh")).casefold()
            )
        ]
    nodes.sort(
        key=lambda node: (
            0 if normalize_text(node.get("genreId")) == normalize_text(keyword) else 1,
            len(collection_genre_ancestor_paths(node.get("genrePath"))),
            normalize_text(node.get("genrePathZh") or node.get("genrePath")).casefold(),
        )
    )
    with session_scope() as session:
        snapshot = collection_genre_policy_snapshot(session, owner_username)
        rule_rows = collection_genre_rule_rows_by_path(session, owner_username)
        return [
            collection_genre_node_to_public(node, snapshot, rule_rows)
            for node in nodes[:bounded_limit]
        ]


def save_user_collection_genre_rule(
    owner_username: str,
    *,
    genre_path: Any,
    genre_id: Any = "",
    policy: Any,
) -> dict[str, Any]:
    normalized_path = normalize_text(genre_path)
    node = rakuten_genre_node_index().get(normalized_path)
    if node is None:
        raise RuntimeError("采集品类不存在或路径无效。")
    normalized_policy = normalize_collection_genre_policy(policy)
    normalized_genre_id = normalize_text(genre_id)
    canonical_genre_id = normalize_text(node.get("genreId"))
    if normalized_genre_id and canonical_genre_id and normalized_genre_id != canonical_genre_id:
        raise RuntimeError("品类编号与品类路径不匹配。")
    path_hash = collection_genre_path_hash(normalized_path)
    with session_scope() as session:
        row = session.scalar(
            select(UserCollectionGenreRuleModel).where(
                UserCollectionGenreRuleModel.owner_username == owner_username,
                UserCollectionGenreRuleModel.genre_path_hash == path_hash,
            )
        )
        if row is None:
            row = UserCollectionGenreRuleModel(
                owner_username=owner_username,
                genre_path=normalized_path,
                genre_path_hash=path_hash,
            )
            session.add(row)
        row.genre_path = normalized_path
        row.genre_path_hash = path_hash
        row.genre_id = canonical_genre_id
        row.policy = normalized_policy
        session.flush()
        snapshot = collection_genre_policy_snapshot(session, owner_username)
        rule_rows = collection_genre_rule_rows_by_path(session, owner_username)
        return collection_genre_node_to_public(node, snapshot, rule_rows)


def delete_user_collection_genre_rule(owner_username: str, rule_id: int) -> bool:
    with session_scope() as session:
        row = session.get(UserCollectionGenreRuleModel, int(rule_id))
        if row is None or row.owner_username != owner_username:
            return False
        session.delete(row)
        return True


def user_collection_genre_pending_impact(owner_username: str, sample_limit: int = 20) -> dict[str, Any]:
    with session_scope() as session:
        snapshot = collection_genre_policy_snapshot(session, owner_username)
        rows = session.scalars(
            select(ProductModel)
            .where(
                ProductModel.owner_username == owner_username,
                ProductModel.review_status == "pending",
            )
            .order_by(ProductModel.id.desc())
        ).all()
    denied_count = 0
    unknown_count = 0
    samples: list[dict[str, Any]] = []
    for row in rows:
        decision = resolve_collection_genre_policy(snapshot, genre_id=row.genre_id)
        if not decision["genrePath"]:
            unknown_count += 1
        if decision["allowed"]:
            continue
        denied_count += 1
        if len(samples) < max(1, min(int(sample_limit or 20), 50)):
            samples.append({
                "productId": int(row.id),
                "title": row.title,
                "genreId": normalize_text(row.genre_id),
                "genrePath": decision["genrePath"],
                "genrePathZh": decision["genrePathZh"],
            })
    return {
        "totalPendingCount": len(rows),
        "deniedCount": denied_count,
        "unknownGenreCount": unknown_count,
        "samples": samples,
    }


def rakuten_attribute_group_rule_for_payload(payload: dict[str, Any]) -> dict[str, Any]:
    genre_id = first_text_from_keys(payload, ("genreId", "genre_id", "genre"))
    if not genre_id:
        return {}
    rules = load_rakuten_attribute_rules()
    genres = rules.get("genres") if isinstance(rules.get("genres"), dict) else {}
    genre = genres.get(genre_id)
    if not isinstance(genre, dict):
        return {}
    group_key = normalize_text(genre.get("groupKey"))
    groups = rules.get("groups") if isinstance(rules.get("groups"), dict) else {}
    group = groups.get(group_key)
    return group if isinstance(group, dict) else {}


def rakuten_attribute_rule_map_for_payload(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    group = rakuten_attribute_group_rule_for_payload(payload)
    attributes = group.get("attributes") if isinstance(group.get("attributes"), dict) else {}
    return {normalize_text(name): rule for name, rule in attributes.items() if isinstance(rule, dict)}


def rakuten_either_required_attribute_names(rule_map: dict[str, dict[str, Any]]) -> list[str]:
    return [
        name
        for name, rule in rule_map.items()
        if "いずれか必須" in normalize_text(rule.get("requirement"))
    ]


def rakuten_attribute_rule_for_name(payload: dict[str, Any], name: str) -> dict[str, Any]:
    return rakuten_attribute_rule_map_for_payload(payload).get(normalize_text(name), {})


def rakuten_attribute_default_unit(payload: dict[str, Any], name: str) -> str:
    rule = rakuten_attribute_rule_for_name(payload, name)
    if bool(rule.get("unitRequired")):
        unit = normalize_rakuten_attribute_unit(rule.get("unit"))
        if unit:
            return unit
    return RAKUTEN_ATTRIBUTE_DEFAULT_UNITS.get(normalize_text(name), "")


def apply_rakuten_attribute_rules_to_payload(payload: dict[str, Any]) -> dict[str, Any]:
    rule_map = rakuten_attribute_rule_map_for_payload(payload)
    if not rule_map:
        return payload
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    patched_variants = patched.get("variants")
    if not isinstance(patched_variants, dict):
        return payload
    changed = False
    for variant in patched_variants.values():
        if not isinstance(variant, dict):
            continue
        attributes = variant.get("attributes")
        if not isinstance(attributes, list):
            attributes = []
            variant["attributes"] = attributes
        existing: dict[str, dict[str, Any]] = {}
        for attribute in list(attributes):
            if not isinstance(attribute, dict):
                attributes.remove(attribute)
                changed = True
                continue
            name = normalize_text(attribute.get("name"))
            if not name or name not in rule_map:
                attributes.remove(attribute)
                changed = True
                continue
            if name in existing:
                attributes.remove(attribute)
                changed = True
                continue
            existing[name] = attribute
        for name, rule in rule_map.items():
            if not bool(rule.get("required")) or name in existing:
                continue
            attribute = infer_missing_mandatory_attribute(name, variant, patched, rule)
            if not attribute:
                continue
            attributes.append(attribute)
            existing[name] = attribute
            changed = True
        either_required_names = rakuten_either_required_attribute_names(rule_map)
        if either_required_names and not any(
            name in existing and rakuten_attribute_has_effective_values(existing[name], rule_map.get(name) or {})
            for name in either_required_names
        ):
            for name in either_required_names:
                attribute = infer_missing_mandatory_attribute(name, variant, patched, rule_map.get(name))
                if not attribute:
                    continue
                attributes.append(attribute)
                existing[name] = attribute
                changed = True
                break
        for name, attribute in list(existing.items()):
            rule = rule_map.get(name)
            if not rule:
                continue
            if bool(rule.get("unitRequired")):
                values = attribute.get("values")
                first_value = first_text_value(values)
                next_values, next_unit = normalize_rakuten_attribute_values_and_unit(
                    name,
                    first_value,
                    normalize_text(attribute.get("unit")),
                    default_unit=normalize_text(rule.get("unit")),
                )
                if next_values and next_unit:
                    if attribute.get("values") != next_values:
                        attribute["values"] = next_values
                        changed = True
                    if attribute.get("unit") != next_unit:
                        attribute["unit"] = next_unit
                        changed = True
                elif not bool(rule.get("required")):
                    attributes.remove(attribute)
                    existing.pop(name, None)
                    changed = True
            elif "unit" in attribute:
                attribute.pop("unit", None)
                changed = True
            next_values = normalize_rakuten_attribute_values_for_rule(attribute.get("values"), rule)
            if next_values:
                if attribute.get("values") != next_values:
                    attribute["values"] = next_values
                    changed = True
            elif not bool(rule.get("required")):
                attributes.remove(attribute)
                existing.pop(name, None)
                changed = True
    return patched if changed else payload


def patch_payload_for_missing_mandatory_attributes(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    missing_attributes = extract_missing_mandatory_attribute_names(error_text)
    rule_map = rakuten_attribute_rule_map_for_payload(payload)
    supported_missing_attributes = [
        attribute for attribute in missing_attributes
        if attribute == RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE
        or attribute in RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES
        or attribute in rule_map
    ]
    if not supported_missing_attributes:
        return payload
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    patched_variants = patched.get("variants")
    if not isinstance(patched_variants, dict):
        return payload
    changed = False
    for variant in patched_variants.values():
        if not isinstance(variant, dict):
            continue
        attributes = variant.get("attributes")
        if not isinstance(attributes, list):
            attributes = []
            variant["attributes"] = attributes
        existing_names = {normalize_text(attribute.get("name")) for attribute in attributes if isinstance(attribute, dict)}
        for attribute_name in supported_missing_attributes:
            if attribute_name in existing_names:
                continue
            attribute = infer_missing_mandatory_attribute(attribute_name, variant, patched, rule_map.get(attribute_name))
            if not attribute:
                continue
            attributes.append(attribute)
            existing_names.add(attribute_name)
            changed = True
    return patched if changed else payload


def patch_payload_for_unknown_attribute_name_errors(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    paths = extract_unknown_attribute_name_error_paths(error_text)
    if not paths:
        return payload
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    patched_variants = patched.get("variants")
    if not isinstance(patched_variants, dict):
        return payload
    changed = False
    by_variant: dict[str, set[int]] = {}
    for variant_id, attribute_index in paths:
        by_variant.setdefault(variant_id, set()).add(attribute_index)
    for variant_id, indexes in by_variant.items():
        variant = patched_variants.get(variant_id)
        if not isinstance(variant, dict):
            continue
        attributes = variant.get("attributes")
        if not isinstance(attributes, list):
            continue
        for attribute_index in sorted(indexes, reverse=True):
            if 0 <= attribute_index < len(attributes):
                attributes.pop(attribute_index)
                changed = True
        if not attributes:
            variant.pop("attributes", None)
    return patched if changed else payload


def patch_payload_for_attribute_string_value_errors(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    errors = extract_attribute_string_value_errors(error_text)
    if not errors:
        return payload
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    patched_variants = patched.get("variants")
    if not isinstance(patched_variants, dict):
        return payload
    rule_map = rakuten_attribute_rule_map_for_payload(patched)
    changed = False
    for error in errors:
        variant_id = normalize_text(error.get("variantId"))
        attribute_index = error.get("attributeIndex")
        attribute_name = normalize_text(error.get("attributeName"))
        max_length = int(error.get("maxLength") or 0)
        target_variants = [patched_variants.get(variant_id)] if variant_id else list(patched_variants.values())
        for variant in target_variants:
            if not isinstance(variant, dict):
                continue
            attributes = variant.get("attributes")
            if not isinstance(attributes, list):
                continue
            target_attributes: list[dict[str, Any]] = []
            if isinstance(attribute_index, int) and 0 <= attribute_index < len(attributes):
                attribute = attributes[attribute_index]
                if isinstance(attribute, dict):
                    target_attributes.append(attribute)
            elif attribute_name:
                target_attributes.extend(
                    attribute for attribute in attributes
                    if isinstance(attribute, dict) and normalize_text(attribute.get("name")) == attribute_name
                )
            for attribute in target_attributes:
                name = normalize_text(attribute.get("name"))
                rule = dict(rule_map.get(name) or {})
                if max_length > 0:
                    rule["maxLength"] = str(max_length)
                next_values = normalize_rakuten_attribute_values_for_rule(attribute.get("values"), rule)
                if next_values:
                    if attribute.get("values") != next_values:
                        attribute["values"] = next_values
                        changed = True
                elif not bool(rule.get("required")) and attribute in attributes:
                    attributes.remove(attribute)
                    changed = True
    return patched if changed else payload


def normalize_rakuten_attribute_values_for_rule(values: Any, rule: dict[str, Any]) -> list[str]:
    value_items: list[str] = []
    raw_values = values if isinstance(values, list) else [values]
    multiple = bool(rule.get("multiple"))
    delimiter = normalize_text(rule.get("delimiter"))
    should_split = multiple and ("|" in delimiter or "バーティカルバー" in delimiter)
    for value in raw_values:
        text = normalize_rakuten_attribute_value(value, allow_placeholder=True)
        if not text:
            continue
        parts = re.split(r"\s*[|｜]\s*", text) if should_split else [text]
        for part in parts:
            normalized = normalize_rakuten_attribute_value(part, allow_placeholder=True)
            if normalized:
                value_items.append(normalized)
    max_length = rakuten_attribute_rule_int(rule, "maxLength")
    if max_length > 0:
        value_items = [value[:max_length] for value in value_items if value[:max_length]]
    max_values = rakuten_attribute_rule_int(rule, "maxValues")
    unique_values = unique_texts(value_items)
    if normalize_text(rule.get("name")) == RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE:
        for value in unique_values:
            color = normalize_rakuten_representative_color(value)
            if color:
                return [color]
        return [RAKUTEN_REPRESENTATIVE_COLOR_FALLBACK]
    if normalize_text(rule.get("inputMethod")) == "選択式":
        unique_values = unique_texts([
            normalize_rakuten_selective_attribute_value(rule.get("name"), value) or value
            for value in unique_values
        ])
        recommended_values = rule.get("recommendedValues")
        if isinstance(recommended_values, list) and recommended_values:
            recommended_set = {normalize_text(value) for value in recommended_values if normalize_text(value)}
            unique_values = [value for value in unique_values if value in recommended_set]
    if max_values > 0:
        unique_values = unique_values[:max_values]
    return unique_values


def normalize_rakuten_selective_attribute_value(attribute_name: Any, value: Any) -> str:
    normalized_name = normalize_text(attribute_name)
    text = unicodedata.normalize("NFKC", normalize_text(value))
    if not normalized_name or not text:
        return ""
    if normalized_name == "チェストの高さ":
        if text in {RAKUTEN_CHEST_HEIGHT_LOW_VALUE, RAKUTEN_CHEST_HEIGHT_HIGH_VALUE}:
            return text
        match = re.search(r"\d+(?:\.\d+)?", text)
        if match:
            try:
                return (
                    RAKUTEN_CHEST_HEIGHT_HIGH_VALUE
                    if Decimal(match.group(0)) >= Decimal("100")
                    else RAKUTEN_CHEST_HEIGHT_LOW_VALUE
                )
            except Exception:
                return ""
    return ""


def rakuten_attribute_has_effective_values(attribute: dict[str, Any], rule: dict[str, Any]) -> bool:
    if not isinstance(attribute, dict):
        return False
    values = normalize_rakuten_attribute_values_for_rule(attribute.get("values"), rule)
    if not values:
        return False
    if bool(rule.get("unitRequired")):
        value = first_text_value(values)
        default_unit = normalize_rakuten_attribute_unit(rule.get("unit"))
        normalized_values, normalized_unit = normalize_rakuten_attribute_values_and_unit(
            normalize_text(attribute.get("name") or rule.get("name")),
            value,
            normalize_text(attribute.get("unit")),
            default_unit=default_unit,
        )
        return bool(normalized_values and normalized_unit)
    recommended_values = rule.get("recommendedValues")
    if (
        normalize_text(rule.get("inputMethod")) == "選択式"
        and isinstance(recommended_values, list)
        and recommended_values
    ):
        recommended_set = {normalize_text(value) for value in recommended_values if normalize_text(value)}
        return all(value in recommended_set for value in values)
    return True


def rakuten_attribute_rule_int(rule: dict[str, Any], key: str) -> int:
    text = normalize_text(rule.get(key))
    if not text:
        return 0
    match = re.search(r"\d+", text)
    if not match:
        return 0
    try:
        return int(match.group(0))
    except ValueError:
        return 0


def infer_missing_mandatory_attribute(
    attribute_name: str,
    variant: dict[str, Any],
    payload: dict[str, Any],
    rule: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    normalized_name = normalize_text(attribute_name)
    if not normalized_name:
        return None
    if normalized_name == RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE:
        value = infer_rakuten_representative_color(variant, payload) or RAKUTEN_REPRESENTATIVE_COLOR_FALLBACK
    elif normalized_name == "ブランド名":
        value = infer_rakuten_brand_name(variant, payload) or RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES[normalized_name]
    elif normalized_name == "シリーズ名":
        value = infer_rakuten_series_name(variant, payload) or RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES[normalized_name]
    elif normalized_name == "メーカー型番":
        value = infer_rakuten_model_number(variant, payload) or RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES[normalized_name]
    elif normalized_name == "原産国／製造国":
        value = infer_rakuten_origin_country(variant, payload) or RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES[normalized_name]
    elif normalized_name == "総本数":
        value = infer_rakuten_total_count(variant, payload) or RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES[normalized_name]
    elif normalized_name == "総個数":
        value = infer_rakuten_total_count(variant, payload) or rakuten_attribute_fallback_value(rule or {})
    elif normalized_name == "総重量":
        value = infer_rakuten_total_weight(variant, payload) or rakuten_attribute_fallback_value(rule or {})
    elif normalized_name == "総容量":
        value = infer_rakuten_total_capacity(variant, payload) or rakuten_attribute_fallback_value(rule or {})
    elif normalized_name == "単品容量":
        value = infer_rakuten_single_capacity(variant, payload) or RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES[normalized_name]
    else:
        value = RAKUTEN_MANDATORY_ATTRIBUTE_FALLBACK_VALUES.get(normalized_name, "")
        if not value:
            value = rakuten_attribute_fallback_value(rule or {})
    if not value:
        return None
    default_unit = normalize_rakuten_attribute_unit((rule or {}).get("unit")) if bool((rule or {}).get("unitRequired")) else ""
    values, unit = normalize_rakuten_attribute_values_and_unit(normalized_name, value, "", default_unit=default_unit)
    if not values:
        values = [normalize_text(value)]
    attribute: dict[str, Any] = {"name": normalized_name, "values": values}
    if unit:
        attribute["unit"] = unit
    return attribute


def rakuten_attribute_fallback_value(rule: dict[str, Any]) -> str:
    example = normalize_text(rule.get("example"))
    if example:
        return example
    value_format = normalize_text(rule.get("format")).lower()
    if value_format == "number":
        return "1"
    return "-"


def rakuten_attribute_context_text(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    values = [
        first_text_from_keys(payload, ("title", "tagline", "itemNumber", "genreId")),
        first_text_from_keys(variant, ("variantId", "merchantDefinedSkuId")),
        first_text_from_keys(variant.get("articleNumber", {}) if isinstance(variant.get("articleNumber"), dict) else {}, ("value",)),
        first_text_value(variant.get("selectorValues")),
        first_text_value(variant.get("attributes")),
    ]
    return " ".join(value for value in unique_texts(values) if value)


def infer_rakuten_brand_name(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    context = rakuten_attribute_context_text(variant, payload)
    normalized_context = unicodedata.normalize("NFKC", context)
    for pattern, brand in RAKUTEN_BRAND_INFERENCE_PATTERNS:
        if re.search(pattern, normalized_context, flags=re.I):
            return brand
    return ""


def infer_rakuten_series_name(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    context = rakuten_attribute_context_text(variant, payload)
    normalized_context = unicodedata.normalize("NFKC", context)
    for pattern, series in RAKUTEN_SERIES_INFERENCE_PATTERNS:
        if re.search(pattern, normalized_context, flags=re.I):
            return series
    return ""


def infer_rakuten_model_number(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    candidates = [
        first_text_from_keys(variant, ("variantId", "merchantDefinedSkuId")),
        first_text_from_keys(variant.get("articleNumber", {}) if isinstance(variant.get("articleNumber"), dict) else {}, ("value",)),
        first_text_from_keys(payload, ("title", "itemNumber")),
    ]
    context = unicodedata.normalize("NFKC", " ".join(candidate for candidate in candidates if candidate)).upper()
    for match in re.finditer(r"\b[A-Z]{1,8}[-_ ]?[A-Z]*\d[A-Z0-9_-]*(?:[-_][A-Z0-9]+)*\b", context):
        model = re.sub(r"\s+", "-", match.group(0)).strip("-_")
        if not model or model.startswith(LISTING_MANAGE_NUMBER_PREFIX.upper()):
            continue
        if model in {"G-SHOCK", "BABY-G"}:
            continue
        return model[:64]
    return ""


def infer_rakuten_origin_country(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    context = unicodedata.normalize("NFKC", rakuten_attribute_context_text(variant, payload))
    for pattern, country in RAKUTEN_ORIGIN_INFERENCE_PATTERNS:
        if re.search(pattern, context, flags=re.I):
            return country
    return ""


def infer_rakuten_total_count(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    context = unicodedata.normalize("NFKC", rakuten_attribute_context_text(variant, payload))
    match = re.search(r"([1-9][0-9]{0,2})\s*(?:本|個|枚|袋|箱|セット)", context)
    return match.group(1) if match else ""


def infer_rakuten_total_weight(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    context = unicodedata.normalize("NFKC", rakuten_attribute_context_text(variant, payload))
    match = re.search(r"([1-9][0-9]{0,5}(?:\.[0-9]+)?)\s*(kg|KG|Kg|g|G|グラム|キログラム)", context)
    if not match:
        return ""
    return f"{match.group(1)}{match.group(2)}"


def infer_rakuten_total_capacity(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    return infer_rakuten_single_capacity(variant, payload)


def infer_rakuten_single_capacity(variant: dict[str, Any], payload: dict[str, Any]) -> str:
    context = unicodedata.normalize("NFKC", rakuten_attribute_context_text(variant, payload))
    match = re.search(r"([1-9][0-9]{0,4}(?:\.[0-9]+)?)\s*(ml|mL|ML|ミリリットル|L|リットル)", context)
    if not match:
        return ""
    return f"{match.group(1)}{match.group(2)}"


def patch_payload_for_attribute_unit_errors(payload: dict[str, Any], error_text: str) -> dict[str, Any]:
    rules = extract_attribute_unit_error_rules(error_text)
    if not rules:
        return payload
    variants = payload.get("variants")
    if not isinstance(variants, dict):
        return payload
    patched = json.loads(json.dumps(payload, ensure_ascii=False))
    patched_variants = patched.get("variants")
    if not isinstance(patched_variants, dict):
        return payload
    changed = False
    for variant in patched_variants.values():
        if not isinstance(variant, dict):
            continue
        attributes = variant.get("attributes")
        if not isinstance(attributes, list):
            continue
        for attribute in attributes:
            if not isinstance(attribute, dict):
                continue
            name = normalize_text(attribute.get("name"))
            action = rules.get(name)
            if not action:
                continue
            values = attribute.get("values")
            first_value = first_text_value(values)
            if action == "remove_unit":
                if "unit" in attribute:
                    attribute.pop("unit", None)
                    changed = True
                continue
            if action == "require_unit":
                next_values, next_unit = normalize_rakuten_attribute_values_and_unit(
                    name,
                    first_value,
                    normalize_text(attribute.get("unit")),
                    default_unit=rakuten_attribute_default_unit(payload, name),
                )
                if next_values and next_unit:
                    if attribute.get("values") != next_values:
                        attribute["values"] = next_values
                        changed = True
                    if attribute.get("unit") != next_unit:
                        attribute["unit"] = next_unit
                        changed = True
    return patched if changed else payload


def extract_attribute_unit_error_rules(error_text: str) -> dict[str, str]:
    rules: dict[str, str] = {}

    def collect_from_detail(detail: Any) -> None:
        if not isinstance(detail, dict):
            return
        code = normalize_text(detail.get("code"))
        properties = detail.get("properties") if isinstance(detail.get("properties"), dict) else {}
        attribute_name = normalize_text(properties.get("attributeName"))
        if not attribute_name:
            return
        if code == "invalidNoInputUnit":
            rules[attribute_name] = "remove_unit"
        elif code == "invalidNoUnitAndValues":
            rules[attribute_name] = "require_unit"

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            collect_from_detail(value)
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    text = normalize_text(error_text)
    try:
        json_start = text.find("{")
        if json_start >= 0:
            walk(json.loads(text[json_start:]))
    except Exception:
        pass
    for match in re.finditer(r'"code"\s*:\s*"(invalidNoInputUnit|invalidNoUnitAndValues)".*?"attributeName"\s*:\s*"([^"]+)"', text):
        code, attribute_name = match.groups()
        rules[normalize_text(attribute_name)] = "remove_unit" if code == "invalidNoInputUnit" else "require_unit"
    return rules


def extract_unknown_attribute_name_error_paths(error_text: str) -> list[tuple[str, int]]:
    paths: list[tuple[str, int]] = []

    def add_path(property_path: Any) -> None:
        match = re.fullmatch(r"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]\.name", normalize_text(property_path))
        if not match:
            return
        variant_id, attribute_index = match.groups()
        path = (normalize_text(variant_id), int(attribute_index))
        if path[0] and path not in paths:
            paths.append(path)

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            if normalize_text(value.get("code")) == "IE1002":
                metadata = value.get("metadata") if isinstance(value.get("metadata"), dict) else {}
                add_path(metadata.get("propertyPath"))
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    text = normalize_text(error_text)
    try:
        json_start = text.find("{")
        if json_start >= 0:
            walk(json.loads(text[json_start:]))
    except Exception:
        pass
    for match in re.finditer(r"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]\.name", text):
        variant_id, attribute_index = match.groups()
        path = (normalize_text(variant_id), int(attribute_index))
        if path[0] and path not in paths:
            paths.append(path)
    return paths


def extract_attribute_string_value_errors(error_text: str) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []

    def add_error(property_path: Any, attribute_name: Any, max_length: Any) -> None:
        normalized_path = normalize_text(property_path)
        match = re.fullmatch(r"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]", normalized_path)
        if not match:
            return
        variant_id, attribute_index = match.groups()
        max_length_value = rakuten_attribute_rule_int({"maxLength": max_length}, "maxLength")
        error = {
            "variantId": normalize_text(variant_id),
            "attributeIndex": int(attribute_index),
            "attributeName": normalize_text(attribute_name),
            "maxLength": max_length_value,
        }
        if error["variantId"] and error not in errors:
            errors.append(error)

    def collect_from_detail(detail: Any, property_path: Any) -> None:
        if not isinstance(detail, dict):
            return
        if normalize_text(detail.get("code")) != "invalidStringValue":
            return
        properties = detail.get("properties") if isinstance(detail.get("properties"), dict) else {}
        add_error(property_path, properties.get("attributeName"), properties.get("maxLength"))

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            metadata = value.get("metadata") if isinstance(value.get("metadata"), dict) else {}
            property_path = metadata.get("propertyPath")
            details = metadata.get("details")
            if isinstance(details, list):
                for detail in details:
                    collect_from_detail(detail, property_path)
            else:
                collect_from_detail(value, property_path)
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    text = normalize_text(error_text)
    try:
        json_start = text.find("{")
        if json_start >= 0:
            walk(json.loads(text[json_start:]))
    except Exception:
        pass
    pattern = (
        r'"code"\s*:\s*"invalidStringValue".*?'
        r'"attributeName"\s*:\s*"([^"]+)".*?'
        r'"maxLength"\s*:\s*"?(\d+)"?.*?'
        r'"propertyPath"\s*:\s*"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]"'
    )
    for match in re.finditer(pattern, text):
        attribute_name, max_length, variant_id, attribute_index = match.groups()
        error = {
            "variantId": normalize_text(variant_id),
            "attributeIndex": int(attribute_index),
            "attributeName": normalize_text(attribute_name),
            "maxLength": int(max_length),
        }
        if error["variantId"] and error not in errors:
            errors.append(error)
    path_pattern = r"variants\.([^.\\\[\]]+)\.attributes\[(\d+)\]"
    name_match = re.search(r'"attributeName"\s*:\s*"([^"]+)"', text)
    max_match = re.search(r'"maxLength"\s*:\s*"?(\d+)"?', text)
    if name_match and max_match:
        for match in re.finditer(path_pattern, text):
            variant_id, attribute_index = match.groups()
            error = {
                "variantId": normalize_text(variant_id),
                "attributeIndex": int(attribute_index),
                "attributeName": normalize_text(name_match.group(1)),
                "maxLength": int(max_match.group(1)),
            }
            if error["variantId"] and error not in errors:
                errors.append(error)
    return errors


def extract_missing_mandatory_attribute_names(error_text: str) -> list[str]:
    text = normalize_text(error_text)
    if not text:
        return []
    names: list[str] = []
    for match in re.finditer(r"attributeNames:\s*([^\"。；;]+)", text):
        raw_names = match.group(1)
        for name in re.split(r"[,、，]\s*", raw_names):
            normalized = normalize_text(name).strip("[]'\" ")
            if normalized and normalized not in names:
                names.append(normalized)
    for match in re.finditer(r'"attributeNames"\s*:\s*\[([^\]]+)\]', text):
        raw_names = match.group(1)
        for name in re.findall(r'"([^"]+)"', raw_names):
            normalized = normalize_text(name)
            if normalized and normalized not in names:
                names.append(normalized)
    try:
        json_start = text.find("{")
        if json_start >= 0:
            payload = json.loads(text[json_start:])
            names.extend(extract_missing_mandatory_attribute_names_from_json(payload))
    except Exception:
        pass
    return unique_texts(names)


def extract_missing_mandatory_attribute_names_from_json(value: Any) -> list[str]:
    names: list[str] = []
    if isinstance(value, dict):
        metadata = value.get("metadata")
        if isinstance(metadata, dict):
            details = metadata.get("details")
            if isinstance(details, list):
                for detail in details:
                    if not isinstance(detail, dict):
                        continue
                    if normalize_text(detail.get("code")) != "invalidAllMandatoryAttributes":
                        continue
                    properties = detail.get("properties") if isinstance(detail.get("properties"), dict) else {}
                    attribute_names = properties.get("attributeNames")
                    if isinstance(attribute_names, list):
                        names.extend(normalize_text(name) for name in attribute_names if normalize_text(name))
        for child in value.values():
            names.extend(extract_missing_mandatory_attribute_names_from_json(child))
    elif isinstance(value, list):
        for item in value:
            names.extend(extract_missing_mandatory_attribute_names_from_json(item))
    return names


def create_store_product_on_rakuten(
    service_secret: str,
    license_key: str,
    store: StoreModel,
    product: ProductModel,
    cabinet_context: dict[str, Any] | None = None,
    cancel_check: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    raw_payload = product_raw_payload(product)
    prepared_image_map = listing_prepared_image_map(product, raw_payload)
    product._listing_store_id = store.id
    manage_number = generate_listing_manage_number(product, raw_payload)
    uploaded_product_images: list[dict[str, str]] = []
    uploaded_description_images: list[dict[str, str]] = []
    product_images = [
        image
        for image in recover_missing_local_product_images(product, product_images_for_edit(product))
        if not is_gif_image_url(image)
    ][:RAKUTEN_LISTING_IMAGE_LIMIT]
    item_write_started = False
    payload: dict[str, Any] = {}
    try:
        if cancel_check and cancel_check():
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        uploaded_product_images = upload_product_images_to_rakuten(
            service_secret,
            license_key,
            store,
            product,
            manage_number,
            cabinet_context=cabinet_context,
            source_images=product_images,
            prepared_image_map=prepared_image_map,
            cancel_check=cancel_check,
            require_complete=True,
        )
        if cancel_check and cancel_check():
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        description_result = upload_product_description_images_to_rakuten(
            service_secret,
            license_key,
            store,
            product,
            manage_number,
            raw_payload,
            cabinet_context=cabinet_context,
            prepared_image_map=prepared_image_map,
            cancel_check=cancel_check,
            require_complete=True,
        )
        raw_payload = description_result["rawPayload"]
        uploaded_description_images = description_result["uploadedImages"]
        if cancel_check and cancel_check():
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        payload = build_rakuten_item_upsert_payload(
            product,
            raw_payload,
            uploaded_product_images,
            manage_number=manage_number,
            hide_item=True,
        )
        if cancel_check and cancel_check():
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        payload = put_rakuten_item_with_attribute_retry(service_secret, license_key, manage_number, payload)
        item_write_started = True
        if cancel_check and cancel_check():
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        inventory_payloads = build_rakuten_inventory_upsert_payloads(
            manage_number,
            payload.get("variants") if isinstance(payload.get("variants"), dict) else {},
        )
        bulk_upsert_rakuten_inventories(service_secret, license_key, inventory_payloads)
        if cancel_check and cancel_check():
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        patch_rakuten_item_visibility(
            service_secret,
            license_key,
            manage_number,
            hide_item=False,
        )
        payload["hideItem"] = False
    except TaskCancelled:
        if item_write_started:
            try:
                delete_rakuten_item(service_secret, license_key, manage_number)
            except Exception:
                pass
        rollback_uploaded_listing_images(
            service_secret,
            license_key,
            [*uploaded_product_images, *uploaded_description_images],
        )
        raise
    except Exception as exc:
        if item_write_started:
            try:
                delete_rakuten_item(service_secret, license_key, manage_number)
            except Exception:
                pass
        rollback_message = rollback_uploaded_listing_images(
            service_secret,
            license_key,
            [*uploaded_product_images, *uploaded_description_images],
        )
        if rollback_message:
            raise RuntimeError(f"{exc}；已回滚本次上传图片：{rollback_message}") from exc
        raise
    now = datetime.now()
    updated_payload = dict(raw_payload)
    updated_payload.update(payload)
    updated_payload["manageNumber"] = manage_number
    updated_payload["itemNumber"] = payload.get("itemNumber") or manage_number
    updated_payload["images"] = uploaded_product_images
    updated_payload["descriptionImages"] = uploaded_description_images
    updated_payload["ltEditedImages"] = [
        build_rakuten_cabinet_image_url(store.store_code, image["location"])
        for image in uploaded_product_images
        if image.get("location")
    ]
    updated_payload["listingStore"] = {
        "storeId": store.id,
        "storeCode": store.store_code,
        "storeName": store.store_name,
        "aliasName": store.alias_name,
    }
    updated_payload["listingImageCompletion"] = {
        "status": "success",
        "completedAt": now.isoformat(timespec="seconds"),
        "totalMainImageCount": len(uploaded_product_images),
        "remainingMainImageCount": 0,
        "descriptionImageCount": len(uploaded_description_images),
    }
    updated_payload["created"] = updated_payload.get("created") or now.isoformat(timespec="seconds")
    updated_payload["updated"] = now.isoformat(timespec="seconds")
    return {
        "manageNumber": manage_number,
        "itemNumber": payload.get("itemNumber") or manage_number,
        "payload": updated_payload,
        "price": price_from_rakuten_item(updated_payload),
        "imageUrl": build_rakuten_cabinet_image_url(store.store_code, uploaded_product_images[0]["location"]) if uploaded_product_images else product.image_url,
    }


def listing_creation_active_for_store(store_id: int) -> bool:
    with session_scope() as session:
        active_rows = session.scalars(
            select(ListingTaskModel).where(
                ListingTaskModel.status.in_(("queued", "running")),
            )
        ).all()
        return any(int(store_id) in listing_task_store_ids(row) for row in active_rows)


def wait_for_listing_creation_priority(
    store_id: int,
    *,
    task_id: str | None = None,
) -> None:
    while listing_creation_active_for_store(store_id):
        if task_id and is_task_cancel_requested(SyncTaskModel, task_id):
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        time.sleep(TASK_START_RETRY_DELAY_SECONDS)


def enrich_listed_product_images(
    owner_username: str,
    store_id: int,
    listed_product_id: int,
    service_secret: str,
    license_key: str,
    *,
    task_id: str | None = None,
) -> None:
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        listed_product = session.get(ProductModel, listed_product_id)
        if (
            store is None
            or listed_product is None
            or listed_product.owner_username != owner_username
            or int(listed_product.store_id or 0) != int(store_id)
        ):
            raise RuntimeError("待完善的店铺商品不存在。")
        source_product = session.get(ProductModel, listed_product.parent_product_id) if listed_product.parent_product_id else None
        if source_product is None or source_product.owner_username != owner_username:
            raise RuntimeError("待完善商品缺少对应的源商品。")
        manage_number = normalize_text(listed_product.rakuten_manage_number)
        if not manage_number:
            raise RuntimeError("待完善商品缺少乐天商品管理编号。")
        source_raw = product_raw_payload(source_product)
        listed_raw = product_raw_payload(listed_product)
        existing_main_images = [
            image
            for image in listed_raw.get("images", [])
            if isinstance(image, dict) and normalize_text(image.get("location"))
        ]
        existing_source_urls = {
            normalize_text(image.get("sourceUrl"))
            for image in existing_main_images
            if normalize_text(image.get("sourceUrl"))
        }
        all_main_sources = [
            image
            for image in recover_missing_local_product_images(
                source_product,
                product_images_for_edit(source_product),
            )
            if not is_gif_image_url(image)
        ][:RAKUTEN_LISTING_IMAGE_LIMIT]
        remaining_main_sources = [
            image
            for image in all_main_sources
            if normalize_text(image) not in existing_source_urls
        ]

        new_main_images: list[dict[str, str]] = []
        new_description_images: list[dict[str, str]] = []
        try:
            if remaining_main_sources:
                wait_for_listing_creation_priority(store_id, task_id=task_id)
                new_main_images = upload_product_images_to_rakuten(
                    service_secret,
                    license_key,
                    store,
                    source_product,
                    manage_number,
                    cabinet_context={},
                    source_images=remaining_main_sources,
                    cancel_check=lambda: is_task_cancel_requested(SyncTaskModel, task_id),
                    start_index=len(existing_main_images) + 1,
                    before_upload=lambda: wait_for_listing_creation_priority(
                        store_id,
                        task_id=task_id,
                    ),
                )
            wait_for_listing_creation_priority(store_id, task_id=task_id)
            description_result = upload_product_description_images_to_rakuten(
                service_secret,
                license_key,
                store,
                source_product,
                manage_number,
                source_raw,
                cabinet_context={},
                cancel_check=lambda: is_task_cancel_requested(SyncTaskModel, task_id),
                before_upload=lambda: wait_for_listing_creation_priority(
                    store_id,
                    task_id=task_id,
                ),
            )
            enriched_raw = description_result["rawPayload"]
            new_description_images = description_result["uploadedImages"]
            combined_main_images = [
                *existing_main_images,
                *new_main_images,
            ][:RAKUTEN_LISTING_IMAGE_LIMIT]
            payload = build_rakuten_item_upsert_payload(
                source_product,
                enriched_raw,
                combined_main_images,
                manage_number=manage_number,
                hide_item=False,
            )
            wait_for_listing_creation_priority(store_id, task_id=task_id)
            payload = put_rakuten_item_with_attribute_retry(
                service_secret,
                license_key,
                manage_number,
                payload,
            )
        except TaskCancelled:
            rollback_uploaded_listing_images(
                service_secret,
                license_key,
                [*new_main_images, *new_description_images],
            )
            raise
        except Exception as exc:
            rollback_message = rollback_uploaded_listing_images(
                service_secret,
                license_key,
                [*new_main_images, *new_description_images],
            )
            if rollback_message:
                raise RuntimeError(f"{exc}；已回滚本次图片完善上传：{rollback_message}") from exc
            raise

        updated_payload = dict(enriched_raw)
        updated_payload.update(payload)
        updated_payload["manageNumber"] = manage_number
        updated_payload["itemNumber"] = payload.get("itemNumber") or listed_product.item_number or manage_number
        updated_payload["images"] = combined_main_images
        updated_payload["descriptionImages"] = new_description_images
        updated_payload["ltEditedImages"] = [
            build_rakuten_cabinet_image_url(store.store_code, image["location"])
            for image in combined_main_images
            if image.get("location")
        ]
        updated_payload["listingStore"] = listed_raw.get("listingStore") or {
            "storeId": store.id,
            "storeCode": store.store_code,
            "storeName": store.store_name,
            "aliasName": store.alias_name,
        }
        created_at = listed_raw.get("created")
        if not created_at and listed_product.listed_at is not None:
            created_at = listed_product.listed_at.isoformat(timespec="seconds")
        updated_payload["created"] = created_at or datetime.now().isoformat(timespec="seconds")
        updated_payload["updated"] = datetime.now().isoformat(timespec="seconds")
        updated_payload["listingImageCompletion"] = {
            "status": "success",
            "completedAt": datetime.now().isoformat(timespec="seconds"),
            "fastImageCount": len(existing_main_images),
            "totalMainImageCount": len(combined_main_images),
            "remainingMainImageCount": 0,
            "descriptionImageCount": len(new_description_images),
            "taskId": task_id,
        }
        listed_product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        if updated_payload["ltEditedImages"]:
            listed_product.image_url = updated_payload["ltEditedImages"][0]
        listed_product.store_last_seen_at = datetime.now()
        listed_product.last_error = None


def perform_listing_image_upload(
    owner_username: str,
    store_id: int,
    listed_product_ids: list[int],
    *,
    task_id: str | None = None,
) -> dict[str, Any]:
    normalized_ids = normalize_product_ids(listed_product_ids)
    success_ids: list[int] = []
    failed_ids: list[int] = []
    errors: list[str] = []
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None or store.owner_username != owner_username:
            raise RuntimeError("图片完善任务关联的店铺不存在。")
        service_secret = decrypt_text(store.rakuten_service_secret_encrypted)
        license_key = decrypt_text(store.rakuten_license_key_encrypted)
    if not service_secret or not license_key:
        raise RuntimeError("图片完善任务关联店铺缺少乐天 Secret 或乐天 Key。")

    for index, product_id in enumerate(normalized_ids, start=1):
        if task_id and is_task_cancel_requested(SyncTaskModel, task_id):
            break
        try:
            enrich_listed_product_images(
                owner_username,
                store_id,
                product_id,
                service_secret,
                license_key,
                task_id=task_id,
            )
            success_ids.append(product_id)
        except TaskCancelled:
            raise
        except Exception as exc:
            failed_ids.append(product_id)
            error_text = str(exc)
            errors.append(f"{product_id}: {error_text}")
            with session_scope() as session:
                product = session.get(ProductModel, product_id)
                if product is not None and product.owner_username == owner_username:
                    raw_payload = product_raw_payload(product)
                    completion = (
                        raw_payload.get("listingImageCompletion")
                        if isinstance(raw_payload.get("listingImageCompletion"), dict)
                        else {}
                    )
                    raw_payload["listingImageCompletion"] = {
                        **completion,
                        "status": "failed",
                        "taskId": task_id,
                        "lastError": error_text,
                    }
                    product.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
                    product.last_error = f"商品已基础上架，图片完善失败：{error_text}"
        if task_id:
            update_task_progress(
                SyncTaskModel,
                task_id,
                total_count=len(normalized_ids),
                success_count=len(success_ids),
                failed_count=len(failed_ids),
                message=f"图片完善中，已处理 {index} / {len(normalized_ids)} 个商品",
            )
    return {
        "totalCount": len(normalized_ids),
        "successCount": len(success_ids),
        "failedCount": len(failed_ids),
        "successIds": success_ids,
        "failedIds": failed_ids,
        "errors": errors,
        "cancelled": bool(task_id and is_task_cancel_requested(SyncTaskModel, task_id)),
    }


def create_listing_image_upload_task(
    owner_username: str,
    store_id: int,
    listing_task_id: str,
    listed_product_ids: list[int],
) -> str | None:
    normalized_ids = normalize_product_ids(listed_product_ids)
    if not normalized_ids:
        return None
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        if store is None or store.owner_username != owner_username:
            raise RuntimeError("不能为其他用户的店铺创建图片完善任务。")
        task_id = uuid.uuid4().hex
        session.add(
            SyncTaskModel(
                id=task_id,
                owner_username=owner_username,
                store_id=store_id,
                store_name=store.alias_name or store.store_name,
                task_name=f"上架图片完善 {listing_task_id[:8]}",
                task_type="listing_image_upload",
                payload_json=json.dumps(
                    {
                        "listingTaskId": listing_task_id,
                        "productIds": normalized_ids,
                    },
                    ensure_ascii=False,
                ),
                status="queued",
                total_count=len(normalized_ids),
                message="等待完善上架商品图片",
            )
        )
        for product_id in normalized_ids:
            product = session.get(ProductModel, product_id)
            if product is None or product.owner_username != owner_username:
                continue
            raw_payload = product_raw_payload(product)
            completion = (
                raw_payload.get("listingImageCompletion")
                if isinstance(raw_payload.get("listingImageCompletion"), dict)
                else {}
            )
            raw_payload["listingImageCompletion"] = {
                **completion,
                "status": "queued",
                "taskId": task_id,
            }
            product.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
    dispatch_sync_task(
        owner_username,
        task_id,
        task_type="listing_image_upload",
        delay_seconds=TASK_START_RETRY_DELAY_SECONDS,
    )
    return task_id


def upsert_listed_store_product_from_listing_result(
    session: Any,
    owner_username: str,
    source_product: ProductModel,
    store: StoreModel,
    listing_result: dict[str, Any],
) -> ProductModel:
    manage_number = normalize_text(listing_result.get("manageNumber"))
    if not manage_number:
        raise RuntimeError("上架结果缺少商品管理编号。")
    row = session.scalar(
        select(ProductModel).where(
            ProductModel.store_id == store.id,
            ProductModel.rakuten_manage_number == manage_number,
        )
    )
    payload = listing_result.get("payload") if isinstance(listing_result.get("payload"), dict) else {}
    source_url = build_public_item_page_url(store.store_code, listing_result.get("itemNumber") or manage_number)
    source_hash_url = f"{source_url}#store={store.id}&manage={quote(manage_number, safe='')}"
    if row is None:
        row = ProductModel(owner_username=owner_username, source_url=source_url, source_url_hash=make_source_url_hash(source_hash_url))
        session.add(row)
    row.owner_username = owner_username
    row.parent_product_id = source_product.id
    row.listing_task_id = None
    row.task_id = source_product.task_id
    row.store_id = store.id
    row.rakuten_manage_number = manage_number
    row.item_number = normalize_text(listing_result.get("itemNumber")) or manage_number
    row.title = source_product.title
    row.source_url = source_url
    row.source_url_hash = make_source_url_hash(source_hash_url)
    row.shop_name = store.store_name or source_product.shop_name
    row.image_url = normalize_text(listing_result.get("imageUrl")) or source_product.image_url
    row.price = Decimal(str(listing_result["price"])) if listing_result.get("price") is not None else source_product.price
    row.currency = source_product.currency or "JPY"
    row.genre_id = source_product.genre_id
    row.review_status = "listed"
    row.store_product_status = "active"
    row.rakuten_listing_status = "listed"
    row.raw_payload_json = json.dumps(payload, ensure_ascii=False)
    row.listed_at = datetime.now()
    row.store_last_seen_at = datetime.now()
    row.last_error = None
    return row


def record_product_listed_store(
    product: ProductModel,
    listed_product: ProductModel,
    store: StoreModel,
    listing_result: dict[str, Any],
) -> None:
    upsert_product_listed_store_record(product, {
        "storeId": store.id,
        "storeCode": store.store_code,
        "storeName": store.store_name,
        "aliasName": store.alias_name,
        "manageNumber": normalize_text(listing_result.get("manageNumber")) or listed_product.rakuten_manage_number,
        "itemNumber": normalize_text(listing_result.get("itemNumber")) or listed_product.item_number,
        "productId": listed_product.id,
        "listedAt": datetime.now().isoformat(sep=" ", timespec="seconds"),
    })


def ensure_product_listed_store_mark_from_store_product(session: Any, store_product: ProductModel, store: StoreModel) -> None:
    if not store_product.parent_product_id:
        return
    parent = session.get(ProductModel, store_product.parent_product_id)
    if parent is None or parent.owner_username != store_product.owner_username:
        return
    upsert_product_listed_store_record(parent, {
        "storeId": store.id,
        "storeCode": store.store_code,
        "storeName": store.store_name,
        "aliasName": store.alias_name,
        "manageNumber": normalize_text(store_product.rakuten_manage_number),
        "itemNumber": normalize_text(store_product.item_number),
        "productId": store_product.id,
        "listedAt": (store_product.listed_at or datetime.now()).isoformat(sep=" ", timespec="seconds"),
    })


def upsert_product_listed_store_record(product: ProductModel, record: dict[str, Any]) -> None:
    raw_payload = product_raw_payload(product)
    listed_stores = raw_payload.get("listedStores") if isinstance(raw_payload.get("listedStores"), list) else []
    try:
        next_store_id = int(record.get("storeId") or 0)
    except (TypeError, ValueError):
        next_store_id = 0
    if not next_store_id:
        return
    next_stores: list[dict[str, Any]] = []
    replaced = False
    for item in listed_stores:
        if not isinstance(item, dict):
            continue
        try:
            store_id = int(item.get("storeId") or 0)
        except (TypeError, ValueError):
            store_id = 0
        if store_id == next_store_id:
            next_stores.append(record)
            replaced = True
        else:
            next_stores.append(item)
    if not replaced:
        next_stores.append(record)
    raw_payload["listedStores"] = next_stores
    product.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
    product.review_status = "listed_master"
    product.listing_task_id = None
    product.last_error = None
    product.listed_at = product.listed_at or datetime.now()


def remove_product_listed_store_mark(product: ProductModel, store_id: int) -> bool:
    raw_payload = product_raw_payload(product)
    listed_stores = raw_payload.get("listedStores") if isinstance(raw_payload.get("listedStores"), list) else []
    next_stores: list[dict[str, Any]] = []
    removed = False
    for item in listed_stores:
        if not isinstance(item, dict):
            continue
        try:
            item_store_id = int(item.get("storeId") or 0)
        except (TypeError, ValueError):
            item_store_id = 0
        if item_store_id == int(store_id):
            removed = True
            continue
        next_stores.append(item)
    if not removed:
        return False
    raw_payload["listedStores"] = next_stores
    product.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
    if next_stores:
        product.review_status = "listed_master"
    elif product.review_status == "listed_master":
        product.review_status = "approved"
    product.listing_task_id = None
    return True


def remove_listed_store_mark_for_store_product(session: Any, store_product: ProductModel) -> None:
    if not store_product.store_id:
        return
    parent = session.get(ProductModel, store_product.parent_product_id) if store_product.parent_product_id else None
    if parent is not None and parent.owner_username == store_product.owner_username:
        remove_product_listed_store_mark(parent, int(store_product.store_id))


def generate_listing_manage_number(product: ProductModel, raw_payload: dict[str, Any]) -> str:
    existing = normalize_text(product.rakuten_manage_number) if product.review_status == "listed" else ""
    if existing:
        return existing[:32]
    store_id = int(getattr(product, "_listing_store_id", 0) or 0)
    return build_listing_manage_number(product.id, store_id=store_id)


def build_listing_manage_number(product_id: int, *, store_id: int = 0, listed_at: datetime | None = None) -> str:
    date_text = (listed_at or datetime.now()).strftime("%Y%m%d")
    store_text = str(max(0, int(store_id or 0)))[-3:] or "0"
    product_text = str(max(0, int(product_id or 0)))[-5:].zfill(5)
    return f"{LISTING_MANAGE_NUMBER_PREFIX}{store_text}{date_text}{product_text}"[:32]


def prepare_rakuten_listing_image(
    image_url: str,
    prepared_image_url: str = "",
) -> dict[str, Any] | None:
    try:
        if prepared_image_url:
            image_data = load_product_image_bytes(
                prepared_image_url,
                max_bytes=RAKUTEN_CABINET_MAX_IMAGE_BYTES,
                size_error_message="缓存图片大小不能超过 2MB。",
            )
        else:
            image_data = prepare_rakuten_cabinet_image(
                load_product_image_bytes(
                    image_url,
                    max_bytes=MAX_PRODUCT_IMAGE_DOWNLOAD_BYTES,
                    size_error_message="图片下载大小不能超过 20MB。",
                )
            )
    except ProductImageUnavailableError:
        return None
    except RuntimeError as exc:
        if should_skip_listing_image_error(exc) or is_listing_image_read_error(exc):
            if is_listing_image_read_error(exc):
                logger.warning(
                    "listing image skipped after read failure url=%s error=%s cause=%s",
                    image_url,
                    exc,
                    exc.__cause__,
                    exc_info=True,
                )
            return None
        raise
    return {
        **image_data,
        "sourceUrl": image_url,
    }


def prepare_rakuten_listing_images(
    image_urls: list[str],
    *,
    prepared_image_map: dict[str, str] | None = None,
    cancel_check: Callable[[], bool] | None = None,
) -> list[dict[str, Any]]:
    if not image_urls:
        return []
    if cancel_check and cancel_check():
        raise TaskCancelled(TASK_CANCELLED_MESSAGE)
    worker_count = min(
        len(image_urls),
        max(1, int(settings.listing_image_prepare_workers)),
    )
    cached_images = prepared_image_map if isinstance(prepared_image_map, dict) else {}

    def prepare_one(image_url: str) -> dict[str, Any] | None:
        prepared_url = normalize_text(cached_images.get(image_url))
        if prepared_url:
            return prepare_rakuten_listing_image(image_url, prepared_url)
        return prepare_rakuten_listing_image(image_url)

    if worker_count <= 1:
        prepared = [prepare_one(image_url) for image_url in image_urls]
    else:
        futures = [
            LISTING_IMAGE_PREPARE_EXECUTOR.submit(
                prepare_one,
                image_url,
            )
            for image_url in image_urls
        ]
        prepared = [future.result() for future in futures]
    if cancel_check and cancel_check():
        raise TaskCancelled(TASK_CANCELLED_MESSAGE)
    return [image for image in prepared if image is not None]


def store_prepared_rakuten_listing_image(
    product_id: int,
    image_data: dict[str, Any],
) -> str:
    content = image_data.get("content") or b""
    if not isinstance(content, bytes) or not content:
        raise RuntimeError("预处理图片内容为空。")
    suffix = normalize_product_image_suffix(image_data.get("suffix"))
    if suffix != ".jpg":
        raise RuntimeError("预处理图片格式不正确。")
    digest = hashlib.sha256(content).hexdigest()[:24]
    filename = f"rakuten-prepared-{digest}.jpg"
    image_url = local_product_image_url(product_id, filename)
    if is_missing_local_product_image_url(image_url):
        store_product_image_content(
            image_url,
            content,
            "image/jpeg",
            LOCAL_PRODUCT_IMAGE_DIR / str(int(product_id)) / filename,
        )
    return image_url


def prepare_collected_product_for_listing(owner_username: str | None, product_id: int) -> dict[str, Any]:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or (owner_username and product.owner_username != owner_username):
            return {}
        if product.review_status not in {"pending", "approved", "error", "listed_master"}:
            return {}
        raw_payload = product_raw_payload(product)
        source_fingerprint = listing_preparation_source_fingerprint(product, raw_payload)
        existing_cache = listing_preparation_cache(product, raw_payload)
        if existing_cache and listing_preparation_ready(product, raw_payload):
            return existing_cache
        main_image_urls = [
            image
            for image in product_images_for_edit(product)
            if not is_gif_image_url(image)
        ][:RAKUTEN_LISTING_IMAGE_LIMIT]
        description_image_urls_to_prepare = unique_texts(
            [
                image_url
                for description in product_descriptions(raw_payload)
                for image_url in description_image_urls(description.get("value"))
                if not is_gif_image_url(image_url)
            ]
        )
        image_urls = unique_texts([*main_image_urls, *description_image_urls_to_prepare])
        preflight = _listing_preflight_product_check_uncached(product, None)

    prepared_images = prepare_rakuten_listing_images(image_urls)
    prepared_by_source = {
        normalize_text(image.get("sourceUrl")): image
        for image in prepared_images
        if normalize_text(image.get("sourceUrl"))
    }
    cached_images: list[dict[str, str]] = []
    for image_url in image_urls:
        prepared = prepared_by_source.get(image_url)
        if prepared is None:
            continue
        cached_images.append(
            {
                "sourceUrl": image_url,
                "preparedUrl": store_prepared_rakuten_listing_image(product_id, prepared),
            }
        )
    missing_images = [image_url for image_url in image_urls if image_url not in prepared_by_source]
    cache_payload = {
        "version": LISTING_PREPARATION_CACHE_VERSION,
        "sourceFingerprint": source_fingerprint,
        "preparedAt": datetime.now().isoformat(timespec="seconds"),
        "preflight": preflight,
        "images": cached_images,
        "imageCount": len(cached_images),
        "missingImageCount": len(missing_images),
    }

    referenced_local_urls: list[str] = []
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or (owner_username and product.owner_username != owner_username):
            return {}
        raw_payload = product_raw_payload(product)
        if listing_preparation_source_fingerprint(product, raw_payload) != source_fingerprint:
            return {}
        raw_payload[LISTING_PREPARATION_CACHE_KEY] = cache_payload
        product.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
        referenced_local_urls = collect_local_product_image_urls(raw_payload)
        session.flush()
    remove_unused_local_product_images(product_id, referenced_local_urls)
    return cache_payload


def upload_product_images_to_rakuten(
    service_secret: str,
    license_key: str,
    store: StoreModel,
    product: ProductModel,
    manage_number: str,
    cabinet_context: dict[str, Any] | None = None,
    source_images: list[str] | None = None,
    prepared_image_map: dict[str, str] | None = None,
    cancel_check: Callable[[], bool] | None = None,
    start_index: int = 1,
    before_upload: Callable[[], None] | None = None,
    require_complete: bool = False,
) -> list[dict[str, str]]:
    images = [
        image for image in recover_missing_local_product_images(product, source_images or product_images_for_edit(product))
        if not is_gif_image_url(image)
    ][:RAKUTEN_LISTING_IMAGE_LIMIT]
    if not images:
        raise RuntimeError("商品缺少图片，不能上架到乐天。")
    uploaded_images: list[dict[str, str]] = []
    upload_cabinet_context = cabinet_context if isinstance(cabinet_context, dict) else {}
    image_alt = sanitize_rakuten_image_alt(product.title) or "商品画像"
    try:
        prepared_images = prepare_rakuten_listing_images(
            images,
            prepared_image_map=prepared_image_map,
            cancel_check=cancel_check,
        )
        if require_complete and len(prepared_images) != len(images):
            missing_count = len(images) - len(prepared_images)
            raise RuntimeError(
                f"商品主图未能完整读取（缺少 {missing_count} 张），已停止上架。"
            )
        for index, image_data in enumerate(prepared_images, start=max(1, int(start_index or 1))):
            if cancel_check and cancel_check():
                raise TaskCancelled(TASK_CANCELLED_MESSAGE)
            if before_upload:
                before_upload()
            image_url = normalize_text(image_data.get("sourceUrl"))
            suffix = image_data["suffix"]
            file_path = listing_cabinet_upload_file_path(manage_number, index, suffix, kind="p")
            file_name = listing_cabinet_upload_file_name(file_path)
            cabinet_folder = ensure_listing_cabinet_folder_for_upload(
                service_secret,
                license_key,
                store,
                1,
                cabinet_context=upload_cabinet_context,
            )
            folder_id = int(cabinet_folder.get("folderId") or 0)
            folder_path = normalize_text(
                cabinet_folder.get("directoryName")
                or cabinet_folder.get("folderPath")
                or cabinet_folder.get("folderName")
            )
            if not folder_id or not folder_path:
                raise RuntimeError("R-Cabinet 上架文件夹不可用。")
            result = insert_rakuten_cabinet_file(
                service_secret,
                license_key,
                file_name=file_name,
                file_path=file_path,
                content=image_data["content"],
                content_type=image_data["contentType"],
                folder_id=folder_id,
                overwrite=True,
            )
            location = cabinet_image_location(folder_path, result.get("filePath") or file_path)
            uploaded_images.append(
                {
                    "type": "CABINET",
                    "location": location,
                    "alt": image_alt,
                    "fileId": str(result.get("fileId") or ""),
                    "folderId": str(folder_id),
                    "folderPath": folder_path,
                    "sourceUrl": image_url,
                    "fileUrl": result.get("fileUrl") or build_rakuten_cabinet_image_url(store.store_code, location),
                }
            )
            reserve_listing_cabinet_folder_slots(upload_cabinet_context, cabinet_folder, 1)
            if cancel_check and cancel_check():
                raise TaskCancelled(TASK_CANCELLED_MESSAGE)
    except TaskCancelled:
        rollback_uploaded_listing_images(service_secret, license_key, uploaded_images)
        raise
    except Exception as exc:
        rollback_message = rollback_uploaded_listing_images(service_secret, license_key, uploaded_images)
        if rollback_message:
            raise RuntimeError(f"{exc}；已回滚本次已上传图片：{rollback_message}") from exc
        raise
    if not uploaded_images:
        raise RuntimeError("商品本地图片文件不存在或已失效，未能上传任何图片。")
    return uploaded_images


def upload_product_description_images_to_rakuten(
    service_secret: str,
    license_key: str,
    store: StoreModel,
    product: ProductModel,
    manage_number: str,
    raw_payload: dict[str, Any],
    cabinet_context: dict[str, Any] | None = None,
    prepared_image_map: dict[str, str] | None = None,
    cancel_check: Callable[[], bool] | None = None,
    before_upload: Callable[[], None] | None = None,
    require_complete: bool = False,
) -> dict[str, Any]:
    description_items = product_descriptions(raw_payload)
    all_description_image_urls = unique_texts([
        url
        for description in description_items
        for url in description_image_urls(description.get("value"))
    ])
    removed_image_urls: list[str] = [url for url in all_description_image_urls if is_gif_image_url(url)]
    image_urls = unique_texts(
        [
            url
            for url in all_description_image_urls
            if should_transfer_description_image(url, store.store_code) and not is_gif_image_url(url)
        ]
    )
    if not image_urls:
        if removed_image_urls:
            raw_payload = remove_product_description_image_urls(raw_payload, removed_image_urls)
        return {"rawPayload": raw_payload, "uploadedImages": []}

    uploaded_images: list[dict[str, str]] = []
    replacement_map: dict[str, str] = {}
    upload_cabinet_context = cabinet_context if isinstance(cabinet_context, dict) else {}
    image_alt = sanitize_rakuten_image_alt(product.title) or "商品画像"
    try:
        prepared_images = prepare_rakuten_listing_images(
            image_urls,
            prepared_image_map=prepared_image_map,
            cancel_check=cancel_check,
        )
        if require_complete and len(prepared_images) != len(image_urls):
            missing_count = len(image_urls) - len(prepared_images)
            raise RuntimeError(
                f"商品说明图未能完整读取（缺少 {missing_count} 张），已停止上架。"
            )
        prepared_source_urls = {
            normalize_text(image_data.get("sourceUrl"))
            for image_data in prepared_images
            if normalize_text(image_data.get("sourceUrl"))
        }
        removed_image_urls.extend(
            image_url
            for image_url in image_urls
            if image_url not in prepared_source_urls
        )
        for index, image_data in enumerate(prepared_images, start=1):
            if cancel_check and cancel_check():
                raise TaskCancelled(TASK_CANCELLED_MESSAGE)
            if before_upload:
                before_upload()
            image_url = normalize_text(image_data.get("sourceUrl"))
            suffix = image_data["suffix"]
            file_path = listing_cabinet_upload_file_path(manage_number, index, suffix, kind="d")
            file_name = listing_cabinet_upload_file_name(file_path)
            cabinet_folder = ensure_listing_cabinet_folder_for_upload(
                service_secret,
                license_key,
                store,
                1,
                cabinet_context=upload_cabinet_context,
            )
            folder_id = int(cabinet_folder.get("folderId") or 0)
            folder_path = normalize_text(
                cabinet_folder.get("directoryName")
                or cabinet_folder.get("folderPath")
                or cabinet_folder.get("folderName")
            )
            if not folder_id or not folder_path:
                raise RuntimeError("R-Cabinet 上架说明图文件夹不可用。")
            result = insert_rakuten_cabinet_file(
                service_secret,
                license_key,
                file_name=file_name,
                file_path=file_path,
                content=image_data["content"],
                content_type=image_data["contentType"],
                folder_id=folder_id,
                overwrite=True,
            )
            location = cabinet_image_location(folder_path, result.get("filePath") or file_path)
            file_url = result.get("fileUrl") or build_rakuten_cabinet_image_url(store.store_code, location)
            uploaded_images.append(
                {
                    "type": "CABINET_DESCRIPTION",
                    "location": location,
                    "alt": image_alt,
                    "fileId": str(result.get("fileId") or ""),
                    "folderId": str(folder_id),
                    "folderPath": folder_path,
                    "sourceUrl": image_url,
                    "fileUrl": file_url,
                }
            )
            replacement_map[image_url] = file_url
            reserve_listing_cabinet_folder_slots(upload_cabinet_context, cabinet_folder, 1)
            if cancel_check and cancel_check():
                raise TaskCancelled(TASK_CANCELLED_MESSAGE)
    except TaskCancelled:
        rollback_uploaded_listing_images(service_secret, license_key, uploaded_images)
        raise
    except Exception as exc:
        rollback_message = rollback_uploaded_listing_images(service_secret, license_key, uploaded_images)
        if rollback_message:
            raise RuntimeError(f"{exc}；已回滚本次已上传说明图：{rollback_message}") from exc
        raise

    if removed_image_urls:
        raw_payload = remove_product_description_image_urls(raw_payload, removed_image_urls)
    if not uploaded_images:
        return {"rawPayload": raw_payload, "uploadedImages": []}
    updated_payload = replace_product_description_image_urls(raw_payload, replacement_map)
    return {"rawPayload": updated_payload, "uploadedImages": uploaded_images}


def description_image_urls(html: Any) -> list[str]:
    soup = BeautifulSoup(str(html or ""), "lxml")
    urls: list[str] = []
    for image in soup.select("img, source"):
        for attr in ("src", "data-src", "data-original", "data-lazy-src"):
            url = normalize_description_image_url(image.get(attr))
            if url:
                urls.append(url)
        srcset = image.get("srcset")
        if srcset:
            for candidate_url, _descriptor in parse_srcset_candidates(srcset):
                url = normalize_description_image_url(candidate_url)
                if url:
                    urls.append(url)
    return unique_texts(urls)


def normalize_description_image_url(value: Any) -> str:
    text = normalize_text(value)
    if not text or text.lower().startswith(("data:", "blob:", "javascript:")):
        return ""
    if text.startswith(LOCAL_PRODUCT_IMAGE_URL_PREFIX):
        return text.split("?", 1)[0].split("#", 1)[0]
    if text.startswith("//"):
        text = f"https:{text}"
    if not text.startswith(("http://", "https://")):
        return ""
    return text


def should_transfer_description_image(url: str, target_shop_code: str) -> bool:
    text = normalize_description_image_url(url)
    if not text:
        return False
    parsed = urlsplit(text)
    host = parsed.netloc.lower()
    path = unquote(parsed.path or "").lower()
    normalized_shop_code = normalize_shop_code(target_shop_code).lower()
    if normalized_shop_code and host == "image.rakuten.co.jp":
        parts = [part for part in path.split("/") if part]
        if parts[:2] and parts[0] == normalized_shop_code and parts[1] == "cabinet":
            return False
    if normalized_shop_code and host in {"shop.r10s.jp", "tshop.r10s.jp"}:
        parts = [part for part in path.split("/") if part]
        if len(parts) >= 2 and parts[0] == normalized_shop_code and parts[1] == "cabinet":
            return False
    return True


def replace_product_description_image_urls(raw_payload: dict[str, Any], replacement_map: dict[str, str]) -> dict[str, Any]:
    if not replacement_map:
        return raw_payload
    updated_payload = json.loads(json.dumps(raw_payload, ensure_ascii=False))
    product_description = updated_payload.get("productDescription")
    if isinstance(product_description, dict):
        for key in ("pc", "sp", "smartphone", "value"):
            if key in product_description:
                product_description[key] = replace_description_html_image_urls(product_description.get(key), replacement_map)
    elif "productDescription" in updated_payload:
        updated_payload["productDescription"] = replace_description_html_image_urls(product_description, replacement_map)

    for key in ("description", "pcDescription", "spDescription", "smartphoneDescription", "salesDescription"):
        if key in updated_payload:
            updated_payload[key] = replace_description_html_image_urls(updated_payload.get(key), replacement_map)

    replace_embedded_item_description_image_urls(updated_payload, replacement_map)

    raw_descriptions = updated_payload.get("descriptions")
    if isinstance(raw_descriptions, list):
        for item in raw_descriptions:
            if isinstance(item, dict) and "value" in item:
                item["value"] = replace_description_html_image_urls(item.get("value"), replacement_map)
    return updated_payload


def replace_embedded_item_description_image_urls(raw_payload: dict[str, Any], replacement_map: dict[str, str]) -> None:
    embedded_item = raw_payload.get("embeddedItem")
    if not isinstance(embedded_item, dict):
        return
    pc_fields = embedded_item.get("pcFields")
    if isinstance(pc_fields, dict) and "productDescription" in pc_fields:
        pc_fields["productDescription"] = replace_description_html_image_urls(pc_fields.get("productDescription"), replacement_map)
    for key in ("newProductDescription", "salesDescription"):
        if key in embedded_item:
            embedded_item[key] = replace_description_html_image_urls(embedded_item.get(key), replacement_map)


def replace_description_html_image_urls(html: Any, replacement_map: dict[str, str]) -> str:
    if not html or not replacement_map:
        return str(html or "")
    soup = BeautifulSoup(str(html), "lxml")
    for image in soup.select("img, source"):
        for attr in ("src", "data-src", "data-original", "data-lazy-src"):
            value = normalize_description_image_url(image.get(attr))
            if value and value in replacement_map:
                image[attr] = replacement_map[value]
        srcset = image.get("srcset")
        if srcset:
            image["srcset"] = replace_srcset_image_urls(srcset, replacement_map)
    body = soup.body
    return body.decode_contents().strip() if body is not None else str(soup).strip()


def remove_product_description_image_urls(raw_payload: dict[str, Any], image_urls: list[str]) -> dict[str, Any]:
    normalized_urls = {normalize_description_image_url(url) for url in image_urls}
    normalized_urls.discard("")
    if not normalized_urls:
        return raw_payload
    updated_payload = json.loads(json.dumps(raw_payload, ensure_ascii=False))
    product_description = updated_payload.get("productDescription")
    if isinstance(product_description, dict):
        for key in ("pc", "sp", "smartphone", "value"):
            if key in product_description:
                product_description[key] = remove_description_html_image_urls(product_description.get(key), normalized_urls)
    elif "productDescription" in updated_payload:
        updated_payload["productDescription"] = remove_description_html_image_urls(product_description, normalized_urls)

    for key in ("description", "pcDescription", "spDescription", "smartphoneDescription", "salesDescription"):
        if key in updated_payload:
            updated_payload[key] = remove_description_html_image_urls(updated_payload.get(key), normalized_urls)

    remove_embedded_item_description_image_urls(updated_payload, normalized_urls)

    raw_descriptions = updated_payload.get("descriptions")
    if isinstance(raw_descriptions, list):
        for item in raw_descriptions:
            if isinstance(item, dict) and "value" in item:
                item["value"] = remove_description_html_image_urls(item.get("value"), normalized_urls)
    return updated_payload


def remove_embedded_item_description_image_urls(raw_payload: dict[str, Any], image_urls: set[str]) -> None:
    embedded_item = raw_payload.get("embeddedItem")
    if not isinstance(embedded_item, dict):
        return
    pc_fields = embedded_item.get("pcFields")
    if isinstance(pc_fields, dict) and "productDescription" in pc_fields:
        pc_fields["productDescription"] = remove_description_html_image_urls(pc_fields.get("productDescription"), image_urls)
    for key in ("newProductDescription", "salesDescription"):
        if key in embedded_item:
            embedded_item[key] = remove_description_html_image_urls(embedded_item.get(key), image_urls)


def remove_description_html_image_urls(html: Any, image_urls: set[str]) -> str:
    if not html or not image_urls:
        return str(html or "")
    soup = BeautifulSoup(str(html), "lxml")
    for image in soup.select("img, source"):
        should_remove = False
        for attr in ("src", "data-src", "data-original", "data-lazy-src"):
            value = normalize_description_image_url(image.get(attr))
            if value and value in image_urls:
                should_remove = True
                break
        srcset = image.get("srcset")
        if srcset:
            srcset_urls = {normalize_description_image_url(url) for url, _descriptor in parse_srcset_candidates(srcset)}
            srcset_urls.discard("")
            if srcset_urls and srcset_urls.issubset(image_urls):
                should_remove = True
        if should_remove:
            image.decompose()
    body = soup.body
    return body.decode_contents().strip() if body is not None else str(soup).strip()


def parse_srcset_candidates(value: str) -> list[tuple[str, str]]:
    candidates: list[tuple[str, str]] = []
    for raw_candidate in str(value or "").split(","):
        candidate = raw_candidate.strip()
        if not candidate:
            continue
        parts = candidate.split()
        url = parts[0] if parts else ""
        descriptor = " ".join(parts[1:]) if len(parts) > 1 else ""
        if url:
            candidates.append((url, descriptor))
    return candidates


def replace_srcset_image_urls(value: str, replacement_map: dict[str, str]) -> str:
    candidates = []
    for url, descriptor in parse_srcset_candidates(value):
        normalized_url = normalize_description_image_url(url)
        next_url = replacement_map.get(normalized_url, url)
        candidates.append(f"{next_url} {descriptor}".strip())
    return ", ".join(candidates)


def ensure_listing_cabinet_folder_for_upload(
    service_secret: str,
    license_key: str,
    store: StoreModel,
    required_slots: int,
    *,
    cabinet_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    context = cabinet_context if isinstance(cabinet_context, dict) else {}
    cached_folder = context.get("currentFolder") if isinstance(context.get("currentFolder"), dict) else None
    if cached_folder and cabinet_folder_remaining_slots(cached_folder) >= required_slots:
        return cached_folder
    folder = ensure_listing_cabinet_folder(
        service_secret,
        license_key,
        store,
        required_slots,
        usage=context.get("usage") if isinstance(context.get("usage"), dict) else None,
    )
    if cabinet_context is not None:
        cabinet_context["currentFolder"] = folder
    return folder


def reserve_listing_cabinet_folder_slots(
    cabinet_context: dict[str, Any] | None,
    folder: dict[str, Any],
    used_slots: int,
) -> None:
    if cabinet_context is None or not isinstance(folder, dict):
        return
    folder["fileCount"] = int(folder.get("fileCount") or 0) + max(0, used_slots)
    cabinet_context["currentFolder"] = folder


def cabinet_image_location(folder_path: str, file_path: str) -> str:
    normalized_folder = normalize_text(folder_path).strip("/")
    normalized_file = normalize_text(file_path).strip("/")
    if not normalized_folder:
        return normalized_file
    if normalized_file.lower().startswith(f"{normalized_folder.lower()}/"):
        return normalized_file
    return f"{normalized_folder}/{normalized_file}"


def listing_cabinet_upload_file_path(manage_number: str, index: int, suffix: str, *, kind: str) -> str:
    normalized_kind = re.sub(r"[^a-z0-9]+", "", normalize_text(kind).lower())[:1] or "p"
    normalized_suffix = normalize_text(suffix).lower()
    if normalized_suffix == ".jpeg":
        normalized_suffix = ".jpg"
    if normalized_suffix not in {".jpg", ".png", ".gif"}:
        normalized_suffix = ".jpg"
    digest = hashlib.sha1(normalize_text(manage_number).encode("utf-8")).hexdigest()[:8]
    stem = f"{normalized_kind}{digest}{max(1, index):03d}"
    return normalize_cabinet_file_path(f"{stem}{normalized_suffix}")


def listing_cabinet_upload_file_name(file_path: str) -> str:
    return normalize_cabinet_file_name(Path(normalize_text(file_path)).stem)


def rollback_uploaded_listing_images(
    service_secret: str,
    license_key: str,
    uploaded_images: list[dict[str, str]],
) -> str:
    if not uploaded_images:
        return ""
    deleted_count = 0
    warnings: list[str] = []
    attempted_ids: set[int] = set()
    for image in uploaded_images:
        file_ids = cabinet_file_ids_from_uploaded_image(service_secret, license_key, image)
        if not file_ids:
            warnings.append(f"{image.get('location') or image.get('filePath') or image.get('fileUrl') or '-'} 未找到文件ID")
            continue
        for file_id in file_ids:
            if file_id in attempted_ids:
                continue
            attempted_ids.add(file_id)
            try:
                delete_rakuten_cabinet_file(service_secret, license_key, file_id)
                deleted_count += 1
            except Exception as exc:
                warnings.append(f"图片 {file_id} 删除失败：{exc}")
    message = f"删除 {deleted_count} 张"
    if warnings:
        message = f"{message}，警告：{'；'.join(warnings[:5])}"
    return message


def cabinet_file_ids_from_uploaded_image(
    service_secret: str,
    license_key: str,
    image: dict[str, str],
) -> list[int]:
    raw_file_id = normalize_text(image.get("fileId"))
    if raw_file_id:
        try:
            return [int(float(raw_file_id))]
        except ValueError:
            pass
    target = {
        "filePath": image.get("location") or image.get("filePath") or "",
        "fileName": Path(normalize_text(image.get("location") or image.get("filePath") or "")).name,
    }
    if target["filePath"] or target["fileName"]:
        try:
            return resolve_cabinet_file_ids(service_secret, license_key, target)
        except Exception:
            return []
    return []


def should_skip_listing_image_error(exc: Exception) -> bool:
    message = str(exc)
    return (
        is_missing_local_product_image_error(exc)
        or "图片文件无法识别" in message
        or "图片文件无法读取" in message
        or "R-Cabinet 图片格式只支持" in message
        or "Unsupported format Error" in message
        or "GIF 图片" in message
    )


def is_listing_image_read_error(exc: Exception) -> bool:
    message = str(exc)
    return any(
        marker in message
        for marker in (
            "读取 OSS 商品图片失败",
            "读取商品图片失败",
            "商品图片文件不存在",
            "本地图片文件不存在",
        )
    )


def recover_missing_local_product_images(product: ProductModel, images: list[str]) -> list[str]:
    normalized_images = unique_texts([image for image in images if normalize_product_image_url(image)])
    if not any(is_missing_local_product_image_url(image) for image in normalized_images):
        return normalized_images
    raw_payload = product_raw_payload(product)
    shop_code = product_shop_code(product, raw_payload)
    original_images = [
        image
        for image in product_original_image_urls(raw_payload, shop_code=shop_code)
        if not is_gif_image_url(image)
    ]
    if not original_images:
        fallback_payload = dict(raw_payload)
        fallback_payload.pop("ltEditedImages", None)
        fallback_payload.pop("images", None)
        original_images = [
            image
            for image in product_image_urls(fallback_payload, shop_code=shop_code)
            if not is_local_product_image_url(image)
            and not is_product_image_draft_url(image)
            and not is_gif_image_url(image)
        ]
    if original_images:
        try:
            image_result = localize_product_image_urls(product.id, original_images, prefix="r")
        except Exception:
            image_result = {"urls": []}
        recovered_urls = image_result.get("urls") if isinstance(image_result, dict) else []
        if recovered_urls:
            existing_urls = [image for image in normalized_images if not is_missing_local_product_image_url(image)]
            next_urls = unique_texts([*existing_urls, *recovered_urls])
            persist_recovered_product_images(product, raw_payload, next_urls)
            return next_urls
    return [image for image in normalized_images if not is_missing_local_product_image_url(image)]


def persist_recovered_product_images(product: ProductModel, raw_payload: dict[str, Any], image_urls: list[str]) -> None:
    if not image_urls:
        return
    try:
        updated_payload = set_product_image_urls(raw_payload, image_urls)
        product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        product.image_url = image_urls[0]
        if product.last_error and any(
            marker in product.last_error
            for marker in ("本地图片文件不存在", "商品图片文件不存在")
        ):
            product.last_error = None
    except Exception:
        return


def product_original_image_urls(raw_payload: dict[str, Any], *, shop_code: str = "") -> list[str]:
    original_images = raw_payload.get("ltOriginalImages")
    if isinstance(original_images, list):
        return unique_texts([
            normalize_product_image_url(image, shop_code=shop_code)
            for image in original_images
            if normalize_product_image_url(image, shop_code=shop_code)
        ])
    return []


def is_missing_local_product_image_url(image_url: str) -> bool:
    local_path = local_product_image_path_from_url(image_url)
    stored_image = parse_product_image_url(image_url)
    if stored_image is not None and product_image_storage.enabled:
        try:
            if product_image_storage.exists(stored_image.object_key):
                return False
        except Exception:
            if local_path and local_path.exists():
                return False
            return False
    return bool(local_path and not local_path.exists())


def is_missing_local_product_image_error(exc: Exception) -> bool:
    message = str(exc)
    return "本地图片文件不存在" in message or "商品图片文件不存在" in message


def load_product_image_bytes(
    image_url: str,
    *,
    max_bytes: int = MAX_PRODUCT_IMAGE_BYTES,
    size_error_message: str = "图片大小不能超过 2MB。",
) -> dict[str, Any]:
    normalized_max_bytes = max(1, int(max_bytes or MAX_PRODUCT_IMAGE_BYTES))
    local_path = local_product_image_path_from_url(image_url)
    stored_image = parse_product_image_url(image_url)
    storage_error: Exception | None = None
    if stored_image is not None and product_image_storage.enabled:
        try:
            content = product_image_storage.read_bytes(
                stored_image.object_key,
                max_bytes=normalized_max_bytes,
            )
            suffix = Path(stored_image.filename).suffix.lower()
            content_type = mimetypes.guess_type(stored_image.filename)[0] or "application/octet-stream"
        except FileNotFoundError:
            content = b""
            suffix = ""
            content_type = ""
        except Exception as exc:
            storage_error = exc
            content = b""
            suffix = ""
            content_type = ""
    else:
        content = b""
        suffix = ""
        content_type = ""
    if not content and local_path:
        if local_path.exists():
            content = local_path.read_bytes()
            suffix = local_path.suffix.lower()
            content_type = mimetypes.guess_type(str(local_path))[0] or "application/octet-stream"
        elif storage_error is not None:
            raise RuntimeError("读取 OSS 商品图片失败。") from storage_error
        elif stored_image is not None:
            raise RuntimeError("商品图片文件不存在。")
        else:
            raise RuntimeError("本地图片文件不存在。")
    if not content:
        remote_image = download_remote_product_image(
            image_url,
            max_bytes=normalized_max_bytes,
            size_error_message=size_error_message,
        )
        content = remote_image["content"]
        suffix = remote_image["suffix"]
        content_type = remote_image["contentType"]
    if suffix == ".jpeg":
        suffix = ".jpg"
    if suffix not in ALLOWED_PRODUCT_IMAGE_EXTENSIONS:
        raise RuntimeError("图片格式只支持 jpg、jpeg、png、gif。")
    if suffix == ".jpg":
        content_type = "image/jpeg"
    elif suffix == ".png":
        content_type = "image/png"
    elif suffix == ".gif":
        content_type = "image/gif"
    if content_type not in ALLOWED_PRODUCT_IMAGE_MIME_TYPES:
        raise RuntimeError("图片文件类型不正确。")
    if not content:
        raise RuntimeError("图片内容为空。")
    if len(content) > normalized_max_bytes:
        raise RuntimeError(size_error_message)
    return {"content": content, "suffix": suffix, "contentType": content_type}


def download_remote_product_image(
    image_url: str,
    *,
    max_bytes: int,
    size_error_message: str,
) -> dict[str, Any]:
    request_url = corrected_remote_product_image_url(image_url)
    proxy_config = crawler_request_proxies()
    attempts: list[dict[str, str] | None] = [None]
    if proxy_config:
        attempts.append(proxy_config)
    last_error: Exception | None = None
    for attempt_index, proxies in enumerate(attempts):
        response: requests.Response | None = None
        try:
            response = requests.get(
                request_url,
                timeout=settings.crawler_timeout_seconds,
                headers={"User-Agent": settings.crawler_user_agent},
                proxies=proxies,
                stream=True,
            )
            status_code = int(response.status_code or 0)
            if status_code in {404, 410}:
                raise ProductImageUnavailableError(status_code)
            if (
                proxies is None
                and proxy_config
                and (status_code in {403, 429} or status_code >= 500)
            ):
                last_error = RuntimeError(f"读取商品图片失败（HTTP {status_code}）。")
                continue
            response.raise_for_status()
            chunks: list[bytes] = []
            size = 0
            for chunk in response.iter_content(chunk_size=1024 * 256):
                if not chunk:
                    continue
                size += len(chunk)
                if size > max_bytes:
                    raise RuntimeError(size_error_message)
                chunks.append(chunk)
            content = b"".join(chunks)
            suffix = Path(urlsplit(request_url).path).suffix.lower()
            content_type = normalize_text(response.headers.get("Content-Type")).split(";", 1)[0].lower()
            if not suffix:
                suffix = product_image_suffix_from_content_type(content_type)
            return {
                "content": content,
                "suffix": suffix,
                "contentType": content_type,
            }
        except ProductImageUnavailableError:
            raise
        except requests.RequestException as exc:
            last_error = exc
            if attempt_index + 1 < len(attempts):
                continue
            break
        finally:
            if response is not None:
                response.close()
    raise RuntimeError("读取商品图片失败。") from last_error


def corrected_remote_product_image_url(image_url: str) -> str:
    parsed = urlsplit(image_url)
    if parsed.path.lower().endswith(".pg"):
        return parsed._replace(path=f"{parsed.path[:-3]}.jpg").geturl()
    return image_url


def prepare_rakuten_cabinet_image(image_data: dict[str, Any]) -> dict[str, Any]:
    content = image_data.get("content") or b""
    suffix = normalize_text(image_data.get("suffix")).lower()
    content_type = normalize_text(image_data.get("contentType")).lower()
    if not content:
        raise RuntimeError("图片内容为空。")
    if suffix == ".jpeg":
        suffix = ".jpg"
    if suffix == ".gif":
        raise RuntimeError("GIF 图片不参与上架，请替换为 jpg/png。")
    if suffix not in {".jpg", ".png"}:
        raise RuntimeError("R-Cabinet 图片格式只支持 jpg、png、gif。")
    try:
        from PIL import Image, ImageOps, UnidentifiedImageError
    except ImportError as exc:
        raise RuntimeError("服务器缺少 Pillow，不能自动处理 R-Cabinet 图片尺寸，请先执行 pip install -r requirements.txt。") from exc

    try:
        with Image.open(BytesIO(content)) as source:
            image = ImageOps.exif_transpose(source)
            image.load()
    except UnidentifiedImageError as exc:
        raise RuntimeError("图片文件无法识别。") from exc
    except OSError as exc:
        raise RuntimeError("图片文件无法读取。") from exc

    image = resize_image_to_max_dimension(image, RAKUTEN_CABINET_MAX_IMAGE_DIMENSION)
    if image.mode in {"RGBA", "LA"} or (image.mode == "P" and "transparency" in image.info):
        background = Image.new("RGB", image.size, (255, 255, 255))
        transparent_image = image.convert("RGBA")
        background.paste(transparent_image, mask=transparent_image.getchannel("A"))
        image = background
    elif image.mode not in {"RGB", "L"}:
        image = image.convert("RGB")
    elif image.mode == "L":
        image = image.convert("RGB")

    quality_values = (88, 82, 76, 70, 64, 58)
    current = image
    for scale_attempt in range(5):
        if scale_attempt:
            next_width = max(1, int(current.width * 0.9))
            next_height = max(1, int(current.height * 0.9))
            current = current.resize((next_width, next_height), Image.Resampling.LANCZOS)
        for quality in quality_values:
            output = BytesIO()
            current.save(output, format="JPEG", quality=quality, optimize=True, progressive=True)
            normalized_content = output.getvalue()
            if len(normalized_content) <= RAKUTEN_CABINET_MAX_IMAGE_BYTES:
                return {"content": normalized_content, "suffix": ".jpg", "contentType": "image/jpeg"}
    raise RuntimeError("图片压缩后仍超过 R-Cabinet 2MB 限制，请替换为更小的图片。")


def validate_rakuten_cabinet_gif(content: bytes) -> None:
    if len(content) > RAKUTEN_CABINET_MAX_IMAGE_BYTES:
        raise RuntimeError("GIF 图片超过 R-Cabinet 2MB 限制，请替换为 jpg/png。")
    try:
        from PIL import Image, UnidentifiedImageError
    except ImportError as exc:
        raise RuntimeError("服务器缺少 Pillow，不能自动检查 R-Cabinet 图片尺寸，请先执行 pip install -r requirements.txt。") from exc
    try:
        with Image.open(BytesIO(content)) as image:
            width, height = image.size
    except UnidentifiedImageError as exc:
        raise RuntimeError("GIF 图片文件无法识别。") from exc
    except OSError as exc:
        raise RuntimeError("GIF 图片文件无法读取。") from exc
    if width > RAKUTEN_CABINET_MAX_IMAGE_DIMENSION or height > RAKUTEN_CABINET_MAX_IMAGE_DIMENSION:
        raise RuntimeError("GIF 图片尺寸超过 R-Cabinet 限制，请替换为 jpg/png。")


def resize_image_to_max_dimension(image: Any, max_dimension: int) -> Any:
    width, height = image.size
    if width <= max_dimension and height <= max_dimension:
        return image.copy()
    scale = min(max_dimension / width, max_dimension / height)
    next_size = (max(1, int(width * scale)), max(1, int(height * scale)))
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("服务器缺少 Pillow，不能自动处理 R-Cabinet 图片尺寸，请先执行 pip install -r requirements.txt。") from exc
    return image.resize(next_size, Image.Resampling.LANCZOS)


def product_image_suffix_from_content_type(content_type: str) -> str:
    normalized = normalize_text(content_type).lower()
    if normalized == "image/jpeg":
        return ".jpg"
    if normalized == "image/png":
        return ".png"
    if normalized == "image/gif":
        return ".gif"
    return ""


def build_rakuten_item_upsert_payload(
    product: ProductModel,
    raw_payload: dict[str, Any],
    uploaded_images: list[dict[str, str]],
    *,
    manage_number: str | None = None,
    hide_item: bool = False,
) -> dict[str, Any]:
    title = first_text_from_keys(raw_payload, ("itemName", "title", "name")) or product.title
    title = normalize_text(title)
    if not title:
        raise RuntimeError("商品标题为空，不能上架到乐天。")
    genre_id = first_text_from_keys(raw_payload, ("genreId", "genre_id", "genre")) or product.genre_id
    if not re.fullmatch(r"\d{6}", normalize_text(genre_id)):
        raise RuntimeError("商品缺少 6 位乐天ジャンルID，不能上架到乐天。")
    variants = build_rakuten_listing_variants(raw_payload, product)
    if not variants:
        raise RuntimeError("商品缺少 SKU 价格信息，不能上架到乐天。")
    item_number = normalize_text(manage_number) or normalize_text(product.rakuten_manage_number) or generate_listing_manage_number(product, raw_payload)
    payload: dict[str, Any] = {
        "itemNumber": item_number[:32],
        "title": truncate_utf8_bytes(title, RAKUTEN_TITLE_MAX_BYTES),
        "tagline": truncate_utf8_bytes(product_tagline(raw_payload), RAKUTEN_TAGLINE_MAX_BYTES),
        "itemType": "NORMAL",
        "genreId": normalize_text(genre_id),
        "hideItem": bool(hide_item),
        "unlimitedInventoryFlag": False,
        "images": build_rakuten_listing_images(uploaded_images, title),
        "productDescription": build_rakuten_product_description(raw_payload),
        "salesDescription": build_rakuten_sales_description(raw_payload),
        "variantSelectors": build_rakuten_variant_selectors(raw_payload, variants),
        "variants": variants,
    }
    payload = apply_rakuten_attribute_rules_to_payload(payload)
    return {key: value for key, value in payload.items() if value not in (None, "", [], {})}


def build_rakuten_listing_images(uploaded_images: list[dict[str, str]], title: str) -> list[dict[str, str]]:
    images: list[dict[str, str]] = []
    for image in uploaded_images:
        location = normalize_rakuten_item_image_location(image.get("location"))
        if not location:
            continue
        alt = sanitize_rakuten_image_alt(image.get("alt") or title) or "商品画像"
        images.append({"type": "CABINET", "location": location, "alt": alt})
        if len(images) >= RAKUTEN_LISTING_IMAGE_LIMIT:
            break
    return images


def normalize_rakuten_item_image_location(value: Any) -> str:
    location = normalize_text(value)
    if not location:
        return ""
    if location.startswith(("http://", "https://")):
        parsed = urlsplit(location)
        path = unquote(parsed.path)
        cabinet_index = path.lower().find("/cabinet/")
        if cabinet_index >= 0:
            location = path[cabinet_index + len("/cabinet/") :]
        else:
            return ""
    location = location.strip()
    if location.lower().startswith("cabinet/"):
        location = location[len("cabinet/") :]
    location = location.lstrip("/")
    if not location or "/" not in location:
        return ""
    return f"/{location}"


def build_rakuten_listing_variants(raw_payload: dict[str, Any], product: ProductModel) -> dict[str, dict[str, Any]]:
    raw_variants = raw_payload.get("variants")
    variant_items: list[tuple[str, dict[str, Any]]] = []
    fallback_price = price_from_rakuten_item_without_variants(raw_payload)
    if fallback_price is None and product.price is not None:
        fallback_price = float(product.price)
    if isinstance(raw_variants, dict):
        variant_items = [(normalize_text(key), value) for key, value in raw_variants.items() if isinstance(value, dict)]
    elif isinstance(raw_variants, list):
        for index, value in enumerate(raw_variants, start=1):
            if isinstance(value, dict):
                variant_id = first_text_from_keys(value, ("variantId", "skuId", "merchantDefinedSkuId")) or f"sku-{index}"
                variant_items.append((variant_id, value))
    if not variant_items:
        price = fallback_price
        if price is None:
            return {}
        variant_items = [("default", {"standardPrice": str(int(price)), "selectorValues": {}})]
    if fallback_price is None:
        variant_price_values = [
            price
            for _, variant in variant_items
            if isinstance(variant, dict)
            for price in [price_from_rakuten_item_without_variants(variant)]
            if price is not None
        ]
        if variant_price_values:
            fallback_price = min(variant_price_values)

    result: dict[str, dict[str, Any]] = {}
    for index, (variant_id, variant) in enumerate(variant_items, start=1):
        normalized_variant_id = re.sub(r"[^A-Za-z0-9_-]+", "-", normalize_text(variant_id)).strip("-_") or f"sku-{index}"
        price_text = first_text_from_keys(variant, RAKUTEN_PRICE_KEYS)
        if not price_text and fallback_price is not None:
            price_text = str(int(fallback_price)) if fallback_price == int(fallback_price) else str(fallback_price)
        normalized_price = normalize_rakuten_price(price_text)
        if not normalized_price:
            continue
        selector_values = variant.get("selectorValues") if isinstance(variant.get("selectorValues"), dict) else {}
        next_variant: dict[str, Any] = {
            "standardPrice": normalized_price,
            "hidden": bool(variant.get("hidden", False)),
            "articleNumber": normalize_article_number(variant.get("articleNumber")),
        }
        merchant_sku = first_text_from_keys(variant, ("merchantDefinedSkuId",))
        if merchant_sku:
            next_variant["merchantDefinedSkuId"] = merchant_sku[:96]
        if selector_values:
            next_variant["selectorValues"] = {
                normalize_text(key): normalize_text(value)[:32]
                for key, value in selector_values.items()
                if normalize_text(key) and normalize_text(value)
            }
        attributes = normalize_rakuten_variant_attributes(variant.get("attributes"))
        attributes = ensure_rakuten_variant_required_attributes(attributes, variant, raw_payload)
        if attributes:
            next_variant["attributes"] = attributes
        result[normalized_variant_id[:32]] = {key: value for key, value in next_variant.items() if value not in (None, "", {}, [])}
    return result


def build_rakuten_inventory_upsert_payloads(
    manage_number: str,
    variants: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    normalized_manage_number = normalize_text(manage_number)
    if not normalized_manage_number or not variants:
        return []
    quantity = max(0, int(settings.rakuten_default_inventory_quantity))
    normal_delivery_time_id = int(settings.rakuten_default_normal_delivery_time_id)
    back_order_delivery_time_id = int(settings.rakuten_default_back_order_delivery_time_id)
    inventories: list[dict[str, Any]] = []
    for variant_id, variant in variants.items():
        normalized_variant_id = normalize_text(variant_id)
        if not normalized_variant_id:
            continue
        inventory: dict[str, Any] = {
            "manageNumber": normalized_manage_number,
            "variantId": normalized_variant_id,
            "mode": "ABSOLUTE",
            "quantity": quantity,
        }
        operation_lead_time: dict[str, int] = {}
        if normal_delivery_time_id > 0:
            operation_lead_time["normalDeliveryTimeId"] = normal_delivery_time_id
        if back_order_delivery_time_id > 0:
            operation_lead_time["backOrderDeliveryTimeId"] = back_order_delivery_time_id
        if operation_lead_time:
            inventory["operationLeadTime"] = operation_lead_time
        ship_from_ids = variant.get("shipFromIds") if isinstance(variant, dict) else None
        if isinstance(ship_from_ids, list):
            normalized_ship_from_ids: list[int] = []
            for value in ship_from_ids:
                try:
                    ship_from_id = int(value)
                except (TypeError, ValueError):
                    continue
                if ship_from_id > 0 and ship_from_id not in normalized_ship_from_ids:
                    normalized_ship_from_ids.append(ship_from_id)
            if normalized_ship_from_ids:
                inventory["shipFromIds"] = normalized_ship_from_ids
        inventories.append(inventory)
    return inventories


def normalize_rakuten_variant_attributes(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    attributes: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        name = first_text_from_keys(item, ("name", "attributeName", "label"))
        attribute_value = first_text_from_keys(item, ("value", "attributeValue", "text", "values"))
        if not name or not attribute_value or name in seen_names:
            continue
        allow_placeholder = normalize_text(name) in RAKUTEN_ATTRIBUTE_ALLOW_PLACEHOLDER_NAMES
        normalized_value = normalize_rakuten_attribute_value(attribute_value, allow_placeholder=allow_placeholder)
        if not normalized_value:
            continue
        unit = normalize_rakuten_attribute_unit(first_text_from_keys(item, ("unit",)))
        values, unit = normalize_rakuten_attribute_values_and_unit(name, normalized_value, unit)
        if not values:
            continue
        seen_names.add(name)
        attribute: dict[str, Any] = {"name": name, "values": values}
        if unit:
            attribute["unit"] = unit
        attributes.append(attribute)
    return attributes


def ensure_rakuten_variant_required_attributes(
    attributes: list[dict[str, Any]],
    variant: dict[str, Any],
    raw_payload: dict[str, Any],
) -> list[dict[str, Any]]:
    return list(attributes)


def infer_rakuten_representative_color(variant: dict[str, Any], raw_payload: dict[str, Any]) -> str:
    candidates: list[Any] = []
    for source in (variant, raw_payload):
        fields: list[Any] = []
        if isinstance(source, dict):
            fields.extend([
                source.get("attributes"),
                source.get("specs"),
            ])
        for attrs in fields:
            if not isinstance(attrs, list):
                continue
            for item in attrs:
                if not isinstance(item, dict):
                    continue
                name = first_text_from_keys(item, ("name", "attributeName", "label", "title"))
                value = first_text_from_keys(item, ("value", "attributeValue", "text", "values"))
                normalized_name = normalize_text(name).lower()
                if normalize_text(name) == RAKUTEN_REPRESENTATIVE_COLOR_ATTRIBUTE:
                    candidates.append(value)
                elif any(keyword.lower() in normalized_name for keyword in RAKUTEN_COLOR_SELECTOR_KEYWORDS):
                    candidates.append(value)
    selector_values = variant.get("selectorValues") if isinstance(variant.get("selectorValues"), dict) else {}
    for key, value in selector_values.items():
        normalized_key = normalize_text(key).lower()
        if any(keyword.lower() in normalized_key for keyword in RAKUTEN_COLOR_SELECTOR_KEYWORDS):
            candidates.append(value)
    candidates.extend(selector_values.values())
    candidates.extend(
        [
            first_text_from_keys(variant, ("variantId", "skuId", "merchantDefinedSkuId")),
            first_text_from_keys(raw_payload, ("title", "itemName", "name")),
        ]
    )
    for candidate in candidates:
        color = normalize_rakuten_representative_color(candidate)
        if color:
            return color
    return ""


def normalize_rakuten_representative_color(value: Any) -> str:
    text = unicodedata.normalize("NFKC", normalize_text(value))
    if not text:
        return ""
    if text in RAKUTEN_ATTRIBUTE_PLACEHOLDER_VALUES:
        return ""
    lowered = text.lower()
    for token, color in sorted(RAKUTEN_COLOR_VALUE_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        if token.lower() in lowered and color in RAKUTEN_REPRESENTATIVE_COLOR_ALLOWED_VALUES:
            return color
    return text if text in RAKUTEN_REPRESENTATIVE_COLOR_ALLOWED_VALUES else ""


def normalize_rakuten_attribute_value(value: Any, *, allow_placeholder: bool = False) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKC", text)
    if normalized in RAKUTEN_ATTRIBUTE_PLACEHOLDER_VALUES and not allow_placeholder:
        return ""
    return normalized


def normalize_rakuten_attribute_values_and_unit(
    name: str,
    value: str,
    unit: str,
    *,
    default_unit: str = "",
) -> tuple[list[str], str]:
    normalized_name = normalize_text(name)
    normalized_value = normalize_rakuten_attribute_value(
        value,
        allow_placeholder=normalized_name in RAKUTEN_ATTRIBUTE_ALLOW_PLACEHOLDER_NAMES,
    )
    normalized_unit = normalize_rakuten_attribute_unit(unit)
    normalized_default_unit = normalize_rakuten_attribute_unit(default_unit)
    if not normalized_value:
        return [], ""
    if normalized_name in RAKUTEN_ATTRIBUTE_TEXT_ONLY_NAMES:
        return [normalized_value], ""
    parsed_number, parsed_unit = parse_rakuten_attribute_number_and_unit(normalized_value)
    if parsed_unit and not normalized_unit:
        normalized_unit = parsed_unit
    if not normalized_unit:
        normalized_unit = normalized_default_unit
    if not normalized_unit:
        normalized_unit = RAKUTEN_ATTRIBUTE_DEFAULT_UNITS.get(normalized_name, "")
    if normalized_unit:
        number = parsed_number or normalize_rakuten_attribute_number(normalized_value)
        if not number:
            return [], ""
        return [number], normalized_unit
    return [normalized_value], ""


def parse_rakuten_attribute_number_and_unit(value: Any) -> tuple[str, str]:
    normalized = unicodedata.normalize("NFKC", normalize_text(value))
    if not normalized:
        return "", ""
    unit_pattern = "|".join(sorted((re.escape(unit) for unit in RAKUTEN_ATTRIBUTE_UNIT_ALIASES), key=len, reverse=True))
    match = re.search(rf"([0-9]+(?:[.,][0-9]+)?)\s*({unit_pattern})\b", normalized, flags=re.I)
    if not match:
        return "", ""
    return normalize_rakuten_attribute_number(match.group(1)), normalize_rakuten_attribute_unit(match.group(2))


def normalize_rakuten_attribute_unit(value: Any) -> str:
    normalized = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    if not normalized:
        return ""
    normalized = re.sub(r"\s+", "", normalized)
    return RAKUTEN_ATTRIBUTE_UNIT_ALIASES.get(normalized, normalized)


def normalize_rakuten_attribute_number(value: Any) -> str:
    normalized = unicodedata.normalize("NFKC", normalize_text(value)).replace(",", ".")
    match = re.search(r"[0-9]+(?:\.[0-9]+)?", normalized)
    if not match:
        return ""
    number = match.group(0)
    if "." in number:
        number = number.rstrip("0").rstrip(".")
    return number


def normalize_rakuten_price(value: Any) -> str:
    text = first_text_value(value)
    normalized = re.sub(r"[^0-9.]", "", text)
    if not normalized:
        return ""
    try:
        price = Decimal(normalized)
    except Exception:
        return ""
    if price <= 0 or price != price.to_integral_value():
        return ""
    return str(int(price))


def normalize_article_number(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        article_value = first_text_from_keys(value, ("value", "articleNumber"))
        exemption_reason = value.get("exemptionReason")
        if article_value:
            return {"value": article_value}
        try:
            reason = int(exemption_reason)
        except (TypeError, ValueError):
            reason = 5
        return {"exemptionReason": reason}
    text = first_text_value(value)
    if text:
        return {"value": text}
    return {"exemptionReason": 5}


def build_rakuten_variant_selectors(raw_payload: dict[str, Any], variants: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    selectors = product_variant_selectors(raw_payload)
    selector_names = {normalize_text(selector.get("key")): normalize_text(selector.get("name")) for selector in selectors}
    normalized_selectors: list[dict[str, Any]] = []
    selector_keys_used = {
        key
        for variant in variants.values()
        for key in (variant.get("selectorValues") if isinstance(variant.get("selectorValues"), dict) else {}).keys()
    }
    for index, selector in enumerate(selectors, start=1):
        key = normalize_text(selector.get("key")) or f"choice-{index}"
        if key not in selector_keys_used:
            continue
        values = []
        seen: set[str] = set()
        for variant in variants.values():
            selector_values = variant.get("selectorValues") if isinstance(variant.get("selectorValues"), dict) else {}
            value = normalize_text(selector_values.get(key))
            if value and value not in seen:
                seen.add(value)
                values.append({"displayValue": value[:32]})
        if values:
            normalized_selectors.append(
                {
                    "key": key[:32],
                    "displayName": (normalize_text(selector.get("name")) or key)[:32],
                    "values": values[:40],
                }
            )
    existing_keys = {selector["key"] for selector in normalized_selectors}
    for key in selector_keys_used:
        if key in existing_keys:
            continue
        values = []
        seen: set[str] = set()
        for variant in variants.values():
            selector_values = variant.get("selectorValues") if isinstance(variant.get("selectorValues"), dict) else {}
            value = normalize_text(selector_values.get(key))
            if value and value not in seen:
                seen.add(value)
                values.append({"displayValue": value[:32]})
        if values:
            normalized_selectors.append(
                {
                    "key": key[:32],
                    "displayName": (selector_names.get(key) or key)[:32],
                    "values": values[:40],
                }
            )
    return normalized_selectors


def build_rakuten_product_description(raw_payload: dict[str, Any]) -> dict[str, str]:
    descriptions = product_descriptions(raw_payload)
    pc_description = first_description_by_label(descriptions, ("PC用 商品説明文",))
    sp_description = first_description_by_label(descriptions, ("スマートフォン用 商品説明文",))
    pc_html = sanitize_rakuten_listing_description_html(pc_description, max_length=10240)
    sp_html = sanitize_rakuten_sp_description_html(sp_description, max_length=10240)
    result: dict[str, str] = {}
    if pc_html:
        result["pc"] = pc_html
    if sp_html:
        result["sp"] = sp_html
    return result


def build_rakuten_sales_description(raw_payload: dict[str, Any]) -> str:
    descriptions = product_descriptions(raw_payload)
    return sanitize_rakuten_listing_description_html(
        first_description_by_label(descriptions, ("PC用 販売説明文",)),
        max_length=10240,
    )


def sanitize_rakuten_listing_description_html(value: Any, *, max_length: int) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = re.sub(r"<\s*thcolspan(\s|>)", r'<th colspan="2"\1', text, flags=re.I)
    text = re.sub(r"</\s*thcolspan\s*>", "</th>", text, flags=re.I)
    text = sanitize_rakuten_pc_description_html(text)
    text = re.sub(r"<\s*thcolspan(\s|>)", r'<th colspan="2"\1', text, flags=re.I)
    text = re.sub(r"</\s*thcolspan\s*>", "</th>", text, flags=re.I)
    text = normalize_rakuten_machine_dependent_characters(text)
    return truncate_text(text, max_length)


def sanitize_rakuten_pc_description_html(value: Any) -> str:
    soup = BeautifulSoup(str(value or ""), "lxml")
    for comment in soup.find_all(string=lambda item: isinstance(item, Comment)):
        comment.extract()
    for element in soup.select("script, object, embed, link, meta, svg, canvas, video, audio, form, input, select, textarea, button"):
        element.decompose()
    for element in soup.select("*"):
        for attribute in list(element.attrs):
            name = normalize_text(attribute).lower()
            attr_values = element.get_attribute_list(attribute)
            value_text = " ".join(str(value) for value in attr_values).strip()
            if is_unsafe_html_attribute_value(name, value_text):
                del element.attrs[attribute]
    body = soup.body
    return body.decode_contents().strip() if body is not None else str(soup).strip()


def sanitize_rakuten_sp_description_html(value: Any, *, max_length: int) -> str:
    text = sanitize_rakuten_listing_description_html(value, max_length=max_length)
    if not text:
        return ""
    soup = BeautifulSoup(text, "lxml")
    sanitize_rakuten_sp_description_soup(soup)
    image_count = 0
    for image in soup.select("img"):
        image_count += 1
        if image_count > RAKUTEN_SP_DESCRIPTION_IMAGE_LIMIT:
            image.decompose()
    body = soup.body
    cleaned = body.decode_contents().strip() if body is not None else str(soup).strip()
    cleaned = normalize_rakuten_machine_dependent_characters(cleaned)
    return truncate_text(cleaned, max_length)


def sanitize_rakuten_sp_description_soup(soup: BeautifulSoup) -> None:
    for comment in soup.find_all(string=lambda value: isinstance(value, Comment)):
        comment.extract()
    for element in list(soup.find_all(True)):
        tag_name = normalize_text(element.name).lower()
        if tag_name in RAKUTEN_SP_DESCRIPTION_DROP_TAGS:
            element.decompose()
            continue
        if tag_name in {"html", "body"}:
            continue
        if tag_name not in RAKUTEN_SP_DESCRIPTION_ALLOWED_TAGS:
            element.unwrap()
            continue
        allowed_attributes = set(RAKUTEN_SP_DESCRIPTION_ALLOWED_ATTRIBUTES.get("*", set()))
        allowed_attributes.update(RAKUTEN_SP_DESCRIPTION_ALLOWED_ATTRIBUTES.get(tag_name, set()))
        for attribute in list(element.attrs):
            attr_name = normalize_text(attribute).lower()
            attr_values = element.get_attribute_list(attribute)
            attr_value = " ".join(str(value) for value in attr_values).strip()
            if attr_name not in allowed_attributes or is_unsafe_html_attribute_value(attr_name, attr_value):
                del element.attrs[attribute]


def is_unsafe_html_attribute_value(name: str, value: str) -> bool:
    normalized_name = normalize_text(name).lower()
    normalized_value = normalize_text(value).lower()
    if normalized_name.startswith("on"):
        return True
    if normalized_value.startswith(("javascript:", "data:", "vbscript:")):
        return True
    if normalized_name in {"src", "href"} and normalized_value and not normalized_value.startswith(("http://", "https://", "/", "#", "mailto:", "tel:")):
        return True
    return False


def truncate_text(value: Any, max_length: int) -> str:
    text = str(value or "")
    return text[:max_length]


def update_product_local_detail(owner_username: str, product_id: int, payload: Any) -> dict[str, Any]:
    cleanup_urls: list[str] = []
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        if product.review_status == "listed":
            raise RuntimeError("店铺商品请使用同步修改。")
        if product.review_status != "pending" and getattr(payload, "imageChanges", None):
            raise RuntimeError("只有待审核商品可以修改图片。")
        genre_id = normalize_text(getattr(payload, "genreId", None))
        if genre_id:
            if product.review_status not in {"pending", "listed_master"}:
                raise RuntimeError("只有待审核商品或已上架商品可以修改品类。")
            if not re.fullmatch(r"\d{6}", genre_id) or not rakuten_genre_path(genre_id):
                raise RuntimeError("请选择有效品类。")

        updated_payload = patch_local_item_detail(
            product_raw_payload(product),
            title=getattr(payload, "title", ""),
            tagline=getattr(payload, "tagline", ""),
            variants=list(getattr(payload, "variants", []) or []),
        )
        if genre_id:
            updated_payload["genreId"] = genre_id
            product.genre_id = genre_id
        image_changes = getattr(payload, "imageChanges", None)
        if product.review_status == "pending":
            updated_payload, cleanup_urls = apply_product_image_changes(product, updated_payload, image_changes)
        product.title = first_text_from_keys(updated_payload, ("itemName", "title", "name")) or product.title
        product.price = price_from_rakuten_item(updated_payload)
        if image_changes:
            images = product_editable_image_urls(updated_payload, shop_code=product_shop_code(product, updated_payload))
            product.image_url = images[0] if images else ""
        product.raw_payload_json = json.dumps(updated_payload, ensure_ascii=False)
        product.last_error = None
        session.flush()
        result = product_detail_to_public(product)
    cleanup_product_image_urls(product_id, cleanup_urls)
    return result


def apply_product_image_changes(
    product: ProductModel,
    raw_payload: dict[str, Any],
    image_changes: Any,
) -> tuple[dict[str, Any], list[str]]:
    if not image_changes:
        return raw_payload, []
    images = unique_texts([
        normalize_product_image_url(image, shop_code=product_shop_code(product, raw_payload))
        for image in list(getattr(image_changes, "images", []) or [])
    ])
    for image_url in images:
        validate_product_image_url_ownership(product.id, image_url)
    old_images = product_images_for_edit(product)
    replacements: dict[str, str] = {}
    for item in list(getattr(image_changes, "replacements", []) or []):
        old_url = normalize_product_image_url(getattr(item, "from_", ""), shop_code=product_shop_code(product, raw_payload))
        new_url = normalize_product_image_url(getattr(item, "to", ""), shop_code=product_shop_code(product, raw_payload))
        if old_url and new_url:
            validate_product_image_url_ownership(product.id, old_url)
            validate_product_image_url_ownership(product.id, new_url)
            replacements[old_url] = new_url
    remove_urls = unique_texts([
        normalize_product_image_url(image, shop_code=product_shop_code(product, raw_payload))
        for image in list(getattr(image_changes, "removeUrls", []) or [])
    ])
    for image_url in remove_urls:
        validate_product_image_url_ownership(product.id, image_url)
    finalized_urls: dict[str, str] = {}
    cleanup_urls: list[str] = []

    def finalize_once(image_url: str) -> str:
        if image_url not in finalized_urls:
            finalized_urls[image_url] = finalize_product_image_url(
                product.id,
                image_url,
                cleanup_urls=cleanup_urls,
            )
        return finalized_urls[image_url]

    normalized_images = [finalize_once(image) for image in images]
    finalized_replacements = {
        old_url: finalize_once(new_url)
        for old_url, new_url in replacements.items()
    }
    updated_payload = set_product_image_urls_with_description_updates(
        raw_payload,
        normalized_images,
        replace_map=finalized_replacements,
        remove_urls=remove_urls,
    )
    current_images = {
        product_image_url_identity(image_url)
        for image_url in normalized_images
    }
    for old_url in old_images:
        if product_image_url_identity(old_url) not in current_images:
            cleanup_urls.append(old_url)
    return updated_payload, unique_texts(cleanup_urls)


def product_images_for_edit(product: ProductModel) -> list[str]:
    raw_payload = product_raw_payload(product)
    images = product_editable_image_urls(raw_payload, shop_code=product_shop_code(product, raw_payload))
    if product.image_url and product.image_url not in images:
        images.insert(0, product.image_url)
    return images


def product_image_download_info(owner_username: str, product_id: int, image_index: int) -> dict[str, Any]:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        images = product_images_for_edit(product)
        image_url = image_url_at_index(images, image_index)
    local_path = local_product_image_path_from_url(image_url)
    filename = product_image_download_name(product_id, image_index, image_url)
    stored_image = parse_product_image_url(image_url)
    if stored_image is not None and product_image_storage.enabled:
        image_info = product_image_http_info(image_url, include_body=True)
        image_info["filename"] = filename
        return image_info
    if local_path and local_path.exists():
        return {
            "type": "local",
            "path": local_path,
            "filename": filename,
            "mediaType": mimetypes.guess_type(str(local_path))[0] or "application/octet-stream",
        }
    return {
        "type": "remote",
        "url": image_url,
        "filename": filename,
        "mediaType": mimetypes.guess_type(filename)[0] or "application/octet-stream",
    }


def product_image_http_info(image_url: str, *, include_body: bool) -> dict[str, Any]:
    local_path = local_product_image_path_from_url(image_url)
    stored_image = parse_product_image_url(image_url)
    storage_error: Exception | None = None
    if stored_image is not None and product_image_storage.enabled:
        try:
            media_type = mimetypes.guess_type(stored_image.filename)[0] or "application/octet-stream"
            if include_body:
                stream = product_image_storage.open_stream(
                    stored_image.object_key,
                    max_bytes=MAX_PRODUCT_IMAGE_DOWNLOAD_BYTES,
                )
                if stream is not None:
                    return {
                        "type": "stream",
                        "body": stream,
                        "size": stream.size,
                        "mediaType": media_type,
                    }
            else:
                fingerprint = product_image_storage.object_fingerprint(stored_image.object_key)
                if fingerprint is not None:
                    return {
                        "type": "metadata",
                        "size": fingerprint.size,
                        "mediaType": media_type,
                    }
        except Exception as exc:
            storage_error = exc
    if local_path and local_path.exists():
        return {
            "type": "local",
            "path": local_path,
            "size": local_path.stat().st_size,
            "mediaType": mimetypes.guess_type(str(local_path))[0] or "application/octet-stream",
        }
    if storage_error is not None:
        raise RuntimeError("读取 OSS 商品图片失败。") from storage_error
    raise RuntimeError("商品图片文件不存在。")


def product_image_content_info(image_url: str) -> dict[str, Any]:
    image_data = load_product_image_bytes(
        image_url,
        max_bytes=MAX_PRODUCT_IMAGE_DOWNLOAD_BYTES,
        size_error_message="图片下载大小不能超过 20MB。",
    )
    return {
        "content": image_data["content"],
        "mediaType": image_data["contentType"],
    }


def replace_product_image(owner_username: str, product_id: int, image_index: int, upload_file: Any) -> dict[str, Any]:
    raise RuntimeError("图片替换请在待审核商品详情中操作，并点击保存后生效。")


def delete_product_image(owner_username: str, product_id: int, image_index: int) -> dict[str, Any]:
    raise RuntimeError("图片删除请在待审核商品详情中操作，并点击保存后生效。")


def image_url_at_index(images: list[str], image_index: int) -> str:
    if image_index < 0 or image_index >= len(images):
        raise RuntimeError("图片不存在。")
    return images[image_index]


def save_product_image_draft(owner_username: str, product_id: int, upload_file: Any) -> str:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        if product.review_status != "pending":
            raise RuntimeError("只有待审核商品可以修改图片。")
    return save_uploaded_product_image_file(
        upload_file,
        LOCAL_PRODUCT_IMAGE_DRAFT_DIR / str(product_id),
        lambda filename: local_product_image_draft_url(product_id, filename),
        name_prefix="draft",
    )


def save_product_image_draft_base64(owner_username: str, product_id: int, image_base64: str, ext: str = "") -> str:
    with session_scope() as session:
        product = session.get(ProductModel, product_id)
        if product is None or product.owner_username != owner_username:
            raise RuntimeError("商品不存在。")
        if product.review_status != "pending":
            raise RuntimeError("只有待审核商品可以修改图片。")
    image_data = decode_product_image_base64(image_base64, ext)
    return save_product_image_bytes(
        image_data["content"],
        image_data["suffix"],
        LOCAL_PRODUCT_IMAGE_DRAFT_DIR / str(product_id),
        lambda filename: local_product_image_draft_url(product_id, filename),
        name_prefix="meitu",
    )


def save_uploaded_product_image(product_id: int, image_index: int, upload_file: Any) -> str:
    return save_uploaded_product_image_file(
        upload_file,
        LOCAL_PRODUCT_IMAGE_DIR / str(product_id),
        lambda filename: local_product_image_url(product_id, filename),
        name_prefix=str(image_index + 1),
    )


def save_uploaded_product_image_file(upload_file: Any, image_dir: Path, url_builder: Callable[[str], str], *, name_prefix: str) -> str:
    filename = normalize_text(getattr(upload_file, "filename", ""))
    suffix = Path(filename).suffix.lower()
    if suffix == ".jpeg":
        suffix = ".jpg"
    content_type = normalize_text(getattr(upload_file, "content_type", ""))
    if suffix not in ALLOWED_PRODUCT_IMAGE_EXTENSIONS:
        raise RuntimeError("图片格式只支持 jpg、jpeg、png、gif。")
    if content_type and content_type not in ALLOWED_PRODUCT_IMAGE_MIME_TYPES:
        raise RuntimeError("图片文件类型不正确。")
    safe_name = f"{name_prefix}-{uuid.uuid4().hex[:12]}{suffix}"
    size = 0
    chunks: list[bytes] = []
    try:
        while True:
            chunk = upload_file.file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_PRODUCT_IMAGE_BYTES:
                raise RuntimeError("图片大小不能超过 2MB。")
            chunks.append(chunk)
    finally:
        try:
            upload_file.file.seek(0)
        except Exception:
            pass
    if size <= 0:
        raise RuntimeError("上传的图片为空。")
    image_url = url_builder(safe_name)
    store_product_image_content(
        image_url,
        b"".join(chunks),
        content_type or product_image_content_type_from_suffix(suffix),
        image_dir / safe_name,
    )
    return image_url


def save_product_image_bytes(content: bytes, suffix: str, image_dir: Path, url_builder: Callable[[str], str], *, name_prefix: str) -> str:
    normalized_suffix = normalize_product_image_suffix(suffix)
    if normalized_suffix not in ALLOWED_PRODUCT_IMAGE_EXTENSIONS:
        raise RuntimeError("图片格式只支持 jpg、jpeg、png、gif。")
    if not content:
        raise RuntimeError("图片内容为空。")
    if len(content) > MAX_PRODUCT_IMAGE_BYTES:
        raise RuntimeError("图片大小不能超过 2MB。")
    safe_name = f"{name_prefix}-{uuid.uuid4().hex[:12]}{normalized_suffix}"
    image_url = url_builder(safe_name)
    store_product_image_content(
        image_url,
        content,
        product_image_content_type_from_suffix(normalized_suffix),
        image_dir / safe_name,
    )
    return image_url


def store_product_image_content(image_url: str, content: bytes, content_type: str, local_path: Path) -> None:
    stored_image = parse_product_image_url(image_url)
    if product_image_storage.enabled:
        if stored_image is None:
            raise RuntimeError("商品图片存储地址无效。")
        product_image_storage.put_bytes(stored_image.object_key, content, content_type)
        return
    local_path.parent.mkdir(parents=True, exist_ok=True)
    local_path.write_bytes(content)


def decode_product_image_base64(image_base64: str, ext: str = "") -> dict[str, Any]:
    raw_value = str(image_base64 or "").strip()
    if not raw_value:
        raise RuntimeError("图片内容为空。")
    data_url_match = re.match(r"^data:(image/[a-z0-9.+-]+);base64,(.+)$", raw_value, flags=re.IGNORECASE | re.DOTALL)
    encoded_value = data_url_match.group(2) if data_url_match else raw_value
    encoded_value = re.sub(r"\s+", "", encoded_value)
    if not encoded_value:
        raise RuntimeError("图片内容为空。")
    padding = len(encoded_value) % 4
    if padding:
        encoded_value += "=" * (4 - padding)
    try:
        content = base64.b64decode(encoded_value, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise RuntimeError("图片 base64 内容无效。") from exc
    if len(content) > MAX_PRODUCT_IMAGE_BYTES:
        raise RuntimeError("图片大小不能超过 2MB。")
    detected_suffix = detect_product_image_suffix(content)
    requested_suffix = normalize_product_image_suffix(ext)
    suffix = detected_suffix or requested_suffix
    if suffix not in ALLOWED_PRODUCT_IMAGE_EXTENSIONS:
        raise RuntimeError("图片格式只支持 jpg、jpeg、png、gif。")
    return {"content": content, "suffix": suffix, "contentType": product_image_content_type_from_suffix(suffix)}


def detect_product_image_suffix(content: bytes) -> str:
    try:
        from PIL import Image, UnidentifiedImageError
        with Image.open(BytesIO(content)) as image:
            image.verify()
            image_format = normalize_text(image.format).upper()
    except (UnidentifiedImageError, OSError, SyntaxError, ValueError) as exc:
        raise RuntimeError("图片文件类型不正确。") from exc
    format_suffixes = {
        "JPEG": ".jpg",
        "PNG": ".png",
        "GIF": ".gif",
    }
    return format_suffixes.get(image_format, "")


def normalize_product_image_suffix(value: str) -> str:
    suffix = normalize_text(value).lower().strip()
    if suffix and not suffix.startswith("."):
        suffix = f".{suffix}"
    if suffix == ".jpeg":
        suffix = ".jpg"
    return suffix


def product_image_content_type_from_suffix(suffix: str) -> str:
    normalized_suffix = normalize_product_image_suffix(suffix)
    if normalized_suffix == ".png":
        return "image/png"
    if normalized_suffix == ".gif":
        return "image/gif"
    return "image/jpeg"


def local_product_image_url(product_id: int, filename: str) -> str:
    return f"{LOCAL_PRODUCT_IMAGE_URL_PREFIX}/{int(product_id)}/{quote(filename, safe='')}"


def local_product_image_draft_url(product_id: int, filename: str) -> str:
    normalized_product_id = int(product_id)
    normalized_filename = str(filename)
    expires_at = int(time.time()) + settings.product_image_draft_url_ttl_seconds
    signature = product_image_draft_signature(
        normalized_product_id,
        normalized_filename,
        expires_at,
    )
    query = urlencode({"expires": expires_at, "signature": signature})
    return (
        f"{LOCAL_PRODUCT_IMAGE_DRAFT_URL_PREFIX}/{normalized_product_id}/"
        f"{quote(normalized_filename, safe='')}?{query}"
    )


def product_image_draft_signature(
    product_id: int,
    filename: str,
    expires_at: int,
) -> str:
    payload = f"{int(product_id)}:{filename}:{int(expires_at)}".encode("utf-8")
    return hmac.new(
        settings.session_secret.encode("utf-8"),
        payload,
        hashlib.sha256,
    ).hexdigest()


def verify_product_image_draft_access(
    product_id: int,
    filename: str,
    expires_at: int,
    signature: str,
    *,
    now: int | None = None,
) -> bool:
    if int(expires_at) <= int(time.time() if now is None else now):
        return False
    expected = product_image_draft_signature(product_id, filename, expires_at)
    return hmac.compare_digest(expected, str(signature or ""))


def validate_product_image_url_ownership(product_id: int, image_url: str) -> None:
    stored_image = parse_product_image_url(image_url)
    if stored_image is not None and stored_image.product_id != int(product_id):
        raise RuntimeError("图片不属于当前商品，不能保存或删除。")


def product_image_url_identity(image_url: str) -> str:
    stored_image = parse_product_image_url(image_url)
    return stored_image.object_key if stored_image is not None else normalize_text(image_url)


def finalize_product_image_url(
    product_id: int,
    image_url: str,
    *,
    cleanup_urls: list[str] | None = None,
) -> str:
    draft_path = local_product_image_path_from_url(image_url)
    if not draft_path or not is_product_image_draft_url(image_url):
        return image_url
    draft_image = parse_product_image_url(image_url)
    validate_product_image_url_ownership(product_id, image_url)
    suffix = Path(draft_image.filename).suffix.lower() if draft_image is not None else draft_path.suffix.lower()
    target_name = f"saved-{uuid.uuid4().hex[:12]}{suffix}"
    target_url = local_product_image_url(product_id, target_name)
    target_image = parse_product_image_url(target_url)
    if (
        product_image_storage.enabled
        and draft_image is not None
        and target_image is not None
        and product_image_storage.exists(draft_image.object_key)
    ):
        product_image_storage.copy(draft_image.object_key, target_image.object_key)
        if cleanup_urls is not None:
            cleanup_urls.append(image_url)
        return target_url
    target_dir = LOCAL_PRODUCT_IMAGE_DIR / str(product_id)
    target_path = target_dir / target_name
    if not draft_path.exists():
        raise RuntimeError("商品草稿图片不存在。")
    if product_image_storage.enabled:
        store_product_image_content(
            target_url,
            draft_path.read_bytes(),
            mimetypes.guess_type(str(draft_path))[0] or product_image_content_type_from_suffix(suffix),
            target_path,
        )
    else:
        target_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(draft_path), str(target_path))
    if cleanup_urls is not None:
        cleanup_urls.append(image_url)
    return target_url


def is_product_image_draft_url(image_url: str) -> bool:
    stored_image = parse_product_image_url(image_url)
    return bool(stored_image and stored_image.kind == DRAFT_IMAGE_OBJECT_PREFIX)


def local_product_image_path_from_url(image_url: str) -> Path | None:
    stored_image = parse_product_image_url(image_url)
    if stored_image is None:
        return None
    if stored_image.kind == DRAFT_IMAGE_OBJECT_PREFIX:
        root_dir = LOCAL_PRODUCT_IMAGE_DRAFT_DIR
    elif stored_image.kind == PRODUCT_IMAGE_OBJECT_PREFIX:
        root_dir = LOCAL_PRODUCT_IMAGE_DIR
    else:
        return None
    candidate = (root_dir / str(stored_image.product_id) / stored_image.filename).resolve()
    root = root_dir.resolve()
    try:
        candidate.relative_to(root)
    except ValueError:
        return None
    return candidate


def remove_local_product_image_if_unused(
    image_url: str,
    current_images: list[str],
    *,
    expected_product_id: int | None = None,
) -> None:
    if image_url in current_images:
        return
    stored_image = parse_product_image_url(image_url)
    current_keys = {
        image.object_key
        for image in (parse_product_image_url(url) for url in current_images)
        if image is not None
    }
    if stored_image is not None and stored_image.object_key in current_keys:
        return
    if (
        stored_image is not None
        and expected_product_id is not None
        and stored_image.product_id != int(expected_product_id)
    ):
        return
    if stored_image is not None and product_image_storage.enabled:
        product_image_storage.delete(stored_image.object_key)
    path = local_product_image_path_from_url(image_url)
    if path and path.exists():
        path.unlink(missing_ok=True)


def cleanup_product_image_urls(product_id: int, image_urls: list[str]) -> None:
    for image_url in unique_texts(image_urls):
        try:
            remove_local_product_image_if_unused(
                image_url,
                [],
                expected_product_id=product_id,
            )
        except Exception:
            logger.exception(
                "Post-commit product image cleanup failed: product_id=%s image_url=%s",
                product_id,
                image_url,
            )


def clear_product_temp_image_files(product_id: int) -> None:
    if product_image_storage.enabled:
        product_image_storage.delete_prefix(f"{PRODUCT_IMAGE_OBJECT_PREFIX}/{int(product_id)}/")
        product_image_storage.delete_prefix(f"{DRAFT_IMAGE_OBJECT_PREFIX}/{int(product_id)}/")
    clear_local_product_image_files(product_id)


def clear_local_product_image_files(product_id: int) -> None:
    for root_dir in (LOCAL_PRODUCT_IMAGE_DIR, LOCAL_PRODUCT_IMAGE_DRAFT_DIR):
        image_dir = (root_dir / str(int(product_id))).resolve()
        root = root_dir.resolve()
        try:
            image_dir.relative_to(root)
        except ValueError:
            continue
        if image_dir.exists() and image_dir.is_dir():
            shutil.rmtree(image_dir, ignore_errors=True)


def cleanup_product_image_ids(product_ids: list[int]) -> None:
    for product_id in dict.fromkeys(int(value) for value in product_ids):
        try:
            clear_product_temp_image_files(product_id)
        except Exception:
            logger.exception(
                "Post-commit product image prefix cleanup failed: product_id=%s",
                product_id,
            )


def cleanup_orphan_product_image_dirs() -> int:
    product_ids, oss_objects = collect_product_image_cleanup_state()
    if not product_ids:
        return 0
    existing_ids: set[int] = set()
    with session_scope() as session:
        for offset in range(0, len(product_ids), 500):
            batch = product_ids[offset:offset + 500]
            existing_ids.update(
                int(value)
                for value in session.scalars(select(ProductModel.id).where(ProductModel.id.in_(batch))).all()
            )
    cutoff = time.time() - settings.product_image_orphan_retention_days * 24 * 60 * 60
    oss_last_modified = {
        product_id: max((item.last_modified for item in objects), default=0)
        for product_id, objects in oss_objects.items()
    }
    orphan_ids = orphan_product_ids_ready_for_cleanup(
        product_ids=product_ids,
        existing_ids=existing_ids,
        oss_last_modified=oss_last_modified,
        cutoff=cutoff,
    )
    if not orphan_ids:
        return 0
    still_existing_ids: set[int] = set()
    with session_scope() as session:
        for offset in range(0, len(orphan_ids), 500):
            batch = orphan_ids[offset:offset + 500]
            still_existing_ids.update(
                int(value)
                for value in session.scalars(select(ProductModel.id).where(ProductModel.id.in_(batch))).all()
            )
    orphan_ids = [
        product_id
        for product_id in orphan_ids
        if product_id not in still_existing_ids
    ]
    for product_id in orphan_ids:
        cleanup_aged_orphan_oss_objects(
            product_id,
            oss_objects.get(product_id, []),
            cutoff=cutoff,
        )
        clear_local_product_image_files(product_id)
    return len(orphan_ids)


def collect_product_image_dir_ids() -> list[int]:
    product_ids, _oss_objects = collect_product_image_cleanup_state()
    return product_ids


def collect_product_image_dir_state() -> tuple[list[int], dict[int, int]]:
    product_ids, oss_objects = collect_product_image_cleanup_state()
    return product_ids, {
        product_id: max((item.last_modified for item in objects), default=0)
        for product_id, objects in oss_objects.items()
    }


def collect_product_image_cleanup_state() -> tuple[list[int], dict[int, list[StoredObject]]]:
    ids: list[int] = []
    seen: set[int] = set()
    oss_objects: dict[int, list[StoredObject]] = {}

    def remember(value: Any, *, stored_object: StoredObject | None = None) -> None:
        try:
            product_id = int(value)
        except (TypeError, ValueError):
            return
        if product_id not in seen:
            seen.add(product_id)
            ids.append(product_id)
        if stored_object is not None:
            oss_objects.setdefault(product_id, []).append(stored_object)

    for root_dir in (LOCAL_PRODUCT_IMAGE_DIR, LOCAL_PRODUCT_IMAGE_DRAFT_DIR):
        if not root_dir.exists() or not root_dir.is_dir():
            continue
        for path in root_dir.iterdir():
            if not path.is_dir():
                continue
            remember(path.name)
    if product_image_storage.enabled:
        for prefix in (PRODUCT_IMAGE_OBJECT_PREFIX, DRAFT_IMAGE_OBJECT_PREFIX):
            for item in product_image_storage.list_objects(f"{prefix}/"):
                parts = item.key.split("/", 2)
                if len(parts) >= 2:
                    remember(parts[1], stored_object=item)
    return ids, oss_objects


def cleanup_aged_orphan_oss_objects(
    product_id: int,
    objects: list[StoredObject],
    *,
    cutoff: float,
) -> None:
    for item in objects:
        if item.last_modified <= 0 or item.last_modified >= cutoff:
            continue
        try:
            current = product_image_storage.object_fingerprint(item.key)
            if (
                current is None
                or current.last_modified <= 0
                or current.last_modified >= cutoff
            ):
                continue
            product_image_storage.delete(item.key)
        except Exception:
            logger.exception(
                "Orphan product image cleanup failed: product_id=%s key=%s",
                product_id,
                item.key,
            )


def orphan_product_ids_ready_for_cleanup(
    *,
    product_ids: list[int],
    existing_ids: set[int],
    oss_last_modified: dict[int, int],
    cutoff: float,
) -> list[int]:
    return [
        product_id
        for product_id in product_ids
        if product_id not in existing_ids
        and (
            product_id not in oss_last_modified
            or oss_last_modified[product_id] < cutoff
        )
    ]


def cleanup_expired_product_image_drafts() -> int:
    cutoff = time.time() - settings.product_image_draft_retention_days * 24 * 60 * 60
    deleted_count = 0
    if product_image_storage.enabled:
        for item in product_image_storage.list_objects(f"{DRAFT_IMAGE_OBJECT_PREFIX}/"):
            if item.last_modified >= cutoff:
                continue
            product_image_storage.delete(item.key)
            deleted_count += 1

    root = LOCAL_PRODUCT_IMAGE_DRAFT_DIR
    if not root.exists():
        return deleted_count
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        try:
            if path.stat().st_mtime >= cutoff:
                continue
            path.unlink(missing_ok=True)
            deleted_count += 1
        except OSError:
            continue
    for directory in sorted((item for item in root.rglob("*") if item.is_dir()), key=lambda item: len(item.parts), reverse=True):
        try:
            directory.rmdir()
        except OSError:
            continue
    return deleted_count


def product_image_download_name(product_id: int, image_index: int, image_url: str) -> str:
    try:
        suffix = Path(urlsplit(image_url).path).suffix.lower()
    except Exception:
        suffix = ""
    if suffix not in ALLOWED_PRODUCT_IMAGE_EXTENSIONS:
        suffix = ".jpg"
    return f"product-{product_id}-{image_index + 1}{suffix}"


def update_store_products_listing_status(
    owner_username: str,
    product_ids: list[int],
    listing_status: str,
) -> dict[str, Any]:
    if listing_status not in {"listed", "unlisted"}:
        raise RuntimeError("上架状态不合法。")
    normalized_ids = [int(value) for value in (product_ids or [])]
    if not normalized_ids:
        raise RuntimeError("请先选择商品。")
    return create_product_listing_status_sync_task(owner_username, normalized_ids, listing_status)


def update_store_all_products_listing_status(
    owner_username: str,
    store_id: int,
    listing_status: str,
) -> dict[str, Any]:
    return create_listing_status_sync_task(owner_username, store_id, listing_status)


def apply_products_listing_status(
    session: Any,
    products: list[ProductModel],
    listing_status: str,
    *,
    progress_callback: Callable[[int, int, int], None] | None = None,
    cancel_check: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    success_ids: list[int] = []
    failed_ids: list[int] = []
    errors: list[str] = []
    credential_cache: dict[int, tuple[str, str]] = {}
    failed_count = 0
    cancelled = False
    for index, product in enumerate(products, start=1):
        if cancel_check and cancel_check():
            cancelled = True
            break
        manage_number = normalize_text(product.rakuten_manage_number or product.item_number)
        if not product.store_id:
            errors.append(f"{product.title} 未关联店铺")
            product.last_error = "未关联店铺，不能更新乐天上架状态。"
            failed_ids.append(product.id)
            failed_count += 1
            if progress_callback:
                progress_callback(index, len(success_ids), failed_count)
            continue
        if not manage_number:
            errors.append(f"{product.title} 缺少商品管理编号")
            product.last_error = "缺少商品管理编号，不能更新乐天上架状态。"
            failed_ids.append(product.id)
            failed_count += 1
            if progress_callback:
                progress_callback(index, len(success_ids), failed_count)
            continue

        credentials = credential_cache.get(product.store_id)
        if credentials is None:
            store = session.get(StoreModel, product.store_id)
            if store is None:
                errors.append(f"{product.title} 关联店铺不存在")
                product.last_error = "关联店铺不存在，不能更新乐天上架状态。"
                failed_ids.append(product.id)
                failed_count += 1
                if progress_callback:
                    progress_callback(index, len(success_ids), failed_count)
                continue
            if not store.enabled:
                errors.append(f"{store.alias_name or store.store_name} 已停用")
                product.last_error = "关联店铺已停用，不能更新乐天上架状态。"
                failed_ids.append(product.id)
                failed_count += 1
                if progress_callback:
                    progress_callback(index, len(success_ids), failed_count)
                continue
            credentials = (
                decrypt_text(store.rakuten_service_secret_encrypted),
                decrypt_text(store.rakuten_license_key_encrypted),
            )
            credential_cache[product.store_id] = credentials

        try:
            patch_rakuten_item_listing_status(
                credentials[0],
                credentials[1],
                manage_number,
                listing_status=listing_status,
            )
        except Exception as exc:
            error_text = str(exc)
            product.last_error = error_text
            errors.append(f"{manage_number}: {error_text}")
            failed_ids.append(product.id)
            failed_count += 1
            if progress_callback:
                progress_callback(index, len(success_ids), failed_count)
            continue

        product.rakuten_listing_status = listing_status
        product.last_error = None
        product.store_product_status = "active"
        product.store_last_seen_at = datetime.now()
        success_ids.append(product.id)
        if progress_callback:
            progress_callback(index, len(success_ids), failed_count)
    return {"successIds": success_ids, "failedIds": failed_ids, "errors": errors, "cancelled": cancelled}


def sync_task_progress_callback(
    task_id: str | None,
    total_count: int,
    action_label: str,
    *,
    initial_failed: int = 0,
) -> Callable[[int, int, int], None] | None:
    if not task_id:
        return None

    def update(processed_count: int, success_count: int, failed_count: int) -> None:
        update_task_progress(
            SyncTaskModel,
            task_id,
            total_count=total_count,
            success_count=success_count,
            failed_count=initial_failed + failed_count,
            message=f"{action_label}中，已处理 {min(total_count, processed_count + initial_failed)} / {total_count} 条",
        )

    return update


def listing_status_result_summary(result: dict[str, Any], total_count: int, *, cancelled: bool = False) -> dict[str, Any]:
    success_ids = list(result.get("successIds") or [])
    failed_ids = list(result.get("failedIds") or [])
    success_count = len(success_ids)
    failed_count = len(failed_ids) if cancelled else max(0, int(total_count) - success_count)
    return {
        "total": int(total_count),
        "successCount": success_count,
        "failedCount": failed_count,
        "successIds": success_ids,
        "failedIds": failed_ids,
        "message": f"完成，成功 {success_count} 个，失败 {failed_count} 个",
        "errors": list(result.get("errors") or [])[:20],
    }


def list_listing_tasks(
    owner_username: str,
    *,
    page: int | None = None,
    page_size: int | None = None,
    task_ids: list[str] | None = None,
) -> list[dict[str, Any]] | dict[str, Any]:
    dispatch_next_listing_task_safely()
    with session_scope() as session:
        finalize_stale_cancel_requested_tasks(session, ListingTaskModel, action_label="上架", owner_username=owner_username)
        reconcile_interrupted_running_tasks(session, ListingTaskModel, owner_username=owner_username)
        query = select(ListingTaskModel).where(ListingTaskModel.owner_username == owner_username)
        normalized_task_ids = normalize_task_ids(task_ids) if task_ids else []
        if normalized_task_ids:
            query = query.where(
                ListingTaskModel.id.in_(normalized_task_ids)
            )
        normalized_page, normalized_page_size = normalize_page_params(page, page_size)
        order_by = task_status_order_by(ListingTaskModel)
        if normalized_task_ids:
            rows = session.scalars(query.order_by(*order_by)).all()
            store_snapshots = listing_task_store_snapshots(session, rows)
            return [
                listing_task_to_public(row, store_snapshots.get(row.store_id, listing_task_store_snapshot(None)))
                for row in rows
            ]
        rows = session.scalars(query).all()
        store_snapshots = listing_task_store_snapshots(session, rows)
        return paginate_task_groups(
            rows,
            page=page,
            page_size=page_size,
            response_key="listingTasks",
            serializer=lambda row: listing_task_to_public(
                row,
                store_snapshots.get(row.store_id, listing_task_store_snapshot(None)),
            ),
            group_kind="listing",
        )


def delete_listing_tasks(owner_username: str, task_ids: list[str]) -> dict[str, Any]:
    normalized_ids = normalize_task_ids(task_ids)
    with session_scope() as session:
        rows = session.scalars(
            select(ListingTaskModel).where(
                ListingTaskModel.owner_username == owner_username,
                ListingTaskModel.id.in_(normalized_ids),
            )
        ).all()
        found_ids = {row.id for row in rows}
        for row in rows:
            session.delete(row)
        deleted_ids = [row.id for row in rows]
        return {
            "deletedIds": deleted_ids,
            "failedIds": [task_id for task_id in normalized_ids if task_id not in found_ids],
            "deletedCount": len(deleted_ids),
        }


def create_listing_task(owner_username: str, payload: Any) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    product_ids = normalize_product_ids([int(value) for value in (getattr(payload, "productIds", None) or [])])
    store_ids = listing_task_payload_store_ids(payload)
    task_name = str(getattr(payload, "taskName", "") or "").strip()
    if not product_ids:
        raise RuntimeError("请选择要上架的商品。")
    if not store_ids:
        raise RuntimeError("请选择上架店铺。")
    with session_scope() as session:
        stores = session.scalars(select(StoreModel).where(StoreModel.id.in_(store_ids))).all()
        stores_by_id = {int(store.id): store for store in stores}
        if len(stores_by_id) != len(store_ids):
            raise RuntimeError("上架店铺不存在。")
        ordered_stores = [stores_by_id[store_id] for store_id in store_ids]
        for store in ordered_stores:
            if store.owner_username != owner_username:
                raise RuntimeError("不能使用其他用户的店铺上架。")
            if not store.enabled:
                raise RuntimeError(f"上架店铺「{store.alias_name or store.store_name}」已停用。")
            if not decrypt_text(store.rakuten_service_secret_encrypted) or not decrypt_text(store.rakuten_license_key_encrypted):
                raise RuntimeError(f"上架店铺「{store.alias_name or store.store_name}」缺少乐天 Secret 或乐天 Key。")
        products = session.scalars(
            select(ProductModel)
            .where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(product_ids),
            )
            .with_for_update()
        ).all()
        if not products:
            raise RuntimeError("没有找到可上架的商品。")
        found_ids = {product.id for product in products}
        missing_ids = [product_id for product_id in product_ids if product_id not in found_ids]
        if missing_ids:
            raise RuntimeError("部分商品不存在，不能创建上架任务。")
        invalid_products = [product for product in products if product.review_status not in {"approved", "listed_master"} or product.listing_task_id]
        if invalid_products:
            names = "、".join(productCodeForError(product) for product in invalid_products[:5])
            raise RuntimeError(f"只有已审核或已上架管理商品可以创建上架任务，且商品不能正在上架中。异常商品：{names}")
        duplicated_messages: list[str] = []
        for product in products:
            listed_store_ids = {int(item.get("storeId") or 0) for item in product_listed_stores(product_raw_payload(product))}
            duplicated_store_names = [
                stores_by_id[store_id].alias_name or stores_by_id[store_id].store_name
                for store_id in store_ids
                if store_id in listed_store_ids
            ]
            if duplicated_store_names:
                duplicated_messages.append(f"{productCodeForError(product)} 已上架过：{'、'.join(duplicated_store_names[:5])}")
        if duplicated_messages:
            raise RuntimeError(f"以下商品已上架过所选店铺，请调整店铺选择：{'；'.join(duplicated_messages[:5])}")
        preflight_checks = [
            listing_preflight_product_check(product, store)
            for store in ordered_stores
            for product in products
        ]
        preflight_blockers = listing_preflight_blocking_messages(preflight_checks)
        if preflight_blockers:
            detail = "；".join(preflight_blockers[:5])
            suffix = "；更多问题请先执行上架前体检。" if len(preflight_blockers) > 5 else ""
            raise RuntimeError(f"上架前体检未通过：{detail}{suffix}")
        product_by_id = {int(product.id): product for product in products}
        ordered_products = [product_by_id[product_id] for product_id in product_ids]
        ready_ids, blocked_ids = partition_product_ids_by_active_task_conflicts(
            session,
            owner_username,
            store_ids,
            product_ids,
        )
        product_chunks = conflict_aware_product_chunks(ready_ids, blocked_ids)
        task_group_id = uuid.uuid4().hex if len(product_chunks) > 1 else None
        base_task_name = task_name or f"上架任务 {datetime.now():%Y-%m-%d %H:%M}"
        task_ids: list[str] = []
        for index, (chunk_ids, blocked) in enumerate(product_chunks, start=1):
            task_id = uuid.uuid4().hex
            product_chunk = [product_by_id[product_id] for product_id in chunk_ids]
            chunk_product_ids = [int(product.id) for product in product_chunk]
            for product in product_chunk:
                product.listing_task_id = task_id
                product.last_error = None
            task = ListingTaskModel(
                id=task_id,
                owner_username=owner_username,
                store_id=ordered_stores[0].id,
                task_group_id=task_group_id,
                task_group_index=index if task_group_id else None,
                task_group_size=len(product_chunks) if task_group_id else None,
                task_name=base_task_name if len(product_chunks) == 1 else f"{base_task_name} {index}/{len(product_chunks)}",
                status="queued",
                total_count=len(product_chunk) * len(ordered_stores),
                success_count=0,
                failed_count=0,
                product_ids_json=json.dumps(
                    listing_task_result_payload(
                        chunk_product_ids,
                        [],
                        [],
                        store_ids=store_ids,
                    ),
                    ensure_ascii=False,
                ),
                message=(
                    "排队中，等待同一商品当前任务完成"
                    if blocked
                    else "等待同步到乐天"
                ),
            )
            session.add(task)
            task_ids.append(task_id)
        session.flush()

    dispatch_next_listing_task()
    with session_scope() as session:
        rows = session.scalars(
            select(ListingTaskModel)
            .where(ListingTaskModel.id.in_(task_ids))
            .order_by(ListingTaskModel.created_at.asc())
        ).all()
        task_by_id = {row.id: row for row in rows}
        tasks = [listing_task_to_public(task_by_id[task_id]) for task_id in task_ids if task_id in task_by_id]
        message = "上架任务已创建" if len(tasks) == 1 else f"上架任务已创建，已拆分为 {len(tasks)} 个任务，每个最多 {BATCH_TASK_PRODUCT_LIMIT} 条"
        return {
            "listingTask": tasks[0] if tasks else {"id": task_ids[0]},
            "listingTasks": tasks,
            "summary": {
                "total": len(product_ids),
                "taskCount": len(tasks),
                "message": message,
            },
        }


def run_listing_task(owner_username: str, task_id: str) -> None:
    try:
        _run_listing_task(owner_username, task_id)
    except TaskCancelled:
        cancel_listing_task_from_worker(owner_username, task_id)
    except Exception as exc:
        fail_listing_task_unexpectedly(owner_username, task_id, exc)
    finally:
        dispatch_next_listing_task_safely()
        dispatch_next_sync_task_safely()


def run_listing_product_attempt(
    owner_username: str,
    task_id: str,
    store_id: int,
    product_id: int,
    service_secret: str,
    license_key: str,
    cabinet_usage: dict[str, Any] | None,
) -> ListingProductAttemptResult:
    with session_scope() as session:
        store = session.get(StoreModel, store_id)
        product = session.get(ProductModel, product_id)
        if store is None or product is None or product.owner_username != owner_username:
            return ListingProductAttemptResult(
                product_id=product_id,
                success=False,
                error=f"{product_id}: 商品或店铺不存在，不能上架。",
            )
        product_label = productCodeForError(product)
        store_label = store.alias_name or store.store_name or f"店铺 {store.id}"
        if product.review_status not in {"approved", "listed_master"} or (
            product.listing_task_id not in {None, task_id}
        ):
            product.last_error = "商品状态已变化或不属于当前上架任务，不能上架。"
            clear_listing_product_lock(product, task_id)
            return ListingProductAttemptResult(
                product_id=product_id,
                success=False,
                error=f"{product_label} / {store_label}: {product.last_error}",
            )
        if any(
            int(item.get("storeId") or 0) == int(store.id)
            for item in product_listed_stores(product_raw_payload(product))
        ):
            product.last_error = None
            return ListingProductAttemptResult(product_id=product_id, success=True)
        try:
            listing_result = create_store_product_on_rakuten(
                service_secret,
                license_key,
                store,
                product,
                cabinet_context={"usage": dict(cabinet_usage or {})},
                cancel_check=lambda: listing_task_cancel_requested(task_id),
            )
            listed_product = upsert_listed_store_product_from_listing_result(
                session,
                owner_username,
                product,
                store,
                listing_result,
            )
            session.flush()
            record_product_listed_store(product, listed_product, store, listing_result)
            return ListingProductAttemptResult(product_id=product_id, success=True)
        except TaskCancelled:
            raise
        except Exception as exc:
            error_text = str(exc)
            clear_listing_product_lock(product, task_id)
            product.last_error = error_text
            return ListingProductAttemptResult(
                product_id=product_id,
                success=False,
                error=f"{product_label} / {store_label}: {error_text}",
            )


def _run_listing_task(owner_username: str, task_id: str) -> None:
    with session_scope() as session:
        task = session.get(ListingTaskModel, task_id)
        if task is None:
            return
        if task.owner_username != owner_username:
            raise RuntimeError("不能执行其他用户的上架任务。")
        if task.status == "cancelled":
            return
        if task.status != "queued":
            return
        if task_cancel_requested(task):
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        wait_reason = listing_task_start_wait_reason(session, task)
        if wait_reason:
            task.message = wait_reason
            session.flush()
            return
        task.status = "running"
        task.message = "上架准备中"
        task.error_detail = None
        task.started_at = datetime.now()
        task.finished_at = None
        product_ids_payload = listing_task_product_ids_payload(task.product_ids_json)
        task_product_ids = product_ids_payload["productIds"]
        retry_product_ids = product_ids_payload["retryIds"]
        product_ids = retry_product_ids or task_product_ids
        if not task_product_ids:
            task_product_ids = product_ids
        task_store_ids = product_ids_payload["storeIds"] or ([int(task.store_id)] if task.store_id else [])
        if not task_store_ids:
            task.status = "failed"
            task.success_count = 0
            task.failed_count = len(product_ids)
            task.total_count = len(product_ids)
            task.message = "上架店铺不存在"
            task.error_detail = task.message
            task.finished_at = datetime.now()
            release_listing_task_locks(session, owner_username, task)
            return
        retry_product_id_set = set(retry_product_ids)
        base_success_ids = (
            [product_id for product_id in product_ids_payload["successIds"] if product_id not in retry_product_id_set]
            if retry_product_ids
            else []
        )
        task.total_count = len(task_product_ids) * len(task_store_ids)
        task.success_count = len(base_success_ids)
        task.failed_count = (len(product_ids) * len(task_store_ids)) if retry_product_ids else 0
        task.message = f"上架中，已处理 0 / {len(product_ids) * len(task_store_ids)} 条"
        session.flush()
        total_count = len(task_product_ids) * len(task_store_ids)
        ordered_product_ids = [int(product_id) for product_id in product_ids]
        run_total_count = len(product_ids) * len(task_store_ids)
        all_product_ids = task_product_ids
        initial_success_ids = set(base_success_ids)

    failed_attempt_count = 0
    success_attempt_count = 0
    success_product_ids: set[int] = set(initial_success_ids)
    failed_product_ids: set[int] = set()
    errors: list[str] = []
    processed_attempts = 0

    update_task_progress(
        ListingTaskModel,
        task_id,
        total_count=total_count,
        success_count=0,
        failed_count=0,
        message=f"上架中，已处理 0 / {run_total_count} 条",
    )

    for task_store_id in task_store_ids:
        raise_if_task_cancelled(ListingTaskModel, task_id)
        with session_scope() as session:
            store = session.get(StoreModel, task_store_id)
            if store is None:
                service_secret = ""
                license_key = ""
                store_label = f"店铺 {task_store_id}"
                store_exists = False
                store_enabled = False
            else:
                service_secret = decrypt_text(store.rakuten_service_secret_encrypted)
                license_key = decrypt_text(store.rakuten_license_key_encrypted)
                store_label = store.alias_name or store.store_name or f"店铺 {store.id}"
                store_exists = True
                store_enabled = bool(store.enabled)
        store_available = store_exists and store_enabled and bool(service_secret and license_key)
        if not store_available:
            reason = "上架店铺不存在"
            if store_exists and not store_enabled:
                reason = "上架店铺已停用"
            elif store_exists and not (service_secret and license_key):
                reason = "上架店铺缺少乐天 Secret 或乐天 Key"
            for product_id in ordered_product_ids:
                processed_attempts += 1
                failed_attempt_count += 1
                failed_product_ids.add(product_id)
                errors.append(f"{product_id} / {store_label}: {reason}")
                update_task_progress(
                    ListingTaskModel,
                    task_id,
                    total_count=total_count,
                    success_count=max(success_attempt_count, len(success_product_ids)),
                    failed_count=failed_attempt_count,
                    message=f"上架中，已处理 {processed_attempts} / {run_total_count} 条",
                )
            continue

        cabinet_usage: dict[str, Any] = {}
        try:
            raise_if_task_cancelled(ListingTaskModel, task_id)
            cabinet_usage = fetch_rakuten_cabinet_usage(service_secret, license_key)
            raise_if_task_cancelled(ListingTaskModel, task_id)
            with session_scope() as session:
                current_store = session.get(StoreModel, task_store_id)
                if current_store is not None:
                    apply_store_cabinet_usage(current_store, cabinet_usage)
        except TaskCancelled:
            raise
        except Exception as exc:
            errors.append(f"{store_label}: R-Cabinet 使用量检测失败: {exc}")
            with session_scope() as session:
                task = session.get(ListingTaskModel, task_id)
                if task is not None:
                    next_error_detail = summarize_task_errors(errors, limit=50)
                    task.error_detail = with_task_cancel_marker(next_error_detail) if task_cancel_requested(task) else next_error_detail

        worker_count = listing_task_product_worker_count(
            len(ordered_product_ids),
            retry=bool(retry_product_ids),
        )
        cancellation_error: TaskCancelled | None = None
        with ThreadPoolExecutor(
            max_workers=worker_count,
            thread_name_prefix=f"listing-products-{task_id[:8]}",
        ) as executor:
            futures = {
                executor.submit(
                    run_listing_product_attempt,
                    owner_username,
                    task_id,
                    task_store_id,
                    product_id,
                    service_secret,
                    license_key,
                    cabinet_usage,
                ): product_id
                for product_id in ordered_product_ids
            }
            for future in as_completed(futures):
                product_id = futures[future]
                processed_attempts += 1
                try:
                    attempt = future.result()
                except TaskCancelled as exc:
                    cancellation_error = exc
                    for pending_future in futures:
                        pending_future.cancel()
                    break
                except Exception as exc:
                    attempt = ListingProductAttemptResult(
                        product_id=product_id,
                        success=False,
                        error=f"{product_id} / {store_label}: {exc}",
                    )
                if attempt.success:
                    success_attempt_count += 1
                    success_product_ids.add(attempt.product_id)
                    failed_product_ids.discard(attempt.product_id)
                else:
                    failed_attempt_count += 1
                    failed_product_ids.add(attempt.product_id)
                    if attempt.error:
                        errors.append(attempt.error)
                cancel_after_progress = False
                with session_scope() as session:
                    task = session.get(ListingTaskModel, task_id)
                    if task is not None:
                        final_failed_product_ids = [
                            failed_product_id
                            for failed_product_id in failed_product_ids
                            if failed_product_id not in success_product_ids
                        ]
                        cancel_requested = task_cancel_requested(task) or listing_task_cancel_requested(task_id)
                        task.total_count = total_count
                        task.success_count = max(success_attempt_count, len(success_product_ids))
                        task.failed_count = failed_attempt_count
                        task.message = (
                            TASK_CANCEL_REQUESTED_MESSAGE
                            if cancel_requested
                            else f"上架中，已处理 {processed_attempts} / {run_total_count} 条"
                        )
                        next_error_detail = summarize_task_errors(errors, limit=50)
                        task.error_detail = (
                            with_task_cancel_marker(next_error_detail)
                            if cancel_requested
                            else next_error_detail
                        )
                        task.product_ids_json = json.dumps(
                            listing_task_result_payload(
                                all_product_ids,
                                list(success_product_ids),
                                final_failed_product_ids,
                                retry_ids=retry_product_ids or None,
                                store_ids=task_store_ids,
                            ),
                            ensure_ascii=False,
                        )
                        cancel_after_progress = cancel_requested
                if cancel_after_progress:
                    cancellation_error = TaskCancelled(TASK_CANCELLED_MESSAGE)
                    for pending_future in futures:
                        pending_future.cancel()
                    break
        if cancellation_error is not None:
            raise cancellation_error
    raise_if_task_cancelled(ListingTaskModel, task_id)
    with session_scope() as session:
        task = session.get(ListingTaskModel, task_id)
        if task is None:
            return
        if task_cancel_requested(task):
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        for product in session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(all_product_ids or [-1]),
            )
        ).all():
            clear_listing_product_lock(product, task_id)
        if task_cancel_requested(task) or listing_task_cancel_requested(task_id):
            raise TaskCancelled(TASK_CANCELLED_MESSAGE)
        final_success_ids = list(success_product_ids)
        final_failed_ids = [product_id for product_id in failed_product_ids if product_id not in success_product_ids]
        final_success_count = max(success_attempt_count, len(success_product_ids))
        task.total_count = total_count
        task.success_count = final_success_count
        task.failed_count = failed_attempt_count
        if final_success_count and failed_attempt_count:
            task.status = "partial"
        elif final_success_count:
            task.status = "success"
        else:
            task.status = "failed"
        task.message = f"完成，上架 {final_success_count} 条，异常 {failed_attempt_count} 条"
        task.error_detail = "\n".join(errors[:50]) if errors else None
        task.product_ids_json = json.dumps(
            listing_task_result_payload(all_product_ids, final_success_ids, final_failed_ids, store_ids=task_store_ids),
            ensure_ascii=False,
        )
        task.finished_at = datetime.now()


def clear_listing_product_lock(product: ProductModel, task_id: str | None = None) -> None:
    if task_id is None or product.listing_task_id == task_id:
        product.listing_task_id = None


def cancel_listing_task_from_worker(owner_username: str, task_id: str) -> None:
    with session_scope() as session:
        task = session.get(ListingTaskModel, task_id)
        if task is None or task.owner_username != owner_username:
            return
        release_listing_task_locks(session, owner_username, task)
        total_count = int(task.total_count or 0)
        success_count = int(task.success_count or 0)
        failed_count = int(task.failed_count or 0)
        task.status = "cancelled"
        task.message = cancelled_task_progress_message("上架", total_count, success_count, failed_count)
        task.error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
        task.finished_at = datetime.now()
    log_event(owner_username, task_id, "info", TASK_CANCELLED_MESSAGE)


def fail_listing_task_unexpectedly(owner_username: str, task_id: str, exc: Exception) -> None:
    message = str(exc) or "上架任务执行失败。"
    with session_scope() as session:
        task = session.get(ListingTaskModel, task_id)
        if task is not None and task.owner_username == owner_username:
            if task.status == "cancelled":
                return
            product_ids_payload = listing_task_product_ids_payload(task.product_ids_json)
            task_product_ids = product_ids_payload["productIds"]
            task_store_ids = product_ids_payload["storeIds"] or ([int(task.store_id)] if task.store_id else [])
            failed_ids = product_ids_payload["retryIds"] or task_product_ids
            if not task_product_ids:
                task_product_ids = failed_ids
            failed_id_set = set(failed_ids)
            success_ids = [product_id for product_id in product_ids_payload["successIds"] if product_id not in failed_id_set]
            products = session.scalars(
                select(ProductModel).where(
                    ProductModel.owner_username == owner_username,
                    ProductModel.id.in_(failed_ids or [-1]),
                )
            ).all()
            for product in products:
                clear_listing_product_lock(product, task_id)
                product.last_error = message
            task.status = "partial" if success_ids and failed_ids else "failed"
            task.success_count = len(success_ids)
            task.failed_count = len(failed_ids) * max(1, len(task_store_ids))
            task.total_count = len(task_product_ids) * max(1, len(task_store_ids))
            task.message = f"完成，上架 {len(success_ids)} 条，异常 {len(failed_ids)} 条" if success_ids else "上架失败"
            task.error_detail = message
            task.product_ids_json = json.dumps(
                listing_task_result_payload(task_product_ids, success_ids, failed_ids, store_ids=task_store_ids),
                ensure_ascii=False,
            )
            task.finished_at = datetime.now()
    log_event(owner_username, task_id, "error", message)


def retry_listing_task(owner_username: str, task_id: str) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    with session_scope() as session:
        task = session.get(ListingTaskModel, task_id)
        if task is None:
            raise RuntimeError("上架任务不存在。")
        if task.owner_username != owner_username:
            raise RuntimeError("不能重试其他用户的上架任务。")
        if task.status in {"queued", "running"}:
            raise RuntimeError("上架任务正在执行中，不能重试。")
        product_ids_payload = listing_task_product_ids_payload(task.product_ids_json)
        retry_product_ids = listing_task_retry_product_ids(task)
        if not retry_product_ids:
            raise RuntimeError("没有可重试的商品。")
        task_product_ids = product_ids_payload["productIds"] or retry_product_ids
        task_store_ids = product_ids_payload["storeIds"] or ([int(task.store_id)] if task.store_id else [])
        retry_product_id_set = set(retry_product_ids)
        base_success_ids = (
            [product_id for product_id in product_ids_payload["successIds"] if product_id not in retry_product_id_set]
            if task.status in {"partial", "failed"} and product_ids_payload["failedIds"]
            else []
        )
        products = session.scalars(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.id.in_(retry_product_ids or [-1]),
            )
        ).all()
        for product in products:
            if product.review_status in {"approved", "listed_master"} and not product.listing_task_id:
                product.listing_task_id = task_id
                product.last_error = None
        task.status = "queued"
        task.total_count = len(task_product_ids) * max(1, len(task_store_ids))
        task.success_count = len(base_success_ids)
        task.failed_count = len(retry_product_ids) * max(1, len(task_store_ids))
        task.message = "等待重新上架"
        task.error_detail = None
        task.product_ids_json = json.dumps(
            listing_task_result_payload(
                task_product_ids,
                base_success_ids,
                retry_product_ids,
                retry_ids=retry_product_ids,
                store_ids=task_store_ids,
            ),
            ensure_ascii=False,
        )
        task.started_at = None
        task.finished_at = None
    dispatch_next_listing_task()
    with session_scope() as session:
        task = session.get(ListingTaskModel, task_id)
        return listing_task_to_public(task) if task else {"id": task_id}


def ensure_default_roles() -> None:
    defaults = [
        {
            "name": "超级管理员",
            "code": "superadmin",
            "scope": "all",
            "permissions": ["users.manage", "roles.manage", "crawler.manage", "products.manage", "stores.manage", "ai.manage"],
            "notes": "系统内置角色，拥有全部管理权限。",
        },
        {
            "name": "运营用户",
            "code": "operator",
            "scope": "own",
            "permissions": ["secrets.manage", "crawler.manage", "products.manage", "stores.manage"],
            "notes": "默认业务角色，可使用公司共享店铺，处理自己的采集任务和商品。",
        },
    ]
    with session_scope() as session:
        for item in defaults:
            row = session.scalar(select(RoleModel).where(RoleModel.code == item["code"]))
            if row is None:
                row = RoleModel(code=item["code"])
                session.add(row)
            row.name = item["name"]
            row.scope = item["scope"]
            row.enabled = True
            row.permissions_json = json.dumps(item["permissions"], ensure_ascii=False)
            row.notes = item["notes"]


def list_roles(*, page: int | None = None, page_size: int | None = None) -> list[dict[str, Any]] | dict[str, Any]:
    ensure_default_roles()
    with session_scope() as session:
        query = select(RoleModel)
        return paginate_query(
            session,
            query,
            order_by=RoleModel.id.asc(),
            page=page,
            page_size=page_size,
            response_key="roles",
            serializer=role_to_public,
        )


def save_role(payload: Any, role_id: int | None = None) -> dict[str, Any]:
    ensure_default_roles()
    with session_scope() as session:
        row = session.get(RoleModel, role_id) if role_id else None
        if row is None:
            row = RoleModel()
            session.add(row)
        code = str(getattr(payload, "code", "") or "").strip()
        if not code:
            raise RuntimeError("角色编码不能为空。")
        if row.code in {"superadmin", "operator"} and code != row.code:
            raise RuntimeError("内置角色编码不能修改。")
        row.name = str(getattr(payload, "name", "") or "").strip()
        row.code = code
        row.scope = str(getattr(payload, "scope", "") or "own").strip()
        row.enabled = bool(getattr(payload, "enabled", True))
        row.permissions_json = json.dumps(getattr(payload, "permissions", None) or [], ensure_ascii=False)
        row.notes = str(getattr(payload, "notes", "") or "").strip()
        if not row.name:
            raise RuntimeError("角色名称不能为空。")
        session.flush()
        return role_to_public(row)


def delete_role(role_id: int) -> None:
    ensure_default_roles()
    with session_scope() as session:
        row = session.get(RoleModel, role_id)
        if row is None:
            return
        if row.code in {"superadmin", "operator"}:
            raise RuntimeError("内置角色不能删除。")
        session.delete(row)


def create_task(owner_username: str, payload: Any) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    source_id = getattr(payload, "sourceId", None)
    scheduled_crawl_id = getattr(payload, "scheduledCrawlId", None)
    source_type = str(getattr(payload, "sourceType", "") or "").strip()
    target = str(getattr(payload, "target", "") or "").strip()
    raw_crawl_price_rule = getattr(payload, "crawlPriceRule", None)
    crawl_price_rule = (
        normalize_crawl_price_rule(raw_crawl_price_rule)
        if raw_crawl_price_rule is not None
        else None
    )
    with session_scope() as session:
        schedule = session.get(ScheduledCrawlModel, scheduled_crawl_id) if scheduled_crawl_id else None
        if scheduled_crawl_id and (
            schedule is None
            or schedule.owner_username != owner_username
            or schedule.source_type != "shop"
        ):
            raise RuntimeError("关联的采集店铺不存在或无权访问。")
        source = session.get(CrawlSourceModel, source_id) if source_id else None
        if source is not None:
            if source.owner_username != owner_username:
                raise RuntimeError("不能使用其他用户的采集源。")
            source_type = source.source_type
            target = source.target
        if not source_type or not target:
            raise RuntimeError("采集类型和目标不能为空。")
        if source_type == "product_url":
            target = "\n".join(normalize_rakuten_product_targets(target))
            crawl_price_rule = {"operator": "all"}
        elif source_type == "whole_shop":
            sid, _ = resolve_whole_shop_identity(target)
            whole_shop_filter = normalize_whole_shop_filter(getattr(payload, "wholeShopFilter", None))
            target = build_whole_shop_task_target(
                sid,
                whole_shop_filter,
                getattr(payload, "crawlLimit", None),
            )
        elif source_type == "shop":
            primary_target, fallback_target = split_shop_fallback_target(target)
            parsed_target, existing_limit, existing_period = parse_ranking_target(strip_shop_ranking_prefix(primary_target))
            raw_target = normalize_rakuten_shop_target(parsed_target)
            if not raw_target:
                raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR)
            period_label = ranking_period_label(getattr(payload, "rankingPeriod", None) or existing_period)
            limit_label = crawl_limit_label(
                getattr(payload, "crawlLimit", None),
                default="全部" if existing_limit is None else f"前 {existing_limit}",
            )
            target = f"店铺:{raw_target} {period_label} {limit_label}"
            target = append_shop_fallback_target(target, fallback_target)
        task = CrawlTaskModel(
            id=uuid.uuid4().hex,
            owner_username=owner_username,
            source_id=source.id if source else None,
            scheduled_crawl_id=int(schedule.id) if schedule is not None else None,
            source_type=source_type,
            target=target,
            crawl_price_rule_json=(
                json.dumps(crawl_price_rule, ensure_ascii=False)
                if crawl_price_rule is not None
                else None
            ),
            mode=str(getattr(payload, "mode", "") or "manual"),
            status="queued",
            total_count=initial_crawl_task_total_count(source_type, target),
            message="等待执行",
        )
        session.add(task)
        session.flush()
        task_public = task_to_public(task)

    if should_use_redis_task_queue():
        dispatch_queued_crawl_tasks_safely(owner_username)
    else:
        dispatch_crawl_task(task_public["id"])
    return task_public


def run_existing_task(owner_username: str, task_id: str) -> dict[str, Any]:
    ensure_system_task_dispatch_allowed(owner_username)
    with session_scope() as session:
        task = session.get(CrawlTaskModel, task_id)
        if task is None:
            raise RuntimeError("采集任务不存在。")
        if task.owner_username != owner_username:
            raise RuntimeError("不能重启其他用户的采集任务。")
        if task.source_type == "product_replace":
            raise RuntimeError("商品替换采集请从店铺商品中重新发起。")
        if task.status in {"queued", "running"}:
            raise RuntimeError("采集任务正在执行中，不能重新采集。")
        task.status = "queued"
        task.queue_job_id = None
        task.total_count = initial_crawl_task_total_count(task.source_type, task.target)
        task.success_count = 0
        task.failed_count = 0
        task.warning_count = 0
        task.saved_count = 0
        task.skipped_count = 0
        task.message = "等待重新执行"
        task.error_detail = None
        task.warning_detail = None
        task.started_at = None
        task.finished_at = None
    if should_use_redis_task_queue():
        dispatch_queued_crawl_tasks_safely(owner_username)
    else:
        dispatch_crawl_task(task_id)
    with session_scope() as session:
        task = session.get(CrawlTaskModel, task_id)
        return task_to_public(task) if task else {"id": task_id}


def collected_item_error(item: dict[str, Any]) -> str | None:
    raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
    detail_error = str(raw.get("detailError") or "").strip()
    if detail_error:
        name = str(item.get("title") or item.get("source_url") or "商品").strip()
        return f"{name}: {detail_error}"
    if raw.get("detailCollected") is False:
        name = str(item.get("title") or item.get("source_url") or "商品").strip()
        return f"{name}: 商品详情采集失败。"
    return None


def summarize_task_errors(errors: list[str], limit: int = 20, *, item_label: str = "错误") -> str | None:
    unique_errors: list[str] = []
    seen: set[str] = set()
    for error in errors:
        normalized = str(error or "").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique_errors.append(normalized)
    if not unique_errors:
        return None
    visible_errors = unique_errors[:limit]
    if len(unique_errors) > limit:
        visible_errors.append(f"另有 {len(unique_errors) - limit} 条{item_label}未显示。")
    return "\n".join(visible_errors)


def run_task(task_id: str, reserved_job_id: str | None = None) -> None:
    owner_username = ""
    source_type = ""
    target = ""
    scheduled_crawl_id: int | None = None
    should_run = False
    should_refill = False
    should_retry_thread_task = False
    success_count = 0
    failed_count = 0
    warning_count = 0
    saved_count = 0
    skipped_count = 0
    total_count = 0
    errors: list[str] = []
    warnings: list[str] = []

    def current_error_detail() -> str | None:
        return summarize_task_errors(errors, limit=50)

    def current_warning_detail() -> str | None:
        return summarize_task_errors(warnings, limit=50, item_label="警告")

    with session_scope() as session:
        task = session.get(CrawlTaskModel, task_id)
        if task is None:
            return
        owner_username = task.owner_username
        should_refill = should_use_redis_task_queue()
        reservation_matches = task.queue_job_id == reserved_job_id
        if should_refill:
            reservation_matches = bool(reserved_job_id) and reservation_matches
        if not reservation_matches:
            pass
        else:
            task.queue_job_id = None
            if task.status == "cancelled":
                pass
            elif task.status != "queued":
                pass
            elif task_cancel_requested(task):
                task.status = "cancelled"
                task.message = TASK_CANCELLED_MESSAGE
                task.error_detail = cancelled_task_error_detail(existing_error_detail=task.error_detail)
                task.warning_detail = cancelled_task_warning_detail(existing_warning_detail=task.warning_detail)
                task.finished_at = datetime.now()
            else:
                wait_reason = crawl_task_start_wait_reason(session, task)
                if wait_reason:
                    task.message = wait_reason
                    should_retry_thread_task = not should_refill
                else:
                    task.status = "running"
                    task.started_at = datetime.now()
                    task.message = "采集中"
                    task.error_detail = None
                    task.warning_detail = None
                    if task.total_count <= 0:
                        task.total_count = initial_crawl_task_total_count(task.source_type, task.target)
                    source_type = task.source_type
                    target = task.target
                    scheduled_crawl_id = task.scheduled_crawl_id
                    should_run = True

    if not should_run:
        if should_refill:
            dispatch_queued_crawl_tasks_safely(owner_username)
        elif should_retry_thread_task:
            dispatch_crawl_task(task_id, delay_seconds=TASK_START_RETRY_DELAY_SECONDS)
        return

    try:
        raise_if_task_cancelled(CrawlTaskModel, task_id)
        item_plan = collect_item_plan(source_type, target, task_id=task_id)
        total_count = item_plan.total_count
        raise_if_task_cancelled(CrawlTaskModel, task_id)
        update_task_progress(
            CrawlTaskModel,
            task_id,
            total_count=total_count,
            success_count=0,
            failed_count=0,
            warning_count=0,
            saved_count=0,
            skipped_count=0,
            message=f"采集中，正在采集详情并入库 0 / {total_count} 条",
            error_detail=current_error_detail(),
            warning_detail=current_warning_detail(),
        )
        processed_count = 0
        active_words = []
        collection_genre_policy = CollectionGenrePolicySnapshot()
        if total_count > 0:
            with session_scope() as session:
                active_words = active_sensitive_words(session)
                collection_genre_policy = collection_genre_policy_snapshot(session, owner_username)
        for item in item_plan.items:
            raise_if_task_cancelled(CrawlTaskModel, task_id)
            processed_count += 1
            item_error = collected_item_error(item)
            save_result = save_collected_item_with_lock_retry(
                owner_username,
                task_id,
                item,
                active_words=active_words,
                collection_genre_policy=collection_genre_policy,
                scheduled_crawl_id=scheduled_crawl_id,
            )
            saved = bool(save_result.get("saved"))
            skipped = bool(save_result.get("skipped"))
            save_error = normalize_text(save_result.get("error"))
            if saved:
                saved_count += 1
                success_count += 1
                item_warnings = [message for message in (item_error, save_error) if message]
                if item_warnings:
                    warning_count += 1
                    warnings.extend(item_warnings)
            elif skipped:
                skipped_count += 1
                success_count += 1
                if save_error:
                    warning_count += 1
                    warnings.append(save_error)
            else:
                failed_count += 1
                if item_error:
                    errors.append(item_error)
                elif save_error:
                    errors.append(save_error)
                else:
                    name = str(item.get("title") or item.get("source_url") or "商品").strip()
                    errors.append(f"{name}: 商品未保存，可能缺少商品标题、商品链接，或已存在于店铺商品中。")
            update_task_progress(
                CrawlTaskModel,
                task_id,
                total_count=total_count,
                success_count=success_count,
                failed_count=failed_count,
                warning_count=warning_count,
                saved_count=saved_count,
                skipped_count=skipped_count,
                message=(
                    f"采集中，正在采集详情并入库 {processed_count} / {total_count} 条，"
                    f"成功 {success_count} 条，失败 {failed_count} 条，"
                    f"入库 {saved_count} 条，跳过 {skipped_count} 条"
                ),
                error_detail=current_error_detail(),
                warning_detail=current_warning_detail(),
            )
        with session_scope() as session:
            task = session.get(CrawlTaskModel, task_id)
            if task is None:
                return
            task.success_count = success_count
            task.failed_count = failed_count
            task.warning_count = warning_count
            task.saved_count = saved_count
            task.skipped_count = skipped_count
            task.total_count = total_count
            task.status = resolve_crawl_task_status("success", total_count, success_count, failed_count)
            task.finished_at = datetime.now()
            task.message = f"完成，采集 {total_count} 条，成功 {success_count} 条，失败 {failed_count} 条，警告 {warning_count} 条，入库 {saved_count} 条，跳过 {skipped_count} 条"
            task.error_detail = current_error_detail()
            task.warning_detail = current_warning_detail()
        log_event(owner_username, task_id, "info", f"任务完成，成功 {success_count} 条，失败 {failed_count} 条，警告 {warning_count} 条，入库 {saved_count} 条，跳过 {skipped_count} 条商品")
    except TaskCancelled:
        with session_scope() as session:
            task = session.get(CrawlTaskModel, task_id)
            if task is None:
                return
            if total_count > 0:
                task.total_count = total_count
            task.success_count = success_count
            task.failed_count = failed_count
            task.warning_count = warning_count
            task.saved_count = saved_count
            task.skipped_count = skipped_count
            task.status = "cancelled"
            task.finished_at = datetime.now()
            task.message = f"已终止，采集已处理 {success_count + failed_count} / {task.total_count or total_count} 条，成功 {success_count} 条，失败 {failed_count} 条，警告 {warning_count} 条，入库 {saved_count} 条，跳过 {skipped_count} 条"
            task.error_detail = cancelled_task_error_detail(errors, task.error_detail)
            task.warning_detail = cancelled_task_warning_detail(warnings, task.warning_detail)
        log_event(owner_username, task_id, "info", TASK_CANCELLED_MESSAGE)
    except Exception as exc:
        error_text = str(exc)
        errors.append(error_text)
        with session_scope() as session:
            task = session.get(CrawlTaskModel, task_id)
            if task is None:
                return
            if task.status == "cancelled":
                return
            task.status = "failed"
            task.failed_count = max(1, failed_count)
            task.warning_count = warning_count
            task.saved_count = saved_count
            task.skipped_count = skipped_count
            task.finished_at = datetime.now()
            task.message = "采集失败"
            task.error_detail = current_error_detail()
            task.warning_detail = current_warning_detail()
        log_event(owner_username, task_id, "error", error_text)
    finally:
        if should_refill:
            dispatch_queued_crawl_tasks_safely(owner_username)


def save_collected_item_with_lock_retry(
    owner_username: str,
    task_id: str,
    item: dict[str, Any],
    *,
    active_words: list[str] | None = None,
    collection_genre_policy: CollectionGenrePolicySnapshot | None = None,
    scheduled_crawl_id: int | None = None,
) -> dict[str, Any]:
    attempts = len(CRAWL_SAVE_LOCK_RETRY_DELAYS_SECONDS) + 1
    for attempt in range(attempts):
        try:
            return save_collected_item(
                owner_username,
                task_id,
                item,
                active_words=active_words,
                collection_genre_policy=collection_genre_policy,
                scheduled_crawl_id=scheduled_crawl_id,
            )
        except OperationalError as exc:
            if not is_mysql_retryable_lock_error(exc):
                raise
            if attempt >= attempts - 1:
                display_name = normalize_text(item.get("title") or item.get("source_url") or "商品")
                return {
                    "saved": False,
                    "skipped": False,
                    "error": f"{display_name}: 商品入库遇到数据库锁冲突，重试 {attempts} 次后仍未成功。",
                }
            raise_if_task_cancelled(CrawlTaskModel, task_id)
            delay_seconds = CRAWL_SAVE_LOCK_RETRY_DELAYS_SECONDS[attempt]
            logger.warning(
                "crawl item save lock conflict; retrying task_id=%s attempt=%s/%s delay=%ss",
                task_id,
                attempt + 1,
                attempts,
                delay_seconds,
            )
            time.sleep(delay_seconds)
    raise RuntimeError("商品入库重试流程异常。")


def save_collected_item(
    owner_username: str,
    task_id: str,
    item: dict[str, Any],
    *,
    active_words: list[str] | None = None,
    collection_genre_policy: CollectionGenrePolicySnapshot | None = None,
    scheduled_crawl_id: int | None = None,
) -> dict[str, Any]:
    product_id: int | None = None
    if item.get("_crawlPriceFiltered"):
        price = item.get("price")
        return {
            "saved": False,
            "skipped": True,
            "error": normalize_text(item.get("_crawlPriceFilterReason"))
            or f"商品价格 {price} 日元不符合用户设置的采集价格条件，已跳过。",
        }
    genre_decision = collected_item_genre_policy_decision(
        item,
        collection_genre_policy or CollectionGenrePolicySnapshot(),
    )
    if not genre_decision["allowed"]:
        display_name = normalize_text(item.get("title") or item.get("source_url") or "商品")
        genre_label = normalize_text(
            genre_decision.get("genrePathZh")
            or genre_decision.get("genrePath")
            or genre_decision.get("genreId")
            or "未识别品类"
        )
        genre_id = normalize_text(genre_decision.get("genreId"))
        if genre_id and genre_id not in genre_label:
            genre_label = f"{genre_label}（{genre_id}）"
        return {
            "saved": False,
            "skipped": True,
            "error": f"{display_name}: 品类“{genre_label}”已被当前用户设置为禁止采集，本次未入库。",
        }
    with session_scope() as session:
        duplicated_product = find_existing_collected_product(session, owner_username, item)
        if duplicated_product is not None:
            display_name = normalize_text(item.get("title") or duplicated_product.title or duplicated_product.source_url)
            status_label = {
                "pending": "待审核商品",
                "approved": "已审核商品",
                "error": "异常商品",
                "listed_master": "已上架商品",
                "rejected": "已拒绝商品",
            }.get(duplicated_product.review_status, "商品管理")
            return {
                "saved": False,
                "skipped": True,
                "error": (
                    f"{display_name}: 已存在于{status_label}，本次未重复入库。"
                ),
            }
        prepared_item = prepare_product_upsert_item(session, item, active_words=active_words)
        saved = upsert_product(
            session,
            owner_username,
            task_id,
            item,
            active_words=active_words,
            prepared_item=prepared_item,
            scheduled_crawl_id=scheduled_crawl_id,
        )
        if saved:
            session.flush()
            source_url = prepared_item.source_url
            if source_url:
                product = session.scalar(
                    select(ProductModel).where(
                        ProductModel.owner_username == owner_username,
                        ProductModel.source_url_hash == make_source_url_hash(source_url),
                    )
                )
                product_id = product.id if product is not None else None
        else:
            return {"saved": False, "error": prepared_item.error}
    image_error = ""
    if product_id is not None:
        try:
            image_error = localize_collected_product_images(owner_username, product_id)
        except Exception as exc:
            image_error = f"图片本地化失败：{exc}"
            mark_product_local_image_error(owner_username, product_id, image_error)
        try:
            prepare_collected_product_for_listing(owner_username, product_id)
        except Exception:
            logger.warning(
                "collected product listing preparation failed owner=%s product_id=%s",
                owner_username,
                product_id,
                exc_info=True,
            )
    return {"saved": True, "error": image_error}


def find_existing_collected_product(session: Any, owner_username: str, item: dict[str, Any]) -> ProductModel | None:
    source_url = str(item.get("source_url") or "").strip()
    if not source_url:
        return None
    source_url_hash_key = str(item.get("source_url_hash_key") or source_url).strip()
    source_url_hash = make_source_url_hash(source_url_hash_key)
    return session.scalar(
        select(ProductModel).where(
            ProductModel.owner_username == owner_username,
            ProductModel.source_url_hash == source_url_hash,
            ProductModel.store_id.is_(None),
        )
    )


def chunk_items(items: list[dict[str, Any]], batch_size: int) -> list[list[dict[str, Any]]]:
    normalized_size = max(1, int(batch_size or 1))
    return [items[index:index + normalized_size] for index in range(0, len(items), normalized_size)]


def initial_crawl_task_total_count(source_type: str, target: str) -> int:
    normalized_source_type = normalize_text(source_type)
    normalized_target = normalize_text(split_shop_fallback_target(target)[0] if normalized_source_type == "shop" else target)
    if not normalized_source_type or not normalized_target:
        return 0
    if normalized_source_type == "product_url":
        return len(normalize_rakuten_product_targets(target))
    if normalized_source_type == "whole_shop":
        _, _, limit = parse_whole_shop_task_options(target)
        return int(limit or 0)
    if normalized_source_type in {"shop", "ranking"}:
        _, limit, _ = parse_ranking_target(
            strip_shop_ranking_prefix(normalized_target) if normalized_source_type == "shop" else normalized_target
        )
        return int(limit or 0)
    return 0


def collect_items(source_type: str, target: str, *, task_id: str | None = None) -> list[dict[str, Any]]:
    plan = collect_item_plan(source_type, target, task_id=task_id)
    return list(plan.items)


def collect_item_plan(source_type: str, target: str, *, task_id: str | None = None) -> CollectedItemPlan:
    if source_type == "shop":
        primary_target, fallback_target = split_shop_fallback_target(target)
        strict_target = fallback_shop_target(primary_target, fallback_target) if fallback_target else primary_target
        return collect_item_plan_for_target(source_type, strict_target, task_id=task_id)
    return collect_item_plan_for_target(source_type, target, task_id=task_id)


def collect_items_for_target(source_type: str, target: str, *, task_id: str | None = None) -> list[dict[str, Any]]:
    plan = collect_item_plan_for_target(source_type, target, task_id=task_id)
    return list(plan.items)


def collect_item_plan_for_target(
    source_type: str,
    target: str,
    *,
    task_id: str | None = None,
) -> CollectedItemPlan:
    raise_if_task_cancelled(CrawlTaskModel, task_id)
    if source_type == "product_url":
        product_urls = normalize_rakuten_product_targets(target)
        items = [
            {
                "title": product_url,
                "source_url": product_url,
                "raw": {},
            }
            for product_url in product_urls
        ]
        return CollectedItemPlan(
            total_count=len(items),
            items=iter_enriched_collected_items_with_detail(
                items,
                task_id=task_id,
                price_rule={"operator": "all"},
            ),
        )
    price_rule = crawl_price_rule_for_task(task_id)
    if source_type == "whole_shop":
        sid, whole_shop_filter, limit = parse_whole_shop_task_options(target)
        items = collect_whole_shop_listing_items(
            sid,
            whole_shop_filter,
            price_rule,
            limit,
            task_id=task_id,
        )
        for item in items:
            mark_collected_item_price_filter(item, price_rule)
        if task_id:
            update_task_progress(
                CrawlTaskModel,
                task_id,
                total_count=len(items),
                message=f"已发现 {len(items)} 个整店商品，开始采集详情",
            )
        existing_source_hashes = existing_collected_source_hashes_for_task(items, task_id)
        return CollectedItemPlan(
            total_count=len(items),
            items=iter_enriched_collected_items_with_detail(
                items,
                task_id=task_id,
                existing_source_hashes=existing_source_hashes,
                price_rule=price_rule,
            ),
        )
    limit: int | None = 30
    shop_code_filter = ""
    if source_type == "shop":
        target, limit, period = parse_ranking_target(strip_shop_ranking_prefix(target))
        normalized_shop_target = normalize_rakuten_shop_target(target)
        if re.fullmatch(r"[0-9]+", normalized_shop_target):
            shop_code_filter, display_name = fetch_rakuten_shop_identity_by_sid(normalized_shop_target)
            if not shop_code_filter:
                if task_id:
                    update_task_progress(
                        CrawlTaskModel,
                        task_id,
                        total_count=0,
                        message="未能确认该店铺的店铺代码，本次榜单采集数量为 0",
                    )
                return CollectedItemPlan(total_count=0, items=iter(()))
            target = display_name or shop_code_filter
        elif looks_like_rakuten_shop_code(normalized_shop_target):
            shop_code_filter = normalize_shop_code(normalized_shop_target)
            target = resolve_rakuten_shop_search_keyword(normalized_shop_target)
        else:
            raise RuntimeError(RAKUTEN_SHOP_TARGET_ERROR)
    elif source_type == "ranking":
        target, limit, period = parse_ranking_target(target)
    else:
        period = "daily"
    if source_type == "ranking":
        url = build_ranking_source_url(target, period)
    elif source_type == "shop" and period == "realtime":
        url = build_ranking_source_url(target, period)
    elif source_type == "shop":
        url = build_ranking_source_url(target, period)
    else:
        url = build_source_url(source_type, target)
    listing_limit = None if shop_code_filter else limit
    items = collect_listing_items(url, listing_limit, task_id=task_id)
    raise_if_task_cancelled(CrawlTaskModel, task_id)
    if source_type in {"ranking", "shop"} and period == "realtime":
        keyword = normalize_text(target).lower()
        items = [item for item in items if keyword in normalize_text(item.get("title")).lower()]
    if source_type == "shop" and shop_code_filter:
        items = [item for item in items if product_url_shop_code(item.get("source_url")) == shop_code_filter]
    limited_items = items if limit is None else items[:limit]
    for item in limited_items:
        mark_collected_item_price_filter(item, price_rule)
    if task_id:
        update_task_progress(
            CrawlTaskModel,
            task_id,
            total_count=len(limited_items),
            message=f"已发现 {len(limited_items)} 个商品，开始采集详情",
        )
    existing_source_hashes = existing_collected_source_hashes_for_task(limited_items, task_id)
    return CollectedItemPlan(
        total_count=len(limited_items),
        items=iter_enriched_collected_items_with_detail(
            limited_items,
            task_id=task_id,
            existing_source_hashes=existing_source_hashes,
            price_rule=price_rule,
        ),
    )


def collect_listing_items(
    url: str,
    requested_limit: int | None,
    *,
    task_id: str | None = None,
    progress_label: str = "",
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    ranking_total: int | None = None
    page_number = 1
    while page_number <= settings.crawler_max_ranking_pages:
        raise_if_task_cancelled(CrawlTaskModel, task_id)
        page_url = ranking_page_url(url, page_number)
        html = fetch_listing_html(page_url)
        raise_if_task_cancelled(CrawlTaskModel, task_id)
        assert_listing_page_available(html, page_url)
        if ranking_total is None:
            ranking_total = parse_ranking_total_count(html)
        page_items = parse_search_items(html, page_url)
        new_count = 0
        for item in page_items:
            source_url = normalize_text(item.get("source_url"))
            if not source_url or source_url in seen:
                continue
            seen.add(source_url)
            items.append(item)
            new_count += 1
            if requested_limit is not None and len(items) >= requested_limit:
                return items
            if ranking_total is not None and len(items) >= ranking_total:
                return items
        if task_id and progress_label:
            progress_total = listing_progress_total_count(
                ranking_total=ranking_total,
                requested_limit=requested_limit,
                collected_count=len(items),
            )
            update_task_progress(
                CrawlTaskModel,
                task_id,
                total_count=progress_total,
                message=f"正在发现{progress_label}：{len(items)} / {progress_total}",
            )
        if not should_fetch_next_ranking_page(
            page_items=page_items,
            new_count=new_count,
            collected_count=len(items),
            requested_limit=requested_limit,
            ranking_total=ranking_total,
        ):
            break
        page_number += 1
    return items


def listing_progress_total_count(
    *,
    ranking_total: int | None,
    requested_limit: int | None,
    collected_count: int,
) -> int:
    if requested_limit is not None:
        if ranking_total is None:
            return requested_limit
        return min(requested_limit, ranking_total)
    if ranking_total is not None:
        return ranking_total
    return collected_count


def should_fetch_next_ranking_page(
    *,
    page_items: list[dict[str, Any]],
    new_count: int,
    collected_count: int,
    requested_limit: int | None,
    ranking_total: int | None,
) -> bool:
    if not page_items or new_count <= 0:
        return False
    if requested_limit is not None and collected_count >= requested_limit:
        return False
    if ranking_total is not None:
        return collected_count < ranking_total
    return requested_limit is None or collected_count < requested_limit


def parse_ranking_total_count(html: str) -> int | None:
    text = normalize_text(BeautifulSoup(html or "", "lxml").get_text(" ", strip=True))
    patterns = (
        r"検索結果\s*[0-9,，]+\s*[〜~\-－]\s*[0-9,，]+\s*件\s*[（(]\s*([0-9,，]+)\s*件\s*[）)]",
        r"(?:共|全)\s*([0-9,，]+)\s*(?:个|件)",
        r"\(\s*(?:共|全)\s*([0-9,，]+)\s*(?:个|件)\s*\)",
        r"([0-9,，]+)\s*(?:個|件)\s*(?:中|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        number = int(match.group(1).replace(",", "").replace("，", ""))
        if number > 0:
            return number
    state_match = re.search(r'"pagination"\s*:\s*\{[^{}]*"numFound"\s*:\s*([0-9]+)', html or "")
    if state_match:
        number = int(state_match.group(1))
        if number >= 0:
            return number
    return None


def normalize_whole_shop_filter(value: Any) -> str:
    normalized = normalize_text(value).lower()
    if normalized in {"", "all", "全店采集", "全部"}:
        return "all"
    if normalized in {"reviewed", "评论采集", "有评论"}:
        return "reviewed"
    raise RuntimeError("整店采集过滤方式不正确。")


def whole_shop_filter_label(value: Any) -> str:
    return "评论采集" if normalize_whole_shop_filter(value) == "reviewed" else "全店采集"


def parse_rakuten_shop_sid(html: str) -> str:
    soup = BeautifulSoup(html or "", "lxml")
    for link in soup.select("a[href*='search.rakuten.co.jp/search/mall']"):
        try:
            sid = normalize_text((parse_qs(urlsplit(str(link.get("href") or "")).query).get("sid") or [""])[0])
        except Exception:
            sid = ""
        if re.fullmatch(r"[0-9]+", sid):
            return sid
    patterns = (
        r'"shopId"\s*:\s*([0-9]+)',
        r"(?:[?&]sid=)([0-9]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, html or "")
        if match:
            return match.group(1)
    return ""


def resolve_whole_shop_identity(target: str) -> tuple[str, str]:
    try:
        normalized_target = normalize_rakuten_shop_target(target)
    except RuntimeError as exc:
        raise RuntimeError(WHOLE_SHOP_TARGET_ERROR) from exc
    if re.fullmatch(r"[0-9]+", normalized_target):
        return normalized_target, ""
    if not looks_like_rakuten_shop_code(normalized_target):
        raise RuntimeError(WHOLE_SHOP_TARGET_ERROR)
    try:
        store_html = fetch_html(build_rakuten_store_url(normalized_target))
    except requests.RequestException as exc:
        raise RuntimeError(f"无法访问该乐天店铺：{exc}") from exc
    sid = parse_rakuten_shop_sid(store_html)
    if not sid:
        raise RuntimeError("未能从店铺页面识别 sid，无法进行整店采集。")
    return sid, parse_rakuten_shop_display_name(store_html)


def build_whole_shop_task_target(
    sid: str,
    whole_shop_filter: str,
    crawl_limit: Any = "all",
) -> str:
    normalized_sid = normalize_text(sid)
    if not re.fullmatch(r"[0-9]+", normalized_sid):
        raise RuntimeError(WHOLE_SHOP_TARGET_ERROR)
    return (
        f"整店:{normalized_sid} {whole_shop_filter_label(whole_shop_filter)} "
        f"{crawl_limit_label(crawl_limit)}"
    )


def parse_whole_shop_task_target(target: str) -> tuple[str, str]:
    sid, whole_shop_filter, _ = parse_whole_shop_task_options(target)
    return sid, whole_shop_filter


def parse_whole_shop_task_options(target: str) -> tuple[str, str, int | None]:
    normalized = normalize_text(target)
    match = re.fullmatch(
        r"整店[:：]\s*([0-9]+)\s+(全店采集|评论采集)(?:\s+(全部|全量|前\s*[0-9]{1,5}))?",
        normalized,
    )
    if not match:
        raise RuntimeError("整店采集任务参数不正确。")
    limit_label = normalize_text(match.group(3))
    limit_match = re.fullmatch(r"前\s*([0-9]{1,5})", limit_label)
    limit = max(1, int(limit_match.group(1))) if limit_match else None
    return match.group(1), normalize_whole_shop_filter(match.group(2)), limit


def whole_shop_price_bounds(price_rule: dict[str, Any] | None) -> tuple[int | None, int | None]:
    normalized_rule = normalize_crawl_price_rule(price_rule or {"operator": "all"})
    operator = normalize_text(normalized_rule.get("operator")).lower()
    if operator == "all":
        return None, None
    if operator == "range":
        return (
            int(normalized_rule.get("minPrice") or 0),
            int(normalized_rule.get("maxPrice") or 0),
        )
    value = int(normalized_rule.get("value") or 0)
    if operator == "gt":
        return value + 1, None
    if operator == "gte":
        return value, None
    if operator == "lt":
        return None, max(0, value - 1)
    if operator == "lte":
        return None, value
    return None, None


def build_whole_shop_search_url(
    sid: str,
    whole_shop_filter: str,
    *,
    price_min: int | None = None,
    price_max: int | None = None,
) -> str:
    normalized_sid = normalize_text(sid)
    if not re.fullmatch(r"[0-9]+", normalized_sid):
        raise RuntimeError(WHOLE_SHOP_TARGET_ERROR)
    query: dict[str, str] = {"sid": normalized_sid}
    if normalize_whole_shop_filter(whole_shop_filter) == "reviewed":
        query["review"] = "1"
    if price_min is not None and int(price_min) > 0:
        query["min"] = str(int(price_min))
    if price_max is not None and int(price_max) >= 0:
        query["max"] = str(int(price_max))
    return f"{RAKUTEN_SEARCH_BASE}?{urlencode(query)}"


def whole_shop_preview_page(url: str) -> tuple[str, int, int]:
    html = fetch_listing_html(url)
    assert_listing_page_available(html, url)
    page_items = parse_search_items(html, url)
    total_count = parse_ranking_total_count(html)
    if total_count is None:
        raise RuntimeError("未能从乐天搜索页读取商品总数，请稍后重试。")
    if total_count > 0 and not page_items:
        raise RuntimeError("乐天搜索页显示存在商品，但未能解析商品列表。")
    page_size = max(1, len(page_items) or 45)
    return html, total_count, math.ceil(total_count / page_size)


def whole_shop_search_count(url: str) -> tuple[int, int]:
    _, total_count, page_count = whole_shop_preview_page(url)
    return total_count, page_count


def split_bounded_whole_shop_price_range(
    sid: str,
    whole_shop_filter: str,
    price_min: int,
    price_max: int,
) -> list[tuple[int, int | None, int]]:
    url = build_whole_shop_search_url(
        sid,
        whole_shop_filter,
        price_min=price_min,
        price_max=price_max,
    )
    total_count, _ = whole_shop_search_count(url)
    if total_count <= WHOLE_SHOP_PARTITION_TARGET_ITEMS:
        return [(price_min, price_max, total_count)]
    if price_min >= price_max:
        raise RuntimeError(
            f"价格 {price_min} 日元的商品超过 {WHOLE_SHOP_PARTITION_TARGET_ITEMS} 件，"
            "乐天搜索无法完整分页，请缩小采集范围。"
        )
    midpoint = (price_min + price_max) // 2
    return [
        *split_bounded_whole_shop_price_range(
            sid,
            whole_shop_filter,
            price_min,
            midpoint,
        ),
        *split_bounded_whole_shop_price_range(
            sid,
            whole_shop_filter,
            midpoint + 1,
            price_max,
        ),
    ]


def whole_shop_price_partitions(
    sid: str,
    whole_shop_filter: str,
    price_rule: dict[str, Any] | None,
) -> list[tuple[int | None, int | None, int]]:
    price_min, price_max = whole_shop_price_bounds(price_rule)
    normalized_min = max(0, int(price_min or 0))
    if price_max is not None:
        return split_bounded_whole_shop_price_range(
            sid,
            whole_shop_filter,
            normalized_min,
            int(price_max),
        )

    partitions: list[tuple[int | None, int | None, int]] = []
    current_min = normalized_min
    width = max(1000, current_min or 1000)
    for _ in range(WHOLE_SHOP_PARTITION_MAX_STEPS):
        tail_url = build_whole_shop_search_url(
            sid,
            whole_shop_filter,
            price_min=current_min,
        )
        tail_total, _ = whole_shop_search_count(tail_url)
        if tail_total <= WHOLE_SHOP_PARTITION_TARGET_ITEMS:
            partitions.append((current_min or None, None, tail_total))
            return partitions
        current_max = current_min + width - 1
        partitions.extend(
            split_bounded_whole_shop_price_range(
                sid,
                whole_shop_filter,
                current_min,
                current_max,
            )
        )
        current_min = current_max + 1
        width *= 2
    raise RuntimeError("整店商品价格范围过大，无法完成自动分段，请缩小采集价格范围。")


def collect_whole_shop_listing_items(
    sid: str,
    whole_shop_filter: str,
    price_rule: dict[str, Any] | None,
    requested_limit: int | None,
    *,
    task_id: str | None = None,
) -> list[dict[str, Any]]:
    price_min, price_max = whole_shop_price_bounds(price_rule)
    direct_url = build_whole_shop_search_url(
        sid,
        whole_shop_filter,
        price_min=price_min,
        price_max=price_max,
    )
    if requested_limit is not None and requested_limit <= WHOLE_SHOP_PARTITION_TARGET_ITEMS:
        return collect_listing_items(
            direct_url,
            requested_limit,
            task_id=task_id,
            progress_label="整店商品",
        )

    total_count, _ = whole_shop_search_count(direct_url)
    if total_count <= WHOLE_SHOP_PARTITION_TARGET_ITEMS:
        return collect_listing_items(
            direct_url,
            requested_limit,
            task_id=task_id,
            progress_label="整店商品",
        )

    partitions = whole_shop_price_partitions(sid, whole_shop_filter, price_rule)
    expected_total = total_count
    if requested_limit is not None:
        expected_total = min(expected_total, requested_limit)
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for partition_min, partition_max, _ in partitions:
        remaining = None if requested_limit is None else max(0, requested_limit - len(items))
        if remaining == 0:
            break
        partition_url = build_whole_shop_search_url(
            sid,
            whole_shop_filter,
            price_min=partition_min,
            price_max=partition_max,
        )
        partition_items = collect_listing_items(
            partition_url,
            remaining,
            task_id=task_id,
        )
        for item in partition_items:
            source_url = normalize_text(item.get("source_url"))
            if not source_url or source_url in seen:
                continue
            seen.add(source_url)
            items.append(item)
            if requested_limit is not None and len(items) >= requested_limit:
                break
        if task_id:
            update_task_progress(
                CrawlTaskModel,
                task_id,
                total_count=expected_total,
                message=f"正在发现整店商品：{len(items)} / {expected_total}",
            )
    return items


def preview_whole_shop_crawl(
    owner_username: str,
    target: str,
    whole_shop_filter: str,
    crawl_price_rule: dict[str, Any] | None = None,
) -> dict[str, Any]:
    del owner_username
    normalized_filter = normalize_whole_shop_filter(whole_shop_filter)
    normalized_price_rule = normalize_crawl_price_rule(crawl_price_rule or {"operator": "all"})
    price_min, price_max = whole_shop_price_bounds(normalized_price_rule)
    sid, resolved_shop_name = resolve_whole_shop_identity(target)
    all_url = build_whole_shop_search_url(sid, "all")
    all_html, total_found, all_page_count = whole_shop_preview_page(all_url)
    shop_name = resolved_shop_name or parse_rakuten_search_shop_name(all_html) or sid
    selected_url = build_whole_shop_search_url(
        sid,
        normalized_filter,
        price_min=price_min,
        price_max=price_max,
    )
    if selected_url == all_url:
        collectable_count = total_found
        page_count = all_page_count
    else:
        _, collectable_count, page_count = whole_shop_preview_page(selected_url)
    reviewed_count = 0
    if normalized_filter == "reviewed":
        reviewed_count = collectable_count
    return {
        "valid": True,
        "shopName": shop_name,
        "sid": sid,
        "filter": normalized_filter,
        "totalFound": total_found,
        "collectableCount": collectable_count,
        "reviewedCount": reviewed_count,
        "pageCount": page_count,
        "message": f"本次预计采集 {collectable_count} 个商品",
    }


def ranking_page_url(url: str, page_number: int) -> str:
    if page_number <= 1:
        return url
    parsed = urlsplit(url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    query["p"] = [str(page_number)]
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urlencode(query, doseq=True), parsed.fragment))


def product_url_shop_code(source_url: Any) -> str:
    parsed = parse_rakuten_product_target(normalize_text(source_url))
    return normalize_shop_code(parsed[0]) if parsed else ""


def existing_collected_source_hashes_for_task(
    items: list[dict[str, Any]],
    task_id: str | None,
) -> set[str]:
    if not task_id or not items:
        return set()
    source_hashes = {
        make_source_url_hash(
            normalize_text(item.get("source_url_hash_key") or item.get("source_url"))
        )
        for item in items
        if normalize_text(item.get("source_url_hash_key") or item.get("source_url"))
    }
    if not source_hashes:
        return set()
    with session_scope() as session:
        task = session.get(CrawlTaskModel, task_id)
        if task is None:
            return set()
        existing_hashes: set[str] = set()
        for hash_batch in chunk_items(list(source_hashes), 1000):
            existing_hashes.update(
                session.scalars(
                    select(ProductModel.source_url_hash).where(
                        ProductModel.owner_username == task.owner_username,
                        ProductModel.store_id.is_(None),
                        ProductModel.source_url_hash.in_(hash_batch),
                    )
                ).all()
            )
        return existing_hashes


def crawl_price_rule_for_task(task_id: str | None) -> dict[str, Any]:
    if not task_id:
        return {"operator": "all"}
    with session_scope() as session:
        task = session.get(CrawlTaskModel, task_id)
        if task is None:
            return {"operator": "all"}
        if task.source_type == "product_url":
            return {"operator": "all"}
        task_rule = task_stored_crawl_price_rule(task)
        if task_rule is not None:
            return task_rule
        account = session.get(UserAccountModel, task.owner_username)
        return account_crawl_price_rule(account) if account is not None else {"operator": "all"}


def collected_item_price_filter_reason(item: dict[str, Any], price_rule: dict[str, Any]) -> str:
    operator = normalize_text(price_rule.get("operator")).lower()
    if operator == "all":
        return ""
    raw_price = item.get("price")
    if raw_price is None or normalize_text(raw_price) == "":
        return ""
    try:
        price = float(raw_price)
    except (TypeError, ValueError):
        return ""
    value = float(price_rule.get("value") or 0)
    min_price = float(price_rule.get("minPrice") or 0)
    max_price = float(price_rule.get("maxPrice") or 0)
    allowed = {
        "gt": price > value,
        "gte": price >= value,
        "lt": price < value,
        "lte": price <= value,
        "range": min_price <= price <= max_price,
    }.get(operator, True)
    if allowed:
        return ""
    condition_text = {
        "gt": f"大于 {int(value)} 日元",
        "gte": f"大于等于 {int(value)} 日元",
        "lt": f"小于 {int(value)} 日元",
        "lte": f"小于等于 {int(value)} 日元",
        "range": f"在 {int(min_price)} 至 {int(max_price)} 日元之间（含上下限）",
    }.get(operator, "符合采集价格条件")
    return f"商品价格 {price:g} 日元不符合“{condition_text}”的采集价格条件，已跳过。"


def should_filter_collected_item_by_price(item: dict[str, Any], price_rule: dict[str, Any]) -> bool:
    return bool(collected_item_price_filter_reason(item, price_rule))


def mark_collected_item_price_filter(item: dict[str, Any], price_rule: dict[str, Any]) -> None:
    reason = collected_item_price_filter_reason(item, price_rule)
    if reason:
        item["_crawlPriceFiltered"] = True
        item["_crawlPriceFilterReason"] = reason


def enrich_collected_item_with_detail(
    item: dict[str, Any],
    *,
    task_id: str | None = None,
    existing_source_hashes: set[str] | None = None,
    price_rule: dict[str, Any] | None = None,
) -> dict[str, Any]:
    known_hashes = existing_source_hashes or set()
    active_price_rule = price_rule or {"operator": "all"}
    raise_if_task_cancelled(CrawlTaskModel, task_id)
    source_url = normalize_text(item.get("source_url"))
    if not source_url or item.get("_crawlPriceFiltered"):
        return item
    source_hash_key = normalize_text(item.get("source_url_hash_key") or source_url)
    if make_source_url_hash(source_hash_key) in known_hashes:
        return item
    try:
        detail = collect_product_detail(source_url)
        raise_if_task_cancelled(CrawlTaskModel, task_id)
    except Exception as exc:
        if isinstance(exc, TaskCancelled):
            raise
        fallback = dict(item)
        raw = fallback.get("raw") if isinstance(fallback.get("raw"), dict) else {}
        fallback["raw"] = {**raw, "detailError": str(exc), "detailCollected": False}
        return fallback
    raw = detail.get("raw") if isinstance(detail.get("raw"), dict) else {}
    list_raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
    merged_raw = {**raw, "listPage": list_raw.get("pageUrl"), "detailCollected": True}
    if "reviewCount" in list_raw:
        merged_raw["reviewCount"] = list_raw.get("reviewCount")
    detail["raw"] = merged_raw
    if not detail.get("price"):
        detail["price"] = item.get("price")
    if not detail.get("image_url"):
        detail["image_url"] = item.get("image_url")
    mark_collected_item_price_filter(detail, active_price_rule)
    return detail


def iter_enriched_collected_items_with_detail(
    items: Iterable[dict[str, Any]],
    *,
    task_id: str | None = None,
    existing_source_hashes: set[str] | None = None,
    price_rule: dict[str, Any] | None = None,
) -> Iterator[dict[str, Any]]:
    item_iterator = iter(items)
    worker_count = max(1, int(settings.crawler_detail_workers or 1))
    if worker_count == 1:
        for item in item_iterator:
            yield enrich_collected_item_with_detail(
                item,
                task_id=task_id,
                existing_source_hashes=existing_source_hashes,
                price_rule=price_rule,
            )
        return

    executor = ThreadPoolExecutor(
        max_workers=worker_count,
        thread_name_prefix="crawl-detail",
    )
    pending: set[Any] = set()

    def submit_next() -> bool:
        try:
            item = next(item_iterator)
        except StopIteration:
            return False
        pending.add(
            executor.submit(
                enrich_collected_item_with_detail,
                item,
                task_id=task_id,
                existing_source_hashes=existing_source_hashes,
                price_rule=price_rule,
            )
        )
        return True

    try:
        for _ in range(worker_count):
            if not submit_next():
                break
        while pending:
            raise_if_task_cancelled(CrawlTaskModel, task_id)
            future = next(as_completed(tuple(pending)))
            pending.remove(future)
            yield future.result()
            submit_next()
    finally:
        for future in pending:
            future.cancel()
        executor.shutdown(wait=False, cancel_futures=True)


def enrich_collected_items_with_detail(
    items: Iterable[dict[str, Any]],
    *,
    task_id: str | None = None,
    existing_source_hashes: set[str] | None = None,
    price_rule: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    return list(
        iter_enriched_collected_items_with_detail(
            items,
            task_id=task_id,
            existing_source_hashes=existing_source_hashes,
            price_rule=price_rule,
        )
    )


def strip_shop_ranking_prefix(target: str) -> str:
    normalized = normalize_text(target)
    if normalized.startswith("店铺:") or normalized.startswith("店铺："):
        return normalized.split(":", 1)[1] if ":" in normalized else normalized.split("：", 1)[1]
    return normalized


def parse_ranking_target(target: str) -> tuple[str, int | None, str]:
    normalized = normalize_text(target)
    limit: int | None = 30
    all_match = re.search(r"(?:^|\s)(全部|全量)\s*$", normalized)
    if all_match:
        limit = None
        normalized = normalized[: all_match.start()].strip()
    match = re.search(r"(?:^|\s)前\s*([0-9]{1,5})\s*$", normalized)
    if match:
        limit = int(match.group(1))
        normalized = normalized[: match.start()].strip()
    period = "daily"
    period_match = re.search(r"(?:^|\s)(实时|实时榜|日榜|每日|每日榜|周榜|週間|週間榜|月榜|月間|月間榜)\s*$", normalized)
    if period_match:
        period_label = period_match.group(1)
        normalized = normalized[: period_match.start()].strip()
        if period_label in {"实时", "实时榜"}:
            period = "realtime"
        elif period_label in {"周榜", "週間", "週間榜"}:
            period = "weekly"
        elif period_label in {"月榜", "月間", "月間榜"}:
            period = "monthly"
        else:
            period = "daily"
    return normalized, None if limit is None else max(1, limit), period


def build_ranking_source_url(keyword: str, period: str) -> str:
    normalized_keyword = normalize_text(keyword)
    if period == "realtime":
        return RAKUTEN_REALTIME_RANKING_URL
    if period == "monthly":
        ptn = "3"
    elif period == "weekly":
        ptn = "2"
    else:
        ptn = "1"
    return f"{RAKUTEN_RANKING_BASE}?stx={quote(normalized_keyword)}&srt=1&ptn={ptn}"


def build_source_url(source_type: str, target: str) -> str:
    target = target.strip()
    if target.startswith("http://") or target.startswith("https://"):
        return target
    if source_type == "shop":
        if target.isdigit():
            return f"{RAKUTEN_SEARCH_BASE}?sid={quote(target)}"
        return f"{RAKUTEN_SEARCH_BASE}{quote(target)}/"
    if source_type == "ranking":
        return f"{RAKUTEN_RANKING_BASE}?stx={quote(target)}&srt=1"
    return f"{RAKUTEN_SEARCH_BASE}{quote(target)}/"


def crawler_browser_headers(url: str = "") -> dict[str, str]:
    parsed = urlsplit(url) if url else None
    origin = f"{parsed.scheme}://{parsed.netloc}" if parsed and parsed.scheme and parsed.netloc else "https://www.rakuten.co.jp"
    return {
        "User-Agent": settings.crawler_user_agent,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "ja-JP,ja;q=0.9,zh-CN;q=0.8,zh;q=0.7,en-US;q=0.6,en;q=0.5",
        "Accept-Encoding": "gzip, deflate",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Connection": "keep-alive",
        "DNT": "1",
        "Referer": origin + "/",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
    }


def crawler_request_proxies() -> dict[str, str] | None:
    proxy_url = normalize_text(settings.crawler_proxy_url)
    if not proxy_url:
        return None
    return {"http": proxy_url, "https": proxy_url}


def crawler_delay_seconds() -> float:
    min_ms = max(0, int(settings.crawler_min_delay_ms or 0))
    max_ms = max(min_ms, int(settings.crawler_max_delay_ms or min_ms))
    if max_ms <= 0:
        return 0.0
    return random.uniform(min_ms / 1000, max_ms / 1000)


def throttle_crawler_request() -> None:
    global CRAWLER_LAST_REQUEST_AT
    delay = crawler_delay_seconds()
    if delay <= 0:
        return
    with CRAWLER_REQUEST_LOCK:
        elapsed = time.monotonic() - CRAWLER_LAST_REQUEST_AT
        wait_seconds = max(0.0, delay - elapsed)
        if wait_seconds:
            time.sleep(wait_seconds)
        CRAWLER_LAST_REQUEST_AT = time.monotonic()


def crawler_backoff_seconds(attempt: int) -> float:
    return min(12.0, (1.5 ** max(0, attempt - 1)) + random.uniform(0.2, 1.2))


def get_crawler_session() -> requests.Session:
    session = getattr(CRAWLER_SESSION_LOCAL, "session", None)
    if isinstance(session, requests.Session):
        return session
    session = requests.Session()
    session.headers.update(crawler_browser_headers())
    CRAWLER_SESSION_LOCAL.session = session
    CRAWLER_SESSION_LOCAL.warmed = False
    return session


def warmup_crawler_session(session: requests.Session) -> None:
    if getattr(CRAWLER_SESSION_LOCAL, "warmed", False):
        return
    CRAWLER_SESSION_LOCAL.warmed = True
    warmup_url = normalize_text(settings.crawler_warmup_url)
    if not warmup_url:
        return
    try:
        throttle_crawler_request()
        session.get(
            warmup_url,
            timeout=settings.crawler_timeout_seconds,
            headers=crawler_browser_headers(warmup_url),
            proxies=crawler_request_proxies(),
        )
    except requests.RequestException:
        return


def fetch_html(url: str) -> str:
    session = get_crawler_session()
    warmup_crawler_session(session)
    max_attempts = max(1, int(settings.crawler_max_retries or 0) + 1)
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            throttle_crawler_request()
            response = session.get(
                url,
                timeout=settings.crawler_timeout_seconds,
                headers=crawler_browser_headers(url),
                proxies=crawler_request_proxies(),
            )
            if response.status_code in CRAWLER_HTTP_RETRY_STATUS_CODES and attempt < max_attempts:
                time.sleep(crawler_backoff_seconds(attempt))
                continue
            response.raise_for_status()
            response.encoding = response.encoding or response.apparent_encoding
            html = response.text
            if is_rakuten_access_limited_page(html) and attempt < max_attempts:
                time.sleep(crawler_backoff_seconds(attempt))
                continue
            return html
        except requests.RequestException as exc:
            last_exc = exc
            if attempt >= max_attempts:
                raise
            time.sleep(crawler_backoff_seconds(attempt))
    if last_exc:
        raise last_exc
    raise RuntimeError("乐天页面采集失败。")


def fetch_listing_html(url: str) -> str:
    try:
        html = fetch_html(url)
    except requests.RequestException as exc:
        if not settings.crawler_browser_fallback_enabled:
            raise
        try:
            return fetch_html_with_browser(url)
        except Exception as browser_exc:
            raise RuntimeError(f"乐天列表页采集失败：{exc}；浏览器兜底采集失败：{browser_exc}") from browser_exc
    if should_retry_listing_with_browser(html):
        try:
            return fetch_html_with_browser(url)
        except Exception as browser_exc:
            if is_rakuten_access_limited_page(html):
                raise RuntimeError(f"乐天列表页返回访问集中/拦截页，浏览器兜底采集失败：{browser_exc}") from browser_exc
            return html
    return html


def should_retry_listing_with_browser(html: str) -> bool:
    if not settings.crawler_browser_fallback_enabled:
        return False
    if is_blocked_or_empty_rakuten_html(html):
        return True
    if "item.rakuten.co.jp" not in (html or "") and "brandavenue.rakuten.co.jp/item/" not in (html or ""):
        return True
    if len(html or "") < 2000:
        text = normalize_text(BeautifulSoup(html or "", "lxml").get_text(" ", strip=True))
        return "Reference #" in text or "楽天" not in text
    return False


def assert_listing_page_available(html: str, url: str) -> None:
    text = normalize_text(BeautifulSoup(html or "", "lxml").get_text(" ", strip=True))
    if is_rakuten_access_limited_page(html):
        raise RuntimeError(f"乐天排行页当前返回访问集中/拦截页，无法采集：{url}")
    if not text:
        raise RuntimeError(f"乐天列表页为空，无法采集：{url}")


def is_rakuten_access_limited_page(html: str) -> bool:
    text = normalize_text(BeautifulSoup(html or "", "lxml").get_text(" ", strip=True))
    return "アクセスが集中しております" in text or re.search(r"Reference\s+#", text) is not None


def parse_search_items(html: str, page_url: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "lxml")
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    review_counts = parse_search_item_review_counts(soup, page_url)
    for item in parse_search_items_from_json_ld(soup, page_url):
        source_url = normalize_text(item.get("source_url"))
        if not source_url or source_url in seen:
            continue
        raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
        item["raw"] = {**raw, "reviewCount": review_counts.get(source_url, 0)}
        seen.add(source_url)
        items.append(item)
    for link in soup.select("a[href*='item.rakuten.co.jp'], a[href*='brandavenue.rakuten.co.jp/item/']"):
        href = normalize_product_href(str(link.get("href") or ""), page_url)
        title = " ".join(link.get_text(" ", strip=True).split())
        if not href or href in seen:
            continue
        seen.add(href)
        container = link.find_parent(["div", "li", "article"]) or link
        image = ""
        image_node = container.select_one("img")
        if image_node:
            image = str(image_node.get("src") or image_node.get("data-src") or "")
        if not title:
            title = normalize_text(image_node.get("alt") if image_node else "")
        price = extract_price(container.get_text(" ", strip=True))
        items.append(
            {
                "title": (title or href)[:500],
                "source_url": href,
                "image_url": image,
                "price": price,
                "shop_name": "",
                "item_number": extract_item_number(href),
                "genre_id": "",
                "raw": {
                    "pageUrl": page_url,
                    "reviewCount": review_counts.get(href, 0),
                },
            }
        )
    return items


def parse_search_item_review_counts(soup: BeautifulSoup, page_url: str) -> dict[str, int]:
    review_counts: dict[str, int] = {}
    for container in soup.select(".searchresultitem"):
        product_link = container.select_one(
            "a[href*='item.rakuten.co.jp'], a[href*='brandavenue.rakuten.co.jp/item/']"
        )
        if product_link is None:
            continue
        source_url = normalize_product_href(str(product_link.get("href") or ""), page_url)
        if not source_url:
            continue
        count = 0
        review_link = container.select_one("a[href*='review.rakuten.co.jp/item/']")
        if review_link is not None:
            match = re.search(r"[（(]\s*([0-9,，]+)\s*件\s*[）)]", normalize_text(review_link.get_text(" ", strip=True)))
            if match:
                count = int(match.group(1).replace(",", "").replace("，", ""))
        review_counts[source_url] = count
    return review_counts


def parse_search_items_from_json_ld(soup: BeautifulSoup, page_url: str) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for entry in extract_json_ld_objects(soup):
        entry_type = entry.get("@type")
        if entry_type == "ItemList" or (isinstance(entry_type, list) and "ItemList" in entry_type):
            for product in json_ld_item_list_products(entry):
                item = search_item_from_json_ld_product(product, page_url)
                if item:
                    items.append(item)
            continue
        if entry_type == "Product" or (isinstance(entry_type, list) and "Product" in entry_type):
            item = search_item_from_json_ld_product(entry, page_url)
            if item:
                items.append(item)
    return items


def json_ld_item_list_products(item_list: dict[str, Any]) -> list[dict[str, Any]]:
    values = item_list.get("itemListElement")
    if not isinstance(values, list):
        return []
    products: list[dict[str, Any]] = []
    for value in values:
        if not isinstance(value, dict):
            continue
        product = value.get("item") if isinstance(value.get("item"), dict) else value
        if isinstance(product, dict):
            products.append(product)
    return products


def search_item_from_json_ld_product(product: dict[str, Any], page_url: str) -> dict[str, Any] | None:
    href = normalize_product_href(
        first_text_from_keys(product, ("url", "@id", "itemUrl", "itemPageUrl")),
        page_url,
    )
    if not href:
        return None
    title = first_text_from_keys(product, ("name", "itemName", "title")) or href
    offers = product.get("offers") if isinstance(product.get("offers"), dict) else {}
    image_url = normalize_product_image_url(first_url_from_keys(product, ("image", "imageUrl", "thumbnailUrl")))
    return {
        "title": title[:500],
        "source_url": href,
        "image_url": image_url,
        "price": price_from_rakuten_item(offers) if isinstance(offers, dict) else price_from_rakuten_item(product),
        "shop_name": "",
        "item_number": extract_item_number(href),
        "genre_id": "",
        "raw": {"pageUrl": page_url, "listSource": "json_ld"},
    }


def normalize_product_href(href: str, page_url: str) -> str:
    normalized = normalize_text(href)
    if not normalized:
        return ""
    absolute = urljoin(page_url, normalized)
    if is_rakuten_product_url(absolute):
        try:
            return normalize_rakuten_product_target(absolute)
        except RuntimeError:
            return absolute.split("?", 1)[0]
    return ""


def collect_product_detail(url: str) -> dict[str, Any]:
    normalized_url = normalize_rakuten_product_target(url)
    try:
        return collect_product_detail_from_html(normalized_url, fetch_html(normalized_url), source="http")
    except Exception as exc:
        if not should_retry_product_detail_with_browser(exc):
            raise
        try:
            html = fetch_html_with_browser(normalized_url)
        except Exception as browser_exc:
            raise RuntimeError(f"{exc}；浏览器兜底采集失败：{browser_exc}") from browser_exc
        try:
            return collect_product_detail_from_html(normalized_url, html, source="browser")
        except Exception as browser_parse_exc:
            raise RuntimeError(f"{exc}；浏览器兜底采集后仍无法解析：{browser_parse_exc}") from browser_parse_exc


def collect_product_detail_from_html(normalized_url: str, html: str, *, source: str = "http") -> dict[str, Any]:
    soup = BeautifulSoup(html, "lxml")
    fashion_url = normalized_url if parse_rakuten_fashion_product_code(normalized_url) else canonical_url(soup)
    if parse_rakuten_fashion_product_code(fashion_url):
        result = collect_rakuten_fashion_product_detail(fashion_url, html, soup)
        if normalized_url != fashion_url:
            result["source_url"] = normalized_url
    else:
        result = collect_rakuten_market_product_detail(normalized_url, html, soup)
    raw = result.get("raw")
    if isinstance(raw, dict):
        raw["detailFetchSource"] = source
    return result


def should_retry_product_detail_with_browser(exc: Exception) -> bool:
    if not settings.crawler_browser_fallback_enabled:
        return False
    text = str(exc)
    retry_markers = (
        "拦截页",
        "后端 HTTP 直接采集",
        "页面被拦截",
        "页面模板不支持",
        "未能从乐天商品详情页解析到",
        "403",
        "429",
    )
    return any(marker in text for marker in retry_markers)


def fetch_html_with_browser(url: str) -> str:
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("服务器未安装 Playwright，请执行 pip install -r requirements.txt，并运行 python -m playwright install chromium。") from exc

    timeout_ms = max(5000, int(settings.crawler_browser_timeout_seconds) * 1000)
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            try:
                context = browser.new_context(
                    user_agent=settings.crawler_user_agent,
                    locale="ja-JP",
                    timezone_id="Asia/Tokyo",
                    viewport={"width": 1366, "height": 900},
                    extra_http_headers={
                        "Accept-Language": "ja-JP,ja;q=0.9,en-US;q=0.8,en;q=0.7",
                    },
                )
                page = context.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                try:
                    page.wait_for_load_state("networkidle", timeout=min(timeout_ms, 12000))
                except PlaywrightTimeoutError:
                    pass
                html = page.content()
                context.close()
                return html
            finally:
                browser.close()
    except PlaywrightTimeoutError as exc:
        raise RuntimeError("浏览器加载乐天页面超时。") from exc
    except Exception as exc:
        message = normalize_text(str(exc))
        if "Executable doesn't exist" in message or "playwright install" in message:
            raise RuntimeError("Playwright 浏览器内核未安装，请运行 python -m playwright install chromium。") from exc
        raise


def collect_rakuten_market_product_detail(normalized_url: str, html: str, soup: BeautifulSoup) -> dict[str, Any]:
    if is_blocked_or_empty_rakuten_html(html):
        raise RuntimeError("乐天商品详情页返回拦截页，无法通过后端 HTTP 直接采集。")
    json_ld = extract_json_ld_objects(soup)
    product_json = first_json_ld_by_type(json_ld, "Product")
    embedded_item = extract_rakuten_market_item_info(soup)
    breadcrumbs = extract_breadcrumbs_from_json_ld(json_ld)
    if isinstance(embedded_item.get("breadcrumbs"), list):
        breadcrumbs = embedded_item["breadcrumbs"] or breadcrumbs
    meta = extract_page_meta(soup)
    json_title = first_text_from_keys(product_json, ("name", "itemName", "title")) if product_json else ""
    title = (
        first_text_from_keys(embedded_item, ("title", "itemName", "name"))
        or json_title
        or first_meta_text(meta, "og:title")
        or page_title(soup)
    )
    parsed_target = parse_rakuten_product_target(normalized_url)
    shop_code = parsed_target[0] if parsed_target else ""
    item_number = extract_item_number(normalized_url)
    tagline = first_text_from_keys(embedded_item, RAKUTEN_TAGLINE_KEYS)
    image_urls = market_item_image_urls(embedded_item, shop_code=shop_code, item_number=item_number)
    if not image_urls and product_json:
        image_urls = product_image_urls(product_json)
    if not image_urls:
        image_urls = extract_image_urls_from_soup(soup, shop_code=shop_code, item_number=item_number)
    image_urls = unique_texts(image_urls)
    descriptions = market_product_descriptions(
        product_json,
        soup,
        embedded_item,
        shop_code=shop_code,
        item_number=item_number,
    )
    offers = product_json.get("offers") if isinstance(product_json, dict) else None
    variants = market_item_variants(embedded_item) or variants_from_json_ld_offers(offers)
    price = price_from_rakuten_item({"variants": variants}) or price_from_rakuten_item(embedded_item)
    if price is None and isinstance(product_json, dict):
        price = price_from_rakuten_item(product_json)
    if price is None:
        price = extract_price(soup.get_text(" ", strip=True))
    if not title or title == normalized_url:
        raise RuntimeError("未能从乐天商品详情页解析到商品标题，可能页面被拦截或页面模板不支持。")
    if not image_urls and price is None and not descriptions:
        raise RuntimeError("未能从乐天商品详情页解析到有效商品数据，可能页面被拦截或页面模板不支持。")
    raw = {
        "sourceType": "rakuten_market_public",
        "url": normalized_url,
        "canonicalUrl": canonical_url(soup) or normalized_url,
        "title": title,
        "name": title,
        "itemName": title,
        "tagline": tagline,
        "catchCopyTrans": first_text_from_keys(embedded_item, ("catchCopyTrans",)) or tagline,
        "itemNumber": extract_item_number(normalized_url),
        "manageNumber": first_text_from_keys(embedded_item, ("manageNumber",)) or item_number,
        "shopCode": shop_code,
        "shopName": infer_market_shop_name(soup, embedded_item, shop_code=shop_code),
        "genreId": first_text_from_keys(embedded_item, ("rCategoryId", "genreId")),
        "price": price,
        "standardPrice": price,
        "images": image_urls,
        "productDescription": descriptions[0]["value"] if descriptions else "",
        "descriptions": descriptions,
        "variantSelectors": market_variant_selectors(embedded_item) or variant_selectors_from_variants(variants),
        "variants": variants,
        "embeddedItem": embedded_item,
        "jsonLd": json_ld,
        "breadcrumbs": breadcrumbs,
        "meta": meta,
        "collectedAt": datetime.now().isoformat(timespec="seconds"),
    }
    raw = remove_cross_item_rakuten_image_links_from_payload(
        raw,
        shop_code=shop_code,
        item_number=item_number,
    )
    return {
        "title": title[:500] or normalized_url,
        "source_url": normalized_url,
        "image_url": image_urls[0] if image_urls else "",
        "price": price,
        "shop_name": raw["shopName"],
        "item_number": raw["itemNumber"],
        "genre_id": raw["genreId"],
        "raw": raw,
    }


def collect_rakuten_fashion_product_detail(normalized_url: str, html: str, soup: BeautifulSoup) -> dict[str, Any]:
    json_ld = extract_json_ld_objects(soup)
    product_json = first_json_ld_by_type(json_ld, "Product")
    breadcrumbs = extract_breadcrumbs_from_json_ld(json_ld)
    meta = extract_page_meta(soup)
    state = extract_initial_state(html)
    product = {}
    if isinstance(state.get("itemDetail"), dict):
        item_data = state["itemDetail"].get("data")
        if isinstance(item_data, dict) and isinstance(item_data.get("product"), dict):
            product = item_data["product"]
    brand_info = state.get("brandInfo", {}).get("data") if isinstance(state.get("brandInfo"), dict) else None
    brand_info = brand_info if isinstance(brand_info, dict) else {}
    model_code = first_text_from_keys(product, ("model_cd",)) or parse_rakuten_fashion_product_code(normalized_url)
    title = (
        first_text_from_keys(product, ("product_name", "itemName", "title", "name"))
        or first_text_from_keys(product_json, ("name",)) if isinstance(product_json, dict) else ""
    ) or page_title(soup)
    image_urls = rakuten_fashion_image_urls(product)
    if isinstance(product_json, dict):
        image_urls.extend(product_image_urls(product_json))
    image_urls.extend(extract_image_urls_from_soup(soup))
    image_urls = unique_texts(image_urls)
    descriptions = rakuten_fashion_descriptions(product, brand_info, product_json)
    variants = rakuten_fashion_variants(product)
    price = (
        numeric_price(first_text_from_keys(product, ("selling_price_no_format", "selling_price")))
        or price_from_rakuten_item({"variants": variants})
        or price_from_rakuten_item(product)
        or extract_price(soup.get_text(" ", strip=True))
    )
    genre_id = first_text_from_keys(product.get("rms_info", {}) if isinstance(product.get("rms_info"), dict) else {}, ("genre_id", "genreId"))
    raw = {
        "sourceType": "rakuten_fashion_public",
        "url": normalized_url,
        "canonicalUrl": canonical_url(soup) or normalized_url,
        "title": title,
        "name": title,
        "itemName": title,
        "tagline": first_text_from_keys(product, RAKUTEN_TAGLINE_KEYS),
        "modelCode": model_code,
        "itemNumber": model_code,
        "manageNumber": model_code,
        "brandNo": first_text_from_keys(product, ("brand_no",)),
        "externalCode": first_text_from_keys(product, ("external_cd",)),
        "brand": first_text_from_keys(product, ("brand_name",)) or first_text_from_keys(brand_info, ("brand_name",)),
        "brandKana": first_text_from_keys(product, ("brand_name_kana",)) or first_text_from_keys(brand_info, ("brand_name_kana",)),
        "makerName": first_text_from_keys(product, ("maker_name",)),
        "shopName": first_text_from_keys(product, ("site_name",)) or "Rakuten Fashion",
        "genreId": genre_id,
        "categoryLName": first_text_from_keys(product, ("category_l_name",)),
        "categoryMName": first_text_from_keys(product, ("category_m_name",)),
        "categoryLCode": first_text_from_keys(product, ("category_l_cd",)),
        "categoryMCode": first_text_from_keys(product, ("category_m_cd",)),
        "price": price,
        "fixedPrice": first_text_from_keys(product, ("fixed_price_no_format", "fixed_price")),
        "sellingPrice": first_text_from_keys(product, ("selling_price_no_format", "selling_price")),
        "discountRate": product.get("discount_rate"),
        "currency": "JPY",
        "images": image_urls,
        "productDescription": {"pc": first_text_from_keys(product, ("product_exp",))},
        "descriptions": descriptions,
        "variantSelectors": variant_selectors_from_variants(variants),
        "variants": variants,
        "inventory": product.get("rms_info", {}).get("inventory_list") if isinstance(product.get("rms_info"), dict) else [],
        "favoriteCount": first_text_from_keys(product, ("favorite_count",)),
        "saleStatus": product.get("sale_status"),
        "saleComment": first_text_from_keys(product, ("sale_comment",)),
        "soldout": product.get("soldout_flg"),
        "soldoutPart": product.get("soldout_part_flg"),
        "preorder": product.get("preorder_flg"),
        "material": rakuten_fashion_first_sku_value(product, "material"),
        "origin": first_text_from_keys(product, ("natives",)),
        "rmsInfo": product.get("rms_info") if isinstance(product.get("rms_info"), dict) else {},
        "coupons": product.get("coupon_list") if isinstance(product.get("coupon_list"), list) else [],
        "brandInfo": brand_info,
        "jsonLd": json_ld,
        "breadcrumbs": breadcrumbs,
        "meta": meta,
        "sourceProduct": product,
        "collectedAt": datetime.now().isoformat(timespec="seconds"),
    }
    return {
        "title": title[:500] or normalized_url,
        "source_url": normalized_url,
        "image_url": image_urls[0] if image_urls else "",
        "price": price,
        "shop_name": raw["brand"] or raw["shopName"],
        "item_number": model_code,
        "genre_id": genre_id,
        "raw": raw,
    }


def extract_initial_state(html: str) -> dict[str, Any]:
    match = re.search(r"window\.__INITIAL_STATE__\s*=\s*(\{.*?\});\s*\nwindow\.__REWIRED_SCHEMAS__", html, re.S)
    if not match:
        return {}
    try:
        state = json.loads(match.group(1))
    except ValueError:
        return {}
    return state if isinstance(state, dict) else {}


def is_blocked_or_empty_rakuten_html(html: str) -> bool:
    text = normalize_text(BeautifulSoup(html or "", "lxml").get_text(" ", strip=True))
    if not text:
        return True
    if len(html or "") < 300 and re.fullmatch(r"Reference\s+#.+", text):
        return True
    return False


def extract_rakuten_market_item_info(soup: BeautifulSoup) -> dict[str, Any]:
    for script in soup.find_all("script"):
        text = script.string or script.get_text() or ""
        if '"itemInfoSku"' not in text:
            continue
        stripped = text.strip()
        if not stripped.startswith("{"):
            continue
        try:
            payload = json.loads(stripped)
        except ValueError:
            continue
        for path in (("api", "data", "itemInfoSku"), ("newApi", "itemInfoSku")):
            value: Any = payload
            for key in path:
                if not isinstance(value, dict):
                    value = None
                    break
                value = value.get(key)
            if isinstance(value, dict):
                return merge_rakuten_market_embedded_payload(value, payload)
    return {}


def merge_rakuten_market_embedded_payload(item_info: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    result = dict(item_info)
    result["embeddedPayload"] = payload
    for key in (
        "title",
        "itemName",
        "name",
        "tagline",
        "catchcopy",
        "catchCopy",
        "catchCopyTrans",
        "subTitle",
        "subtitle",
        "newProductDescription",
        "salesDescription",
        "rCategoryId",
        "genreId",
        "manageNumber",
    ):
        if not has_description_source(result.get(key)):
            value = first_value_by_key(payload, key)
            if has_description_source(value):
                result[key] = value
    pc_fields = result.get("pcFields") if isinstance(result.get("pcFields"), dict) else {}
    pc_description = pc_fields.get("productDescription") if isinstance(pc_fields, dict) else None
    if not has_description_source(pc_description):
        value = first_value_by_key(payload, "productDescription")
        if has_description_source(value):
            next_pc_fields = dict(pc_fields)
            next_pc_fields["productDescription"] = value
            result["pcFields"] = next_pc_fields
    return result


def first_value_by_key(source: Any, target_key: str) -> Any:
    if isinstance(source, dict):
        if target_key in source and has_description_source(source.get(target_key)):
            return source.get(target_key)
        for child in source.values():
            value = first_value_by_key(child, target_key)
            if has_description_source(value):
                return value
    elif isinstance(source, list):
        for child in source:
            value = first_value_by_key(child, target_key)
            if has_description_source(value):
                return value
    return None


def extract_json_ld_objects(soup: BeautifulSoup) -> list[dict[str, Any]]:
    objects: list[dict[str, Any]] = []
    for script in soup.find_all("script", type="application/ld+json"):
        text = script.string or script.get_text() or ""
        try:
            value = json.loads(text)
        except ValueError:
            continue
        for item in flatten_json_ld(value):
            if isinstance(item, dict):
                objects.append(item)
    return objects


def flatten_json_ld(value: Any) -> list[Any]:
    if isinstance(value, list):
        result: list[Any] = []
        for item in value:
            result.extend(flatten_json_ld(item))
        return result
    if isinstance(value, dict) and isinstance(value.get("@graph"), list):
        return flatten_json_ld(value.get("@graph"))
    return [value]


def first_json_ld_by_type(objects: list[dict[str, Any]], target_type: str) -> dict[str, Any]:
    for item in objects:
        item_type = item.get("@type")
        if isinstance(item_type, list) and target_type in item_type:
            return item
        if item_type == target_type:
            return item
    return {}


def extract_breadcrumbs_from_json_ld(objects: list[dict[str, Any]]) -> list[dict[str, str]]:
    breadcrumb = first_json_ld_by_type(objects, "BreadcrumbList")
    values = breadcrumb.get("itemListElement") if isinstance(breadcrumb, dict) else None
    result: list[dict[str, str]] = []
    if not isinstance(values, list):
        return result
    for item in values:
        if not isinstance(item, dict):
            continue
        child = item.get("item") if isinstance(item.get("item"), dict) else {}
        result.append(
            {
                "name": first_text_from_keys(child, ("name",)) or first_text_from_keys(item, ("name",)),
                "url": first_text_from_keys(child, ("@id", "url")),
            }
        )
    return [item for item in result if item.get("name") or item.get("url")]


def extract_page_meta(soup: BeautifulSoup) -> dict[str, str]:
    meta: dict[str, str] = {}
    for node in soup.select("meta"):
        key = normalize_text(node.get("property") or node.get("name"))
        content = normalize_text(node.get("content"))
        if key and content and key not in meta:
            meta[key] = content
    return meta


def first_meta_text(meta: dict[str, str], key: str) -> str:
    return normalize_text(meta.get(key))


def page_title(soup: BeautifulSoup) -> str:
    title_node = soup.select_one("h1") or soup.select_one("title")
    return normalize_text(title_node.get_text(" ", strip=True) if title_node else "")


def canonical_url(soup: BeautifulSoup) -> str:
    node = soup.find("link", rel="canonical")
    return normalize_text(node.get("href") if node else "")


def infer_market_shop_name(soup: BeautifulSoup, embedded_item: dict[str, Any] | None = None, shop_code: str = "") -> str:
    if embedded_item:
        shop_status = embedded_item.get("shopStatus")
        if isinstance(shop_status, dict):
            name = first_text_from_keys(shop_status, ("shopName", "name"))
            if name:
                return name
    if soup.title:
        title = normalize_text(soup.title.get_text(" ", strip=True))
        if "：" in title:
            candidate = title.rsplit("：", 1)[-1].strip()
            if candidate and candidate != "楽天市場":
                return candidate
    for selector in ("meta[property='og:site_name']", "#shopName", ".shopName"):
        node = soup.select_one(selector)
        value = normalize_text(node.get("content") if node and node.name == "meta" else node.get_text(" ", strip=True) if node else "")
        if value and value != "楽天市場":
            return value
    return normalize_shop_code(shop_code)


def market_product_descriptions(
    product_json: dict[str, Any],
    soup: BeautifulSoup,
    embedded_item: dict[str, Any] | None = None,
    *,
    shop_code: str = "",
    item_number: str = "",
) -> list[dict[str, str]]:
    descriptions: list[dict[str, str]] = []
    description = first_text_from_keys(product_json, ("description",)) if isinstance(product_json, dict) else ""
    if description:
        descriptions.append({"label": "商品说明", "value": description})
    if embedded_item:
        embedded_descriptions = (
            ("スマートフォン用 商品説明文", embedded_item.get("newProductDescription")),
            ("PC用 商品説明文", (embedded_item.get("pcFields") or {}).get("productDescription") if isinstance(embedded_item.get("pcFields"), dict) else None),
            ("PC用 販売説明文", embedded_item.get("salesDescription")),
        )
        for label, value in embedded_descriptions:
            html_value = normalize_detail_html(
                remove_cross_item_rakuten_image_links(
                    value,
                    shop_code=shop_code,
                    item_number=item_number,
                )
            )
            if html_value and all(description_content_key(item["value"]) != description_content_key(html_value) for item in descriptions):
                descriptions.append({"label": label, "value": html_value})
    for selector, label in (
        ("#itemCaption", "商品说明"),
        ("#itemDescription", "商品详情"),
        ("[class*='description']", "商品说明"),
    ):
        node = soup.select_one(selector)
        if node:
            text_value = strip_low_quality_description_lines(str(node))
            if text_value and all(description_content_key(item["value"]) != description_content_key(text_value) for item in descriptions):
                descriptions.append({"label": label, "value": text_value})
    return clean_market_product_descriptions(descriptions, keep_empty_labels=RAKUTEN_DESCRIPTION_FIELD_LABELS)


def remove_cross_item_rakuten_image_links(
    value: Any,
    *,
    shop_code: str,
    item_number: str,
) -> str:
    html = str(value or "")
    normalized_item_number = normalize_text(item_number).lower()
    if not html.strip() or not normalized_item_number:
        return html
    normalized_shop_code = normalize_shop_code(shop_code).lower()
    soup = BeautifulSoup(html, "lxml")
    removed = False
    for link in soup.select("a[href]"):
        if link.select_one("img, source") is None:
            continue
        target = linked_rakuten_item_target(link.get("href"))
        if target is None:
            continue
        target_shop_code, target_item_number = target
        if normalized_shop_code and target_shop_code.lower() != normalized_shop_code:
            continue
        if target_item_number.lower() != normalized_item_number:
            link.decompose()
            removed = True
    if not removed:
        return html
    body = soup.body
    if body is not None:
        return body.decode_contents()
    return str(soup)


def remove_cross_item_rakuten_image_links_from_payload(
    payload: dict[str, Any],
    *,
    shop_code: str,
    item_number: str,
) -> dict[str, Any]:
    updated_payload = deepcopy(payload)
    description_keys = {
        "description",
        "descriptions",
        "newproductdescription",
        "pcdescription",
        "productdescription",
        "salesdescription",
        "smartphonedescription",
        "spdescription",
    }

    def clean(value: Any) -> str:
        return remove_cross_item_rakuten_image_links(
            value,
            shop_code=shop_code,
            item_number=item_number,
        )

    def clean_description_value(value: Any) -> Any:
        if isinstance(value, str):
            return clean(value)
        if isinstance(value, dict):
            return {
                key: clean_description_value(child)
                for key, child in value.items()
            }
        if isinstance(value, list):
            return [clean_description_value(child) for child in value]
        return value

    def walk(value: Any) -> Any:
        if isinstance(value, dict):
            return {
                key: (
                    clean_description_value(child)
                    if str(key).lower() in description_keys
                    else walk(child)
                )
                for key, child in value.items()
            }
        if isinstance(value, list):
            return [walk(child) for child in value]
        return value

    return walk(updated_payload)


def linked_rakuten_item_target(value: Any) -> tuple[str, str] | None:
    url = normalize_text(value)
    if not url:
        return None
    try:
        parsed = urlsplit(url)
    except Exception:
        return None
    if parsed.netloc.lower() not in {"item.rakuten.co.jp", "soko.rms.rakuten.co.jp"}:
        return None
    parts = [unquote(part.strip()) for part in parsed.path.split("/") if part.strip()]
    if len(parts) < 2:
        return None
    return parts[0], parts[1]


RAKUTEN_DESCRIPTION_FIELD_LABELS = (
    "PC用 商品説明文",
    "スマートフォン用 商品説明文",
    "PC用 販売説明文",
    "PC用商品说明文",
    "智能手机用商品说明文",
    "PC用销售说明文",
)


RAKUTEN_STANDARD_DESCRIPTION_LABELS = (
    "PC用 商品説明文",
    "スマートフォン用 商品説明文",
    "PC用 販売説明文",
)


LOW_QUALITY_DESCRIPTION_KEYWORDS = (
    "キャンセルポリシー",
    "メーカー希望小売価格",
    "メーカーカタログ",
    "メーカーサイトに基づいて掲載",
    "メーカーサイトTOP",
    "メーカーサイト会社概要",
    "特定商取引法表示",
    "会社概要",
    "有効期間",
    "年間ランキング",
    "ランキング",
    "受賞",
    "買い物かご",
    "商品レビュー",
    "ショップレビュー",
)
PRODUCT_DESCRIPTION_KEYWORDS = (
    "商品ポイント",
    "デザイン",
    "シルエット",
    "コーディネート",
    "サイズ",
    "カラー",
    "素材",
    "原産国",
    "生産国",
    "商品名",
    "商品コード",
    "洗濯",
    "着用",
    "アイテム",
    "セットアップ",
    "パンツ",
    "トップス",
    "ウエスト",
    "伸縮",
    "仕様",
    "重量",
    "長さ",
)


def clean_market_product_descriptions(
    descriptions: list[dict[str, str]],
    *,
    keep_empty_labels: tuple[str, ...] = (),
) -> list[dict[str, str]]:
    if not descriptions:
        return []
    normalized: list[dict[str, str]] = []
    seen_by_content: dict[str, int] = {}
    seen_empty_official: set[str] = set()
    keep_empty_label_set = {normalize_text(label) for label in keep_empty_labels}
    for item in descriptions:
        label = normalize_text(item.get("label")) or "商品说明"
        is_official_rakuten_field = label in keep_empty_label_set
        value = normalize_listing_detail_html(item.get("value")) if is_official_rakuten_field else normalize_detail_html(item.get("value"))
        if not is_official_rakuten_field:
            value = strip_low_quality_description_lines(value)
        value_key = description_content_key(value)
        if (not value or not value_key) and not is_official_rakuten_field:
            continue
        if not is_official_rakuten_field and is_low_quality_product_description(value):
            continue
        if not value_key:
            if label in seen_empty_official:
                continue
            seen_empty_official.add(label)
            normalized.append({"label": label, "value": value})
            continue
        content_key = f"{label}\0{value_key}" if is_official_rakuten_field else value_key
        previous_index = seen_by_content.get(content_key)
        current = {"label": label, "value": value}
        if previous_index is not None:
            previous = normalized[previous_index]
            if description_label_priority(label, keep_empty_label_set) > description_label_priority(previous["label"], keep_empty_label_set):
                normalized[previous_index] = current
            continue
        seen_by_content[content_key] = len(normalized)
        normalized.append(current)
    if not normalized:
        return []
    return normalized


def normalize_rakuten_description_fields(descriptions: list[dict[str, str]]) -> list[dict[str, str]]:
    fields: dict[str, str] = {label: "" for label in RAKUTEN_STANDARD_DESCRIPTION_LABELS}
    extras: list[dict[str, str]] = []
    for item in descriptions:
        label = normalize_text(item.get("label")) or "商品说明"
        target_label = standard_rakuten_description_label(label)
        value = normalize_listing_detail_html(item.get("value")) if target_label else normalize_detail_html(item.get("value"))
        if target_label:
            if value and not fields[target_label]:
                fields[target_label] = value
            continue
        if value:
            extras.append({"label": label, "value": value})

    for item in extras:
        label = item["label"]
        value = item["value"]
        target_label = fallback_rakuten_description_label(label, value)
        if value and not fields[target_label]:
            fields[target_label] = value

    return [{"label": label, "value": fields[label]} for label in RAKUTEN_STANDARD_DESCRIPTION_LABELS]


def standard_rakuten_description_label(label: str) -> str:
    normalized = normalize_text(label).replace(" ", "")
    if normalized in {"PC用商品説明文", "PC用商品说明文", "PC商品说明", "PC用商品説明"}:
        return "PC用 商品説明文"
    if normalized in {"スマートフォン用商品説明文", "スマートフォン用商品说明文", "智能手机商品说明", "智能手机用商品说明文", "移动端商品说明"}:
        return "スマートフォン用 商品説明文"
    if normalized in {"PC用販売説明文", "PC用销售说明文", "销售说明", "販売説明文"}:
        return "PC用 販売説明文"
    return ""


def fallback_rakuten_description_label(label: str, value: str) -> str:
    normalized = normalize_text(label)
    if "販売" in normalized or "销售" in normalized or "sale" in normalized.lower():
        return "PC用 販売説明文"
    if "スマートフォン" in normalized or "智能手机" in normalized or "移动" in normalized:
        return "スマートフォン用 商品説明文"
    return "PC用 商品説明文"


def description_label_priority(label: str, official_labels: set[str]) -> int:
    normalized_label = normalize_text(label)
    if normalized_label in official_labels:
        return 20
    if normalized_label in {"结构化商品说明", "商品详情", "商品说明"}:
        return 10
    return 0


def first_description_by_label(descriptions: list[dict[str, str]], labels: tuple[str, ...]) -> str:
    normalized_labels = {normalize_text(label) for label in labels}
    for description in descriptions:
        if normalize_text(description.get("label")) in normalized_labels:
            return normalize_listing_detail_html(description.get("value"))
    return ""


def best_product_description(descriptions: list[dict[str, str]]) -> str:
    if not descriptions:
        return ""
    return max(descriptions, key=product_description_quality_score).get("value") or ""


def strip_low_quality_description_lines(value: str) -> str:
    text = detail_html_plain_text(normalize_detail_html(value))
    if not text:
        return ""
    lines: list[str] = []
    for raw_line in re.split(r"[\r\n]+", text):
        line = normalize_text(raw_line)
        if not line:
            continue
        if is_low_quality_description_line(line):
            continue
        lines.append(line)
    return "\n".join(lines).strip()


def is_low_quality_description_line(line: str) -> bool:
    if not line or line in {"?", "？", "-", "ー"}:
        return True
    keyword_hits = sum(1 for keyword in LOW_QUALITY_DESCRIPTION_KEYWORDS if keyword in line)
    product_hits = sum(1 for keyword in PRODUCT_DESCRIPTION_KEYWORDS if keyword in line)
    if keyword_hits and product_hits:
        return False
    if keyword_hits >= 1 and len(line) <= 120:
        return True
    if keyword_hits >= 2 and len(line) <= 240:
        return True
    return False


def is_low_quality_product_description(value: str) -> bool:
    plain_text = detail_html_plain_text(value)
    if not plain_text:
        return True
    if plain_text in {"?", "？", "-", "ー"}:
        return True
    keyword_hits = sum(1 for keyword in LOW_QUALITY_DESCRIPTION_KEYWORDS if keyword in plain_text)
    product_hits = sum(1 for keyword in PRODUCT_DESCRIPTION_KEYWORDS if keyword in plain_text)
    if len(plain_text) < 40 and product_hits == 0:
        return True
    if keyword_hits >= 2 and product_hits == 0:
        return True
    if len(plain_text) < 120 and keyword_hits >= 1 and product_hits == 0:
        return True
    return False


def is_near_duplicate_description(left: str, right: str) -> bool:
    left_text = normalize_text(left)
    right_text = normalize_text(right)
    if not left_text or not right_text:
        return False
    if left_text == right_text:
        return True
    shorter, longer = sorted((left_text, right_text), key=len)
    return len(shorter) >= 80 and shorter in longer and (len(longer) - len(shorter) <= 80)


def product_description_quality_score(item: dict[str, str]) -> int:
    label = normalize_text(item.get("label"))
    value = normalize_detail_html(item.get("value"))
    plain_text = detail_html_plain_text(value)
    score = min(len(plain_text), 2000)
    score += sum(250 for keyword in PRODUCT_DESCRIPTION_KEYWORDS if keyword in plain_text)
    score -= sum(350 for keyword in LOW_QUALITY_DESCRIPTION_KEYWORDS if keyword in plain_text)
    if "销售" in label or "sales" in label.lower():
        score += 300
    if "PC" in label or "商品说明" in label:
        score += 80
    return score


def detail_html_plain_text(value: Any) -> str:
    return normalize_text(BeautifulSoup(str(value or ""), "lxml").get_text(" ", strip=True))


def description_content_key(value: Any) -> str:
    html = str(value or "")
    plain_text = detail_html_plain_text(html)
    if plain_text:
        return plain_text
    soup = BeautifulSoup(html, "lxml")
    image_sources: list[str] = []
    for image in soup.select("img, source"):
        src = image.get("src") or image.get("data-src") or image.get("data-original") or image.get("srcset")
        src_text = normalize_text(src)
        if src_text:
            image_sources.append(src_text)
    if image_sources:
        return "|".join(image_sources)
    return normalize_text(html)


def market_item_image_urls(item: dict[str, Any], *, shop_code: str, item_number: str) -> list[str]:
    urls: list[str] = []
    media = item.get("media") if isinstance(item.get("media"), dict) else {}
    pc_fields = item.get("pcFields") if isinstance(item.get("pcFields"), dict) else {}

    def collect(value: Any) -> None:
        if isinstance(value, str):
            url = normalize_product_image_url(value, shop_code=shop_code)
            if is_relevant_market_item_image(url, shop_code=shop_code, item_number=item_number) and url not in urls:
                urls.append(url)
            return
        if isinstance(value, dict):
            for key in ("location", "url", "imageUrl", "src"):
                collect(value.get(key))
            for child in value.values():
                collect(child)
            return
        if isinstance(value, list):
            for child in value:
                collect(child)

    collect(pc_fields.get("images"))
    collect(media.get("images"))
    collect(media.get("skuImages"))
    collect(item.get("picImageUrl"))
    for sku in item.get("sku") if isinstance(item.get("sku"), list) else []:
        if isinstance(sku, dict):
            collect(sku.get("images"))
    return urls


def is_relevant_market_item_image(url: str, *, shop_code: str, item_number: str) -> bool:
    if not url:
        return False
    try:
        parsed = urlsplit(url)
    except Exception:
        return False
    host = parsed.netloc.lower()
    path = parsed.path.lower()
    normalized_shop = normalize_shop_code(shop_code).lower()
    if host == "r.r10s.jp":
        return False
    if normalized_shop and normalized_shop not in path and normalized_shop not in host:
        return False
    if "/cabinet/" in path and normalized_shop and (normalized_shop in path or normalized_shop in host):
        return True
    item_tokens = item_number_image_tokens(item_number)
    if item_tokens:
        return any(token in path for token in item_tokens)
    if host in {"image.rakuten.co.jp", "tshop.r10s.jp", "cabinet.rms.rakuten.co.jp"}:
        return True
    return False


def item_number_image_tokens(item_number: str) -> list[str]:
    normalized = normalize_text(item_number).lower()
    tokens = [normalized] if normalized else []
    for part in re.split(r"[^a-z0-9]+", normalized):
        if len(part) >= 4 and part not in tokens:
            tokens.append(part)
    return tokens


def market_item_variants(item: dict[str, Any]) -> dict[str, dict[str, Any]]:
    skus = item.get("sku")
    if not isinstance(skus, list):
        return {}
    selectors = item.get("variantSelectors") if isinstance(item.get("variantSelectors"), list) else []
    purchase_info = item.get("purchaseInfo") if isinstance(item.get("purchaseInfo"), dict) else {}
    purchase_skus = purchase_info.get("sku") if isinstance(purchase_info.get("sku"), list) else []
    purchase_by_variant = {
        first_text_from_keys(row, ("variantId",)): row for row in purchase_skus if isinstance(row, dict)
    }
    inventories = purchase_info.get("variantMappedInventories") if isinstance(purchase_info.get("variantMappedInventories"), list) else []
    inventory_by_variant = {
        first_text_from_keys(row, ("sku",)): row for row in inventories if isinstance(row, dict)
    }
    result: dict[str, dict[str, Any]] = {}
    for index, sku in enumerate(skus, start=1):
        if not isinstance(sku, dict):
            continue
        variant_id = first_text_from_keys(sku, ("variantId",)) or f"sku-{index}"
        purchase_row = purchase_by_variant.get(variant_id, {})
        inventory_row = inventory_by_variant.get(variant_id, {})
        purchase_sku = purchase_row.get("newPurchaseSku") if isinstance(purchase_row.get("newPurchaseSku"), dict) else {}
        selector_values = market_selector_values(sku.get("selectorValues"), selectors)
        price = (
            first_text_from_keys(sku, ("taxIncludedPrice", "standardPrice", "price"))
            or first_text_from_keys(purchase_row, ("taxIncludedPrice", "standardPrice", "price"))
        )
        result[variant_id] = {
            "variantId": variant_id,
            "merchantDefinedSkuId": first_text_from_keys(sku, ("merchantDefinedSkuId",)),
            "articleNumber": first_text_value(sku.get("articleNumber")),
            "standardPrice": price,
            "hidden": bool(sku.get("hidden")),
            "selectorValues": selector_values,
            "specs": market_named_values(sku.get("specs")),
            "attributes": market_named_values(sku.get("attributes")),
            "inventoryId": first_text_from_keys(inventory_row, ("inventoryId",)),
            "material": first_attribute_value(sku.get("attributes"), ("素材", "素材（生地・毛糸）")),
            "images": product_image_urls({"images": sku.get("images")}),
            "referencePrice": first_text_from_keys(sku.get("referencePrice", {}) if isinstance(sku.get("referencePrice"), dict) else {}, ("value",)),
        }
    return result


def market_variant_selectors(item: dict[str, Any]) -> list[dict[str, Any]]:
    selectors = item.get("variantSelectors") if isinstance(item.get("variantSelectors"), list) else []
    result: list[dict[str, Any]] = []
    for index, selector in enumerate(selectors, start=1):
        if not isinstance(selector, dict):
            continue
        key = first_text_from_keys(selector, ("key",)) or f"k{index}"
        result.append(
            {
                "key": key,
                "name": first_text_from_keys(selector, ("label", "name", "displayName")) or key,
                "values": selector_values_to_public(selector.get("values")),
            }
        )
    return result


def market_selector_values(values: Any, selectors: list[Any]) -> dict[str, str]:
    result: dict[str, str] = {}
    if not isinstance(values, list):
        return result
    for index, value in enumerate(values):
        selector = selectors[index] if index < len(selectors) and isinstance(selectors[index], dict) else {}
        key = first_text_from_keys(selector, ("key",)) or f"k{index + 1}"
        result[key] = first_text_value(value)
    return result


def market_named_values(values: Any) -> list[dict[str, str]]:
    if not isinstance(values, list):
        return []
    result: list[dict[str, str]] = []
    for item in values:
        if not isinstance(item, dict):
            continue
        name = first_text_from_keys(item, ("title", "name", "label"))
        value = first_text_from_keys(item, ("value", "text"))
        if name or value:
            result.append({"name": name, "value": value})
    return result


def first_attribute_value(values: Any, labels: tuple[str, ...]) -> str:
    if not isinstance(values, list):
        return ""
    for item in values:
        if not isinstance(item, dict):
            continue
        title = first_text_from_keys(item, ("title", "name", "label"))
        if any(label in title for label in labels):
            return first_text_from_keys(item, ("value", "text"))
    return ""


def variants_from_json_ld_offers(offers: Any) -> dict[str, dict[str, Any]]:
    offer_items = offers if isinstance(offers, list) else [offers] if isinstance(offers, dict) else []
    variants: dict[str, dict[str, Any]] = {}
    for index, offer in enumerate(offer_items, start=1):
        if not isinstance(offer, dict):
            continue
        sku = first_text_from_keys(offer, ("sku", "mpn")) or f"sku-{index}"
        variants[sku] = {
            "variantId": sku,
            "merchantDefinedSkuId": sku,
            "standardPrice": first_text_from_keys(offer, ("price",)),
            "hidden": "OutOfStock" in first_text_from_keys(offer, ("availability",)),
            "selectorValues": {},
            "images": [first_text_from_keys(offer, ("image",))] if first_text_from_keys(offer, ("image",)) else [],
            "availability": first_text_from_keys(offer, ("availability",)),
            "itemCondition": first_text_from_keys(offer, ("itemCondition",)),
        }
    return variants


def rakuten_fashion_image_urls(product: dict[str, Any]) -> list[str]:
    filenames: list[str] = []

    def remember_filename(value: Any) -> None:
        text = normalize_text(value)
        if not text or not re.search(r"\.(?:jpe?g|png|webp|gif)$", text, flags=re.I):
            return
        if text not in filenames:
            filenames.append(text)

    for key in ("product_img_path",):
        remember_filename(product.get(key))
    sku_images = product.get("product_sku_img_path")
    if isinstance(sku_images, dict):
        for value in sku_images.values():
            remember_filename(value)
    sku_sub_images = product.get("product_sku_img_path_sub")
    if isinstance(sku_sub_images, list):
        for value in sku_sub_images:
            remember_filename(value)
    model_info = product.get("product_sku_img_model_info")
    if isinstance(model_info, dict):
        for key in model_info.keys():
            remember_filename(key)
    return unique_texts([rakuten_fashion_image_url(filename) for filename in filenames])


def rakuten_fashion_image_url(filename: str) -> str:
    normalized = normalize_text(filename).lower()
    match = re.search(r"([a-z0-9]+)-", normalized)
    directory = ""
    if match:
        code = match.group(1)
        directory = code[-3:]
    if not directory:
        directory = "000"
    return f"{RAKUTEN_FASHION_IMAGE_BASE}/{directory}/{normalized}"


def rakuten_fashion_descriptions(
    product: dict[str, Any],
    brand_info: dict[str, Any],
    product_json: dict[str, Any] | None,
) -> list[dict[str, str]]:
    descriptions: list[dict[str, str]] = []
    product_exp = first_text_from_keys(product, ("product_exp",))
    if product_exp:
        descriptions.append({"label": "商品说明", "value": normalize_detail_html(product_exp)})
    json_description = first_text_from_keys(product_json or {}, ("description",))
    if json_description and all(item["value"] != json_description for item in descriptions):
        descriptions.append({"label": "结构化商品说明", "value": json_description})
    brand_exp = first_text_from_keys(brand_info, ("brand_exp",))
    if brand_exp:
        descriptions.append({"label": "品牌说明", "value": normalize_detail_html(brand_exp)})
    return descriptions


def rakuten_fashion_variants(product: dict[str, Any]) -> dict[str, dict[str, Any]]:
    skus = product.get("product_sku")
    if not isinstance(skus, list):
        return {}
    variants: dict[str, dict[str, Any]] = {}
    for index, sku in enumerate(skus, start=1):
        if not isinstance(sku, dict):
            continue
        inventory = sku.get("inventory_info") if isinstance(sku.get("inventory_info"), dict) else {}
        variant_id = first_text_from_keys(sku, ("inventory_id",)) or first_text_from_keys(inventory, ("variant_id", "inventory_id")) or f"sku-{index}"
        color = first_text_from_keys(sku, ("rms_v_choise_name", "product_color_name")) or first_text_from_keys(inventory, ("color_name",))
        size = first_text_from_keys(sku, ("rms_h_choise_name", "product_size_name")) or first_text_from_keys(inventory, ("size",))
        image_path = first_text_from_keys(sku, ("product_img_path",))
        variants[variant_id] = {
            "variantId": variant_id,
            "merchantDefinedSkuId": variant_id,
            "articleNumber": "",
            "standardPrice": first_text_from_keys(sku, ("selling_price_with_tax", "selling_price_tax_included", "tax_included_selling_price"))
            or first_text_from_keys(product, ("selling_price_no_format",))
            or first_text_from_keys(sku, ("selling_price", "fixed_price")),
            "displayPrice": first_text_from_keys(product, ("selling_price",)) or first_text_from_keys(sku, ("selling_price",)),
            "fixedPrice": first_text_from_keys(sku, ("fixed_price",)),
            "hidden": first_text_from_keys(sku, ("inventory_exist_flg",)) == "0",
            "selectorValues": {"color": color, "size": size},
            "specs": [
                {"name": "素材", "value": first_text_from_keys(sku, ("material",))},
                {"name": "発送予定", "value": first_text_from_keys(sku, ("inventory_status_message",))},
            ],
            "attributes": [
                {"name": "颜色代码", "value": first_text_from_keys(sku, ("product_color_cd",))},
            ],
            "inventoryId": first_text_from_keys(sku, ("inventory_id",)),
            "material": first_text_from_keys(sku, ("material",)),
            "images": [rakuten_fashion_image_url(image_path)] if image_path else [],
        }
    return variants


def rakuten_fashion_first_sku_value(product: dict[str, Any], key: str) -> str:
    skus = product.get("product_sku")
    if not isinstance(skus, list):
        return ""
    for sku in skus:
        if isinstance(sku, dict):
            value = first_text_from_keys(sku, (key,))
            if value:
                return value
    return ""


def variant_selectors_from_variants(variants: Any) -> list[dict[str, Any]]:
    variant_values = variants.values() if isinstance(variants, dict) else variants if isinstance(variants, list) else []
    selectors: dict[str, list[str]] = {}
    for variant in variant_values:
        if not isinstance(variant, dict):
            continue
        selector_values = variant.get("selectorValues")
        if not isinstance(selector_values, dict):
            continue
        for key, value in selector_values.items():
            text = normalize_text(value)
            if not text:
                continue
            selectors.setdefault(str(key), [])
            if text not in selectors[str(key)]:
                selectors[str(key)].append(text)
    return [{"key": key, "name": selector_display_name(key), "values": values} for key, values in selectors.items()]


def selector_display_name(key: str) -> str:
    return {"color": "颜色", "size": "尺码"}.get(key, key)


def extract_image_urls_from_soup(
    soup: BeautifulSoup,
    *,
    shop_code: str = "",
    item_number: str = "",
) -> list[str]:
    urls: list[str] = []
    for node in soup.select("img"):
        for attr in ("src", "data-src", "data-original", "data-lazy-src"):
            value = node.get(attr)
            if value:
                normalized = normalize_product_image_url(value)
                if normalized and (
                    not shop_code
                    or is_relevant_market_item_image(normalized, shop_code=shop_code, item_number=item_number)
                ):
                    urls.append(normalized)
    return unique_texts(urls)


def unique_texts(values: list[Any]) -> list[str]:
    result: list[str] = []
    for value in values:
        text = normalize_text(value)
        if text and text not in result:
            result.append(text)
    return result


def normalize_detail_html(value: Any) -> str:
    text = str(value or "").replace("\\/", "/")
    text = unescape(text)
    text = re.sub(r"<script[\s\S]*?</script>", "", text, flags=re.I)
    text = re.sub(r"<style[\s\S]*?</style>", "", text, flags=re.I)
    return sanitize_description_html(text).strip()


def normalize_listing_detail_html(value: Any) -> str:
    text = str(value or "").replace("\\/", "/")
    text = unescape(text)
    if not text.strip():
        return ""
    return sanitize_rakuten_pc_description_html(text).strip()


def sanitize_description_html(value: str) -> str:
    soup = BeautifulSoup(value or "", "lxml")
    for element in soup.select("script, style, iframe, object, embed, link, meta, svg, canvas, video, audio"):
        element.decompose()
    for element in soup.select("*"):
        for attribute in list(element.attrs):
            name = attribute.lower()
            value_text = " ".join(element.get_attribute_list(attribute)).strip()
            if name.startswith("on") or value_text.lower().startswith("javascript:"):
                del element.attrs[attribute]
    body = soup.body
    if body is not None:
        return body.decode_contents().strip()
    return str(soup).strip()


def has_description_source(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, tuple, set, dict)):
        return bool(value)
    return True


def numeric_price(value: Any) -> float | None:
    text = first_text_value(value)
    if not text:
        return None
    normalized = re.sub(r"[^0-9.]", "", text)
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def extract_price(text: str) -> float | None:
    matches = re.findall(r"([0-9][0-9,]{2,})\s*円", text)
    if not matches:
        matches = re.findall(r"￥\s*([0-9][0-9,]{2,})", text)
    if not matches:
        return None
    try:
        return float(matches[0].replace(",", ""))
    except ValueError:
        return None


def extract_item_number(url: str) -> str:
    parsed_target = parse_rakuten_product_target(url)
    if parsed_target is not None:
        return parsed_target[1][:255]
    parts = [part for part in url.rstrip("/").split("/") if part]
    return parts[-1][:255] if parts else ""


def upsert_store_product(
    session: Any,
    owner_username: str,
    store: StoreModel,
    item: dict[str, Any],
    *,
    active_words: list[str] | None = None,
) -> bool:
    item_number = first_text_from_keys(item, ("itemNumber", "manageNumber"))
    manage_number = first_text_from_keys(item, ("manageNumber", "itemNumber"))
    source_url = (
        first_url_from_keys(item, ("itemUrl", "itemPageUrl", "url"))
        or build_public_item_page_url(store.store_code, item_number or manage_number)
    )
    source_hash_url = f"{source_url}#store={store.id}&manage={quote(manage_number or item_number, safe='')}"
    title = first_text_from_keys(item, ("itemName", "title", "name"))
    normalized = {
        "title": title,
        "source_url": source_url,
        "source_url_hash_key": source_hash_url,
        "image_url": first_rakuten_image_url(item, store.store_code),
        "price": price_from_rakuten_item(item),
        "shop_name": store.store_name,
        "item_number": item_number or manage_number,
        "rakuten_manage_number": manage_number,
        "rakuten_listing_status": rakuten_listing_status_from_item(item),
        "genre_id": first_text_from_keys(item, ("genreId", "genre_id", "genre")),
        "raw": item,
    }
    saved = upsert_product(
        session,
        owner_username,
        None,
        normalized,
        review_status="listed",
        store_id=store.id,
        active_words=active_words,
    )
    if saved and manage_number:
        row = session.scalar(
            select(ProductModel).where(
                ProductModel.store_id == store.id,
                ProductModel.rakuten_manage_number == manage_number,
                ProductModel.review_status == "listed",
            )
        )
        if row is not None:
            ensure_product_listed_store_mark_from_store_product(session, row, store)
    return saved


def upsert_product(
    session: Any,
    owner_username: str,
    task_id: str | None,
    item: dict[str, Any],
    *,
    review_status: str = "pending",
    store_id: int | None = None,
    active_words: list[str] | None = None,
    prepared_item: PreparedProductUpsertItem | None = None,
    scheduled_crawl_id: int | None = None,
) -> bool:
    prepared = prepared_item or prepare_product_upsert_item(session, item, active_words=active_words)
    if prepared.error or not prepared.source_url or not prepared.title:
        return False
    source_url_hash = make_source_url_hash(prepared.source_url_hash_key)
    row = None
    if store_id is not None and prepared.rakuten_manage_number:
        row = session.scalar(
            select(ProductModel).where(
                ProductModel.store_id == store_id,
                ProductModel.rakuten_manage_number == prepared.rakuten_manage_number,
            )
        )
    if row is None and store_id is None:
        row = session.scalar(
            select(ProductModel).where(
                ProductModel.owner_username == owner_username,
                ProductModel.source_url_hash == source_url_hash,
            )
        )
    if row is None:
        row = ProductModel(owner_username=owner_username, source_url=prepared.source_url, source_url_hash=source_url_hash)
        session.add(row)
    elif store_id is None and row.store_id is not None and row.review_status == "listed":
        return False
    row.source_url = prepared.source_url
    row.source_url_hash = source_url_hash
    row.task_id = task_id
    if scheduled_crawl_id is not None:
        row.scheduled_crawl_id = scheduled_crawl_id
    row.collection_source = "scheduled" if scheduled_crawl_id is not None else "manual"
    row.store_id = store_id
    row.rakuten_manage_number = prepared.rakuten_manage_number
    row.rakuten_listing_status = str(prepared.item.get("rakuten_listing_status") or row.rakuten_listing_status or "")
    row.title = prepared.title[:500]
    row.image_url = str(prepared.item.get("image_url") or "")
    row.item_number = str(prepared.item.get("item_number") or "")
    row.shop_name = str(prepared.item.get("shop_name") or "")
    row.genre_id = str(prepared.item.get("genre_id") or "")
    price = prepared.item.get("price")
    row.price = Decimal(str(price)) if price is not None else None
    row.currency = "JPY"
    row.review_status = review_status
    if store_id is not None and review_status == "listed":
        row.store_product_status = "active"
        row.store_last_seen_at = datetime.now()
    raw_payload = prepared.item.get("raw") or prepared.item
    if isinstance(raw_payload, dict) and "reviewCount" in raw_payload:
        row.review_count = product_review_count(raw_payload)
    row.listed_at = parse_rakuten_datetime_value(raw_payload.get("created") if isinstance(raw_payload, dict) else None) or row.listed_at
    row.raw_payload_json = json.dumps(raw_payload, ensure_ascii=False)
    row.last_error = None
    return True


def prepare_product_upsert_item(
    session: Any,
    item: dict[str, Any],
    *,
    active_words: list[str] | None = None,
) -> PreparedProductUpsertItem:
    original_title = str(item.get("title") or "").strip()
    words = active_words if active_words is not None else active_sensitive_words(session)
    cleaned_item = item
    if words:
        cleaned_item, _ = sanitize_product_payload(item, words)
    source_url = str(cleaned_item.get("source_url") or "").strip()
    title = str(cleaned_item.get("title") or "").strip()
    source_url_hash_key = str(cleaned_item.get("source_url_hash_key") or source_url).strip()
    rakuten_manage_number = str(cleaned_item.get("rakuten_manage_number") or "").strip() or None
    error = EMPTY_SENSITIVE_TITLE_SAVE_ERROR if original_title and not title else ""
    return PreparedProductUpsertItem(
        item=cleaned_item,
        source_url=source_url,
        title=title,
        source_url_hash_key=source_url_hash_key,
        rakuten_manage_number=rakuten_manage_number,
        error=error,
    )
