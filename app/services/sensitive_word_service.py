from __future__ import annotations

import json
import re
from io import BytesIO
from zipfile import BadZipFile
from collections.abc import Iterable
from typing import Any
from xml.etree.ElementTree import ParseError

from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError

from app.db.database import SessionLocal
from app.db.models import ProductModel, SensitiveWordModel

BRACKET_RULE = "【】"
BRACKET_SEGMENT_RE = re.compile(r"【[^】]*】")
WHITESPACE_RE = re.compile(r"\s+")
MAX_PAGE_SIZE = 500
SENSITIVE_WORD_TEMPLATE_SHEET_NAME = "敏感词导入"
SENSITIVE_WORD_TEMPLATE_HEADER = "敏感词"
DEFAULT_SENSITIVE_WORDS: tuple[str, ...] = (
    "500円OFFクーポン",
    "【全店2点購入で10％OFF】",
    "300円OFFクーポン",
    "【お買い物マラソン 当店ポイント5倍】",
    "翌日出荷",
    "翌日配達",
    "楽天1位",
    "楽天2位",
    "楽天3位",
    "【】",
    "【P10&最大600円OFF】",
    "【楽天倉庫出荷】",
    "【日本国内発送】",
    "【楽天1位】",
    "【楽天2位】",
    "【楽天3位】",
    "新生活",
    "即納",
    "【10%OFFクーポン】",
    "【8%OFFクーポン】",
    "【期間限定5％OFF】",
    "一部即納",
    "【P5倍期間限定】",
    "【P10倍期間限定】",
    "お買い物マラソン",
    "【P5倍】",
    "【LINE追加で5%OFF】",
    "【限定10%OFF】",
    "【スーパーDEAL&お買い物マラソンP10】",
    "【お買い物マラソン最大2000円OFF】",
    "＼7%OFFクーポン利用可2.18まで／",
    "【短納期】",
)
RECURSIVE_SANITIZE_FIELD_NAMES = {
    "title",
    "itemName",
    "tagline",
    "subtitle",
    "subTitle",
    "catchCopy",
    "catchcopy",
    "catchCopyTrans",
    "saleComment",
    "sale_comment",
}
ROOT_ONLY_SANITIZE_FIELD_NAMES = {
    "name",
}


def normalize_sensitive_word(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sanitize_sensitive_text(value: Any, words: Iterable[str]) -> str:
    text = normalize_sensitive_word(value)
    if not text:
        return ""

    bracket_rule_enabled, normalized_words = _prepare_sensitive_words(words)

    if bracket_rule_enabled:
        text = BRACKET_SEGMENT_RE.sub("", text)

    for word in normalized_words:
        text = text.replace(word, "")

    return WHITESPACE_RE.sub(" ", text).strip()


def sanitize_product_payload(payload: dict[str, Any], words: Iterable[str]) -> tuple[dict[str, Any], bool]:
    bracket_rule_enabled, normalized_words = _prepare_sensitive_words(words)
    return _sanitize_product_payload_with_prepared_words(payload, normalized_words, bracket_rule_enabled)


def seed_default_sensitive_words(session: Any) -> int:
    existing_words = {
        normalize_sensitive_word(word)
        for word in session.scalars(select(SensitiveWordModel.word)).all()
        if normalize_sensitive_word(word)
    }
    created_count = 0
    for word in DEFAULT_SENSITIVE_WORDS:
        normalized = normalize_sensitive_word(word)
        if not normalized or normalized in existing_words:
            continue
        session.add(SensitiveWordModel(word=normalized, enabled=True))
        existing_words.add(normalized)
        created_count += 1
    if created_count:
        session.flush()
    return created_count


def list_sensitive_words(page: int, page_size: int, keyword: str = "") -> dict[str, Any]:
    normalized_page, normalized_page_size = _normalize_page_params(page, page_size)
    normalized_keyword = normalize_sensitive_word(keyword)

    with SessionLocal() as session:
        query = select(SensitiveWordModel)
        if normalized_keyword:
            query = query.where(SensitiveWordModel.word.like(f"%{normalized_keyword}%"))
        total = int(session.scalar(select(func.count()).select_from(query.order_by(None).subquery())) or 0)
        if total:
            max_page = max(1, (total + normalized_page_size - 1) // normalized_page_size)
            normalized_page = min(normalized_page, max_page)
        rows = session.scalars(
            query.order_by(SensitiveWordModel.created_at.asc(), SensitiveWordModel.id.asc())
            .offset((normalized_page - 1) * normalized_page_size)
            .limit(normalized_page_size)
        ).all()
        return {
            "items": [_sensitive_word_to_public(row) for row in rows],
            "total": total,
            "page": normalized_page,
            "pageSize": normalized_page_size,
        }


def create_sensitive_word(word: str, enabled: bool = True) -> dict[str, Any]:
    normalized_word = normalize_sensitive_word(word)
    if not normalized_word:
        raise RuntimeError("敏感词不能为空。")

    with SessionLocal() as session:
        row = SensitiveWordModel(word=normalized_word, enabled=bool(enabled))
        session.add(row)
        try:
            session.commit()
        except IntegrityError as exc:
            session.rollback()
            raise RuntimeError("敏感词已存在。") from exc
        session.refresh(row)
        return _sensitive_word_to_public(row)


def update_sensitive_word(word_id: int, word: str, enabled: bool) -> dict[str, Any]:
    normalized_word = normalize_sensitive_word(word)
    if not normalized_word:
        raise RuntimeError("敏感词不能为空。")

    with SessionLocal() as session:
        row = session.get(SensitiveWordModel, int(word_id))
        if row is None:
            raise RuntimeError("敏感词不存在。")
        row.word = normalized_word
        row.enabled = bool(enabled)
        try:
            session.commit()
        except IntegrityError as exc:
            session.rollback()
            raise RuntimeError("敏感词已存在。") from exc
        session.refresh(row)
        return _sensitive_word_to_public(row)


def delete_sensitive_word(word_id: int) -> bool:
    with SessionLocal() as session:
        row = session.get(SensitiveWordModel, int(word_id))
        if row is None:
            return False
        session.delete(row)
        session.commit()
        return True


def active_sensitive_words(session: Any) -> list[str]:
    return [
        normalize_sensitive_word(word)
        for word in session.scalars(
            select(SensitiveWordModel.word)
            .where(SensitiveWordModel.enabled.is_(True))
            .order_by(func.length(SensitiveWordModel.word).desc(), SensitiveWordModel.word.asc(), SensitiveWordModel.id.asc())
        ).all()
        if normalize_sensitive_word(word)
    ]


def build_sensitive_word_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法生成敏感词导入模板。") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SENSITIVE_WORD_TEMPLATE_SHEET_NAME
    sheet.append([SENSITIVE_WORD_TEMPLATE_HEADER])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_sensitive_words(content: bytes, filename: str) -> dict[str, int]:
    normalized_filename = normalize_sensitive_word(filename).lower()
    if not content:
        raise RuntimeError("导入文件为空。")
    if not normalized_filename.endswith(".xlsx"):
        raise RuntimeError("敏感词导入只支持 .xlsx 文件。")

    workbook = _load_sensitive_word_workbook(content)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("导入文件没有内容。")

    header = _sensitive_word_import_header_key(rows[0][0] if rows[0] else "")
    if header != SENSITIVE_WORD_TEMPLATE_HEADER:
        raise RuntimeError("导入文件表头必须包含：敏感词。")

    created_count = 0
    duplicate_count = 0
    invalid_count = 0

    with SessionLocal() as session:
        existing_words = {
            normalize_sensitive_word(word)
            for word in session.scalars(select(SensitiveWordModel.word)).all()
            if normalize_sensitive_word(word)
        }

        for row in rows[1:]:
            raw_value = row[0] if row else ""
            normalized_word = normalize_sensitive_word(raw_value)
            if not normalized_word:
                invalid_count += 1
                continue
            if normalized_word in existing_words:
                duplicate_count += 1
                continue
            try:
                with session.begin_nested():
                    session.add(SensitiveWordModel(word=normalized_word, enabled=True))
                    session.flush()
            except IntegrityError as exc:
                if _is_sensitive_word_unique_conflict(exc):
                    duplicate_count += 1
                    existing_words.add(normalized_word)
                    continue
                raise
            existing_words.add(normalized_word)
            created_count += 1

        if created_count:
            session.commit()

    return {
        "createdCount": created_count,
        "duplicateCount": duplicate_count,
        "invalidCount": invalid_count,
    }


def cleanup_pending_products(session: Any, *, apply: bool = False) -> dict[str, int]:
    active_words = active_sensitive_words(session)
    bracket_rule_enabled, literal_words = _prepare_sensitive_words(active_words)
    scanned_count = 0
    matched_count = 0
    updated_count = 0
    empty_title_count = 0

    pending_products = session.scalars(
        select(ProductModel)
        .where(ProductModel.review_status == "pending")
        .order_by(ProductModel.id.asc())
    ).all()

    for product in pending_products:
        scanned_count += 1

        cleaned_title = _sanitize_text_with_prepared_words(product.title, literal_words, bracket_rule_enabled)
        title_changed = cleaned_title != product.title

        cleaned_payload = None
        payload_changed = False
        raw_payload, payload_supported = _load_cleanup_payload(product.raw_payload_json)
        if payload_supported:
            cleaned_payload, payload_changed = _sanitize_product_payload_with_prepared_words(
                raw_payload,
                literal_words,
                bracket_rule_enabled,
            )
        elif title_changed:
            matched_count += 1
            if not cleaned_title:
                empty_title_count += 1
            continue

        if not title_changed and not payload_changed:
            continue

        matched_count += 1
        if not cleaned_title:
            empty_title_count += 1
            continue

        updated_count += 1
        if apply:
            product.title = cleaned_title
            if payload_changed and cleaned_payload is not None:
                product.raw_payload_json = json.dumps(cleaned_payload, ensure_ascii=False)

    return {
        "scannedCount": scanned_count,
        "matchedCount": matched_count,
        "updatedCount": updated_count,
        "emptyTitleCount": empty_title_count,
    }


def _prepare_sensitive_words(words: Iterable[str]) -> tuple[bool, list[str]]:
    bracket_rule_enabled = False
    literal_words: set[str] = set()

    for word in words:
        normalized = normalize_sensitive_word(word)
        if not normalized:
            continue
        if normalized == BRACKET_RULE:
            bracket_rule_enabled = True
            continue
        literal_words.add(normalized)

    return bracket_rule_enabled, sorted(literal_words, key=lambda item: (-len(item), item))


def _normalize_page_params(page: int | None, page_size: int | None) -> tuple[int, int]:
    normalized_page = max(1, int(page or 1))
    normalized_page_size = min(MAX_PAGE_SIZE, max(1, int(page_size or 1)))
    return normalized_page, normalized_page_size


def _rule_type_for_word(word: str) -> str:
    return "bracket" if word == BRACKET_RULE else "literal"


def _sensitive_word_to_public(row: SensitiveWordModel) -> dict[str, Any]:
    return {
        "id": row.id,
        "word": row.word,
        "ruleType": _rule_type_for_word(row.word),
        "enabled": bool(row.enabled),
        "createdAt": row.created_at.isoformat() if row.created_at else "",
        "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
    }


def _sanitize_text_with_prepared_words(value: Any, literal_words: list[str], bracket_rule_enabled: bool) -> str:
    text = normalize_sensitive_word(value)
    if not text:
        return ""

    if bracket_rule_enabled:
        text = BRACKET_SEGMENT_RE.sub("", text)

    for word in literal_words:
        text = text.replace(word, "")

    return WHITESPACE_RE.sub(" ", text).strip()


def _sanitize_product_payload_with_prepared_words(
    payload: dict[str, Any],
    literal_words: list[str],
    bracket_rule_enabled: bool,
) -> tuple[dict[str, Any], bool]:
    cleaned, changed = _sanitize_payload_node(payload, literal_words, bracket_rule_enabled)

    for key in ROOT_ONLY_SANITIZE_FIELD_NAMES:
        if key not in cleaned:
            continue
        value = cleaned.get(key)
        if isinstance(value, (dict, list)):
            continue
        sanitized = _sanitize_text_with_prepared_words(value, literal_words, bracket_rule_enabled)
        if sanitized != value:
            cleaned[key] = sanitized
            changed = True

    return cleaned, changed


def _load_cleanup_payload(raw_payload_json: str) -> tuple[dict[str, Any], bool]:
    normalized_payload_json = normalize_sensitive_word(raw_payload_json)
    if not normalized_payload_json:
        return {}, True
    try:
        payload = json.loads(normalized_payload_json)
    except json.JSONDecodeError:
        return {}, False
    if not isinstance(payload, dict):
        return {}, False
    return payload, True


def _load_sensitive_word_workbook(content: bytes) -> Any:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法读取 xlsx 文件。") from exc

    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, KeyError, OSError, ParseError, ValueError, InvalidFileException) as exc:
        raise RuntimeError("无法读取敏感词导入文件，请确认文件是有效的 .xlsx 格式。") from exc


def _sensitive_word_import_header_key(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_sensitive_word(value))


def _is_sensitive_word_unique_conflict(exc: IntegrityError) -> bool:
    orig = getattr(exc, "orig", None)
    if getattr(orig, "pgcode", None) == "23505":
        return True

    orig_args = getattr(orig, "args", ())
    if orig_args and orig_args[0] == 1062:
        return True

    message = " ".join(
        part for part in (
            str(orig).lower() if orig is not None else "",
            str(exc).lower(),
        ) if part
    )
    return any(
        token in message
        for token in (
            "unique constraint",
            "duplicate entry",
            "uq_lt_sensitive_word",
            "lt_sensitive_words.word",
        )
    )


def _sanitize_payload_node(value: Any, words: list[str], bracket_rule_enabled: bool) -> tuple[Any, bool]:
    if isinstance(value, dict):
        changed = False
        cleaned: dict[str, Any] = {}
        for key, item in value.items():
            if key in RECURSIVE_SANITIZE_FIELD_NAMES and not isinstance(item, (dict, list)):
                sanitized = _sanitize_text_with_prepared_words(item, words, bracket_rule_enabled)
                cleaned[key] = sanitized
                changed = changed or sanitized != item
                continue
            cleaned_item, item_changed = _sanitize_payload_node(item, words, bracket_rule_enabled)
            cleaned[key] = cleaned_item
            changed = changed or item_changed
        return cleaned, changed

    if isinstance(value, list):
        changed = False
        cleaned_items: list[Any] = []
        for item in value:
            cleaned_item, item_changed = _sanitize_payload_node(item, words, bracket_rule_enabled)
            cleaned_items.append(cleaned_item)
            changed = changed or item_changed
        return cleaned_items, changed

    return value, False
