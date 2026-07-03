from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


SKIPPED_SHEETS = {"はじめに", "付則", "ジャンル一覧", "推奨値シート"}


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "-"} else text


def raw_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_genres(workbook: Any, source_name: str) -> dict[str, dict[str, str]]:
    if "ジャンル一覧" not in workbook.sheetnames:
        return {}
    sheet = workbook["ジャンル一覧"]
    genres: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        genre_id = cell_text(row[0] if len(row) > 0 else None)
        genre_path = cell_text(row[1] if len(row) > 1 else None)
        group_name = cell_text(row[2] if len(row) > 2 else None)
        if not genre_id or not group_name:
            continue
        genres[genre_id] = {
            "genrePath": genre_path,
            "group": group_name,
            "groupKey": f"{source_name}::{group_name}",
            "source": source_name,
        }
    return genres


def parse_group_sheet(sheet: Any, source_name: str) -> tuple[str, dict[str, Any]] | None:
    group_name = cell_text(sheet.cell(row=1, column=2).value) or sheet.title
    attributes: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if not row or not cell_text(row[0] if len(row) > 0 else None):
            continue
        name = cell_text(row[1] if len(row) > 1 else None)
        if not name:
            continue
        required_text = raw_text(row[2] if len(row) > 2 else None)
        unit_required_text = raw_text(row[8] if len(row) > 8 else None)
        required = required_text == "必須"
        unit_required = unit_required_text not in {"", "-"}
        if not required and not unit_required:
            continue
        recommended_unit = cell_text(row[9] if len(row) > 9 else None)
        multiple_text = raw_text(row[10] if len(row) > 10 else None)
        delimiter = cell_text(row[12] if len(row) > 12 else None)
        example = cell_text(row[14] if len(row) > 14 else None)
        attributes[name] = {
            "name": name,
            "required": required,
            "requirement": required_text,
            "inputMethod": cell_text(row[3] if len(row) > 3 else None),
            "hasRecommendedValues": cell_text(row[4] if len(row) > 4 else None),
            "format": cell_text(row[5] if len(row) > 5 else None),
            "maxLength": cell_text(row[6] if len(row) > 6 else None),
            "unitRequired": unit_required,
            "unit": recommended_unit,
            "multiple": multiple_text == "可",
            "maxValues": cell_text(row[11] if len(row) > 11 else None),
            "delimiter": delimiter,
            "example": example,
            "sameValueTarget": cell_text(row[15] if len(row) > 15 else None),
        }
    if not attributes:
        return None
    return group_name, {
        "source": source_name,
        "group": group_name,
        "attributes": attributes,
    }


def parse_recommended_values(workbook: Any) -> dict[str, dict[str, list[str]]]:
    if "推奨値シート" not in workbook.sheetnames:
        return {}
    sheet = workbook["推奨値シート"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}
    recommended: dict[str, dict[str, list[str]]] = {}
    group_row = rows[0]
    name_row = rows[1]
    max_column = max((len(row) for row in rows), default=0)
    for column_index in range(1, max_column):
        group_name = cell_text(group_row[column_index] if column_index < len(group_row) else None) or "*"
        attribute_name = cell_text(name_row[column_index] if column_index < len(name_row) else None)
        if not attribute_name:
            continue
        values: list[str] = []
        for row in rows[2:]:
            value = cell_text(row[column_index] if column_index < len(row) else None)
            if value and value not in values:
                values.append(value)
        if values:
            recommended.setdefault(group_name, {})[attribute_name] = values
    return recommended


def merge_recommended_values(
    target: dict[str, dict[str, list[str]]],
    source_name: str,
    group_rules: dict[str, dict[str, Any]],
    recommended: dict[str, dict[str, list[str]]],
) -> None:
    for group_name, values_by_attribute in recommended.items():
        keys: list[str]
        if group_name == "*":
            keys = [key for key, rule in group_rules.items() if rule.get("source") == source_name]
        else:
            keys = [f"{source_name}::{group_name}"]
        for key in keys:
            for attribute_name, values in values_by_attribute.items():
                target.setdefault(key, {})[attribute_name] = values


def build_rules(source_dir: Path, *, include_recommended_values: bool = False) -> dict[str, Any]:
    genres: dict[str, dict[str, str]] = {}
    group_rules: dict[str, dict[str, Any]] = {}
    recommended_values: dict[str, dict[str, list[str]]] = {}
    files = sorted(source_dir.glob("*.xlsx"))
    for path in files:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        source_name = path.name
        genres.update(parse_genres(workbook, source_name))
        for sheet in workbook.worksheets:
            if sheet.title in SKIPPED_SHEETS:
                continue
            parsed = parse_group_sheet(sheet, source_name)
            if not parsed:
                continue
            group_name, group_rule = parsed
            group_rules[f"{source_name}::{group_name}"] = group_rule
        if include_recommended_values:
            merge_recommended_values(
                recommended_values,
                source_name,
                group_rules,
                parse_recommended_values(workbook),
            )
    for group_key, group_recommended in recommended_values.items():
        rule = group_rules.get(group_key)
        if not rule:
            continue
        for attribute_name, values in group_recommended.items():
            attribute = rule["attributes"].get(attribute_name)
            if attribute:
                attribute["recommendedValues"] = values
    return {
        "schemaVersion": 1,
        "generatedAt": datetime.now().isoformat(sep=" ", timespec="seconds"),
        "sourceDir": source_dir.name,
        "sourceFileCount": len(files),
        "genreCount": len(genres),
        "groupCount": len(group_rules),
        "genres": genres,
        "groups": group_rules,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Rakuten RMS product attribute definition workbooks.")
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("--output", type=Path, default=Path("app/resources/rakuten_attribute_rules.json"))
    parser.add_argument("--include-recommended-values", action="store_true")
    args = parser.parse_args()
    rules = build_rules(args.source_dir, include_recommended_values=args.include_recommended_values)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(rules, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {args.output} "
        f"({rules['sourceFileCount']} files, {rules['genreCount']} genres, {rules['groupCount']} groups)"
    )


if __name__ == "__main__":
    main()
