from __future__ import annotations

from datetime import datetime
from io import BytesIO
import tempfile
import unittest
from contextlib import contextmanager
from unittest.mock import Mock, patch
from zipfile import ZipFile
import json

from sqlalchemy import event
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker

from app.db.database import Base


class SensitiveWordModelTests(unittest.TestCase):
    def test_sensitive_word_model_has_expected_schema(self) -> None:
        from app.services.sensitive_word_service import normalize_sensitive_word
        from app.db.models import SensitiveWordModel

        id_column = SensitiveWordModel.__table__.columns["id"]
        enabled_column = SensitiveWordModel.__table__.columns["enabled"]
        created_at_column = SensitiveWordModel.__table__.columns["created_at"]
        updated_at_column = SensitiveWordModel.__table__.columns["updated_at"]

        self.assertEqual(normalize_sensitive_word("  即納  "), "即納")
        self.assertEqual(SensitiveWordModel.__tablename__, "lt_sensitive_words")
        self.assertIn("id", SensitiveWordModel.__table__.columns)
        self.assertIn("word", SensitiveWordModel.__table__.columns)
        self.assertIn("enabled", SensitiveWordModel.__table__.columns)
        self.assertIn("created_at", SensitiveWordModel.__table__.columns)
        self.assertIn("updated_at", SensitiveWordModel.__table__.columns)
        self.assertTrue(id_column.primary_key)
        self.assertTrue(id_column.autoincrement)
        self.assertEqual(SensitiveWordModel.__table__.columns["word"].type.length, 500)
        self.assertFalse(SensitiveWordModel.__table__.columns["word"].nullable)
        self.assertFalse(enabled_column.nullable)
        self.assertIsNotNone(enabled_column.default)
        self.assertTrue(enabled_column.default.arg)
        self.assertIsNotNone(enabled_column.server_default)
        self.assertEqual(str(enabled_column.server_default.arg), "1")
        self.assertFalse(created_at_column.nullable)
        self.assertIsNotNone(created_at_column.server_default)
        self.assertFalse(updated_at_column.nullable)
        self.assertIsNotNone(updated_at_column.server_default)
        self.assertIsNotNone(updated_at_column.onupdate)
        self.assertTrue(
            any(constraint.name == "uq_lt_sensitive_word" for constraint in SensitiveWordModel.__table__.constraints)
        )


class SensitiveWordSanitizerTests(unittest.TestCase):
    def test_removes_normal_words_and_collapses_whitespace(self) -> None:
        from app.services.sensitive_word_service import sanitize_sensitive_text

        self.assertEqual(
            sanitize_sensitive_text(
                "楽天1位  春物  即納",
                ["楽天1位", "即納"],
            ),
            "春物",
        )

    def test_empty_bracket_rule_removes_every_bracketed_segment(self) -> None:
        from app.services.sensitive_word_service import sanitize_sensitive_text

        self.assertEqual(
            sanitize_sensitive_text(
                "【楽天1位】【日本国内発送】 春物",
                ["【】"],
            ),
            "春物",
        )

    def test_overlapping_literal_words_are_deduplicated_and_applied_longest_first(self) -> None:
        from app.services.sensitive_word_service import sanitize_sensitive_text

        self.assertEqual(
            sanitize_sensitive_text(
                "期間限定",
                [" 限定 ", "期間限定", "限定", "期間限定"],
            ),
            "",
        )

    def test_payload_sanitizer_updates_title_and_tagline_fields_recursively(self) -> None:
        from app.services.sensitive_word_service import sanitize_product_payload

        payload = {
            "title": "【楽天1位】 春物",
            "itemName": "【楽天1位】 春物",
            "tagline": "即納 おすすめ",
            "item": {"subtitle": "【期間限定】 即納"},
            "description": "【期間限定】 即納",
        }

        cleaned, changed = sanitize_product_payload(payload, ["【】", "即納"])

        self.assertTrue(changed)
        self.assertEqual(cleaned["title"], "春物")
        self.assertEqual(cleaned["itemName"], "春物")
        self.assertEqual(cleaned["tagline"], "おすすめ")
        self.assertEqual(cleaned["item"]["subtitle"], "")
        self.assertEqual(cleaned["description"], "【期間限定】 即納")

    def test_payload_sanitizer_only_cleans_root_name_but_keeps_nested_generic_names(self) -> None:
        from app.services.sensitive_word_service import sanitize_product_payload

        payload = {
            "name": "即納 春物",
            "metadata": {
                "name": "即納 サイズ名",
                "tagline": "即納 おすすめ",
            },
            "attributes": [
                {"name": "即納 カラー", "value": "赤"},
                {"name": "即納 素材", "value": "綿"},
            ],
            "item": {
                "itemName": "即納 ワンピース",
                "subtitle": "即納 人気",
            },
        }

        cleaned, changed = sanitize_product_payload(payload, ["即納"])

        self.assertTrue(changed)
        self.assertEqual(cleaned["name"], "春物")
        self.assertEqual(cleaned["metadata"]["name"], "即納 サイズ名")
        self.assertEqual(cleaned["attributes"][0]["name"], "即納 カラー")
        self.assertEqual(cleaned["attributes"][1]["name"], "即納 素材")
        self.assertEqual(cleaned["metadata"]["tagline"], "おすすめ")
        self.assertEqual(cleaned["item"]["itemName"], "ワンピース")
        self.assertEqual(cleaned["item"]["subtitle"], "人気")


class SensitiveWordDatabaseTestCase(unittest.TestCase):
    def setUp(self) -> None:
        import app.db.models  # noqa: F401

        self.engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
        Base.metadata.create_all(self.engine)
        self.session_factory = sessionmaker(
            bind=self.engine,
            expire_on_commit=False,
            future=True,
        )

    def tearDown(self) -> None:
        self.engine.dispose()

    @contextmanager
    def session_scope(self):
        session = self.session_factory()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    def list_words(self) -> list[str]:
        from app.db.models import SensitiveWordModel

        with Session(self.engine, future=True) as session:
            return session.scalars(
                select(SensitiveWordModel.word).order_by(SensitiveWordModel.id.asc())
            ).all()


class SensitiveWordPersistenceTests(SensitiveWordDatabaseTestCase):
    def test_seed_is_idempotent_and_deduplicates_default_words(self) -> None:
        from app.db.models import SensitiveWordModel
        from app.services.sensitive_word_service import DEFAULT_SENSITIVE_WORDS, seed_default_sensitive_words

        expected_default_words = (
            "500円OFFクーポン",
            "【全店2点購入で10％OFF】",
            "300円OFFクーポン",
            "【お買い物マラソン 当店ポイント5倍】",
            "翌日出荷",
            "翌日配達",
            "楽天1位",
            "楽天2位",
            "楽天3位",
            "【】",
            "【P10&最大600円OFF】",
            "【楽天倉庫出荷】",
            "【日本国内発送】",
            "【楽天1位】",
            "【楽天2位】",
            "【楽天3位】",
            "新生活",
            "即納",
            "【10%OFFクーポン】",
            "【8%OFFクーポン】",
            "【期間限定5％OFF】",
            "一部即納",
            "【P5倍期間限定】",
            "【P10倍期間限定】",
            "お買い物マラソン",
            "【P5倍】",
            "【LINE追加で5%OFF】",
            "【限定10%OFF】",
            "【スーパーDEAL&お買い物マラソンP10】",
            "【お買い物マラソン最大2000円OFF】",
            "＼7%OFFクーポン利用可2.18まで／",
            "【短納期】",
        )

        self.assertEqual(DEFAULT_SENSITIVE_WORDS, expected_default_words)

        with self.session_scope() as session:
            created_count = seed_default_sensitive_words(session)
            self.assertGreater(created_count, 0)

        with self.session_scope() as session:
            self.assertEqual(seed_default_sensitive_words(session), 0)
            total = session.scalar(select(func.count()).select_from(SensitiveWordModel))

        self.assertEqual(total, len(set(DEFAULT_SENSITIVE_WORDS)))
        self.assertEqual(len(DEFAULT_SENSITIVE_WORDS), len(set(DEFAULT_SENSITIVE_WORDS)))
        self.assertEqual(list(self.list_words()), list(expected_default_words))
        self.assertIn("【】", DEFAULT_SENSITIVE_WORDS)
        self.assertIn("即納", DEFAULT_SENSITIVE_WORDS)
        self.assertIn("楽天1位", DEFAULT_SENSITIVE_WORDS)
        self.assertIn("翌日配達", DEFAULT_SENSITIVE_WORDS)

    def test_init_database_seeds_defaults_without_breaking_existing_bootstrap_steps(self) -> None:
        import app.db.database as database_module
        from app.db.models import SensitiveWordModel

        with (
            patch.object(database_module, "engine", self.engine),
            patch.object(database_module, "SessionLocal", self.session_factory),
            patch.object(database_module, "ensure_mysql_database_exists") as ensure_database_exists,
            patch.object(database_module, "ensure_schema_compatibility") as ensure_schema_compatibility,
            patch.object(database_module.settings, "database_auto_create", False),
            patch("app.services.user_service.ensure_initial_superadmin") as ensure_initial_superadmin,
            patch("app.services.crawler_service.ensure_default_roles") as ensure_default_roles,
        ):
            database_module.init_database()

        ensure_database_exists.assert_not_called()
        ensure_schema_compatibility.assert_called_once_with()
        ensure_initial_superadmin.assert_called_once_with()
        ensure_default_roles.assert_called_once_with()

        with Session(self.engine, future=True) as session:
            total = session.scalar(select(func.count()).select_from(SensitiveWordModel))
            self.assertEqual(total, session.scalar(select(func.count()).select_from(SensitiveWordModel)))
            self.assertGreater(total or 0, 0)

        with (
            patch.object(database_module, "engine", self.engine),
            patch.object(database_module, "SessionLocal", self.session_factory),
            patch.object(database_module, "ensure_mysql_database_exists"),
            patch.object(database_module, "ensure_schema_compatibility"),
            patch.object(database_module.settings, "database_auto_create", False),
            patch("app.services.user_service.ensure_initial_superadmin"),
            patch("app.services.crawler_service.ensure_default_roles"),
        ):
            database_module.init_database()

        with Session(self.engine, future=True) as session:
            second_total = session.scalar(select(func.count()).select_from(SensitiveWordModel))

        self.assertEqual(second_total, total)

    def test_crud_normalizes_words_and_rejects_duplicates(self) -> None:
        from app.services import sensitive_word_service

        with patch.object(sensitive_word_service, "SessionLocal", self.session_factory):
            created = sensitive_word_service.create_sensitive_word("  即納  ")

            self.assertEqual(created["word"], "即納")
            self.assertEqual(created["ruleType"], "literal")
            self.assertTrue(created["enabled"])

            with self.assertRaisesRegex(RuntimeError, "已存在"):
                sensitive_word_service.create_sensitive_word("即納")

            with self.assertRaisesRegex(RuntimeError, "不能为空"):
                sensitive_word_service.create_sensitive_word("   ")

    def test_list_filters_and_paginates_sensitive_words(self) -> None:
        from app.services import sensitive_word_service

        with patch.object(sensitive_word_service, "SessionLocal", self.session_factory):
            sensitive_word_service.create_sensitive_word("翌日配達")
            sensitive_word_service.create_sensitive_word("即納")
            sensitive_word_service.create_sensitive_word("【】")

            page_one = sensitive_word_service.list_sensitive_words(page=1, page_size=2)
            filtered = sensitive_word_service.list_sensitive_words(page=1, page_size=10, keyword="即")

        self.assertEqual(page_one["total"], 3)
        self.assertEqual(page_one["page"], 1)
        self.assertEqual(page_one["pageSize"], 2)
        self.assertEqual(len(page_one["items"]), 2)
        self.assertEqual([item["word"] for item in page_one["items"]], ["翌日配達", "即納"])
        self.assertEqual(filtered["total"], 1)
        self.assertEqual(filtered["items"][0]["word"], "即納")
        self.assertEqual(filtered["items"][0]["ruleType"], "literal")

    def test_update_delete_and_active_word_ordering_respect_enabled_state(self) -> None:
        from app.db.models import SensitiveWordModel
        from app.services import sensitive_word_service

        with patch.object(sensitive_word_service, "SessionLocal", self.session_factory):
            created_bracket = sensitive_word_service.create_sensitive_word("【】")
            created_short = sensitive_word_service.create_sensitive_word("即納")
            created_long = sensitive_word_service.create_sensitive_word("期間限定")

            updated = sensitive_word_service.update_sensitive_word(created_short["id"], "  翌日配達  ", False)
            self.assertEqual(updated["word"], "翌日配達")
            self.assertFalse(updated["enabled"])

            with self.assertRaisesRegex(RuntimeError, "已存在"):
                sensitive_word_service.update_sensitive_word(created_long["id"], "【】", True)

            self.assertTrue(sensitive_word_service.delete_sensitive_word(created_bracket["id"]))
            self.assertFalse(sensitive_word_service.delete_sensitive_word(created_bracket["id"]))

        with Session(self.engine, future=True) as session:
            active_words = sensitive_word_service.active_sensitive_words(session)
            rows = session.scalars(select(SensitiveWordModel).order_by(SensitiveWordModel.id.asc())).all()

        self.assertEqual(active_words, ["期間限定"])
        self.assertEqual([row.word for row in rows], ["翌日配達", "期間限定"])
        self.assertFalse(rows[0].enabled)
        self.assertTrue(rows[1].enabled)


class SensitiveWordExcelImportTests(SensitiveWordDatabaseTestCase):
    def test_template_contains_expected_sheet_and_header(self) -> None:
        from openpyxl import load_workbook

        from app.services.sensitive_word_service import build_sensitive_word_template

        content = build_sensitive_word_template()
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames, ["敏感词导入"])
        sheet = workbook["敏感词导入"]
        self.assertEqual(sheet["A1"].value, "敏感词")

    def test_import_counts_created_duplicate_and_invalid_rows(self) -> None:
        from openpyxl import Workbook

        from app.services import sensitive_word_service

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "敏感词导入"
        sheet.append(["敏感词"])
        sheet.append(["春物"])
        sheet.append(["即納"])
        sheet.append(["春物"])
        sheet.append(["   "])
        sheet.append(["翌日配達"])
        buffer = BytesIO()
        workbook.save(buffer)

        with patch.object(sensitive_word_service, "SessionLocal", self.session_factory):
            sensitive_word_service.create_sensitive_word("即納")
            result = sensitive_word_service.import_sensitive_words(
                content=buffer.getvalue(),
                filename="sensitive-words.xlsx",
            )

        self.assertEqual(
            result,
            {
                "createdCount": 2,
                "duplicateCount": 2,
                "invalidCount": 1,
            },
        )
        self.assertEqual(self.list_words(), ["即納", "春物", "翌日配達"])

    def test_import_rejects_non_xlsx_files(self) -> None:
        from app.services.sensitive_word_service import import_sensitive_words

        with self.assertRaisesRegex(RuntimeError, r"\.xlsx"):
            import_sensitive_words(content=b"not-an-excel-file", filename="sensitive-words.csv")

    def test_import_rejects_valid_zip_that_is_not_a_real_xlsx_workbook(self) -> None:
        from app.services.sensitive_word_service import import_sensitive_words

        buffer = BytesIO()
        with ZipFile(buffer, "w") as archive:
            archive.writestr("not-a-workbook.txt", "plain text")

        with self.assertRaisesRegex(RuntimeError, r"有效的 \.xlsx"):
            import_sensitive_words(content=buffer.getvalue(), filename="sensitive-words.xlsx")

    def test_import_requires_sensitive_word_header(self) -> None:
        from openpyxl import Workbook

        from app.services.sensitive_word_service import import_sensitive_words

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "敏感词导入"
        sheet.append(["词汇"])
        sheet.append(["春物"])
        buffer = BytesIO()
        workbook.save(buffer)

        with self.assertRaisesRegex(RuntimeError, "表头"):
            import_sensitive_words(content=buffer.getvalue(), filename="sensitive-words.xlsx")

    def test_import_counts_flush_time_uniqueness_race_as_duplicate_and_keeps_valid_rows(self) -> None:
        from openpyxl import Workbook

        from app.db.models import SensitiveWordModel
        from app.services import sensitive_word_service

        race_engine = None
        race_session_factory = None
        inject_competing_insert = None

        with tempfile.TemporaryDirectory() as temp_dir:
            database_path = f"{temp_dir}\\sensitive-word-race.sqlite3"
            race_engine = create_engine(f"sqlite+pysqlite:///{database_path}", future=True)
            try:
                Base.metadata.create_all(race_engine)
                race_session_factory = sessionmaker(
                    bind=race_engine,
                    expire_on_commit=False,
                    future=True,
                )

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "敏感词导入"
                sheet.append(["敏感词"])
                sheet.append(["春物"])
                sheet.append(["競合語"])
                sheet.append(["翌日配達"])
                buffer = BytesIO()
                workbook.save(buffer)

                injected_words: list[str] = []

                @event.listens_for(race_session_factory.class_, "before_flush")
                def inject_competing_insert(session, flush_context, instances) -> None:
                    if not session.info.get("inject_race_once") or session.info.get("race_injected"):
                        return

                    pending_words = {
                        row.word
                        for row in session.new
                        if isinstance(row, SensitiveWordModel)
                    }
                    if "競合語" not in pending_words:
                        return

                    session.info["race_injected"] = True
                    with race_session_factory() as competing_session:
                        competing_session.add(SensitiveWordModel(word="競合語", enabled=True))
                        competing_session.commit()
                    injected_words.append("競合語")

                def flagged_session_local():
                    session = race_session_factory()
                    session.info["inject_race_once"] = True
                    return session

                with patch.object(sensitive_word_service, "SessionLocal", flagged_session_local):
                    result = sensitive_word_service.import_sensitive_words(
                        content=buffer.getvalue(),
                        filename="sensitive-words.xlsx",
                    )

                self.assertEqual(injected_words, ["競合語"])
                self.assertEqual(
                    result,
                    {
                        "createdCount": 2,
                        "duplicateCount": 1,
                        "invalidCount": 0,
                    },
                )

                with Session(race_engine, future=True) as session:
                    words = session.scalars(
                        select(SensitiveWordModel.word).order_by(SensitiveWordModel.id.asc())
                    ).all()

                self.assertEqual(words, ["春物", "競合語", "翌日配達"])
            finally:
                if race_session_factory is not None and inject_competing_insert is not None:
                    event.remove(race_session_factory.class_, "before_flush", inject_competing_insert)
                if race_engine is not None:
                    race_engine.dispose()


class SensitiveWordUpsertTests(SensitiveWordDatabaseTestCase):
    def test_upsert_product_sanitizes_stored_title_and_payload_with_active_words(self) -> None:
        from app.db.models import ProductModel, SensitiveWordModel
        from app.services import crawler_service

        item = {
            "source_url": "https://example.test/item/1",
            "title": "【楽天1位】 即納 春物",
            "raw": {
                "title": "【楽天1位】 即納 春物",
                "itemName": "【楽天1位】 即納 春物",
                "tagline": "翌日配達 おすすめ",
            },
        }

        with self.session_scope() as session:
            session.add_all(
                [
                    SensitiveWordModel(word="【】", enabled=True),
                    SensitiveWordModel(word="即納", enabled=True),
                    SensitiveWordModel(word="翌日配達", enabled=True),
                ]
            )
            session.flush()

            saved = crawler_service.upsert_product(session, "alice", "task-1", item)
            self.assertTrue(saved)
            session.flush()

            row = session.scalar(select(ProductModel))
            self.assertIsNotNone(row)
            assert row is not None
            self.assertEqual(row.title, "春物")
            self.assertEqual(
                json.loads(row.raw_payload_json),
                {
                    "title": "春物",
                    "itemName": "春物",
                    "tagline": "おすすめ",
                },
            )

        self.assertEqual(item["title"], "【楽天1位】 即納 春物")
        self.assertEqual(item["raw"]["title"], "【楽天1位】 即納 春物")
        self.assertEqual(item["raw"]["itemName"], "【楽天1位】 即納 春物")
        self.assertEqual(item["raw"]["tagline"], "翌日配達 おすすめ")

    def test_upsert_product_returns_false_when_sanitized_title_becomes_empty(self) -> None:
        from app.db.models import ProductModel, SensitiveWordModel
        from app.services import crawler_service

        item = {
            "source_url": "https://example.test/item/2",
            "title": "【楽天1位】 即納",
            "raw": {
                "title": "【楽天1位】 即納",
                "itemName": "【楽天1位】 即納",
            },
        }

        with self.session_scope() as session:
            session.add_all(
                [
                    SensitiveWordModel(word="【】", enabled=True),
                    SensitiveWordModel(word="即納", enabled=True),
                ]
            )
            session.flush()

            saved = crawler_service.upsert_product(session, "alice", "task-2", item)

            self.assertFalse(saved)
            self.assertEqual(session.scalar(select(func.count()).select_from(ProductModel)), 0)

    def test_save_collected_item_returns_specific_error_when_sensitive_words_clear_title(self) -> None:
        from app.db.models import ProductModel
        from app.services import crawler_service

        item = {
            "source_url": "https://example.test/item/3",
            "title": "【楽天1位】 即納",
            "raw": {
                "title": "【楽天1位】 即納",
                "itemName": "【楽天1位】 即納",
            },
        }

        with patch.object(crawler_service, "session_scope", self.session_scope):
            result = crawler_service.save_collected_item(
                "alice",
                "task-3",
                item,
                active_words=["【】", "即納"],
            )

        self.assertEqual(
            result,
            {
                "saved": False,
                "error": "商品标题命中敏感词后为空，商品未保存。",
            },
        )
        with Session(self.engine, future=True) as session:
            self.assertEqual(session.scalar(select(func.count()).select_from(ProductModel)), 0)

    def test_save_collected_item_propagates_active_word_loading_failure(self) -> None:
        from app.services import crawler_service

        item = {
            "source_url": "https://example.test/item/4",
            "title": "春物",
            "raw": {
                "title": "春物",
                "itemName": "春物",
            },
        }

        with (
            patch.object(crawler_service, "session_scope", self.session_scope),
            patch.object(crawler_service, "active_sensitive_words", side_effect=RuntimeError("敏感词加载失败")),
        ):
            with self.assertRaisesRegex(RuntimeError, "敏感词加载失败"):
                crawler_service.save_collected_item("alice", "task-4", item)

    def test_upsert_store_product_preserves_dedupe_and_listed_store_mark_with_preloaded_words(self) -> None:
        from app.db.models import ProductModel, StoreModel
        from app.services import crawler_service

        linked_at = datetime(2026, 7, 13, 12, 0, 0)

        with self.session_scope() as session:
            store = StoreModel(owner_username="alice", store_code="demo", store_name="Demo Store")
            parent = ProductModel(
                owner_username="alice",
                title="原商品",
                source_url="https://source.example/item/1",
                source_url_hash="parent-hash-1",
                review_status="approved",
                raw_payload_json="{}",
            )
            session.add_all([store, parent])
            session.flush()

            previous_store_links = {
                "MN-1": {
                    "parentProductId": parent.id,
                    "listedAt": linked_at,
                }
            }
            first_item = {
                "manageNumber": "MN-1",
                "itemNumber": "IT-1",
                "itemUrl": "https://item.rakuten.co.jp/demo/it-1/",
                "itemName": "【楽天1位】 即納 春物",
            }
            second_item = {
                **first_item,
                "itemName": "【楽天1位】 即納 夏物",
            }

            with patch.object(crawler_service, "active_sensitive_words", side_effect=AssertionError("should not load")):
                first_saved = crawler_service.upsert_store_product(
                    session,
                    "alice",
                    store,
                    first_item,
                    previous_store_links=previous_store_links,
                    active_words=["【】", "即納"],
                )
                second_saved = crawler_service.upsert_store_product(
                    session,
                    "alice",
                    store,
                    second_item,
                    previous_store_links=previous_store_links,
                    active_words=["【】", "即納"],
                )
            session.flush()

            store_products = session.scalars(
                select(ProductModel).where(ProductModel.store_id == store.id).order_by(ProductModel.id.asc())
            ).all()
            parent_row = session.get(ProductModel, parent.id)

        self.assertTrue(first_saved)
        self.assertTrue(second_saved)
        self.assertEqual(len(store_products), 1)
        self.assertEqual(store_products[0].title, "夏物")
        self.assertEqual(store_products[0].rakuten_manage_number, "MN-1")
        self.assertEqual(store_products[0].parent_product_id, parent.id)
        self.assertEqual(store_products[0].listed_at, linked_at)
        self.assertIsNotNone(parent_row)
        assert parent_row is not None
        self.assertEqual(parent_row.review_status, "listed_master")
        parent_payload = json.loads(parent_row.raw_payload_json)
        self.assertEqual(len(parent_payload["listedStores"]), 1)
        self.assertEqual(parent_payload["listedStores"][0]["storeId"], store.id)

    def test_run_task_loads_active_words_once_and_records_specific_empty_title_error(self) -> None:
        from app.db.models import CrawlTaskModel, ProductModel, SensitiveWordModel
        from app.services import crawler_service

        with self.session_scope() as session:
            session.add(
                CrawlTaskModel(
                    id="task-sensitive-empty-title",
                    owner_username="alice",
                    source_type="shop",
                    target="店铺:https://www.rakuten.co.jp/demo/ 全部",
                    mode="manual",
                    status="queued",
                    total_count=0,
                    success_count=0,
                    failed_count=0,
                    warning_count=0,
                    message="等待执行",
                )
            )
            session.add_all(
                [
                    SensitiveWordModel(word="【】", enabled=True),
                    SensitiveWordModel(word="即納", enabled=True),
                ]
            )

        collected_items = [
            {
                "source_url": "https://example.test/item/ok",
                "title": "春物",
                "raw": {
                    "title": "春物",
                    "itemName": "春物",
                },
            },
            {
                "source_url": "https://example.test/item/empty",
                "title": "【楽天1位】 即納",
                "raw": {
                    "title": "【楽天1位】 即納",
                    "itemName": "【楽天1位】 即納",
                },
            },
        ]

        with (
            patch.object(crawler_service, "session_scope", self.session_scope),
            patch.object(crawler_service.settings, "task_queue_mode", "thread"),
            patch.object(crawler_service, "collect_items", return_value=collected_items),
            patch.object(crawler_service, "localize_collected_product_images", return_value=""),
            patch.object(crawler_service, "log_event", Mock()),
            patch.object(crawler_service, "dispatch_queued_crawl_tasks_safely", Mock()),
            patch.object(crawler_service, "reconcile_interrupted_running_tasks", return_value=0),
            patch.object(crawler_service, "active_sensitive_words", wraps=crawler_service.active_sensitive_words) as mock_active_words,
        ):
            crawler_service.run_task("task-sensitive-empty-title")

        with Session(self.engine, future=True) as session:
            task = session.get(CrawlTaskModel, "task-sensitive-empty-title")
            products = session.scalars(select(ProductModel).order_by(ProductModel.id.asc())).all()

        self.assertEqual(mock_active_words.call_count, 1)
        self.assertIsNotNone(task)
        assert task is not None
        self.assertEqual(task.status, "partial")
        self.assertEqual(task.success_count, 1)
        self.assertEqual(task.failed_count, 1)
        self.assertIn("商品标题命中敏感词后为空，商品未保存。", task.error_detail or "")
        self.assertEqual(len(products), 1)
        self.assertEqual(products[0].title, "春物")


class SensitiveWordCleanupTests(SensitiveWordDatabaseTestCase):
    def seed_pending_cleanup_rows(self) -> dict[str, int]:
        from app.db.models import ProductModel, SensitiveWordModel

        with self.session_scope() as session:
            session.add_all(
                [
                    SensitiveWordModel(word="【】", enabled=True),
                    SensitiveWordModel(word="即納", enabled=True),
                ]
            )
            session.flush()

            rows = [
                ProductModel(
                    owner_username="alice",
                    title="【楽天1位】 即納 春物",
                    source_url="https://example.test/pending-updatable",
                    source_url_hash="cleanup-pending-updatable",
                    review_status="pending",
                    raw_payload_json=json.dumps(
                        {
                            "title": "【楽天1位】 即納 春物",
                            "itemName": "【楽天1位】 即納 春物",
                            "tagline": "即納 おすすめ",
                        },
                        ensure_ascii=False,
                    ),
                ),
                ProductModel(
                    owner_username="alice",
                    title="【楽天1位】 即納",
                    source_url="https://example.test/pending-empty-title",
                    source_url_hash="cleanup-pending-empty",
                    review_status="pending",
                    raw_payload_json=json.dumps(
                        {
                            "title": "【楽天1位】 即納",
                            "itemName": "【楽天1位】 即納",
                            "tagline": "即納",
                        },
                        ensure_ascii=False,
                    ),
                ),
                ProductModel(
                    owner_username="alice",
                    title="常规商品",
                    source_url="https://example.test/pending-clean",
                    source_url_hash="cleanup-pending-clean",
                    review_status="pending",
                    raw_payload_json=json.dumps(
                        {
                            "title": "常规商品",
                            "itemName": "常规商品",
                            "tagline": "普通推荐",
                        },
                        ensure_ascii=False,
                    ),
                ),
                ProductModel(
                    owner_username="alice",
                    title="【楽天1位】 即納 已审核商品",
                    source_url="https://example.test/approved",
                    source_url_hash="cleanup-approved",
                    review_status="approved",
                    raw_payload_json=json.dumps(
                        {
                            "title": "【楽天1位】 即納 已审核商品",
                            "itemName": "【楽天1位】 即納 已审核商品",
                        },
                        ensure_ascii=False,
                    ),
                ),
                ProductModel(
                    owner_username="alice",
                    title="【楽天1位】 即納 已上架商品",
                    source_url="https://example.test/listed",
                    source_url_hash="cleanup-listed",
                    review_status="listed",
                    raw_payload_json=json.dumps(
                        {
                            "title": "【楽天1位】 即納 已上架商品",
                            "itemName": "【楽天1位】 即納 已上架商品",
                        },
                        ensure_ascii=False,
                    ),
                ),
                ProductModel(
                    owner_username="alice",
                    title="【楽天1位】 即納 异常商品",
                    source_url="https://example.test/error",
                    source_url_hash="cleanup-error",
                    review_status="error",
                    raw_payload_json=json.dumps(
                        {
                            "title": "【楽天1位】 即納 异常商品",
                            "itemName": "【楽天1位】 即納 异常商品",
                        },
                        ensure_ascii=False,
                    ),
                ),
            ]
            session.add_all(rows)
            session.flush()
            return {
                "pending_updatable": rows[0].id,
                "pending_empty": rows[1].id,
                "pending_clean": rows[2].id,
                "approved": rows[3].id,
                "listed": rows[4].id,
                "error": rows[5].id,
            }

    def test_cleanup_pending_products_dry_run_reports_without_persisting(self) -> None:
        from app.db.models import ProductModel
        from app.services.sensitive_word_service import cleanup_pending_products

        row_ids = self.seed_pending_cleanup_rows()

        with self.session_scope() as session:
            summary = cleanup_pending_products(session, apply=False)

        self.assertEqual(
            summary,
            {
                "scannedCount": 3,
                "matchedCount": 2,
                "updatedCount": 1,
                "emptyTitleCount": 1,
            },
        )

        with Session(self.engine, future=True) as session:
            pending_updatable = session.get(ProductModel, row_ids["pending_updatable"])
            pending_empty = session.get(ProductModel, row_ids["pending_empty"])
            approved = session.get(ProductModel, row_ids["approved"])
            listed = session.get(ProductModel, row_ids["listed"])
            error = session.get(ProductModel, row_ids["error"])

        self.assertIsNotNone(pending_updatable)
        self.assertIsNotNone(pending_empty)
        self.assertIsNotNone(approved)
        self.assertIsNotNone(listed)
        self.assertIsNotNone(error)
        assert pending_updatable is not None
        assert pending_empty is not None
        assert approved is not None
        assert listed is not None
        assert error is not None
        self.assertEqual(pending_updatable.title, "【楽天1位】 即納 春物")
        self.assertEqual(json.loads(pending_updatable.raw_payload_json)["title"], "【楽天1位】 即納 春物")
        self.assertEqual(pending_empty.title, "【楽天1位】 即納")
        self.assertEqual(json.loads(pending_empty.raw_payload_json)["title"], "【楽天1位】 即納")
        self.assertEqual(approved.title, "【楽天1位】 即納 已审核商品")
        self.assertEqual(listed.title, "【楽天1位】 即納 已上架商品")
        self.assertEqual(error.title, "【楽天1位】 即納 异常商品")

    def test_cleanup_pending_products_apply_updates_only_pending_nonempty_titles(self) -> None:
        from app.db.models import ProductModel
        from app.services.sensitive_word_service import cleanup_pending_products

        row_ids = self.seed_pending_cleanup_rows()

        with self.session_scope() as session:
            summary = cleanup_pending_products(session, apply=True)
            session.flush()

        self.assertEqual(
            summary,
            {
                "scannedCount": 3,
                "matchedCount": 2,
                "updatedCount": 1,
                "emptyTitleCount": 1,
            },
        )

        with Session(self.engine, future=True) as session:
            pending_updatable = session.get(ProductModel, row_ids["pending_updatable"])
            pending_empty = session.get(ProductModel, row_ids["pending_empty"])
            pending_clean = session.get(ProductModel, row_ids["pending_clean"])
            approved = session.get(ProductModel, row_ids["approved"])
            listed = session.get(ProductModel, row_ids["listed"])
            error = session.get(ProductModel, row_ids["error"])

        self.assertIsNotNone(pending_updatable)
        self.assertIsNotNone(pending_empty)
        self.assertIsNotNone(pending_clean)
        self.assertIsNotNone(approved)
        self.assertIsNotNone(listed)
        self.assertIsNotNone(error)
        assert pending_updatable is not None
        assert pending_empty is not None
        assert pending_clean is not None
        assert approved is not None
        assert listed is not None
        assert error is not None
        self.assertEqual(pending_updatable.title, "春物")
        self.assertEqual(
            json.loads(pending_updatable.raw_payload_json),
            {
                "title": "春物",
                "itemName": "春物",
                "tagline": "おすすめ",
            },
        )
        self.assertEqual(pending_empty.title, "【楽天1位】 即納")
        self.assertEqual(
            json.loads(pending_empty.raw_payload_json),
            {
                "title": "【楽天1位】 即納",
                "itemName": "【楽天1位】 即納",
                "tagline": "即納",
            },
        )
        self.assertEqual(pending_clean.title, "常规商品")
        self.assertEqual(json.loads(pending_clean.raw_payload_json)["title"], "常规商品")
        self.assertEqual(approved.title, "【楽天1位】 即納 已审核商品")
        self.assertEqual(listed.title, "【楽天1位】 即納 已上架商品")
        self.assertEqual(error.title, "【楽天1位】 即納 异常商品")

    def test_cleanup_pending_products_apply_skips_malformed_and_non_object_payloads_without_partial_title_updates(self) -> None:
        from app.db.models import ProductModel, SensitiveWordModel
        from app.services.sensitive_word_service import cleanup_pending_products

        malformed_payload = '{"title":"【楽天1位】 即納 春物"'
        non_object_payload = '["【楽天1位】 即納 夏物"]'

        with self.session_scope() as session:
            session.add_all(
                [
                    SensitiveWordModel(word="【】", enabled=True),
                    SensitiveWordModel(word="即納", enabled=True),
                ]
            )
            session.flush()

            malformed_row = ProductModel(
                owner_username="alice",
                title="【楽天1位】 即納 春物",
                source_url="https://example.test/pending-malformed-payload",
                source_url_hash="cleanup-pending-malformed-payload",
                review_status="pending",
                raw_payload_json=malformed_payload,
            )
            non_object_row = ProductModel(
                owner_username="alice",
                title="【楽天1位】 即納 夏物",
                source_url="https://example.test/pending-non-object-payload",
                source_url_hash="cleanup-pending-non-object-payload",
                review_status="pending",
                raw_payload_json=non_object_payload,
            )
            session.add_all([malformed_row, non_object_row])
            session.flush()

            malformed_id = malformed_row.id
            non_object_id = non_object_row.id

            summary = cleanup_pending_products(session, apply=True)
            session.flush()

        self.assertEqual(
            summary,
            {
                "scannedCount": 2,
                "matchedCount": 2,
                "updatedCount": 0,
                "emptyTitleCount": 0,
            },
        )

        with Session(self.engine, future=True) as session:
            malformed_row = session.get(ProductModel, malformed_id)
            non_object_row = session.get(ProductModel, non_object_id)

        self.assertIsNotNone(malformed_row)
        self.assertIsNotNone(non_object_row)
        assert malformed_row is not None
        assert non_object_row is not None
        self.assertEqual(malformed_row.title, "【楽天1位】 即納 春物")
        self.assertEqual(malformed_row.raw_payload_json, malformed_payload)
        self.assertEqual(non_object_row.title, "【楽天1位】 即納 夏物")
        self.assertEqual(non_object_row.raw_payload_json, non_object_payload)


if __name__ == "__main__":
    unittest.main()
